"""
COA Master Import Service
Parses ანგარიშები.xlsx (1C Chart of Accounts) into COAMasterAccount records.
Derives IFRS mappings from account type (А/П/АП) + class rules + GEORGIAN_COA.
"""
import re, os
from openpyxl import load_workbook

# Account type translation
_TYPE_MAP = {"А": "Active", "П": "Passive", "АП": "Mixed"}

# Georgian boolean flags
_YES_VALUES = {"დიახ", "yes", "да", "true", "1"}

# Class → default IFRS derivation rules (side, sub, bs_line/pl_line)
_CLASS_RULES = {
    "0": {"ifrs_side": "asset",     "ifrs_sub": "current",    "ifrs_bs_line": "Other Assets"},
    "1": {"ifrs_side": "asset",     "ifrs_sub": "current",    "ifrs_bs_line": "Current Assets"},
    "2": {"ifrs_side": "asset",     "ifrs_sub": "noncurrent", "ifrs_bs_line": "Noncurrent Assets"},
    "3": {"ifrs_side": "liability", "ifrs_sub": "current",    "ifrs_bs_line": "Current Liabilities"},
    "4": {"ifrs_side": "liability", "ifrs_sub": "noncurrent", "ifrs_bs_line": "Noncurrent Liabilities"},
    "5": {"ifrs_side": "equity",    "ifrs_sub": "equity",     "ifrs_bs_line": "Equity"},
    "6": {"ifrs_side": "income",    "ifrs_sub": "",           "ifrs_pl_line": "Revenue"},
    "7": {"ifrs_side": "expense",   "ifrs_sub": "",           "ifrs_pl_line": "Expenses"},
    "8": {"ifrs_side": "expense",   "ifrs_sub": "",           "ifrs_pl_line": "Other Income/Expense"},
    "9": {"ifrs_side": "expense",   "ifrs_sub": "",           "ifrs_pl_line": "Other P&L"},
}

# Override side based on account type for mixed scenarios
_TYPE_SIDE_OVERRIDE = {
    # (class_digit, account_type) → overridden side
    ("3", "А"):  "asset",      # e.g. 3340 Input VAT — type А in liability class
    ("4", "А"):  "asset",      # e.g. 4210 DTA — type А in noncurrent liabilities
    ("1", "П"):  "liability",  # unlikely but handle
    ("2", "П"):  "liability",
}


def _parse_bool_flag(val) -> bool:
    if val is None: return False
    return str(val).strip().lower() in _YES_VALUES


def _normalize_code(code: str) -> str:
    """Strip non-digit characters for prefix matching."""
    return re.sub(r'[^0-9]', '', code)


def _split_bilingual(name: str):
    """Split 'Georgian // Russian' into (ka, ru). If no separator, return (name, '')."""
    if not name: return ("", "")
    if "//" in name:
        parts = name.split("//", 1)
        return (parts[0].strip(), parts[1].strip())
    return (name.strip(), "")


def parse_coa_xlsx(file_path: str) -> list:
    """Parse ანგარიშები.xlsx → list of dicts ready for COAMasterAccount insertion."""
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.active
    accounts = []
    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        code_raw = row[1]  # Column B: Код
        if code_raw is None: continue
        code = str(code_raw).strip()
        if not code: continue
        # Skip pure Cyrillic prefix codes like "Я96.01" — rarely used
        normalized = _normalize_code(code)
        if not normalized: continue
        name_full = str(row[3] or "").strip()  # Column D: Наименование
        name_ka, name_ru = _split_bilingual(name_full)
        acct_type = str(row[5] or "").strip()  # Column F: Акт.
        acct_type_en = _TYPE_MAP.get(acct_type, "Unknown")
        is_off_bal = _parse_bool_flag(row[4])   # Column E: Заб.
        tracks_cur = _parse_bool_flag(row[6])    # Column G: Вал.
        tracks_qty = _parse_bool_flag(row[7])    # Column H: Кол.
        sub1 = str(row[8] or "").strip() or None  # Column I
        sub2 = str(row[9] or "").strip() or None  # Column J
        sub3 = str(row[10] or "").strip() or None # Column K
        prefix = normalized[:2] if len(normalized) >= 2 else normalized
        accounts.append({
            "account_code": code,
            "account_code_normalized": normalized,
            "account_prefix": prefix,
            "name_ka": name_ka or None,
            "name_ru": name_ru or None,
            "account_type": acct_type,
            "account_type_en": acct_type_en,
            "is_off_balance": is_off_bal,
            "tracks_currency": tracks_cur,
            "tracks_quantity": tracks_qty,
            "subconto_1": sub1,
            "subconto_2": sub2,
            "subconto_3": sub3,
        })
    wb.close()
    return accounts


def derive_ifrs_mappings(accounts: list, georgian_coa: dict) -> list:
    """
    For each account, derive IFRS fields using 3-step priority:
    1. Exact/prefix match in GEORGIAN_COA → inherit curated mappings
    2. Account type (А/П/АП) + class rules → auto-derive
    3. Pure class fallback

    Modifies accounts in-place and returns them.
    """
    for acct in accounts:
        norm = acct["account_code_normalized"]
        prefix = acct["account_prefix"]
        acct_type = acct.get("account_type", "")
        class_digit = norm[0] if norm else "0"
        # Step 1: Try GEORGIAN_COA match (longest prefix first)
        coa_entry = None
        for length in range(len(norm), 0, -1):
            test_prefix = norm[:length]
            if test_prefix in georgian_coa:
                coa_entry = georgian_coa[test_prefix]
                break
        if coa_entry:
            # Inherit from curated GEORGIAN_COA
            if "bs" in coa_entry:
                acct["ifrs_bs_line"] = coa_entry.get("bs", "")
                acct["ifrs_side"] = coa_entry.get("bs_side", "asset")
                acct["ifrs_sub"] = coa_entry.get("bs_sub", "current")
                acct["is_contra"] = coa_entry.get("contra", False)
            elif "pl" in coa_entry:
                acct["ifrs_pl_line"] = coa_entry.get("pl", "")
                acct["ifrs_side"] = coa_entry.get("side", "expense")
                acct["ifrs_pl_category"] = coa_entry.get("pl_line", "")
            continue
        # Step 2: Derive from account type + class rules
        rules = _CLASS_RULES.get(class_digit, _CLASS_RULES["0"])
        acct["ifrs_side"] = rules.get("ifrs_side", "asset")
        acct["ifrs_sub"] = rules.get("ifrs_sub", "")
        # Check type-based override (e.g. class 3 type А → asset)
        override_key = (class_digit, acct_type)
        if override_key in _TYPE_SIDE_OVERRIDE:
            acct["ifrs_side"] = _TYPE_SIDE_OVERRIDE[override_key]
            if acct["ifrs_side"] == "asset":
                acct["ifrs_sub"] = "current" if class_digit in ("1", "3") else "noncurrent"
        if class_digit in ("6", "7", "8", "9"):
            acct["ifrs_pl_line"] = rules.get("ifrs_pl_line", "")
            # Derive pl_category from class
            if class_digit == "7":
                sub2 = norm[1] if len(norm) > 1 else ""
                if sub2 == "1": acct["ifrs_pl_category"] = "COGS"
                elif sub2 == "2": acct["ifrs_pl_category"] = "SGA"
                elif sub2 in ("3", "4"): acct["ifrs_pl_category"] = "SGA"
                elif sub2 == "5": acct["ifrs_pl_category"] = "Finance"
                elif sub2 == "7": acct["ifrs_pl_category"] = "Tax"
                else: acct["ifrs_pl_category"] = "SGA"
            elif class_digit == "6":
                acct["ifrs_side"] = "income"
        else:
            acct["ifrs_bs_line"] = rules.get("ifrs_bs_line", "Other Assets")
        # Contra detection: accounts starting with 22, 26 are accumulated depreciation/amortization
        if norm.startswith("22") or norm.startswith("26"):
            acct["is_contra"] = True
    return accounts
