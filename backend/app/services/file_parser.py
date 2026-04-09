"""
file_parser.py — Intelligent Excel/CSV parser with Georgian COA mapping.
Parses 1C accounting exports: Base (transactions), Revenue Breakdown, COGS Breakdown.
Product classification for Wholesale/Retail/Other with sub-categories.
Integrates Semantic Layer for enhanced transaction classification.
"""
import contextvars
import re
import logging
from typing import Optional
import openpyxl
import csv
import io
from app.services.schema_registry import SchemaRegistry

logger = logging.getLogger(__name__)

# ── Per-request state — isolated via ContextVar (async-safe, no cross-request mutation) ──
# Each async request gets its own snapshot; concurrent requests cannot overwrite each other.
_cv_revenue_mappings: contextvars.ContextVar[dict] = contextvars.ContextVar(
    "_user_revenue_mappings", default={}
)
_cv_cogs_mappings: contextvars.ContextVar[dict] = contextvars.ContextVar(
    "_user_cogs_mappings", default={}
)
_cv_coa_overrides: contextvars.ContextVar[dict] = contextvars.ContextVar(
    "_user_coa_overrides", default={}
)

# ── Process-wide COA master cache — same for all requests, read-only after first load ──
_coa_master_cache: dict = {}        # normalized_code → COA master entry (from ანგარიშები.xlsx)


def load_coa_overrides(overrides: list):
    """Load COAMappingOverride records for map_coa() to check first.

    Stores into a ContextVar so concurrent requests don't overwrite each other.
    """
    new_overrides: dict = {}
    for o in overrides:
        code = (o.get("account_code") or "").strip()
        if not code:
            continue
        entry = {"raw_code": code, "prefix": code}
        bs_side = o.get("bs_side", "")
        if bs_side in ("asset", "liability", "equity"):
            entry["bs"] = o.get("ifrs_line_item", "")
            entry["bs_side"] = bs_side
            entry["bs_sub"] = o.get("bs_sub", "current")
        else:
            entry["pl"] = o.get("ifrs_line_item", "")
            entry["side"] = bs_side or "expense"
            entry["pl_line"] = o.get("pl_line", "")
        new_overrides[code] = entry
    _cv_coa_overrides.set(new_overrides)
    logger.info(f"Loaded {len(new_overrides)} COA mapping overrides")


def load_coa_master(accounts: list):
    """Load COAMasterAccount records into cache for map_coa() tier-2 lookup."""
    global _coa_master_cache
    _coa_master_cache = {}
    for a in accounts:
        code = (a.get("account_code_normalized") or a.get("account_code") or "").strip()
        if not code: continue
        entry = {"raw_code": code, "prefix": a.get("account_prefix", code[:2])}
        side = a.get("ifrs_side", "")
        if side in ("asset", "liability", "equity"):
            entry["bs"] = a.get("ifrs_bs_line") or ""
            entry["bs_side"] = side
            entry["bs_sub"] = a.get("ifrs_sub") or "current"
            if a.get("is_contra"): entry["contra"] = True
        else:
            entry["pl"] = a.get("ifrs_pl_line") or ""
            entry["side"] = side or "expense"
            entry["pl_line"] = a.get("ifrs_pl_category") or ""
        entry["account_type"] = a.get("account_type", "")
        entry["account_type_en"] = a.get("account_type_en", "")
        entry["name_ka"] = a.get("name_ka", "")
        entry["name_ru"] = a.get("name_ru", "")
        entry["is_off_balance"] = a.get("is_off_balance", False)
        _coa_master_cache[code] = entry
    logger.info(f"Loaded {len(_coa_master_cache)} COA master accounts")


def load_user_mappings(approved_mappings: list):
    """Load user-approved ProductMapping records into per-request classification cache.

    Uses ContextVar so concurrent upload requests cannot overwrite each other's mappings.
    Call this before parsing a new file.
    Each mapping dict should have: product_name, revenue_category, cogs_category.
    """
    new_revenue: dict = {}
    new_cogs: dict = {}
    for m in approved_mappings:
        name = (m.get("product_name") or "").strip()
        if not name:
            continue
        rev_cat = m.get("revenue_category")
        cogs_cat = m.get("cogs_category")
        if rev_cat and rev_cat != "Other Revenue":
            new_revenue[name] = rev_cat
        if cogs_cat and cogs_cat != "Other COGS":
            new_cogs[name] = cogs_cat
    _cv_revenue_mappings.set(new_revenue)
    _cv_cogs_mappings.set(new_cogs)
    if new_revenue or new_cogs:
        logger.info(f"Loaded {len(new_revenue)} revenue + {len(new_cogs)} COGS user-approved mappings")


# ── Georgian Chart of Accounts ─────────────────────────────────────────────
def validate_schema(filename: str, content: bytes, sample_rows: int = 50):
    """Validate workbook schema deterministically using SchemaRegistry."""
    ext = filename.lower().rsplit('.', 1)[-1]
    registry = SchemaRegistry()
    sheet_rows = {}

    if ext in ('xlsx', 'xls', 'xlsm'):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if sample_rows and i >= sample_rows:
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheet_rows[name] = rows
    elif ext == 'csv':
        decoded = content.decode('utf-8', errors='replace')
        reader = list(csv.reader(io.StringIO(decoded)))
        sheet_rows["Sheet1"] = reader[:sample_rows] if sample_rows else reader
    else:
        return SchemaRegistry().validate_workbook({}, file_hint=ext)

    return registry.validate_workbook(sheet_rows, file_hint=filename)


GEORGIAN_COA = {
    # ══════════════════════════════════════════════════════════════════
    # Built from AccountN.xlsx (1C master COA, 134 base accounts).
    # Georgian names (bs_ka/pl_ka) from 1C, English names translated.
    # Account types verified: А=Asset, П=Passive/Liability, АП=Both.
    # ══════════════════════════════════════════════════════════════════

    # ── CLASS 6: REVENUE (შემოსავლები) ────────────────────────────────
    "6":    {"pl": "Revenue",              "pl_ka": "შემოსავლები",                    "side": "income"},
    "61":   {"pl": "Sales Revenue",        "pl_ka": "შემოსავალი რეალიზაციიდან",       "side": "income"},
    "611":  {"pl": "Sales Revenue",        "pl_ka": "შემოსავალი რეალიზაციიდან",       "side": "income", "segment": "Retail"},
    "612":  {"pl": "Discounts & Returns",  "pl_ka": "ფასდათმობა/დაბრუნება",           "side": "income", "contra_revenue": True},
    "613":  {"pl": "Wholesale Revenue",    "pl_ka": "საბითუმო შემოსავალი",            "side": "income", "segment": "Wholesale"},
    "614":  {"pl": "LPG/CNG Revenue",      "pl_ka": "თხევადი/ბუნებრივი აირი",        "side": "income", "segment": "Retail"},
    "62":   {"pl": "Service Revenue",      "pl_ka": "შემოსავალი მომსახურებიდან",      "side": "income"},

    # ── CLASS 7: EXPENSES (ხარჯები) ───────────────────────────────────
    "7":    {"pl": "Expenses",             "pl_ka": "ხარჯები",                        "side": "expense"},
    # 71xx — COGS (გაყიდული პროდუქციის თვითღირებულება)
    "71":   {"pl": "Cost of Sales",        "pl_ka": "თვითღირებულება",                  "side": "expense", "pl_line": "COGS"},
    "711":  {"pl": "COGS - Products",      "pl_ka": "პროდუქციის თვითღირებულება",      "side": "expense", "pl_line": "COGS"},
    "712":  {"pl": "COGS - Transport",     "pl_ka": "ტრანსპორტირება",                 "side": "expense", "pl_line": "COGS"},
    "713":  {"pl": "COGS - Fuel",          "pl_ka": "საწვავის ხარჯი",                 "side": "expense", "pl_line": "COGS"},
    "714":  {"pl": "COGS - Other",         "pl_ka": "სხვა თვითღირებულება",            "side": "expense", "pl_line": "COGS"},
    # 72xx — Labour (შრომის ხარჯები)
    "72":   {"pl": "Labour & HR",          "pl_ka": "შრომის ხარჯები",                 "side": "expense", "pl_line": "SGA", "sub": "Labour"},
    "721":  {"pl": "Salaries",             "pl_ka": "ხელფასები",                      "side": "expense", "pl_line": "SGA", "sub": "Labour"},
    "722":  {"pl": "Social Security",      "pl_ka": "სოციალური უზრუნველყოფა",        "side": "expense", "pl_line": "SGA", "sub": "Labour"},
    # 73xx — Selling Expenses (გაყიდვების ხარჯები) — per AccountN.xlsx 7310
    "73":   {"pl": "Selling Expenses",     "pl_ka": "გაყიდვების ხარჯები",            "side": "expense", "pl_line": "SGA", "sub": "Selling"},
    "731":  {"pl": "Selling Expenses",     "pl_ka": "გაყიდვების ხარჯები",            "side": "expense", "pl_line": "SGA", "sub": "Selling"},
    "732":  {"pl": "Selling Expenses",     "pl_ka": "გაყიდვების ხარჯები",            "side": "expense", "pl_line": "SGA", "sub": "Selling"},
    # 74xx — General Admin (საერთო ადმინისტრაციული) — D&A posted here in 1C
    "74":   {"pl": "General & Admin",      "pl_ka": "ადმინისტრაციული ხარჯები",       "side": "expense", "pl_line": "DA", "is_da": True},
    "741":  {"pl": "General Admin",        "pl_ka": "საერთო ადმინისტრაციული",        "side": "expense", "pl_line": "DA", "is_da": True},
    "742":  {"pl": "Management Expenses",  "pl_ka": "მართვის ხარჯები",               "side": "expense", "pl_line": "DA", "is_da": True},
    # 75xx — Finance Expense
    "75":   {"pl": "Finance Expense",      "pl_ka": "ფინანსური ხარჯი",               "side": "expense", "pl_line": "Finance"},
    "751":  {"pl": "Interest Expense",     "pl_ka": "პროცენტის ხარჯი",               "side": "expense", "pl_line": "Finance"},
    # 76xx — Finance Income
    "76":   {"pl": "Finance Income",       "pl_ka": "ფინანსური შემოსავალი",           "side": "income",  "pl_line": "Finance"},
    # 77xx — Income Tax
    "77":   {"pl": "Income Tax",           "pl_ka": "მოგების გადასახადი",             "side": "expense", "pl_line": "Tax"},
    "771":  {"pl": "Current Tax",          "pl_ka": "მიმდინარე გადასახადი",           "side": "expense", "pl_line": "Tax"},

    # ── CLASS 8: NON-OPERATING (არასაოპერაციო) ────────────────────────
    "8":    {"pl": "Other Income/Expense", "pl_ka": "არასაოპერაციო",                 "side": "expense"},
    "81":   {"pl": "Non-operating Income", "pl_ka": "არასაოპერაციო შემოსავლები",     "side": "income"},
    "811":  {"pl": "Non-operating Income", "pl_ka": "არასაოპერაციო შემოსავლები",     "side": "income"},
    "82":   {"pl": "Non-operating Expense","pl_ka": "არასაოპერაციო ხარჯები",         "side": "expense", "pl_line": "SGA", "sub": "Other"},
    "822":  {"pl": "Non-operating Expense","pl_ka": "არასაოპერაციო ხარჯები",         "side": "expense", "pl_line": "SGA", "sub": "Other"},
    "823":  {"pl": "Shortages & Losses",   "pl_ka": "დანაკლისი და დანაკარგები",      "side": "expense", "pl_line": "SGA", "sub": "Other"},

    # ── CLASS 9: OTHER P&L ────────────────────────────────────────────
    "92":   {"pl": "Other P&L",            "pl_ka": "სხვა მოგება-ზარალი",            "side": "expense", "pl_line": "SGA", "sub": "Other"},
    "921":  {"pl": "Other P&L",            "pl_ka": "სხვა მოგება-ზარალი",            "side": "expense", "pl_line": "SGA", "sub": "Other"},

    # ══════════════════════════════════════════════════════════════════
    # BALANCE SHEET — CLASS 1: CURRENT ASSETS (მოკლევადიანი აქტივები)
    # All 1xxx accounts are CURRENT per Georgian COA standard.
    # ══════════════════════════════════════════════════════════════════
    "1":    {"bs": "Current Assets",           "bs_ka": "მოკლევადიანი აქტივები",         "bs_side": "asset", "bs_sub": "current"},
    # 11xx — Cash (ნაღდი ფული)
    "11":   {"bs": "Cash & Equivalents",       "bs_ka": "ფულადი სახსრები",                "bs_side": "asset", "bs_sub": "current"},
    "111":  {"bs": "Cash in Hand (GEL)",       "bs_ka": "ნაღდი ფული ეროვნულ ვალუტაში",   "bs_side": "asset", "bs_sub": "current"},
    "112":  {"bs": "Cash in Hand (FX)",        "bs_ka": "ნაღდი ფული უცხოურ ვალუტაში",    "bs_side": "asset", "bs_sub": "current"},
    # 12xx — Bank Accounts (საბანკო ანგარიშები) — NOT "Receivables"!
    "12":   {"bs": "Bank Accounts",            "bs_ka": "საბანკო ანგარიშები",             "bs_side": "asset", "bs_sub": "current"},
    "121":  {"bs": "Bank Accounts (GEL)",      "bs_ka": "ეროვნული ვალუტა ბანკში",        "bs_side": "asset", "bs_sub": "current"},
    "122":  {"bs": "Bank Accounts (FX)",       "bs_ka": "უცხოური ვალუტა ბანკში",         "bs_side": "asset", "bs_sub": "current"},
    "129":  {"bs": "Money in Transit",         "bs_ka": "ფული გზაში",                     "bs_side": "asset", "bs_sub": "current"},
    # 13xx — Short-term Investments
    "13":   {"bs": "Short-term Investments",   "bs_ka": "მოკლევადიანი ინვესტიციები",      "bs_side": "asset", "bs_sub": "current"},
    # 14xx — Trade Receivables (მოთხოვნები) — NOT "Prepayments"!
    "14":   {"bs": "Trade Receivables",        "bs_ka": "მოთხოვნები",                     "bs_side": "asset", "bs_sub": "current"},
    "141":  {"bs": "Trade Receivables",        "bs_ka": "მოთხოვნები მიწოდებიდან",         "bs_side": "asset", "bs_sub": "current"},
    "1415": {"bs": "Doubtful Debt Allowance",  "bs_ka": "საეჭვო ვალების რეზერვი",        "bs_side": "asset", "bs_sub": "current", "contra": True},
    "143":  {"bs": "Employee Receivables",     "bs_ka": "მოთხოვნები თანამშრომლების მიმართ","bs_side": "asset", "bs_sub": "current"},
    "148":  {"bs": "Advances to Suppliers",    "bs_ka": "ავანსები მომწოდებლებზე",         "bs_side": "asset", "bs_sub": "current"},
    "149":  {"bs": "Other Receivables",        "bs_ka": "სხვა მოკლევადიანი მოთხოვნები",   "bs_side": "asset", "bs_sub": "current"},
    # 15xx — Other Current Assets
    "15":   {"bs": "Other Current Assets",     "bs_ka": "სხვა მოკლევადიანი აქტივები",     "bs_side": "asset", "bs_sub": "current"},
    # 16xx — Inventory (მარაგები) — CURRENT, not noncurrent!
    "16":   {"bs": "Inventory",                "bs_ka": "მარაგები",                        "bs_side": "asset", "bs_sub": "current"},
    "160":  {"bs": "Goods in Transit",         "bs_ka": "საქონელი გზაში",                 "bs_side": "asset", "bs_sub": "current"},
    "161":  {"bs": "Merchandise",              "bs_ka": "საქონელი",                        "bs_side": "asset", "bs_sub": "current"},
    "162":  {"bs": "Raw Materials & Fuel",     "bs_ka": "ნედლეული და საწვავი",            "bs_side": "asset", "bs_sub": "current"},
    "163":  {"bs": "Work in Progress",         "bs_ka": "დაუმთავრებელი წარმოება",         "bs_side": "asset", "bs_sub": "current"},
    "164":  {"bs": "Finished Goods",           "bs_ka": "მზა პროდუქცია",                  "bs_side": "asset", "bs_sub": "current"},
    # 17xx — Prepaid Taxes (წინასწარ გადახდილი გადასახადი) — CURRENT!
    "17":   {"bs": "Prepaid Taxes",            "bs_ka": "წინასწარ გადახდილი გადასახადი",   "bs_side": "asset", "bs_sub": "current"},
    "179":  {"bs": "Prepaid VAT",              "bs_ka": "წინასწარ გადახდილი დღგ",          "bs_side": "asset", "bs_sub": "current"},
    # 18xx — Dividends & Interest Receivable — CURRENT!
    "18":   {"bs": "Dividends & Interest Recv","bs_ka": "მისაღები დივიდენდები/პროცენტები", "bs_side": "asset", "bs_sub": "current"},
    "181":  {"bs": "Dividends Receivable",     "bs_ka": "მისაღები დივიდენდები",            "bs_side": "asset", "bs_sub": "current"},
    "182":  {"bs": "Interest Receivable",      "bs_ka": "მისაღები პროცენტები",             "bs_side": "asset", "bs_sub": "current"},
    # 19xx — Other Current Assets
    "19":   {"bs": "Other Current Assets",     "bs_ka": "სხვა მოკლევადიანი აქტივები",     "bs_side": "asset", "bs_sub": "current"},

    # ══════════════════════════════════════════════════════════════════
    # CLASS 2: NONCURRENT ASSETS (გრძელვადიანი აქტივები)
    # ══════════════════════════════════════════════════════════════════
    "2":    {"bs": "Noncurrent Assets",        "bs_ka": "გრძელვადიანი აქტივები",          "bs_side": "asset", "bs_sub": "noncurrent"},
    # 21xx — Fixed Assets (ძირითადი საშუალებები)
    "21":   {"bs": "Fixed Assets (PP&E)",      "bs_ka": "ძირითადი საშუალებები",            "bs_side": "asset", "bs_sub": "noncurrent"},
    "211":  {"bs": "Land",                     "bs_ka": "მიწა",                            "bs_side": "asset", "bs_sub": "noncurrent"},
    "212":  {"bs": "Construction in Progress", "bs_ka": "დაუმთავრებელი მშენებლობა",       "bs_side": "asset", "bs_sub": "noncurrent"},
    "213":  {"bs": "Fixed Assets",             "bs_ka": "ძირითადი საშუალებები",            "bs_side": "asset", "bs_sub": "noncurrent"},
    "214":  {"bs": "Investment Property",      "bs_ka": "საინვესტიციო ქონება",            "bs_side": "asset", "bs_sub": "noncurrent"},
    "215":  {"bs": "Land Acquisition",         "bs_ka": "მიწის შესყიდვა",                 "bs_side": "asset", "bs_sub": "noncurrent"},
    "216":  {"bs": "Fixed Asset Acquisition",  "bs_ka": "ძირ. საშუალების შესყიდვა",       "bs_side": "asset", "bs_sub": "noncurrent"},
    # 22xx — Accumulated Depreciation (contra)
    "22":   {"bs": "Accumulated Depreciation", "bs_ka": "ცვეთა და ამორტიზაცია",           "bs_side": "asset", "bs_sub": "noncurrent", "contra": True},
    "223":  {"bs": "Acc. Depr. - Fixed Assets","bs_ka": "ძირ. საშუალების ამორტიზაცია",    "bs_side": "asset", "bs_sub": "noncurrent", "contra": True},
    # 23xx — Deferred Tax Assets
    "23":   {"bs": "Deferred Tax Assets",      "bs_ka": "გადავადებული საგადასახადო აქტივი","bs_side": "asset", "bs_sub": "noncurrent"},
    # 24xx — Long-term Investments
    "24":   {"bs": "Long-term Investments",    "bs_ka": "გრძელვადიანი ინვესტიციები",      "bs_side": "asset", "bs_sub": "noncurrent"},
    # 25xx — Intangible Assets
    "25":   {"bs": "Intangible Assets",        "bs_ka": "არამატერიალური აქტივები",        "bs_side": "asset", "bs_sub": "noncurrent"},
    "251":  {"bs": "Intangible Assets",        "bs_ka": "არამატერიალური აქტივები",        "bs_side": "asset", "bs_sub": "noncurrent"},
    # 26xx — Accumulated Amortization (contra)
    "26":   {"bs": "Accumulated Amortization", "bs_ka": "არამატერ. აქტივის ამორტიზაცია",  "bs_side": "asset", "bs_sub": "noncurrent", "contra": True},
    "261":  {"bs": "Acc. Amort. - Intangibles","bs_ka": "არამატერ. აქტივის ამორტიზაცია",  "bs_side": "asset", "bs_sub": "noncurrent", "contra": True},

    # ══════════════════════════════════════════════════════════════════
    # CLASS 3: CURRENT LIABILITIES (მოკლევადიანი ვალდებულებები)
    # ══════════════════════════════════════════════════════════════════
    "3":    {"bs": "Current Liabilities",      "bs_ka": "მოკლევადიანი ვალდებულებები",     "bs_side": "liability", "bs_sub": "current"},
    # 31xx — Trade Payables
    "31":   {"bs": "Trade Payables",           "bs_ka": "ვალდებულებები მოწოდებიდან",      "bs_side": "liability", "bs_sub": "current"},
    "311":  {"bs": "Trade Payables",           "bs_ka": "ვალდებულებები მოწოდებიდან",      "bs_side": "liability", "bs_sub": "current"},
    "312":  {"bs": "Advances Received",        "bs_ka": "მიღებული ავანსები",              "bs_side": "liability", "bs_sub": "current"},
    "313":  {"bs": "Wages Payable",            "bs_ka": "გადასახდელი ხელფასი",            "bs_side": "liability", "bs_sub": "current"},
    "319":  {"bs": "Other Trade Payables",     "bs_ka": "სხვა ვალდებულებები მოწოდებიდან", "bs_side": "liability", "bs_sub": "current"},
    # 32xx — Short-term Debt
    "32":   {"bs": "Short-term Debt",          "bs_ka": "მოკლევადიანი სესხები",           "bs_side": "liability", "bs_sub": "current"},
    "321":  {"bs": "Short-term Loans",         "bs_ka": "მოკლევადიანი კრედიტები",         "bs_side": "liability", "bs_sub": "current"},
    "323":  {"bs": "Current Lease Liability",  "bs_ka": "იჯარის მიმდინარე ნაწილი",        "bs_side": "liability", "bs_sub": "current"},
    # 33xx — Tax Payables
    "33":   {"bs": "Tax Payables",             "bs_ka": "საგადასახადო ვალდებულებები",      "bs_side": "liability", "bs_sub": "current"},
    "331":  {"bs": "Income Tax Payable",       "bs_ka": "გადასახდელი მოგების გადასახადი",  "bs_side": "liability", "bs_sub": "current"},
    "332":  {"bs": "Revenue Tax Payable",      "bs_ka": "გადასახდელი საშემოსავლო",        "bs_side": "liability", "bs_sub": "current"},
    "333":  {"bs": "VAT Payable",              "bs_ka": "გადასახდელი დღგ",                 "bs_side": "liability", "bs_sub": "current"},
    "334":  {"bs": "Other Tax Payables",       "bs_ka": "სხვა საგადასახადო",              "bs_side": "liability", "bs_sub": "current"},
    # SPECIAL: 3340 is type А (Asset) — Input VAT receivable from government
    "3340": {"bs": "Input VAT (Asset)",        "bs_ka": "გადახდილი დღგ",                   "bs_side": "asset",     "bs_sub": "current"},
    "335":  {"bs": "Excise Payable",           "bs_ka": "გადასახდელი აქციზი",             "bs_side": "liability", "bs_sub": "current"},
    "337":  {"bs": "Pension Obligations",      "bs_ka": "საპენსიო ვალდებულებები",         "bs_side": "liability", "bs_sub": "current"},
    "338":  {"bs": "Property Tax Payable",     "bs_ka": "ქონების გადასახადი",             "bs_side": "liability", "bs_sub": "current"},
    "339":  {"bs": "Other Tax Liabilities",    "bs_ka": "სხვა საგადასახადო ვალდებულებები", "bs_side": "liability", "bs_sub": "current"},
    # 34xx — Accrued Liabilities
    "34":   {"bs": "Accrued Liabilities",      "bs_ka": "დარიცხული ვალდებულებები",        "bs_side": "liability", "bs_sub": "current"},
    "341":  {"bs": "Interest Payable",         "bs_ka": "გადასახდელი პროცენტები",         "bs_side": "liability", "bs_sub": "current"},
    "342":  {"bs": "Dividends Payable",        "bs_ka": "გადასახდელი დივიდენდები",        "bs_side": "liability", "bs_sub": "current"},
    "349":  {"bs": "Other Accrued Liabilities","bs_ka": "სხვა დარიცხული ვალდებულებები",   "bs_side": "liability", "bs_sub": "current"},

    # ══════════════════════════════════════════════════════════════════
    # CLASS 4: NONCURRENT LIABILITIES (გრძელვადიანი ვალდებულებები)
    # ══════════════════════════════════════════════════════════════════
    "4":    {"bs": "Noncurrent Liabilities",   "bs_ka": "გრძელვადიანი ვალდებულებები",     "bs_side": "liability", "bs_sub": "noncurrent"},
    "41":   {"bs": "Long-term Debt",           "bs_ka": "გრძელვადიანი სესხები",           "bs_side": "liability", "bs_sub": "noncurrent"},
    "413":  {"bs": "Usufruct Obligations",     "bs_ka": "უზუფრუქტის ვალდებულებები",      "bs_side": "liability", "bs_sub": "noncurrent"},
    "414":  {"bs": "Long-term Loans",          "bs_ka": "გრძელვადიანი სესხები",           "bs_side": "liability", "bs_sub": "noncurrent"},
    "417":  {"bs": "Long-term Lease Liability","bs_ka": "იჯარის გრძელვადიანი ვალდებულება","bs_side": "liability", "bs_sub": "noncurrent"},
    "419":  {"bs": "Other LT Liabilities",     "bs_ka": "სხვა გრძელვადიანი ვალდებულებები","bs_side": "liability", "bs_sub": "noncurrent"},
    # 4210=А (DTA asset), 4211=П (DTL liability)
    "4210": {"bs": "Deferred Tax Assets",      "bs_ka": "გადავადებული საგადასახადო აქტივი","bs_side": "asset",     "bs_sub": "noncurrent"},
    "4211": {"bs": "Deferred Tax Liabilities", "bs_ka": "გადავადებული საგადასახადო ვალდ.", "bs_side": "liability", "bs_sub": "noncurrent"},

    # ══════════════════════════════════════════════════════════════════
    # CLASS 5: EQUITY (საკუთარი კაპიტალი)
    # ══════════════════════════════════════════════════════════════════
    "5":    {"bs": "Equity",                   "bs_ka": "საკუთარი კაპიტალი",              "bs_side": "equity", "bs_sub": "equity"},
    "51":   {"bs": "Share Capital",            "bs_ka": "საწესდებო კაპიტალი",             "bs_side": "equity", "bs_sub": "equity"},
    "53":   {"bs": "Retained Earnings",        "bs_ka": "გაუნაწილებელი მოგება",            "bs_side": "equity", "bs_sub": "equity"},
    "54":   {"bs": "Reserves",                 "bs_ka": "რეზერვები",                       "bs_side": "equity", "bs_sub": "equity"},
}

# ── G&A Account Codes ──────────────────────────────────────────────────────
GA_ACCOUNT_CODES = {"7310.02.1", "8220.01.1", "9210"}
DA_ACCOUNT_CODES = {"7410", "7410.01"}

GA_ACCOUNT_NAMES = {
    "7310.01.1": "Circulation Expenses (Non-ENVD)",
    "7310.01.99":"Non-deductible Expenses (Commercial)",
    "7310.02.1": "Commercial Expenses in Production (Non-ENVD)",
    "8220.01.1": "Non-operating Expenses (Non-ENVD)",
    "8220.01.2": "Non-operating Expenses (Distributable)",
    "8230":      "Shortages & Losses from Spoilage",
    "9210":      "Other P&L Account",
}

DA_ACCOUNT_NAMES = {
    "7410":      "General Admin (D&A)",        # 7410 = საერთო ადმინისტრაციული ხარჯები
    "7410.01":   "Depreciation & Amortization",
}

# ── Product Dimension Table ────────────────────────────────────────────────
# Structured product ontology: each product has fuel type, segment, trade type.
# Categories are DERIVED from dimensions — not hardcoded per product.
# This replaces the old REVENUE_PRODUCT_MAP / COGS_PRODUCT_MAP approach.

PRODUCT_DIMENSIONS = [
    # ── Wholesale products (identified by trade type suffixes) ──
    {"name_ka": "ევრო რეგულარი (იმპორტი)",   "name_en": "Euro Regular (Import)",     "fuel": "Petrol",  "segment": "Wholesale", "trade": "Import",    "unit": "kg"},
    {"name_ka": "პრემიუმი (რეექსპორტი)",      "name_en": "Premium (Re-export)",       "fuel": "Petrol",  "segment": "Wholesale", "trade": "Re-export", "unit": "kg"},
    {"name_ka": "სუპერი (რეექსპორტი)",        "name_en": "Super (Re-export)",         "fuel": "Petrol",  "segment": "Wholesale", "trade": "Re-export", "unit": "kg"},
    {"name_ka": "ევრო რეგულარი (საბითუმო)",    "name_en": "Euro Regular (Wholesale)",  "fuel": "Petrol",  "segment": "Wholesale", "trade": "Wholesale", "unit": "L"},
    {"name_ka": "პრემიუმი (საბითუმო)",         "name_en": "Premium (Wholesale)",       "fuel": "Petrol",  "segment": "Wholesale", "trade": "Wholesale", "unit": "L"},
    {"name_ka": "დიზელი (საბითუმო)",           "name_en": "Diesel (Wholesale)",        "fuel": "Diesel",  "segment": "Wholesale", "trade": "Wholesale", "unit": "L"},
    {"name_ka": "ევროდიზელი  (ექსპორტი)",     "name_en": "Euro Diesel (Export)",      "fuel": "Diesel",  "segment": "Wholesale", "trade": "Export",    "unit": "kg"},
    {"name_ka": "ბიტუმი (საბითუმო)",           "name_en": "Bitumen (Wholesale)",       "fuel": "Bitumen", "segment": "Wholesale", "trade": "Wholesale", "unit": "kg"},
    # ── Retail products (no trade type suffix) ──
    {"name_ka": "ევრო რეგულარი",               "name_en": "Euro Regular",              "fuel": "Petrol",  "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "პრემიუმი",                    "name_en": "Premium",                   "fuel": "Petrol",  "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "სუპერი",                      "name_en": "Super",                     "fuel": "Petrol",  "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "დიზელი",                      "name_en": "Diesel",                    "fuel": "Diesel",  "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "ევრო დიზელი",                 "name_en": "Euro Diesel",               "fuel": "Diesel",  "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "ბუნებრივი აირი",               "name_en": "Natural Gas",               "fuel": "CNG",     "segment": "Retail",    "trade": "Retail",    "unit": "m3"},
    {"name_ka": "ბუნებრივი აირი (საბითუმო)",    "name_en": "Natural Gas (Wholesale)",   "fuel": "CNG",     "segment": "Retail",    "trade": "Retail",    "unit": "m3"},
    {"name_ka": "თხევადი აირი",                 "name_en": "LPG",                       "fuel": "LPG",     "segment": "Retail",    "trade": "Retail",    "unit": "L"},
    {"name_ka": "თხევადი გაზი (LPG) საბითუმო",  "name_en": "LPG (Wholesale)",           "fuel": "LPG",     "segment": "Wholesale", "trade": "Wholesale", "unit": "L"},
    {"name_ka": "დიზელი (იმპორტი)",             "name_en": "Diesel (Import)",           "fuel": "Diesel",  "segment": "Wholesale", "trade": "Import",    "unit": "kg"},
]

# Segment abbreviations used in category strings (matches income_statement.py)
_SEG_ABBREV = {"Wholesale": "Whsale", "Retail": "Retial"}

# Keyword sets for intelligent product suggestion (AI ontology)
_FUEL_KEYWORDS = {
    "Petrol":  ["რეგულარი", "პრემიუმი", "სუპერი", "ბენზინ", "petrol", "gasoline", "regular", "premium", "super", "95", "98"],
    "Diesel":  ["დიზელი", "diesel", "ევროდიზელი", "eurodiesel"],
    "CNG":     ["ბუნებრივი აირი", "natural gas", "cng", "მეთანი", "methane"],
    "LPG":     ["თხევადი აირი", "lpg", "პროპანი", "propane"],
    "Bitumen": ["ბიტუმი", "bitumen", "ასფალტ"],
}
_WHOLESALE_KEYWORDS = ["იმპორტი", "ექსპორტი", "რეექსპორტი", "საბითუმო", "wholesale", "import", "export", "re-export"]

# Service/non-fuel indicators — products containing these are NOT fuel products,
# even if they contain fuel-related substrings (e.g. "ბენზინმზიდები" = fuel tankers)
_SERVICE_KEYWORDS = [
    "მომსახურება", "იჯარა", "service", "rental", "rent", "lease",
    "მზიდ",       # carrier/transporter (ბენზინმზიდები = fuel tankers)
    "ტრანზიტ",    # transit
    "სასტუმრო",   # hotel
    "სამრეცხაო",  # car wash
    "დაცვ",       # security/protection
    "ელექტრო",    # electricity
    "წყალი",      # water
    "კომუნალ",    # communal/utility
    "ანაზღაურება", # compensation/reimbursement
]


def _find_product_dimension(product_name: str) -> Optional[dict]:
    """Find the best matching product dimension using structured lookup."""
    clean = product_name.strip()
    if not clean:
        return None
    # 1. Exact match (fastest)
    for dim in PRODUCT_DIMENSIONS:
        if dim["name_ka"] == clean or dim["name_en"] == clean:
            return dim
    # 2. Substring match — longest name first for specificity
    #    (e.g. "ევრო რეგულარი (იმპორტი)" matches before "ევრო რეგულარი")
    for dim in sorted(PRODUCT_DIMENSIONS, key=lambda d: -len(d["name_ka"])):
        if dim["name_ka"] in clean:
            return dim
    # 3. AI-like keyword analysis for unknown products
    return _suggest_dimension(clean)


def _suggest_dimension(product_name: str) -> Optional[dict]:
    """Semantic keyword analysis to suggest product dimension for unknown products."""
    p = product_name.lower()
    # Guard: skip service/non-fuel products that may contain fuel-related substrings
    # e.g. "იჯარა ბენზინმზიდების" (fuel tanker rental) contains "ბენზინ" but is NOT fuel
    if any(kw in p for kw in _SERVICE_KEYWORDS):
        return None
    # Detect fuel type
    fuel = None
    for fuel_type, keywords in _FUEL_KEYWORDS.items():
        if any(kw in p for kw in keywords):
            fuel = fuel_type
            break
    if not fuel:
        return None
    # Detect segment
    segment = "Retail"  # default
    if any(kw in p for kw in _WHOLESALE_KEYWORDS):
        segment = "Wholesale"
    return {
        "name_ka": product_name, "name_en": product_name,
        "fuel": fuel, "segment": segment, "trade": "Unknown", "unit": "?",
        "_suggested": True,  # flag: this is AI-suggested, not from known dimensions
    }


def _derive_category(dim: dict, source: str = "revenue") -> str:
    """Derive Revenue/COGS category string from product dimension attributes."""
    seg = _SEG_ABBREV.get(dim["segment"], "Retial")
    prefix = "Revenue" if source == "revenue" else "COGS"
    return f"{prefix} {seg} {dim['fuel']}"

# ── Georgian → English product name translation ──────────────────────────────
# Auto-generated from Product Dimension Table + manual unit-suffix variants
PRODUCT_ENGLISH_MAP = {dim["name_ka"]: dim["name_en"] for dim in PRODUCT_DIMENSIONS}
# Additional variants with comma/unit suffixes (from 1C export format)
PRODUCT_ENGLISH_MAP.update({
    "თხევადი აირი (მხოლოდ SGP !!!)": "LPG (SGP only)",
    "ევრო რეგულარი, ლ": "Euro Regular (L)",
    "პრემიუმი , ლ": "Premium (L)",
    "სუპერი , ლ": "Super (L)",
    "დიზელი, ლ": "Diesel (L)",
    "ევრო დიზელი, ლ": "Euro Diesel (L)",
    "ბუნებრივი აირი, მ3": "Natural Gas (m³)",
    "ბუნებრივი აირი (საბითუმო), მ3": "Natural Gas Wholesale (m³)",
    "ბიტუმი (საბითუმო), კგ": "Bitumen Wholesale (kg)",
    "ევრო რეგულარი (იმპორტი), კგ": "Euro Regular Import (kg)",
    "პრემიუმი (რეექსპორტი), კგ": "Premium Re-export (kg)",
    "სუპერი (რეექსპორტი), კგ": "Super Re-export (kg)",
    "დიზელი (საბითუმო), ლ": "Diesel Wholesale (L)",
    "ევროდიზელი  (ექსპორტი), კგ": "Euro Diesel Export (kg)",
    "თხევადი აირი (მხოლოდ SGP !!!), ლ": "LPG SGP (L)",
})
# ── Non-fuel / service product translations (Other Revenue items) ────────────
PRODUCT_ENGLISH_MAP.update({
    "სხვა ადმინისტრაციული (ტრანზიტული ბარათები), მომსახურება": "Other Administrative (Transit Cards)",
    "BP გადაზიდვის ღირებულება, მომსახურება": "BP Transportation Cost",
    "DRY GOODS SMARKETS, ცალი": "Dry Goods Smarkets (pcs)",
    "გადასახდელი იჯარა, მომსახურება": "Payable Lease",
    "გაცემული იჯარა (ინვოისისთ), მომსახურება": "Issued Lease (Invoiced)",
    "ელექტროენერგია (კომუნალური), მომსახურება": "Electricity (Utility)",
    "ზარალის ანაზღაურება, მომსახურება": "Damage Compensation",
    "იჯარა ბენზინმზიდების, მომსახურება": "Fuel Tanker Rental",
    "იჯარა ფართის  (მიწა და საწყობი), მომსახურება": "Property Lease (Land & Warehouse)",
    "იჯარა ჩიხით სარგებლობის, მომსახურება": "Dead-end Access Lease",
    "კომერციული ავტოტექმომსახურება, მომსახურება": "Commercial Auto Maintenance",
    "მოლეკულური საცერი molecular sieve, კგ": "Molecular Sieve (kg)",
    "მომსახურების ღირებულება,": "Service Fee",
    "მომსახურების ღირებულება ა/ფ გარეშე,": "Service Fee (w/o Invoice)",
    "პირადი დაცვის მომსახურება, მომსახურება": "Personal Security Service",
    "სამრეცხაო მომსახურება, მომსახურება": "Car Wash Service",
    "სხვა მომსახურება, მომსახურება": "Other Service",
    "სხვა მომსახურება Z - ებით ( სასტუმრო ტირ-პარკი ), მომსახურება": "Other Service via Z-Reports (Hotel, Truck Parking)",
    "ძრავის ზეთი მეორადი, ლ": "Used Motor Oil (L)",
    "წყალი, მ3": "Water (m³)",
    # Common service-related stems for substring matching
    "გადაზიდვა": "Transportation",
    "მომსახურება": "Service",
    "იჯარა": "Lease",
    "ანაზღაურება": "Compensation",
    "ელექტროენერგია": "Electricity",
    "კომუნალური": "Utility",
    "სამრეცხაო": "Car Wash",
    "დაცვა": "Security",
    "ავტოტექმომსახურება": "Auto Maintenance",
})


def get_english_name(georgian_name: str) -> str:
    """Translate a Georgian product name to English using Product Dimension Table."""
    if not georgian_name:
        return ""
    clean = georgian_name.strip()
    # 1. Product Dimension Table lookup
    dim = _find_product_dimension(clean)
    if dim and dim["name_en"] != clean:
        return dim["name_en"]
    # 2. Manual map (unit suffix variants)
    if clean in PRODUCT_ENGLISH_MAP:
        return PRODUCT_ENGLISH_MAP[clean]
    for ka, en in sorted(PRODUCT_ENGLISH_MAP.items(), key=lambda x: -len(x[0])):
        if ka in clean:
            return en
    return clean


def map_coa(code: str) -> Optional[dict]:
    """
    Map an account code to its COA entry using hierarchical fallback.

    Fallback chain for dotted sub-accounts (e.g., 7110.01/1):
      1. Strip non-digits → try longest prefix up to 4 digits (existing behavior)
      2. If code has dots/slashes, split on separators and try progressively
         shorter segment joins: "7110.01/1" → "711001" → "71100" → "7110" → match!
      3. Return None if nothing matches (caller should use class-level rules)
    """
    if not code:
        return None
    raw = str(code).strip()

    # Priority 0: User override (exact match) — read from per-request ContextVar
    _overrides = _cv_coa_overrides.get()
    if raw in _overrides:
        return {**_overrides[raw]}

    # Priority 1: COA Master exact match (406 accounts from ანგარიშები.xlsx)
    clean = re.sub(r'[^0-9]', '', raw)
    if clean in _coa_master_cache:
        return {**_coa_master_cache[clean]}

    # Priority 2: GEORGIAN_COA prefix match (114 curated entries)
    for length in range(min(len(clean), 4), 0, -1):
        prefix = clean[:length]
        if prefix in GEORGIAN_COA:
            return {**GEORGIAN_COA[prefix], "prefix": prefix, "raw_code": code}

    # Method 2: handle dotted/slashed codes (e.g., "7110.01.1", "7110.01/1")
    # Split on dots and slashes, then try joining progressively fewer parts
    parts = re.split(r'[./]', raw)
    if len(parts) > 1:
        for num_parts in range(len(parts), 0, -1):
            joined = ''.join(re.sub(r'[^0-9]', '', p) for p in parts[:num_parts])
            for length in range(min(len(joined), 4), 0, -1):
                prefix = joined[:length]
                if prefix in GEORGIAN_COA:
                    return {**GEORGIAN_COA[prefix], "prefix": prefix, "raw_code": code}

    return None


def infer_txn_type(acct_dr: str, acct_cr: str) -> str:
    """Infer transaction type from account codes."""
    dr_map = map_coa(acct_dr)
    cr_map = map_coa(acct_cr)
    if dr_map and dr_map.get("side") == "expense":
        return "Expense"
    if cr_map and cr_map.get("side") == "income":
        return "Income"
    if dr_map and dr_map.get("bs_side") == "asset":
        return "Asset Movement"
    return "Other"


# ── Classification helpers ─────────────────────────────────────────────────

def _classify_revenue_product(product_name: str) -> str:
    """Map a revenue product name to its category using Product Dimension Table."""
    clean = product_name.strip()
    # 1. Check user-approved mappings (from ProductMapping table) — per-request ContextVar
    _rev = _cv_revenue_mappings.get()
    if clean in _rev:
        return _rev[clean]
    for name, cat in _rev.items():
        if name in clean:
            return cat
    # 2. Product Dimension Table — derive category from (segment, fuel)
    dim = _find_product_dimension(clean)
    if dim:
        return _derive_category(dim, "revenue")
    return "Other Revenue"


def _classify_cogs_product(product_name: str, classification_col: str = "") -> tuple:
    """
    Map a COGS product to (segment, category) using Product Dimension Table.
    Returns (segment, category) tuple.
    """
    clean = product_name.strip()
    # 1. Check user-approved COGS mappings first — per-request ContextVar
    _cogs = _cv_cogs_mappings.get()
    if clean in _cogs:
        cat = _cogs[clean]
        seg = "COGS Wholesale" if "Whsale" in cat else "COGS Retail" if "Retial" in cat else "Other COGS"
        return seg, cat
    for name, cat in _cogs.items():
        if name in clean:
            seg = "COGS Wholesale" if "Whsale" in cat else "COGS Retail" if "Retial" in cat else "Other COGS"
            return seg, cat
    # 2. Product Dimension Table — derive category from (segment, fuel)
    dim = _find_product_dimension(clean)
    if dim:
        # Override: if product name contains wholesale keywords, force Wholesale segment
        # This handles cases like "დიზელი (იმპორტი)" matching "დიზელი" (Retail) by substring
        effective_segment = dim["segment"]
        if effective_segment == "Retail" and any(kw in clean.lower() for kw in _WHOLESALE_KEYWORDS):
            effective_segment = "Wholesale"
        cat = _derive_category({**dim, "segment": effective_segment}, "cogs")
        seg = "COGS Wholesale" if effective_segment == "Wholesale" else "COGS Retail"
        return seg, cat
    # 3. Fallback to R column classification
    if classification_col and classification_col not in ("0", "???", ""):
        return classification_col, "Other COGS"
    return "Other COGS", "Other COGS"


def _is_ga_account(acct_code: str) -> bool:
    """Check if an account code is a G&A expense account."""
    return str(acct_code).strip() in GA_ACCOUNT_CODES


def _is_da_account(acct_code: str) -> bool:
    """Check if an account code is a D&A account."""
    return str(acct_code).strip() in DA_ACCOUNT_CODES


def _detect_period(filename: str) -> Optional[str]:
    """Auto-detect period from filename."""
    fn = filename.lower()
    months = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    full_months = ['January','February','March','April','May','June',
                   'July','August','September','October','November','December']
    m = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[_\- ]?(20\d\d)', fn)
    if not m:
        m = re.search(r'(20\d\d)[_\- ]?(0[1-9]|1[0-2])', fn)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
    if m:
        mo = m.group(1)[:3]
        yr = m.group(2) if len(m.groups()) > 1 else m.group(1)
        mi = months.index(mo) if mo in months else -1
        if mi >= 0:
            return f"{full_months[mi]} {yr}"
    return None


def _safe_float(val) -> float:
    try:
        return round(float(str(val).replace(',', '').replace(' ', '') or 0), 2)
    except (ValueError, TypeError):
        return 0.0


def _find_col(headers: list, *keywords) -> int:
    """Find column index by keyword match. Prefers exact matches over substring matches."""
    # First pass: exact match
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        if any(k == hl for k in keywords):
            return i
    # Second pass: substring match
    for i, h in enumerate(headers):
        hl = h.lower().strip()
        if any(k in hl for k in keywords):
            return i
    return -1


# ── COGS Breakdown Parser ─────────────────────────────────────────────────

def _parse_cogs_breakdown(rows: list) -> list:
    """
    Parse COGS Breakdown sheet from 1C accounting export.
    Supports multiple column layouts:
      - Layout A (original): columns 6, 7310, 8230 + classification in last col
      - Layout B (new):      mixed account columns + classification in last non-empty col
    The parser auto-detects the header row (may not be row 0 if title rows exist)
    and dynamically finds cost account columns by their numeric header labels.
    """
    cogs_items = []
    if not rows or len(rows) < 2:
        return cogs_items

    # ── Step 1: Find the actual header row (scan first 15 rows for "Субконто") ─
    # The real header row has "Субконто" as a standalone cell value (column name),
    # NOT embedded in descriptive text like "Детализация по субконто: ..."
    header_row_idx = 0
    for ri, row in enumerate(rows[:15]):
        for c in row:
            cell_val = str(c or '').strip().lower()
            # Must be a standalone "субконто" (column header), not a description
            if cell_val == 'субконто':
                header_row_idx = ri
                break
        if header_row_idx == ri and header_row_idx > 0:
            break

    headers = [str(h or '').strip() for h in rows[header_row_idx]]

    # ── Step 2: Find product column ─────────────────────────────────────────
    product_idx = 0
    for i, h in enumerate(headers):
        if 'субконто' in h.lower():
            product_idx = i
            break

    # ── Step 3: Find "Кред. оборот" (Credit turnover = actual COGS amount) ──
    # This is the most reliable single column for total COGS per product.
    # Fallback: sum specific account columns (6, 7310, 8230, etc.)
    credit_turnover_idx = -1
    for i, h in enumerate(headers):
        h_lower = h.lower().strip()
        if 'кред' in h_lower and 'оборот' in h_lower:
            credit_turnover_idx = i
            break

    # ── Step 4: Find specific account columns (6, 7310, 8230, 1610, etc.) ───
    # These are sub-columns under "Кред. оборот" in 1C export format
    col6_idx = -1       # Account 6 (Revenue/COGS account)
    col7310_idx = -1    # Account 7310 (selling expenses)
    col8230_idx = -1    # Account 8230 (other expenses)
    col1610_idx = -1    # Account 1610 (inventory — used in some layouts)

    # Scan headers after "Кред. оборот" for sub-account columns
    for i, h in enumerate(headers):
        h_stripped = h.strip()
        if h_stripped == '6':
            col6_idx = i
        elif h_stripped == '7310':
            col7310_idx = i
        elif h_stripped == '8230':
            col8230_idx = i
        elif h_stripped == '1610':
            # There may be two 1610 columns (debit/credit); use the one after Кред. оборот
            if credit_turnover_idx >= 0 and i > credit_turnover_idx:
                col1610_idx = i
            elif col1610_idx < 0:
                col1610_idx = i

    # ── Step 5: Find classification column (last meaningful column) ──────────
    # Usually the rightmost column contains "COGS Wholesale", "COGS Retail", etc.
    class_idx = -1
    for i in range(len(headers) - 1, product_idx, -1):
        h = headers[i].strip().lower()
        if h in ('q', '') or 'cogs' in h or 'wholesale' in h or 'retail' in h:
            class_idx = i
            break
    # If not found by keyword, use the very last column
    if class_idx < 0:
        class_idx = len(headers) - 1

    # ── Step 6: Determine COGS value strategy ───────────────────────────────
    # The COGS Breakdown is "Обороты счета 1610" (turnovers of inventory account).
    # Credit turnover = ALL inventory outflows, which includes:
    #   - Cr 1610 → Dr 6/7310/8230: actual COGS (goods sold → expense accounts)
    #   - Cr 1610 → Dr 1610: INTERNAL transfers between warehouses (NOT COGS!)
    # We must EXCLUDE the 1610→1610 internal transfers from COGS totals.
    #
    # Strategy A: credit_turnover minus 1610 internal transfers
    # Strategy B: Sum specific account columns (6 + 7310 + 8230) — already excludes 1610
    # Strategy C: Fallback to debit turnover
    use_credit_turnover = credit_turnover_idx >= 0
    has_old_layout = col6_idx >= 0 and (col7310_idx >= 0 or col8230_idx >= 0)

    # Detect credit-side 1610 sub-column (for subtracting internal transfers)
    # col1610_idx after the loop points to credit-side 1610 if it exists after Кред. оборот
    has_credit_1610 = col1610_idx >= 0 and credit_turnover_idx >= 0 and col1610_idx > credit_turnover_idx

    logger.info(f"COGS parser: header_row={header_row_idx}, product_col={product_idx}, "
                f"credit_turnover_col={credit_turnover_idx}, col6={col6_idx}, "
                f"col7310={col7310_idx}, col8230={col8230_idx}, col1610={col1610_idx}, "
                f"has_credit_1610={has_credit_1610}, "
                f"class_col={class_idx}, use_credit_turnover={use_credit_turnover}")

    # ── Step 7: Parse data rows ─────────────────────────────────────────────
    data_rows = rows[header_row_idx + 1:]
    for row_offset, row in enumerate(data_rows):
        if not row or product_idx >= len(row):
            continue
        product = str(row[product_idx] or '').strip()
        if not product or product.lower() in ('итого', 'итог', 'total', ''):
            continue
        # Skip title/description rows (no numeric data)
        if all(not _safe_float(row[j]) for j in range(1, min(len(row), 6))):
            continue

        # Extract COGS value
        if use_credit_turnover and not has_old_layout:
            # Strategy A: Use credit turnover MINUS 1610→1610 internal transfers
            # Credit turnover includes ALL outflows; 1610 sub-column = warehouse transfers (not COGS)
            total = abs(_safe_float(row[credit_turnover_idx]) if credit_turnover_idx < len(row) else 0.0)
            if has_credit_1610:
                internal_1610 = abs(_safe_float(row[col1610_idx]) if col1610_idx < len(row) else 0.0)
                total = max(0, round(total - internal_1610, 2))
            k_val = abs(_safe_float(row[col6_idx]) if col6_idx >= 0 and col6_idx < len(row) else 0.0)
            l_val = 0.0
            o_val = 0.0
        elif has_old_layout:
            # Strategy B: Sum account columns 6 + 7310 + 8230 (original layout)
            k_val = abs(_safe_float(row[col6_idx]) if col6_idx >= 0 and col6_idx < len(row) else 0.0)
            l_val = abs(_safe_float(row[col7310_idx]) if col7310_idx >= 0 and col7310_idx < len(row) else 0.0)
            o_val = abs(_safe_float(row[col8230_idx]) if col8230_idx >= 0 and col8230_idx < len(row) else 0.0)
            total = round(k_val + l_val + o_val, 2)
        else:
            # Strategy C: No recognizable layout — try credit turnover or debit оборот
            deb_idx = -1
            for i, h in enumerate(headers):
                if 'деб' in h.lower().strip() and 'оборот' in h.lower().strip():
                    deb_idx = i
                    break
            total = abs(_safe_float(row[deb_idx]) if deb_idx >= 0 and deb_idx < len(row) else 0.0)
            k_val = total
            l_val = 0.0
            o_val = 0.0

        if total == 0:
            continue

        # Extract classification from the last column
        classification = ""
        if class_idx >= 0 and class_idx < len(row):
            classification = str(row[class_idx] or '').strip()
        segment, category = _classify_cogs_product(product, classification)

        cogs_row_num = header_row_idx + 1 + row_offset + 1  # 1-based
        cogs_items.append({
            "product": product,
            "product_en": get_english_name(product),
            "col6": k_val,
            "col7310": l_val,
            "col8230": o_val,
            "total_cogs": total,
            "segment": segment,
            "category": category,
            "_lineage": {
                "source_sheet": "COGS Breakdown",
                "source_row": cogs_row_num,
                "classification_rule": "product_dimension" if category != "Other COGS" else "classification_col",
                "confidence": 0.95 if category != "Other COGS" else 0.6,
            },
        })

    logger.info(f"COGS parser: extracted {len(cogs_items)} items")
    return cogs_items


# ── G&A and D&A Expense Extractors ─────────────────────────────────────────

def _extract_ga_expenses(transactions: list) -> list:
    """
    Extract G&A expense items from parsed transactions.
    Groups by Account Dr code and sums amounts for GA_ACCOUNT_CODES.
    D&A codes are now EXCLUDED — use _extract_da_expenses() separately.
    """
    ga_by_account = {}
    for txn in transactions:
        acct_dr = str(txn.get("acct_dr", "")).strip()
        if _is_ga_account(acct_dr):
            if acct_dr not in ga_by_account:
                ga_by_account[acct_dr] = 0.0
            ga_by_account[acct_dr] += abs(float(txn.get("amount", 0)))

    return [
        {
            "account_code": code,
            "account_name": GA_ACCOUNT_NAMES.get(code, f"G&A ({code})"),
            "amount": round(amount, 2),
        }
        for code, amount in ga_by_account.items()
    ]


def _extract_da_expenses(transactions: list) -> list:
    """
    Extract D&A (Depreciation & Amortization) items from parsed transactions.
    Groups by Account Dr code and sums amounts for DA_ACCOUNT_CODES.
    Separated from G&A for proper P&L waterfall (EBITDA → D&A → EBIT).
    """
    da_by_account = {}
    for txn in transactions:
        acct_dr = str(txn.get("acct_dr", "")).strip()
        if _is_da_account(acct_dr):
            if acct_dr not in da_by_account:
                da_by_account[acct_dr] = 0.0
            da_by_account[acct_dr] += abs(float(txn.get("amount", 0)))

    return [
        {
            "account_code": code,
            "account_name": DA_ACCOUNT_NAMES.get(code, f"D&A ({code})"),
            "amount": round(amount, 2),
        }
        for code, amount in da_by_account.items()
    ]


# ══════════════════════════════════════════════════════════════════════════
# TDSheet (Trial Balance) Parser
# ══════════════════════════════════════════════════════════════════════════

# Account-code to G&A sub-category mapping (from Mapping sheet analysis)
_GA_SUBCATEGORY_MAP = {
    "7310":      "Selling Expenses (Total)",
    "7310.01":   "Selling Expenses (General)",
    "7310.01.1": "Circulation Expenses (Non-ENVD)",           # Was incorrectly "Bank Commissions"
    "7310.01.99":"Non-deductible Expenses (Commercial)",
    "7310.02":   "Commercial Expenses in Production",
    "7310.02.1": "Commercial Expenses in Production (Non-ENVD)",  # Was "Fuel Transportation" (too narrow)
    "7410":      "General Admin Expenses",
    "7410.01":   "Depreciation & Amortization",
    "7410.01.1": "Depreciation (PPE)",
    "8110":      "Non-operating Income",
    "8220":      "Non-operating Expenses",
    "8220.01.1": "Non-operating Expenses (Non-ENVD)",
    "8220.01.2": "Non-operating Expenses (Distributable)",
    "8230":      "Shortages & Losses from Spoilage",
    "9210":      "Other P&L Account",
}

def _parse_tdsheet(rows: list) -> dict:
    """
    Parse TDSheet (Оборотно-сальдовая ведомость / Trial Balance) from 1C export.
    Extracts: all account turnovers, G&A, D&A, Finance, Tax, Labour.
    """
    # Step 1: Find header row (rows 6-7 typically, merged)
    header_row_idx = 0
    for ri, row in enumerate(rows[:10]):
        row_text = ' '.join(str(c or '').lower() for c in row)
        if 'счет' in row_text or 'оборот' in row_text and 'дебет' in row_text:
            header_row_idx = ri
            break
        if any(str(c or '').strip().lower() in ('код', 'наименование') for c in row):
            header_row_idx = ri
            break

    # Step 2: Map columns by scanning header area
    # TDSheet has merged headers across rows 6-7. Column positions are fixed:
    # C(2)=Code, D(3)=Name, E(4)=SubAcct1, F(5)=SubAcct2,
    # G(6)=OpenDr, H(7)=OpenCr, K(10)=TurnDr, N(13)=TurnCr,
    # O(14)=CloseDr, Q(16)=CloseCr, T(19)=NetPL
    # But we detect dynamically by scanning headers
    code_idx = 2    # Column C
    name_idx = 3    # Column D
    sub1_idx = 4    # Column E
    sub2_idx = 5    # Column F

    # Try to detect column positions from header keywords
    for ri in range(max(0, header_row_idx - 1), min(len(rows), header_row_idx + 3)):
        if ri >= len(rows):
            break
        for ci, cell in enumerate(rows[ri]):
            cv = str(cell or '').strip().lower()
            if cv in ('код', 'code'):
                code_idx = ci
            elif cv in ('наименование', 'name'):
                name_idx = ci

    # For turnover columns, we need to find them by header keywords
    # Typical layout: G=Open Dr, H=Open Cr, (I,J skipped), K=Turn Dr, (L,M skipped), N=Turn Cr, O=Close Dr, (P skipped), Q=Close Cr
    open_dr_idx = 6
    open_cr_idx = 7
    turn_dr_idx = 10
    turn_cr_idx = 13
    close_dr_idx = 14
    close_cr_idx = 16
    net_pl_idx = 19

    # Scan headers to find turnover columns more reliably
    for ri in range(max(0, header_row_idx - 1), min(len(rows), header_row_idx + 3)):
        if ri >= len(rows):
            break
        for ci, cell in enumerate(rows[ri]):
            cv = str(cell or '').strip().lower()
            if 'сальдо на начало' in cv or 'сальдо на нач' in cv:
                open_dr_idx = ci
                open_cr_idx = ci + 1
            elif 'оборот за период' in cv or 'обороты за период' in cv:
                turn_dr_idx = ci
            elif 'сальдо на конец' in cv or 'сальдо на кон' in cv:
                close_dr_idx = ci
                close_cr_idx = ci + 2  # Usually has a gap column

    # Step 3: Parse data rows
    data_start = header_row_idx + 2  # Skip header + sub-header row
    trial_balance_items = []
    ga_expenses = []
    da_expenses = []
    da_detail = {}  # Deduplicate D&A by code (TDSheet may have duplicate rows)
    finance_income = 0.0
    finance_expense = 0.0
    tax_expense = 0.0
    labour_costs = 0.0
    total_revenue_tb = 0.0
    total_cogs_tb = 0.0
    total_inventory_credit_tb = 0.0  # 16xx (inventory) credit turnover for COGS reconciliation

    # G&A detail accumulator (for Mapping-style sub-categories)
    ga_detail = {}

    for row in rows[data_start:]:
        if not row or len(row) < 5:
            continue

        code = str(row[code_idx] if code_idx < len(row) else '').strip()
        name = str(row[name_idx] if name_idx < len(row) else '').strip()
        sub1 = str(row[sub1_idx] if sub1_idx < len(row) else '').strip()

        # Skip empty rows and grand total
        if not code and not name and not sub1:
            continue
        if name.lower() in ('итого', 'итог', 'total'):
            continue

        # Determine hierarchy level
        level = 1
        if not code and sub1:
            level = 3  # Detail row (counterparty/station)
        elif '.' in code:
            level = 2  # Sub-account
        elif code:
            level = 1  # Parent account

        # Extract numeric values
        o_dr = _safe_float(row[open_dr_idx] if open_dr_idx < len(row) else 0)
        o_cr = _safe_float(row[open_cr_idx] if open_cr_idx < len(row) else 0)
        t_dr = _safe_float(row[turn_dr_idx] if turn_dr_idx < len(row) else 0)
        t_cr = _safe_float(row[turn_cr_idx] if turn_cr_idx < len(row) else 0)
        c_dr = _safe_float(row[close_dr_idx] if close_dr_idx < len(row) else 0)
        c_cr = _safe_float(row[close_cr_idx] if close_cr_idx < len(row) else 0)
        net_pl = _safe_float(row[net_pl_idx] if net_pl_idx < len(row) else 0)

        # Skip rows with no data
        if o_dr == 0 and o_cr == 0 and t_dr == 0 and t_cr == 0 and c_dr == 0 and c_cr == 0:
            continue

        # Determine account class
        acct_class = ''
        effective_code = code if code else ''
        if effective_code:
            acct_class = effective_code[0] if effective_code[0].isdigit() else ''

        item = {
            "account_code": effective_code,
            "account_name": name,
            "sub_account_detail": sub1,
            "opening_debit": o_dr,
            "opening_credit": o_cr,
            "turnover_debit": t_dr,
            "turnover_credit": t_cr,
            "closing_debit": c_dr,
            "closing_credit": c_cr,
            "net_pl_impact": net_pl,
            "account_class": acct_class,
            "hierarchy_level": level,
        }
        trial_balance_items.append(item)

        # ── Extract P&L components from parent/sub-account turnovers ──
        # Only process accounts that have a code (not detail counterparty rows)
        if not effective_code or level == 3:
            continue

        # Revenue (6xxx accounts) — credit turnover
        if effective_code.startswith('6') and level <= 2:
            if t_cr > 0:
                total_revenue_tb += t_cr

        # COGS (71xx accounts) — debit turnover
        # Only capture level 1 (parent summary) to avoid double-counting with children
        if effective_code.startswith('71') and level == 1:
            if t_dr > 0:
                total_cogs_tb += t_dr

        # Inventory (16xx accounts) — credit turnover (outflow from warehouse → COGS)
        # Only capture account 1610 specifically (merchandise), not all 16xx
        # Level 1 only to avoid double-counting parent+child (1610 vs 1610.01)
        if effective_code == '1610' and level == 1:
            if t_cr > 0:
                total_inventory_credit_tb += t_cr

        # Labour (72xx accounts) — net P&L impact → included in G&A
        if effective_code.startswith('72') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            labour_costs += amt
            if amt > 10 and '.' in effective_code:
                ga_detail[effective_code] = {
                    "account_code": effective_code,
                    "account_name": name or f"Labour Costs ({effective_code})",
                    "amount": round(amt, 2),
                }

        # Selling Expenses (73xx accounts) — these are G&A
        # Only capture leaf-level sub-accounts (e.g. 7310.01.1, 7310.02.1) not parent summaries
        # Use net_pl (DR - CR) as the correct expense amount, not gross t_dr
        if effective_code.startswith('73') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            # Skip parent accounts with tiny amounts (they're just summary headers)
            if amt > 10 and '.' in effective_code:
                subcat = _GA_SUBCATEGORY_MAP.get(effective_code, f"Selling Expenses ({effective_code})")
                ga_detail[effective_code] = {"account_code": effective_code, "account_name": subcat, "amount": round(amt, 2)}

        # D&A (7410.xx accounts) — use net_pl for correct amounts
        # Capture all 7410 sub-accounts; parent filtering will remove summaries later
        # Use da_detail dict to deduplicate (TDSheet may have duplicate rows per subconto)
        if effective_code.startswith('7410') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            if amt > 0 and ('.' in effective_code or '/' in effective_code):
                da_detail[effective_code] = {
                    "account_code": effective_code,
                    "account_name": name or "Depreciation & Amortization",
                    "amount": round(amt, 2),
                }

        # Other Admin Expenses (74xx but not 7410.xx)
        if effective_code.startswith('74') and not effective_code.startswith('7410') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            if amt > 10:
                ga_detail[effective_code] = {"account_code": effective_code, "account_name": f"Admin Expenses ({effective_code})", "amount": round(amt, 2)}

        # Finance Expense (75xx accounts) — use net_pl for accuracy
        if effective_code.startswith('75') and level <= 2:
            finance_expense += abs(net_pl) if net_pl else abs(t_dr)

        # Finance Income (76xx accounts) — use net_pl for accuracy
        if effective_code.startswith('76') and level <= 2:
            finance_income += abs(net_pl) if net_pl else abs(t_cr)

        # Tax (77xx accounts) — use net_pl for accuracy
        if effective_code.startswith('77') and level <= 2:
            tax_expense += abs(net_pl) if net_pl else abs(t_dr)

        # Non-operating Income (81xx accounts) — separate from finance_income
        # These are FX gains, asset sales, etc. — NOT finance income (76xx)
        if effective_code.startswith('81') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_cr)
            if amt > 0 and '.' in effective_code:
                ga_detail[f'NOI:{effective_code}'] = {
                    "account_code": f'NOI:{effective_code}',
                    "account_name": name or f"Non-operating Income ({effective_code})",
                    "amount": round(amt, 2),
                }

        # Non-operating Expenses (82xx, 83xx accounts) — use net_pl, only leaf sub-accounts
        if (effective_code.startswith('82') or effective_code.startswith('83')) and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            if amt > 10 and '.' in effective_code:
                ga_detail[effective_code] = {"account_code": effective_code, "account_name": _GA_SUBCATEGORY_MAP.get(effective_code, f"Non-operating Expenses ({effective_code})"), "amount": round(amt, 2)}

        # Other P&L (92xx accounts) — leaf only
        if effective_code.startswith('92') and level <= 2:
            amt = abs(net_pl) if net_pl else abs(t_dr)
            if amt > 10 and '.' in effective_code:
                ga_detail[effective_code] = {"account_code": effective_code, "account_name": f"Other P&L ({effective_code})", "amount": round(amt, 2)}

    # Build parent codes from ALL TB items to detect parent-child via both '.' and '/'
    parent_codes = _build_parent_codes(trial_balance_items)

    # Deduplicate G&A: remove parent accounts whose children are also captured
    # Check both '.' and '/' separators (1C uses '/' for subconto hierarchy)
    ga_codes = sorted(ga_detail.keys())
    ga_leaf_codes = set(ga_codes)
    for code in ga_codes:
        # Strip NOI: prefix for parent check
        raw_code = code[4:] if code.startswith('NOI:') else code
        # If this code is a parent of any TB item, discard it
        if raw_code in parent_codes:
            ga_leaf_codes.discard(code)
            continue
        # Also check if any other GA code starts with this code + '.' or '/'
        for other in ga_codes:
            if other != code and (other.startswith(code + '.') or other.startswith(code + '/')):
                ga_leaf_codes.discard(code)
                break
    ga_expenses = [ga_detail[c] for c in sorted(ga_leaf_codes)]

    # Convert da_detail dict to da_expenses list, filtering parent codes
    da_expenses = []
    for da_code in sorted(da_detail.keys()):
        if da_code not in parent_codes:
            da_expenses.append(da_detail[da_code])
    # If all were filtered (unlikely), fall back to original da_detail values
    if not da_expenses and da_detail:
        # If all were filtered as parents, use them as-is (fallback)
        da_expenses = list(da_detail.values())

    logger.info(f"TDSheet parser: {len(trial_balance_items)} items, "
                f"{len(ga_expenses)} G&A, {len(da_expenses)} D&A, "
                f"finance_inc={finance_income:,.0f}, finance_exp={finance_expense:,.0f}, "
                f"tax={tax_expense:,.0f}, labour={labour_costs:,.0f}, "
                f"rev_tb={total_revenue_tb:,.0f}, cogs_tb={total_cogs_tb:,.0f}, "
                f"inv_credit_tb={total_inventory_credit_tb:,.0f}")

    return {
        "trial_balance_items": trial_balance_items,
        "ga_expenses": ga_expenses,
        "da_expenses": da_expenses,
        "finance_income": round(finance_income, 2),
        "finance_expense": round(finance_expense, 2),
        "tax_expense": round(tax_expense, 2),
        "labour_costs": round(labour_costs, 2),
        "total_revenue_tb": round(total_revenue_tb, 2),
        "total_cogs_tb": round(total_cogs_tb, 2),
        "total_inventory_credit_tb": round(total_inventory_credit_tb, 2),
    }


# ══════════════════════════════════════════════════════════════════════════
# Balance Sheet Parser (MAPPING GRP / IFRS classification)
# ══════════════════════════════════════════════════════════════════════════

def _parse_balance_sheet(rows: list) -> dict:
    """
    Parse "Balance" sheet — enriched trial balance with IFRS MAPPING GRP column.
    Returns balance_sheet_items and aggregated BS by IFRS line.
    """
    if not rows or len(rows) < 2:
        return {"balance_sheet_items": [], "bs_aggregated": {}}

    # Detect header row by looking for key columns
    header_idx = 0
    for ri, row in enumerate(rows[:5]):
        row_text = ' '.join(str(c or '').lower() for c in row)
        if 'mapping' in row_text or 'code' in row_text and 'name' in row_text:
            header_idx = ri
            break

    raw_headers = [str(h or '').strip().lower() for h in rows[header_idx]]

    # Map columns dynamically
    def _fc(keywords):
        for i, h in enumerate(raw_headers):
            if any(k in h for k in keywords):
                return i
        return -1

    code_i    = _fc(['code'])
    name_i    = _fc(['name'])
    start_dr_i= _fc(['start dr', 'start_dr'])
    start_cr_i= _fc(['start cr', 'start_cr'])
    start_bal_i= _fc(['starting balance', 'starting_balance', 'start balance'])
    turn_dr_i = _fc(['turnover dr', 'turnover_dr'])
    turn_cr_i = _fc(['turnover cr', 'turnover_cr'])
    turn_bal_i= _fc(['turnover balance', 'turnover_balance'])
    end_dr_i  = _fc(['end dr', 'end_dr'])
    end_cr_i  = _fc(['end cr', 'end_cr'])
    end_bal_i = _fc(['end balance', 'end_balance'])
    type_i    = _fc(['type'])
    map_grp_i = _fc(['mapping grp', 'mapping_grp'])
    map_st_i  = _fc(['maping st', 'mapping_st', 'maping_st'])
    map_baku_i = _fc(['maping baku', 'mapping baku', 'baku'])
    ic_entity_i = _fc(['intercompany', 'ic entity', 'ic_entity'])

    # If "MAPPING GRP" not found by keyword, try exact column position (P=15 in 0-indexed)
    if map_grp_i < 0:
        for i, h in enumerate(raw_headers):
            if 'mapping' in h and 'grp' in h:
                map_grp_i = i
                break
    if map_grp_i < 0 and len(raw_headers) > 15:
        # Try column P (index 15) as fallback
        if 'mapping' in raw_headers[15] or raw_headers[15] == '':
            map_grp_i = 15

    balance_sheet_items = []
    bs_aggregated = {}

    for row in rows[header_idx + 1:]:
        if not row or len(row) < 5:
            continue

        code = str(row[code_i] if code_i >= 0 and code_i < len(row) else '').strip()
        name = str(row[name_i] if name_i >= 0 and name_i < len(row) else '').strip()
        if not code and not name:
            continue

        ifrs_line = str(row[map_grp_i] if map_grp_i >= 0 and map_grp_i < len(row) else '').strip()
        ifrs_stmt = str(row[map_st_i] if map_st_i >= 0 and map_st_i < len(row) else '').strip()
        row_type = str(row[type_i] if type_i >= 0 and type_i < len(row) else '').strip()

        opening_bal = _safe_float(row[start_bal_i] if start_bal_i >= 0 and start_bal_i < len(row) else 0)
        turn_dr = _safe_float(row[turn_dr_i] if turn_dr_i >= 0 and turn_dr_i < len(row) else 0)
        turn_cr = _safe_float(row[turn_cr_i] if turn_cr_i >= 0 and turn_cr_i < len(row) else 0)
        closing_bal = _safe_float(row[end_bal_i] if end_bal_i >= 0 and end_bal_i < len(row) else 0)

        baku_mapping = str(row[map_baku_i] if map_baku_i >= 0 and map_baku_i < len(row) else '').strip()
        ic_entity = str(row[ic_entity_i] if ic_entity_i >= 0 and ic_entity_i < len(row) else '').strip()

        # Normalize ifrs_statement: fix garbage values "0", "" from source Excel
        # Account classes 1-5 = Balance Sheet, 6-9 = Profit & Loss (Income Statement)
        clean_code = code.replace('X', '').replace('x', '')
        first_digit = clean_code[:1] if clean_code else ''
        if ifrs_stmt.upper() in ('BS', 'PL', 'P&L', 'IS'):
            normalized_stmt = 'BS' if ifrs_stmt.upper() == 'BS' else 'PL'
        elif first_digit in ('1', '2', '3', '4', '5'):
            normalized_stmt = 'BS'
        elif first_digit in ('6', '7', '8', '9'):
            normalized_stmt = 'PL'
        else:
            normalized_stmt = ifrs_stmt  # Preserve original if unknown

        item = {
            "account_code": code,
            "account_name": name,
            "ifrs_line_item": ifrs_line,
            "ifrs_statement": normalized_stmt,
            "baku_bs_mapping": baku_mapping,
            "intercompany_entity": ic_entity,
            "opening_balance": opening_bal,
            "turnover_debit": turn_dr,
            "turnover_credit": turn_cr,
            "closing_balance": closing_bal,
            "row_type": row_type,
        }

        # Filter logic for BalanceSheetItem storage:
        # - BS summary rows (class 1-5, row_type='სხვა'): ALWAYS store — needed for BS
        # - BS counterparty detail (class 1-5, with /N codes): SKIP — massive duplication
        # - P&L items (class 6-9) WITH baku_bs_mapping: STORE — needed for MR expense sub-breakdowns
        #   These carry category tags (Payroll, Depreciation, Rent, etc.) that the MR engine
        #   uses to distribute expenses into P&L sub-items.
        # - P&L items WITHOUT baku_bs_mapping: SKIP — no useful classification
        is_pl_class = first_digit in ('6', '7', '8', '9')
        is_counterparty_bs = ('/' in code and row_type == 'საქვეანგარიშგებო' and not is_pl_class)

        if is_pl_class:
            # P&L item: store all P&L rows so they are available for MR/diagnostics.
            # If baku mapping is missing, keep the item but mark it and log for later mapping.
            if not baku_mapping or baku_mapping in ('0', 'nan', 'NaN', 'None', ''):
                item["unmapped_pl"] = True
                logger.debug(f"Parsing BS: P&L row without baku mapping (code={code}, name={name})")
            balance_sheet_items.append(item)
        elif not is_counterparty_bs:
            # BS item (class 1-5): store summary rows, skip counterparty detail
            balance_sheet_items.append(item)

        # Aggregate by IFRS line item for BS rows
        # Track both summary and detail totals so callers can prefer summary
        # when present and otherwise fall back to summing detail rows.
        # (This avoids losing values that only exist in detail rows.)
        if ifrs_line and ifrs_stmt.upper() == 'BS':
            # Use two buckets: summary_rows (row_type == 'სხვა') and detail_rows (others)
            # We'll finalize bs_aggregated after parsing all rows below.
            # Use keys with suffix to avoid nested dicts in loop.
            key_summary = f"__summary__::{ifrs_line}"
            key_detail = f"__detail__::{ifrs_line}"
            bs_aggregated[key_summary] = bs_aggregated.get(key_summary, 0) + (closing_bal if row_type == 'სხვა' else 0)
            bs_aggregated[key_detail] = bs_aggregated.get(key_detail, 0) + (closing_bal if row_type != 'სხვა' else 0)

    # Finalize bs_aggregated: prefer summary totals when present, else fall back to detail totals
    final_bs = {}
    # collect unique lines from the summary/detail keys
    seen_lines = set()
    for k in list(bs_aggregated.keys()):
        if k.startswith("__summary__::") or k.startswith("__detail__::"):
            _, line = k.split("::", 1)
            seen_lines.add(line)

    for line in seen_lines:
        summary_val = bs_aggregated.get(f"__summary__::{line}", 0.0)
        detail_val = bs_aggregated.get(f"__detail__::{line}", 0.0)
        if abs(summary_val) > 0.0001:
            final_bs[line] = final_bs.get(line, 0) + summary_val
        else:
            final_bs[line] = final_bs.get(line, 0) + detail_val

    logger.info(f"Balance sheet parser: {len(balance_sheet_items)} items (filtered), "
                f"{len(final_bs)} IFRS BS line items (aggregated)")
    return {"balance_sheet_items": balance_sheet_items, "bs_aggregated": final_bs}


# ══════════════════════════════════════════════════════════════════════════
# BS Sheet Parser (Pre-formatted IFRS Statement of Financial Position)
# ══════════════════════════════════════════════════════════════════════════

# IFRS BS line → section mapping
_BS_SECTION_MAP = {
    "Property, plant and equipment":     ("noncurrent_assets", 1),
    "PPE Cost":                          ("noncurrent_assets", 2),
    "PPE Depreciation":                  ("noncurrent_assets", 2),
    "Right of use asset":                ("noncurrent_assets", 1),
    "Investment properties":             ("noncurrent_assets", 1),
    "Investments":                       ("noncurrent_assets", 1),
    "Intangible assets":                 ("noncurrent_assets", 1),
    "Intangible Assets COST":            ("noncurrent_assets", 2),
    "Intangible Assets Amortisations":   ("noncurrent_assets", 2),
    "Trade receivables LT":              ("noncurrent_assets", 1),
    "Inventories":                       ("current_assets", 1),
    "Trade receivables":                 ("current_assets", 1),
    "Tax assets":                        ("current_assets", 1),
    "Prepayments and other receivables": ("current_assets", 1),
    "Short term loans receivable":       ("current_assets", 1),
    "Cash and cash equivalents":         ("current_assets", 1),
    "Lease liability non current":       ("noncurrent_liabilities", 1),
    "Government Grants non current":     ("noncurrent_liabilities", 1),
    "Short-term loans and borrowings":   ("current_liabilities", 1),
    "Other taxes payable":               ("current_liabilities", 1),
    "Trade and other payables":          ("current_liabilities", 1),
    "Advances received":                 ("current_liabilities", 1),
    "Government grant liability":        ("current_liabilities", 1),
    "Long-Term Loans Payable":           ("noncurrent_liabilities", 1),
    "Lease liability":                   ("current_liabilities", 1),
    "Government grants":                 ("current_liabilities", 1),
    "Share capital":                     ("equity", 1),
    "Additional Paid-in Capital":        ("equity", 1),
    "Unpaid Capital":                    ("equity", 1),
    "Revaluation reserve":               ("equity", 1),
    "Retained earnings":                 ("equity", 1),
    "Net income for the Period":         ("equity", 1),
}

def _parse_bs_sheet(rows: list) -> dict:
    """
    Parse the BS sheet — pre-formatted IFRS Statement of Financial Position.
    Simple 2-column format: line item name | amount.
    """
    items = []
    sections = {
        "noncurrent_assets": [],
        "current_assets": [],
        "noncurrent_liabilities": [],
        "current_liabilities": [],
        "equity": [],
    }
    totals = {}
    current_section = None
    net_income = 0.0

    for row in rows:
        if not row or len(row) < 1:
            continue
        label = str(row[0] or '').strip()
        amount = _safe_float(row[1] if len(row) > 1 else 0)

        if not label:
            continue

        label_lower = label.lower()

        # Detect section headers
        if 'non-current assets' in label_lower or 'non current assets' in label_lower:
            current_section = 'noncurrent_assets'
            continue
        elif 'current assets' in label_lower and 'non' not in label_lower and 'total' not in label_lower:
            current_section = 'current_assets'
            continue
        elif 'non-current liabilities' in label_lower or 'non current liabilities' in label_lower:
            current_section = 'noncurrent_liabilities'
            continue
        elif 'current liabilities' in label_lower and 'non' not in label_lower and 'total' not in label_lower:
            current_section = 'current_liabilities'
            continue
        elif label_lower == 'equity':
            current_section = 'equity'
            continue

        # Detect totals
        if 'total non-current assets' in label_lower or 'total non current assets' in label_lower:
            totals['total_noncurrent_assets'] = amount
            continue
        elif 'total current assets' in label_lower and 'non' not in label_lower:
            totals['total_current_assets'] = amount
            continue
        elif 'total assets' in label_lower:
            totals['total_assets'] = amount
            continue
        elif 'total non-current liabilities' in label_lower or 'total non current liabilities' in label_lower:
            totals['total_noncurrent_liabilities'] = amount
            continue
        elif 'total short-term liabilities' in label_lower or 'total current liabilities' in label_lower:
            totals['total_current_liabilities'] = amount
            continue
        elif 'total liabilities' in label_lower and 'equity' not in label_lower:
            totals['total_liabilities'] = amount
            continue
        elif 'total equity' in label_lower:
            totals['total_equity'] = amount
            continue
        elif 'total liabilities and equity' in label_lower:
            totals['total_liabilities_equity'] = amount
            continue
        elif label_lower in ('check', 'statement of financial position', 'sgp'):
            continue

        # Net income special handling
        if 'net income' in label_lower:
            net_income = amount

        item = {"label": label, "amount": amount, "section": current_section}
        items.append(item)

        # Add to section
        section_info = _BS_SECTION_MAP.get(label)
        if section_info and section_info[0] in sections:
            sections[section_info[0]].append({"label": label, "amount": amount, "level": section_info[1]})
        elif current_section and current_section in sections:
            sections[current_section].append({"label": label, "amount": amount, "level": 1})

    logger.info(f"BS sheet parser: {len(items)} items, net_income={net_income:,.2f}, "
                f"total_assets={totals.get('total_assets', 0):,.2f}")
    return {"items": items, "sections": sections, "totals": totals, "net_income": net_income}


# ══════════════════════════════════════════════════════════════════════════
# Mapping Sheet Parser
# ══════════════════════════════════════════════════════════════════════════

def _parse_mapping_sheet(rows: list) -> list:
    """
    Parse the Mapping sheet — account code → IFRS category + sub-category detail.
    Format: B=Account Code, C=Account Name, D=Amount, E=MAPPING GRP, F=Sub-category
    """
    mappings = []
    for row in rows:
        if not row or len(row) < 3:
            continue
        # Mapping sheet starts at column B (index 1)
        code = str(row[1] if len(row) > 1 else '').strip()
        name = str(row[2] if len(row) > 2 else '').strip()
        amount = _safe_float(row[3] if len(row) > 3 else 0)
        mapping_grp = str(row[4] if len(row) > 4 else '').strip()
        sub_category = str(row[5] if len(row) > 5 else '').strip()

        if not code and not name:
            continue

        mappings.append({
            "account_code": code,
            "account_name": name,
            "amount": amount,
            "mapping_grp": mapping_grp,
            "sub_category": sub_category,
        })

    logger.info(f"Mapping sheet parser: {len(mappings)} entries")
    return mappings


# ── FALLBACK: Derive Revenue/COGS from Transaction Ledger via COA ──────────

def _derive_revenue_from_transactions(transactions: list) -> tuple:
    """
    When no Revenue Breakdown sheet exists, derive revenue from Transaction Ledger.
    Uses Account Cr codes starting with 6xx (Georgian COA revenue accounts).
    Returns (revenue_items, source_description).
    """
    revenue_items = []
    revenue_by_code = {}

    for txn in transactions:
        acct_cr = str(txn.get("acct_cr", "")).strip()
        if not acct_cr:
            continue
        cr_map = map_coa(acct_cr)
        if cr_map and cr_map.get("side") == "income":
            amt = abs(float(txn.get("amount", 0)))
            if amt == 0:
                continue
            prefix = cr_map.get("prefix", "6")
            label = cr_map.get("pl", "Revenue")
            segment = cr_map.get("segment", "Other")
            key = f"{prefix}_{label}"
            if key not in revenue_by_code:
                revenue_by_code[key] = {
                    "product": label,
                    "product_en": label,
                    "gross": 0.0, "vat": 0.0, "net": 0.0,
                    "segment": f"Revenue {segment}" if segment != "Other" else "Other Revenue",
                    "category": "Other Revenue",
                }
            revenue_by_code[key]["net"] += amt
            revenue_by_code[key]["gross"] += amt  # no VAT split available from ledger

    for item in revenue_by_code.values():
        item["net"] = round(item["net"], 2)
        item["gross"] = round(item["gross"], 2)
        revenue_items.append(item)

    return revenue_items, "Derived from Transaction Ledger (Account Cr 6xx)"


def _derive_cogs_from_transactions(transactions: list) -> tuple:
    """
    When no COGS Breakdown sheet exists, derive COGS from Transaction Ledger.
    Uses Account Dr codes starting with 71x (Georgian COA cost of sales).
    Returns (cogs_items, source_description).
    """
    cogs_items = []
    cogs_by_code = {}

    for txn in transactions:
        acct_dr = str(txn.get("acct_dr", "")).strip()
        if not acct_dr:
            continue
        dr_map = map_coa(acct_dr)
        if dr_map and dr_map.get("pl_line") == "COGS":
            amt = abs(float(txn.get("amount", 0)))
            if amt == 0:
                continue
            prefix = dr_map.get("prefix", "71")
            label = dr_map.get("pl", "Cost of Sales")
            key = f"{prefix}_{label}"
            if key not in cogs_by_code:
                cogs_by_code[key] = {
                    "product": label,
                    "product_en": label,
                    "col6": 0.0, "col7310": 0.0, "col8230": 0.0,
                    "total_cogs": 0.0,
                    "segment": "Other COGS",
                    "category": "Other COGS",
                }
            cogs_by_code[key]["total_cogs"] += amt

    for item in cogs_by_code.values():
        item["total_cogs"] = round(item["total_cogs"], 2)
        cogs_items.append(item)

    return cogs_items, "Derived from Transaction Ledger (Account Dr 71x)"


# ══════════════════════════════════════════════════════════════════════════════
# LLM CLASSIFICATION FALLBACK — Called when weighted scoring fails
# ══════════════════════════════════════════════════════════════════════════════

# Module-level cache for LLM classification results (avoid re-classifying same sheet)
_llm_classify_cache: dict = {}


def _llm_classify_sheet(rows: list, sheet_name: str, scores: dict) -> dict:
    """Synchronous wrapper for LLM sheet classification.

    Returns dict: {"sheet_type": str, "confidence": float, "reasoning": str}
    or None if LLM is unavailable.

    Uses asyncio to call the DataAgent's async classify_unknown_sheet() method.
    Results are cached by sheet_name to avoid duplicate API calls.
    """
    cache_key = f"{sheet_name}:{len(rows)}"
    if cache_key in _llm_classify_cache:
        return _llm_classify_cache[cache_key]

    try:
        import asyncio
        from app.agents.registry import registry

        data_agent = registry.get("data")
        if not data_agent:
            logger.debug("DataAgent not registered — skipping LLM classification")
            return None

        # Check if we're already in an async context
        try:
            loop = asyncio.get_running_loop()
            # We're inside an async context — can't use asyncio.run()
            # Instead, we'll do a synchronous LLM call via the data agent's method
            # This should only happen during upload (which is already async)
            # The caller should use the async version directly
            logger.debug("LLM classify: inside async loop — deferring to async caller")
            return None
        except RuntimeError:
            # No running loop — we can create one
            result = asyncio.run(data_agent.classify_unknown_sheet(rows, sheet_name, scores))
            result_dict = {
                "sheet_type": result.sheet_type,
                "confidence": result.confidence,
                "reasoning": result.reasoning,
                "column_mapping": result.column_mapping,
            }
            _llm_classify_cache[cache_key] = result_dict
            return result_dict

    except Exception as e:
        logger.warning(f"LLM classification unavailable: {e}")
        return None


async def classify_sheet_async(rows: list, sheet_name: str, scores: dict) -> dict:
    """Async version of LLM sheet classification.

    Call this from async contexts (like the upload endpoint) when
    weighted scoring fails.
    """
    cache_key = f"{sheet_name}:{len(rows)}"
    if cache_key in _llm_classify_cache:
        return _llm_classify_cache[cache_key]

    try:
        from app.agents.registry import registry
        data_agent = registry.get("data")
        if not data_agent:
            return None

        result = await data_agent.classify_unknown_sheet(rows, sheet_name, scores)
        result_dict = {
            "sheet_type": result.sheet_type,
            "confidence": result.confidence,
            "reasoning": result.reasoning,
            "column_mapping": result.column_mapping,
        }
        _llm_classify_cache[cache_key] = result_dict
        return result_dict

    except Exception as e:
        logger.warning(f"Async LLM classification failed: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# CONTENT-BASED SHEET DETECTION — Structural fingerprinting
# ══════════════════════════════════════════════════════════════════════════════

def _detect_trial_balance(rows: list) -> float:
    """Detect if sheet is a Trial Balance (TDSheet) by content structure. Returns confidence 0.0-1.0."""
    if len(rows) < 10:
        return 0.0
    score = 0.0

    # 1. Account codes (weight 0.3): Look for 4+ digit account codes in first 200 data rows
    code_count = 0
    for row in rows[5:min(205, len(rows))]:  # Skip header area
        for cell in row[:5]:  # Check first 5 columns
            clean = re.sub(r'[^0-9.]', '', str(cell or ''))
            digits_only = re.sub(r'[^0-9]', '', clean)
            if len(digits_only) >= 4 and digits_only[:4].isdigit():
                code_val = int(digits_only[:4])
                if 1000 <= code_val <= 9999:
                    code_count += 1
                    break
    if code_count >= 20:
        score += 0.3
    elif code_count >= 10:
        score += 0.15

    # 2. Debit/Credit column pairs (weight 0.3): Header keywords in first 10 rows
    dc_keywords_ru = ['дебет', 'кредит', 'оборот', 'сальдо']
    dc_keywords_en = ['debit', 'credit', 'turnover', 'balance']
    dc_keywords_ge = ['დებეტ', 'კრედიტ']
    all_dc = dc_keywords_ru + dc_keywords_en + dc_keywords_ge
    dc_found = 0
    for row in rows[:10]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        for kw in all_dc:
            if kw in row_text:
                dc_found += 1
    if dc_found >= 3:
        score += 0.3
    elif dc_found >= 2:
        score += 0.2

    # 3. Russian/Georgian TB-specific keywords (weight 0.2)
    tb_keywords = ['оборотно', 'сальдовая', 'ведомость', 'счет', 'საცდელი', 'ბალანსი',
                   'trial balance', 'начальное', 'конечное']
    kw_found = 0
    for row in rows[:8]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        for kw in tb_keywords:
            if kw in row_text:
                kw_found += 1
    if kw_found >= 2:
        score += 0.2
    elif kw_found >= 1:
        score += 0.1

    # 4. Column count (weight 0.1): 1C TB exports typically have 15-25 columns
    max_cols = max((len(r) for r in rows[:10]), default=0)
    if 14 <= max_cols <= 26:
        score += 0.1

    # 5. Row count (weight 0.1): TBs typically have 100+ data rows
    if len(rows) >= 100:
        score += 0.1
    elif len(rows) >= 50:
        score += 0.05

    return min(score, 1.0)


def _detect_balance_mapped(rows: list) -> float:
    """Detect if sheet is an IFRS-mapped Balance sheet (with MAPPING GRP). Returns confidence 0.0-1.0."""
    if len(rows) < 5:
        return 0.0
    score = 0.0

    # 1. MAPPING GRP column (weight 0.5)
    for row in rows[:5]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        if 'mapping' in row_text and 'grp' in row_text:
            score += 0.5
            break

    # 2. MAPING ST column with BS/IS values (weight 0.3)
    for row in rows[:5]:
        for cell in row:
            cell_str = str(cell or '').lower()
            if 'maping' in cell_str and 'st' in cell_str:
                score += 0.3
                break
        if score >= 0.8:
            break
    if score < 0.8:
        # Check if data rows contain BS/IS values
        bs_is_count = 0
        for row in rows[1:min(50, len(rows))]:
            for cell in row:
                cv = str(cell or '').strip()
                if cv in ('BS', 'IS'):
                    bs_is_count += 1
        if bs_is_count >= 10:
            score += 0.2

    # 3. Row type column (weight 0.2): Georgian row types
    ge_types = ['სხვა', 'საქვეანგარიშგებო']
    ge_found = 0
    for row in rows[1:min(50, len(rows))]:
        for cell in row:
            if str(cell or '').strip() in ge_types:
                ge_found += 1
    if ge_found >= 5:
        score += 0.2

    return min(score, 1.0)


def _detect_cogs_breakdown(rows: list) -> float:
    """Detect if sheet is a COGS Breakdown (1C export). Returns confidence 0.0-1.0."""
    if len(rows) < 5:
        return 0.0
    score = 0.0

    # 1. "Субконто" keyword (weight 0.4)
    for row in rows[:15]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        if 'субконто' in row_text:
            score += 0.4
            break

    # 2. Account 1610 reference (weight 0.3): COGS inventory account
    for row in rows[:8]:
        row_text = ' '.join(str(c or '') for c in row)
        if '1610' in row_text:
            score += 0.3
            break

    # 3. Turnover columns (weight 0.3): "Оборот" in headers
    for row in rows[:10]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        if 'оборот' in row_text:
            score += 0.3
            break

    return min(score, 1.0)


def _detect_revenue_breakdown(rows: list, headers: list) -> float:
    """Detect if sheet is a Revenue Breakdown. Returns confidence 0.0-1.0."""
    if len(rows) < 3:
        return 0.0
    score = 0.0
    headers_lower = [h.lower() for h in headers]

    # 1. Product/Amount columns (weight 0.5): existing detection pattern
    has_product = any('product' in h or 'პროდუქტი' in h or 'продукт' in h or 'наименование' in h for h in headers_lower)
    has_net = any('net' in h or 'წმინდა' in h or 'нетто' in h for h in headers_lower)
    has_amount = any('amount' in h or 'gel' in h or 'თანხა' in h or 'сумма' in h for h in headers_lower)
    if has_product and (has_net or has_amount):
        score += 0.5

    # 2. VAT column (weight 0.3)
    has_vat = any('vat' in h or 'ვატ' in h or 'ддс' in h or 'დღგ' in h for h in headers_lower)
    if has_vat:
        score += 0.3

    # 3. Revenue keywords in first rows (weight 0.2)
    rev_keywords = ['შემოსავალი', 'revenue', 'выручка', 'реализация', 'გაყიდვ']
    for row in rows[:5]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        if any(kw in row_text for kw in rev_keywords):
            score += 0.2
            break

    return min(score, 1.0)


def _detect_pl_extract(rows: list) -> float:
    """Detect if sheet is a P&L Account Summary / Extract (single net amount per account).
    Common 1C export format: account_code + account_name + single net_amount column.
    Returns confidence 0.0-1.0.
    """
    if len(rows) < 10:
        return 0.0
    score = 0.0

    # 1. Account codes (weight 0.3): 4+ digit codes in classes 1-9, found in first 5 columns
    code_count = 0
    for row in rows[1:min(101, len(rows))]:
        for cell in row[:5]:
            raw = str(cell or '').strip()
            if not raw:
                continue
            # Strip X placeholders, dots, slashes to get digits
            digits = re.sub(r'[^0-9]', '', raw.split('.')[0].split('/')[0].replace('X', '').replace('x', ''))
            if len(digits) >= 4 and digits[:4].isdigit():
                code_val = int(digits[:4])
                if 1000 <= code_val <= 9999:
                    code_count += 1
                    break
    if code_count >= 15:
        score += 0.3
    elif code_count >= 8:
        score += 0.15

    # 2. Few columns (weight 0.25): P&L extract has 3-5 columns, NOT 15-25 like TDSheet
    max_cols = max((len(r) for r in rows[:10]), default=0)
    if 3 <= max_cols <= 6:
        score += 0.25
    elif max_cols <= 8:
        score += 0.1

    # 3. NO debit/credit header keywords (weight 0.2): distinguishes from TDSheet
    dc_keywords = ['дебет', 'кредит', 'debit', 'credit', 'დებეტ', 'კრედიტ',
                    'оборот', 'turnover', 'сальдо', 'balance']
    dc_found = 0
    for row in rows[:10]:
        row_text = ' '.join(str(c or '').lower() for c in row)
        for kw in dc_keywords:
            if kw in row_text:
                dc_found += 1
    if dc_found == 0:
        score += 0.2

    # 4. Single numeric column pattern (weight 0.15): most rows have exactly 1 numeric value
    single_num_rows = 0
    total_data_rows = 0
    for row in rows[1:min(50, len(rows))]:
        if not row or not any(str(c or '').strip() for c in row):
            continue
        total_data_rows += 1
        num_count = 0
        for cell in row:
            v = _safe_float(cell)
            if v != 0:
                num_count += 1
        if num_count == 1:
            single_num_rows += 1
    if total_data_rows > 0 and single_num_rows / total_data_rows >= 0.6:
        score += 0.15
    elif total_data_rows > 0 and single_num_rows / total_data_rows >= 0.4:
        score += 0.07

    # 5. Row count 30-300 (weight 0.1): typical P&L extract size
    if 30 <= len(rows) <= 300:
        score += 0.1
    elif len(rows) >= 15:
        score += 0.05

    return min(score, 1.0)


def _parse_pl_extract(rows: list, sheet_name: str) -> dict:
    """Parse P&L Account Summary / Extract into TrialBalanceItem-compatible dicts.
    Input: rows with [empty?, account_code, account_name, net_amount] structure.
    Output: same shape as _parse_tdsheet() return value.
    """
    # Step 1: Detect column positions by scanning first 10 data rows
    code_idx = -1
    name_idx = -1
    amount_idx = -1

    # Find code column: first column with 4+ digit account codes
    for ci in range(min(5, max((len(r) for r in rows[:10]), default=0))):
        code_hits = 0
        for row in rows[1:min(20, len(rows))]:
            if ci >= len(row):
                continue
            raw = str(row[ci] or '').strip()
            if not raw:
                continue
            digits = re.sub(r'[^0-9]', '', raw.split('.')[0].split('/')[0].replace('X', '').replace('x', ''))
            if len(digits) >= 4 and digits[:1].isdigit():
                val = int(digits[:4])
                if 1000 <= val <= 9999:
                    code_hits += 1
        if code_hits >= 5:
            code_idx = ci
            break

    if code_idx < 0:
        logger.warning(f"P&L Extract '{sheet_name}': could not find account code column")
        return {"trial_balance_items": [], "ga_expenses": [], "da_expenses": [],
                "finance_income": 0, "finance_expense": 0, "tax_expense": 0,
                "labour_costs": 0, "total_cogs_tb": 0, "total_inventory_credit_tb": 0}

    # Find amount column: last column with numeric values
    max_data_cols = max((len(r) for r in rows[:20]), default=0)
    for ci in range(max_data_cols - 1, code_idx, -1):
        num_hits = 0
        for row in rows[1:min(30, len(rows))]:
            if ci >= len(row):
                continue
            v = _safe_float(row[ci])
            if v != 0:
                num_hits += 1
        if num_hits >= 5:
            amount_idx = ci
            break

    if amount_idx < 0:
        logger.warning(f"P&L Extract '{sheet_name}': could not find amount column")
        return {"trial_balance_items": [], "ga_expenses": [], "da_expenses": [],
                "finance_income": 0, "finance_expense": 0, "tax_expense": 0,
                "labour_costs": 0, "total_cogs_tb": 0, "total_inventory_credit_tb": 0}

    # Name column: between code and amount
    name_idx = code_idx + 1 if code_idx + 1 < amount_idx else -1

    logger.info(f"P&L Extract '{sheet_name}': code_col={code_idx}, name_col={name_idx}, amount_col={amount_idx}")

    # Step 2: Parse data rows
    trial_balance_items = []
    ga_expenses = []
    da_expenses = []
    finance_income = 0.0
    finance_expense = 0.0
    tax_expense = 0.0
    labour_costs = 0.0
    total_cogs_tb = 0.0

    # Labour keywords
    labour_kw = ['ხელფასი', 'salary', 'salaries', 'wages', 'зарплат', 'оплата труд',
                 'შრომის ანაზღაურება', 'პერსონალ']
    # D&A keywords
    da_kw = ['ცვეთა', 'deprec', 'amort', 'амортиз', 'износ']

    for row in rows[1:]:
        if not row or code_idx >= len(row):
            continue

        code = str(row[code_idx] or '').strip()
        if not code:
            continue

        # Skip summary placeholders (61XX, 81XX, etc.)
        if 'X' in code.upper():
            continue

        # Skip total rows
        name = str(row[name_idx] or '').strip() if name_idx >= 0 and name_idx < len(row) else ''
        if code.lower() in ('итого', 'итог', 'total') or name.lower() in ('итого', 'итог', 'total'):
            continue

        amount = _safe_float(row[amount_idx]) if amount_idx < len(row) else 0.0

        # Determine hierarchy level
        level = 1  # parent
        if '.' in code or '/' in code:
            level = 2  # sub-account

        # Extract root digits for account class
        digits_only = re.sub(r'[^0-9]', '', code.split('.')[0].split('/')[0])
        acct_class = digits_only[0] if digits_only and digits_only[0].isdigit() else ''

        # Map net amount to turnover fields
        # Convention: negative = income (credit), positive = expense (debit)
        if amount < 0:
            t_dr = 0.0
            t_cr = abs(amount)
        else:
            t_dr = amount
            t_cr = 0.0

        item = {
            "account_code": code,
            "account_name": name,
            "sub_account_detail": "",
            "opening_debit": 0.0,
            "opening_credit": 0.0,
            "turnover_debit": t_dr,
            "turnover_credit": t_cr,
            "closing_debit": 0.0,
            "closing_credit": 0.0,
            "net_pl_impact": amount,
            "account_class": acct_class,
            "hierarchy_level": level,
        }
        trial_balance_items.append(item)

        # ── Sub-account depreciation detection (level 2) ──
        # The Mapping sheet has column F (sub_category) with explicit classification
        # like "Depreciation". Sub-accounts like 7310.02.1/1, 7410.01/1 contain the
        # actual depreciation amounts that the parent-level accounts don't break out.
        # We must scan these BEFORE the `level != 1` skip to capture depreciation.
        if level == 2 and digits_only:
            sub_two_digit = digits_only[:2] if len(digits_only) >= 2 else ''
            # Check column F (sub_category at index 5) for depreciation classification
            sub_category_f = str(row[5] if len(row) > 5 and row[5] else '').strip()
            sub_name_lower = name.lower()
            # Depreciation: from column F ("Depreciation") or name keywords
            is_depreciation = (
                (sub_category_f and 'depreci' in sub_category_f.lower()) or
                (sub_category_f and 'amortiz' in sub_category_f.lower()) or
                (sub_category_f and 'ამორტიზ' in sub_category_f.lower()) or
                any(kw in sub_name_lower for kw in da_kw)
            )
            if is_depreciation and sub_two_digit in ('73', '74') and abs(amount) > 0:
                da_expenses.append({
                    "account_code": code,
                    "account_name": name or sub_category_f,
                    "amount": abs(amount),
                    "_source": "sub_account_col_f",
                })

        # Skip further extraction for non-parent or non-coded rows
        if not digits_only or level != 1:
            continue

        two_digit = digits_only[:2] if len(digits_only) >= 2 else ''
        four_digit = digits_only[:4] if len(digits_only) >= 4 else ''
        name_lower = name.lower()

        # COGS (71xx) — accumulate for reconciliation
        if two_digit == '71':
            total_cogs_tb += t_dr

        # G&A: Selling (73xx) and Admin (74xx) — store as ga_expenses
        if two_digit in ('73', '74'):
            ga_expenses.append({
                "account_code": code,
                "account_name": name,
                "amount": abs(amount),
            })
            # Check for D&A within selling/admin (parent-level name check)
            if any(kw in name_lower for kw in da_kw):
                da_expenses.append({
                    "account_code": code,
                    "account_name": name,
                    "amount": abs(amount),
                })

        # Non-operating income (8110) — stored as ga_expense with NOI: prefix
        # so build_income_statement routes it to other_income (NOT finance_income)
        if four_digit == '8110':
            ga_expenses.append({
                "account_code": f"NOI:{code}",
                "account_name": name or "Non-operating Income",
                "amount": abs(t_cr) if t_cr else abs(amount),
            })

        # Non-operating expenses (8220, 8230) — stored as ga_expense
        # so build_income_statement routes it to other_expense (NOT finance_expense)
        if four_digit in ('8220', '8230'):
            ga_expenses.append({
                "account_code": code,
                "account_name": name or "Non-operating Expenses",
                "amount": abs(t_dr) if t_dr else abs(amount),
            })

        # Tax expense (9210) — stored as ga_expense
        # so build_income_statement routes it to tax_expense
        if four_digit == '9210':
            ga_expenses.append({
                "account_code": code,
                "account_name": name or "Other P&L",
                "amount": abs(t_dr) if t_dr else abs(amount),
            })

        # Labour costs
        if any(kw in name_lower for kw in labour_kw):
            labour_costs += abs(amount)

    logger.info(f"P&L Extract '{sheet_name}': parsed {len(trial_balance_items)} TB items, "
                f"{len(ga_expenses)} GA, COGS={total_cogs_tb:,.0f}, "
                f"FinIncome={finance_income:,.0f}, FinExp={finance_expense:,.0f}")

    return {
        "trial_balance_items": trial_balance_items,
        "ga_expenses": ga_expenses,
        "da_expenses": da_expenses,
        "finance_income": finance_income,
        "finance_expense": finance_expense,
        "tax_expense": tax_expense,
        "labour_costs": labour_costs,
        "total_cogs_tb": total_cogs_tb,
        "total_inventory_credit_tb": 0.0,  # Not available in P&L extract
    }


# ══════════════════════════════════════════════════════════════════════════════
# AUTO-GENERATE BS FROM TDSHEET — Using GEORGIAN_COA mapping
# ══════════════════════════════════════════════════════════════════════════════

def _build_parent_codes(tb_items: list) -> set:
    """Build set of ALL ancestor codes from TB items for leaf detection.
    For code '8220.01.1/1', generates parents: '8220', '8220.01', '8220.01.1'.
    This prevents double-counting when both parent totals and child details exist.
    """
    parent_codes = set()
    for item in tb_items:
        code = item.get('account_code', '')
        # Generate ALL ancestor prefixes by finding each separator
        for i, ch in enumerate(code):
            if ch in ('.', '/'):
                parent_codes.add(code[:i])
    return parent_codes


def _generate_bs_from_tdsheet(tb_items: list) -> list:
    """Generate BalanceSheetItem-compatible dicts from TDSheet closing balances.
    Uses map_coa() → GEORGIAN_COA to map account codes (1xxx-5xxx) to IFRS BS line items.
    Only processes leaf accounts to avoid double-counting parent+child.
    """
    if not tb_items:
        return []

    # Build set of ALL ancestor codes to identify leaf accounts
    parent_codes = _build_parent_codes(tb_items)

    # Also detect summary rows: 1C uses "XX", "X" suffixes for aggregate rows
    # e.g., "21XX" = total for all 21xx accounts, "141X" = total for 141x
    # These are NOT leaf accounts — they double-count if included with their children
    summary_codes = set()
    for item in tb_items:
        code = item.get('account_code', '')
        if not code:
            continue
        # Codes containing X are summary placeholders (e.g., "21XX", "53XX", "141X", "129X")
        if 'X' in code.upper():
            summary_codes.add(code)

    # Filter to leaf accounts only
    leaf_items = [i for i in tb_items if i.get('account_code', '') and
                  i.get('account_code', '') not in parent_codes and
                  i.get('account_code', '') not in summary_codes and
                  i.get('hierarchy_level', 0) != 3]  # Also skip detail rows (counterparty/station level)

    # Build name lookup for parent name inheritance (child /1 rows often have empty names)
    name_lookup = {}
    for item in tb_items:
        c = item.get('account_code', '')
        n = (item.get('account_name') or '').strip()
        if c and n:
            name_lookup[c] = n

    bs_items = []
    for item in leaf_items:
        code = item.get('account_code', '')
        if not code:
            continue

        coa = map_coa(code)
        if not coa or 'bs' not in coa:
            continue  # Skip P&L accounts (6xxx-9xxx) — they have 'pl' not 'bs'

        # Net closing balance
        c_dr = item.get('closing_debit') or 0
        c_cr = item.get('closing_credit') or 0
        bs_side = coa.get('bs_side', '')

        # Sign convention:
        # Assets (normal): DR balance = positive → closing = c_dr - c_cr
        # Assets (contra, e.g. accum. depreciation): CR balance → closing = c_dr - c_cr (naturally negative)
        # Liabilities/Equity: CR balance = positive → closing = c_cr - c_dr
        if bs_side in ('liability', 'equity'):
            closing = c_cr - c_dr
        else:
            # Assets (both normal and contra) use DR - CR
            # Contra-assets naturally get negative values (CR > DR)
            closing = c_dr - c_cr

        # Opening balance with same convention
        o_dr = item.get('opening_debit') or 0
        o_cr = item.get('opening_credit') or 0
        if bs_side in ('liability', 'equity'):
            opening = o_cr - o_dr
        else:
            opening = o_dr - o_cr

        ifrs_line = coa['bs']  # e.g. "Cash & Equivalents", "Fixed Assets (PP&E)"
        ifrs_statement = 'BS'  # All class 1-5 accounts are BS by definition

        # Inherit parent name if child has empty name
        acct_name = (item.get('account_name') or '').strip()
        if not acct_name:
            if '/' in code:
                acct_name = name_lookup.get(code.rsplit('/', 1)[0], '')
            if not acct_name and '.' in code:
                acct_name = name_lookup.get(code.rsplit('.', 1)[0], '')
            if not acct_name:
                acct_name = coa.get('name_ka', '') or coa.get('name_ru', '') or code

        bs_items.append({
            'account_code': code,
            'account_name': acct_name,
            'ifrs_line_item': ifrs_line,
            'ifrs_statement': ifrs_statement,
            'opening_balance': round(opening, 2),
            'turnover_debit': item.get('turnover_debit') or 0,
            'turnover_credit': item.get('turnover_credit') or 0,
            'closing_balance': round(closing, 2),
            'row_type': 'COA_DERIVED',  # Mark as auto-generated
            'currency': 'GEL',
            'period': '',
        })

    logger.info(f"Auto-generated {len(bs_items)} BalanceSheetItem records from TDSheet via GEORGIAN_COA mapping")
    return bs_items


def _generate_pl_from_tdsheet(tb_items: list) -> dict:
    """Generate P&L structure from TDSheet class 6-9 accounts using map_coa().

    Returns dict with:
      - revenue_items: list of revenue line dicts (class 6)
      - cogs_items: list of COGS line dicts (71xx)
      - ga_expenses: list of G&A expense dicts (72xx-77xx, 8xx, 9xx)
      - pl_summary: aggregated totals
    """
    if not tb_items:
        return {}

    # Build set of ALL ancestor codes to identify leaf accounts
    parent_codes = _build_parent_codes(tb_items)
    summary_codes = set()
    for item in tb_items:
        code = item.get('account_code', '')
        if 'X' in code.upper():
            summary_codes.add(code)

    # Build name lookup for parent name inheritance (child /1 rows often have empty names)
    name_lookup = {}
    for item in tb_items:
        code = item.get('account_code', '')
        name = (item.get('account_name') or '').strip()
        if code and name:
            name_lookup[code] = name

    revenue_items = []
    cogs_items = []
    ga_expenses = []
    totals = {'revenue': 0, 'cogs': 0, 'sga': 0, 'da': 0, 'finance_income': 0, 'finance_expense': 0, 'tax': 0, 'other': 0}

    for item in tb_items:
        code = item.get('account_code', '')
        if not code or code in parent_codes or code in summary_codes:
            continue
        if (item.get('hierarchy_level', 0) or 0) >= 3:
            continue

        coa = map_coa(code)
        if not coa or 'pl' not in coa:
            continue

        # For P&L: turnover amounts are the key figures
        # Revenue (class 6): credit-side = revenue earned → net = credit - debit
        # Expenses (class 7-9): debit-side = expenses incurred → net = debit - credit
        t_dr = item.get('turnover_debit') or 0
        t_cr = item.get('turnover_credit') or 0
        side = coa.get('side', 'expense')
        pl_line = coa.get('pl_line', '')
        pl_name = coa.get('pl', '')
        acct_name = (item.get('account_name') or '').strip()
        # Inherit parent name if child has empty name (e.g. 6110/1 inherits from 6110)
        if not acct_name:
            if '/' in code:
                parent_code = code.rsplit('/', 1)[0]
                acct_name = name_lookup.get(parent_code, '')
            if not acct_name and '.' in code:
                parent_code = code.rsplit('.', 1)[0]
                acct_name = name_lookup.get(parent_code, '')
            if not acct_name:
                # Use COA master name if available
                acct_name = coa.get('name_ka', '') or coa.get('name_ru', '') or code

        if side == 'income':
            net = t_cr - t_dr  # Revenue: credit balance = positive
            # Determine if this is operating revenue (class 6) or non-operating income (class 8+)
            acct_class = code[:1] if code else ''
            if pl_line == 'Finance':
                totals['finance_income'] += net
            elif acct_class == '6':
                # Class 6 = Operating Revenue → goes into revenue_items
                revenue_items.append({
                    'product': acct_name, 'account_code': code,
                    'gross': round(net, 2), 'vat': 0, 'net': round(net, 2),
                    'segment': 'TB-Derived', 'category': pl_name,
                    'pl_line': pl_line,
                })
                totals['revenue'] += net
            else:
                # Class 8+ income accounts = Non-operating Income → stored as GA item
                # Prefixed with NOI: so income_statement.py can separate from G&A
                ga_expenses.append({
                    'category': 'Other Income', 'account_code': f'NOI:{code}',
                    'account_name': acct_name, 'amount': round(net, 2),  # positive = income amount
                    'ga_type': 'OTHER_INCOME',
                })
                totals['other_income'] = totals.get('other_income', 0) + net
        else:
            net = t_dr - t_cr  # Expense: debit balance = positive
            is_contra = coa.get('contra_revenue', False)
            if is_contra:
                # Contra-revenue (discounts/returns) reduces revenue
                revenue_items.append({
                    'product': acct_name, 'account_code': code,
                    'gross': round(-net, 2), 'vat': 0, 'net': round(-net, 2),
                    'segment': 'TB-Derived', 'category': 'Discounts & Returns',
                    'pl_line': pl_line,
                })
                totals['revenue'] -= net
            elif pl_line == 'COGS':
                cogs_items.append({
                    'product': acct_name, 'account_code': code,
                    'total_cogs': round(net, 2),
                    'segment': 'TB-Derived', 'category': pl_name,
                })
                totals['cogs'] += net
            elif pl_line == 'Finance':
                ga_expenses.append({
                    'category': 'Finance Expense', 'account_code': code,
                    'account_name': acct_name, 'amount': round(net, 2),
                    'ga_type': 'FINANCE_EXPENSE',
                })
                totals['finance_expense'] += net
            elif pl_line == 'Tax':
                ga_expenses.append({
                    'category': 'Income Tax', 'account_code': code,
                    'account_name': acct_name, 'amount': round(net, 2),
                    'ga_type': 'TAX_EXPENSE',
                })
                totals['tax'] += net
            elif pl_line == 'DA':
                ga_expenses.append({
                    'category': pl_name or 'General & Admin', 'account_code': code,
                    'account_name': acct_name, 'amount': round(net, 2),
                    'ga_type': 'DA_EXPENSE',
                })
                totals['da'] += net
            elif pl_line == 'SGA':
                ga_expenses.append({
                    'category': pl_name or 'SGA Expense', 'account_code': code,
                    'account_name': acct_name, 'amount': round(net, 2),
                    'ga_type': 'SGA_EXPENSE',
                })
                totals['sga'] += net
            else:
                ga_expenses.append({
                    'category': pl_name or 'Other Expense', 'account_code': code,
                    'account_name': acct_name, 'amount': round(net, 2),
                    'ga_type': 'OTHER_EXPENSE',
                })
                totals['other'] += net

    other_income = totals.get('other_income', 0)
    gross_profit = totals['revenue'] - totals['cogs']
    ebitda = gross_profit - totals['sga'] - totals['other']
    ebit = ebitda - totals['da']
    # Other income (non-operating: FX gains, asset sales, etc.) goes below EBIT
    ebt = ebit + other_income + totals['finance_income'] - totals['finance_expense']
    net_profit = ebt - totals['tax']

    pl_summary = {
        'revenue': round(totals['revenue'], 2),
        'cogs': round(totals['cogs'], 2),
        'gross_profit': round(gross_profit, 2),
        'sga': round(totals['sga'], 2),
        'da': round(totals['da'], 2),
        'ebitda': round(ebitda, 2),
        'ebit': round(ebit, 2),
        'other_income': round(other_income, 2),
        'finance_income': round(totals['finance_income'], 2),
        'finance_expense': round(totals['finance_expense'], 2),
        'ebt': round(ebt, 2),
        'tax': round(totals['tax'], 2),
        'net_profit': round(net_profit, 2),
        'other': round(totals['other'], 2),
    }

    logger.info(f"Auto-generated P&L from TDSheet: {len(revenue_items)} revenue, {len(cogs_items)} COGS, {len(ga_expenses)} expenses. Net profit: {net_profit:,.0f}")
    return {
        'revenue_items': revenue_items,
        'cogs_items': cogs_items,
        'ga_expenses': ga_expenses,
        'pl_summary': pl_summary,
    }


# ══════════════════════════════════════════════════════════════════════════
# COGS ↔ Inventory Reconciliation (3-way cross-check)
# ══════════════════════════════════════════════════════════════════════════

def _reconcile_cogs(cogs_items: list, total_cogs_tb: float, total_inventory_credit_tb: float) -> dict:
    """
    Cross-validate COGS from independent sources.

    After the 1610→1610 internal transfer fix, COGS Breakdown only includes
    actual cost outflows (Cr 1610 → Dr 6/7310/8230), excluding internal
    warehouse-to-warehouse transfers.

    TB 71xx captures account 7110 debit turnovers from the Trial Balance.
    These two should be close — both measure actual cost of goods sold.

    1610 total credit is informational — shows total inventory movement
    (including internal transfers that are NOT COGS).

    Checks:
      1. PRIMARY: COGS Breakdown ↔ TB 71xx (both measure actual COGS)
      2. INFO: 1610 total credit vs COGS (shows how much was internal transfers)
    """
    cogs_breakdown_total = sum(c.get('total_cogs', 0) for c in cogs_items)
    checks = []

    # ── PRIMARY CHECK: COGS Breakdown vs TB 71xx ──
    # After excluding 1610→1610 internal transfers, COGS Breakdown should closely
    # match TB 71xx debits — both represent actual COGS recognized in the period.
    # Small differences are normal (rounding, timing, minor account class differences).
    if cogs_breakdown_total > 0 and total_cogs_tb > 0:
        var = abs(cogs_breakdown_total - total_cogs_tb)
        pct = (var / max(cogs_breakdown_total, total_cogs_tb)) * 100
        checks.append({
            "check": "COGS Breakdown vs TB 71xx",
            "check_ka": "COGS Breakdown vs ბრუნვა 71xx",
            "source_a": {"label": "COGS Breakdown Sheet", "label_ka": "COGS Breakdown ფურცელი", "value": round(cogs_breakdown_total, 2)},
            "source_b": {"label": "TB 71xx Debit", "label_ka": "ბრუნვა 71xx დებეტი", "value": round(total_cogs_tb, 2)},
            "variance": round(var, 2),
            "variance_pct": round(pct, 2),
            "severity": "info" if pct < 2 else "warning" if pct < 5 else "critical",
            "status": "match" if pct < 2 else "mismatch",
        })

    # ── INFO CHECK: 1610 total credit (total inventory movement) ──
    # 1610 credit includes ALL inventory outflows: actual COGS + internal transfers.
    # Show what fraction of total movement was actual COGS vs internal transfers.
    if cogs_breakdown_total > 0 and total_inventory_credit_tb > 0:
        cogs_pct_of_movement = (cogs_breakdown_total / total_inventory_credit_tb) * 100
        internal_transfers = round(total_inventory_credit_tb - cogs_breakdown_total, 2)
        checks.append({
            "check": "1610 Inventory Movement",
            "check_ka": "1610 მარაგის მოძრაობა",
            "source_a": {"label": "COGS (actual sales)", "label_ka": "COGS (რეალური გაყიდვა)", "value": round(cogs_breakdown_total, 2)},
            "source_b": {"label": "1610 Total Credit", "label_ka": "1610 ჯამური კრედიტი", "value": round(total_inventory_credit_tb, 2)},
            "variance": internal_transfers,
            "variance_pct": round(cogs_pct_of_movement, 2),
            "severity": "info",
            "status": "component",
            "is_component_check": True,
            "note": f"COGS = {cogs_pct_of_movement:.1f}% of total 1610 credit. Internal transfers (1610→1610) = {100-cogs_pct_of_movement:.1f}%",
            "note_ka": f"COGS = 1610 კრედიტის {cogs_pct_of_movement:.1f}%. შიდა ტრანსფერები (1610→1610) = {100-cogs_pct_of_movement:.1f}%",
        })

    has_mismatch = any(c["status"] == "mismatch" for c in checks)
    severity_order = {"info": 0, "warning": 1, "critical": 2}
    worst_severity = max((c["severity"] for c in checks), key=lambda s: severity_order.get(s, 0), default="info") if checks else "info"

    if checks:
        logger.info(f"COGS Reconciliation: {len(checks)} checks, "
                    f"{'MISMATCH' if has_mismatch else 'ALL MATCH'} "
                    f"(breakdown={cogs_breakdown_total:,.0f}, tb_71xx={total_cogs_tb:,.0f}, inv_16xx={total_inventory_credit_tb:,.0f})")

    return {
        "checks": checks,
        "has_mismatch": has_mismatch,
        "worst_severity": worst_severity,
        "sources": {
            "cogs_breakdown_total": round(cogs_breakdown_total, 2),
            "total_cogs_tb": round(total_cogs_tb, 2),
            "total_inventory_credit_tb": round(total_inventory_credit_tb, 2),
        }
    }


def parse_file(filename: str, content: bytes, strict: bool = True) -> dict:
    """
    Parse an Excel or CSV file into structured financial data.
    Returns: {transactions, revenue, budget, cogs_items, ga_expenses, file_type, period}
    """
    transactions = []
    revenue = []
    budget = {}
    cogs_items = []
    ga_expenses = []
    da_expenses = []
    trial_balance_items = []
    balance_sheet_items = []
    bs_preformatted = {}
    mapping_items = []
    finance_income = 0.0
    finance_expense = 0.0
    tax_expense = 0.0
    labour_costs = 0.0
    total_cogs_tb = 0.0
    total_inventory_credit_tb = 0.0
    file_type = "Financial Data"
    # ── Processing pipeline tracking ──────────────────────────────────
    pipeline_steps = []
    detected_sheets_info = []
    _semantic_stats = {}

    ext = filename.lower().rsplit('.', 1)[-1]

    if ext in ('xlsx', 'xls', 'xlsm'):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = {name: wb[name] for name in wb.sheetnames}
    elif ext == 'csv':
        decoded = content.decode('utf-8', errors='replace')
        reader = list(csv.reader(io.StringIO(decoded)))
        sheets = {"Sheet1": reader}
    else:
        return {"error": f"Unsupported file type: {ext}"}

    file_size_kb = round(len(content) / 1024, 1)
    pipeline_steps.append({"step": "upload", "detail": f"{filename} uploaded ({file_size_kb} KB, {len(sheets)} sheets)"})

    for sheet_name, sheet in sheets.items():
        if ext in ('xlsx', 'xls', 'xlsm'):
            rows = [[str(cell.value or '').strip() for cell in row] for row in sheet.iter_rows()]
        else:
            rows = sheet

        if not rows:
            continue
        raw_headers = [str(h or '').strip() for h in rows[0]]
        headers = [h.lower() for h in raw_headers]
        sheet_name_lower = sheet_name.lower()

        # ═══════════════════════════════════════════════════════════
        # NAME-BASED PRIORITY — Handle known sheet names FIRST to
        # prevent misclassification (e.g. "Mapping" → pl_extract)
        # ═══════════════════════════════════════════════════════════
        if sheet_name_lower == 'mapping' and not ('cogs' in sheet_name_lower or 'revenue' in sheet_name_lower):
            mapping_items = _parse_mapping_sheet(rows)
            # Also parse as P&L Extract for supplementary finance data
            _pl = _parse_pl_extract(rows, sheet_name)
            if _pl.get("finance_income") and finance_income == 0:
                finance_income = _pl["finance_income"]
            if _pl.get("finance_expense") and finance_expense == 0:
                finance_expense = _pl["finance_expense"]
            if _pl.get("tax_expense") and tax_expense == 0:
                tax_expense = _pl["tax_expense"]
            # ALWAYS use Mapping's GA/DA — it's the PRIMARY source for operating expenses.
            # The Mapping sheet contains the full chart of accounts with aggregated amounts
            # for 7310 (Selling), 7410 (Admin), 8110 (NonOp Income), 8220 (NonOp Expense).
            # Previously this was gated on `if not trial_balance_items`, which silently
            # dropped G&A when TDSheet existed alongside Mapping — causing EBITDA = GP.
            if _pl.get("ga_expenses"):
                ga_expenses = _pl["ga_expenses"]
                logger.info(f"Mapping sheet: extracted {len(ga_expenses)} GA/expense items (ALWAYS — primary source)")
            if _pl.get("da_expenses"):
                da_expenses = _pl["da_expenses"]
                logger.info(f"Mapping sheet: extracted {len(da_expenses)} D&A items")
            # TB items from Mapping as fallback (TDSheet is more detailed if available)
            if not trial_balance_items:
                trial_balance_items = _pl.get("trial_balance_items", [])
            logger.info(f"Sheet '{sheet_name}' handled as Mapping sheet (name-based priority)")
            detected_sheets_info.append({"name": sheet_name, "type": "mapping", "confidence": 1.0, "method": "name-based"})
            pipeline_steps.append({"step": "parse_mapping", "detail": f"'{sheet_name}' → Mapping sheet (GA/DA/finance data extracted)"})
            continue

        if sheet_name_lower == 'bs':
            bs_preformatted = _parse_bs_sheet(rows)
            file_type = "Full Report"
            logger.info(f"Sheet '{sheet_name}' handled as pre-formatted BS (name-based priority)")
            detected_sheets_info.append({"name": sheet_name, "type": "balance_sheet", "confidence": 1.0, "method": "name-based"})
            pipeline_steps.append({"step": "parse_balance", "detail": f"'{sheet_name}' → Pre-formatted Balance Sheet"})
            continue

        if 'budget' in sheet_name_lower:
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                val_raw = row[1] if len(row) > 1 else (row[2] if len(row) > 2 else 0)
                v = _safe_float(val_raw)
                if v != 0:
                    budget[str(row[0]).strip()] = v
            continue

        # ═══════════════════════════════════════════════════════════
        # CONTENT-BASED DETECTION — Score each sheet type by structure
        # ═══════════════════════════════════════════════════════════
        td_score = _detect_trial_balance(rows)
        bal_score = _detect_balance_mapped(rows)
        cogs_score = _detect_cogs_breakdown(rows)
        rev_score = _detect_revenue_breakdown(rows, raw_headers)
        pl_score = _detect_pl_extract(rows)

        # Also keep name-based hints as tie-breakers
        if 'tdsheet' in sheet_name_lower or 'td sheet' in sheet_name_lower:
            td_score = max(td_score, 0.7)
        if sheet_name_lower == 'balance':
            bal_score = max(bal_score, 0.7)
        if 'cogs' in sheet_name_lower:
            cogs_score = max(cogs_score, 0.5)

        # Find the best match above threshold
        detections = [
            ('tdsheet', td_score, 0.5),
            ('balance', bal_score, 0.5),
            ('cogs', cogs_score, 0.4),
            ('revenue', rev_score, 0.5),
            ('pl_extract', pl_score, 0.55),
        ]
        # Sort by score descending, pick highest above threshold
        detections.sort(key=lambda x: x[1], reverse=True)
        detected_type = None
        detected_conf = 0.0
        for dtype, dscore, thresh in detections:
            if dscore >= thresh:
                detected_type = dtype
                detected_conf = dscore
                break

        # ── LLM Classification Fallback ───────────────────────────
        # When no detector meets its threshold, ask the DataAgent's LLM
        # to classify the sheet. This handles non-standard formats
        # that the weighted scoring was not trained for.
        if not strict and detected_type is None and len(rows) > 3:
            try:
                llm_result = _llm_classify_sheet(
                    rows, sheet_name,
                    {"tdsheet": td_score, "balance": bal_score,
                     "cogs": cogs_score, "revenue": rev_score, "pl_extract": pl_score}
                )
                if llm_result and llm_result.get("sheet_type") != "unknown":
                    # Map LLM types to our internal types
                    type_map = {
                        "trial_balance": "tdsheet",
                        "revenue": "revenue",
                        "cogs": "cogs",
                        "balance_sheet": "balance",
                        "pl_extract": "pl_extract",
                        "transaction_ledger": None,  # handled separately below
                        "budget": None,              # handled separately below
                    }
                    mapped = type_map.get(llm_result["sheet_type"])
                    if mapped:
                        detected_type = mapped
                        detected_conf = llm_result.get("confidence", 0.6)
                        logger.info(f"LLM classified sheet '{sheet_name}' as '{detected_type}' (conf: {detected_conf:.2f})")
                        pipeline_steps.append({
                            "step": "llm_classify",
                            "detail": f"'{sheet_name}' → {detected_type} via LLM (conf {detected_conf:.0%}): {llm_result.get('reasoning', '')[:100]}"
                        })
            except Exception as e:
                logger.warning(f"LLM classification fallback failed for '{sheet_name}': {e}")

        # ── TDSheet (Trial Balance / Оборотно-сальдовая ведомость) ────
        if detected_type == 'tdsheet':
            logger.info(f"Sheet '{sheet_name}' detected as TDSheet (confidence: {detected_conf:.2f})")
            td_result = _parse_tdsheet(rows)
            trial_balance_items = td_result.get("trial_balance_items", [])
            # TDSheet-derived G&A/D&A (overrides transaction-based extraction)
            if td_result.get("ga_expenses"):
                ga_expenses = td_result["ga_expenses"]
            if td_result.get("da_expenses"):
                da_expenses = td_result["da_expenses"]
            finance_income = td_result.get("finance_income", 0)
            finance_expense = td_result.get("finance_expense", 0)
            tax_expense = td_result.get("tax_expense", 0)
            labour_costs = td_result.get("labour_costs", 0)
            total_cogs_tb = td_result.get("total_cogs_tb", 0)
            total_inventory_credit_tb = td_result.get("total_inventory_credit_tb", 0)
            file_type = "Full Report"
            detected_sheets_info.append({"name": sheet_name, "type": "tdsheet", "confidence": round(detected_conf, 2), "records": len(trial_balance_items), "method": "content-based"})
            _ga_count = len(ga_expenses) if ga_expenses else 0
            pipeline_steps.append({"step": "parse_tdsheet", "detail": f"'{sheet_name}' → {len(trial_balance_items)} TB items, {_ga_count} G&A, {len(da_expenses)} D&A (conf {detected_conf:.0%})"})
            continue

        # ── Balance sheet (IFRS-mapped with MAPPING GRP) ──────────
        if detected_type == 'balance':
            logger.info(f"Sheet '{sheet_name}' detected as IFRS Balance (confidence: {detected_conf:.2f})")
            bal_result = _parse_balance_sheet(rows)
            balance_sheet_items = bal_result.get("balance_sheet_items", [])
            file_type = "Full Report"
            detected_sheets_info.append({"name": sheet_name, "type": "balance_sheet", "confidence": round(detected_conf, 2), "records": len(balance_sheet_items), "method": "content-based"})
            pipeline_steps.append({"step": "parse_balance", "detail": f"'{sheet_name}' → {len(balance_sheet_items)} BS items with IFRS mapping (conf {detected_conf:.0%})"})
            continue

        # ── COGS Breakdown sheet ──────────────────────────────
        if detected_type == 'cogs':
            logger.info(f"Sheet '{sheet_name}' detected as COGS Breakdown (confidence: {detected_conf:.2f})")
            cogs_items = _parse_cogs_breakdown(rows)
            file_type = "Full Report" if transactions else "COGS Report"
            detected_sheets_info.append({"name": sheet_name, "type": "cogs", "confidence": round(detected_conf, 2), "records": len(cogs_items), "method": "content-based"})
            pipeline_steps.append({"step": "parse_cogs", "detail": f"'{sheet_name}' → {len(cogs_items)} COGS products (conf {detected_conf:.0%})"})
            continue

        # ── Revenue Breakdown sheet ───────────────────────────
        if detected_type == 'revenue':
            logger.info(f"Sheet '{sheet_name}' detected as Revenue Breakdown (confidence: {detected_conf:.2f})")
            detected_sheets_info.append({"name": sheet_name, "type": "revenue", "confidence": round(detected_conf, 2), "method": "content-based"})
            # Revenue parsing happens below in the existing flow (uses same has_product logic)
            # Just log the detection; actual parsing continues in the revenue block below

        # ── P&L Account Summary / Extract ─────────────────────
        if detected_type == 'pl_extract':
            logger.info(f"Sheet '{sheet_name}' detected as P&L Extract (confidence: {detected_conf:.2f})")
            pl_result = _parse_pl_extract(rows, sheet_name)
            trial_balance_items = pl_result.get("trial_balance_items", [])
            if pl_result.get("ga_expenses"):
                ga_expenses = pl_result["ga_expenses"]
            if pl_result.get("da_expenses"):
                da_expenses = pl_result["da_expenses"]
            finance_income = pl_result.get("finance_income", 0)
            finance_expense = pl_result.get("finance_expense", 0)
            tax_expense = pl_result.get("tax_expense", 0)
            labour_costs = pl_result.get("labour_costs", 0)
            total_cogs_tb = pl_result.get("total_cogs_tb", 0)
            file_type = "P&L Statement"
            detected_sheets_info.append({"name": sheet_name, "type": "pl_extract", "confidence": round(detected_conf, 2), "records": len(trial_balance_items), "method": "content-based"})
            pipeline_steps.append({"step": "parse_pl", "detail": f"'{sheet_name}' → P&L Extract with {len(trial_balance_items)} items (conf {detected_conf:.0%})"})
            continue

        # (BS, Mapping, and Budget sheets are now handled by name-based
        #  priority above, before content-based detection)

        has_period  = _find_col(headers, 'period', 'date') >= 0
        has_amount  = _find_col(headers, 'amount', 'gel', 'сумма') >= 0
        has_acct    = _find_col(headers, 'account', 'acct', 'debit', 'credit') >= 0
        has_product = _find_col(headers, 'product') >= 0
        has_net     = _find_col(headers, 'net') >= 0

        # ── Transaction sheet (Base) ──────────────────────────
        if (has_period or has_acct) and has_amount:
            file_type = "Transaction Ledger" if file_type == "Financial Data" else "Full Report"
            idx = {
                "date":  _find_col(headers, 'period', 'date'),
                "amt":   _find_col(headers, 'сумма', 'amount gel', 'gel', 'amount'),
                "cls":   _find_col(headers, 'classification', 'type'),
                "dept":  _find_col(headers, 'department eng', 'dept', 'department'),
                "cp":    _find_col(headers, 'counterparty eng', 'counterparty', 'partner'),
                "cat":   _find_col(headers, 'cost classification', 'cost class', 'category'),
                "adr":   _find_col(headers, 'account dr', 'acct dr', 'debit'),
                "acr":   _find_col(headers, 'account cr', 'acct cr', 'credit'),
                "vat":   _find_col(headers, 'vat'),
                "rec":   _find_col(headers, 'recorder', 'name'),
                "month": _find_col(headers, 'month'),
            }

            for row in rows[1:]:
                if not row or idx["date"] < 0 or not row[idx["date"]]:
                    continue
                if idx["amt"] < 0:
                    continue
                amt = _safe_float(row[idx["amt"]]) if idx["amt"] < len(row) else 0
                if amt == 0:
                    continue

                raw_dr = row[idx["adr"]] if idx["adr"] >= 0 and idx["adr"] < len(row) else ""
                raw_cr = row[idx["acr"]] if idx["acr"] >= 0 and idx["acr"] < len(row) else ""

                txn_type = row[idx["cls"]] if idx["cls"] >= 0 and idx["cls"] < len(row) else ""
                if not txn_type:
                    txn_type = infer_txn_type(raw_dr, raw_cr)

                row_num = rows.index(row) + 1 if row in rows else len(transactions) + 2
                transactions.append({
                    "date":         str(row[idx["date"]]).split(' ')[0] if idx["date"] < len(row) else "",
                    "recorder":     (row[idx["rec"]][:50] if idx["rec"] >= 0 and idx["rec"] < len(row) else ""),
                    "acct_dr":      raw_dr,
                    "acct_cr":      raw_cr,
                    "dept":         row[idx["dept"]] if idx["dept"] >= 0 and idx["dept"] < len(row) else "",
                    "counterparty": row[idx["cp"]] if idx["cp"] >= 0 and idx["cp"] < len(row) else "",
                    "cost_class":   row[idx["cat"]] if idx["cat"] >= 0 and idx["cat"] < len(row) else "",
                    "type":         txn_type,
                    "amount":       round(amt, 2),
                    "vat":          _safe_float(row[idx["vat"]]) if idx["vat"] >= 0 and idx["vat"] < len(row) else 0.0,
                    "_lineage": {
                        "source_sheet": sheet_name,
                        "source_row": row_num,
                        "classification_rule": "COA_infer" if not (idx["cls"] >= 0 and idx["cls"] < len(row) and row[idx["cls"]]) else "explicit",
                        "confidence": 0.95 if (idx["cls"] >= 0 and idx["cls"] < len(row) and row[idx["cls"]]) else 0.75,
                    },
                })

        # ── Revenue Breakdown sheet ───────────────────────────
        if has_product and (has_net or has_amount):
            file_type = "Full Report" if transactions else "Revenue Report"
            pi = _find_col(headers, 'product')
            gi = _find_col(headers, 'gross', 'amount gel')
            vi = _find_col(headers, 'vat')
            ni = _find_col(headers, 'net')
            si = _find_col(headers, 'segment', 'category', 'q')

            # ── Auto-detect segment column even if header is missing ──
            # Some Revenue Breakdown sheets have an unlabeled segment column
            # (e.g., col 5 with values "Revenue Retail", "Revenue Wholesale", "Other Revenue")
            # The column may be WITHIN the headers range but with a None/empty header,
            # or it may be BEYOND the headers (extra unlabeled columns).
            if si < 0:
                _SEG_KEYWORDS = {'revenue retail', 'revenue wholesale', 'other revenue',
                                 'wholesale', 'retail', 'cogs retail', 'cogs wholesale'}
                # Build candidate columns: empty-header columns + any beyond headers
                max_data_cols = max((len(r) for r in rows[:10]), default=len(headers))
                candidate_cols = [i for i, h in enumerate(headers) if not h or str(h).strip() == '']
                candidate_cols += list(range(len(headers), min(len(headers) + 4, max_data_cols)))
                for scan_col in candidate_cols:
                    matches = 0
                    for scan_row in rows[1:min(8, len(rows))]:
                        if scan_col < len(scan_row) and scan_row[scan_col]:
                            val = str(scan_row[scan_col]).strip().lower()
                            if val in _SEG_KEYWORDS:
                                matches += 1
                    if matches >= 2:
                        si = scan_col
                        logger.info(f"Revenue: auto-detected segment column at index {si} (no header)")
                        break

            # ── Auto-detect "Eliminated" flag column ──
            # Some items marked "Eliminated" should be excluded (intercompany eliminations)
            elim_idx = -1
            max_data_cols_e = max((len(r) for r in rows[:10]), default=len(headers))
            candidate_cols_e = [i for i, h in enumerate(headers) if not h or str(h).strip() == '']
            candidate_cols_e += list(range(len(headers), min(len(headers) + 4, max_data_cols_e)))
            for scan_col in candidate_cols_e:
                if scan_col == si:
                    continue
                for scan_row in rows[1:min(8, len(rows))]:
                    if scan_col < len(scan_row) and scan_row[scan_col]:
                        val = str(scan_row[scan_col]).strip().lower()
                        if val == 'eliminated':
                            elim_idx = scan_col
                            logger.info(f"Revenue: found 'Eliminated' flag column at index {elim_idx}")
                            break
                if elim_idx >= 0:
                    break

            for row in rows[1:]:
                if not row or pi < 0 or pi >= len(row) or not row[pi]:
                    continue
                product_name = str(row[pi]).strip()
                if product_name.lower() in ('итог', 'итого', 'total', ''):
                    continue

                # Skip eliminated items (intercompany eliminations)
                is_eliminated = False
                if elim_idx >= 0 and elim_idx < len(row) and row[elim_idx]:
                    if str(row[elim_idx]).strip().lower() == 'eliminated':
                        is_eliminated = True

                segment = str(row[si]).strip() if si >= 0 and si < len(row) and row[si] else "Other Revenue"
                category = _classify_revenue_product(product_name)
                rev_row_num = rows.index(row) + 1 if row in rows else len(revenue) + 2
                revenue.append({
                    "product":  product_name,
                    "product_en": get_english_name(product_name),
                    "gross":    _safe_float(row[gi]) if gi >= 0 and gi < len(row) else 0.0,
                    "vat":      _safe_float(row[vi]) if vi >= 0 and vi < len(row) else 0.0,
                    "net":      _safe_float(row[ni]) if ni >= 0 and ni < len(row) else 0.0,
                    "segment":  segment,
                    "category": category,
                    "eliminated": is_eliminated,
                    "_lineage": {
                        "source_sheet": sheet_name,
                        "source_row": rev_row_num,
                        "classification_rule": "product_dimension" if category != "Other Revenue" else "fallback",
                        "confidence": 0.95 if category != "Other Revenue" else 0.5,
                    },
                })

    # Extract G&A and D&A from transactions IF not already provided by TDSheet parser
    if transactions and not ga_expenses:
        ga_expenses = _extract_ga_expenses(transactions)
    if transactions and not da_expenses:
        da_expenses = _extract_da_expenses(transactions)
    if transactions and cogs_items:
        file_type = "Full Report"

    # ── AUTO-GENERATE BS from TDSheet if no Balance sheet was parsed ──────────
    if strict:
        if not trial_balance_items and not balance_sheet_items and not bs_preformatted:
            raise ValueError("Strict parsing requires a Trial Balance or Balance Sheet sheet.")
        if not trial_balance_items and not revenue and not cogs_items:
            raise ValueError("Strict parsing requires Revenue/COGS breakdown or a Trial Balance sheet.")

    # ── CONVERT bs_preformatted (from BS sheet) into balance_sheet_items ────
    # The BS sheet parser produces a clean dict with items/sections/totals,
    # but datasets.py only stores balance_sheet_items (list of BalanceSheetItem dicts).
    # Convert the preformatted BS into the same format so it gets persisted.
    if bs_preformatted and not balance_sheet_items:
        bs_items_from_preformatted = []
        for item in bs_preformatted.get("items", []):
            label = item.get("label", "")
            amount = item.get("amount", 0)
            section = item.get("section", "")
            # Map section to IFRS statement type
            ifrs_statement = "BS"
            bs_items_from_preformatted.append({
                "account_code": "",
                "account_name": label,
                "ifrs_line_item": label,
                "ifrs_statement": ifrs_statement,
                "baku_bs_mapping": "",
                "intercompany_entity": "",
                "opening_balance": 0.0,
                "turnover_debit": 0.0,
                "turnover_credit": 0.0,
                "closing_balance": amount,
                "row_type": "detail" if section else "total",
                "_lineage": {
                    "source_sheet": "BS",
                    "source_row": None,
                    "classification_rule": "bs_preformatted",
                    "confidence": 1.0,
                },
            })
        # Also add total rows from the totals dict
        for key, amount in bs_preformatted.get("totals", {}).items():
            label = key.replace("_", " ").replace("total ", "Total ").title()
            bs_items_from_preformatted.append({
                "account_code": "",
                "account_name": label,
                "ifrs_line_item": label,
                "ifrs_statement": "BS",
                "baku_bs_mapping": "",
                "intercompany_entity": "",
                "opening_balance": 0.0,
                "turnover_debit": 0.0,
                "turnover_credit": 0.0,
                "closing_balance": amount,
                "row_type": "total",
            })
        if bs_items_from_preformatted:
            balance_sheet_items = bs_items_from_preformatted
            logger.info(f"Converted {len(bs_items_from_preformatted)} BS preformatted items to balance_sheet_items")

    if trial_balance_items and not balance_sheet_items:
        auto_bs = _generate_bs_from_tdsheet(trial_balance_items)
        if auto_bs:
            for item in auto_bs:
                item["_lineage"] = {
                    "source_sheet": "TDSheet",
                    "source_row": None,
                    "classification_rule": "derive_bs_from_tdsheet",
                    "confidence": 1.0,
                    "transform_chain": [
                        {"step": "derive_bs_from_tdsheet", "detail": "GEORGIAN_COA mapping from Trial Balance"},
                    ],
                }
            balance_sheet_items = auto_bs
            logger.info(f"No Balance sheet found — auto-generated {len(auto_bs)} BS items from TDSheet via GEORGIAN_COA")

    # ── AUTO-GENERATE P&L from TDSheet class 6-9 accounts if no Revenue/COGS parsed ──
    if trial_balance_items and not revenue and not cogs_items:
        auto_pl = _generate_pl_from_tdsheet(trial_balance_items)
        if auto_pl and auto_pl.get('revenue_items'):
            revenue = auto_pl['revenue_items']
            cogs_items = auto_pl.get('cogs_items', [])
            # Auto-generated GA from TDSheet is more comprehensive (leaf-level
            # with parent filtering) — use it exclusively, don't merge with
            # P&L Extract GA which may have parent-level summaries
            auto_ga = auto_pl.get('ga_expenses', [])
            if auto_ga:
                ga_expenses = auto_ga
            for r in revenue:
                r["_lineage"] = {
                    "source_sheet": "TDSheet",
                    "source_row": None,
                    "classification_rule": "derive_pl_from_tdsheet",
                    "confidence": 1.0,
                    "transform_chain": [
                        {"step": "derive_pl_from_tdsheet", "detail": "GEORGIAN_COA mapping from Trial Balance"},
                    ],
                }
            for c in cogs_items:
                c["_lineage"] = {
                    "source_sheet": "TDSheet",
                    "source_row": None,
                    "classification_rule": "derive_pl_from_tdsheet",
                    "confidence": 1.0,
                    "transform_chain": [
                        {"step": "derive_pl_from_tdsheet", "detail": "GEORGIAN_COA mapping from Trial Balance"},
                    ],
                }
            for g in ga_expenses:
                g["_lineage"] = {
                    "source_sheet": "TDSheet",
                    "source_row": None,
                    "classification_rule": "derive_pl_from_tdsheet",
                    "confidence": 1.0,
                    "transform_chain": [
                        {"step": "derive_pl_from_tdsheet", "detail": "GEORGIAN_COA mapping from Trial Balance"},
                    ],
                }
            # Store PL summary for downstream use
            finance_income = auto_pl['pl_summary'].get('finance_income', finance_income)
            finance_expense = auto_pl['pl_summary'].get('finance_expense', finance_expense)
            tax_expense = auto_pl['pl_summary'].get('tax', tax_expense)
            labour_costs = auto_pl['pl_summary'].get('sga', 0) if not labour_costs else labour_costs
            logger.info(f"No Revenue/COGS sheets — auto-generated P&L from TDSheet: {len(revenue)} revenue, {len(cogs_items)} COGS items")

    # ── FALLBACK: Derive Revenue & COGS from Transactions when sheets are missing ──
    # Layer 1: Pure COA-based derivation
    if not strict and transactions and not revenue:
        revenue, _fb_rev = _derive_revenue_from_transactions(transactions)
        if revenue:
            file_type = file_type + " (Revenue derived from ledger)"

    if not strict and transactions and not cogs_items:
        cogs_items, _fb_cogs = _derive_cogs_from_transactions(transactions)
        if cogs_items:
            file_type = file_type if "derived" in file_type else file_type + " (COGS derived from ledger)"

    # Layer 2: Semantic Layer enhancement — enrich with counterparty/dept/cost_class signals
    if not strict and transactions:
        try:
            from app.services.semantic_layer import (
                derive_enhanced_financials, get_pattern_store
            )
            semantic_data = derive_enhanced_financials(transactions)

            # If we have revenue/cogs from sheets, teach the pattern store
            if (revenue and any(r.get("category") != "Other Revenue" for r in revenue)) or \
               (cogs_items and any(c.get("category") != "Other COGS" for c in cogs_items)):
                store = get_pattern_store()
                store.learn_from_full_report(transactions, revenue, cogs_items)
                logger.info(f"Semantic layer learned from full report: {store.get_stats()}")

            # If COA fallback produced only "Other" categories, semantic layer may do better
            if revenue and all(r.get("category") == "Other Revenue" for r in revenue):
                sem_rev = semantic_data.get("revenue_items", [])
                if sem_rev:
                    # Merge semantic revenue with COA-derived — prefer items with better segments
                    for sr in sem_rev:
                        if sr.get("segment") != "Other Revenue":
                            revenue.append(sr)
                    logger.info(f"Semantic layer enriched revenue with {len(sem_rev)} items")

            if cogs_items and all(c.get("category") == "Other COGS" for c in cogs_items):
                sem_cogs = semantic_data.get("cogs_items", [])
                if sem_cogs:
                    for sc in sem_cogs:
                        if sc.get("segment") != "Other COGS":
                            cogs_items.append(sc)
                    logger.info(f"Semantic layer enriched COGS with {len(sem_cogs)} items")

            # If no G&A was found, use semantic G&A extraction
            if not ga_expenses and semantic_data.get("ga_expenses"):
                ga_expenses = semantic_data["ga_expenses"]
                logger.info(f"Semantic layer provided {len(ga_expenses)} G&A items")

            # Attach semantic analysis stats to result
            _semantic_stats = semantic_data.get("stats", {})
        except Exception as e:
            logger.warning(f"Semantic layer enhancement failed (non-critical): {e}")
            _semantic_stats = {}

    # File name overrides
    fn_lower = filename.lower()
    if 'mr_' in fn_lower or 'management' in fn_lower:
        file_type = "Management Report"
    elif 'budget' in fn_lower:
        file_type = "Budget Report"
    elif 'p&l' in fn_lower or 'pl_' in fn_lower:
        file_type = "P&L Statement"
    elif 'balance' in fn_lower:
        file_type = "Balance Sheet"

    period = _detect_period(filename)

    # ── Period detection from sheet content (Russian months) ─────────
    # If filename detection failed, scan TDSheet/COGS headers for "Период: Январь 2026 г."
    if not period:
        _RUSSIAN_MONTHS = {
            'январь': '01', 'февраль': '02', 'март': '03', 'апрель': '04',
            'май': '05', 'июнь': '06', 'июль': '07', 'август': '08',
            'сентябрь': '09', 'октябрь': '10', 'ноябрь': '11', 'декабрь': '12',
        }
        for sname, sheet in sheets.items():
            if ext in ('xlsx', 'xls', 'xlsm'):
                _first_rows = [[str(cell.value or '') for cell in row] for row in list(sheet.iter_rows(min_row=1, max_row=10))]
            else:
                _first_rows = sheet[:10]
            for _row in _first_rows:
                for _cell in _row:
                    _cell_lower = str(_cell).lower().strip()
                    for ru_month, num in _RUSSIAN_MONTHS.items():
                        if ru_month in _cell_lower:
                            _yr_match = re.search(r'20\d{2}', str(_cell))
                            if _yr_match:
                                _EN_MONTHS = ['January','February','March','April','May','June',
                                              'July','August','September','October','November','December']
                                period = f"{_EN_MONTHS[int(num)-1]} {_yr_match.group()}"
                                logger.info(f"Period from Russian text in '{sname}': {period}")
                                break
                    if period:
                        break
                if period:
                    break
            if period:
                break

    total_records = len(transactions) + len(revenue) + len(budget) + len(cogs_items) + len(trial_balance_items) + len(balance_sheet_items)

    # ── COA mapping coverage step ──────────────────────────────────
    if trial_balance_items:
        _tb_codes = set()
        _tb_mapped = 0
        for tbi in trial_balance_items:
            code = tbi.get("account_code", "")
            if code and code not in _tb_codes:
                _tb_codes.add(code)
                if map_coa(code) is not None:
                    _tb_mapped += 1
        _tb_total = len(_tb_codes)
        _cov = round(_tb_mapped / _tb_total * 100, 1) if _tb_total else 100.0
        pipeline_steps.append({"step": "coa_mapping", "detail": f"COA mapped {_tb_mapped}/{_tb_total} accounts ({_cov}%)"})

    # ── COGS ↔ Inventory Reconciliation ──────────────────────────────
    cogs_reconciliation = _reconcile_cogs(cogs_items, total_cogs_tb, total_inventory_credit_tb)
    if cogs_reconciliation and cogs_reconciliation.get("checks"):
        _has_mismatch = cogs_reconciliation.get("has_mismatch", False)
        pipeline_steps.append({"step": "reconcile", "detail": f"COGS reconciliation: {'⚠ mismatch detected' if _has_mismatch else '✓ matched'}"})

    # ── Data quality flags ────────────────────────────────────────────
    data_quality_flags = []
    if not ga_expenses:
        data_quality_flags.append({"code": "MISSING_GA", "severity": "CRITICAL",
            "message": "No G&A expenses found — EBITDA will equal Gross Profit"})
    if not da_expenses:
        data_quality_flags.append({"code": "MISSING_DEPRECIATION", "severity": "HIGH",
            "message": "No depreciation/amortization data found"})
    if not balance_sheet_items and not bs_preformatted:
        data_quality_flags.append({"code": "MISSING_BS", "severity": "HIGH",
            "message": "No Balance Sheet data found in file"})
    if not period or period == "January 2025":
        data_quality_flags.append({"code": "NO_PERIOD_DETECTED", "severity": "HIGH",
            "message": "Could not detect period from file — using default"})
    # Revenue cross-validation (Mapping vs Revenue Breakdown)
    if mapping_items and revenue:
        mapping_rev = 0
        for m in mapping_items:
            code = m.get("account_code", "")
            if code.startswith("61") and "XX" in code:
                mapping_rev = abs(m.get("amount", 0))
        if mapping_rev > 0:
            rev_bd_total = sum(r.get("net", 0) for r in revenue)
            if rev_bd_total > 0 and abs(mapping_rev - rev_bd_total) > 1000:
                data_quality_flags.append({"code": "REVENUE_MISMATCH", "severity": "MEDIUM",
                    "message": f"Revenue Breakdown ({rev_bd_total:,.0f}) != Mapping 61XX ({mapping_rev:,.0f})"})
    # Compute score
    _dq_score = 100
    for f in data_quality_flags:
        if f["severity"] == "CRITICAL": _dq_score -= 25
        elif f["severity"] == "HIGH": _dq_score -= 15
        elif f["severity"] == "MEDIUM": _dq_score -= 5
    _dq_score = max(0, min(100, _dq_score))

    # ── Final summary step ──────────────────────────────────────────
    # Professional Integrity Check: Balance Sheet Delta
    bs_delta = 0.0
    _bs_items = []
    if trial_balance_items:
        # Filter for Balance Sheet accounts using mapping hints
        for tbi in trial_balance_items:
            side = tbi.get("bs_side")
            if side in ["asset", "liability", "equity"]:
                _bs_items.append(tbi)
    
    if _bs_items:
        assets = sum(abs(i.get("closing_balance", 0)) for i in _bs_items if i.get("bs_side") == "asset")
        liabs = sum(abs(i.get("closing_balance", 0)) for i in _bs_items if i.get("bs_side") == "liability")
        equity = sum(abs(i.get("closing_balance", 0)) for i in _bs_items if i.get("bs_side") == "equity")
        bs_delta = abs(assets - (liabs + equity))
        
        if bs_delta > 10.0: # 10 GEL materiality threshold
            data_quality_flags.append({
                "code": "BALANCE_SHEET_DISCREPANCY", 
                "severity": "CRITICAL",
                "message": f"Balance sheet is non-compliant. Delta: {bs_delta:,.2f} GEL (Assets: {assets:,.0f} | L+E: {liabs+equity:,.0f})"
            })
            _dq_score = max(0, _dq_score - 40)
        else:
            pipeline_steps.append({"step": "integrity", "detail": "✓ Balance Sheet verified (Compliance Checked)"})

    pipeline_steps.append({"step": "complete", "detail": f"Ready: {total_records} records across {len(detected_sheets_info)} sheet types (quality: {_dq_score}/100)"})

    return {
        "file_type":    file_type,
        "period":       period,
        "sheet_count":  len(sheets),
        "transactions": transactions,
        "revenue":      revenue,
        "budget":       budget,
        "cogs_items":   cogs_items,
        "ga_expenses":  ga_expenses,
        "da_expenses":  da_expenses,
        "trial_balance_items": trial_balance_items,
        "balance_sheet_items": balance_sheet_items,
        "bs_preformatted": bs_preformatted,
        "mapping_items": mapping_items,
        "finance_income": finance_income,
        "finance_expense": finance_expense,
        "tax_expense": tax_expense,
        "labour_costs": labour_costs,
        "total_cogs_tb": total_cogs_tb,
        "total_inventory_credit_tb": total_inventory_credit_tb,
        "cogs_reconciliation": cogs_reconciliation,
        "data_quality_flags": data_quality_flags,
        "data_quality_score": _dq_score,
        "record_count": total_records,
        "semantic_stats": _semantic_stats,
        "processing_pipeline": pipeline_steps,
        "detected_sheets": detected_sheets_info,
    }
