"""
FinAI Excel Report Bytes Generator
===================================
Generates Excel report bytes for email attachments.
Reuses the same styling as the export endpoints but returns bytes instead of StreamingResponse.
"""

import io
import logging
from typing import Optional, Dict, Any

from app.config import settings

logger = logging.getLogger(__name__)


async def generate_pl_excel_bytes(dataset_id: int, prior_dataset_id: Optional[int], db) -> tuple[bytes, str]:
    """Generate P&L comparison Excel bytes for email attachment."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from app.routers.journal_router import _auto_find_prior
    from sqlalchemy import select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not dataset_id:
        result = await db.execute(
            select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000)
            .order_by(Dataset.id.desc()).limit(1)
        )
        ds = result.scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise ValueError("No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.full_pl(dataset_id, prior_dataset_id, db)
    summary = data.get("summary", {})
    rows = data.get("rows", [])

    # Modern style
    PRIMARY = "1B3A5C"
    ACCENT = "2563EB"
    LIGHT_BG = "F0F4F8"
    WHITE = "FFFFFF"
    RED = "DC2626"
    GREEN = "059669"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    # Row 1: Header
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = data.get("company", settings.COMPANY_NAME)
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Row 2: Subtitle
    ws.merge_cells("A2:G2")
    period_text = f"Income Statement — {data.get('period', '')}"
    if data.get("prior_period"):
        period_text += f"  vs  {data['prior_period']}"
    period_text += f"  |  Currency: {data.get('currency', 'GEL')}"
    c = ws["A2"]
    c.value = period_text
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8

    # KPI row
    kpi_data = [
        ("Revenue", summary.get("revenue", 0)),
        ("Gross Profit", summary.get("gross_profit", 0)),
        ("EBITDA", summary.get("ebitda", 0)),
        ("Net Profit", summary.get("net_profit", 0)),
    ]
    for i, (label, value) in enumerate(kpi_data):
        col = i + 2
        ws.cell(row=4, column=col, value=label).font = Font(name="Calibri", size=9, color="64748B")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        cell = ws.cell(row=5, column=col, value=value)
        color = RED if isinstance(value, (int, float)) and value < 0 else PRIMARY
        cell.font = Font(name="Calibri", bold=True, size=14, color=color)
        cell.number_format = '#,##0;(#,##0);"-"'
        cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[6].height = 8

    # Column headers
    headers = ["Code", "Line Item", "Prior Year", "Actual", "% of Revenue", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=i, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="center")
    ws.row_dimensions[7].height = 26

    # Data rows
    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'
    row_num = 8
    revenue = summary.get("revenue", 1) or 1

    for idx, r in enumerate(rows):
        ac = r.get("ac", 0) or 0
        pr = r.get("pr", 0) or 0
        var = r.get("var", 0) or 0
        var_pct = r.get("var_pct", 0) or 0
        is_bold = r.get("bold", False)
        level = r.get("lvl", 0)

        indent = "  " * level
        ws.cell(row=row_num, column=1, value=r.get("c", ""))
        ws.cell(row=row_num, column=2, value=f"{indent}{r.get('l', '')}")
        ws.cell(row=row_num, column=3, value=pr if pr != 0 else None)
        ws.cell(row=row_num, column=4, value=ac if ac != 0 else None)
        ws.cell(row=row_num, column=5, value=ac / revenue if revenue and ac else None)
        ws.cell(row=row_num, column=6, value=var if var != 0 else None)
        ws.cell(row=row_num, column=7, value=var_pct / 100 if var_pct else None)

        row_fill = PatternFill("solid", fgColor=LIGHT_BG) if is_bold else (PatternFill("solid", fgColor="F8FAFC") if idx % 2 else PatternFill("solid", fgColor=WHITE))
        bold_font = Font(name="Calibri", bold=True, size=10)
        normal_font = Font(name="Calibri", size=10)
        row_font = bold_font if is_bold else normal_font
        border = Border(bottom=Side(style="medium" if r.get("sep") else "thin", color="D1D5DB"))

        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = border
            if col <= 2:
                cell.font = row_font
            elif col in (3, 4, 6):
                cell.number_format = num_fmt
                val = cell.value
                if isinstance(val, (int, float)) and val < 0:
                    cell.font = Font(name="Calibri", bold=is_bold, size=10, color=RED)
                else:
                    cell.font = row_font
            elif col == 5:
                cell.number_format = pct_fmt
                cell.font = Font(name="Calibri", size=9, color="64748B")
            elif col == 7:
                cell.number_format = pct_fmt
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = Font(name="Calibri", bold=is_bold, size=10, color=GREEN if val > 0 else RED)
                else:
                    cell.font = row_font

        ws.cell(row=row_num, column=1).font = Font(name="Calibri", size=9, color="9CA3AF")
        row_num += 1

    # Footer
    row_num += 1
    ws.merge_cells(f"A{row_num}:G{row_num}")
    footer = ws.cell(row=row_num, column=1)
    footer.value = f"Generated by FinAI Foundry  |  {data.get('company', '')}  |  {data.get('period', '')}  |  All amounts in {data.get('currency', 'GEL')}"
    footer.font = Font(name="Calibri", size=8, color="9CA3AF", italic=True)
    footer.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10
    ws.freeze_panes = "A8"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"PL_Statement_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return buf.getvalue(), filename


async def generate_bs_excel_bytes(dataset_id: int, prior_dataset_id: Optional[int], db) -> tuple[bytes, str]:
    """Generate Balance Sheet Excel bytes for email attachment."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from app.routers.journal_router import _auto_find_prior
    from sqlalchemy import select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise ValueError("No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.balance_sheet_comparison(dataset_id, prior_dataset_id, db)

    PRIMARY = "1B3A5C"
    ACCENT = "2563EB"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    ws.merge_cells("A1:E1")
    ws["A1"].value = settings.COMPANY_NAME
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    period_text = f"Balance Sheet — {data.get('period', '')}"
    if data.get("prior_period"):
        period_text += f"  vs  {data['prior_period']}"
    ws["A2"].value = period_text + "  |  Currency: GEL"
    ws["A2"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=ACCENT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8

    headers = ["IFRS Line Item", "Prior Year", "Actual", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left")

    rows_data = data.get("rows", [])
    row_num = 5
    current_section = None
    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'

    for r in rows_data:
        if r.get("section") != current_section:
            current_section = r.get("section")
            ws.merge_cells(f"A{row_num}:E{row_num}")
            cell = ws.cell(row=row_num, column=1, value=current_section)
            cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=ACCENT)
            row_num += 1

        actual = r.get("actual", 0)
        prior = r.get("prior", 0)
        variance = r.get("variance", 0)
        variance_pct = r.get("variance_pct", 0)
        is_bold = r.get("bold", False)
        level = r.get("level", 0)

        ws.cell(row=row_num, column=1, value=("  " * level) + r.get("ifrs_line", ""))
        ws.cell(row=row_num, column=2, value=prior if prior else None)
        ws.cell(row=row_num, column=3, value=actual if actual else None)
        ws.cell(row=row_num, column=4, value=variance if variance else None)
        ws.cell(row=row_num, column=5, value=variance_pct / 100 if variance_pct else None)

        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = Border(bottom=Side(style="thin", color="D1D5DB"))
            if col == 1:
                cell.font = Font(name="Calibri", bold=is_bold, size=10)
            elif col in (2, 3, 4):
                cell.number_format = num_fmt
                val = cell.value
                cell.font = Font(name="Calibri", bold=is_bold, size=10, color="DC2626" if isinstance(val, (int, float)) and val < 0 else "1E293B")
            elif col == 5:
                cell.number_format = pct_fmt
                val = cell.value
                cell.font = Font(name="Calibri", bold=is_bold, size=10, color="059669" if isinstance(val, (int, float)) and val > 0 else ("DC2626" if isinstance(val, (int, float)) and val < 0 else "64748B"))

        row_num += 1

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 10
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"Balance_Sheet_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return buf.getvalue(), filename


async def generate_revenue_excel_bytes(dataset_id: int, prior_dataset_id: Optional[int], db) -> tuple[bytes, str]:
    """Generate Revenue comparison Excel bytes."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from app.routers.journal_router import _auto_find_prior
    from sqlalchemy import select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise ValueError("No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.revenue_comparison(dataset_id, prior_dataset_id, db)

    PRIMARY = "1B3A5C"
    ACCENT = "2563EB"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue Analysis"

    ws.merge_cells("A1:G1")
    ws["A1"].value = settings.COMPANY_NAME
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Revenue Analysis — {data.get('period', '')} vs {data.get('prior_period', '')}"
    ws["A2"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=ACCENT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8

    headers = ["Product", "Segment", "Prior Year", "Actual", "% of Total", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=PRIMARY)
        c.alignment = Alignment(horizontal="center")

    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'
    rows = data.get('rows', [])
    for rindex, r in enumerate(rows):
        row = rindex + 5
        ws.cell(row=row, column=1, value=r.get('product', '')).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=r.get('segment', '')).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=3, value=r.get('prior_net', 0) or None)
        ws.cell(row=row, column=4, value=r.get('actual_net', 0) or None)
        ws.cell(row=row, column=5, value=(r.get('pct_of_total', 0) or 0) / 100)
        ws.cell(row=row, column=6, value=r.get('variance', 0) or None)
        ws.cell(row=row, column=7, value=(r.get('variance_pct', 0) or 0) / 100)

        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.number_format = num_fmt if col in (3, 4, 6) else pct_fmt
            cell.border = Border(bottom=Side(style="thin", color="D1D5DB"))
            val = cell.value
            if col in (6, 7) and isinstance(val, (int, float)):
                cell.font = Font(name="Calibri", size=10, color="059669" if val > 0 else "DC2626")

    # Totals row
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_row, column=3, value=data.get('total_revenue_prior', 0))
    ws.cell(row=total_row, column=4, value=data.get('total_revenue_actual', 0))
    ws.cell(row=total_row, column=6, value=data.get('total_variance', 0))
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = Border(top=Side(style="medium", color=PRIMARY), bottom=Side(style="double", color=PRIMARY))
        if col in (3, 4, 6):
            cell.number_format = num_fmt

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"Revenue_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return buf.getvalue(), filename


async def generate_cogs_excel_bytes(dataset_id: int, prior_dataset_id: Optional[int], db) -> tuple[bytes, str]:
    """Generate COGS comparison Excel bytes."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from app.routers.journal_router import _auto_find_prior
    from sqlalchemy import select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise ValueError("No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.cogs_comparison(dataset_id, prior_dataset_id, db)

    PRIMARY = "1B3A5C"
    ACCENT = "2563EB"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COGS Analysis"

    ws.merge_cells("A1:G1")
    ws["A1"].value = settings.COMPANY_NAME
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=PRIMARY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"COGS Analysis — {data.get('period', '')} vs {data.get('prior_period', '')}"
    ws["A2"].font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=ACCENT)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 8

    headers = ["Category", "Segment", "Prior Year", "Actual", "% of Total", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=PRIMARY)
        c.alignment = Alignment(horizontal="center")

    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'
    rows = data.get('rows', [])
    total_actual = data.get('total_actual', 1) or 1

    for rindex, r in enumerate(rows):
        row = rindex + 5
        ws.cell(row=row, column=1, value=r.get('category', '')).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=r.get('segment', '')).font = Font(name="Calibri", size=10)
        ws.cell(row=row, column=3, value=r.get('prior', 0) or None)
        ws.cell(row=row, column=4, value=r.get('actual', 0) or None)
        ws.cell(row=row, column=5, value=(r.get('actual', 0) or 0) / total_actual if total_actual else 0)
        ws.cell(row=row, column=6, value=r.get('variance', 0) or None)
        ws.cell(row=row, column=7, value=(r.get('variance_pct', 0) or 0) / 100)

        for col in range(3, 8):
            cell = ws.cell(row=row, column=col)
            cell.number_format = num_fmt if col in (3, 4, 6) else pct_fmt
            cell.border = Border(bottom=Side(style="thin", color="D1D5DB"))
            val = cell.value
            # For COGS, negative variance (cost reduction) is good
            if col in (6, 7) and isinstance(val, (int, float)):
                cell.font = Font(name="Calibri", size=10, color="059669" if val < 0 else "DC2626")

    # Totals
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="TOTAL COGS").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=total_row, column=3, value=data.get('total_prior', 0))
    ws.cell(row=total_row, column=4, value=data.get('total_actual', 0))
    ws.cell(row=total_row, column=6, value=data.get('total_variance', 0))
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(name="Calibri", bold=True, size=11)
        cell.border = Border(top=Side(style="medium", color=PRIMARY), bottom=Side(style="double", color=PRIMARY))
        if col in (3, 4, 6):
            cell.number_format = num_fmt

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"COGS_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return buf.getvalue(), filename
