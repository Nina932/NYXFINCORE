"""
Phase N-1: SmartExcelParser — Universal Financial Excel/CSV Parser
===================================================================
Accepts ANY Excel/CSV file and automatically:
  1. Detects header row (even if not row 1)
  2. Fuzzy-matches columns to canonical financial fields
  3. Supports English, Georgian, and Russian column names
  4. Handles multi-sheet workbooks, merged cells, summary rows
  5. Returns normalized financial records with confidence scores

Canonical fields:
  revenue, cogs, gross_profit, ga_expenses, ebitda, net_profit,
  depreciation, finance_expense, tax_expense, gross_margin_pct,
  net_margin_pct, ebitda_margin_pct, cogs_to_revenue_pct,
  total_assets, total_liabilities, total_equity,
  total_current_assets, total_current_liabilities, cash, period

All derived metrics (margins, percentages) are computed deterministically
if missing — NO LLM involvement.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════
# CANONICAL FIELD DEFINITIONS + FUZZY KEYWORD MAP
# ═══════════════════════════════════════════════════════════════════

# Each canonical field → list of possible column names (EN, KA, RU)
_FIELD_KEYWORDS: Dict[str, List[str]] = {
    "revenue": [
        "revenue", "sales", "income", "turnover", "total revenue",
        "net revenue", "net sales", "gross sales", "total sales",
        "შემოსავალი", "გაყიდვები", "რეალიზაცია",
        "выручка", "доход", "продажи", "реализация",
    ],
    "cogs": [
        "cogs", "cost of goods sold", "cost of sales", "cost of revenue",
        "direct costs", "cost of goods", "cos",
        "თვითღირებულება", "გაყიდული საქონლის თვითღირებულება",
        "себестоимость", "себестоимость продаж",
    ],
    "gross_profit": [
        "gross profit", "gross income", "gross margin",
        "მთლიანი მოგება",
        "валовая прибыль", "валовой доход",
    ],
    "ga_expenses": [
        "g&a", "ga expenses", "general and administrative",
        "admin expenses", "administrative expenses",
        "operating expenses", "opex", "selling general",
        "sga", "sg&a", "selling expenses",
        "ადმინისტრაციული ხარჯები", "საოპერაციო ხარჯები",
        "административные расходы", "коммерческие расходы",
        "управленческие расходы",
    ],
    "ebitda": [
        "ebitda", "operating profit before depreciation",
        "ებითდა", "ебітда",
    ],
    "net_profit": [
        "net profit", "net income", "net earnings",
        "profit after tax", "pat", "bottom line",
        "წმინდა მოგება", "წმინდა შემოსავალი",
        "чистая прибыль", "чистый доход",
    ],
    "depreciation": [
        "depreciation", "depreciation and amortization",
        "d&a", "amortization",
        "ცვეთა", "ამორტიზაცია",
        "амортизация", "износ",
    ],
    "finance_expense": [
        "finance expense", "interest expense", "finance costs",
        "financial expenses", "interest",
        "საფინანსო ხარჯი", "პროცენტი",
        "финансовые расходы", "процентные расходы",
    ],
    "tax_expense": [
        "tax expense", "income tax", "tax",
        "გადასახადი", "საშემოსავლო გადასახადი",
        "налог на прибыль", "налог",
    ],
    # Balance sheet fields
    "total_assets": [
        "total assets", "assets total",
        "სულ აქტივები", "აქტივები სულ",
        "итого активы", "всего активов",
    ],
    "total_liabilities": [
        "total liabilities", "liabilities total",
        "სულ ვალდებულებები",
        "итого обязательства",
    ],
    "total_equity": [
        "total equity", "shareholders equity", "stockholders equity",
        "equity total", "net worth",
        "სულ კაპიტალი", "საკუთარი კაპიტალი",
        "собственный капитал", "итого капитал",
    ],
    "total_current_assets": [
        "current assets", "total current assets",
        "მიმდინარე აქტივები",
        "оборотные активы",
    ],
    "total_current_liabilities": [
        "current liabilities", "total current liabilities",
        "მიმდინარე ვალდებულებები",
        "краткосрочные обязательства",
    ],
    "cash": [
        "cash", "cash and equivalents", "cash & equivalents",
        "ფულადი სახსრები", "ნაღდი ფული",
        "денежные средства", "касса",
    ],
    "period": [
        "period", "date", "month", "quarter", "year",
        "პერიოდი", "თარიღი", "თვე",
        "период", "дата", "месяц",
    ],
}

# Reverse index for fast lookup
_KEYWORD_TO_FIELD: Dict[str, str] = {}
for _field, _keywords in _FIELD_KEYWORDS.items():
    for _kw in _keywords:
        _KEYWORD_TO_FIELD[_kw.lower()] = _field


# ═══════════════════════════════════════════════════════════════════
# FUZZY MATCHER
# ═══════════════════════════════════════════════════════════════════

def _normalize_header(text: str) -> str:
    """Normalize column header for matching."""
    if not text:
        return ""
    text = str(text).strip().lower()
    # Remove special chars except &
    text = re.sub(r"[^\w\s&]", "", text)
    # Collapse whitespace
    text = re.sub(r"\s+", " ", text)
    return text


def _fuzzy_match_column(header: str) -> Tuple[Optional[str], int]:
    """
    Match a column header to a canonical field.

    Returns:
        (field_name, confidence_pct) or (None, 0)
    """
    norm = _normalize_header(header)
    if not norm:
        return None, 0

    # 1. Exact match (100%)
    if norm in _KEYWORD_TO_FIELD:
        return _KEYWORD_TO_FIELD[norm], 100

    # 2. Contains match (85%)
    for keyword, field_name in _KEYWORD_TO_FIELD.items():
        if keyword in norm or norm in keyword:
            return field_name, 85

    # 3. Token overlap (70%)
    norm_tokens = set(norm.split())
    best_field = None
    best_overlap = 0
    for keyword, field_name in _KEYWORD_TO_FIELD.items():
        kw_tokens = set(keyword.split())
        if not kw_tokens:
            continue
        overlap = len(norm_tokens & kw_tokens) / len(kw_tokens)
        if overlap > best_overlap and overlap >= 0.5:
            best_overlap = overlap
            best_field = field_name

    if best_field:
        return best_field, int(70 * best_overlap)

    return None, 0


# ═══════════════════════════════════════════════════════════════════
# DATA CLASSES
# ═══════════════════════════════════════════════════════════════════

@dataclass
class ColumnMapping:
    """Mapping of a source column to a canonical field."""
    source_index: int
    source_header: str
    canonical_field: Optional[str]
    confidence: int  # 0-100

    def to_dict(self) -> Dict[str, Any]:
        return {
            "source_index": self.source_index,
            "source_header": self.source_header,
            "canonical_field": self.canonical_field,
            "confidence": self.confidence,
        }


@dataclass
class ParsedSheet:
    """Result of parsing one sheet."""
    sheet_name: str
    header_row: int           # 0-indexed
    column_mappings: List[ColumnMapping]
    records: List[Dict[str, Any]]
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "column_count": len(self.column_mappings),
            "mapped_columns": len([m for m in self.column_mappings if m.canonical_field]),
            "unmapped_columns": len([m for m in self.column_mappings if not m.canonical_field]),
            "record_count": len(self.records),
            "column_mappings": [m.to_dict() for m in self.column_mappings],
            "warnings": self.warnings,
        }


@dataclass
class ParseResult:
    """Complete parse result for a file."""
    filename: str
    file_type: str            # xlsx, xls, csv
    sheets: List[ParsedSheet]
    normalized_financials: Dict[str, Any]  # merged + derived metrics
    confidence_score: int     # overall 0-100
    auto_corrections: List[str]
    warnings: List[str]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "file_type": self.file_type,
            "sheet_count": len(self.sheets),
            "sheets": [s.to_dict() for s in self.sheets],
            "normalized_financials": self.normalized_financials,
            "confidence_score": self.confidence_score,
            "auto_corrections": self.auto_corrections,
            "warnings": self.warnings,
        }


# ═══════════════════════════════════════════════════════════════════
# HEADER DETECTOR
# ═══════════════════════════════════════════════════════════════════

def _detect_header_row(rows: List[List[Any]], max_scan: int = 20) -> int:
    """
    Find the header row by scoring each row for financial keywords.

    Returns 0-indexed row number.
    """
    if not rows:
        return 0

    best_row = 0
    best_score = 0

    for i, row in enumerate(rows[:max_scan]):
        score = 0
        text_cells = 0
        for cell in row:
            if cell is None:
                continue
            val = str(cell).strip().lower()
            if not val:
                continue
            text_cells += 1
            # Check if cell matches any known keyword
            if val in _KEYWORD_TO_FIELD:
                score += 10
            else:
                for kw in _KEYWORD_TO_FIELD:
                    if kw in val or val in kw:
                        score += 5
                        break
                else:
                    # General text (not numeric) → mild header signal
                    try:
                        float(val.replace(",", "").replace(" ", ""))
                    except ValueError:
                        score += 1

        # Penalize rows with too few cells
        if text_cells < 2:
            score = 0

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


# ═══════════════════════════════════════════════════════════════════
# METRIC COMPUTER
# ═══════════════════════════════════════════════════════════════════

def compute_derived_metrics(data: Dict[str, Any]) -> Tuple[Dict[str, Any], List[str]]:
    """
    Compute any missing derived metrics from available data.
    All computation is deterministic — no LLM.

    Returns:
        (enriched_data, list_of_auto_corrections)
    """
    d = dict(data)
    corrections = []

    def _get(key: str) -> Optional[float]:
        v = d.get(key)
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    revenue = _get("revenue")
    cogs = _get("cogs")
    gross_profit = _get("gross_profit")
    ga_expenses = _get("ga_expenses")
    ebitda = _get("ebitda")
    net_profit = _get("net_profit")
    depreciation = _get("depreciation")
    finance_expense = _get("finance_expense")
    tax_expense = _get("tax_expense")

    # Gross profit
    if gross_profit is None and revenue is not None and cogs is not None:
        gross_profit = revenue - cogs
        d["gross_profit"] = gross_profit
        corrections.append("Computed gross_profit = revenue - cogs")

    # EBITDA
    if ebitda is None and gross_profit is not None and ga_expenses is not None:
        ebitda = gross_profit - ga_expenses
        d["ebitda"] = ebitda
        corrections.append("Computed ebitda = gross_profit - ga_expenses")

    # Net profit (approximate)
    if net_profit is None and ebitda is not None:
        dep = depreciation or 0
        fin = finance_expense or 0
        tax = tax_expense or 0
        net_profit = ebitda - dep - fin - tax
        d["net_profit"] = net_profit
        corrections.append("Computed net_profit = ebitda - depreciation - finance - tax")

    # Percentages
    if revenue and revenue != 0:
        if _get("gross_margin_pct") is None and gross_profit is not None:
            d["gross_margin_pct"] = round(gross_profit / revenue * 100, 2)
            corrections.append("Computed gross_margin_pct")
        if _get("net_margin_pct") is None and net_profit is not None:
            d["net_margin_pct"] = round(net_profit / revenue * 100, 2)
            corrections.append("Computed net_margin_pct")
        if _get("ebitda_margin_pct") is None and ebitda is not None:
            d["ebitda_margin_pct"] = round(ebitda / revenue * 100, 2)
            corrections.append("Computed ebitda_margin_pct")
        if _get("cogs_to_revenue_pct") is None and cogs is not None:
            d["cogs_to_revenue_pct"] = round(cogs / revenue * 100, 2)
            corrections.append("Computed cogs_to_revenue_pct")

    return d, corrections


# ═══════════════════════════════════════════════════════════════════
# SMART EXCEL PARSER
# ═══════════════════════════════════════════════════════════════════

class SmartExcelParser:
    """
    Universal financial Excel/CSV parser.

    Automatically detects headers, maps columns via fuzzy matching,
    normalizes data, and computes derived metrics.
    """

    def parse_file(self, path: str) -> ParseResult:
        """Parse an Excel or CSV file from disk."""
        p = Path(path)
        if not p.exists():
            raise FileNotFoundError(f"File not found: {path}")

        ext = p.suffix.lower()
        if ext in (".xlsx", ".xls"):
            return self._parse_excel(path)
        elif ext == ".csv":
            return self._parse_csv(path)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    def parse_bytes(self, data: bytes, filename: str) -> ParseResult:
        """Parse file from bytes (HTTP upload)."""
        ext = Path(filename).suffix.lower()
        if ext in (".xlsx", ".xls"):
            return self._parse_excel_bytes(data, filename)
        elif ext == ".csv":
            return self._parse_csv_bytes(data, filename)
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    # ── Excel parsing ───────────────────────────────────────────────

    def _parse_excel(self, path: str) -> ParseResult:
        """Parse Excel file from disk path."""
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return self._process_workbook(wb, Path(path).name)

    def _parse_excel_bytes(self, data: bytes, filename: str) -> ParseResult:
        """Parse Excel from bytes."""
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return self._process_workbook(wb, filename)

    def _process_workbook(self, wb, filename: str) -> ParseResult:
        """Process all sheets in a workbook."""
        import openpyxl
        sheets = []
        all_warnings = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))

            if not rows or all(all(c is None for c in r) for r in rows):
                all_warnings.append(f"Sheet '{sheet_name}' is empty — skipped")
                continue

            parsed = self._process_rows(rows, sheet_name)
            sheets.append(parsed)
            all_warnings.extend(parsed.warnings)

        try:
            wb.close()
        except Exception:
            pass

        return self._build_result(sheets, filename, "xlsx", all_warnings)

    # ── CSV parsing ─────────────────────────────────────────────────

    def _parse_csv(self, path: str) -> ParseResult:
        """Parse CSV file from disk."""
        rows = self._read_csv_rows(Path(path).read_text(encoding="utf-8", errors="replace"))
        sheets = [self._process_rows(rows, "Sheet1")]
        return self._build_result(sheets, Path(path).name, "csv", sheets[0].warnings)

    def _parse_csv_bytes(self, data: bytes, filename: str) -> ParseResult:
        """Parse CSV from bytes."""
        text = data.decode("utf-8", errors="replace")
        rows = self._read_csv_rows(text)
        sheets = [self._process_rows(rows, "Sheet1")]
        return self._build_result(sheets, filename, "csv", sheets[0].warnings)

    def _read_csv_rows(self, text: str) -> List[List[Any]]:
        """Read CSV with auto-delimiter detection."""
        # Detect delimiter
        first_line = text.split("\n")[0] if text else ""
        if ";" in first_line and "," not in first_line:
            delimiter = ";"
        elif "\t" in first_line:
            delimiter = "\t"
        else:
            delimiter = ","

        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = []
        for row in reader:
            # Convert numeric strings
            parsed_row = []
            for cell in row:
                cell = cell.strip()
                if not cell:
                    parsed_row.append(None)
                    continue
                # Try numeric
                try:
                    # Handle comma as decimal separator
                    num_str = cell.replace(" ", "").replace("\xa0", "")
                    if "," in num_str and "." not in num_str:
                        num_str = num_str.replace(",", ".")
                    parsed_row.append(float(num_str))
                except ValueError:
                    parsed_row.append(cell)
            rows.append(parsed_row)
        return rows

    # ── Core processing ─────────────────────────────────────────────

    def _process_rows(self, rows: List[List[Any]], sheet_name: str) -> ParsedSheet:
        """Process raw rows from any source into a ParsedSheet."""
        warnings = []

        # Step 1: Detect header row
        header_row = _detect_header_row(rows)
        if header_row > 0:
            warnings.append(f"Header detected at row {header_row + 1} (not row 1)")

        headers = rows[header_row] if header_row < len(rows) else []

        # Step 2: Map columns via fuzzy matching
        mappings: List[ColumnMapping] = []
        mapped_fields: Dict[str, int] = {}  # field -> column index

        for i, header in enumerate(headers):
            header_str = str(header).strip() if header else ""
            field_name, confidence = _fuzzy_match_column(header_str)

            # Avoid duplicate field mappings — keep higher confidence
            if field_name and field_name in mapped_fields:
                existing = mappings[mapped_fields[field_name]]
                if confidence <= existing.confidence:
                    field_name = None
                    confidence = 0
                else:
                    # Replace existing mapping
                    mappings[mapped_fields[field_name]].canonical_field = None
                    mappings[mapped_fields[field_name]].confidence = 0

            mapping = ColumnMapping(
                source_index=i,
                source_header=header_str,
                canonical_field=field_name,
                confidence=confidence,
            )
            mappings.append(mapping)

            if field_name:
                mapped_fields[field_name] = i

        # Warn about unmapped columns
        unmapped = [m for m in mappings if not m.canonical_field and m.source_header]
        for m in unmapped:
            warnings.append(f"Unmapped column [{m.source_index}]: '{m.source_header}'")

        # Step 3: Extract data rows
        records = []
        data_start = header_row + 1
        for row_idx in range(data_start, len(rows)):
            row = rows[row_idx]
            if not row or all(c is None for c in row):
                continue

            record: Dict[str, Any] = {}
            has_data = False

            for mapping in mappings:
                if not mapping.canonical_field:
                    continue
                idx = mapping.source_index
                if idx >= len(row):
                    continue

                val = row[idx]
                if val is None:
                    continue

                # Try to convert to float for numeric fields
                if mapping.canonical_field != "period":
                    try:
                        if isinstance(val, str):
                            num_str = val.strip().replace(" ", "").replace("\xa0", "")
                            num_str = num_str.replace(",", "")
                            val = float(num_str)
                        elif isinstance(val, (int, float)):
                            val = float(val)
                    except (ValueError, TypeError):
                        pass

                record[mapping.canonical_field] = val
                has_data = True

            # Skip summary/total rows (heuristic)
            first_cell = str(row[0]).strip().lower() if row and row[0] else ""
            if first_cell in ("total", "итого", "сул", "grand total", "subtotal"):
                continue

            if has_data:
                records.append(record)

        return ParsedSheet(
            sheet_name=sheet_name,
            header_row=header_row,
            column_mappings=mappings,
            records=records,
            warnings=warnings,
        )

    # ── Result building ─────────────────────────────────────────────

    def _build_result(
        self,
        sheets: List[ParsedSheet],
        filename: str,
        file_type: str,
        warnings: List[str],
    ) -> ParseResult:
        """Build final ParseResult by merging sheets and computing metrics."""
        # Merge all records from all sheets
        all_records = []
        for sheet in sheets:
            all_records.extend(sheet.records)

        # If we have a single-row P&L/BS (common), take it directly
        if len(all_records) == 1:
            merged = dict(all_records[0])
        elif all_records:
            # Multiple records: try to sum numeric fields or take last
            merged: Dict[str, Any] = {}
            for rec in all_records:
                for k, v in rec.items():
                    if k == "period":
                        merged[k] = v  # keep last period
                    elif isinstance(v, (int, float)):
                        merged[k] = merged.get(k, 0) + v
                    else:
                        merged[k] = v
        else:
            merged = {}

        # Compute derived metrics
        enriched, corrections = compute_derived_metrics(merged)

        # Overall confidence
        all_mappings = []
        for s in sheets:
            all_mappings.extend(s.column_mappings)
        mapped = [m for m in all_mappings if m.canonical_field]
        if mapped:
            avg_conf = sum(m.confidence for m in mapped) / len(mapped)
        else:
            avg_conf = 0

        return ParseResult(
            filename=filename,
            file_type=file_type,
            sheets=sheets,
            normalized_financials=enriched,
            confidence_score=int(avg_conf),
            auto_corrections=corrections,
            warnings=warnings,
        )


# Module-level singleton
smart_parser = SmartExcelParser()
