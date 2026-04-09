"""
scheduler.py -- Report Scheduler & Email Sender for FinAI Platform

Periodically checks for scheduled reports that are due, generates
financial summaries from the active dataset, and delivers them via
SMTP email.  Uses only Python stdlib (smtplib, email.mime) so there
are zero extra dependencies beyond what the project already ships.

Typical lifecycle:
    scheduler = ReportScheduler()
    await scheduler.start()       # called once at app startup
    ...
    await scheduler.stop()        # called at app shutdown
"""

import asyncio
import logging
import json
import smtplib
import os
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from app.models.all_models import (
    ScheduledReport,
    Dataset,
    RevenueItem,
    COGSItem,
    GAExpenseItem,
    BudgetLine,
    Anomaly,
    Transaction,
)
from app.config import settings
from app.database import AsyncSessionLocal

logger = logging.getLogger("finai.scheduler")

# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _fmt(value: float, currency: str = "GEL") -> str:
    """Format a monetary value with thousands separator and currency symbol."""
    sign = "-" if value < 0 else ""
    abs_val = abs(value)
    if abs_val >= 1_000_000:
        formatted = f"{abs_val:,.0f}"
    elif abs_val >= 1:
        formatted = f"{abs_val:,.2f}"
    else:
        formatted = f"{abs_val:.2f}"
    return f"{sign}{formatted} {currency}"


def _pct(numerator: float, denominator: float) -> str:
    """Return a percentage string, guarding against division by zero."""
    if denominator == 0:
        return "N/A"
    return f"{(numerator / denominator) * 100:.1f}%"


# ═══════════════════════════════════════════════════════════════════════════
# ReportScheduler
# ═══════════════════════════════════════════════════════════════════════════

class ReportScheduler:
    """Background scheduler that checks for due reports and emails them.

    The scheduler runs a simple polling loop (every ``check_interval``
    seconds) instead of relying on heavy task-queue infrastructure,
    keeping deployment simple for small-to-mid-scale deployments.
    """

    def __init__(self) -> None:
        self.is_running: bool = False
        self._task: Optional[asyncio.Task] = None
        self.check_interval: int = 60  # seconds between schedule checks

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    async def start(self) -> None:
        """Start the background scheduler loop."""
        if self.is_running:
            logger.warning("Scheduler is already running -- ignoring duplicate start()")
            return
        self.is_running = True
        self._task = asyncio.create_task(self._scheduler_loop())
        logger.info(
            "Report scheduler started (checking every %ds)", self.check_interval
        )

    async def stop(self) -> None:
        """Gracefully stop the scheduler loop."""
        self.is_running = False
        if self._task is not None and not self._task.done():
            self._task.cancel()
            try:
                await self._task
            except asyncio.CancelledError:
                pass
        self._task = None
        logger.info("Report scheduler stopped")

    # ------------------------------------------------------------------
    # Core loop
    # ------------------------------------------------------------------

    async def _scheduler_loop(self) -> None:
        """Poll for due schedules in an infinite loop."""
        logger.debug("Scheduler loop entered")
        while self.is_running:
            try:
                async with AsyncSessionLocal() as db:
                    await self._check_schedules(db)
                    await self._resolve_open_predictions(db)
            except asyncio.CancelledError:
                break
            except Exception:
                logger.exception("Error during schedule check")
            try:
                await asyncio.sleep(self.check_interval)
            except asyncio.CancelledError:
                break

    async def _check_schedules(self, db: AsyncSession) -> None:
        """Find all active schedules whose next_run is in the past and process them."""
        now = datetime.utcnow()
        stmt = select(ScheduledReport).where(
            and_(
                ScheduledReport.is_active == True,  # noqa: E712
                ScheduledReport.next_run <= now,
            )
        )
        result = await db.execute(stmt)
        due_schedules: List[ScheduledReport] = list(result.scalars().all())

        if due_schedules:
            logger.info("Found %d due schedule(s) to process", len(due_schedules))

        for schedule in due_schedules:
            try:
                await self._generate_and_send(schedule, db)
            except Exception:
                logger.exception(
                    "Failed to process schedule id=%s name='%s'",
                    schedule.id,
                    schedule.name,
                )
                schedule.last_status = "failed"
                schedule.last_run = datetime.utcnow()
                schedule.next_run = self._compute_next_run(schedule.frequency)
                await db.commit()

    async def _resolve_open_predictions(self, db: AsyncSession) -> None:
        """Evaluate and resolve any open predictions using the active dataset dynamically."""
        from app.models.all_models import PredictionRecord
        from app.services.v2.prediction_tracker import prediction_tracker
        from app.services.ai_agent import agent
        from sqlalchemy import text
        
        q = select(PredictionRecord).where(PredictionRecord.resolved == False)
        open_preds = (await db.execute(q)).scalars().all()
        if not open_preds: return

        active_ds = await db.execute(select(Dataset).where(Dataset.is_active == True))
        active_ds = active_ds.scalar_one_or_none()
        if not active_ds: return

        # Load complete P&L context
        try:
            stmt = await agent._build_stmt(db)
            metrics_dict = stmt.to_dict()
        except:
            return  # skip if statement can't be built

        # Preload segments / products
        rev_q = select(RevenueItem).where(RevenueItem.dataset_id == active_ds.id)
        rev_items = (await db.execute(rev_q)).scalars().all()
        cogs_q = select(COGSItem).where(COGSItem.dataset_id == active_ds.id)
        cogs_items = (await db.execute(cogs_q)).scalars().all()
        ga_q = select(GAExpenseItem).where(GAExpenseItem.dataset_id == active_ds.id)
        ga_items = (await db.execute(ga_q)).scalars().all()

        for p in open_preds:
            actual = None
            raw_metric = p.metric.lower().strip()
            key_metric = raw_metric.replace(" ", "_").replace("-", "_")
            
            # 1. Exact match in standard P&L dict (total_revenue, gross_profit, margin_wholesale_total...)
            if key_metric in metrics_dict:
                actual = metrics_dict[key_metric]
            
            # 2. Dynamic Metric Search: check Revenue items
            elif actual is None:
                for r in rev_items:
                    if r.product and r.product.lower() in raw_metric:
                        actual = r.net
                        break
                    if r.segment and r.segment.lower() in raw_metric:
                        actual = (actual or 0) + (r.net or 0)
                        break

            # 3. Dynamic Metric Search: check COGS items
            if actual is None:
                for c in cogs_items:
                    if c.product and c.product.lower() in raw_metric:
                        actual = c.total_cogs
                        break
                    if c.segment and c.segment.lower() in raw_metric:
                        actual = (actual or 0) + (c.total_cogs or 0)
                        break
                        
            # 4. Dynamic Metric Search: check GA Accounts
            if actual is None:
                for g in ga_items:
                    if g.account_code and g.account_code.lower() in raw_metric:
                        actual = g.amount
                        break
                    if g.account_name and g.account_name.lower() in raw_metric:
                        actual = g.amount
                        break
                        
            # If successfully found actual data from active state, resolve it.
            if actual is not None and not isinstance(actual, (list, dict)):
                try:
                    await prediction_tracker.resolve_prediction(
                        prediction_id=p.id,
                        actual_value=float(actual),
                        db=db
                    )
                    await db.commit()
                except Exception as e:
                    logger.error(f"Error resolving prediction {p.id}: {e}")

    # ------------------------------------------------------------------
    # Report generation & sending
    # ------------------------------------------------------------------

    async def _generate_and_send(
        self, schedule: ScheduledReport, db: AsyncSession
    ) -> None:
        """Generate the report for *schedule*, email it, and update metadata."""
        logger.info(
            "Generating report for schedule id=%s type='%s'",
            schedule.id,
            schedule.report_type,
        )

        # Determine which dataset to pull from (active or parameter-specified)
        dataset = await self._resolve_dataset(schedule, db)
        data = await self._gather_report_data(schedule.report_type, dataset, db)
        data["schedule_name"] = schedule.name
        data["generated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        data["period"] = dataset.period if dataset else "N/A"
        data["currency"] = dataset.currency if dataset else "GEL"
        data["company"] = dataset.company if dataset else getattr(
            settings, "COMPANY_NAME", "NYX Core Thinker LLC"
        )

        html_body = self._build_report_html(schedule.report_type, data)
        excel_path = self._build_report_excel(schedule.report_type, data)

        # Parse recipients
        recipients = schedule.recipients
        if isinstance(recipients, str):
            try:
                recipients = json.loads(recipients)
            except (json.JSONDecodeError, TypeError):
                recipients = [r.strip() for r in recipients.split(",") if r.strip()]

        subject = (
            f"[FinAI] {schedule.name} -- "
            f"{data.get('period', '')} "
            f"({datetime.utcnow().strftime('%Y-%m-%d')})"
        )

        success = self._send_email(
            recipients=recipients,
            subject=subject,
            html_body=html_body,
            attachment_path=excel_path,
        )

        # Update schedule metadata
        now = datetime.utcnow()
        schedule.last_run = now
        schedule.last_status = "success" if success else "failed"
        schedule.next_run = self._compute_next_run(schedule.frequency, from_time=now)
        await db.commit()

        logger.info(
            "Schedule id=%s processed -- status=%s, next_run=%s",
            schedule.id,
            schedule.last_status,
            schedule.next_run.isoformat() if schedule.next_run else "None",
        )

    # ------------------------------------------------------------------
    # Data resolution helpers
    # ------------------------------------------------------------------

    async def _resolve_dataset(
        self, schedule: ScheduledReport, db: AsyncSession
    ) -> Optional[Dataset]:
        """Return the dataset to use for this schedule.

        Priority:
          1. ``schedule.parameters["dataset_id"]`` if specified
          2. The currently active dataset (``is_active=True``)
          3. The most recently created dataset
        """
        params = schedule.parameters or {}
        dataset_id = params.get("dataset_id")

        if dataset_id:
            result = await db.execute(
                select(Dataset).where(Dataset.id == dataset_id)
            )
            ds = result.scalar_one_or_none()
            if ds:
                return ds

        # Fall back to active dataset
        result = await db.execute(
            select(Dataset).where(Dataset.is_active == True).limit(1)  # noqa: E712
        )
        ds = result.scalar_one_or_none()
        if ds:
            return ds

        # Last resort -- most recent dataset
        result = await db.execute(
            select(Dataset).order_by(Dataset.created_at.desc()).limit(1)
        )
        return result.scalar_one_or_none()

    async def _gather_report_data(
        self, report_type: str, dataset: Optional[Dataset], db: AsyncSession
    ) -> Dict:
        """Collect financial figures from the DB for the given report type."""
        data: Dict = {}
        if dataset is None:
            data["error"] = "No dataset available for report generation."
            return data

        ds_id = dataset.id

        if report_type in ("pl_summary", "full_report"):
            data.update(await self._gather_pl_data(ds_id, db))

        if report_type in ("full_report",):
            data.update(await self._gather_budget_data(ds_id, db))

        if report_type in ("anomaly_alert", "full_report"):
            data.update(await self._gather_anomaly_data(ds_id, db))

        if report_type == "custom":
            # Custom reports get everything
            data.update(await self._gather_pl_data(ds_id, db))
            data.update(await self._gather_budget_data(ds_id, db))
            data.update(await self._gather_anomaly_data(ds_id, db))

        return data

    async def _gather_pl_data(self, dataset_id: int, db: AsyncSession) -> Dict:
        """Gather revenue, COGS, G&A, and compute P&L summary figures."""
        # Revenue
        res = await db.execute(
            select(RevenueItem).where(RevenueItem.dataset_id == dataset_id)
        )
        revenue_items = list(res.scalars().all())
        total_revenue = sum(r.net for r in revenue_items)

        rev_by_segment: Dict[str, float] = {}
        for r in revenue_items:
            seg = r.segment or "Other Revenue"
            rev_by_segment[seg] = rev_by_segment.get(seg, 0.0) + r.net

        # COGS
        res = await db.execute(
            select(COGSItem).where(COGSItem.dataset_id == dataset_id)
        )
        cogs_items = list(res.scalars().all())
        total_cogs = sum(c.total_cogs for c in cogs_items)

        cogs_by_segment: Dict[str, float] = {}
        for c in cogs_items:
            seg = c.segment or "Other COGS"
            cogs_by_segment[seg] = cogs_by_segment.get(seg, 0.0) + c.total_cogs

        # G&A
        res = await db.execute(
            select(GAExpenseItem).where(GAExpenseItem.dataset_id == dataset_id)
        )
        ga_items = list(res.scalars().all())
        total_ga = sum(g.amount for g in ga_items)

        gross_profit = total_revenue - total_cogs
        ebitda = gross_profit - total_ga

        return {
            "total_revenue": total_revenue,
            "rev_by_segment": rev_by_segment,
            "total_cogs": total_cogs,
            "cogs_by_segment": cogs_by_segment,
            "gross_profit": gross_profit,
            "gross_margin_pct": (gross_profit / total_revenue * 100)
            if total_revenue
            else 0.0,
            "total_ga": total_ga,
            "ga_items": [
                {"code": g.account_code, "name": g.account_name, "amount": g.amount}
                for g in ga_items
            ],
            "ebitda": ebitda,
            "ebitda_margin_pct": (ebitda / total_revenue * 100)
            if total_revenue
            else 0.0,
        }

    async def _gather_budget_data(self, dataset_id: int, db: AsyncSession) -> Dict:
        """Gather budget vs. actual figures."""
        res = await db.execute(
            select(BudgetLine).where(BudgetLine.dataset_id == dataset_id)
        )
        budget_lines = list(res.scalars().all())
        budget_rows = []
        for b in budget_lines:
            actual = b.actual_amount if b.actual_amount is not None else 0.0
            budget_rows.append(
                {
                    "item": b.line_item,
                    "budget": b.budget_amount,
                    "actual": actual,
                    "variance": actual - b.budget_amount,
                }
            )
        return {"budget_lines": budget_rows}

    async def _gather_anomaly_data(self, dataset_id: int, db: AsyncSession) -> Dict:
        """Gather anomaly detection results."""
        res = await db.execute(
            select(Anomaly)
            .where(Anomaly.dataset_id == dataset_id)
            .order_by(Anomaly.severity.desc(), Anomaly.score.desc())
            .limit(20)
        )
        anomalies = list(res.scalars().all())
        return {
            "anomalies": [a.to_dict() for a in anomalies],
            "anomaly_count": len(anomalies),
        }

    # ------------------------------------------------------------------
    # HTML builder
    # ------------------------------------------------------------------

    def _build_report_html(self, report_type: str, data: Dict) -> str:
        """Build a styled HTML email body for the given report type and data.

        Returns a complete HTML document string suitable for email delivery.
        """
        company = data.get("company", settings.COMPANY_NAME)
        period = data.get("period", "N/A")
        currency = data.get("currency", "GEL")
        generated_at = data.get("generated_at", "")
        schedule_name = data.get("schedule_name", "Financial Report")

        # ---------- shared CSS ----------
        style = """
        <style>
            body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                   margin: 0; padding: 0; background: #f4f6f9; color: #333; }
            .container { max-width: 700px; margin: 20px auto; background: #fff;
                         border-radius: 8px; overflow: hidden;
                         box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
            .header { background: linear-gradient(135deg, #1a237e, #283593);
                      padding: 28px 32px; color: #fff; }
            .header h1 { margin: 0 0 4px; font-size: 22px; letter-spacing: 0.5px; }
            .header .subtitle { opacity: 0.85; font-size: 13px; }
            .body { padding: 28px 32px; }
            h2 { font-size: 17px; color: #1a237e; border-bottom: 2px solid #e8eaf6;
                 padding-bottom: 6px; margin-top: 28px; }
            table { width: 100%; border-collapse: collapse; margin: 12px 0 24px; }
            th { background: #e8eaf6; color: #1a237e; font-size: 12px;
                 text-transform: uppercase; padding: 8px 12px; text-align: left; }
            td { padding: 8px 12px; border-bottom: 1px solid #eee; font-size: 14px; }
            tr:hover td { background: #f5f7ff; }
            .number { text-align: right; font-variant-numeric: tabular-nums; }
            .total-row td { font-weight: 700; border-top: 2px solid #1a237e;
                            background: #f0f1fa; }
            .positive { color: #2e7d32; }
            .negative { color: #c62828; }
            .kpi-grid { display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0; }
            .kpi-card { flex: 1 1 140px; background: #f5f7ff; border-radius: 6px;
                        padding: 14px 16px; text-align: center; }
            .kpi-label { font-size: 11px; text-transform: uppercase; color: #666;
                         margin-bottom: 4px; }
            .kpi-value { font-size: 20px; font-weight: 700; color: #1a237e; }
            .badge { display: inline-block; padding: 2px 8px; border-radius: 4px;
                     font-size: 11px; font-weight: 600; }
            .badge-critical { background: #ffcdd2; color: #b71c1c; }
            .badge-high { background: #ffe0b2; color: #e65100; }
            .badge-medium { background: #fff9c4; color: #f57f17; }
            .badge-low { background: #c8e6c9; color: #1b5e20; }
            .footer { padding: 16px 32px; background: #f4f6f9; font-size: 11px;
                      color: #999; text-align: center; border-top: 1px solid #eee; }
            .error-box { background: #fff3e0; border-left: 4px solid #ff9800;
                         padding: 12px 16px; margin: 12px 0; border-radius: 4px; }
        </style>
        """

        # ---------- header ----------
        header_html = f"""
        <div class="header">
            <h1>{company} &mdash; FinAI Report</h1>
            <div class="subtitle">
                {schedule_name} &bull; Period: {period} &bull; Generated: {generated_at}
            </div>
        </div>
        """

        # ---------- body sections ----------
        sections: List[str] = []

        # Error fallback
        if "error" in data:
            sections.append(
                f'<div class="error-box">{data["error"]}</div>'
            )

        # ---- KPI cards (for pl_summary / full_report) ----
        if "total_revenue" in data:
            sections.append(self._html_kpi_cards(data, currency))

        # ---- P&L section ----
        if report_type in ("pl_summary", "full_report", "custom") and "total_revenue" in data:
            sections.append(self._html_pl_section(data, currency))

        # ---- Budget variance (full_report / custom) ----
        if report_type in ("full_report", "custom") and data.get("budget_lines"):
            sections.append(self._html_budget_section(data, currency))

        # ---- Anomalies (anomaly_alert / full_report / custom) ----
        if report_type in ("anomaly_alert", "full_report", "custom") and data.get("anomalies"):
            sections.append(self._html_anomaly_section(data))

        # ---- Custom note ----
        if report_type == "custom":
            sections.append(
                "<p><em>This is a custom report. "
                "Contact finance@nyxcore.tech to adjust contents.</em></p>"
            )

        # ---------- footer ----------
        footer_html = """
        <div class="footer">
            Automated report by FinAI Financial Intelligence Platform<br>
            &copy; {settings.COMPANY_NAME} &mdash; Confidential
        </div>
        """

        body_content = "\n".join(sections) if sections else (
            '<p style="color:#999;">No report data available for this run.</p>'
        )

        return (
            "<!DOCTYPE html>"
            "<html lang='en'><head><meta charset='UTF-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            f"<title>{schedule_name}</title>{style}</head>"
            f"<body><div class='container'>{header_html}"
            f"<div class='body'>{body_content}</div>"
            f"{footer_html}</div></body></html>"
        )

    # ---- HTML sub-sections ----

    @staticmethod
    def _html_kpi_cards(data: Dict, currency: str) -> str:
        """Render the top-line KPI cards."""
        cards = [
            ("Total Revenue", _fmt(data.get("total_revenue", 0), currency)),
            ("Gross Profit", _fmt(data.get("gross_profit", 0), currency)),
            ("Gross Margin", _pct(data.get("gross_profit", 0), data.get("total_revenue", 1))),
            ("EBITDA", _fmt(data.get("ebitda", 0), currency)),
            ("EBITDA Margin", _pct(data.get("ebitda", 0), data.get("total_revenue", 1))),
        ]
        html = '<div class="kpi-grid">'
        for label, value in cards:
            html += (
                f'<div class="kpi-card">'
                f'<div class="kpi-label">{label}</div>'
                f'<div class="kpi-value">{value}</div>'
                f'</div>'
            )
        html += "</div>"
        return html

    @staticmethod
    def _html_pl_section(data: Dict, currency: str) -> str:
        """Render the P&L summary table."""
        html = "<h2>Profit &amp; Loss Summary</h2><table>"
        html += "<tr><th>Line Item</th><th class='number'>Amount</th></tr>"

        # Revenue by segment
        for seg, val in sorted(data.get("rev_by_segment", {}).items()):
            html += (
                f"<tr><td>&nbsp;&nbsp;{seg}</td>"
                f"<td class='number'>{_fmt(val, currency)}</td></tr>"
            )
        html += (
            f"<tr class='total-row'><td>Total Revenue</td>"
            f"<td class='number'>{_fmt(data.get('total_revenue', 0), currency)}</td></tr>"
        )

        # COGS by segment
        for seg, val in sorted(data.get("cogs_by_segment", {}).items()):
            css = "negative" if val > 0 else ""
            html += (
                f"<tr><td>&nbsp;&nbsp;{seg}</td>"
                f"<td class='number {css}'>({_fmt(val, currency)})</td></tr>"
            )
        html += (
            f"<tr class='total-row'><td>Total COGS</td>"
            f"<td class='number negative'>({_fmt(data.get('total_cogs', 0), currency)})</td></tr>"
        )

        # Gross profit
        gp = data.get("gross_profit", 0)
        gp_css = "positive" if gp >= 0 else "negative"
        html += (
            f"<tr class='total-row'><td>Gross Profit</td>"
            f"<td class='number {gp_css}'>{_fmt(gp, currency)}</td></tr>"
        )

        # G&A
        total_ga = data.get("total_ga", 0)
        if total_ga:
            html += (
                f"<tr><td>General &amp; Administrative Expenses</td>"
                f"<td class='number negative'>({_fmt(total_ga, currency)})</td></tr>"
            )

        # EBITDA
        ebitda = data.get("ebitda", 0)
        ebitda_css = "positive" if ebitda >= 0 else "negative"
        html += (
            f"<tr class='total-row'><td>EBITDA</td>"
            f"<td class='number {ebitda_css}'>{_fmt(ebitda, currency)}</td></tr>"
        )

        html += "</table>"
        return html

    @staticmethod
    def _html_budget_section(data: Dict, currency: str) -> str:
        """Render the budget vs. actual variance table."""
        html = "<h2>Budget vs. Actual</h2><table>"
        html += (
            "<tr><th>Line Item</th><th class='number'>Budget</th>"
            "<th class='number'>Actual</th><th class='number'>Variance</th></tr>"
        )
        for row in data.get("budget_lines", []):
            var_val = row.get("variance", 0)
            var_css = "positive" if var_val >= 0 else "negative"
            html += (
                f"<tr>"
                f"<td>{row.get('item', '')}</td>"
                f"<td class='number'>{_fmt(row.get('budget', 0), currency)}</td>"
                f"<td class='number'>{_fmt(row.get('actual', 0), currency)}</td>"
                f"<td class='number {var_css}'>{_fmt(var_val, currency)}</td>"
                f"</tr>"
            )
        html += "</table>"
        return html

    @staticmethod
    def _html_anomaly_section(data: Dict) -> str:
        """Render the anomaly alerts table."""
        anomalies = data.get("anomalies", [])
        count = data.get("anomaly_count", len(anomalies))
        html = f"<h2>Anomaly Alerts ({count})</h2><table>"
        html += (
            "<tr><th>Severity</th><th>Type</th>"
            "<th>Description</th><th class='number'>Score</th></tr>"
        )
        for a in anomalies:
            sev = a.get("severity", "medium")
            badge_cls = f"badge-{sev}"
            html += (
                f"<tr>"
                f"<td><span class='badge {badge_cls}'>{sev.upper()}</span></td>"
                f"<td>{a.get('anomaly_type', 'N/A')}</td>"
                f"<td>{a.get('description', '')}</td>"
                f"<td class='number'>{a.get('score', 0)}</td>"
                f"</tr>"
            )
        html += "</table>"
        return html

    # ------------------------------------------------------------------
    # Excel attachment builder
    # ------------------------------------------------------------------

    def _build_report_excel(self, report_type: str, data: Dict) -> Optional[str]:
        """Build an Excel workbook summarising the report data.

        Returns the file path on success, ``None`` on failure.
        The caller is responsible for cleanup (the file lives in ./exports/).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            logger.warning("openpyxl not installed -- skipping Excel attachment")
            return None

        try:
            export_dir = os.path.join(".", "exports")
            os.makedirs(export_dir, exist_ok=True)

            wb = Workbook()
            company = data.get("company", settings.COMPANY_NAME)
            period = data.get("period", "N/A")
            currency = data.get("currency", "GEL")
            generated = data.get("generated_at", "")

            hdr_fill = PatternFill("solid", fgColor="1A237E")
            hdr_font = Font(color="FFFFFF", bold=True, size=11)
            total_fill = PatternFill("solid", fgColor="E8EAF6")
            total_font = Font(bold=True, size=11)
            num_align = Alignment(horizontal="right")
            thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

            def _write_header(ws, title: str, cols: int):
                ws.append([company] + [""] * (cols - 1))
                ws.append([title] + [""] * (cols - 1))
                ws.append([f"Period: {period}  |  Generated: {generated}"] + [""] * (cols - 1))
                ws.append([])
                for c in range(1, cols + 1):
                    for r in range(1, 4):
                        cell = ws.cell(r, c)
                        cell.font = Font(bold=True, size=12 if r == 1 else 10,
                                         color="1A237E")

            def _style_header_row(ws, row_num, cols):
                for c in range(1, cols + 1):
                    cell = ws.cell(row_num, c)
                    cell.fill = hdr_fill
                    cell.font = hdr_font

            # ── Sheet 1: P&L Summary ──
            if "total_revenue" in data:
                ws = wb.active
                ws.title = "P&L Summary"
                cols = 2
                _write_header(ws, "Profit & Loss Summary", cols)
                ws.append(["Line Item", f"Amount ({currency})"])
                _style_header_row(ws, ws.max_row, cols)

                for seg, val in sorted(data.get("rev_by_segment", {}).items()):
                    ws.append([f"  {seg}", val])
                ws.append(["Total Revenue", data.get("total_revenue", 0)])
                r = ws.max_row
                for c in range(1, cols + 1):
                    ws.cell(r, c).fill = total_fill
                    ws.cell(r, c).font = total_font
                ws.append([])

                for seg, val in sorted(data.get("cogs_by_segment", {}).items()):
                    ws.append([f"  {seg}", -abs(val)])
                ws.append(["Total COGS", -abs(data.get("total_cogs", 0))])
                r = ws.max_row
                for c in range(1, cols + 1):
                    ws.cell(r, c).fill = total_fill
                    ws.cell(r, c).font = total_font
                ws.append([])

                ws.append(["Gross Profit", data.get("gross_profit", 0)])
                r = ws.max_row
                for c in range(1, cols + 1):
                    ws.cell(r, c).fill = total_fill
                    ws.cell(r, c).font = total_font

                if data.get("total_ga"):
                    ws.append(["G&A Expenses", -abs(data["total_ga"])])

                ws.append(["EBITDA", data.get("ebitda", 0)])
                r = ws.max_row
                for c in range(1, cols + 1):
                    ws.cell(r, c).fill = total_fill
                    ws.cell(r, c).font = total_font

                # G&A detail
                if data.get("ga_items"):
                    ws.append([])
                    ws.append(["G&A Expense Detail", ""])
                    ws.append(["Account", f"Amount ({currency})"])
                    _style_header_row(ws, ws.max_row, cols)
                    for g in data["ga_items"]:
                        label = f"{g.get('code', '')} {g.get('name', '')}".strip()
                        ws.append([label, g.get("amount", 0)])

                ws.column_dimensions["A"].width = 40
                ws.column_dimensions["B"].width = 22
                for row in ws.iter_rows(min_row=6, max_col=2):
                    row[1].number_format = '#,##0.00'
                    row[1].alignment = num_align
            else:
                ws = wb.active
                ws.title = "Report"

            # ── Sheet 2: Budget vs Actual ──
            if data.get("budget_lines"):
                ws2 = wb.create_sheet("Budget vs Actual")
                cols = 4
                _write_header(ws2, "Budget vs Actual", cols)
                ws2.append(["Line Item", "Budget", "Actual", "Variance"])
                _style_header_row(ws2, ws2.max_row, cols)
                for row in data["budget_lines"]:
                    ws2.append([
                        row.get("item", ""),
                        row.get("budget", 0),
                        row.get("actual", 0),
                        row.get("variance", 0),
                    ])
                ws2.column_dimensions["A"].width = 40
                for col in ["B", "C", "D"]:
                    ws2.column_dimensions[col].width = 20
                for row in ws2.iter_rows(min_row=6, max_col=4):
                    for cell in row[1:]:
                        cell.number_format = '#,##0.00'
                        cell.alignment = num_align

            # ── Sheet 3: Anomalies ──
            if data.get("anomalies"):
                ws3 = wb.create_sheet("Anomalies")
                cols = 4
                _write_header(ws3, "Anomaly Alerts", cols)
                ws3.append(["Severity", "Type", "Description", "Score"])
                _style_header_row(ws3, ws3.max_row, cols)
                for a in data["anomalies"]:
                    ws3.append([
                        a.get("severity", "").upper(),
                        a.get("anomaly_type", ""),
                        a.get("description", ""),
                        a.get("score", 0),
                    ])
                ws3.column_dimensions["A"].width = 12
                ws3.column_dimensions["B"].width = 20
                ws3.column_dimensions["C"].width = 50
                ws3.column_dimensions["D"].width = 10

            # Save
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            safe_type = report_type.replace(" ", "_")
            filename = f"FinAI_{safe_type}_{ts}.xlsx"
            filepath = os.path.join(export_dir, filename)
            wb.save(filepath)
            logger.info("Excel report saved to %s", filepath)
            return filepath

        except Exception:
            logger.exception("Failed to build Excel report attachment")
            return None

    # ------------------------------------------------------------------
    # Email sending
    # ------------------------------------------------------------------

    def _send_email(
        self,
        recipients: List[str],
        subject: str,
        html_body: str,
        attachment_path: Optional[str] = None,
    ) -> bool:
        """Send an HTML email via SMTP.

        All SMTP settings are read through ``getattr(settings, ...)`` with
        safe defaults so the application never crashes if they are absent.

        Returns ``True`` on success, ``False`` on any failure.
        """
        smtp_enabled = (
            getattr(settings, "SMTP_ENABLED", False)
            or os.getenv("SMTP_ENABLED", "false").lower() == "true"
        )
        if not smtp_enabled:
            logger.warning(
                "SMTP not enabled -- email not sent (subject='%s', recipients=%s)",
                subject,
                recipients,
            )
            return False

        if not recipients:
            logger.warning("No recipients specified -- email not sent")
            return False

        smtp_host = getattr(settings, "SMTP_HOST", None) or os.getenv(
            "SMTP_HOST", "smtp.gmail.com"
        )
        smtp_port = int(
            getattr(settings, "SMTP_PORT", None) or os.getenv("SMTP_PORT", "587")
        )
        smtp_user = getattr(settings, "SMTP_USER", None) or os.getenv(
            "SMTP_USER", ""
        )
        smtp_password = getattr(settings, "SMTP_PASSWORD", None) or os.getenv(
            "SMTP_PASSWORD", ""
        )
        smtp_from = getattr(settings, "SMTP_FROM", None) or os.getenv(
            "SMTP_FROM", "FinAI Reports <reports@nyxcore.tech>"
        )

        try:
            msg = MIMEMultipart("mixed")
            msg["From"] = smtp_from
            msg["To"] = ", ".join(recipients)
            msg["Subject"] = subject

            # HTML body
            html_part = MIMEText(html_body, "html", "utf-8")
            msg.attach(html_part)

            # Optional file attachment
            if attachment_path and os.path.isfile(attachment_path):
                filename = os.path.basename(attachment_path)
                with open(attachment_path, "rb") as fp:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(fp.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition", f"attachment; filename={filename}"
                )
                msg.attach(part)

            # Send
            with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
                server.ehlo()
                server.starttls()
                server.ehlo()
                if smtp_user and smtp_password:
                    server.login(smtp_user, smtp_password)
                server.sendmail(smtp_from, recipients, msg.as_string())

            logger.info(
                "Email sent successfully -- subject='%s' recipients=%s",
                subject,
                recipients,
            )
            return True

        except smtplib.SMTPAuthenticationError:
            logger.error(
                "SMTP authentication failed for user='%s' host='%s'",
                smtp_user,
                smtp_host,
            )
            return False
        except smtplib.SMTPConnectError:
            logger.error("Cannot connect to SMTP server %s:%s", smtp_host, smtp_port)
            return False
        except Exception:
            logger.exception("Unexpected error sending email")
            return False

    # ------------------------------------------------------------------
    # Schedule timing
    # ------------------------------------------------------------------

    @staticmethod
    def _compute_next_run(
        frequency: str, from_time: Optional[datetime] = None
    ) -> datetime:
        """Compute the next run datetime based on frequency.

        Supported frequencies: ``daily``, ``weekly``, ``monthly``.
        """
        base = from_time or datetime.utcnow()

        if frequency == "daily":
            return base + timedelta(days=1)

        if frequency == "weekly":
            return base + timedelta(weeks=1)

        if frequency == "monthly":
            # Advance to the same day next month, clamping to month-end if needed
            month = base.month + 1
            year = base.year
            if month > 12:
                month = 1
                year += 1
            # Clamp day to the last valid day of the target month
            import calendar

            max_day = calendar.monthrange(year, month)[1]
            day = min(base.day, max_day)
            return base.replace(year=year, month=month, day=day)

        # Unknown frequency -- default to daily
        logger.warning(
            "Unknown frequency '%s' -- defaulting to daily interval", frequency
        )
        return base + timedelta(days=1)

    # ------------------------------------------------------------------
    # Public API: test email
    # ------------------------------------------------------------------

    async def send_test_email(
        self, db: AsyncSession, schedule_id: int
    ) -> Dict:
        """Send a test email for a given schedule to its first recipient.

        Returns a dict with ``success``, ``message``, and ``recipient`` keys.
        """
        result = await db.execute(
            select(ScheduledReport).where(ScheduledReport.id == schedule_id)
        )
        schedule = result.scalar_one_or_none()
        if schedule is None:
            return {
                "success": False,
                "message": f"Schedule id={schedule_id} not found",
                "recipient": None,
            }

        # Parse recipients
        recipients = schedule.recipients
        if isinstance(recipients, str):
            try:
                recipients = json.loads(recipients)
            except (json.JSONDecodeError, TypeError):
                recipients = [r.strip() for r in recipients.split(",") if r.strip()]

        if not recipients:
            return {
                "success": False,
                "message": "No recipients configured for this schedule",
                "recipient": None,
            }

        first_recipient = recipients[0]

        # Generate a lightweight test report
        dataset = await self._resolve_dataset(schedule, db)
        data = await self._gather_report_data(schedule.report_type, dataset, db)
        data["schedule_name"] = f"[TEST] {schedule.name}"
        data["generated_at"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        data["period"] = dataset.period if dataset else "N/A"
        data["currency"] = dataset.currency if dataset else "GEL"
        data["company"] = dataset.company if dataset else getattr(
            settings, "COMPANY_NAME", "NYX Core Thinker LLC"
        )

        html_body = self._build_report_html(schedule.report_type, data)
        excel_path = self._build_report_excel(schedule.report_type, data)
        subject = (
            f"[FinAI TEST] {schedule.name} -- "
            f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
        )

        success = self._send_email(
            recipients=[first_recipient],
            subject=subject,
            html_body=html_body,
            attachment_path=excel_path,
        )

        return {
            "success": success,
            "message": "Test email sent successfully" if success else "Failed to send test email (check SMTP settings)",
            "recipient": first_recipient,
        }

    # ------------------------------------------------------------------
    # Public API: CRUD
    # ------------------------------------------------------------------

    async def create_schedule(
        self,
        db: AsyncSession,
        name: str,
        report_type: str,
        frequency: str,
        recipients: List[str],
        is_active: bool = True,
        parameters: Optional[Dict] = None,
    ) -> Dict:
        """Create a new ScheduledReport in the database.

        Returns the newly created schedule as a dict.
        """
        next_run = self._compute_next_run(frequency)

        schedule = ScheduledReport(
            name=name,
            report_type=report_type,
            frequency=frequency,
            recipients=recipients,
            parameters=parameters or {},
            next_run=next_run,
            last_status="pending",
            is_active=is_active,
        )
        db.add(schedule)
        await db.commit()
        await db.refresh(schedule)

        logger.info(
            "Created schedule id=%s name='%s' type='%s' freq='%s' next_run=%s",
            schedule.id,
            name,
            report_type,
            frequency,
            next_run.isoformat(),
        )
        return schedule.to_dict()

    async def update_schedule(
        self, db: AsyncSession, schedule_id: int, updates: Dict
    ) -> Dict:
        """Update an existing schedule with the provided field values.

        Supports updating: ``name``, ``report_type``, ``frequency``,
        ``recipients``, ``is_active``, ``parameters``.

        If ``frequency`` changes, ``next_run`` is automatically recomputed.

        Returns the updated schedule as a dict, or an error dict if not found.
        """
        result = await db.execute(
            select(ScheduledReport).where(ScheduledReport.id == schedule_id)
        )
        schedule = result.scalar_one_or_none()
        if schedule is None:
            return {"error": f"Schedule id={schedule_id} not found"}

        allowed_fields = {
            "name", "report_type", "frequency", "recipients",
            "is_active", "parameters",
        }
        frequency_changed = False

        for key, value in updates.items():
            if key not in allowed_fields:
                logger.debug("Ignoring unknown update field '%s'", key)
                continue
            if key == "frequency" and value != schedule.frequency:
                frequency_changed = True
            setattr(schedule, key, value)

        if frequency_changed:
            schedule.next_run = self._compute_next_run(schedule.frequency)
            logger.info(
                "Schedule id=%s frequency changed to '%s' -- next_run=%s",
                schedule_id,
                schedule.frequency,
                schedule.next_run.isoformat(),
            )

        await db.commit()
        await db.refresh(schedule)

        logger.info("Updated schedule id=%s fields=%s", schedule_id, list(updates.keys()))
        return schedule.to_dict()


# ═══════════════════════════════════════════════════════════════════════════
# Module-level singleton
# ═══════════════════════════════════════════════════════════════════════════

report_scheduler = ReportScheduler()
