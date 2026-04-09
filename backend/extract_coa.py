"""
Extract 4-digit base accounts from AccountN.xlsx
Filters out sub-accounts (those with dots in the code).
Extracts the Georgian part of bilingual account names (before //).
"""

import sys
import io
import re
from openpyxl import load_workbook

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

FILEPATH = r"C:\Users\Nino\OneDrive\Desktop\AccountN.xlsx"

wb = load_workbook(FILEPATH, read_only=True, data_only=True)
ws = wb.active

print(f"{'CODE':<8} | {'TYPE':<30} | {'NAME'}")
print("-" * 90)

count = 0
for row in ws.iter_rows(min_row=1, values_only=False):
    # Column B = index 1, Column D = index 3, Column F = index 5
    account_code_cell = row[1].value if len(row) > 1 else None
    account_name_cell = row[3].value if len(row) > 3 else None
    account_type_cell = row[5].value if len(row) > 5 else None

    if account_code_cell is None:
        continue

    code_str = str(account_code_cell).strip()

    # Only pure 4-digit numbers (no dots, no letters, no sub-accounts)
    if not re.fullmatch(r"\d{4}", code_str):
        continue

    # Extract Georgian part (before //)
    name = str(account_name_cell).strip() if account_name_cell else ""
    if "//" in name:
        name = name.split("//")[0].strip()

    acc_type = str(account_type_cell).strip() if account_type_cell else ""
    if "//" in acc_type:
        acc_type = acc_type.split("//")[0].strip()

    print(f"{code_str:<8} | {acc_type:<30} | {name}")
    count += 1

wb.close()
print("-" * 90)
print(f"Total base accounts found: {count}")
