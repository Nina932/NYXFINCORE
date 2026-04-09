"""
FinAI Modern Excel Report Generator
===================================
Generates top-quality, modern Excel reports with professional styling,
advanced formatting, and AI-enhanced design suitable for executive presentations.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers, GradientFill
    )
    from openpyxl.styles.fills import Stop
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from openpyxl.chart import BarChart, LineChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════
# MODERN STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════

# Modern color palette
MODERN_COLORS = {
    "primary": "2563eb",        # Modern blue
    "primary_dark": "1d4ed8",   # Darker blue
    "secondary": "64748b",      # Slate gray
    "accent": "f59e0b",         # Amber
    "success": "10b981",        # Emerald
    "warning": "f59e0b",        # Amber
    "danger": "ef4444",         # Red
    "light_bg": "f8fafc",       # Very light gray
    "card_bg": "ffffff",        # White
    "text_dark": "1e293b",      # Dark slate
    "text_light": "64748b",     # Light slate
    "border_light": "e2e8f0",   # Light border
    "gradient_start": "2563eb", # Gradient start
    "gradient_end": "1d4ed8",   # Gradient end
}

# Modern fonts
MODERN_FONTS = {
    "title": Font(name="Inter", size=18, bold=True, color="FFFFFF"),
    "subtitle": Font(name="Inter", size=12, bold=True, color="FFFFFF"),
    "header": Font(name="Inter", size=11, bold=True, color="FFFFFF"),
    "bold": Font(name="Inter", size=10, bold=True, color=MODERN_COLORS["text_dark"]),
    "normal": Font(name="Inter", size=10, color=MODERN_COLORS["text_dark"]),
    "small": Font(name="Inter", size=9, color=MODERN_COLORS["text_light"]),
    "kpi_value": Font(name="Inter", size=16, bold=True, color=MODERN_COLORS["primary"]),
    "kpi_label": Font(name="Inter", size=9, color=MODERN_COLORS["text_light"]),
    "positive": Font(name="Inter", size=10, bold=True, color=MODERN_COLORS["success"]),
    "negative": Font(name="Inter", size=10, bold=True, color=MODERN_COLORS["danger"]),
}

# Modern fills
MODERN_FILLS = {
    "primary_gradient": GradientFill(
        stop=[
            Stop(MODERN_COLORS["gradient_start"], 0),
            Stop(MODERN_COLORS["gradient_end"], 1)
        ],
        degree=90
    ),
    "primary_solid": PatternFill("solid", fgColor=MODERN_COLORS["primary"]),
    "secondary_solid": PatternFill("solid", fgColor=MODERN_COLORS["secondary"]),
    "accent_solid": PatternFill("solid", fgColor=MODERN_COLORS["accent"]),
    "light_bg": PatternFill("solid", fgColor=MODERN_COLORS["light_bg"]),
    "white": PatternFill("solid", fgColor=MODERN_COLORS["card_bg"]),
    "success_light": PatternFill("solid", fgColor="d1fae5"),  # Light green
    "danger_light": PatternFill("solid", fgColor="fee2e2"),   # Light red
}

# Modern borders
MODERN_BORDERS = {
    "thin": Border(
        left=Side(style='thin', color=MODERN_COLORS["border_light"]),
        right=Side(style='thin', color=MODERN_COLORS["border_light"]),
        top=Side(style='thin', color=MODERN_COLORS["border_light"]),
        bottom=Side(style='thin', color=MODERN_COLORS["border_light"]),
    ),
    "medium": Border(
        left=Side(style='medium', color=MODERN_COLORS["primary"]),
        right=Side(style='medium', color=MODERN_COLORS["primary"]),
        top=Side(style='medium', color=MODERN_COLORS["primary"]),
        bottom=Side(style='medium', color=MODERN_COLORS["primary"]),
    ),
    "thick": Border(
        left=Side(style='thick', color=MODERN_COLORS["primary_dark"]),
        right=Side(style='thick', color=MODERN_COLORS["primary_dark"]),
        top=Side(style='thick', color=MODERN_COLORS["primary_dark"]),
        bottom=Side(style='thick', color=MODERN_COLORS["primary_dark"]),
    ),
}

# Number formats
MODERN_FORMATS = {
    "currency": '#,##0;(#,##0);"-"',
    "percentage": '0.0%;(0.0%);"-"',
    "number": '#,##0;(#,##0);"-"',
    "decimal": '#,##0.00;(#,##0.00);"-"',
}


def _auto_width(ws, min_width=12, max_width=50):
    """Auto-fit column widths with modern constraints."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value or ""))
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)


class ModernExcelGenerator:
    """Generates modern, professional Excel reports."""

    def generate_cash_runway_report(
        self,
        company_name: str,
        period: str,
        cash_balance: float,
        monthly_burn: float,
        runway_months: float,
        projection_data: List[Dict[str, Any]],
        risk_level: str,
    ) -> bytes:
        """
        Generate a modern Cash Runway Analysis Excel report.

        Args:
            company_name: Name of the company
            period: Reporting period
            cash_balance: Current cash balance
            monthly_burn: Monthly cash burn rate
            runway_months: Number of months of runway
            projection_data: Monthly cash projection data
            risk_level: Risk assessment level

        Returns:
            Excel file bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required for Excel generation")

        wb = openpyxl.Workbook()

        # Executive Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        self._build_modern_summary_sheet(
            ws_summary, company_name, period, cash_balance,
            monthly_burn, runway_months, risk_level
        )

        # Cash Projection Sheet
        ws_projection = wb.create_sheet("Cash Projection")
        self._build_modern_projection_sheet(ws_projection, projection_data, period)

        # Analysis Sheet
        ws_analysis = wb.create_sheet("Analysis")
        self._build_modern_analysis_sheet(ws_analysis, cash_balance, monthly_burn, runway_months)

        # Apply modern styling to all sheets
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            _auto_width(ws)

        # Save to bytes
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _build_modern_summary_sheet(
        self, ws, company_name, period, cash_balance,
        monthly_burn, runway_months, risk_level
    ):
        """Build modern executive summary sheet."""
        # Header with gradient
        ws.merge_cells("A1:F1")
        ws["A1"] = f"💰 Cash Runway Analysis - {company_name}"
        ws["A1"].font = MODERN_FONTS["title"]
        ws["A1"].fill = MODERN_FILLS["primary_gradient"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 50

        # Period info
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Analysis Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = MODERN_FONTS["small"]
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 25

        # KPI Cards
        kpis = [
            ("Current Cash Balance", cash_balance, MODERN_COLORS["success"], "₾"),
            ("Monthly Burn Rate", monthly_burn, MODERN_COLORS["danger"], "₾"),
            ("Runway Months", runway_months, MODERN_COLORS["primary"], "months"),
            ("Risk Level", risk_level.upper(), self._get_risk_color(risk_level), ""),
        ]

        row = 4
        for i, (label, value, color, suffix) in enumerate(kpis):
            col_start = (i % 2) * 3 + 1
            col_end = col_start + 2

            # KPI Card background
            for c in range(col_start, col_end + 1):
                ws.cell(row=row, column=c).fill = MODERN_FILLS["light_bg"]
                ws.cell(row=row + 1, column=c).fill = MODERN_FILLS["white"]
                ws.cell(row=row + 2, column=c).fill = MODERN_FILLS["white"]

            # KPI Value
            ws.merge_cells(f"{get_column_letter(col_start)}{row}:{get_column_letter(col_end)}{row}")
            cell = ws.cell(row=row, column=col_start)
            if isinstance(value, str):
                cell.value = value
                cell.font = Font(name="Inter", size=14, bold=True, color=color)
            else:
                cell.value = value
                cell.font = Font(name="Inter", size=18, bold=True, color=color)
                cell.number_format = MODERN_FORMATS["currency"] if "₾" in suffix else MODERN_FORMATS["decimal"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # KPI Label
            ws.merge_cells(f"{get_column_letter(col_start)}{row+1}:{get_column_letter(col_end)}{row+1}")
            cell = ws.cell(row=row+1, column=col_start)
            cell.value = label
            cell.font = MODERN_FONTS["kpi_label"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Suffix
            if suffix:
                ws.cell(row=row+2, column=col_start + 1).value = suffix
                ws.cell(row=row+2, column=col_start + 1).font = MODERN_FONTS["small"]
                ws.cell(row=row+2, column=col_start + 1).alignment = Alignment(horizontal="center")

            if i % 2 == 1:
                row += 4

        # Risk Assessment
        row += 2
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1).value = "🎯 Risk Assessment & Recommendations"
        ws.cell(row=row, column=1).font = Font(name="Inter", size=14, bold=True, color=MODERN_COLORS["text_dark"])
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 35

        row += 1
        recommendations = self._get_runway_recommendations(risk_level, runway_months)
        for rec in recommendations:
            ws.merge_cells(f"A{row}:F{row}")
            ws.cell(row=row, column=1).value = f"• {rec}"
            ws.cell(row=row, column=1).font = MODERN_FONTS["normal"]
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
            ws.row_dimensions[row].height = 25
            row += 1

    def _build_modern_projection_sheet(self, ws, projection_data, period):
        """Build modern cash projection sheet with charts."""
        # Header
        ws.merge_cells("A1:D1")
        ws["A1"] = f"📈 Cash Flow Projection - {period}"
        ws["A1"].font = MODERN_FONTS["title"]
        ws["A1"].fill = MODERN_FILLS["primary_gradient"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Column headers
        headers = ["Month", "Cash Balance", "Monthly Burn", "Risk Indicator"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i)
            cell.value = header
            cell.font = MODERN_FONTS["header"]
            cell.fill = MODERN_FILLS["secondary_solid"]
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = MODERN_BORDERS["thin"]

        # Data rows
        for i, data_point in enumerate(projection_data[:24], 4):  # Limit to 24 months
            month = data_point.get("month", i-3)
            cash = data_point.get("cash", 0)
            burn = data_point.get("burn", 0)
            risk = "High" if cash <= 0 else "Medium" if cash < burn * 3 else "Low"

            ws.cell(row=i, column=1).value = f"Month {month}"
            ws.cell(row=i, column=2).value = cash
            ws.cell(row=i, column=3).value = burn
            ws.cell(row=i, column=4).value = risk

            # Formatting
            ws.cell(row=i, column=2).number_format = MODERN_FORMATS["currency"]
            ws.cell(row=i, column=3).number_format = MODERN_FORMATS["currency"]

            # Conditional formatting for risk
            if risk == "High":
                ws.cell(row=i, column=4).font = MODERN_FONTS["negative"]
                for col in [1, 2, 3, 4]:
                    ws.cell(row=i, column=col).fill = MODERN_FILLS["danger_light"]
            elif risk == "Medium":
                ws.cell(row=i, column=4).font = Font(name="Inter", size=10, bold=True, color=MODERN_COLORS["warning"])

            # Borders
            for col in range(1, 5):
                ws.cell(row=i, column=col).border = MODERN_BORDERS["thin"]

    def _build_modern_analysis_sheet(self, ws, cash_balance, monthly_burn, runway_months):
        """Build modern analysis sheet with insights."""
        # Header
        ws.merge_cells("A1:C1")
        ws["A1"] = "🔍 Financial Analysis & Insights"
        ws["A1"].font = MODERN_FONTS["title"]
        ws["A1"].fill = MODERN_FILLS["primary_gradient"]
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        # Key Metrics Analysis
        row = 3
        analyses = [
            ("Cash Position", self._analyze_cash_position(cash_balance)),
            ("Burn Rate", self._analyze_burn_rate(monthly_burn)),
            ("Runway Duration", self._analyze_runway(runway_months)),
            ("Recommendations", self._get_runway_insights(cash_balance, monthly_burn, runway_months)),
        ]

        for title, content in analyses:
            # Section header
            ws.merge_cells(f"A{row}:C{row}")
            ws.cell(row=row, column=1).value = title
            ws.cell(row=row, column=1).font = Font(name="Inter", size=12, bold=True, color=MODERN_COLORS["primary"])
            ws.cell(row=row, column=1).fill = MODERN_FILLS["light_bg"]
            ws.row_dimensions[row].height = 30
            row += 1

            # Content
            if isinstance(content, list):
                for item in content:
                    ws.cell(row=row, column=1).value = f"• {item}"
                    ws.cell(row=row, column=1).font = MODERN_FONTS["normal"]
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
                    ws.row_dimensions[row].height = 25
                    row += 1
            else:
                ws.merge_cells(f"A{row}:C{row+1}")
                ws.cell(row=row, column=1).value = content
                ws.cell(row=row, column=1).font = MODERN_FONTS["normal"]
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
                ws.row_dimensions[row].height = 40
                row += 2

            row += 1  # Spacing

    def _get_risk_color(self, risk_level: str) -> str:
        """Get color for risk level."""
        colors = {
            "safe": MODERN_COLORS["success"],
            "low": MODERN_COLORS["success"],
            "medium": MODERN_COLORS["warning"],
            "high": MODERN_COLORS["danger"],
            "critical": MODERN_COLORS["danger"],
        }
        return colors.get(risk_level.lower(), MODERN_COLORS["secondary"])

    def _get_runway_recommendations(self, risk_level: str, runway_months: float) -> List[str]:
        """Get runway-specific recommendations."""
        if risk_level.lower() in ["safe", "low"] and runway_months > 12:
            return [
                "Maintain current cash management strategy",
                "Consider strategic investments or acquisitions",
                "Build cash reserves for future opportunities",
            ]
        elif risk_level.lower() == "medium" or (runway_months > 6 and runway_months <= 12):
            return [
                "Monitor cash flow closely and implement cost controls",
                "Explore additional funding options if needed",
                "Optimize working capital management",
            ]
        else:
            return [
                "URGENT: Implement immediate cash preservation measures",
                "Reduce discretionary spending and delay non-essential projects",
                "Secure additional funding or financing immediately",
                "Consider asset sales or cost restructuring",
            ]

    def _analyze_cash_position(self, cash_balance: float) -> str:
        """Analyze cash position."""
        if cash_balance > 1000000:
            return "Strong cash position provides significant financial flexibility."
        elif cash_balance > 500000:
            return "Adequate cash reserves for current operations."
        elif cash_balance > 100000:
            return "Limited cash reserves - monitor closely."
        else:
            return "Critical cash position requiring immediate attention."

    def _analyze_burn_rate(self, monthly_burn: float) -> str:
        """Analyze burn rate."""
        if monthly_burn < 100000:
            return "Conservative burn rate allows for sustainable operations."
        elif monthly_burn < 300000:
            return "Moderate burn rate - monitor efficiency improvements."
        else:
            return "High burn rate requires immediate cost optimization measures."

    def _analyze_runway(self, runway_months: float) -> str:
        """Analyze runway duration."""
        if runway_months > 18:
            return "Excellent runway provides ample time for strategic initiatives."
        elif runway_months > 12:
            return "Good runway duration with comfortable planning horizon."
        elif runway_months > 6:
            return "Moderate runway - begin contingency planning."
        else:
            return "Critical runway duration requiring immediate action."

    def _get_runway_insights(self, cash_balance: float, monthly_burn: float, runway_months: float) -> List[str]:
        """Get comprehensive runway insights."""
        insights = []

        # Cash efficiency
        cash_efficiency = cash_balance / monthly_burn if monthly_burn > 0 else 0
        if cash_efficiency > 12:
            insights.append("Excellent cash efficiency ratio")
        elif cash_efficiency > 6:
            insights.append("Good cash efficiency - room for optimization")
        else:
            insights.append("Poor cash efficiency - focus on cash management")

        # Runway benchmarks
        if runway_months < 3:
            insights.append("Immediate funding required within 3 months")
        elif runway_months < 6:
            insights.append("Secure additional funding within 6 months")
        elif runway_months < 12:
            insights.append("Plan for funding needs within 12 months")

        # Burn rate analysis
        if monthly_burn > cash_balance * 0.1:
            insights.append("Burn rate exceeds 10% of cash reserves monthly")

        return insights if insights else ["Cash position and burn rate are within normal ranges"]


# Global instance
modern_excel_generator = ModernExcelGenerator()