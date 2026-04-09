"""
FinAI Report Agent — Unified report assembly with narrative, charts, and enhanced export.
═══════════════════════════════════════════════════════════════════════════════════════════
The Report Agent orchestrates the final output layer. It:

  1. Assembles reports from CalcAgent data + InsightAgent narrative
  2. Generates enhanced Excel exports with Executive Summary sheets + charts
  3. Manages chart generation for inline chat visualization
  4. Saves reports to database with full metadata

Tools owned (migrated from legacy):
  - generate_mr_report  → management report with narrative commentary
  - save_report_to_db   → persist report to database
  - generate_chart      → inline chart for chat

Architecture:
  CalcAgent.data + InsightAgent.narrative → ReportAgent.assemble()
                                          → Excel with narrative sheet + charts
                                          → DB report with metadata + narrative
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.agents.base import BaseAgent, AgentTask, AgentContext, AgentResult
from app.models.all_models import (
    Dataset, Report, RevenueItem, COGSItem, GAExpenseItem,
    BudgetLine, Transaction,
)

from app.config import settings

logger = logging.getLogger(__name__)

EXPORT_DIR = Path("./exports")
EXPORT_DIR.mkdir(exist_ok=True)


def _fgel(v: float) -> str:
    """Format financial value with GEL symbol."""
    if v is None:
        return "N/A"
    if abs(v) >= 1_000_000:
        return f"{v:,.0f}"
    return f"{v:,.2f}"


# Tools this agent owns
REPORT_TOOLS = [
    "generate_mr_report",
    "save_report_to_db",
    "generate_chart",
]


class ReportAgent(BaseAgent):
    """Report assembly and export specialist.

    Orchestrates the full report pipeline:
    CalcAgent (data) + InsightAgent (narrative) → assemble → export
    """

    name = "report"
    description = "Report assembly specialist — MR reports, Excel export, charts, narrative integration"
    capabilities = ["report", "export", "chart"]
    tools = []  # Tool defs still in legacy; ReportAgent delegates for execution

    def can_handle(self, task: AgentTask) -> bool:
        return (
            task.task_type in self.capabilities
            or task.parameters.get("tool_name") in REPORT_TOOLS
        )

    async def execute(self, task: AgentTask, context: AgentContext) -> AgentResult:
        """Execute a report generation/export task."""
        tool_name = task.parameters.get("tool_name", "")
        params = task.parameters.get("tool_input", task.parameters)
        db = context.db

        try:
            if tool_name == "generate_mr_report":
                result_text = await self._generate_mr_report_with_narrative(params, db)
            elif tool_name == "save_report_to_db":
                result_text = await self._save_report(params, db)
            elif tool_name == "generate_chart":
                result_text = self._generate_chart(params)
            elif task.task_type == "export":
                result_text = await self._enhanced_excel_export(params, db)
            else:
                return await self._delegate_to_legacy(task, context)

            return self._make_result(
                status="success",
                data={"tool_result": result_text, "tool_name": tool_name},
                narrative=result_text if isinstance(result_text, str) else str(result_text),
            )
        except Exception as e:
            logger.error(f"ReportAgent error on {tool_name}: {e}", exc_info=True)
            return self._error_result(str(e))

    # ── Tool Implementations ──────────────────────────────────────────────────

    async def _generate_mr_report_with_narrative(self, params: Dict, db: AsyncSession) -> str:
        """Generate Management Report WITH narrative commentary from InsightAgent.

        Enhancement over legacy: adds AI Commentary section to the MR output.
        """
        from app.services.income_statement import build_income_statement
        from app.services.narrative_engine import narrative_engine

        active_ds = (await db.execute(
            select(Dataset).where(Dataset.is_active == True)
        )).scalar_one_or_none()
        period = params.get("period") or (active_ds.period if active_ds else "January 2025")
        ds_name = active_ds.name if active_ds else settings.COMPANY_NAME

        # Build income statement
        ds_id = active_ds.id if active_ds else None
        if not ds_id:
            return "No active dataset. Upload financial data first."

        rev = (await db.execute(select(RevenueItem).where(RevenueItem.dataset_id == ds_id))).scalars().all()
        cogs = (await db.execute(select(COGSItem).where(COGSItem.dataset_id == ds_id))).scalars().all()
        ga = (await db.execute(select(GAExpenseItem).where(GAExpenseItem.dataset_id == ds_id))).scalars().all()

        stmt = build_income_statement(rev, cogs, ga, period)

        # Budget data
        bud_result = await db.execute(select(BudgetLine))
        budget = {b.line_item: b.actual_amount if b.actual_amount is not None else b.budget_amount
                  for b in bud_result.scalars().all()}
        bud_rev = budget.get("Revenue", 0)

        # Expense data
        exp_result = await db.execute(select(Transaction).where(Transaction.type == "Expense"))
        expenses = exp_result.scalars().all()
        total_exp = sum(e.amount or 0 for e in expenses)

        # Build structured rows
        from app.services.coa_engine import build_structured_pl_rows
        rows = build_structured_pl_rows(stmt, budget)

        # ── Generate narrative commentary ─────────────────────────────
        stmt_dict = stmt.to_dict()
        narrative = narrative_engine.generate_income_statement_narrative(stmt_dict, period=period)
        narrative_dict = narrative.to_dict()

        # Build summary text
        summary = f"""Management Report \u2014 {period}
Source: {ds_name} | {len(stmt.revenue_by_product)} revenue products | {len(stmt.cogs_by_product)} COGS products

KEY METRICS
\u2022 Net Revenue: {_fgel(stmt.total_revenue)} (Budget: {_fgel(bud_rev)}, Variance: {_fgel(stmt.total_revenue - bud_rev)})
\u2022 Revenue Wholesale: {_fgel(stmt.revenue_wholesale_total)} | Revenue Retail: {_fgel(stmt.revenue_retail_total)}
\u2022 COGS: {_fgel(stmt.total_cogs)} ({stmt.total_cogs/stmt.total_revenue*100:.1f}% of revenue)
\u2022 Gross Margin: {_fgel(stmt.total_gross_margin)} ({stmt.total_gross_margin/stmt.total_revenue*100:.1f}%)
  - Wholesale: {_fgel(stmt.margin_wholesale_total)}{' NEGATIVE' if stmt.margin_wholesale_total < 0 else ''}
  - Retail: {_fgel(stmt.margin_retail_total)}
\u2022 Other Revenue: {_fgel(stmt.other_revenue_total)}
\u2022 Total Gross Profit: {_fgel(stmt.total_gross_profit)}
\u2022 G&A Expenses: {_fgel(stmt.ga_expenses)}
\u2022 EBITDA: {_fgel(stmt.ebitda)}
\u2022 Total OpEx: {_fgel(total_exp)} ({len(expenses)} transactions)"""

        # Save report with narrative metadata
        report = Report(
            title=f"Management Report \u2014 {period}",
            report_type="mr", period=period, currency="GEL",
            rows=rows, summary=summary,
            kpis={
                "revenue": stmt.total_revenue, "cogs": stmt.total_cogs,
                "gross_margin": stmt.total_gross_margin,
                "total_gross_profit": stmt.total_gross_profit,
                "ga_expenses": stmt.ga_expenses, "ebitda": stmt.ebitda,
                "wholesale_margin": stmt.margin_wholesale_total,
                "retail_margin": stmt.margin_retail_total,
                "net_profit": stmt.net_profit,
            },
            metadata_json={
                "narrative": narrative_dict,
                "generated_by_agent": "report",
                "has_narrative": True,
            },
            generated_by="agent",
            source_dataset_id=ds_id,
        )
        db.add(report)
        await db.commit()

        # ── Build enhanced Excel export ──────────────────────────────
        try:
            export_path = self._build_enhanced_excel(report, stmt, narrative_dict)
            report.export_path = export_path
            await db.commit()
            export_note = f"\nExcel exported: {export_path}"
        except Exception as e:
            logger.warning(f"Enhanced Excel export skipped: {e}")
            export_note = ""

        # Format output text
        output = f"__NAVIGATE_TO__mr__END__\n"
        output += f"**Management Report generated from {ds_name}**\n"
        output += f"Period: {period}\n\n{summary}\n"
        output += f"\nSaved to database (ID: {report.id}){export_note}"

        # Add narrative commentary
        if narrative_dict.get("executive_summary"):
            output += f"\n\n**AI Commentary:**\n{narrative_dict['executive_summary']}"
            warnings = narrative_dict.get("warnings", [])
            if warnings:
                output += "\n" + "\n".join(f"  ⚠ {w}" for w in warnings[:5])
            recs = narrative_dict.get("recommendations", [])
            if recs:
                output += "\n" + "\n".join(f"  \u2192 {r}" for r in recs[:5])

        return output

    async def _save_report(self, params: Dict, db: AsyncSession) -> str:
        """Save a report to the database."""
        report = Report(
            title=params.get("title", "Report"),
            report_type=params.get("type", "custom"),
            period=params.get("period", "January 2025"),
            summary=params.get("summary", ""),
            generated_by="agent",
        )
        db.add(report)
        await db.commit()
        return f"Report saved to database (ID: {report.id}): {report.title}"

    def _generate_chart(self, params: Dict) -> str:
        """Generate an inline chart for WebSocket chat display."""
        chart_data = json.dumps({
            "type": params.get("type", "bar"),
            "title": params.get("title", "Chart"),
            "labels": params.get("labels", []),
            "data": params.get("data", []),
        })
        return f"__CHART__{chart_data}__END__\nChart generated: {params.get('title', '')}"

    async def _enhanced_excel_export(self, params: Dict, db: AsyncSession) -> str:
        """Export a report with enhanced narrative sheet.

        Can be called directly or via API for any report_id.
        """
        report_id = params.get("report_id")
        if not report_id:
            return "report_id is required for export."

        report = (await db.execute(
            select(Report).where(Report.id == report_id)
        )).scalar_one_or_none()
        if not report:
            return f"Report {report_id} not found."

        # Check if narrative exists in metadata
        narrative_dict = {}
        if report.metadata_json and isinstance(report.metadata_json, dict):
            narrative_dict = report.metadata_json.get("narrative", {})

        # If no narrative, generate one from the income statement
        if not narrative_dict and report.kpis:
            try:
                from app.services.narrative_engine import narrative_engine
                from app.services.income_statement import build_income_statement

                ds_id = report.source_dataset_id
                if ds_id:
                    rev = (await db.execute(select(RevenueItem).where(RevenueItem.dataset_id == ds_id))).scalars().all()
                    cogs_items = (await db.execute(select(COGSItem).where(COGSItem.dataset_id == ds_id))).scalars().all()
                    ga = (await db.execute(select(GAExpenseItem).where(GAExpenseItem.dataset_id == ds_id))).scalars().all()
                    stmt = build_income_statement(rev, cogs_items, ga, report.period or "")
                    narrative_obj = narrative_engine.generate_income_statement_narrative(
                        stmt.to_dict(), period=report.period or ""
                    )
                    narrative_dict = narrative_obj.to_dict()
            except Exception as e:
                logger.warning(f"Narrative generation for export skipped: {e}")

        # Build enhanced Excel
        try:
            from app.services.income_statement import build_income_statement
            stmt = None
            ds_id = report.source_dataset_id
            if ds_id:
                rev = (await db.execute(select(RevenueItem).where(RevenueItem.dataset_id == ds_id))).scalars().all()
                cogs_items = (await db.execute(select(COGSItem).where(COGSItem.dataset_id == ds_id))).scalars().all()
                ga = (await db.execute(select(GAExpenseItem).where(GAExpenseItem.dataset_id == ds_id))).scalars().all()
                stmt = build_income_statement(rev, cogs_items, ga, report.period or "")

            path = self._build_enhanced_excel(report, stmt, narrative_dict)
            report.export_path = path
            await db.commit()
            return {"export_path": path}
        except Exception as e:
            # Fallback to basic export
            from app.utils.excel_export import excel_exporter
            path = excel_exporter.export_report(report)
            report.export_path = path
            await db.commit()
            return {"export_path": path, "fallback": True}

    # ── Enhanced Excel Builder ────────────────────────────────────────────────

    def _build_enhanced_excel(
        self,
        report: Report,
        stmt: Any = None,
        narrative_dict: Optional[Dict] = None,
    ) -> str:
        """Build an enhanced Excel file with Executive Summary sheet + charts.

        Args:
            report: Report model instance
            stmt: IncomeStatement (optional, for chart data)
            narrative_dict: FinancialNarrative.to_dict() output

        Returns:
            File path of the exported Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl not installed")

        wb = Workbook()

        # ── Sheet 1: Executive Summary ────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Colours
        BG_DARK = "0A0D1A"
        BG_CARD = "111827"
        BG_HEADER = "0F172A"
        ACCENT = "0EA5E9"
        TEXT = "E2E8F0"
        MUTED = "64748B"
        GREEN = "10B981"
        RED = "EF4444"
        WHITE = "FFFFFF"

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def font(color=None, bold=False, size=11):
            return Font(color=color or TEXT, bold=bold, size=size, name="Calibri")

        ws_summary.column_dimensions["A"].width = 4
        ws_summary.column_dimensions["B"].width = 48
        ws_summary.column_dimensions["C"].width = 20
        ws_summary.column_dimensions["D"].width = 20

        # Title
        row = 1
        ws_summary.merge_cells(f"A{row}:D{row}")
        c = ws_summary[f"A{row}"]
        c.value = report.company or settings.COMPANY_NAME
        c.font = font(ACCENT, bold=True, size=14)
        c.fill = fill(BG_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[row].height = 32

        row = 2
        ws_summary.merge_cells(f"A{row}:D{row}")
        c = ws_summary[f"A{row}"]
        c.value = f"Executive Summary - {report.period or ''}"
        c.font = font(WHITE, bold=True, size=12)
        c.fill = fill(BG_HEADER)
        c.alignment = Alignment(horizontal="center")
        ws_summary.row_dimensions[row].height = 26

        row = 3
        ws_summary.merge_cells(f"A{row}:D{row}")
        ws_summary[f"A{row}"].fill = fill(BG_DARK)
        ws_summary.row_dimensions[row].height = 8

        # ── KPI Cards ─────────────────────────────────────────────────
        row = 4
        ws_summary.merge_cells(f"A{row}:D{row}")
        c = ws_summary[f"A{row}"]
        c.value = "KEY PERFORMANCE INDICATORS"
        c.font = font(ACCENT, bold=True, size=11)
        c.fill = fill(BG_HEADER)
        ws_summary.row_dimensions[row].height = 22

        kpis = report.kpis or {}
        kpi_items = [
            ("Net Revenue", kpis.get("revenue", 0)),
            ("Cost of Goods Sold", kpis.get("cogs", 0)),
            ("Gross Margin", kpis.get("gross_margin", 0)),
            ("Total Gross Profit", kpis.get("total_gross_profit", 0)),
            ("G&A Expenses", kpis.get("ga_expenses", 0)),
            ("EBITDA", kpis.get("ebitda", 0)),
            ("Net Profit", kpis.get("net_profit", 0)),
            ("Wholesale Margin", kpis.get("wholesale_margin", 0)),
            ("Retail Margin", kpis.get("retail_margin", 0)),
        ]

        for label, value in kpi_items:
            row += 1
            ws_summary.row_dimensions[row].height = 20
            ws_summary[f"B{row}"].value = label
            ws_summary[f"B{row}"].font = font(TEXT)
            ws_summary[f"B{row}"].fill = fill(BG_CARD)
            ws_summary[f"C{row}"].value = value or 0
            ws_summary[f"C{row}"].number_format = '#,##0.00'
            color = GREEN if (value or 0) >= 0 else RED
            ws_summary[f"C{row}"].font = font(color, bold=True)
            ws_summary[f"C{row}"].fill = fill(BG_CARD)
            ws_summary[f"A{row}"].fill = fill(BG_CARD)
            ws_summary[f"D{row}"].fill = fill(BG_CARD)

        # ── Narrative Commentary ──────────────────────────────────────
        row += 2
        ws_summary.merge_cells(f"A{row}:D{row}")
        c = ws_summary[f"A{row}"]
        c.value = "AI COMMENTARY"
        c.font = font(ACCENT, bold=True, size=11)
        c.fill = fill(BG_HEADER)
        ws_summary.row_dimensions[row].height = 22

        if narrative_dict:
            # Executive summary
            exec_summary = narrative_dict.get("executive_summary", "")
            if exec_summary:
                row += 1
                ws_summary.merge_cells(f"A{row}:D{row}")
                c = ws_summary[f"A{row}"]
                c.value = exec_summary
                c.font = font(TEXT, size=10)
                c.fill = fill(BG_CARD)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws_summary.row_dimensions[row].height = 60

            # Sections
            for section in narrative_dict.get("sections", []):
                row += 1
                ws_summary.merge_cells(f"A{row}:D{row}")
                c = ws_summary[f"A{row}"]
                sev = section.get("severity", "info")
                icon = {"warning": "[!]", "critical": "[!!]", "positive": "[OK]"}.get(sev, "")
                c.value = f"{icon} {section.get('title', '')}"
                c.font = font(
                    RED if sev == "critical" else
                    "F59E0B" if sev == "warning" else
                    GREEN if sev == "positive" else WHITE,
                    bold=True, size=10
                )
                c.fill = fill(BG_CARD)

                row += 1
                ws_summary.merge_cells(f"A{row}:D{row}")
                c = ws_summary[f"A{row}"]
                c.value = section.get("body", "")
                c.font = font(TEXT, size=9)
                c.fill = fill(BG_CARD)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                ws_summary.row_dimensions[row].height = 45

            # Warnings
            warnings = narrative_dict.get("warnings", [])
            if warnings:
                row += 1
                ws_summary.merge_cells(f"A{row}:D{row}")
                c = ws_summary[f"A{row}"]
                c.value = "WARNINGS"
                c.font = font(RED, bold=True, size=10)
                c.fill = fill(BG_HEADER)

                for w in warnings:
                    row += 1
                    ws_summary.merge_cells(f"A{row}:D{row}")
                    c = ws_summary[f"A{row}"]
                    c.value = f"  [!] {w}"
                    c.font = font("F59E0B", size=9)
                    c.fill = fill(BG_CARD)

            # Recommendations
            recs = narrative_dict.get("recommendations", [])
            if recs:
                row += 1
                ws_summary.merge_cells(f"A{row}:D{row}")
                c = ws_summary[f"A{row}"]
                c.value = "RECOMMENDATIONS"
                c.font = font(ACCENT, bold=True, size=10)
                c.fill = fill(BG_HEADER)

                for r in recs:
                    row += 1
                    ws_summary.merge_cells(f"A{row}:D{row}")
                    c = ws_summary[f"A{row}"]
                    c.value = f"  > {r}"
                    c.font = font(TEXT, size=9)
                    c.fill = fill(BG_CARD)
        else:
            row += 1
            ws_summary.merge_cells(f"A{row}:D{row}")
            ws_summary[f"A{row}"].value = "No narrative available for this report."
            ws_summary[f"A{row}"].font = font(MUTED)
            ws_summary[f"A{row}"].fill = fill(BG_CARD)

        # Footer
        row += 2
        ws_summary.merge_cells(f"A{row}:D{row}")
        c = ws_summary[f"A{row}"]
        c.value = f"Generated by FinAI ReportAgent | {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        c.font = font(MUTED, size=9)
        c.fill = fill(BG_DARK)

        # Fill empty cells with dark bg
        for r in range(1, row + 1):
            for col_letter in ["A", "B", "C", "D"]:
                cell = ws_summary[f"{col_letter}{r}"]
                if cell.fill.fill_type is None or cell.fill.fill_type == "none":
                    cell.fill = fill(BG_DARK)

        # ── Sheet 2: Charts ───────────────────────────────────────────
        if stmt and hasattr(stmt, "revenue_wholesale_total"):
            ws_charts = wb.create_sheet("Charts")

            # Revenue data for chart
            ws_charts["A1"] = "Segment"
            ws_charts["B1"] = "Revenue"
            ws_charts["C1"] = "COGS"
            ws_charts["D1"] = "Margin"

            segments = [
                ("Wholesale Petrol", stmt.revenue_wholesale_petrol, stmt.cogs_wholesale_petrol, stmt.margin_wholesale_petrol),
                ("Wholesale Diesel", stmt.revenue_wholesale_diesel, stmt.cogs_wholesale_diesel, stmt.margin_wholesale_diesel),
                ("Retail Petrol", stmt.revenue_retail_petrol, stmt.cogs_retail_petrol, stmt.margin_retail_petrol),
                ("Retail Diesel", stmt.revenue_retail_diesel, stmt.cogs_retail_diesel, stmt.margin_retail_diesel),
                ("Retail CNG", stmt.revenue_retail_cng, stmt.cogs_retail_cng, stmt.margin_retail_cng),
                ("Retail LPG", stmt.revenue_retail_lpg, stmt.cogs_retail_lpg, stmt.margin_retail_lpg),
            ]
            # Filter out zero-revenue segments
            segments = [(s, r, c, m) for s, r, c, m in segments if r > 0]

            for i, (seg, rev_val, cogs_val, margin_val) in enumerate(segments, 2):
                ws_charts[f"A{i}"] = seg
                ws_charts[f"B{i}"] = rev_val
                ws_charts[f"C{i}"] = cogs_val
                ws_charts[f"D{i}"] = margin_val

            # Revenue vs COGS bar chart
            if segments:
                chart = BarChart()
                chart.type = "col"
                chart.title = "Revenue vs COGS by Segment"
                chart.y_axis.title = "GEL"
                chart.x_axis.title = "Segment"
                chart.style = 10

                data_ref = Reference(ws_charts, min_col=2, max_col=3, min_row=1, max_row=len(segments) + 1)
                cats = Reference(ws_charts, min_col=1, min_row=2, max_row=len(segments) + 1)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                chart.width = 20
                chart.height = 12
                ws_charts.add_chart(chart, "F1")

                # Margin bar chart
                margin_chart = BarChart()
                margin_chart.type = "col"
                margin_chart.title = "Gross Margin by Segment"
                margin_chart.y_axis.title = "GEL"
                margin_chart.style = 10

                margin_data = Reference(ws_charts, min_col=4, max_col=4, min_row=1, max_row=len(segments) + 1)
                margin_chart.add_data(margin_data, titles_from_data=True)
                margin_chart.set_categories(cats)
                margin_chart.width = 20
                margin_chart.height = 12
                ws_charts.add_chart(margin_chart, "F18")

            # Revenue pie chart data
            pie_row = len(segments) + 4
            ws_charts[f"A{pie_row}"] = "Segment"
            ws_charts[f"B{pie_row}"] = "Revenue Share"
            for i, (seg, rev_val, _, _) in enumerate(segments, 1):
                ws_charts[f"A{pie_row + i}"] = seg
                ws_charts[f"B{pie_row + i}"] = rev_val

            if segments:
                pie = PieChart()
                pie.title = "Revenue Mix by Segment"
                pie.style = 10
                pie_data = Reference(ws_charts, min_col=2, min_row=pie_row, max_row=pie_row + len(segments))
                pie_cats = Reference(ws_charts, min_col=1, min_row=pie_row + 1, max_row=pie_row + len(segments))
                pie.add_data(pie_data, titles_from_data=True)
                pie.set_categories(pie_cats)
                pie.width = 16
                pie.height = 12
                ws_charts.add_chart(pie, "F35")

        # ── Sheet 3: Financial Data (original report rows) ────────────
        if report.rows:
            from app.utils.excel_export import excel_exporter
            ws_data = wb.create_sheet("Financial Data")

            # Simplified data dump
            headers = ["Code", "Line Item", "Actual", "Plan"]
            for j, h in enumerate(headers, 1):
                c = ws_data.cell(row=1, column=j)
                c.value = h
                c.font = Font(color=WHITE, bold=True, name="Calibri")
                c.fill = fill(ACCENT)

            ws_data.column_dimensions["A"].width = 10
            ws_data.column_dimensions["B"].width = 40
            ws_data.column_dimensions["C"].width = 18
            ws_data.column_dimensions["D"].width = 18

            for i, item in enumerate(report.rows, 2):
                if not isinstance(item, dict):
                    continue
                ws_data.cell(row=i, column=1).value = item.get("c", item.get("code", ""))
                ws_data.cell(row=i, column=2).value = item.get("l", item.get("label", ""))
                ac = item.get("ac", item.get("actual", 0)) or 0
                pl_val = item.get("pl", item.get("plan", 0)) or 0
                ws_data.cell(row=i, column=3).value = ac
                ws_data.cell(row=i, column=3).number_format = '#,##0.00'
                ws_data.cell(row=i, column=4).value = pl_val
                ws_data.cell(row=i, column=4).number_format = '#,##0.00'

        # Save
        fname = f"enhanced_{report.report_type}_{(report.period or 'report').replace(' ', '_')}_{report.id}.xlsx"
        path = str(EXPORT_DIR / fname)
        wb.save(path)
        logger.info(f"Enhanced Excel exported: {path}")
        return path

    # ── Helpers ───────────────────────────────────────────────────────────────

    async def _delegate_to_legacy(self, task: AgentTask, context: AgentContext) -> AgentResult:
        """Delegate to legacy agent for tools not yet fully migrated."""
        from app.agents.registry import registry
        legacy = registry.get("legacy")
        if legacy:
            return await legacy.execute(task, context)
        return self._error_result("Legacy agent not available")
