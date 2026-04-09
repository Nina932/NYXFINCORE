"""
hypothesis_parser.py — Hypothesis-Driven Parsing (HDP) Engine
==============================================================
Replaces the single-pass early-commitment ingestion pipeline with a
multi-hypothesis search architecture.

Instead of:
    Excel → Classify → Detect → Parse  (sequential, irreversible)

The HDP engine does:
    Excel → Generate Hypotheses → Parse Under Each → Score → Best Fit

This architecture is used in serious financial data platforms (Palantir,
Bloomberg-class systems) because spreadsheet ingestion is a SEARCH problem,
not a classification problem.

Core idea:
    "What possible financial structure could generate this table?"
    Instead of: "What is this column?"

Scoring uses accounting invariants:
    - ∑debit ≈ ∑credit  (trial balance / GL)
    - COA prefix match rate
    - Numeric column consistency
    - Structural density (null ratio)
    - Date monotonicity
    - Account code format consistency

Key classes:
    SchemaHypothesis       — one candidate interpretation
    HypothesisGenerator    — produces N candidate interpretations
    HypothesisValidator    — validates using accounting invariants
    HypothesisScorer       — composite scoring engine
    HypothesisDrivenParser — orchestrator (replaces pipeline)
"""
from __future__ import annotations

import re
import logging
import math
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from app.services.ingestion_intelligence import (
    SchemaType,
    ColumnMapping,
    DetectionResult,
    ParsedFile,
    FileStructureDetector,
    ColumnSemanticClassifier,
    SchemaDetector,
    _COLUMN_ROLE_KEYWORDS,
    _SCHEMA_SIGNALS,
)

logger = logging.getLogger(__name__)


# ── Data structures ──────────────────────────────────────────────────────────

@dataclass
class ColumnHypothesis:
    """A single column-role assignment hypothesis."""
    col_idx: int
    role: str
    confidence: float = 0.0
    source: str = ""  # "keyword", "pattern", "fallback", "override"


@dataclass
class ParsedHypothesisResult:
    """The result of parsing data under one hypothesis."""
    records: List[Dict[str, Any]] = field(default_factory=list)
    total_rows: int = 0
    parsed_rows: int = 0
    dropped_rows: int = 0
    drop_reasons: Dict[str, int] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)


@dataclass
class ValidationSignals:
    """Accounting validation signals for a hypothesis."""
    debit_credit_balance_ratio: float = 1.0      # |Σd - Σc| / max(Σd, Σc); 0 = perfect
    coa_prefix_match_rate: float = 0.0            # % of codes matching known prefixes
    numeric_consistency: float = 0.0               # % of amount cols that are numeric
    null_ratio: float = 1.0                        # % of key cols that are null (lower = better)
    date_monotonicity: float = 0.0                 # 1.0 = dates are monotonic
    row_completeness: float = 0.0                  # % of rows with all required fields
    account_code_format_consistency: float = 0.0   # % of codes matching dominant format
    value_magnitude_consistency: float = 0.0        # 1 - normalized_std of amounts
    total_amount_coverage: float = 0.0             # % of rows with valid amounts


@dataclass
class SchemaHypothesis:
    """
    A single candidate interpretation of a spreadsheet sheet.

    Each hypothesis proposes:
        - What schema type this is (TB, GL, IS, BS, COA, etc.)
        - How columns map to semantic roles
        - What constraints should hold

    The scorer evaluates how well the data actually fits.
    """
    schema_type: str
    column_mapping: ColumnMapping
    parse_result: Optional[ParsedHypothesisResult] = None
    validation: Optional[ValidationSignals] = None
    composite_score: float = 0.0
    scoring_breakdown: Dict[str, float] = field(default_factory=dict)
    rank: int = 0
    is_winner: bool = False

    @property
    def confidence(self) -> float:
        return min(self.composite_score, 1.0)


@dataclass
class HypothesisParseResult:
    """Final result of hypothesis-driven parsing."""
    winner: SchemaHypothesis
    all_hypotheses: List[SchemaHypothesis]
    detection: DetectionResult
    records: List[Dict[str, Any]]
    metadata: Dict[str, Any] = field(default_factory=dict)
    ingestion_warnings: List[str] = field(default_factory=list)

    @property
    def schema_type(self) -> str:
        return self.winner.schema_type

    @property
    def confidence(self) -> float:
        return self.winner.confidence

    def to_parsed_file(self) -> ParsedFile:
        """Convert to legacy ParsedFile format for backward compatibility."""
        return ParsedFile(
            detection=self.detection,
            records=self.records,
            metadata=self.metadata,
        )


# ── Known COA prefix patterns ───────────────────────────────────────────────

# Georgian IFRS 4-digit prefixes
_KNOWN_ACCOUNT_PREFIXES = {
    "1": "current_assets",
    "2": "noncurrent_assets",
    "3": "current_liabilities",
    "4": "noncurrent_liabilities",
    "5": "equity",
    "6": "revenue",
    "7": "cogs_opex",
    "8": "other_income_expense",
    "9": "tax_extraordinary",
}

# Russian 1C 2-digit account ranges
_RUSSIAN_1C_RANGES = {
    (1, 9):   "noncurrent_assets",
    (10, 19): "materials_inventory",
    (20, 29): "production",
    (40, 49): "finished_goods",
    (50, 59): "cash",
    (60, 69): "settlements",
    (70, 79): "payroll_settlements",
    (80, 89): "capital",
    (90, 99): "financial_results",
}


# ── HypothesisGenerator ─────────────────────────────────────────────────────

class HypothesisGenerator:
    """
    Generates candidate schema interpretations from raw sheet data.

    Strategy:
    1. Run column classification with MULTIPLE interpretations per column
    2. Generate one hypothesis per supported schema type
    3. For ambiguous columns, create variant hypotheses

    This replaces the single-pass classifier + detector with a
    multi-hypothesis search space.
    """

    SCHEMA_TYPES = [
        SchemaType.TRIAL_BALANCE,
        SchemaType.GENERAL_LEDGER,
        SchemaType.CHART_OF_ACCOUNTS,
        SchemaType.INCOME_STATEMENT,
        SchemaType.BALANCE_SHEET,
        SchemaType.CASH_FLOW,
        SchemaType.BUDGET,
        SchemaType.KPI_DASHBOARD,
    ]

    def __init__(self):
        self._structure = FileStructureDetector()
        self._classifier = ColumnSemanticClassifier()

    def generate(
        self,
        header_row: tuple,
        data_rows: List[tuple],
        filename: str = "",
        sheet_name: str = "",
    ) -> List[SchemaHypothesis]:
        """
        Generate all plausible hypotheses for the given sheet.

        Returns 8 hypotheses (one per schema type), each with its own
        column mapping tailored to what that schema type expects.
        """
        hypotheses: List[SchemaHypothesis] = []

        # Step 1: Get base column classification (keyword + pattern)
        base_mapping = self._classifier.classify(header_row, data_rows)

        # Step 2: Detect ALL possible column-role assignments (soft)
        soft_assignments = self._soft_classify_columns(header_row, data_rows)

        # Step 3: Generate one hypothesis per schema type
        for schema_type in self.SCHEMA_TYPES:
            col_map = self._build_mapping_for_schema(
                schema_type, base_mapping, soft_assignments,
                header_row, data_rows
            )
            hypothesis = SchemaHypothesis(
                schema_type=schema_type,
                column_mapping=col_map,
            )
            hypotheses.append(hypothesis)

        return hypotheses

    def _soft_classify_columns(
        self,
        header_row: tuple,
        data_rows: List[tuple],
    ) -> Dict[int, List[ColumnHypothesis]]:
        """
        For each column, generate ALL plausible role assignments with
        confidence scores. Unlike the hard classifier which picks one,
        this returns a ranked list per column.
        """
        assignments: Dict[int, List[ColumnHypothesis]] = {}

        for col_idx, cell in enumerate(header_row):
            col_hyps: List[ColumnHypothesis] = []

            # Keyword-based assignments
            if cell is not None:
                cell_str = str(cell).lower().strip()
                for role, keywords in _COLUMN_ROLE_KEYWORDS.items():
                    for kw in keywords:
                        if kw in cell_str:
                            specificity = len(kw) / max(len(cell_str), 1)
                            col_hyps.append(ColumnHypothesis(
                                col_idx=col_idx,
                                role=role,
                                confidence=min(0.5 + specificity * 0.5, 0.99),
                                source="keyword",
                            ))

            # Data-pattern-based assignments
            if data_rows:
                sample = data_rows[:min(20, len(data_rows))]
                vals = [
                    row[col_idx]
                    for row in sample
                    if col_idx < len(row) and row[col_idx] is not None
                ]
                if vals:
                    pattern_hyps = self._infer_roles_from_data(col_idx, vals)
                    col_hyps.extend(pattern_hyps)

            # Sort by confidence descending
            col_hyps.sort(key=lambda h: h.confidence, reverse=True)
            assignments[col_idx] = col_hyps

        return assignments

    def _infer_roles_from_data(
        self,
        col_idx: int,
        values: List[Any],
    ) -> List[ColumnHypothesis]:
        """Infer possible roles from column data patterns."""
        hyps: List[ColumnHypothesis] = []
        total = len(values)
        if total == 0:
            return hyps

        num_count = sum(1 for v in values if isinstance(v, (int, float)))
        str_count = sum(1 for v in values if isinstance(v, str))
        num_ratio = num_count / total
        str_ratio = str_count / total

        # Mostly numeric → could be amount, debit, credit, balance
        if num_ratio > 0.6:
            nums = [float(v) for v in values if isinstance(v, (int, float))]
            has_negatives = any(n < 0 for n in nums)
            avg_abs = sum(abs(n) for n in nums) / len(nums)

            # Large values → amount
            if avg_abs > 10:
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="amount",
                    confidence=0.5 + min(num_ratio * 0.3, 0.3),
                    source="pattern"
                ))
                # Could also be debit or credit
                if not has_negatives:
                    hyps.append(ColumnHypothesis(
                        col_idx=col_idx, role="debit",
                        confidence=0.3 + min(num_ratio * 0.2, 0.2),
                        source="pattern"
                    ))
                    hyps.append(ColumnHypothesis(
                        col_idx=col_idx, role="credit",
                        confidence=0.3 + min(num_ratio * 0.2, 0.2),
                        source="pattern"
                    ))
            # Small values → could be quantity or balance
            if avg_abs <= 100:
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="quantity",
                    confidence=0.2,
                    source="pattern"
                ))

            # Balance: could be debit-credit net
            if has_negatives:
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="balance",
                    confidence=0.4 + min(num_ratio * 0.2, 0.2),
                    source="pattern"
                ))

        # Mostly strings → could be account_code, account_name, counterparty
        if str_ratio > 0.6:
            str_vals = [str(v) for v in values if isinstance(v, str)]
            avg_len = sum(len(s) for s in str_vals) / len(str_vals) if str_vals else 0

            if avg_len <= 8:
                # Short codes
                code_like = sum(
                    1 for v in str_vals
                    if re.match(r'^[\d.Xx\-]+$', v.strip())
                )
                if code_like / len(str_vals) > 0.4:
                    hyps.append(ColumnHypothesis(
                        col_idx=col_idx, role="account_code",
                        confidence=0.6 + min(code_like / len(str_vals) * 0.3, 0.3),
                        source="pattern"
                    ))
                else:
                    hyps.append(ColumnHypothesis(
                        col_idx=col_idx, role="account_code",
                        confidence=0.3,
                        source="pattern"
                    ))
                    hyps.append(ColumnHypothesis(
                        col_idx=col_idx, role="cost_center",
                        confidence=0.2,
                        source="pattern"
                    ))
            elif avg_len > 8 and avg_len <= 40:
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="account_name",
                    confidence=0.5 + min(str_ratio * 0.3, 0.3),
                    source="pattern"
                ))
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="counterparty",
                    confidence=0.3,
                    source="pattern"
                ))
            elif avg_len > 40:
                hyps.append(ColumnHypothesis(
                    col_idx=col_idx, role="account_name",
                    confidence=0.6,
                    source="pattern"
                ))

        # Date detection
        from datetime import datetime as dt
        date_count = sum(
            1 for v in values
            if isinstance(v, dt) or (isinstance(v, str) and re.match(
                r'\d{4}[-/]\d{2}[-/]\d{2}', str(v)
            ))
        )
        if date_count / total > 0.5:
            hyps.append(ColumnHypothesis(
                col_idx=col_idx, role="date",
                confidence=0.7 + min(date_count / total * 0.2, 0.2),
                source="pattern"
            ))

        return hyps

    def _build_mapping_for_schema(
        self,
        schema_type: str,
        base_mapping: ColumnMapping,
        soft_assignments: Dict[int, List[ColumnHypothesis]],
        header_row: tuple,
        data_rows: List[tuple],
    ) -> ColumnMapping:
        """
        Build the best column mapping for a specific schema type.

        For each required/bonus role of the schema type, find the
        highest-confidence column assignment. This allows different
        hypotheses to interpret the same column differently.
        """
        config = _SCHEMA_SIGNALS.get(schema_type, {})
        required_roles = config.get("required", set())
        bonus_roles = config.get("bonus", set())
        target_roles = required_roles | bonus_roles

        roles: Dict[str, Optional[int]] = {role: None for role in _COLUMN_ROLE_KEYWORDS}
        used_cols: set = set()

        # Priority 1: Use base mapping for roles that have keyword matches
        for role, idx in base_mapping.roles.items():
            if idx is not None and role in target_roles:
                roles[role] = idx
                used_cols.add(idx)

        # Priority 2: Fill missing required roles from soft assignments
        for role in target_roles:
            if roles.get(role) is not None:
                continue
            # Find the best unassigned column for this role
            best_col = None
            best_conf = 0.0
            for col_idx, col_hyps in soft_assignments.items():
                if col_idx in used_cols:
                    continue
                for hyp in col_hyps:
                    if hyp.role == role and hyp.confidence > best_conf:
                        best_conf = hyp.confidence
                        best_col = col_idx
            if best_col is not None:
                roles[role] = best_col
                used_cols.add(best_col)

        # Priority 3: Fill remaining roles from base mapping
        for role, idx in base_mapping.roles.items():
            if idx is not None and roles.get(role) is None and idx not in used_cols:
                roles[role] = idx
                used_cols.add(idx)

        detected = [r for r, i in roles.items() if i is not None]
        return ColumnMapping(
            roles=roles,
            header_row=base_mapping.header_row,
            detected_columns=detected,
        )


# ── HypothesisValidator ─────────────────────────────────────────────────────

class HypothesisValidator:
    """
    Validates hypotheses using accounting invariants and data consistency checks.

    This is the core innovation of HDP: instead of classifying THEN parsing,
    we parse under each hypothesis THEN validate which interpretation
    produces the most accounting-consistent result.
    """

    def parse_and_validate(
        self,
        hypothesis: SchemaHypothesis,
        header_row: tuple,
        data_rows: List[tuple],
    ) -> SchemaHypothesis:
        """Parse data under this hypothesis and compute validation signals."""
        # Step 1: Parse records using this hypothesis's column mapping
        parse_result = self._parse_under_hypothesis(hypothesis, data_rows)
        hypothesis.parse_result = parse_result

        # Step 2: Compute validation signals
        validation = self._compute_validation_signals(
            hypothesis, parse_result, header_row, data_rows
        )
        hypothesis.validation = validation

        return hypothesis

    def _parse_under_hypothesis(
        self,
        hypothesis: SchemaHypothesis,
        data_rows: List[tuple],
    ) -> ParsedHypothesisResult:
        """Parse data rows using this hypothesis's column mapping."""
        records: List[Dict[str, Any]] = []
        drop_reasons: Dict[str, int] = {}
        total_rows = len(data_rows)

        col_map = hypothesis.column_mapping
        schema_type = hypothesis.schema_type

        # Determine which roles are required for this schema type
        config = _SCHEMA_SIGNALS.get(schema_type, {})
        required_roles = config.get("required", set())

        for row in data_rows:
            if not any(row):
                drop_reasons["blank_row"] = drop_reasons.get("blank_row", 0) + 1
                continue

            record: Dict[str, Any] = {}
            for role, idx in col_map.roles.items():
                if idx is None or idx >= len(row):
                    continue
                val = row[idx]
                if val is not None:
                    record[role] = val

            if not record:
                drop_reasons["empty_record"] = drop_reasons.get("empty_record", 0) + 1
                continue

            # Check required fields
            missing_required = [r for r in required_roles if r not in record]
            if missing_required:
                reason = f"missing_{','.join(missing_required)}"
                drop_reasons[reason] = drop_reasons.get(reason, 0) + 1
                continue

            records.append(record)

        parsed_rows = len(records)
        dropped_rows = total_rows - parsed_rows

        warnings: List[str] = []
        if dropped_rows > 0:
            drop_pct = dropped_rows / max(total_rows, 1) * 100
            warnings.append(
                f"{dropped_rows}/{total_rows} rows dropped ({drop_pct:.1f}%)"
            )
            for reason, count in drop_reasons.items():
                warnings.append(f"  - {reason}: {count}")

        return ParsedHypothesisResult(
            records=records,
            total_rows=total_rows,
            parsed_rows=parsed_rows,
            dropped_rows=dropped_rows,
            drop_reasons=drop_reasons,
            warnings=warnings,
        )

    def _compute_validation_signals(
        self,
        hypothesis: SchemaHypothesis,
        parse_result: ParsedHypothesisResult,
        header_row: tuple,
        data_rows: List[tuple],
    ) -> ValidationSignals:
        """Compute accounting invariant signals for scoring."""
        records = parse_result.records
        signals = ValidationSignals()

        if not records:
            return signals

        # 1. Debit/Credit balance ratio
        signals.debit_credit_balance_ratio = self._check_debit_credit_balance(records)

        # 2. COA prefix match rate
        signals.coa_prefix_match_rate = self._check_coa_prefixes(records)

        # 3. Numeric consistency of amount columns
        signals.numeric_consistency = self._check_numeric_consistency(
            records, hypothesis.schema_type
        )

        # 4. Null ratio for key columns
        signals.null_ratio = self._check_null_ratio(records, hypothesis.schema_type)

        # 5. Date monotonicity
        signals.date_monotonicity = self._check_date_monotonicity(records)

        # 6. Row completeness
        signals.row_completeness = parse_result.parsed_rows / max(
            parse_result.total_rows, 1
        )

        # 7. Account code format consistency
        signals.account_code_format_consistency = self._check_code_format(records)

        # 8. Value magnitude consistency
        signals.value_magnitude_consistency = self._check_value_magnitude(records)

        # 9. Amount coverage
        signals.total_amount_coverage = self._check_amount_coverage(records)

        return signals

    def _check_debit_credit_balance(self, records: List[Dict]) -> float:
        """
        Check if debits ≈ credits (accounting invariant).
        Returns ratio where 0.0 = perfectly balanced, 1.0 = completely unbalanced.
        """
        total_debit = 0.0
        total_credit = 0.0
        has_dc = False

        for r in records:
            dr = r.get("debit")
            cr = r.get("credit")
            if dr is not None:
                try:
                    total_debit += abs(float(dr))
                    has_dc = True
                except (ValueError, TypeError):
                    pass
            if cr is not None:
                try:
                    total_credit += abs(float(cr))
                    has_dc = True
                except (ValueError, TypeError):
                    pass

        if not has_dc:
            return 1.0  # No debit/credit data → neutral

        maximum = max(total_debit, total_credit)
        if maximum == 0:
            return 0.0  # Both zero → balanced

        imbalance = abs(total_debit - total_credit) / maximum
        return imbalance

    def _check_coa_prefixes(self, records: List[Dict]) -> float:
        """Check what % of account codes match known COA prefix patterns."""
        codes = []
        for r in records:
            code = r.get("account_code")
            if code is not None:
                codes.append(str(code).strip())

        if not codes:
            return 0.0

        matched = 0
        for code in codes:
            # Check Georgian 4-digit pattern
            if len(code) >= 1 and code[0] in _KNOWN_ACCOUNT_PREFIXES:
                matched += 1
                continue
            # Check Russian 1C 2-digit pattern
            try:
                num = int(code[:2])
                for (lo, hi), _ in _RUSSIAN_1C_RANGES.items():
                    if lo <= num <= hi:
                        matched += 1
                        break
            except (ValueError, IndexError):
                pass

        return matched / len(codes) if codes else 0.0

    def _check_numeric_consistency(
        self,
        records: List[Dict],
        schema_type: str,
    ) -> float:
        """Check that amount/debit/credit fields contain valid numbers."""
        numeric_roles = {"amount", "debit", "credit", "balance"}
        total_checks = 0
        valid_checks = 0

        for r in records:
            for role in numeric_roles:
                val = r.get(role)
                if val is not None:
                    total_checks += 1
                    try:
                        float(val)
                        valid_checks += 1
                    except (ValueError, TypeError):
                        pass

        return valid_checks / max(total_checks, 1)

    def _check_null_ratio(self, records: List[Dict], schema_type: str) -> float:
        """Check proportion of null values in key columns (lower = better)."""
        config = _SCHEMA_SIGNALS.get(schema_type, {})
        key_roles = config.get("required", set()) | config.get("bonus", set())
        if not key_roles:
            key_roles = {"account_code", "amount"}

        total = 0
        nulls = 0
        for r in records:
            for role in key_roles:
                total += 1
                if r.get(role) is None:
                    nulls += 1

        return nulls / max(total, 1)

    def _check_date_monotonicity(self, records: List[Dict]) -> float:
        """Check if date column values are monotonically increasing."""
        from datetime import datetime as dt

        dates = []
        for r in records:
            d = r.get("date")
            if d is not None:
                if isinstance(d, dt):
                    dates.append(d)
                elif isinstance(d, str):
                    try:
                        dates.append(dt.fromisoformat(d.replace("/", "-")))
                    except (ValueError, AttributeError):
                        pass

        if len(dates) < 2:
            return 0.5  # Not enough data to judge

        monotonic_pairs = sum(
            1 for i in range(len(dates) - 1) if dates[i] <= dates[i + 1]
        )
        return monotonic_pairs / (len(dates) - 1)

    def _check_code_format(self, records: List[Dict]) -> float:
        """
        Check if account codes follow a consistent format pattern.
        E.g., all 4-digit, all 2-digit, etc.
        """
        codes = []
        for r in records:
            code = r.get("account_code")
            if code is not None:
                codes.append(str(code).strip())

        if not codes:
            return 0.0

        # Count format patterns
        patterns: Dict[str, int] = {}
        for code in codes:
            # Normalize to pattern: d=digit, a=alpha, x=other
            pat = ""
            for ch in code:
                if ch.isdigit():
                    pat += "d"
                elif ch.isalpha():
                    pat += "a"
                else:
                    pat += "x"
            # Collapse runs: "dddd" → "d4"
            collapsed = ""
            i = 0
            while i < len(pat):
                ch = pat[i]
                count = 1
                while i + count < len(pat) and pat[i + count] == ch:
                    count += 1
                collapsed += f"{ch}{count}"
                i += count
            patterns[collapsed] = patterns.get(collapsed, 0) + 1

        # Dominant pattern ratio
        if patterns:
            dominant_count = max(patterns.values())
            return dominant_count / len(codes)
        return 0.0

    def _check_value_magnitude(self, records: List[Dict]) -> float:
        """
        Check if financial values are within a consistent magnitude range.
        Financial data from one source tends to be within 2-3 orders of magnitude.
        """
        amounts = []
        for r in records:
            for role in ("amount", "debit", "credit"):
                val = r.get(role)
                if val is not None:
                    try:
                        v = abs(float(val))
                        if v > 0:
                            amounts.append(math.log10(v))
                    except (ValueError, TypeError):
                        pass

        if len(amounts) < 3:
            return 0.5  # Not enough data

        mean = sum(amounts) / len(amounts)
        variance = sum((x - mean) ** 2 for x in amounts) / len(amounts)
        std = math.sqrt(variance) if variance > 0 else 0

        # Financial data typically spans ~3 orders of magnitude
        # std < 1.5 is consistent, > 3 is chaotic
        consistency = max(0, 1 - std / 3)
        return consistency

    def _check_amount_coverage(self, records: List[Dict]) -> float:
        """Check what proportion of records have at least one valid amount."""
        if not records:
            return 0.0

        has_amount = 0
        for r in records:
            for role in ("amount", "debit", "credit", "balance"):
                val = r.get(role)
                if val is not None:
                    try:
                        float(val)
                        has_amount += 1
                        break
                    except (ValueError, TypeError):
                        pass

        return has_amount / len(records)


# ── HypothesisScorer ─────────────────────────────────────────────────────────

class HypothesisScorer:
    """
    Composite scoring engine for hypothesis ranking.

    Combines validation signals with schema-specific weights to produce
    a single composite score per hypothesis. Different schema types
    emphasize different signals.

    Scoring philosophy:
        - Accounting invariants (debit=credit) are strongest signal for TB/GL
        - COA prefix match rate is strongest for COA/GL/TB
        - Row completeness penalizes bad column assignments
        - Numeric consistency catches type mismatches
    """

    # Schema-specific weight profiles
    _WEIGHT_PROFILES: Dict[str, Dict[str, float]] = {
        SchemaType.TRIAL_BALANCE: {
            "debit_credit_balance":     0.30,  # Most important: TB MUST balance
            "coa_prefix_match":         0.20,
            "row_completeness":         0.15,
            "numeric_consistency":      0.15,
            "code_format_consistency":  0.10,
            "amount_coverage":          0.05,
            "value_magnitude":          0.05,
        },
        SchemaType.GENERAL_LEDGER: {
            "debit_credit_balance":     0.15,  # GL may not balance per-page
            "coa_prefix_match":         0.20,
            "row_completeness":         0.15,
            "numeric_consistency":      0.15,
            "date_monotonicity":        0.15,  # GL is chronological
            "code_format_consistency":  0.10,
            "amount_coverage":          0.10,
        },
        SchemaType.CHART_OF_ACCOUNTS: {
            "coa_prefix_match":         0.30,  # COA = account codes are key
            "code_format_consistency":  0.25,
            "row_completeness":         0.20,
            "null_ratio":              0.15,
            "numeric_consistency":      0.05,  # COA may not have amounts
            "amount_coverage":          0.05,
        },
        SchemaType.INCOME_STATEMENT: {
            "amount_coverage":          0.25,
            "numeric_consistency":      0.25,
            "row_completeness":         0.20,
            "value_magnitude":          0.15,
            "null_ratio":              0.15,
        },
        SchemaType.BALANCE_SHEET: {
            "amount_coverage":          0.25,
            "numeric_consistency":      0.25,
            "row_completeness":         0.20,
            "value_magnitude":          0.15,
            "null_ratio":              0.15,
        },
        SchemaType.CASH_FLOW: {
            "amount_coverage":          0.25,
            "numeric_consistency":      0.25,
            "row_completeness":         0.20,
            "value_magnitude":          0.15,
            "null_ratio":              0.15,
        },
        SchemaType.BUDGET: {
            "amount_coverage":          0.25,
            "numeric_consistency":      0.20,
            "row_completeness":         0.20,
            "coa_prefix_match":         0.15,
            "null_ratio":              0.10,
            "value_magnitude":          0.10,
        },
        SchemaType.KPI_DASHBOARD: {
            "row_completeness":         0.30,
            "numeric_consistency":      0.25,
            "amount_coverage":          0.20,
            "null_ratio":              0.15,
            "value_magnitude":          0.10,
        },
    }

    def score(self, hypothesis: SchemaHypothesis) -> SchemaHypothesis:
        """
        Compute composite score for a hypothesis using weighted signals.
        """
        v = hypothesis.validation
        if v is None:
            hypothesis.composite_score = 0.0
            return hypothesis

        pr = hypothesis.parse_result
        if pr is None or pr.parsed_rows == 0:
            hypothesis.composite_score = 0.0
            return hypothesis

        profile = self._WEIGHT_PROFILES.get(
            hypothesis.schema_type,
            self._WEIGHT_PROFILES[SchemaType.KPI_DASHBOARD],
        )

        breakdown: Dict[str, float] = {}
        composite = 0.0

        # Map signal names to values
        signal_values = {
            "debit_credit_balance":    1.0 - v.debit_credit_balance_ratio,  # Invert: 0=bad, 1=balanced
            "coa_prefix_match":        v.coa_prefix_match_rate,
            "numeric_consistency":     v.numeric_consistency,
            "null_ratio":             1.0 - v.null_ratio,  # Invert: low nulls = good
            "date_monotonicity":       v.date_monotonicity,
            "row_completeness":        v.row_completeness,
            "code_format_consistency": v.account_code_format_consistency,
            "value_magnitude":         v.value_magnitude_consistency,
            "amount_coverage":         v.total_amount_coverage,
        }

        for signal_name, weight in profile.items():
            value = signal_values.get(signal_name, 0.5)
            weighted = value * weight
            breakdown[signal_name] = round(weighted, 4)
            composite += weighted

        # Bonus: if schema type was also detected by filename/keyword
        # (incorporate legacy SchemaDetector signals)
        # This gives a small boost to hypotheses that have keyword evidence
        schema_config = _SCHEMA_SIGNALS.get(hypothesis.schema_type, {})
        # We don't have filename here, so skip filename bonus

        hypothesis.composite_score = round(composite, 4)
        hypothesis.scoring_breakdown = breakdown
        return hypothesis

    def rank_hypotheses(
        self,
        hypotheses: List[SchemaHypothesis],
    ) -> List[SchemaHypothesis]:
        """Rank hypotheses by composite score, assign ranks."""
        hypotheses.sort(key=lambda h: h.composite_score, reverse=True)
        for i, h in enumerate(hypotheses):
            h.rank = i + 1
            h.is_winner = (i == 0)
        return hypotheses


# ── HypothesisDrivenParser ───────────────────────────────────────────────────

class HypothesisDrivenParser:
    """
    Orchestrates the full Hypothesis-Driven Parsing pipeline.

    Flow:
        1. Detect table structure (header, regions)
        2. Generate N hypotheses (one per schema type)
        3. Parse data under each hypothesis
        4. Validate using accounting invariants
        5. Score and rank hypotheses
        6. Select winner
        7. Apply schema memory boost (if available)

    This replaces the sequential ingestion pipeline with a
    self-correcting multi-hypothesis search.
    """

    def __init__(self, schema_memory=None):
        self._structure = FileStructureDetector()
        self._generator = HypothesisGenerator()
        self._validator = HypothesisValidator()
        self._scorer = HypothesisScorer()
        self._schema_memory = schema_memory
        self._legacy_detector = SchemaDetector()

    def parse_sheet(
        self,
        rows: List[tuple],
        filename: str = "unknown.xlsx",
        sheet_name: str = "Sheet1",
    ) -> HypothesisParseResult:
        """
        Parse a single sheet using hypothesis-driven approach.

        Returns HypothesisParseResult with the winning interpretation,
        all scored alternatives, and full diagnostic data.
        """
        if not rows:
            empty_det = DetectionResult(
                schema_type=SchemaType.UNKNOWN,
                confidence=0.0,
                filename=filename,
                sheet_name=sheet_name,
            )
            return HypothesisParseResult(
                winner=SchemaHypothesis(
                    schema_type=SchemaType.UNKNOWN,
                    column_mapping=ColumnMapping(),
                ),
                all_hypotheses=[],
                detection=empty_det,
                records=[],
                ingestion_warnings=["Empty sheet"],
            )

        # Step 1: Detect structure
        header_idx = self._structure.find_header_row(rows)
        header_row = rows[header_idx]
        start, end = self._structure.detect_table_region(rows, header_idx)
        data_rows = rows[start:end + 1]

        # Step 2: Generate hypotheses
        hypotheses = self._generator.generate(
            header_row, data_rows, filename, sheet_name
        )

        # Step 3: Parse + Validate each hypothesis
        for hyp in hypotheses:
            self._validator.parse_and_validate(hyp, header_row, data_rows)

        # Step 4: Score each hypothesis
        for hyp in hypotheses:
            self._scorer.score(hyp)

        # Step 5: Apply legacy keyword/filename bonus
        self._apply_keyword_bonus(hypotheses, filename, sheet_name, header_row, data_rows)

        # Step 6: Apply schema memory boost if available
        if self._schema_memory is not None:
            self._apply_schema_memory_boost(
                hypotheses, header_row, data_rows, filename
            )

        # Step 7: Rank and select winner
        hypotheses = self._scorer.rank_hypotheses(hypotheses)

        winner = hypotheses[0] if hypotheses else SchemaHypothesis(
            schema_type=SchemaType.UNKNOWN,
            column_mapping=ColumnMapping(),
        )

        # Build detection result (backward compatible)
        detection = DetectionResult(
            schema_type=winner.schema_type,
            confidence=winner.confidence,
            filename=filename,
            sheet_name=sheet_name,
            header_row=header_idx,
            columns={
                k: v for k, v in winner.column_mapping.roles.items()
                if v is not None
            },
            row_count=len(data_rows),
            signals=[
                f"{k}={v:.3f}" for k, v in winner.scoring_breakdown.items()
            ],
            warnings=(
                winner.parse_result.warnings if winner.parse_result else []
            ),
        )

        records = winner.parse_result.records if winner.parse_result else []

        # Collect ingestion warnings
        ingestion_warnings: List[str] = []
        if winner.parse_result and winner.parse_result.dropped_rows > 0:
            pr = winner.parse_result
            ingestion_warnings.append(
                f"Parsed {pr.parsed_rows}/{pr.total_rows} rows "
                f"({pr.dropped_rows} dropped)"
            )
            for reason, count in pr.drop_reasons.items():
                ingestion_warnings.append(f"  Drop reason: {reason} ({count})")

        # Add runner-up information
        if len(hypotheses) >= 2:
            h2 = hypotheses[1]
            gap = winner.composite_score - h2.composite_score
            if gap < 0.05:
                ingestion_warnings.append(
                    f"AMBIGUOUS: Runner-up {h2.schema_type} scored "
                    f"{h2.composite_score:.3f} vs winner {winner.composite_score:.3f} "
                    f"(gap={gap:.3f}). Consider manual review."
                )

        # Build metadata
        metadata: Dict[str, Any] = {
            "hdp_version": "1.0",
            "hypotheses_evaluated": len(hypotheses),
            "winner_schema": winner.schema_type,
            "winner_score": winner.composite_score,
            "scoring_breakdown": winner.scoring_breakdown,
            "runner_up": {
                "schema": hypotheses[1].schema_type if len(hypotheses) >= 2 else None,
                "score": hypotheses[1].composite_score if len(hypotheses) >= 2 else 0,
            },
            "all_scores": {
                h.schema_type: {
                    "score": h.composite_score,
                    "rank": h.rank,
                    "parsed_rows": h.parse_result.parsed_rows if h.parse_result else 0,
                    "dropped_rows": h.parse_result.dropped_rows if h.parse_result else 0,
                }
                for h in hypotheses
            },
        }

        # Record fingerprint in schema memory
        if self._schema_memory is not None and records:
            try:
                self._schema_memory.record_successful_parse(
                    header_row=header_row,
                    data_rows=data_rows,
                    schema_type=winner.schema_type,
                    confidence=winner.confidence,
                    column_mapping=winner.column_mapping,
                    filename=filename,
                )
            except Exception as e:
                logger.warning("Schema memory recording failed: %s", e)

        return HypothesisParseResult(
            winner=winner,
            all_hypotheses=hypotheses,
            detection=detection,
            records=records,
            metadata=metadata,
            ingestion_warnings=ingestion_warnings,
        )

    def _apply_keyword_bonus(
        self,
        hypotheses: List[SchemaHypothesis],
        filename: str,
        sheet_name: str,
        header_row: tuple,
        data_rows: List[tuple],
    ) -> None:
        """
        Apply keyword/filename evidence as a bonus to composite scores.
        This integrates the legacy SchemaDetector's keyword matching.
        """
        filename_lower = filename.lower()
        sheet_lower = sheet_name.lower()
        header_text = " ".join(
            str(c).lower() for c in header_row if c is not None
        )
        all_text = " ".join(
            str(c).lower()
            for row in data_rows[:30]
            for c in row
            if c is not None and isinstance(c, str)
        )

        for hyp in hypotheses:
            # Never boost hypotheses that parsed zero rows —
            # keyword evidence alone cannot overcome zero data parse
            if hyp.parse_result is not None and hyp.parse_result.parsed_rows == 0:
                continue

            config = _SCHEMA_SIGNALS.get(hyp.schema_type, {})
            bonus = 0.0

            # Filename keyword match
            for kw in config.get("keywords", []):
                if kw in filename_lower or kw in sheet_lower:
                    bonus += 0.08  # ~8% boost for filename match
                    break

            # Header keyword match
            for kw in config.get("keywords", []):
                if kw in header_text:
                    bonus += 0.05  # ~5% boost for header match
                    break

            # Row content keyword matches
            row_kws = config.get("row_keyword", [])
            hits = sum(1 for kw in row_kws if kw in all_text)
            bonus += hits * 0.02  # ~2% per row keyword hit

            hyp.composite_score += bonus
            if bonus > 0:
                hyp.scoring_breakdown["keyword_bonus"] = round(bonus, 4)

    def _apply_schema_memory_boost(
        self,
        hypotheses: List[SchemaHypothesis],
        header_row: tuple,
        data_rows: List[tuple],
        filename: str,
    ) -> None:
        """
        Boost hypotheses that match known successful parse patterns.
        This is where Schema Memory integrates with HDP.
        """
        if self._schema_memory is None:
            return

        try:
            match = self._schema_memory.find_match(header_row, data_rows, filename)
            if match is not None:
                matched_type, match_confidence = match
                for hyp in hypotheses:
                    if hyp.schema_type == matched_type:
                        boost = match_confidence * 0.15  # Up to 15% boost
                        hyp.composite_score += boost
                        hyp.scoring_breakdown["schema_memory_boost"] = round(boost, 4)
                        break
        except Exception as e:
            logger.warning("Schema memory lookup failed: %s", e)

    # ── Segmented parsing (multi-table sheets) ─────────────────────

    def parse_sheet_segmented(
        self,
        rows: List[tuple],
        filename: str = "unknown.xlsx",
        sheet_name: str = "Sheet1",
    ) -> List[HypothesisParseResult]:
        """
        Parse a sheet that may contain multiple table regions.

        Strategy:
        1. Run TableSegmenter to detect regions
        2. If multiple table regions found, run HDP per region
        3. Run ConstraintGraph to validate cross-table relationships
        4. Apply constraint penalties/bonuses to hypothesis scores
        5. Return one HypothesisParseResult per table region

        Falls back to single-table parse_sheet() if:
        - Only one region detected
        - Segmentation fails
        """
        try:
            from app.services.table_segmenter import TableSegmenter
            segmenter = TableSegmenter()
            seg_result = segmenter.segment(rows)
        except Exception as e:
            logger.warning("Table segmentation failed: %s, using single-table parse", e)
            return [self.parse_sheet(rows, filename, sheet_name)]

        tables = seg_result.get_tables()

        # If only one table (or none), use standard parse
        if len(tables) <= 1:
            result = self.parse_sheet(rows, filename, sheet_name)
            # Attach segmentation metadata
            result.metadata["segmentation"] = seg_result.to_dict()
            return [result]

        # Multiple tables detected — parse each independently
        results: List[HypothesisParseResult] = []
        for region in tables:
            region_rows = region.to_rows()
            region_name = f"{sheet_name}:R{region.region_id}"
            result = self.parse_sheet(region_rows, filename, region_name)

            # Attach region metadata
            result.metadata["region"] = region.to_dict()
            result.metadata["segmentation"] = {
                "total_regions": len(seg_result.regions),
                "table_regions": len(tables),
                "title_block": seg_result.title_block,
            }
            results.append(result)

        # Run constraint graph validation across all parsed tables
        try:
            from app.services.constraint_graph import ConstraintGraph
            graph = ConstraintGraph()

            parsed_tables = [
                {
                    "schema_type": r.schema_type,
                    "records": r.records,
                    "confidence": r.confidence,
                }
                for r in results
            ]
            validation = graph.validate(parsed_tables)

            # Attach validation report to each result
            for r in results:
                r.metadata["constraint_validation"] = validation.to_dict()

            # Apply confidence adjustments
            if validation.overall_confidence_adjustment != 0:
                for r in results:
                    adj = validation.overall_confidence_adjustment
                    if r.winner:
                        old_score = r.winner.composite_score
                        r.winner.composite_score = max(0, old_score + adj)
                        r.winner.scoring_breakdown["constraint_adjustment"] = adj

            # Log validation results
            logger.info(
                "ConstraintGraph: %d constraints checked, %d passed, "
                "%d failed, model_valid=%s",
                validation.total_constraints,
                validation.passed,
                validation.failed,
                validation.financial_model_valid,
            )
        except Exception as e:
            logger.warning("Constraint graph validation failed: %s", e)

        return results

    def process_file(self, path: str) -> List[HypothesisParseResult]:
        """Process all sheets in an Excel file using HDP with segmentation."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl required")

        from pathlib import Path as P
        p = P(path)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        all_results = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            results = self.parse_sheet_segmented(rows, p.name, sheet_name)
            all_results.extend(results)

        wb.close()
        return all_results

    def process_bytes(self, data: bytes, filename: str) -> List[HypothesisParseResult]:
        """Process Excel bytes (HTTP upload) using HDP with segmentation."""
        import io
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl required")

        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=True, read_only=True
        )
        all_results = []

        from pathlib import Path as P
        p = P(filename)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            results = self.parse_sheet_segmented(rows, p.name, sheet_name)
            all_results.extend(results)

        wb.close()
        return all_results


# ── Module-level singleton ───────────────────────────────────────────────────
# Schema memory will be attached after schema_memory module loads
hdp_parser = HypothesisDrivenParser(schema_memory=None)
