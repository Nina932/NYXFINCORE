"""
Dynamic schema registry backed by DB profiles + versions.
Supports strict validation with adaptive, versioned schemas.
"""
from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
import io
import csv
import openpyxl

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.all_models import SchemaProfile, SchemaVersion, SchemaProposal
from app.services.schema_registry import SchemaRegistry, SheetSchema, _find_first_non_empty_row, _norm_header


@dataclass
class DynamicSchemaValidationResult:
    ok: bool
    file_type: str
    sheet_results: List[Dict]
    errors: List[str]
    warnings: List[str]
    profile_id: Optional[int] = None
    proposal_id: Optional[int] = None


def _load_sheet_rows(filename: str, content: bytes, sample_rows: int = 50) -> Dict[str, List[List]]:
    ext = filename.lower().rsplit('.', 1)[-1]
    sheet_rows: Dict[str, List[List]] = {}
    if ext in ("xlsx", "xls", "xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if sample_rows and i >= sample_rows:
                    break
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheet_rows[name] = rows
    elif ext == "csv":
        decoded = content.decode("utf-8", errors="replace")
        reader = list(csv.reader(io.StringIO(decoded)))
        sheet_rows["Sheet1"] = reader[:sample_rows] if sample_rows else reader
    return sheet_rows


def _serialize_registry() -> Dict:
    registry = SchemaRegistry()
    rules = []
    for ftype, schemas in registry.schemas.items():
        for s in schemas:
            rules.append({
                "name": ftype,
                "required_all": s.required_all,
                "required_any_groups": s.required_any_groups,
                "numeric_columns_any": s.numeric_columns_any,
            })
    return {"header_scan_rows": 200, "sheets": rules}


def _looks_like_data_row(row: List[str]) -> bool:
    vals = [str(c).strip() for c in row if str(c).strip()]
    if not vals:
        return False
    if not any(any(ch.isalpha() for ch in v) for v in vals):
        return True
    data_like = 0
    for v in vals:
        compact = v.replace(".", "").replace("/", "").replace("-", "")
        if compact.isdigit():
            data_like += 1
        elif any(ch.isdigit() for ch in v) and len(v) <= 10:
            data_like += 1
    return (data_like / len(vals)) >= 0.5


def _build_rules_from_headers(sheet_rows: Dict[str, List[List]]) -> Dict:
    rules = []
    for sheet_name, rows in sheet_rows.items():
        _, header = _find_first_non_empty_row(rows)
        norm = [_norm_header(h) for h in header if _norm_header(h)]
        alpha_headers = sum(1 for h in norm if any(ch.isalpha() for ch in h))
        headerless = len(norm) == 0 or _looks_like_data_row(header)
        if not headerless and alpha_headers < 2:
            if any(_looks_like_data_row(r) for r in rows[:10]):
                headerless = True
        col_count = max((len(r) for r in rows[:10]), default=len(header))
        rules.append({
            "name": sheet_name,
            "required_all": [] if headerless else norm,
            "required_any_groups": [],
            "numeric_columns_any": [],
            "headerless": headerless,
            "column_count_min": col_count,
        })
    return {"header_scan_rows": 200, "sheets": rules}


def _has_structured_data(rows: List[List]) -> bool:
    for row in rows:
        nonempty = [c for c in row if str(c).strip()]
        if len(nonempty) >= 3 and any(any(ch.isdigit() for ch in str(c)) for c in nonempty):
            return True
    return False


def _validate_with_rules(sheet_rows: Dict[str, List[List]], rules_json: Dict) -> Tuple[bool, List[Dict], List[str]]:
    errors: List[str] = []
    sheet_results: List[Dict] = []
    rules = rules_json.get("sheets", [])
    scan_rows = int(rules_json.get("header_scan_rows", 50) or 50)

    schemas = []
    schema_extras = {}
    for r in rules:
        schemas.append(SheetSchema(
            name=r.get("name", ""),
            required_all=r.get("required_all", []),
            required_any_groups=r.get("required_any_groups", []),
            numeric_columns_any=r.get("numeric_columns_any", []),
        ))
        schema_extras[r.get("name", "")] = {
            "headerless": bool(r.get("headerless")),
            "column_count_min": int(r.get("column_count_min", 0) or 0),
        }

    for sheet_name, rows in sheet_rows.items():
        matched = None
        header_row_idx = None
        if not _has_structured_data(rows[:scan_rows]):
            sheet_results.append({
                "sheet": sheet_name,
                "matched": "ignored",
                "header_row": None,
                "errors": [],
                "warnings": ["sheet_empty_or_decorative"],
            })
            continue
        for i, row in enumerate(rows[:scan_rows]):
            if not row or not any(str(c).strip() for c in row):
                continue
            for schema in schemas:
                schema_name = (schema.name or "").strip().lower()
                if schema_name and schema_name not in ("*", "any") and schema_name != sheet_name.lower():
                    continue
                extras = schema_extras.get(schema.name or "", {})
                if extras.get("headerless"):
                    col_min = extras.get("column_count_min", 0)
                    if len(row) >= col_min and sum(1 for c in row if str(c).strip()) >= 3:
                        matched = schema.name or "matched"
                        header_row_idx = i
                        break
                errs = schema.validate_headers(row)
                if not errs:
                    matched = schema.name or "matched"
                    header_row_idx = i
                    break
            if matched:
                break

        if matched is None:
            errors.append(f"Sheet '{sheet_name}' does not match any active schema profile.")
        sheet_results.append({
            "sheet": sheet_name,
            "matched": matched,
            "header_row": header_row_idx,
            "errors": [] if matched else ["schema_not_matched"],
            "warnings": [],
        })

    return len(errors) == 0, sheet_results, errors


async def _ensure_default_profile(db: AsyncSession) -> SchemaProfile:
    res = await db.execute(select(SchemaProfile).where(SchemaProfile.is_active == True))
    existing = res.scalars().all()
    if existing:
        return existing[0]

    profile = SchemaProfile(name="Default 1C Schema", file_type="generic", is_active=True)
    db.add(profile)
    await db.flush()
    rules = _serialize_registry()
    db.add(SchemaVersion(profile_id=profile.id, version=1, rules_json=rules, is_active=True))
    await db.flush()
    return profile


async def validate_schema_db(filename: str, content: bytes, db: AsyncSession, sample_rows: int = 50) -> DynamicSchemaValidationResult:
    sheet_rows = _load_sheet_rows(filename, content, sample_rows)
    await _ensure_default_profile(db)

    profiles = (await db.execute(select(SchemaProfile).where(SchemaProfile.is_active == True))).scalars().all()
    active_versions = (await db.execute(select(SchemaVersion).where(SchemaVersion.is_active == True))).scalars().all()
    version_by_profile = {v.profile_id: v for v in active_versions}

    for p in profiles:
        v = version_by_profile.get(p.id)
        if not v or not v.rules_json:
            continue
        ok, sheet_results, errors = _validate_with_rules(sheet_rows, v.rules_json)
        if ok:
            return DynamicSchemaValidationResult(
                ok=True,
                file_type=p.file_type or "Financial Data",
                sheet_results=sheet_results,
                errors=[],
                warnings=[],
                profile_id=p.id,
            )

    # No match — create proposal
    existing = (await db.execute(
        select(SchemaProposal).where(
            SchemaProposal.file_name == filename,
            SchemaProposal.status == "pending"
        )
    )).scalars().first()
    if existing:
        proposal = existing
    else:
        proposal = SchemaProposal(
            file_name=filename,
            sheet_names=list(sheet_rows.keys()),
            header_samples={k: v[:5] for k, v in sheet_rows.items()},
            suggested_rules_json=_build_rules_from_headers(sheet_rows),
            status="pending",
        )
        db.add(proposal)
        await db.flush()

    ok, sheet_results, errors = _validate_with_rules(sheet_rows, _build_rules_from_headers(sheet_rows))
    return DynamicSchemaValidationResult(
        ok=False,
        file_type="unknown",
        sheet_results=sheet_results,
        errors=errors,
        warnings=[],
        proposal_id=proposal.id,
    )
