"""
FinAI Data Connectors Framework
=================================
Pluggable connectors for external financial data sources:

  1. OneCConnector   — 1C:Enterprise v8 via HTTP web services
  2. BankCSVConnector — Georgian bank CSV statements (TBC, BOG, Liberty)
  3. SAPConnector    — SAP ERP via OData v4
  4. ExcelConnector  — Wraps existing upload pipeline

All connectors implement BaseConnector and produce data in the same
format as the smart-upload pipeline (Dict[str, float] financials,
GL transaction rows, COA rows, etc.).
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import re
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime, date
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

import time

try:
    import httpx
    HAS_HTTPX = True
except ImportError:
    HAS_HTTPX = False

try:
    import paramiko
    HAS_PARAMIKO = True
except ImportError:
    HAS_PARAMIKO = False

import aiohttp

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
# Connection Config
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ConnectorConfig:
    """Universal connector configuration."""
    connector_type: str              # "1c", "bank_csv", "sap", "excel"
    name: str = ""                   # Human-readable label
    server_url: str = ""             # Base URL for HTTP-based connectors
    database: str = ""               # 1C database name, SAP system ID
    username: str = ""
    password: str = ""
    # SAP-specific
    host: str = ""
    system_number: str = "00"
    client: str = "100"
    # Bank CSV-specific
    bank_format: str = ""            # "tbc", "bog", "liberty", "auto"
    encoding: str = "utf-8"
    # Extra params
    extra: Dict[str, Any] = field(default_factory=dict)

    def to_safe_dict(self) -> Dict[str, Any]:
        """Return config without secrets."""
        return {
            "connector_type": self.connector_type,
            "name": self.name,
            "server_url": self.server_url,
            "database": self.database,
            "username": self.username,
            "host": self.host,
            "bank_format": self.bank_format,
            "has_password": bool(self.password),
        }


@dataclass
class ConnectorResult:
    """Standardized result from any connector fetch operation."""
    success: bool
    connector_type: str
    data_type: str                    # "trial_balance", "journal_entries", "chart_of_accounts", "bank_statement"
    records: List[Dict[str, Any]]     # Raw records
    record_count: int = 0
    period: str = ""
    currency: str = "GEL"
    source: str = ""
    warnings: List[str] = field(default_factory=list)
    error: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)

    # Pre-processed financials dict (for direct pipeline ingestion)
    financials: Optional[Dict[str, float]] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "connector_type": self.connector_type,
            "data_type": self.data_type,
            "record_count": self.record_count,
            "period": self.period,
            "currency": self.currency,
            "source": self.source,
            "warnings": self.warnings,
            "error": self.error,
            "metadata": self.metadata,
            "records_sample": self.records[:20],
            "has_financials": self.financials is not None,
        }


# ═══════════════════════════════════════════════════════════════════════════════
# Base Connector (Abstract)
# ═══════════════════════════════════════════════════════════════════════════════

class BaseConnector(ABC):
    """Abstract base class for all data connectors."""

    CONNECTOR_TYPE: str = "base"
    DISPLAY_NAME: str = "Base Connector"
    DESCRIPTION: str = ""

    def __init__(self, config: ConnectorConfig):
        self.config = config

    @abstractmethod
    async def test_connection(self) -> Dict[str, Any]:
        """Test the connection to the data source.

        Returns: {"connected": bool, "message": str, "details": {...}}
        """
        ...

    @abstractmethod
    async def fetch_data(
        self, data_type: str, params: Dict[str, Any] = None
    ) -> ConnectorResult:
        """Fetch data from the source.

        Args:
            data_type: "trial_balance", "journal_entries", "chart_of_accounts"
            params: Type-specific params (period, date_from, date_to, etc.)

        Returns: ConnectorResult with records and optional financials dict.
        """
        ...

    @abstractmethod
    async def get_schema(self) -> Dict[str, Any]:
        """Return the schema/structure of available data.

        Returns: {"data_types": [...], "fields": {...}, "periods": [...]}
        """
        ...

    def info(self) -> Dict[str, Any]:
        """Return public connector info."""
        return {
            "connector_type": self.CONNECTOR_TYPE,
            "display_name": self.DISPLAY_NAME,
            "description": self.DESCRIPTION,
            "config": self.config.to_safe_dict(),
        }


# ═══════════════════════════════════════════════════════════════════════════════
# 1C:Enterprise Connector (HTTP Web Services)
# ═══════════════════════════════════════════════════════════════════════════════

class OneCConnector(BaseConnector):
    """Connect to 1C:Enterprise v8 via published HTTP web services.

    1C publishes web services (SOAP/REST) via its built-in HTTP service.
    Common endpoints:
      /hs/odata/standard.odata/AccountingRegister_...
      /hs/accounting/TrialBalance
      /hs/accounting/JournalEntries
      /hs/accounting/ChartOfAccounts

    The exact URL paths depend on the 1C configuration publication.
    This connector supports the common HTTP service patterns used
    by Georgian 1C installations.
    """

    CONNECTOR_TYPE = "1c"
    DISPLAY_NAME = "1C:Enterprise"
    DESCRIPTION = "Connect to 1C:Enterprise v8 accounting via HTTP web services"

    def _base_url(self) -> str:
        url = self.config.server_url.rstrip("/")
        if self.config.database:
            url = f"{url}/{self.config.database}"
        return url

    def _auth(self) -> aiohttp.BasicAuth:
        return aiohttp.BasicAuth(
            login=self.config.username,
            password=self.config.password,
        )

    async def _request(
        self, path: str, params: Dict[str, str] = None, method: str = "GET"
    ) -> Optional[Dict[str, Any]]:
        """Make an authenticated HTTP request to 1C."""
        url = f"{self._base_url()}{path}"
        try:
            async with aiohttp.ClientSession(auth=self._auth()) as session:
                async with session.request(
                    method,
                    url,
                    params=params,
                    headers={"Accept": "application/json", "Content-Type": "application/json"},
                    timeout=aiohttp.ClientTimeout(total=30),
                    ssl=False,  # Many 1C installations use self-signed certs
                ) as resp:
                    if resp.status == 200:
                        text = await resp.text()
                        try:
                            return json.loads(text)
                        except json.JSONDecodeError:
                            # 1C sometimes returns XML even when JSON is requested
                            return {"raw_response": text, "status": resp.status}
                    else:
                        body = await resp.text()
                        logger.warning("1C API %s returned %d: %s", path, resp.status, body[:300])
                        return {"error": f"HTTP {resp.status}", "body": body[:300]}
        except aiohttp.ClientConnectorError as e:
            logger.debug("1C connection failed: %s", e)
            return None
        except Exception as e:
            logger.debug("1C request failed: %s", e)
            return None

    async def test_connection(self) -> Dict[str, Any]:
        """Test connection to 1C HTTP service."""
        # Try the standard OData metadata endpoint
        for path in ["/hs/odata/standard.odata/$metadata", "/hs/accounting/ping", "/hs/odata/"]:
            result = await self._request(path)
            if result is not None and "error" not in result:
                return {
                    "connected": True,
                    "message": f"Connected to 1C at {self._base_url()}",
                    "details": {"endpoint": path, "response_type": type(result).__name__},
                }
        return {
            "connected": False,
            "message": f"Cannot connect to 1C at {self._base_url()}",
            "details": {"tried_endpoints": ["/hs/odata/standard.odata/$metadata", "/hs/accounting/ping"]},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["trial_balance", "journal_entries", "chart_of_accounts"],
            "fields": {
                "trial_balance": ["account_code", "account_name", "debit_opening", "credit_opening",
                                  "debit_turnover", "credit_turnover", "debit_closing", "credit_closing"],
                "journal_entries": ["date", "document", "account_debit", "account_credit",
                                    "amount", "description", "currency"],
                "chart_of_accounts": ["code", "name", "type", "parent_code", "is_group"],
            },
            "source": self._base_url(),
        }

    async def fetch_data(
        self, data_type: str, params: Dict[str, Any] = None
    ) -> ConnectorResult:
        params = params or {}

        if data_type == "trial_balance":
            return await self._fetch_trial_balance(params)
        elif data_type == "journal_entries":
            return await self._fetch_journal_entries(params)
        elif data_type == "chart_of_accounts":
            return await self._fetch_chart_of_accounts(params)
        else:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=[],
                error=f"Unknown data_type: {data_type}",
            )

    async def _fetch_trial_balance(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch trial balance from 1C.

        Tries multiple endpoint patterns:
          1. /hs/accounting/TrialBalance?period=YYYY-MM
          2. /hs/odata/standard.odata/AccountingRegister_Хозрасчетный_BalanceAndTurnover
        """
        period = params.get("period", datetime.now().strftime("%Y-%m"))

        # Try custom HTTP service first
        data = await self._request("/hs/accounting/TrialBalance", {"period": period})
        if data and "error" not in data:
            records = data.get("accounts", data.get("rows", data.get("value", [])))
            if isinstance(records, list):
                return self._build_tb_result(records, period)

        # Try OData endpoint
        data = await self._request(
            "/hs/odata/standard.odata/AccountingRegister_Хозрасчетный_BalanceAndTurnover",
            {"$filter": f"Period eq '{period}'", "$format": "json"},
        )
        if data and "error" not in data:
            records = data.get("value", data.get("rows", []))
            if isinstance(records, list):
                return self._build_tb_result(records, period)

        return ConnectorResult(
            success=False, connector_type=self.CONNECTOR_TYPE,
            data_type="trial_balance", records=[],
            period=period, source=self._base_url(),
            error="Could not fetch trial balance from 1C — check endpoint configuration",
        )

    def _build_tb_result(self, records: List[Dict], period: str) -> ConnectorResult:
        """Normalize 1C TB records to standard format."""
        normalized = []
        for r in records:
            normalized.append({
                "account_code": str(r.get("Account_Code", r.get("code", r.get("Счет_Code", "")))),
                "account_name": r.get("Account_Description", r.get("name", r.get("Счет_Description", ""))),
                "debit_opening": float(r.get("OpeningBalanceDr", r.get("debit_opening", 0)) or 0),
                "credit_opening": float(r.get("OpeningBalanceCr", r.get("credit_opening", 0)) or 0),
                "debit_turnover": float(r.get("TurnoverDr", r.get("debit_turnover", 0)) or 0),
                "credit_turnover": float(r.get("TurnoverCr", r.get("credit_turnover", 0)) or 0),
                "debit_closing": float(r.get("ClosingBalanceDr", r.get("debit_closing", 0)) or 0),
                "credit_closing": float(r.get("ClosingBalanceCr", r.get("credit_closing", 0)) or 0),
            })

        # Build financials dict from TB data for pipeline ingestion
        financials = self._tb_to_financials(normalized)

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="trial_balance", records=normalized,
            record_count=len(normalized), period=period,
            currency="GEL", source=self._base_url(),
            financials=financials,
            metadata={"raw_count": len(records)},
        )

    def _tb_to_financials(self, records: List[Dict]) -> Dict[str, float]:
        """Convert trial balance records to financials dict.

        Maps account codes to P&L lines using standard Georgian/1C ranges:
          6xxx = Revenue, 7xxx = COGS, 8xxx = OpEx, 9xxx = Other
        """
        revenue = 0.0
        cogs = 0.0
        opex = 0.0
        other_income = 0.0
        other_expense = 0.0

        for r in records:
            code = r.get("account_code", "")
            credit_turnover = r.get("credit_turnover", 0)
            debit_turnover = r.get("debit_turnover", 0)

            if code.startswith("6"):
                revenue += credit_turnover - debit_turnover
            elif code.startswith("7"):
                cogs += debit_turnover - credit_turnover
            elif code.startswith("8"):
                opex += debit_turnover - credit_turnover
            elif code.startswith("91") or code.startswith("92"):
                other_income += credit_turnover - debit_turnover
            elif code.startswith("93") or code.startswith("94") or code.startswith("95"):
                other_expense += debit_turnover - credit_turnover

        gross_profit = revenue - cogs
        operating_income = gross_profit - opex
        net_income = operating_income + other_income - other_expense

        return {
            "revenue": revenue,
            "cost_of_goods_sold": cogs,
            "gross_profit": gross_profit,
            "operating_expenses": opex,
            "operating_income": operating_income,
            "other_income": other_income,
            "other_expenses": other_expense,
            "net_income": net_income,
        }

    async def _fetch_journal_entries(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch journal entries (GL) from 1C."""
        date_from = params.get("date_from", "")
        date_to = params.get("date_to", "")

        query_params = {}
        if date_from and date_to:
            query_params["$filter"] = f"Date ge datetime'{date_from}T00:00:00' and Date le datetime'{date_to}T23:59:59'"
        query_params["$format"] = "json"
        query_params["$top"] = str(params.get("limit", 5000))

        # Try custom HTTP service
        data = await self._request(
            "/hs/accounting/JournalEntries",
            {"date_from": date_from, "date_to": date_to, "limit": str(params.get("limit", 5000))},
        )
        if data and "error" not in data:
            records = data.get("entries", data.get("rows", data.get("value", [])))
            if isinstance(records, list):
                normalized = []
                for r in records:
                    normalized.append({
                        "date": r.get("Date", r.get("date", "")),
                        "document": r.get("Document", r.get("Recorder", r.get("document", ""))),
                        "account_debit": str(r.get("AccountDr_Code", r.get("account_debit", ""))),
                        "account_credit": str(r.get("AccountCr_Code", r.get("account_credit", ""))),
                        "amount": float(r.get("Amount", r.get("amount", 0)) or 0),
                        "description": r.get("Content", r.get("description", r.get("Comment", ""))),
                        "currency": r.get("Currency", r.get("currency", "GEL")),
                    })
                return ConnectorResult(
                    success=True, connector_type=self.CONNECTOR_TYPE,
                    data_type="journal_entries", records=normalized,
                    record_count=len(normalized),
                    period=f"{date_from} to {date_to}",
                    source=self._base_url(),
                )

        # Try OData
        data = await self._request(
            "/hs/odata/standard.odata/AccountingRegister_Хозрасчетный",
            query_params,
        )
        if data and "error" not in data:
            records = data.get("value", [])
            if isinstance(records, list):
                normalized = []
                for r in records:
                    normalized.append({
                        "date": r.get("Period", r.get("Date", "")),
                        "document": r.get("Recorder", ""),
                        "account_debit": str(r.get("AccountDr_Key", "")),
                        "account_credit": str(r.get("AccountCr_Key", "")),
                        "amount": float(r.get("Amount", 0) or 0),
                        "description": r.get("Content", ""),
                        "currency": r.get("Currency_Key", "GEL"),
                    })
                return ConnectorResult(
                    success=True, connector_type=self.CONNECTOR_TYPE,
                    data_type="journal_entries", records=normalized,
                    record_count=len(normalized),
                    period=f"{date_from} to {date_to}",
                    source=self._base_url(),
                )

        return ConnectorResult(
            success=False, connector_type=self.CONNECTOR_TYPE,
            data_type="journal_entries", records=[],
            error="Could not fetch journal entries from 1C",
            source=self._base_url(),
        )

    async def _fetch_chart_of_accounts(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch chart of accounts from 1C."""
        # Try custom endpoint
        data = await self._request("/hs/accounting/ChartOfAccounts")
        if data and "error" not in data:
            records = data.get("accounts", data.get("rows", data.get("value", [])))
            if isinstance(records, list):
                normalized = []
                for r in records:
                    normalized.append({
                        "code": str(r.get("Code", r.get("code", ""))),
                        "name": r.get("Description", r.get("name", r.get("Наименование", ""))),
                        "type": r.get("Type", r.get("type", "")),
                        "parent_code": str(r.get("Parent_Code", r.get("parent_code", ""))),
                        "is_group": bool(r.get("IsFolder", r.get("is_group", False))),
                    })
                return ConnectorResult(
                    success=True, connector_type=self.CONNECTOR_TYPE,
                    data_type="chart_of_accounts", records=normalized,
                    record_count=len(normalized),
                    source=self._base_url(),
                )

        # Try OData catalog
        data = await self._request(
            "/hs/odata/standard.odata/ChartOfAccounts_Хозрасчетный",
            {"$format": "json"},
        )
        if data and "error" not in data:
            records = data.get("value", [])
            if isinstance(records, list):
                normalized = []
                for r in records:
                    normalized.append({
                        "code": str(r.get("Code", "")),
                        "name": r.get("Description", ""),
                        "type": r.get("Type", ""),
                        "parent_code": str(r.get("Parent_Key", "")),
                        "is_group": bool(r.get("IsFolder", False)),
                    })
                return ConnectorResult(
                    success=True, connector_type=self.CONNECTOR_TYPE,
                    data_type="chart_of_accounts", records=normalized,
                    record_count=len(normalized),
                    source=self._base_url(),
                )

        return ConnectorResult(
            success=False, connector_type=self.CONNECTOR_TYPE,
            data_type="chart_of_accounts", records=[],
            error="Could not fetch chart of accounts from 1C",
            source=self._base_url(),
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Georgian Bank CSV Connector
# ═══════════════════════════════════════════════════════════════════════════════

class BankCSVConnector(BaseConnector):
    """Import bank statements from Georgian banks in CSV format.

    Supports auto-detection of format from:
      - TBC Bank (თიბისი ბანკი)
      - Bank of Georgia (საქართველოს ბანკი)
      - Liberty Bank (ლიბერთი ბანკი)

    Outputs standardized transaction records compatible with the GL pipeline.
    """

    CONNECTOR_TYPE = "bank_csv"
    DISPLAY_NAME = "Georgian Bank CSV"
    DESCRIPTION = "Import CSV bank statements from TBC Bank, Bank of Georgia, and Liberty Bank"

    # Column name patterns for auto-detection (Georgian + English)
    _BANK_SIGNATURES = {
        "tbc": {
            "markers": ["თიბისი", "TBC", "tbc bank"],
            "date_cols": ["თარიღი", "Date", "ოპერაციის თარიღი", "Transaction Date"],
            "desc_cols": ["აღწერა", "Description", "დანიშნულება", "Purpose"],
            "debit_cols": ["დებეტი", "Debit", "გასავალი", "Withdrawal"],
            "credit_cols": ["კრედიტი", "Credit", "შემოსავალი", "Deposit"],
            "balance_cols": ["ნაშთი", "Balance", "ბალანსი"],
            "date_format": "%d/%m/%Y",
        },
        "bog": {
            "markers": ["საქართველოს ბანკი", "Bank of Georgia", "BOG"],
            "date_cols": ["თარიღი", "Date", "Transaction date"],
            "desc_cols": ["აღწერა", "Description", "მიზანი", "Narrative"],
            "debit_cols": ["დებეტი", "Debit", "თანხა (დებეტი)"],
            "credit_cols": ["კრედიტი", "Credit", "თანხა (კრედიტი)"],
            "balance_cols": ["ნაშთი", "Balance", "საბოლოო ნაშთი"],
            "date_format": "%d.%m.%Y",
        },
        "liberty": {
            "markers": ["ლიბერთი", "Liberty", "liberty bank"],
            "date_cols": ["თარიღი", "Date", "Txn Date"],
            "desc_cols": ["აღწერა", "Description", "დეტალები", "Details"],
            "debit_cols": ["დებეტი", "Debit", "გასავალი"],
            "credit_cols": ["კრედიტი", "Credit", "შემოსავალი"],
            "balance_cols": ["ნაშთი", "Balance"],
            "date_format": "%Y-%m-%d",
        },
    }

    async def test_connection(self) -> Dict[str, Any]:
        """For CSV connector, 'connection' means we have valid CSV data."""
        return {
            "connected": True,
            "message": "Bank CSV connector ready — provide CSV data via fetch_data",
            "details": {"supported_banks": ["TBC Bank", "Bank of Georgia", "Liberty Bank"]},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["bank_statement"],
            "fields": {
                "bank_statement": ["date", "description", "debit", "credit", "balance", "currency", "reference"],
            },
            "supported_banks": list(self._BANK_SIGNATURES.keys()),
        }

    async def fetch_data(
        self, data_type: str, params: Dict[str, Any] = None
    ) -> ConnectorResult:
        """Parse a bank CSV.

        params:
          csv_content: str — raw CSV text
          csv_path: str — path to CSV file (alternative to csv_content)
          bank_format: str — "tbc", "bog", "liberty", "auto"
          encoding: str — file encoding (default: utf-8)
        """
        params = params or {}
        csv_content = params.get("csv_content", "")
        csv_path = params.get("csv_path", "")
        encoding = params.get("encoding", self.config.encoding or "utf-8")

        if not csv_content and csv_path:
            try:
                with open(csv_path, "r", encoding=encoding, errors="replace") as f:
                    csv_content = f.read()
            except Exception as e:
                return ConnectorResult(
                    success=False, connector_type=self.CONNECTOR_TYPE,
                    data_type="bank_statement", records=[],
                    error=f"Cannot read file: {e}",
                )

        if not csv_content:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="bank_statement", records=[],
                error="No CSV data provided — supply csv_content or csv_path",
            )

        # Detect bank format
        bank_format = params.get("bank_format", self.config.bank_format or "auto")
        if bank_format == "auto":
            bank_format = self._detect_bank_format(csv_content)

        if not bank_format or bank_format not in self._BANK_SIGNATURES:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="bank_statement", records=[],
                error=f"Could not detect bank format. Supported: {', '.join(self._BANK_SIGNATURES.keys())}",
            )

        # Parse CSV
        records, warnings = self._parse_bank_csv(csv_content, bank_format)

        # Build financials summary
        total_income = sum(r.get("credit", 0) for r in records)
        total_expense = sum(r.get("debit", 0) for r in records)
        financials = {
            "bank_income": total_income,
            "bank_expense": total_expense,
            "bank_net": total_income - total_expense,
            "transaction_count": len(records),
        }

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="bank_statement", records=records,
            record_count=len(records),
            currency=params.get("currency", "GEL"),
            source=f"bank_csv:{bank_format}",
            warnings=warnings,
            financials=financials,
            metadata={"bank_format": bank_format, "total_income": total_income, "total_expense": total_expense},
        )

    def _detect_bank_format(self, csv_content: str) -> str:
        """Auto-detect which Georgian bank produced the CSV."""
        content_lower = csv_content[:2000].lower()
        header_line = csv_content.split("\n")[0] if csv_content else ""

        for bank_id, sig in self._BANK_SIGNATURES.items():
            # Check for bank name markers
            for marker in sig["markers"]:
                if marker.lower() in content_lower:
                    logger.info("Bank CSV auto-detected: %s (marker: %s)", bank_id, marker)
                    return bank_id

        # Fallback: check column names in header
        for bank_id, sig in self._BANK_SIGNATURES.items():
            match_count = 0
            for col_group in [sig["date_cols"], sig["desc_cols"], sig["debit_cols"], sig["credit_cols"]]:
                for col_name in col_group:
                    if col_name.lower() in header_line.lower():
                        match_count += 1
                        break
            if match_count >= 3:
                logger.info("Bank CSV auto-detected by columns: %s", bank_id)
                return bank_id

        logger.warning("Could not auto-detect bank format")
        return ""

    def _parse_bank_csv(
        self, csv_content: str, bank_format: str
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse bank statement CSV into standardized records."""
        sig = self._BANK_SIGNATURES[bank_format]
        warnings: List[str] = []
        records: List[Dict[str, Any]] = []

        # Parse CSV
        reader = csv.reader(io.StringIO(csv_content))
        rows = list(reader)
        if len(rows) < 2:
            warnings.append("CSV has fewer than 2 rows")
            return records, warnings

        # Find header row (might not be the first row)
        header_idx = 0
        header = rows[0]
        for i, row in enumerate(rows[:5]):
            row_lower = [c.lower().strip() for c in row]
            match_count = sum(
                1 for col_group in [sig["date_cols"], sig["desc_cols"]]
                for col_name in col_group
                if col_name.lower() in row_lower
            )
            if match_count >= 2:
                header_idx = i
                header = row
                break

        header_lower = [h.lower().strip() for h in header]

        # Map column indices
        def find_col(candidates: List[str]) -> int:
            for c in candidates:
                cl = c.lower()
                for i, h in enumerate(header_lower):
                    if cl in h or h in cl:
                        return i
            return -1

        date_idx = find_col(sig["date_cols"])
        desc_idx = find_col(sig["desc_cols"])
        debit_idx = find_col(sig["debit_cols"])
        credit_idx = find_col(sig["credit_cols"])
        balance_idx = find_col(sig["balance_cols"])

        if date_idx < 0 or desc_idx < 0:
            warnings.append(f"Could not find required columns (date, description) in header: {header}")
            return records, warnings

        # Parse data rows
        date_format = sig["date_format"]
        for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if not row or all(c.strip() == "" for c in row):
                continue

            try:
                date_str = row[date_idx].strip() if date_idx < len(row) else ""
                description = row[desc_idx].strip() if desc_idx < len(row) else ""
                debit = self._parse_amount(row[debit_idx]) if debit_idx >= 0 and debit_idx < len(row) else 0.0
                credit = self._parse_amount(row[credit_idx]) if credit_idx >= 0 and credit_idx < len(row) else 0.0
                balance = self._parse_amount(row[balance_idx]) if balance_idx >= 0 and balance_idx < len(row) else 0.0

                # Parse date
                parsed_date = ""
                if date_str:
                    for fmt in [date_format, "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"]:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                            break
                        except ValueError:
                            continue

                if not parsed_date and not description:
                    continue  # Skip completely empty rows

                records.append({
                    "date": parsed_date or date_str,
                    "description": description,
                    "debit": abs(debit),
                    "credit": abs(credit),
                    "balance": balance,
                    "currency": "GEL",
                    "reference": "",
                    "row_number": row_num,
                })
            except Exception as e:
                warnings.append(f"Row {row_num}: parse error — {e}")

        logger.info("Bank CSV parsed: %s, %d records, %d warnings", bank_format, len(records), len(warnings))
        return records, warnings

    @staticmethod
    def _parse_amount(value: str) -> float:
        """Parse a monetary amount string, handling Georgian/European formats."""
        if not value:
            return 0.0
        # Remove currency symbols and whitespace
        cleaned = re.sub(r'[₾$€£\s]', '', value.strip())
        # Handle European comma decimal (1.234,56 -> 1234.56)
        if ',' in cleaned and '.' in cleaned:
            if cleaned.rindex(',') > cleaned.rindex('.'):
                cleaned = cleaned.replace('.', '').replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            # Single comma could be decimal separator
            parts = cleaned.split(',')
            if len(parts) == 2 and len(parts[1]) <= 2:
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# SAP Connector (OData v4)
# ═══════════════════════════════════════════════════════════════════════════════

class SAPConnector(BaseConnector):
    """Connect to SAP ERP via OData v4 APIs.

    SAP publishes financial data via OData services:
      - /sap/opu/odata/sap/API_JOURNALENTRYITEM_SRV
      - /sap/opu/odata4/sap/API_TRIALBALANCE/srvd_a2x/...
      - /sap/opu/odata/sap/API_GLACCOUNTINCHARTOFACCOUNTS_SRV

    This connector uses OData v4 format which is simpler than RFC.
    Authentication: Basic auth or OAuth2 bearer token.
    """

    CONNECTOR_TYPE = "sap"
    DISPLAY_NAME = "SAP ERP"
    DESCRIPTION = "Connect to SAP ERP via OData v4 API"

    def _base_url(self) -> str:
        host = self.config.host or self.config.server_url
        return host.rstrip("/")

    def _auth_headers(self) -> Dict[str, str]:
        """Build authentication headers for SAP OData."""
        import base64
        credentials = base64.b64encode(
            f"{self.config.username}:{self.config.password}".encode()
        ).decode()
        return {
            "Authorization": f"Basic {credentials}",
            "Accept": "application/json",
            "Content-Type": "application/json",
            "sap-client": self.config.client,
        }

    async def _request(
        self, path: str, params: Dict[str, str] = None
    ) -> Optional[Dict[str, Any]]:
        """Make an authenticated OData request to SAP."""
        url = f"{self._base_url()}{path}"
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    url,
                    params=params,
                    headers=self._auth_headers(),
                    timeout=aiohttp.ClientTimeout(total=30),
                    ssl=False,
                ) as resp:
                    if resp.status == 200:
                        return await resp.json()
                    else:
                        body = await resp.text()
                        logger.warning("SAP OData %s returned %d: %s", path, resp.status, body[:300])
                        return {"error": f"HTTP {resp.status}", "body": body[:300]}
        except aiohttp.ClientConnectorError as e:
            logger.debug("SAP connection failed: %s", e)
            return None
        except Exception as e:
            logger.debug("SAP request failed: %s", e)
            return None

    async def test_connection(self) -> Dict[str, Any]:
        """Test SAP OData connection."""
        # Try metadata endpoint
        data = await self._request(
            "/sap/opu/odata4/sap/API_JOURNALENTRYITEM_SRV/$metadata"
        )
        if data is not None and "error" not in (data or {}):
            return {
                "connected": True,
                "message": f"Connected to SAP at {self._base_url()}",
                "details": {"client": self.config.client, "system": self.config.system_number},
            }
        return {
            "connected": False,
            "message": f"Cannot connect to SAP at {self._base_url()}",
            "details": {"client": self.config.client},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["trial_balance", "gl_entries", "chart_of_accounts"],
            "fields": {
                "trial_balance": ["GLAccount", "GLAccountName", "DebitAmount", "CreditAmount",
                                  "DebitBalance", "CreditBalance", "CompanyCode", "FiscalYear"],
                "gl_entries": ["AccountingDocument", "FiscalYear", "GLAccount", "AmountInCompanyCodeCurrency",
                               "DebitCreditCode", "PostingDate", "DocumentItemText"],
                "chart_of_accounts": ["GLAccount", "GLAccountName", "ChartOfAccounts",
                                      "GLAccountType", "IsBalanceSheetAccount"],
            },
            "source": self._base_url(),
        }

    async def fetch_data(
        self, data_type: str, params: Dict[str, Any] = None
    ) -> ConnectorResult:
        params = params or {}

        if data_type == "trial_balance":
            return await self._fetch_trial_balance(params)
        elif data_type in ("gl_entries", "journal_entries"):
            return await self._fetch_gl_entries(params)
        elif data_type == "chart_of_accounts":
            return await self._fetch_chart_of_accounts(params)
        else:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=[],
                error=f"Unknown data_type: {data_type}",
            )

    async def _fetch_trial_balance(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch trial balance from SAP via OData."""
        company_code = params.get("company_code", "1000")
        fiscal_year = params.get("fiscal_year", str(datetime.now().year))
        period = params.get("period", "")

        odata_filter = f"CompanyCode eq '{company_code}' and FiscalYear eq '{fiscal_year}'"
        if period:
            odata_filter += f" and FiscalPeriod eq '{period}'"

        data = await self._request(
            "/sap/opu/odata4/sap/API_TRIALBALANCE/srvd_a2x/TrialBalance",
            {"$filter": odata_filter, "$format": "json", "$top": "10000"},
        )

        if not data or "error" in data:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="trial_balance", records=[],
                error=f"SAP trial balance fetch failed: {data.get('error', 'connection error') if data else 'no response'}",
                source=self._base_url(),
            )

        raw_records = data.get("value", data.get("d", {}).get("results", []))
        normalized = []
        for r in raw_records:
            normalized.append({
                "account_code": r.get("GLAccount", ""),
                "account_name": r.get("GLAccountName", r.get("GLAccountLongName", "")),
                "debit_opening": float(r.get("StartingBalanceAmtInCoCodeCrcy", 0) or 0) if r.get("DebitCreditCode") == "S" else 0,
                "credit_opening": float(r.get("StartingBalanceAmtInCoCodeCrcy", 0) or 0) if r.get("DebitCreditCode") == "H" else 0,
                "debit_turnover": float(r.get("DebitAmountInCoCodeCrcy", 0) or 0),
                "credit_turnover": float(r.get("CreditAmountInCoCodeCrcy", 0) or 0),
                "debit_closing": float(r.get("EndingBalanceAmtInCoCodeCrcy", 0) or 0) if r.get("DebitCreditCode") == "S" else 0,
                "credit_closing": float(r.get("EndingBalanceAmtInCoCodeCrcy", 0) or 0) if r.get("DebitCreditCode") == "H" else 0,
            })

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="trial_balance", records=normalized,
            record_count=len(normalized),
            period=f"{fiscal_year}-{period}" if period else fiscal_year,
            source=self._base_url(),
            metadata={"company_code": company_code, "fiscal_year": fiscal_year},
        )

    async def _fetch_gl_entries(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch GL journal entries from SAP."""
        company_code = params.get("company_code", "1000")
        date_from = params.get("date_from", "")
        date_to = params.get("date_to", "")

        odata_filter = f"CompanyCode eq '{company_code}'"
        if date_from:
            odata_filter += f" and PostingDate ge '{date_from}'"
        if date_to:
            odata_filter += f" and PostingDate le '{date_to}'"

        data = await self._request(
            "/sap/opu/odata4/sap/API_JOURNALENTRYITEM_SRV/A_JournalEntryItem",
            {"$filter": odata_filter, "$format": "json", "$top": str(params.get("limit", 5000))},
        )

        if not data or "error" in data:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="gl_entries", records=[],
                error="SAP GL entries fetch failed",
                source=self._base_url(),
            )

        raw_records = data.get("value", data.get("d", {}).get("results", []))
        normalized = []
        for r in raw_records:
            amount = float(r.get("AmountInCompanyCodeCurrency", 0) or 0)
            is_debit = r.get("DebitCreditCode") == "S"
            normalized.append({
                "date": r.get("PostingDate", ""),
                "document": r.get("AccountingDocument", ""),
                "account_code": r.get("GLAccount", ""),
                "account_name": r.get("GLAccountName", ""),
                "debit": abs(amount) if is_debit else 0,
                "credit": abs(amount) if not is_debit else 0,
                "description": r.get("DocumentItemText", ""),
                "currency": r.get("CompanyCodeCurrency", ""),
            })

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="gl_entries", records=normalized,
            record_count=len(normalized),
            period=f"{date_from} to {date_to}",
            source=self._base_url(),
            metadata={"company_code": company_code},
        )

    async def _fetch_chart_of_accounts(self, params: Dict[str, Any]) -> ConnectorResult:
        """Fetch chart of accounts from SAP."""
        chart = params.get("chart_of_accounts", "YCOA")

        data = await self._request(
            "/sap/opu/odata4/sap/API_GLACCOUNTINCHARTOFACCOUNTS_SRV/A_GLAccountInChartOfAccounts",
            {"$filter": f"ChartOfAccounts eq '{chart}'", "$format": "json", "$top": "10000"},
        )

        if not data or "error" in data:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="chart_of_accounts", records=[],
                error="SAP COA fetch failed",
                source=self._base_url(),
            )

        raw_records = data.get("value", data.get("d", {}).get("results", []))
        normalized = []
        for r in raw_records:
            normalized.append({
                "code": r.get("GLAccount", ""),
                "name": r.get("GLAccountName", ""),
                "type": "BS" if r.get("IsBalanceSheetAccount") else "PL",
                "chart": r.get("ChartOfAccounts", ""),
                "is_group": bool(r.get("GLAccountGroup")),
            })

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="chart_of_accounts", records=normalized,
            record_count=len(normalized),
            source=self._base_url(),
            metadata={"chart_of_accounts": chart},
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Connector (wraps existing upload logic)
# ═══════════════════════════════════════════════════════════════════════════════

class ExcelConnector(BaseConnector):
    """Wraps the existing smart-upload pipeline for Excel/CSV files.

    This connector reads a local file and parses it using the same
    logic as the /agents/smart-upload endpoint.
    """

    CONNECTOR_TYPE = "excel"
    DISPLAY_NAME = "Excel / CSV Upload"
    DESCRIPTION = "Import financial data from Excel (.xlsx, .xls) or CSV files"

    async def test_connection(self) -> Dict[str, Any]:
        """For Excel, check that the file path exists."""
        file_path = self.config.extra.get("file_path", "")
        if file_path and os.path.isfile(file_path):
            size = os.path.getsize(file_path)
            return {
                "connected": True,
                "message": f"File found: {os.path.basename(file_path)} ({size:,} bytes)",
                "details": {"file_path": file_path, "size_bytes": size},
            }
        return {
            "connected": False,
            "message": "No file specified or file not found",
            "details": {"file_path": file_path},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["auto_detect"],
            "fields": {
                "auto_detect": ["The schema depends on the file content — trial balance, P&L, or bank statement"],
            },
            "supported_formats": [".xlsx", ".xls", ".csv"],
        }

    async def fetch_data(
        self, data_type: str, params: Dict[str, Any] = None
    ) -> ConnectorResult:
        """Read and parse an Excel/CSV file.

        params:
          file_path: str — path to the file
          sheet_name: str — specific sheet (optional)
        """
        params = params or {}
        file_path = params.get("file_path", self.config.extra.get("file_path", ""))

        if not file_path or not os.path.isfile(file_path):
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="auto_detect", records=[],
                error=f"File not found: {file_path}",
            )

        try:
            # Use the document intelligence service for type detection
            from app.services.document_intelligence import doc_intelligence
            doc_analysis = doc_intelligence.analyze(file_path)

            # Use the TB parser if it looks like a trial balance
            if doc_analysis.doc_type == "trial_balance" and doc_analysis.confidence >= 0.5:
                from app.services.tb_parser import tb_parser
                tb_result = tb_parser.detect_and_parse(file_path)
                if tb_result and tb_result.detected and tb_result.postable_count > 0:
                    records = []
                    for acc in tb_result.accounts:
                        records.append({
                            "account_code": acc.get("code", ""),
                            "account_name": acc.get("name", ""),
                            "debit_opening": acc.get("debit_opening", 0),
                            "credit_opening": acc.get("credit_opening", 0),
                            "debit_turnover": acc.get("debit_turnover", 0),
                            "credit_turnover": acc.get("credit_turnover", 0),
                            "debit_closing": acc.get("debit_closing", 0),
                            "credit_closing": acc.get("credit_closing", 0),
                        })
                    return ConnectorResult(
                        success=True, connector_type=self.CONNECTOR_TYPE,
                        data_type="trial_balance", records=records,
                        record_count=len(records),
                        period=tb_result.period or "",
                        source=file_path,
                        metadata={
                            "company": tb_result.company,
                            "balanced": tb_result.is_balanced,
                            "doc_type": "trial_balance",
                        },
                    )

            # Fallback: read raw data with openpyxl or csv
            ext = os.path.splitext(file_path)[1].lower()
            if ext in (".xlsx", ".xls"):
                records = self._read_excel(file_path, params.get("sheet_name"))
            elif ext == ".csv":
                records = self._read_csv(file_path)
            else:
                return ConnectorResult(
                    success=False, connector_type=self.CONNECTOR_TYPE,
                    data_type="auto_detect", records=[],
                    error=f"Unsupported file format: {ext}",
                )

            return ConnectorResult(
                success=True, connector_type=self.CONNECTOR_TYPE,
                data_type=doc_analysis.doc_type or "unknown",
                records=records,
                record_count=len(records),
                source=file_path,
                metadata={
                    "doc_type": doc_analysis.doc_type,
                    "confidence": doc_analysis.confidence,
                    "company": doc_analysis.detected_company,
                },
            )

        except Exception as e:
            logger.error("Excel connector error: %s", e)
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="auto_detect", records=[],
                error=str(e),
            )

    @staticmethod
    def _read_excel(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Read Excel file into list of dicts."""
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return []

        headers = [str(h or f"col_{i}").strip() for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            if all(c is None for c in row):
                continue
            record = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                record[key] = val
            records.append(record)
        return records

    @staticmethod
    def _read_csv(file_path: str) -> List[Dict[str, Any]]:
        """Read CSV file into list of dicts."""
        records = []
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(dict(row))
        return records


# ═══════════════════════════════════════════════════════════════════════════════
# NBG (National Bank of Georgia) Connector — Live Exchange Rates
# ═══════════════════════════════════════════════════════════════════════════════

class NBGConnector(BaseConnector):
    """Fetch live exchange rates from the National Bank of Georgia API.

    Uses the official NBG monetary policy API:
      https://nbg.gov.ge/gw/api/ct/monetarypolicy/currencies/

    Returns rates for major currencies: USD, EUR, GBP, TRY, RUB, AZN.
    Rates are cached for 1 hour to avoid excessive API calls.
    """

    CONNECTOR_TYPE = "nbg"
    DISPLAY_NAME = "National Bank of Georgia"
    DESCRIPTION = "Live exchange rates from the National Bank of Georgia (NBG)"

    _NBG_BASE_URL = "https://nbg.gov.ge/gw/api/ct/monetarypolicy/currencies"
    _DEFAULT_CURRENCIES = "USD,EUR,GBP,TRY,RUB,AZN"
    _CACHE_TTL = 3600  # 1 hour

    # Class-level cache
    _rate_cache: Optional[Dict[str, Any]] = None
    _cache_ts: float = 0.0

    async def test_connection(self) -> Dict[str, Any]:
        """Test connectivity to the NBG API."""
        try:
            result = await self._fetch_rates(currencies="USD")
            if result and "error" not in result:
                return {
                    "connected": True,
                    "message": "Connected to NBG API (live exchange rates)",
                    "details": {"api": self._NBG_BASE_URL, "sample_rate": result},
                }
        except Exception as e:
            logger.debug("NBG connection test failed: %s", e)
        return {
            "connected": False,
            "message": "Cannot connect to NBG API — check internet connectivity",
            "details": {"api": self._NBG_BASE_URL},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["exchange_rates"],
            "fields": {
                "exchange_rates": ["currency", "rate", "diff", "date", "validFromDate"],
            },
            "supported_currencies": self._DEFAULT_CURRENCIES.split(","),
            "cache_ttl_seconds": self._CACHE_TTL,
        }

    async def fetch_data(
        self, data_type: str = "exchange_rates", params: Dict[str, Any] = None
    ) -> ConnectorResult:
        params = params or {}
        currencies = params.get("currencies", self._DEFAULT_CURRENCIES)
        result = await self._fetch_rates(currencies=currencies)

        if result is None or "error" in (result or {}):
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="exchange_rates", records=[],
                error=result.get("error", "NBG API unreachable") if result else "NBG API unreachable",
            )

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="exchange_rates", records=result.get("rates", []),
            record_count=len(result.get("rates", [])),
            currency="GEL",
            source="nbg.gov.ge",
            metadata=result.get("metadata", {}),
        )

    async def _fetch_rates(self, currencies: str = None) -> Optional[Dict[str, Any]]:
        """Fetch exchange rates from NBG API with caching.

        NBG API returns:
        [
          {
            "date": "2025-01-15T...",
            "currencies": [
              {"code": "USD", "quantity": 1, "rateFormated": "2.7250", "diffFormated": "-0.0050", ...},
              ...
            ]
          }
        ]
        """
        currencies = currencies or self._DEFAULT_CURRENCIES

        # Check cache
        cache_key = currencies
        now = time.time()
        if (NBGConnector._rate_cache is not None
                and NBGConnector._cache_ts > now - self._CACHE_TTL
                and NBGConnector._rate_cache.get("_cache_key") == cache_key):
            logger.debug("NBG rates served from cache")
            return NBGConnector._rate_cache

        url = f"{self._NBG_BASE_URL}/?currencies={currencies}"

        try:
            if HAS_HTTPX:
                async with httpx.AsyncClient(timeout=15.0, verify=True) as client:
                    resp = await client.get(url)
                    resp.raise_for_status()
                    raw = resp.json()
            else:
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                        if resp.status != 200:
                            return {"error": f"NBG API returned HTTP {resp.status}"}
                        raw = await resp.json()

            # Parse the NBG response structure
            if not isinstance(raw, list) or len(raw) == 0:
                return {"error": "Unexpected NBG response format", "raw_type": type(raw).__name__}

            first_entry = raw[0]
            date_str = first_entry.get("date", "")
            currencies_data = first_entry.get("currencies", [])

            rates = []
            rate_map = {}  # For currency_engine compatibility
            for cur in currencies_data:
                code = cur.get("code", "")
                quantity = float(cur.get("quantity", 1) or 1)
                # rate is per `quantity` units — normalize to per-1-unit
                rate_raw = float(cur.get("rate", 0) or cur.get("rateFormated", "0").replace(",", ""))
                rate_per_unit = rate_raw / quantity if quantity > 0 else rate_raw
                diff_raw = cur.get("diff", cur.get("diffFormated", "0"))
                if isinstance(diff_raw, str):
                    diff_raw = diff_raw.replace(",", "")
                diff = float(diff_raw or 0) / quantity if quantity > 0 else 0

                rates.append({
                    "currency": code,
                    "rate": round(rate_per_unit, 6),
                    "diff": round(diff, 6),
                    "quantity": int(quantity),
                    "raw_rate": rate_raw,
                    "date": date_str,
                    "name": cur.get("name", ""),
                    "nameEn": cur.get("nameEn", ""),
                })
                rate_map[code] = round(rate_per_unit, 6)

            result = {
                "rates": rates,
                "rate_map": rate_map,
                "date": date_str,
                "metadata": {
                    "source": "nbg.gov.ge",
                    "fetched_at": datetime.now().isoformat(),
                    "cache_ttl": self._CACHE_TTL,
                    "currencies_requested": currencies,
                },
                "_cache_key": currencies,
            }

            # Update cache
            NBGConnector._rate_cache = result
            NBGConnector._cache_ts = now
            logger.info("NBG rates fetched: %d currencies, date=%s", len(rates), date_str[:10])
            return result

        except Exception as e:
            err_type = type(e).__name__
            if "HTTPStatusError" in err_type:
                logger.warning("NBG API HTTP error: %s", e)
                return {"error": f"NBG API HTTP error: {e}"}
            elif "RequestError" in err_type or "ConnectError" in err_type:
                logger.warning("NBG API request error: %s", e)
                return {"error": f"NBG API connection error: {e}"}
            logger.warning("NBG rate fetch failed: %s", e)
            return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
# Georgian Revenue Service (rs.ge) Connector
# ═══════════════════════════════════════════════════════════════════════════════

class RevenueServiceConnector(BaseConnector):
    """Look up Georgian companies on the Revenue Service (rs.ge).

    Public API for company lookup by tax identification number (TIN):
      https://www.rs.ge/tax-free-api

    Note: Full financial data requires authentication. This connector
    implements the public lookup endpoint only.
    """

    CONNECTOR_TYPE = "rs_ge"
    DISPLAY_NAME = "Georgian Revenue Service (rs.ge)"
    DESCRIPTION = "Look up Georgian companies by tax ID on rs.ge"

    _RS_BASE_URL = "https://www.rs.ge"

    async def test_connection(self) -> Dict[str, Any]:
        """Test connectivity to rs.ge."""
        try:
            if HAS_HTTPX:
                async with httpx.AsyncClient(timeout=10.0, verify=True) as client:
                    resp = await client.get(f"{self._RS_BASE_URL}/", follow_redirects=True)
                    reachable = resp.status_code < 500
            else:
                async with aiohttp.ClientSession() as session:
                    async with session.get(self._RS_BASE_URL, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                        reachable = resp.status < 500
            return {
                "connected": reachable,
                "message": "rs.ge reachable" if reachable else "rs.ge unreachable",
                "details": {"api": self._RS_BASE_URL, "note": "Public company lookup only"},
            }
        except Exception as e:
            return {
                "connected": False,
                "message": f"Cannot reach rs.ge: {e}",
                "details": {"api": self._RS_BASE_URL},
            }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["company_lookup"],
            "fields": {
                "company_lookup": ["tax_id", "name", "name_en", "status",
                                   "registration_date", "legal_form", "address"],
            },
            "note": "Financial data requires rs.ge authentication (not implemented)",
        }

    async def fetch_data(
        self, data_type: str = "company_lookup", params: Dict[str, Any] = None
    ) -> ConnectorResult:
        params = params or {}
        tax_id = params.get("tax_id", "")
        if not tax_id:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="company_lookup", records=[],
                error="tax_id parameter is required",
            )
        return await self._lookup_company(tax_id)

    async def _lookup_company(self, tax_id: str) -> ConnectorResult:
        """Look up a company by tax identification number.

        Tries multiple public endpoints:
          1. /tax-free-api/taxpayer/{tax_id}
          2. /api/tax-payer?tin={tax_id}
        """
        endpoints = [
            f"/tax-free-api/taxpayer/{tax_id}",
            f"/api/tax-payer?tin={tax_id}",
        ]

        for endpoint in endpoints:
            url = f"{self._RS_BASE_URL}{endpoint}"
            try:
                if HAS_HTTPX:
                    async with httpx.AsyncClient(timeout=15.0, verify=True) as client:
                        resp = await client.get(url, follow_redirects=True)
                        if resp.status_code == 200:
                            data = resp.json()
                            return self._parse_company_result(data, tax_id)
                else:
                    async with aiohttp.ClientSession() as session:
                        async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                            if resp.status == 200:
                                data = await resp.json()
                                return self._parse_company_result(data, tax_id)
            except Exception as e:
                logger.debug("rs.ge endpoint %s failed: %s", endpoint, e)
                continue

        # If API endpoints fail, return structured error with guidance
        return ConnectorResult(
            success=False, connector_type=self.CONNECTOR_TYPE,
            data_type="company_lookup", records=[],
            error=f"Could not look up tax ID {tax_id} — rs.ge public API may require a different endpoint or format",
            metadata={"tax_id": tax_id, "endpoints_tried": endpoints},
        )

    def _parse_company_result(self, data: Any, tax_id: str) -> ConnectorResult:
        """Parse rs.ge API response into standardized format."""
        if isinstance(data, dict):
            records = [data]
        elif isinstance(data, list):
            records = data
        else:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="company_lookup", records=[],
                error=f"Unexpected response type: {type(data).__name__}",
            )

        normalized = []
        for r in records:
            normalized.append({
                "tax_id": r.get("tin", r.get("taxPayerId", r.get("identificationNumber", tax_id))),
                "name": r.get("name", r.get("organizationName", r.get("nameKa", ""))),
                "name_en": r.get("nameEn", r.get("name_en", "")),
                "status": r.get("status", r.get("registrationStatus", "")),
                "registration_date": r.get("registrationDate", r.get("regDate", "")),
                "legal_form": r.get("legalForm", r.get("organizationType", "")),
                "address": r.get("address", r.get("legalAddress", "")),
            })

        return ConnectorResult(
            success=True, connector_type=self.CONNECTOR_TYPE,
            data_type="company_lookup", records=normalized,
            record_count=len(normalized),
            source="rs.ge",
            metadata={"tax_id": tax_id},
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Generic REST API Connector
# ═══════════════════════════════════════════════════════════════════════════════

class GenericRESTConnector(BaseConnector):
    """Connect to any REST API with configurable auth.

    Supports:
      - No auth
      - Basic auth (username:password)
      - Bearer token
      - API key (header or query param)

    Auto-detects financial data structure in JSON responses.
    """

    CONNECTOR_TYPE = "rest"
    DISPLAY_NAME = "Generic REST API"
    DESCRIPTION = "Connect to any REST API with configurable authentication"

    def _get_auth_headers(self) -> Dict[str, str]:
        """Build authentication headers based on config."""
        auth_type = self.config.extra.get("auth_type", "none")
        headers = dict(self.config.extra.get("headers", {}))
        headers.setdefault("Accept", "application/json")

        if auth_type == "basic":
            import base64
            credentials = base64.b64encode(
                f"{self.config.username}:{self.config.password}".encode()
            ).decode()
            headers["Authorization"] = f"Basic {credentials}"
        elif auth_type == "bearer":
            token = self.config.extra.get("token", self.config.password)
            headers["Authorization"] = f"Bearer {token}"
        elif auth_type == "api_key":
            key_name = self.config.extra.get("api_key_header", "X-API-Key")
            key_value = self.config.extra.get("api_key", self.config.password)
            if self.config.extra.get("api_key_in") == "query":
                pass  # Handled in params
            else:
                headers[key_name] = key_value

        return headers

    def _get_auth_params(self) -> Dict[str, str]:
        """Build query params for API key auth."""
        auth_type = self.config.extra.get("auth_type", "none")
        if auth_type == "api_key" and self.config.extra.get("api_key_in") == "query":
            key_name = self.config.extra.get("api_key_param", "api_key")
            key_value = self.config.extra.get("api_key", self.config.password)
            return {key_name: key_value}
        return {}

    async def test_connection(self) -> Dict[str, Any]:
        """Test the REST API connection."""
        base_url = self.config.server_url.rstrip("/")
        health_paths = self.config.extra.get("health_paths", ["/", "/health", "/api/health"])

        for path in health_paths:
            url = f"{base_url}{path}"
            try:
                if HAS_HTTPX:
                    async with httpx.AsyncClient(timeout=10.0, verify=not self.config.extra.get("no_ssl_verify")) as client:
                        resp = await client.get(url, headers=self._get_auth_headers(),
                                                params=self._get_auth_params(), follow_redirects=True)
                        if resp.status_code < 400:
                            return {
                                "connected": True,
                                "message": f"Connected to {base_url}",
                                "details": {"endpoint": path, "status": resp.status_code},
                            }
                else:
                    async with aiohttp.ClientSession() as session:
                        async with session.get(url, headers=self._get_auth_headers(),
                                               params=self._get_auth_params(),
                                               timeout=aiohttp.ClientTimeout(total=10)) as resp:
                            if resp.status < 400:
                                return {
                                    "connected": True,
                                    "message": f"Connected to {base_url}",
                                    "details": {"endpoint": path, "status": resp.status},
                                }
            except Exception:
                continue

        return {
            "connected": False,
            "message": f"Cannot connect to {base_url}",
            "details": {"tried": health_paths},
        }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["json_data"],
            "auth_types": ["none", "basic", "bearer", "api_key"],
            "source": self.config.server_url,
            "note": "Schema auto-detected from response",
        }

    async def fetch_data(
        self, data_type: str = "json_data", params: Dict[str, Any] = None
    ) -> ConnectorResult:
        """Fetch data from the REST API.

        params:
          endpoint: str — API path (appended to base_url)
          method: str — HTTP method (default: GET)
          query_params: dict — query string parameters
          body: dict — request body (for POST/PUT)
          data_path: str — JSON path to data array (e.g. "data.items")
        """
        params = params or {}
        base_url = self.config.server_url.rstrip("/")
        endpoint = params.get("endpoint", "/")
        method = params.get("method", "GET").upper()
        query_params = {**self._get_auth_params(), **params.get("query_params", {})}
        body = params.get("body")
        data_path = params.get("data_path", "")

        url = f"{base_url}{endpoint}"

        try:
            if HAS_HTTPX:
                async with httpx.AsyncClient(
                    timeout=30.0,
                    verify=not self.config.extra.get("no_ssl_verify"),
                ) as client:
                    if method == "GET":
                        resp = await client.get(url, headers=self._get_auth_headers(),
                                                params=query_params, follow_redirects=True)
                    elif method == "POST":
                        resp = await client.post(url, headers=self._get_auth_headers(),
                                                 params=query_params, json=body, follow_redirects=True)
                    elif method == "PUT":
                        resp = await client.put(url, headers=self._get_auth_headers(),
                                                params=query_params, json=body, follow_redirects=True)
                    else:
                        return ConnectorResult(
                            success=False, connector_type=self.CONNECTOR_TYPE,
                            data_type=data_type, records=[],
                            error=f"Unsupported method: {method}",
                        )

                    resp.raise_for_status()
                    raw = resp.json()
            else:
                async with aiohttp.ClientSession() as session:
                    kwargs = {
                        "headers": self._get_auth_headers(),
                        "params": query_params,
                        "timeout": aiohttp.ClientTimeout(total=30),
                    }
                    if body and method in ("POST", "PUT"):
                        kwargs["json"] = body

                    async with session.request(method, url, **kwargs) as resp:
                        if resp.status >= 400:
                            return ConnectorResult(
                                success=False, connector_type=self.CONNECTOR_TYPE,
                                data_type=data_type, records=[],
                                error=f"HTTP {resp.status}: {(await resp.text())[:200]}",
                            )
                        raw = await resp.json()

            # Extract data array from response using data_path
            records = self._extract_data(raw, data_path)

            # Auto-detect financial structure
            financial_hints = self._detect_financial_structure(records)

            return ConnectorResult(
                success=True, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=records,
                record_count=len(records),
                source=url,
                metadata={"financial_hints": financial_hints, "method": method, "endpoint": endpoint},
            )

        except Exception as e:
            logger.warning("REST fetch failed (%s): %s", url, e)
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=[],
                error=str(e),
            )

    def _extract_data(self, raw: Any, data_path: str = "") -> List[Dict[str, Any]]:
        """Extract the data array from a JSON response using a dot-notation path."""
        data = raw
        if data_path:
            for key in data_path.split("."):
                if isinstance(data, dict):
                    data = data.get(key, data)
                elif isinstance(data, list) and key.isdigit():
                    idx = int(key)
                    data = data[idx] if idx < len(data) else data
                else:
                    break

        if isinstance(data, list):
            return [r for r in data if isinstance(r, dict)]
        elif isinstance(data, dict):
            # Look for common array keys
            for array_key in ["data", "results", "items", "records", "rows", "value", "entries"]:
                if array_key in data and isinstance(data[array_key], list):
                    return [r for r in data[array_key] if isinstance(r, dict)]
            return [data]
        return []

    @staticmethod
    def _detect_financial_structure(records: List[Dict]) -> Dict[str, Any]:
        """Auto-detect if the records contain financial data."""
        if not records:
            return {"detected": False}

        sample = records[0]
        keys_lower = {k.lower() for k in sample.keys()}

        hints = {"detected": False, "fields": {}}
        financial_patterns = {
            "amount": ["amount", "value", "total", "sum", "balance"],
            "account": ["account", "code", "gl_account", "account_code"],
            "date": ["date", "period", "posting_date", "transaction_date"],
            "description": ["description", "text", "memo", "narrative", "name"],
            "debit": ["debit", "dr", "debit_amount"],
            "credit": ["credit", "cr", "credit_amount"],
            "currency": ["currency", "curr", "ccy"],
        }

        for field_type, patterns in financial_patterns.items():
            for pattern in patterns:
                matching = [k for k in keys_lower if pattern in k]
                if matching:
                    hints["fields"][field_type] = list(matching)
                    hints["detected"] = True
                    break

        return hints


# ═══════════════════════════════════════════════════════════════════════════════
# SFTP Connector — Automated file ingestion from 1C exports
# ═══════════════════════════════════════════════════════════════════════════════

class SFTPConnector(BaseConnector):
    """Connect to SFTP servers for automated file ingestion.

    Designed for nightly pulls of 1C file exports (Excel/CSV).
    Requires paramiko (optional dependency).

    Config:
      host, port (default 22), username, password or key_path, remote_path
    """

    CONNECTOR_TYPE = "sftp"
    DISPLAY_NAME = "SFTP File Ingestion"
    DESCRIPTION = "Automated file ingestion from SFTP servers (1C exports, bank files)"

    async def test_connection(self) -> Dict[str, Any]:
        """Test SFTP connection."""
        if not HAS_PARAMIKO:
            return {
                "connected": False,
                "message": "paramiko not installed — run: pip install paramiko",
                "details": {"dependency": "paramiko"},
            }

        try:
            transport = paramiko.Transport((
                self.config.host or self.config.server_url,
                int(self.config.extra.get("port", 22)),
            ))
            key_path = self.config.extra.get("key_path", "")
            if key_path and os.path.exists(key_path):
                pkey = paramiko.RSAKey.from_private_key_file(key_path)
                transport.connect(username=self.config.username, pkey=pkey)
            else:
                transport.connect(username=self.config.username, password=self.config.password)

            sftp = paramiko.SFTPClient.from_transport(transport)
            remote_path = self.config.extra.get("remote_path", "/")
            listing = sftp.listdir(remote_path)
            sftp.close()
            transport.close()

            return {
                "connected": True,
                "message": f"Connected to SFTP {self.config.host}",
                "details": {
                    "host": self.config.host,
                    "remote_path": remote_path,
                    "files_found": len(listing),
                    "sample_files": listing[:10],
                },
            }
        except Exception as e:
            return {
                "connected": False,
                "message": f"SFTP connection failed: {e}",
                "details": {"host": self.config.host},
            }

    async def get_schema(self) -> Dict[str, Any]:
        return {
            "data_types": ["file_list", "file_download"],
            "fields": {
                "file_list": ["filename", "size", "modified", "path"],
                "file_download": ["filename", "content_bytes", "size"],
            },
            "note": "Requires paramiko for SFTP" + (" (installed)" if HAS_PARAMIKO else " (NOT installed)"),
        }

    async def fetch_data(
        self, data_type: str = "file_list", params: Dict[str, Any] = None
    ) -> ConnectorResult:
        """List or download files from SFTP.

        params:
          pattern: str — glob pattern for filtering files (e.g. "*.xlsx")
          remote_path: str — override remote directory
          download_file: str — specific file to download (for file_download type)
        """
        if not HAS_PARAMIKO:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=[],
                error="paramiko not installed — run: pip install paramiko",
            )

        params = params or {}

        if data_type == "file_list":
            return await self._list_files(params)
        elif data_type == "file_download":
            return await self._download_file(params)
        else:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type=data_type, records=[],
                error=f"Unknown data_type: {data_type}. Use 'file_list' or 'file_download'",
            )

    async def _list_files(self, params: Dict[str, Any]) -> ConnectorResult:
        """List files on the SFTP server matching a pattern."""
        import fnmatch

        pattern = params.get("pattern", "*")
        remote_path = params.get("remote_path", self.config.extra.get("remote_path", "/"))

        try:
            transport = self._connect_transport()
            sftp = paramiko.SFTPClient.from_transport(transport)

            all_files = sftp.listdir_attr(remote_path)
            records = []
            for f in all_files:
                if fnmatch.fnmatch(f.filename, pattern):
                    records.append({
                        "filename": f.filename,
                        "size": f.st_size,
                        "modified": datetime.fromtimestamp(f.st_mtime).isoformat() if f.st_mtime else "",
                        "path": f"{remote_path.rstrip('/')}/{f.filename}",
                        "is_dir": f.longname.startswith("d") if f.longname else False,
                    })

            sftp.close()
            transport.close()

            return ConnectorResult(
                success=True, connector_type=self.CONNECTOR_TYPE,
                data_type="file_list", records=records,
                record_count=len(records),
                source=f"sftp://{self.config.host}{remote_path}",
                metadata={"pattern": pattern, "total_files": len(all_files)},
            )
        except Exception as e:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="file_list", records=[],
                error=f"SFTP list failed: {e}",
            )

    async def _download_file(self, params: Dict[str, Any]) -> ConnectorResult:
        """Download a file from the SFTP server."""
        remote_file = params.get("download_file", params.get("remote_path", ""))
        if not remote_file:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="file_download", records=[],
                error="download_file parameter is required",
            )

        try:
            transport = self._connect_transport()
            sftp = paramiko.SFTPClient.from_transport(transport)

            import io as _io
            buf = _io.BytesIO()
            sftp.getfo(remote_file, buf)
            buf.seek(0)
            file_bytes = buf.read()
            file_size = len(file_bytes)

            sftp.close()
            transport.close()

            filename = os.path.basename(remote_file)
            return ConnectorResult(
                success=True, connector_type=self.CONNECTOR_TYPE,
                data_type="file_download",
                records=[{
                    "filename": filename,
                    "size": file_size,
                    "path": remote_file,
                    "content_preview": f"<{file_size} bytes>",
                }],
                record_count=1,
                source=f"sftp://{self.config.host}{remote_file}",
                metadata={
                    "filename": filename,
                    "size_bytes": file_size,
                    "content_available": True,
                },
            )
        except Exception as e:
            return ConnectorResult(
                success=False, connector_type=self.CONNECTOR_TYPE,
                data_type="file_download", records=[],
                error=f"SFTP download failed: {e}",
            )

    def _connect_transport(self) -> "paramiko.Transport":
        """Create and authenticate an SFTP transport."""
        transport = paramiko.Transport((
            self.config.host or self.config.server_url,
            int(self.config.extra.get("port", 22)),
        ))
        key_path = self.config.extra.get("key_path", "")
        if key_path and os.path.exists(key_path):
            pkey = paramiko.RSAKey.from_private_key_file(key_path)
            transport.connect(username=self.config.username, pkey=pkey)
        else:
            transport.connect(username=self.config.username, password=self.config.password)
        return transport


# ═══════════════════════════════════════════════════════════════════════════════
# Connector Registry
# ═══════════════════════════════════════════════════════════════════════════════

CONNECTOR_CLASSES = {
    "1c": OneCConnector,
    "bank_csv": BankCSVConnector,
    "sap": SAPConnector,
    "excel": ExcelConnector,
    "nbg": NBGConnector,
    "rs_ge": RevenueServiceConnector,
    "rest": GenericRESTConnector,
    "sftp": SFTPConnector,
}


class ConnectorRegistry:
    """Registry for managing active connector instances."""

    def __init__(self):
        self._connectors: Dict[str, BaseConnector] = {}

    def create_connector(self, config: ConnectorConfig) -> BaseConnector:
        """Create a connector from configuration."""
        cls = CONNECTOR_CLASSES.get(config.connector_type)
        if not cls:
            raise ValueError(
                f"Unknown connector type: {config.connector_type}. "
                f"Available: {', '.join(CONNECTOR_CLASSES.keys())}"
            )
        connector = cls(config)
        key = f"{config.connector_type}_{config.name or 'default'}"
        self._connectors[key] = connector
        logger.info("Connector created: %s (%s)", key, cls.DISPLAY_NAME)
        return connector

    def get_connector(self, key: str) -> Optional[BaseConnector]:
        return self._connectors.get(key)

    def list_available(self) -> List[Dict[str, Any]]:
        """List all available connector types and active instances."""
        types = []
        for type_id, cls in CONNECTOR_CLASSES.items():
            types.append({
                "type": type_id,
                "display_name": cls.DISPLAY_NAME,
                "description": cls.DESCRIPTION,
            })

        active = []
        for key, conn in self._connectors.items():
            active.append({
                "key": key,
                "type": conn.CONNECTOR_TYPE,
                "display_name": conn.DISPLAY_NAME,
                "config": conn.config.to_safe_dict(),
            })

        return {
            "connector_types": types,
            "active_connectors": active,
        }


# Module-level singleton
connector_registry = ConnectorRegistry()
