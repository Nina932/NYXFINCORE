"""
Phase S-1: End-to-End Pipeline Test
======================================
Simulates the full user journey WITHOUT a running server.
Direct Python calls to all services.

Usage: python e2e_test.py
"""

import os
import sys
import tempfile
import time

# Colors
G = "\033[92m"
R = "\033[91m"
Y = "\033[93m"
C = "\033[96m"
B = "\033[1m"
X = "\033[0m"

results = []

def step(name, condition, detail=""):
    results.append((name, condition))
    icon = f"{G}PASS{X}" if condition else f"{R}FAIL{X}"
    det = f"  ({detail})" if detail else ""
    print(f"  [{icon}] {name}{det}")


print(f"\n{B}{C}{'=' * 65}{X}")
print(f"{B}{C}  FINAI E2E PIPELINE TEST{X}")
print(f"{B}{C}{'=' * 65}{X}\n")

t0 = time.time()

# ── Step 1: DataStore — Create Company ────────────────────────────
print(f"{B}Step 1: DataStore — Create Company{X}")
try:
    from app.services.data_store import DataStore
    _db = os.path.join(tempfile.gettempdir(), "finai_e2e_test.db")
    store = DataStore(_db)
    store.reset()
    company_id = store.create_company("E2E Test Corp", "fuel_distribution", "GEL")
    step("Company created", company_id > 0, f"id={company_id}")
    co = store.get_company(company_id)
    step("Company retrieved", co is not None and co["name"] == "E2E Test Corp")
except Exception as e:
    step("DataStore", False, str(e))

# ── Step 2: Create Test Excel ─────────────────────────────────────
print(f"\n{B}Step 2: Create Test Excel{X}")
try:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.append(["Revenue", "COGS", "G&A Expenses", "Net Profit", "Depreciation", "Period"])
    ws.append([50000000, 42000000, 3000000, 3500000, 500000, "Jan 2025"])
    _xlsx_path = os.path.join(tempfile.gettempdir(), "finai_e2e_upload.xlsx")
    wb.save(_xlsx_path)
    wb.close()
    step("Excel created", os.path.exists(_xlsx_path))
except Exception as e:
    step("Excel creation", False, str(e))

# ── Step 3: Parse with SmartExcelParser ───────────────────────────
print(f"\n{B}Step 3: SmartExcelParser{X}")
try:
    from app.services.smart_excel_parser import smart_parser
    parse_result = smart_parser.parse_file(_xlsx_path)
    step("File parsed", parse_result is not None)
    step("Revenue mapped", parse_result.normalized_financials.get("revenue") == 50000000)
    step("COGS mapped", parse_result.normalized_financials.get("cogs") == 42000000)
    step("Confidence > 80%", parse_result.confidence_score >= 80,
         f"confidence={parse_result.confidence_score}%")
    financials = parse_result.normalized_financials
except Exception as e:
    step("SmartExcelParser", False, str(e))
    financials = {}

# ── Step 4: Validate with DataValidator ───────────────────────────
print(f"\n{B}Step 4: DataValidator{X}")
try:
    from app.services.data_validator import data_validator
    vr = data_validator.validate(financials)
    step("Validation passed (no errors)", vr.valid)
    step("Auto-corrections or valid", len(vr.auto_corrections) >= 0,
         f"{len(vr.auto_corrections)} corrections")
    step("Corrected data has fields", len(vr.corrected_data) >= 5,
         f"{len(vr.corrected_data)} fields")
    financials = vr.corrected_data  # use corrected data going forward
except Exception as e:
    step("DataValidator", False, str(e))

# ── Step 5: Save to DataStore ─────────────────────────────────────
print(f"\n{B}Step 5: Persist to DataStore{X}")
try:
    period_id = store.save_financials(company_id, "Jan 2025", financials, "e2e_upload.xlsx")
    step("Financials saved", period_id > 0)
    retrieved = store.get_financials(company_id, "Jan 2025")
    step("Financials retrieved", retrieved.get("revenue") == 50000000)
    step("All fields stored", len(retrieved) >= 5, f"{len(retrieved)} fields")
except Exception as e:
    step("DataStore save", False, str(e))

# ── Step 6: Run Orchestrator ──────────────────────────────────────
print(f"\n{B}Step 6: Financial Orchestrator (7-Stage Pipeline){X}")
try:
    from app.services.orchestrator import FinancialOrchestrator
    orch = FinancialOrchestrator()
    orch_result = orch.run(
        current_financials=financials,
        previous_financials={"revenue": 45000000, "cogs": 38000000, "gross_margin_pct": 15.6},
        balance_sheet={
            "total_current_assets": 12000000, "total_current_liabilities": 7000000,
            "total_assets": 30000000, "total_liabilities": 18000000,
            "total_equity": 12000000, "cash": 5000000,
        },
        monte_carlo_iterations=50,
    )
    step("Orchestrator ran", orch_result is not None)
    step("7 stages completed", len(orch_result.stages_completed) == 7,
         f"stages={orch_result.stages_completed}")
    step("0 stages failed", len(orch_result.stages_failed) == 0)
    step("Health score > 0", orch_result.health_score > 0,
         f"score={orch_result.health_score:.0f} ({orch_result.health_grade})")
    step("Strategy assigned", orch_result.strategy_name is not None and len(orch_result.strategy_name) > 0,
         orch_result.strategy_name)
    step("Execution < 5s", orch_result.execution_time_ms < 5000,
         f"{orch_result.execution_time_ms:.0f}ms")

    # Save to DataStore
    rd = orch_result.to_dict()
    run_id = store.save_orchestrator_result(company_id, rd)
    step("Result persisted", run_id > 0)
except Exception as e:
    step("Orchestrator", False, str(e))
    rd = {}

# ── Step 7: Generate PDF Report ───────────────────────────────────
print(f"\n{B}Step 7: Professional PDF Report{X}")
try:
    from app.services.professional_pdf import professional_pdf
    pdf_bytes = professional_pdf.generate(rd, "E2E Test Corp", "Jan 2025")
    step("PDF generated", pdf_bytes is not None and len(pdf_bytes) > 0)
    step("PDF starts with %PDF", pdf_bytes[:5] == b"%PDF-")
    step("PDF > 5KB", len(pdf_bytes) > 5000, f"{len(pdf_bytes):,} bytes")

    _pdf_path = os.path.join(tempfile.gettempdir(), "finai_e2e_report.pdf")
    with open(_pdf_path, "wb") as f:
        f.write(pdf_bytes)
    step("PDF saved", os.path.exists(_pdf_path))
except Exception as e:
    step("PDF Report", False, str(e))

# ── Step 8: Generate Excel Report ─────────────────────────────────
print(f"\n{B}Step 8: Excel Report{X}")
try:
    from app.services.excel_report import excel_report
    import io as _io
    excel_bytes = excel_report.generate(rd, "E2E Test Corp", "Jan 2025")
    step("Excel generated", excel_bytes is not None and len(excel_bytes) > 0)

    _wb_check = openpyxl.load_workbook(_io.BytesIO(excel_bytes))
    step("Excel has 5+ sheets", len(_wb_check.sheetnames) >= 5,
         str(_wb_check.sheetnames))
    _wb_check.close()
except Exception as e:
    step("Excel Report", False, str(e))

# ── Step 9: Generate Executive Brief ──────────────────────────────
print(f"\n{B}Step 9: Executive Brief{X}")
try:
    from app.services.executive_brief import executive_brief
    brief = executive_brief.generate(rd, "E2E Test Corp", "Jan 2025", "landscape")
    step("Brief generated", brief is not None and len(brief) > 0)
    step("Brief is valid PDF", brief[:5] == b"%PDF-")
except Exception as e:
    step("Executive Brief", False, str(e))

# ── Step 10: Chat Engine — English ────────────────────────────────
print(f"\n{B}Step 10: Financial Chat Engine (English){X}")
try:
    from app.services.financial_chat import chat_engine
    chat_engine.set_context(
        financials,
        {"revenue": 45000000, "gross_margin_pct": 15.6},
    )
    r1 = chat_engine.query("What is our gross margin?")
    step("Metric query works", r1.intent == "metric_query")
    step("Correct metric identified", r1.data.get("metric") == "gross_margin_pct")
    step("Value present", r1.data.get("value") is not None)

    r2 = chat_engine.query("What are our biggest risks?")
    step("Diagnostic query works", r2.intent == "diagnostic_query")

    r3 = chat_engine.query("What if revenue increases 20%?")
    step("What-if query works", r3.intent == "whatif_query")
    step("Simulation data present", "simulated" in r3.data)
except Exception as e:
    step("Chat Engine EN", False, str(e))

# ── Step 11: Chat Engine — Georgian ───────────────────────────────
print(f"\n{B}Step 11: Financial Chat Engine (Georgian){X}")
try:
    r_ka = chat_engine.query("რა არის ჩვენი შემოსავალი?")
    step("Georgian query works", r_ka.intent == "metric_query")
    step("Georgian language detected", r_ka.language == "ka")
    step("Revenue identified in Georgian", r_ka.data.get("metric") == "revenue")
except Exception as e:
    step("Chat Engine KA", False, str(e))

# ── Step 12: Alert Manager ────────────────────────────────────────
print(f"\n{B}Step 12: Persistent Alert Manager{X}")
try:
    from app.services.persistent_alerts import PersistentAlertManager
    _alert_db = os.path.join(tempfile.gettempdir(), "finai_e2e_alerts.db")
    am = PersistentAlertManager(_alert_db)
    am.clear_all()

    # Trigger on distressed data
    distress = {"net_margin_pct": -25.0, "cash_runway_months": 2.0}
    triggered = am.evaluate_financials(distress)
    step("Alerts triggered on distress", len(triggered) >= 1,
         f"{len(triggered)} alerts")
    step("Critical alert present", any(a.severity == "critical" for a in triggered))

    # Acknowledge
    if triggered:
        ack = am.acknowledge_alert(triggered[0].alert_id)
        step("Alert acknowledged", ack)

    os.remove(_alert_db)
except Exception as e:
    step("Alert Manager", False, str(e))

# ── Step 13: Analogy Search ───────────────────────────────────────
print(f"\n{B}Step 13: Analogy Base Search{X}")
try:
    from app.services.analogy_base import analogy_base
    if not analogy_base._initialized:
        analogy_base.initialize()
    matches = analogy_base.find_analogies(financials, top_k=3)
    step("Analogy search returned matches", len(matches) > 0, f"{len(matches)} matches")
    step("Match has similarity score", hasattr(matches[0], "similarity_score"))
    step("Match has snapshot data", hasattr(matches[0], "snapshot"))
except Exception as e:
    step("Analogy Search", False, str(e))

# ── Cleanup ───────────────────────────────────────────────────────
try:
    os.remove(_xlsx_path)
    os.remove(_db)
except Exception:
    pass

# ── Results ───────────────────────────────────────────────────────
elapsed = time.time() - t0
total = len(results)
passed = sum(1 for _, ok in results if ok)
failed = total - passed

print(f"\n{'=' * 65}")
if failed == 0:
    print(f"  {G}{B}E2E RESULT: {passed}/{total} steps passed ({elapsed:.1f}s){X}")
    print(f"  {G}{B}FULL PIPELINE OPERATIONAL{X}")
else:
    print(f"  {R}{B}E2E RESULT: {passed}/{total} steps passed  |  {failed} FAILED ({elapsed:.1f}s){X}")
    print()
    for name, ok in results:
        if not ok:
            print(f"  {R}[FAIL] {name}{X}")
print(f"{'=' * 65}\n")

sys.exit(0 if failed == 0 else 1)
