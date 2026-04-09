"""
FinAI MR Report Generator
==========================
Reads the NYX Core Thinker MR Report template (ANNUAL_MR_REPORT_2023.xlsx),
fills it with parsed financial data converted to thousands USD,
and exports a ready-to-submit Excel file.

Features:
- GEL → USD conversion using NBG rate (user-controllable)
- Period selection (which month)
- Fills P&L, BS, Product Revenue/COGS sheets
- Preserves template formatting, formulas, and structure

Usage:
    from app.services.mr_report_generator import generate_mr_report
    
    excel_bytes = await generate_mr_report(
        template_path="data/ANNUAL_MR_REPORT_2023.xlsx",
        parsed_data=parse_result.to_dict(),
        period="2026-01",
        gel_usd_rate=2.72,        # Average rate for the month
        gel_usd_rate_eop=2.69,    # End-of-period rate (for BS)
        company_name="NYX Core Thinker LLC",
    )
"""

import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, Optional, Any
from copy import copy
import logging

from app.config import settings

logger = logging.getLogger(__name__)


# ── P&L line mapping: MR Report code → parser field ──
PL_MAPPING = {
    # Revenue
    '01':       lambda d: d.get('revenue', 0),
    '01.A':     lambda d: d.get('revenue', 0) - d.get('revenue_other', 0),  # Product revenue = total - services
    '01.A.02':  lambda d: d.get('revenue_wholesale', 0),  # Oil products (wholesale)
    '01.A.04':  lambda d: 0,  # Natural Gas — needs product-level mapping
    '01.A.06':  lambda d: 0,  # Gases (retail) — needs product-level mapping
    '01.A.07':  lambda d: d.get('revenue_other', 0),  # Other products
    '01.B':     lambda d: 0,  # Services revenue — subset of Other Revenue
    
    # Expenses
    '02':       lambda d: -(d.get('cogs', 0) + d.get('selling_expenses', 0) + d.get('admin_expenses', 0)),
    '02.A':     lambda d: -d.get('cogs', 0),
    '02.A.03':  lambda d: -d.get('cogs', 0),  # Raw materials = main COGS
    'GP':       lambda d: d.get('gross_profit', 0),
    
    # Admin expenses breakdown
    '02.B':     lambda d: -d.get('admin_expenses', 0),
    '02.B.01':  lambda d: -_get_expense_cat(d, 'admin', 'Payroll'),
    '02.B.02':  lambda d: -_get_expense_cat(d, 'admin', 'Depreciation'),
    '02.B.02.01': lambda d: -_get_expense_cat(d, 'admin', 'Depreciation'),
    '02.B.04':  lambda d: -_get_expense_cat(d, 'admin', 'Utilities Expense'),
    '02.B.05':  lambda d: -_get_expense_cat(d, 'admin', 'Maintenance/Renovation'),
    '02.B.06':  lambda d: -_get_expense_cat(d, 'admin', 'Taxes Other'),
    '02.B.07':  lambda d: -_get_expense_cat(d, 'admin', 'Rent'),
    '02.B.08':  lambda d: -_get_expense_cat(d, 'admin', 'Insurance'),
    '02.B.09':  lambda d: -_get_expense_cat(d, 'admin', 'Business Trip'),
    '02.B.11':  lambda d: -_get_expense_cat(d, 'admin', 'Security Expense'),
    '02.B.12':  lambda d: -_get_expense_cat(d, 'admin', 'IT and communication'),
    '02.B.13':  lambda d: -_get_expense_cat(d, 'admin', 'Auto Park Cost'),
    '02.B.15':  lambda d: -(_get_expense_cat(d, 'admin', 'Consulting') + _get_expense_cat(d, 'admin', 'Legal')),
    '02.B.16':  lambda d: -_get_expense_cat(d, 'admin', 'Other G&A'),
    
    # Operating profit
    'OP':       lambda d: d.get('ebit', 0) + d.get('depreciation', 0),  # OP = EBIT + D&A
    
    # Below operating profit
    '06':       lambda d: -d.get('fx_gain_loss', 0) if d.get('fx_gain_loss', 0) < 0 else 0,
    '06.А':     lambda d: -d.get('fx_gain_loss', 0) if d.get('fx_gain_loss', 0) < 0 else 0,
    '07':       lambda d: d.get('fx_gain_loss', 0) if d.get('fx_gain_loss', 0) > 0 else 0,
    '08':       lambda d: d.get('ebit', 0) + d.get('depreciation', 0) + d.get('non_operating_income', 0) - d.get('non_operating_expense', 0),
    'EBITDA':   lambda d: d.get('ebitda', 0),
    '09':       lambda d: d.get('interest_income', 0) - d.get('interest_expense', 0),
    '09.A':     lambda d: d.get('interest_income', 0),
    '09.B':     lambda d: -d.get('interest_expense', 0),
    '10':       lambda d: d.get('profit_before_tax', 0),
    '12':       lambda d: d.get('net_profit', 0),
    '14':       lambda d: d.get('net_profit', 0),
}

# ── BS line mapping ──
BS_MAPPING = {
    '10.B':         lambda bs: bs.get('current_assets', 0),
    '10.B.01':      lambda bs: bs.get('inventories', 0),
    '10.B.01.01':   lambda bs: bs.get('inventories', 0),
    '10.B.02':      lambda bs: bs.get('trade_receivables', 0) + bs.get('prepayments', 0),
    '10.B.02.01':   lambda bs: bs.get('trade_receivables', 0) + bs.get('prepayments', 0),
    '10.B.02.01.01': lambda bs: bs.get('trade_receivables', 0),
    '10.B.02.01.03': lambda bs: bs.get('prepayments', 0),
    '10.B.03':      lambda bs: bs.get('cash', 0),
    '10.B.03.01':   lambda bs: bs.get('cash', 0),
    '10.A':         lambda bs: bs.get('non_current_assets', 0),
    '10.A.01':      lambda bs: bs.get('ppe_net', 0),
    '10.A.01.01':   lambda bs: bs.get('ppe_cost', 0),
    '10.A.01.02':   lambda bs: bs.get('ppe_depreciation', 0),
    '10.A.01.02.02': lambda bs: bs.get('ppe_depreciation', 0),
    '10.A.03':      lambda bs: bs.get('intangible_assets', 0),
    '10.A.04':      lambda bs: bs.get('investments', 0),
    '20.B':         lambda bs: bs.get('current_liabilities', 0),
    '20.B.02':      lambda bs: bs.get('trade_payables', 0),
    '20.A':         lambda bs: bs.get('non_current_liabilities', 0),
    '20.A.01':      lambda bs: bs.get('long_term_loans', 0),
    '30':           lambda bs: bs.get('total_equity', 0),
    '30.01':        lambda bs: bs.get('share_capital', 0),
    '30.03':        lambda bs: bs.get('retained_earnings', 0),
}


def _get_expense_cat(data: dict, expense_type: str, category: str) -> float:
    """Get expense amount from expense detail by category name."""
    detail = data.get(f'{expense_type}_expense_detail', {})
    if isinstance(detail, dict):
        return detail.get(category, 0)
    return 0


def _to_thousands_usd(gel_amount: float, rate: float) -> float:
    """Convert GEL amount to thousands USD."""
    if rate <= 0:
        return 0
    return gel_amount / rate / 1000


async def generate_mr_report(
    template_path: str,
    parsed_data: Dict[str, Any],
    period: str = "2026-01",
    gel_usd_rate: float = 2.72,
    gel_usd_rate_eop: float = 2.69,
    gel_usd_rate_boy: float = 2.70,
    company_name: str = None,
    include_sheets: Optional[list] = None,
) -> bytes:
    """
    Generate MR Report Excel from template + parsed data.

    Args:
        template_path: Path to ANNUAL_MR_REPORT_2023.xlsx template
        parsed_data: Output from parse_nyx_excel().to_dict() or smart-upload response
        period: Period string like "2026-01"
        gel_usd_rate: Average GEL/USD rate for the month (for P&L conversion)
        gel_usd_rate_eop: End-of-period rate (for BS conversion)
        gel_usd_rate_boy: Beginning-of-year rate
        company_name: Company name to put in headers
        include_sheets: List of sheet names to fill (None = all available)

    Returns:
        Excel file bytes
    """
    company_name = company_name or settings.COMPANY_NAME

    wb = openpyxl.load_workbook(template_path)
    
    pnl_data = parsed_data.get('pnl', {})
    if isinstance(pnl_data, dict) and hasattr(pnl_data, '__dict__'):
        pnl_data = pnl_data.__dict__
    elif not isinstance(pnl_data, dict):
        pnl_data = {}
    
    bs_data = parsed_data.get('balance_sheet', {})
    if isinstance(bs_data, dict) and hasattr(bs_data, '__dict__'):
        bs_data = bs_data.__dict__
    elif not isinstance(bs_data, dict):
        bs_data = {}
    
    # Add expense detail to pnl_data for category lookups
    pnl_data['admin_expense_detail'] = parsed_data.get('admin_expense_detail', {})
    pnl_data['selling_expense_detail'] = parsed_data.get('selling_expense_detail', {})
    
    # ── Fill P&L Sheet ──
    if 'P&L' in wb.sheetnames and (include_sheets is None or 'P&L' in include_sheets):
        ws = wb['P&L']
        _fill_pl_sheet(ws, pnl_data, gel_usd_rate, company_name, period)
        logger.info("Filled P&L sheet")
    
    # ── Fill BS Sheet ──
    if 'BS' in wb.sheetnames and bs_data and (include_sheets is None or 'BS' in include_sheets):
        ws = wb['BS']
        _fill_bs_sheet(ws, bs_data, gel_usd_rate_eop, company_name, period)
        logger.info("Filled BS sheet")
    
    # ── Fill Currency Sheet ──
    if 'Currency' in wb.sheetnames:
        ws = wb['Currency']
        _fill_currency_sheet(ws, gel_usd_rate, gel_usd_rate_eop, gel_usd_rate_boy)
    
    # ── Remove Azerbaijani text ──
    _clean_azerbaijani(wb)
    
    # ── Save to bytes ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _clean_azerbaijani(wb):
    """Remove all Azerbaijani text from the workbook, keeping only English."""
    az_chars = set('əıöüşçğƏİÖÜŞÇĞ')
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                v = cell.value
                
                # Split bilingual "English / Azerbaijani" patterns
                if ' / ' in v:
                    parts = v.split(' / ')
                    if len(parts) == 2 and len(parts[1]) < 100:
                        cell.value = parts[0].strip()
                        continue
                
                # Remove pure Azerbaijani cells
                if sum(1 for c in v if c in az_chars) >= 2:
                    cell.value = None
                    continue
                
                # Clean multiline with AZ
                if '\n' in v:
                    lines = v.split('\n')
                    en_lines = [l for l in lines if not any(c in l for c in az_chars)]
                    if len(en_lines) < len(lines):
                        cell.value = '\n'.join(en_lines).strip() or None
                
                # Remove AZ guidance text
                if v.strip() in ('Təlimat:', 'Guidance:') or v.strip().startswith('Zəhmət') or v.strip().startswith('! Zəhmət'):
                    cell.value = None


def _fill_pl_sheet(ws, pnl_data: dict, rate: float, company: str, period: str):
    """Fill the P&L sheet with converted values."""
    
    # Update company name (row 4, col C)
    ws['C4'] = company
    
    # Find data rows and fill column F (Current period actual)
    # The P&L sheet has codes in column A starting from row 11
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row):
        code_cell = row[0]  # Column A
        code = str(code_cell.value or '').strip()
        
        if code in PL_MAPPING:
            gel_value = PL_MAPPING[code](pnl_data)
            usd_thousands = _to_thousands_usd(gel_value, rate)
            
            # Column F (index 5) = Current period actual
            actual_cell = row[5]
            if usd_thousands != 0:
                actual_cell.value = round(usd_thousands, 2)
            
            # Column Q (index 16) = Cumulative actual (same as monthly for single month)
            if len(row) > 16:
                cum_cell = row[16]
                if usd_thousands != 0:
                    cum_cell.value = round(usd_thousands, 2)


def _fill_bs_sheet(ws, bs_data: dict, rate: float, company: str, period: str):
    """Fill the BS sheet with converted values."""
    
    ws['C4'] = company
    
    for row in ws.iter_rows(min_row=13, max_row=ws.max_row):
        code_cell = row[0]  # Column A
        code = str(code_cell.value or '').strip()
        
        if code in BS_MAPPING:
            gel_value = BS_MAPPING[code](bs_data)
            usd_thousands = _to_thousands_usd(gel_value, rate)
            
            # Column G (index 6) = Current period actual
            actual_cell = row[6]
            if usd_thousands != 0:
                actual_cell.value = round(usd_thousands, 2)


def _fill_currency_sheet(ws, avg_rate: float, eop_rate: float, boy_rate: float):
    """Fill the Currency sheet with exchange rates."""
    # Row 5: GEL rates
    # B=Plan, C=BOY, D=BOM, E=Avg month, F=Avg cumulative, G=EOP
    for row in ws.iter_rows(min_row=5, max_row=5):
        if str(row[0].value or '').strip() == 'GEL':
            row[2].value = boy_rate     # C = Beginning of year
            row[3].value = eop_rate     # D = Beginning of month (approx)
            row[4].value = avg_rate     # E = Average for month
            row[5].value = avg_rate     # F = Average cumulative
            row[6].value = eop_rate     # G = End of period


# ── FastAPI endpoint ──

async def mr_report_endpoint(request_data: dict, template_path: str = "data/ANNUAL_MR_REPORT_2023.xlsx") -> bytes:
    """
    API endpoint handler for MR Report generation.
    
    Expected request_data:
    {
        "period": "2026-01",
        "gel_usd_rate": 2.72,           # User can override
        "gel_usd_rate_eop": 2.69,       # User can override
        "gel_usd_rate_boy": 2.70,       # User can override
        "company_name": "NYX Core Thinker LLC",
        "include_sheets": ["P&L", "BS", "Currency"],  # Optional filter
        "parsed_data": { ... }          # From smart-upload or DB
    }
    """
    return await generate_mr_report(
        template_path=template_path,
        parsed_data=request_data.get('parsed_data', {}),
        period=request_data.get('period', '2026-01'),
        gel_usd_rate=request_data.get('gel_usd_rate', 2.72),
        gel_usd_rate_eop=request_data.get('gel_usd_rate_eop', 2.69),
        gel_usd_rate_boy=request_data.get('gel_usd_rate_boy', 2.70),
        company_name=request_data.get('company_name', settings.COMPANY_NAME),
        include_sheets=request_data.get('include_sheets'),
    )
