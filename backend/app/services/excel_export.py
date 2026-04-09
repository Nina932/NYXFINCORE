"""
FinAI Excel Export — Professional Financial Reports (.xlsx)
============================================================
Generates a modern, styled multi-sheet Excel workbook with:
  - Executive Summary (cover sheet with KPI dashboard)
  - P&L Statement (hierarchical, formulas, conditional bars)
  - Balance Sheet (sections, equation check, ratios)
  - Revenue Breakdown (sorted, % bars, category grouping)
  - COGS Breakdown (sorted, % contribution)
  - KPI Dashboard (status indicators, benchmarks)
  - Data Dictionary (field definitions)
"""

from __future__ import annotations
import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle,
)
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════════════════
# DESIGN SYSTEM
# ══════════════════════════════════════════════════════════════════

# Colors — modern dark finance palette
C_NAVY      = "0C1527"
C_DARK_BLUE = "152238"
C_BLUE      = "1E3A5F"
C_SKY       = "38BDF8"
C_WHITE     = "FFFFFF"
C_OFFWHITE  = "F8FAFC"
C_LIGHT     = "E2E8F0"
C_MUTED     = "94A3B8"
C_GREEN     = "10B981"
C_GREEN_BG  = "ECFDF5"
C_RED       = "EF4444"
C_RED_BG    = "FEF2F2"
C_AMBER     = "F59E0B"
C_AMBER_BG  = "FFFBEB"
C_VIOLET    = "8B5CF6"

# Fills
FILL_HEADER     = PatternFill(start_color=C_NAVY, end_color=C_NAVY, fill_type="solid")
FILL_SUBHEADER  = PatternFill(start_color=C_DARK_BLUE, end_color=C_DARK_BLUE, fill_type="solid")
FILL_SECTION    = PatternFill(start_color=C_BLUE, end_color=C_BLUE, fill_type="solid")
FILL_TOTAL      = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
FILL_SUBTOTAL   = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_GREEN_BG   = PatternFill(start_color=C_GREEN_BG, end_color=C_GREEN_BG, fill_type="solid")
FILL_RED_BG     = PatternFill(start_color=C_RED_BG, end_color=C_RED_BG, fill_type="solid")
FILL_AMBER_BG   = PatternFill(start_color=C_AMBER_BG, end_color=C_AMBER_BG, fill_type="solid")
FILL_WHITE      = PatternFill(start_color=C_WHITE, end_color=C_WHITE, fill_type="solid")
FILL_OFFWHITE   = PatternFill(start_color=C_OFFWHITE, end_color=C_OFFWHITE, fill_type="solid")
FILL_COVER_BG   = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

# Fonts
FN_TITLE      = Font(name="Calibri", size=18, bold=True, color=C_NAVY)
FN_SUBTITLE   = Font(name="Calibri", size=11, color=C_MUTED)
FN_HEADER     = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
FN_SECTION    = Font(name="Calibri", size=11, bold=True, color=C_WHITE)
FN_LABEL      = Font(name="Calibri", size=10, color="334155")
FN_LABEL_BOLD = Font(name="Calibri", size=10, bold=True, color="1E293B")
FN_MONEY      = Font(name="Consolas", size=10, color="1E293B")
FN_MONEY_BOLD = Font(name="Consolas", size=10, bold=True, color="0F172A")
FN_MONEY_NEG  = Font(name="Consolas", size=10, color=C_RED)
FN_PCT        = Font(name="Consolas", size=9, color=C_MUTED)
FN_STATUS_OK  = Font(name="Calibri", size=10, bold=True, color=C_GREEN)
FN_STATUS_BAD = Font(name="Calibri", size=10, bold=True, color=C_RED)
FN_STATUS_MED = Font(name="Calibri", size=10, bold=True, color=C_AMBER)
FN_COVER_T    = Font(name="Calibri", size=28, bold=True, color=C_WHITE)
FN_COVER_S    = Font(name="Calibri", size=14, color=C_SKY)
FN_COVER_D    = Font(name="Calibri", size=11, color=C_MUTED)
FN_COVER_KPI_LABEL = Font(name="Calibri", size=9, color=C_MUTED)
FN_COVER_KPI_VALUE = Font(name="Consolas", size=20, bold=True, color=C_WHITE)
FN_SMALL_MUTED = Font(name="Calibri", size=8, color=C_MUTED)

# Borders
BORDER_BOTTOM = Border(bottom=Side(style="thin", color=C_LIGHT))
BORDER_TOTAL  = Border(
    top=Side(style="medium", color=C_NAVY),
    bottom=Side(style="double", color=C_NAVY),
)
BORDER_SECTION = Border(bottom=Side(style="medium", color=C_BLUE))

# Alignments
AL_LEFT   = Alignment(horizontal="left", vertical="center")
AL_RIGHT  = Alignment(horizontal="right", vertical="center")
AL_CENTER = Alignment(horizontal="center", vertical="center")
AL_INDENT1 = Alignment(horizontal="left", vertical="center", indent=1)
AL_INDENT2 = Alignment(horizontal="left", vertical="center", indent=2)
AL_INDENT3 = Alignment(horizontal="left", vertical="center", indent=3)
AL_WRAP    = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
FMT_MONEY    = '#,##0'
FMT_MONEY_DEC = '#,##0.00'
FMT_PCT      = '0.0%'
FMT_RATIO    = '0.00"x"'

INDENTS = [AL_LEFT, AL_INDENT1, AL_INDENT2, AL_INDENT3]


# ══════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════

def _set_col_widths(ws, widths: Dict[int, float]):
    """Set specific column widths."""
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _row_fill(ws, row: int, cols: int, fill: PatternFill):
    """Fill entire row with a color."""
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = fill


def _stripe_rows(ws, start_row: int, end_row: int, cols: int):
    """Apply zebra striping to rows."""
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = FILL_OFFWHITE


def _write_header_row(ws, row: int, headers: List[str], aligns: Optional[List[Alignment]] = None) -> int:
    """Write a styled header row with navy background."""
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = FN_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = aligns[i - 1] if aligns else AL_CENTER
    ws.row_dimensions[row].height = 30
    return row + 1


def _write_section_row(ws, row: int, title: str, cols: int) -> int:
    """Write a colored section separator."""
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FN_SECTION
    cell.fill = FILL_SECTION
    for c in range(2, cols + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
    ws.row_dimensions[row].height = 28
    return row + 1


def _write_money(ws, row: int, col: int, value, is_total=False, is_negative=False):
    """Write a formatted money cell."""
    cell = ws.cell(row=row, column=col, value=value if value else 0)
    cell.number_format = FMT_MONEY
    cell.alignment = AL_RIGHT
    if is_total:
        cell.font = FN_MONEY_BOLD
        cell.fill = FILL_TOTAL
        cell.border = BORDER_TOTAL
    elif is_negative or (isinstance(value, (int, float)) and value < 0):
        cell.font = FN_MONEY_NEG
    else:
        cell.font = FN_MONEY
    return cell


def _write_pct(ws, row: int, col: int, value: float):
    """Write a percentage cell."""
    cell = ws.cell(row=row, column=col, value=value / 100 if abs(value) <= 100 else value)
    cell.number_format = FMT_PCT
    cell.alignment = AL_RIGHT
    cell.font = FN_PCT
    return cell


def _status_cell(ws, row: int, col: int, status: str, value: str = ""):
    """Write a status indicator cell with colored background."""
    cell = ws.cell(row=row, column=col, value=value or status)
    cell.alignment = AL_CENTER
    if status in ("Good", "Healthy", "Profitable", "Positive", "Safe", "Strong"):
        cell.font = FN_STATUS_OK
        cell.fill = FILL_GREEN_BG
    elif status in ("Bad", "Loss", "Negative", "High", "Risk", "Weak", "Critical"):
        cell.font = FN_STATUS_BAD
        cell.fill = FILL_RED_BG
    else:
        cell.font = FN_STATUS_MED
        cell.fill = FILL_AMBER_BG
    return cell


# ══════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════

class ExcelExporter:
    """Generate professional styled Excel financial reports."""

    def generate(
        self,
        pnl: Dict[str, Any],
        balance_sheet: Optional[Dict[str, Any]] = None,
        revenue_breakdown: Optional[List[Dict]] = None,
        cogs_breakdown: Optional[List[Dict]] = None,
        pl_line_items: Optional[List[Dict]] = None,
        company: str = "Company",
        period: str = "",
    ) -> bytes:
        """Generate a multi-sheet Excel workbook. Returns bytes."""
        wb = Workbook()
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

        # ── Sheet 1: Executive Summary (Cover) ──
        ws_cover = wb.active
        ws_cover.title = "Executive Summary"
        self._write_cover(ws_cover, pnl, balance_sheet, company, period, now)

        # ── Sheet 2: P&L Statement ──
        ws_pnl = wb.create_sheet("P&L Statement")
        self._write_pnl(ws_pnl, pnl, pl_line_items, company, period)

        # ── Sheet 3: Balance Sheet ──
        if balance_sheet:
            ws_bs = wb.create_sheet("Balance Sheet")
            self._write_balance_sheet(ws_bs, balance_sheet, company, period)

        # ── Sheet 4: Revenue Breakdown ──
        if revenue_breakdown:
            ws_rev = wb.create_sheet("Revenue")
            self._write_revenue(ws_rev, revenue_breakdown, company, period)

        # ── Sheet 5: COGS Breakdown ──
        if cogs_breakdown:
            ws_cogs = wb.create_sheet("COGS")
            self._write_cogs(ws_cogs, cogs_breakdown, company, period)

        # ── Sheet 6: KPI Dashboard ──
        ws_kpi = wb.create_sheet("KPI Dashboard")
        self._write_kpis(ws_kpi, pnl, balance_sheet, company, period)

        # Global workbook settings
        for ws in wb.worksheets:
            ws.sheet_properties.tabColor = "38BDF8" if ws.title == "Executive Summary" else "1E3A5F"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # ──────────────────────────────────────────────────────────────
    # EXECUTIVE SUMMARY (Cover Sheet)
    # ──────────────────────────────────────────────────────────────
    def _write_cover(self, ws, pnl: Dict, bs: Optional[Dict], company: str, period: str, generated: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 3, 2: 22, 3: 22, 4: 22, 5: 22, 6: 3})

        # Dark background for entire cover
        for r in range(1, 35):
            _row_fill(ws, r, 6, FILL_COVER_BG)

        # Company name
        ws.merge_cells("B3:E3")
        ws.cell(row=3, column=2, value=company).font = FN_COVER_T
        ws.cell(row=3, column=2).alignment = AL_LEFT

        # Subtitle
        ws.merge_cells("B4:E4")
        ws.cell(row=4, column=2, value="Financial Report").font = FN_COVER_S
        ws.cell(row=4, column=2).alignment = AL_LEFT

        # Period + Generated
        ws.merge_cells("B5:E5")
        ws.cell(row=5, column=2, value=f"Period: {period}   |   Generated: {generated}").font = FN_COVER_D
        ws.cell(row=5, column=2).alignment = AL_LEFT

        # ── KPI Cards (row 8-11) ──
        revenue = pnl.get("revenue", pnl.get("total_revenue", 0))
        gross_profit = pnl.get("gross_profit", 0)
        ebitda = pnl.get("ebitda", 0)
        net_profit = pnl.get("net_profit", 0)
        gross_margin = pnl.get("gross_margin_pct", 0)
        net_margin = pnl.get("net_margin_pct", 0)

        kpi_cards = [
            ("REVENUE", f"₾{revenue / 1e6:,.1f}M", C_SKY),
            ("GROSS PROFIT", f"₾{gross_profit / 1e6:,.1f}M", C_GREEN),
            ("EBITDA", f"₾{ebitda / 1e6:,.1f}M", C_VIOLET),
            ("NET PROFIT", f"₾{net_profit / 1e6:,.1f}M", C_GREEN if net_profit >= 0 else C_RED),
        ]

        for i, (label, value, color) in enumerate(kpi_cards):
            col = i + 2
            # Label
            cell_l = ws.cell(row=8, column=col, value=label)
            cell_l.font = FN_COVER_KPI_LABEL
            cell_l.alignment = AL_LEFT
            # Value
            cell_v = ws.cell(row=9, column=col, value=value)
            cell_v.font = Font(name="Consolas", size=20, bold=True, color=color)
            cell_v.alignment = AL_LEFT

        # ── Margin Cards (row 12-14) ──
        margin_cards = [
            ("GROSS MARGIN", f"{gross_margin:.1f}%"),
            ("NET MARGIN", f"{net_margin:.1f}%"),
            ("EBITDA MARGIN", f"{(ebitda / revenue * 100 if revenue else 0):.1f}%"),
        ]

        for i, (label, value) in enumerate(margin_cards):
            col = i + 2
            ws.cell(row=12, column=col, value=label).font = FN_COVER_KPI_LABEL
            ws.cell(row=12, column=col).alignment = AL_LEFT
            ws.cell(row=13, column=col, value=value).font = Font(name="Consolas", size=16, bold=True, color=C_WHITE)
            ws.cell(row=13, column=col).alignment = AL_LEFT

        # ── Balance Sheet Summary (row 16+) ──
        if bs:
            ws.cell(row=16, column=2, value="BALANCE SHEET SUMMARY").font = Font(name="Calibri", size=10, bold=True, color=C_SKY)
            bs_items = [
                ("Total Assets", bs.get("total_assets", 0)),
                ("Total Liabilities", bs.get("total_liabilities", 0)),
                ("Total Equity", bs.get("total_equity", 0)),
            ]
            for i, (label, val) in enumerate(bs_items):
                r = 17 + i
                ws.cell(row=r, column=2, value=label).font = FN_COVER_KPI_LABEL
                ws.cell(row=r, column=3, value=f"₾{val / 1e6:,.1f}M").font = Font(name="Consolas", size=12, bold=True, color=C_WHITE)

        # ── Sheet Index (row 22+) ──
        ws.cell(row=22, column=2, value="REPORT CONTENTS").font = Font(name="Calibri", size=10, bold=True, color=C_SKY)
        sheets = [
            ("01", "P&L Statement", "Income statement with line-item detail"),
            ("02", "Balance Sheet", "Assets, liabilities, equity breakdown"),
            ("03", "Revenue", "Product-level revenue analysis"),
            ("04", "COGS", "Cost of goods sold breakdown"),
            ("05", "KPI Dashboard", "Key metrics with status indicators"),
        ]
        for i, (num, name, desc) in enumerate(sheets):
            r = 23 + i
            ws.cell(row=r, column=2, value=f"  {num}  {name}").font = Font(name="Consolas", size=10, color=C_WHITE)
            ws.cell(row=r, column=4, value=desc).font = FN_COVER_D

        # Footer
        ws.cell(row=30, column=2, value="Generated by FinAI Financial Intelligence Platform").font = FN_SMALL_MUTED
        ws.cell(row=31, column=2, value="© Confidential — For internal use only").font = FN_SMALL_MUTED

        # Print setup
        ws.print_area = "A1:F32"

    # ──────────────────────────────────────────────────────────────
    # P&L STATEMENT
    # ──────────────────────────────────────────────────────────────
    def _write_pnl(self, ws, pnl: Dict, line_items: Optional[List[Dict]], company: str, period: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 12, 2: 45, 3: 18, 4: 12})

        # Title
        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value=f"{company}").font = FN_TITLE
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value=f"Income Statement  —  {period}").font = FN_SUBTITLE
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 8  # spacer

        row = 4

        if line_items:
            row = _write_header_row(ws, row, ["Code", "Description", "Amount (₾)", "% Rev"],
                                     [AL_CENTER, AL_LEFT, AL_RIGHT, AL_RIGHT])
            ws.freeze_panes = f"A{row}"

            revenue = abs(pnl.get("revenue", pnl.get("total_revenue", 1)))
            data_start = row

            for item in line_items:
                level = item.get("level", 0)
                item_type = item.get("type", "")
                amount = item.get("amount", 0)
                is_total = item_type in ("total", "subtotal", "grand_total")
                is_header = item_type == "header"

                # Code column
                code_cell = ws.cell(row=row, column=1, value=item.get("code", ""))
                code_cell.font = Font(name="Consolas", size=8, color=C_MUTED)
                code_cell.alignment = AL_CENTER

                # Label column with indent
                label = item.get("label", "")
                label_cell = ws.cell(row=row, column=2, value=label)
                label_cell.alignment = INDENTS[min(level, 3)]

                if is_total or is_header:
                    label_cell.font = FN_LABEL_BOLD
                    if is_total:
                        _row_fill(ws, row, 4, FILL_TOTAL)
                else:
                    label_cell.font = FN_LABEL

                # Amount
                _write_money(ws, row, 3, amount, is_total=is_total)

                # % of Revenue
                if revenue and amount != 0:
                    pct = amount / revenue * 100
                    _write_pct(ws, row, 4, pct)

                # Row border
                for c in range(1, 5):
                    ws.cell(row=row, column=c).border = BORDER_BOTTOM

                row += 1

            # Zebra stripe non-total rows
            _stripe_rows(ws, data_start, row - 1, 4)

        else:
            # Fallback: build from pnl dict
            row = _write_header_row(ws, row, ["", "Line Item", "Amount (₾)", "% Rev"],
                                     [AL_CENTER, AL_LEFT, AL_RIGHT, AL_RIGHT])
            ws.freeze_panes = f"A{row}"

            revenue = pnl.get("revenue", pnl.get("total_revenue", 0))
            cogs = pnl.get("cogs", pnl.get("total_cogs", 0))
            gross_profit = pnl.get("gross_profit", 0)
            selling = pnl.get("selling_expenses", 0)
            admin = pnl.get("admin_expenses", pnl.get("ga_expenses", 0))
            ebitda = pnl.get("ebitda", 0)
            depreciation = pnl.get("depreciation", 0)
            net_profit = pnl.get("net_profit", 0)

            lines = [
                ("", "Revenue", revenue, True, 0),
                ("", "  Cost of Goods Sold", -abs(cogs) if cogs else 0, False, 1),
                ("", "Gross Profit", gross_profit, True, 0),
                ("", "  Selling Expenses", -abs(selling) if selling else 0, False, 1),
                ("", "  Administrative Expenses", -abs(admin) if admin else 0, False, 1),
                ("", "EBITDA", ebitda, True, 0),
                ("", "  Depreciation", -abs(depreciation) if depreciation else 0, False, 1),
                ("", "Net Profit / (Loss)", net_profit, True, 0),
            ]

            for code, label, amount, is_total, level in lines:
                if amount == 0 and not is_total:
                    continue
                ws.cell(row=row, column=1, value=code).font = Font(name="Consolas", size=8, color=C_MUTED)
                label_cell = ws.cell(row=row, column=2, value=label)
                label_cell.alignment = INDENTS[min(level, 3)]
                label_cell.font = FN_LABEL_BOLD if is_total else FN_LABEL
                if is_total:
                    _row_fill(ws, row, 4, FILL_TOTAL)
                _write_money(ws, row, 3, amount, is_total=is_total)
                if revenue:
                    _write_pct(ws, row, 4, amount / revenue * 100)
                for c in range(1, 5):
                    ws.cell(row=row, column=c).border = BORDER_BOTTOM
                row += 1

        # Print settings
        ws.print_area = f"A1:D{row}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1

    # ──────────────────────────────────────────────────────────────
    # BALANCE SHEET
    # ──────────────────────────────────────────────────────────────
    def _write_balance_sheet(self, ws, bs: Dict, company: str, period: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 40, 2: 20, 3: 16})

        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value=company).font = FN_TITLE
        ws.merge_cells("A2:C2")
        ws.cell(row=2, column=1, value=f"Statement of Financial Position  —  {period}").font = FN_SUBTITLE
        ws.row_dimensions[3].height = 8

        row = 4
        total_assets = bs.get("total_assets", 0)

        sections = [
            ("ASSETS", [
                ("Cash & Cash Equivalents", bs.get("cash", bs.get("cash_and_equivalents", 0))),
                ("Accounts Receivable", bs.get("accounts_receivable", bs.get("receivables", 0))),
                ("Inventory", bs.get("inventory", 0)),
                ("Other Current Assets", bs.get("other_current_assets", 0)),
                ("Total Current Assets", bs.get("total_current_assets", 0)),
                (None, None),  # separator
                ("Fixed Assets (Net)", bs.get("fixed_assets_net", bs.get("fixed_assets", 0))),
                ("Investments", bs.get("investments", 0)),
                ("Intangible Assets", bs.get("intangible_assets", 0)),
                ("Other Non-Current Assets", bs.get("other_noncurrent_assets", 0)),
                (None, None),
                ("TOTAL ASSETS", total_assets),
            ]),
            ("LIABILITIES", [
                ("Accounts Payable", bs.get("accounts_payable", bs.get("payables", 0))),
                ("Short-term Debt", bs.get("short_term_debt", 0)),
                ("Other Current Liabilities", bs.get("other_current_liabilities", 0)),
                ("Total Current Liabilities", bs.get("total_current_liabilities", 0)),
                (None, None),
                ("Long-term Debt", bs.get("long_term_debt", 0)),
                ("Other Non-Current Liabilities", bs.get("other_noncurrent_liabilities", 0)),
                (None, None),
                ("TOTAL LIABILITIES", bs.get("total_liabilities", 0)),
            ]),
            ("EQUITY", [
                ("Share Capital", bs.get("share_capital", 0)),
                ("Retained Earnings", bs.get("retained_earnings", 0)),
                ("Other Equity", bs.get("other_equity", 0)),
                (None, None),
                ("TOTAL EQUITY", bs.get("total_equity", 0)),
            ]),
        ]

        for section_name, items in sections:
            row = _write_section_row(ws, row, section_name, 3)
            for label, amount in items:
                if label is None:
                    row += 1  # blank separator
                    continue
                if amount == 0 and "TOTAL" not in label and "Total" not in label:
                    continue
                is_total = label.startswith("TOTAL") or label.startswith("Total")
                ws.cell(row=row, column=1, value=label).font = FN_LABEL_BOLD if is_total else FN_LABEL
                ws.cell(row=row, column=1).alignment = AL_LEFT if is_total else AL_INDENT1
                _write_money(ws, row, 2, amount, is_total=is_total)
                # % of total assets
                if total_assets and amount:
                    _write_pct(ws, row, 3, amount / total_assets * 100)
                for c in range(1, 4):
                    ws.cell(row=row, column=c).border = BORDER_BOTTOM
                if is_total:
                    _row_fill(ws, row, 3, FILL_TOTAL)
                row += 1

        # ── Equation Check ──
        row += 1
        total_le = bs.get("total_liabilities", 0) + bs.get("total_equity", 0)
        balanced = abs(total_assets - total_le) < 1
        ws.cell(row=row, column=1, value="Balance Sheet Equation Check").font = FN_LABEL_BOLD
        _status_cell(ws, row, 2, "Good" if balanced else "Bad", "✓ BALANCED" if balanced else "✗ UNBALANCED")

        # ── Key Ratios ──
        row += 2
        row = _write_section_row(ws, row, "KEY RATIOS", 3)
        ratios = [
            ("Current Ratio", bs.get("current_ratio", 0), "x", 1.5, 1.0),
            ("Debt to Equity", bs.get("debt_to_equity", 0), "x", 2.0, 3.0),
        ]
        for label, val, suffix, good_thresh, bad_thresh in ratios:
            ws.cell(row=row, column=1, value=label).font = FN_LABEL
            ws.cell(row=row, column=2, value=f"{val:.2f}{suffix}").font = FN_MONEY_BOLD
            ws.cell(row=row, column=2).alignment = AL_RIGHT
            status = "Good" if (val >= good_thresh if suffix == "x" and "Current" in label else val <= good_thresh) else "Bad" if (val < bad_thresh if "Current" in label else val > bad_thresh) else "OK"
            _status_cell(ws, row, 3, status)
            row += 1

        ws.print_area = f"A1:C{row}"

    # ──────────────────────────────────────────────────────────────
    # REVENUE BREAKDOWN
    # ──────────────────────────────────────────────────────────────
    def _write_revenue(self, ws, breakdown: List[Dict], company: str, period: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 4, 2: 35, 3: 18, 4: 18, 5: 12})

        ws.merge_cells("A1:E1")
        ws.cell(row=1, column=1, value=company).font = FN_TITLE
        ws.merge_cells("A2:E2")
        ws.cell(row=2, column=1, value=f"Revenue Analysis  —  {period}").font = FN_SUBTITLE
        ws.row_dimensions[3].height = 8

        row = 4
        row = _write_header_row(ws, row, ["#", "Product", "Category", "Net Revenue (₾)", "% Total"],
                                 [AL_CENTER, AL_LEFT, AL_LEFT, AL_RIGHT, AL_RIGHT])
        ws.freeze_panes = f"A{row}"

        total = sum(abs(item.get("net_revenue", 0)) for item in breakdown)
        sorted_items = sorted(breakdown, key=lambda x: abs(x.get("net_revenue", 0)), reverse=True)
        data_start = row

        for idx, item in enumerate(sorted_items, 1):
            rev = item.get("net_revenue", 0)
            ws.cell(row=row, column=1, value=idx).font = Font(name="Consolas", size=9, color=C_MUTED)
            ws.cell(row=row, column=1).alignment = AL_CENTER
            ws.cell(row=row, column=2, value=item.get("product", "")).font = FN_LABEL
            cat = item.get("category", "")
            cat_cell = ws.cell(row=row, column=3, value=cat)
            cat_cell.font = Font(name="Calibri", size=9, color=C_MUTED)
            _write_money(ws, row, 4, rev)
            if total:
                _write_pct(ws, row, 5, abs(rev) / total * 100)
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = BORDER_BOTTOM
            row += 1

        _stripe_rows(ws, data_start, row - 1, 5)

        # Total row
        ws.cell(row=row, column=2, value="TOTAL REVENUE").font = FN_LABEL_BOLD
        _write_money(ws, row, 4, total, is_total=True)
        ws.cell(row=row, column=5, value="100.0%").font = FN_MONEY_BOLD
        ws.cell(row=row, column=5).alignment = AL_RIGHT
        _row_fill(ws, row, 5, FILL_TOTAL)

        # Data bars on revenue column
        if row > data_start + 1:
            try:
                rule = DataBarRule(start_type="min", end_type="max", color=C_SKY)
                ws.conditional_formatting.add(f"D{data_start}:D{row - 1}", rule)
            except Exception:
                pass

        # Category summary below
        row += 2
        row = _write_section_row(ws, row, "REVENUE BY CATEGORY", 5)
        cats: Dict[str, float] = {}
        for item in breakdown:
            c = item.get("category", "Other")
            cats[c] = cats.get(c, 0) + abs(item.get("net_revenue", 0))
        for cat_name, cat_total in sorted(cats.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=2, value=cat_name).font = FN_LABEL
            _write_money(ws, row, 4, cat_total)
            if total:
                _write_pct(ws, row, 5, cat_total / total * 100)
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = BORDER_BOTTOM
            row += 1

        ws.print_area = f"A1:E{row}"

    # ──────────────────────────────────────────────────────────────
    # COGS BREAKDOWN
    # ──────────────────────────────────────────────────────────────
    def _write_cogs(self, ws, breakdown: List[Dict], company: str, period: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 4, 2: 35, 3: 18, 4: 12})

        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value=company).font = FN_TITLE
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value=f"Cost of Goods Sold  —  {period}").font = FN_SUBTITLE
        ws.row_dimensions[3].height = 8

        row = 4
        row = _write_header_row(ws, row, ["#", "Product", "Amount (₾)", "% COGS"],
                                 [AL_CENTER, AL_LEFT, AL_RIGHT, AL_RIGHT])
        ws.freeze_panes = f"A{row}"

        total = sum(abs(item.get("amount", 0)) for item in breakdown)
        sorted_items = sorted(breakdown, key=lambda x: abs(x.get("amount", 0)), reverse=True)
        data_start = row

        for idx, item in enumerate(sorted_items, 1):
            amt = abs(item.get("amount", 0))
            ws.cell(row=row, column=1, value=idx).font = Font(name="Consolas", size=9, color=C_MUTED)
            ws.cell(row=row, column=1).alignment = AL_CENTER
            ws.cell(row=row, column=2, value=item.get("product", "")).font = FN_LABEL
            _write_money(ws, row, 3, amt)
            if total:
                _write_pct(ws, row, 4, amt / total * 100)
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = BORDER_BOTTOM
            row += 1

        _stripe_rows(ws, data_start, row - 1, 4)

        ws.cell(row=row, column=2, value="TOTAL COGS").font = FN_LABEL_BOLD
        _write_money(ws, row, 3, total, is_total=True)
        ws.cell(row=row, column=4, value="100.0%").font = FN_MONEY_BOLD
        ws.cell(row=row, column=4).alignment = AL_RIGHT
        _row_fill(ws, row, 4, FILL_TOTAL)

        if row > data_start + 1:
            try:
                rule = DataBarRule(start_type="min", end_type="max", color=C_RED)
                ws.conditional_formatting.add(f"C{data_start}:C{row - 1}", rule)
            except Exception:
                pass

        ws.print_area = f"A1:D{row}"

    # ──────────────────────────────────────────────────────────────
    # KPI DASHBOARD
    # ──────────────────────────────────────────────────────────────
    def _write_kpis(self, ws, pnl: Dict, bs: Optional[Dict], company: str, period: str):
        ws.sheet_view.showGridLines = False
        _set_col_widths(ws, {1: 30, 2: 18, 3: 14, 4: 30})

        ws.merge_cells("A1:D1")
        ws.cell(row=1, column=1, value=company).font = FN_TITLE
        ws.merge_cells("A2:D2")
        ws.cell(row=2, column=1, value=f"Key Performance Indicators  —  {period}").font = FN_SUBTITLE
        ws.row_dimensions[3].height = 8

        row = 4
        row = _write_header_row(ws, row, ["Metric", "Value", "Status", "Benchmark / Note"],
                                 [AL_LEFT, AL_RIGHT, AL_CENTER, AL_LEFT])
        ws.freeze_panes = f"A{row}"

        revenue = pnl.get("revenue", pnl.get("total_revenue", 0))
        net_profit = pnl.get("net_profit", 0)
        gross_margin = pnl.get("gross_margin_pct", 0)
        net_margin = pnl.get("net_margin_pct", 0)
        ebitda = pnl.get("ebitda", 0)
        ebitda_margin = ebitda / revenue * 100 if revenue else 0

        # ── Profitability ──
        row = _write_section_row(ws, row, "PROFITABILITY", 4)
        kpis = [
            ("Revenue", f"₾{revenue:,.0f}", "—", "Top-line income"),
            ("Gross Margin", f"{gross_margin:.1f}%", "Good" if gross_margin > 15 else "OK" if gross_margin > 8 else "Bad", "Target: >15%"),
            ("EBITDA", f"₾{ebitda:,.0f}", "Good" if ebitda > 0 else "Bad", f"Margin: {ebitda_margin:.1f}%"),
            ("EBITDA Margin", f"{ebitda_margin:.1f}%", "Good" if ebitda_margin > 10 else "OK" if ebitda_margin > 0 else "Bad", "Target: >10%"),
            ("Net Profit", f"₾{net_profit:,.0f}", "Good" if net_profit > 0 else "Bad", "After all expenses"),
            ("Net Margin", f"{net_margin:.1f}%", "Good" if net_margin > 5 else "OK" if net_margin > 0 else "Bad", "Target: >5%"),
        ]

        for label, value, status, note in kpis:
            ws.cell(row=row, column=1, value=label).font = FN_LABEL
            ws.cell(row=row, column=2, value=value).font = FN_MONEY_BOLD
            ws.cell(row=row, column=2).alignment = AL_RIGHT
            _status_cell(ws, row, 3, status)
            ws.cell(row=row, column=4, value=note).font = FN_SMALL_MUTED
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = BORDER_BOTTOM
            row += 1

        # ── Balance Sheet Ratios ──
        if bs:
            row += 1
            row = _write_section_row(ws, row, "BALANCE SHEET & LIQUIDITY", 4)

            de = bs.get("debt_to_equity") or 0
            cr = bs.get("current_ratio") or 0
            total_assets = bs.get("total_assets", 0)
            roa = net_profit / total_assets * 100 if total_assets else 0

            bs_kpis = [
                ("Total Assets", f"₾{total_assets:,.0f}", "—", "Statement of financial position"),
                ("Debt to Equity", f"{de:.2f}x", "Good" if de < 2 else "OK" if de < 3 else "Bad", "Target: <2.0x"),
                ("Current Ratio", f"{cr:.2f}x", "Good" if cr > 1.5 else "OK" if cr > 1 else "Bad", "Target: >1.5x"),
                ("Return on Assets", f"{roa:.1f}%", "Good" if roa > 5 else "OK" if roa > 0 else "Bad", f"Net Profit / Total Assets"),
            ]

            for label, value, status, note in bs_kpis:
                ws.cell(row=row, column=1, value=label).font = FN_LABEL
                ws.cell(row=row, column=2, value=value).font = FN_MONEY_BOLD
                ws.cell(row=row, column=2).alignment = AL_RIGHT
                _status_cell(ws, row, 3, status)
                ws.cell(row=row, column=4, value=note).font = FN_SMALL_MUTED
                for c in range(1, 5):
                    ws.cell(row=row, column=c).border = BORDER_BOTTOM
                row += 1

        # ── Cost Efficiency ──
        row += 1
        row = _write_section_row(ws, row, "COST EFFICIENCY", 4)
        cogs = abs(pnl.get("cogs", pnl.get("total_cogs", 0)))
        cogs_pct = cogs / revenue * 100 if revenue else 0
        opex = abs(pnl.get("selling_expenses", 0)) + abs(pnl.get("admin_expenses", pnl.get("ga_expenses", 0)))
        opex_pct = opex / revenue * 100 if revenue else 0

        cost_kpis = [
            ("COGS", f"₾{cogs:,.0f}", "OK" if cogs_pct < 80 else "Bad", f"{cogs_pct:.1f}% of revenue"),
            ("Operating Expenses", f"₾{opex:,.0f}", "Good" if opex_pct < 15 else "OK" if opex_pct < 25 else "Bad", f"{opex_pct:.1f}% of revenue"),
        ]
        for label, value, status, note in cost_kpis:
            ws.cell(row=row, column=1, value=label).font = FN_LABEL
            ws.cell(row=row, column=2, value=value).font = FN_MONEY_BOLD
            ws.cell(row=row, column=2).alignment = AL_RIGHT
            _status_cell(ws, row, 3, status)
            ws.cell(row=row, column=4, value=note).font = FN_SMALL_MUTED
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = BORDER_BOTTOM
            row += 1

        ws.print_area = f"A1:D{row}"


excel_exporter = ExcelExporter()
