"""
FinAI LangGraph Nodes — Each node is an agent function.
Nodes are pure functions that take state and return state updates.

STRICT RULES:
- Only reasoner_node uses LLM. All others are deterministic.
- CalcAgent uses Decimal for all financial math.
- No node ever fabricates financial numbers.
"""

import logging
import time
from decimal import Decimal
from typing import Any, Dict

from app.graph.state import FinAIState

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════
# NODE 1: DATA EXTRACTOR (parses uploaded file)
# ═══════════════════════════════════════════════════════════════════

def data_extractor_node(state: FinAIState) -> Dict[str, Any]:
    """Extract financial data from uploaded file using MultiSheetAnalyzer."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    file_path = state.get("file_path")
    if not file_path:
        # No file — use existing financials from DataStore
        from app.services.data_store import data_store
        company_id = state.get("company_id", 0)
        periods = data_store.get_all_periods(company_id) if company_id else []
        if periods:
            fin = data_store.get_financials(company_id, periods[-1])
            trace.append(f"[DataExtractor] Loaded {len(fin)} fields from DataStore (period: {periods[-1]})")
            stages.append("data_extraction")
            return {
                "extracted_financials": fin,
                "period": periods[-1],
                "data_type": "full_financials" if fin.get("revenue") and fin.get("cogs") else "partial",
                "status": "extracted",
                "stages_completed": stages,
                "reasoning_trace": trace,
            }
        trace.append("[DataExtractor] No file and no stored data")
        return {"data_type": "empty", "status": "no_data", "reasoning_trace": trace}

    try:
        from app.services.multi_sheet_analyzer import multi_sheet_analyzer
        from app.services.data_validator import data_validator

        extracted = multi_sheet_analyzer.analyze_file(file_path)
        validation = data_validator.validate(extracted.current_financials)

        has_rev = bool(validation.corrected_data.get("revenue"))
        has_cogs = bool(validation.corrected_data.get("cogs"))
        has_opex = bool(validation.corrected_data.get("selling_expenses") or validation.corrected_data.get("admin_expenses"))

        if has_rev and has_cogs:
            dtype = "full_financials"
        elif has_rev:
            dtype = "revenue_only"
        elif has_opex:
            dtype = "expenses_only"
        else:
            dtype = "unknown"

        trace.append(f"[DataExtractor] Parsed {len(extracted.sheet_analyses)} sheets, "
                     f"type={dtype}, {len(extracted.revenue_breakdown)} revenue items, "
                     f"{len(extracted.cogs_breakdown)} COGS items")
        stages.append("data_extraction")

        return {
            "extracted_financials": validation.corrected_data,
            "line_items": [
                {"code": item.get("code", ""), "name": item.get("product", item.get("name", "")),
                 "amount": item.get("amount", item.get("net_revenue", 0)),
                 "category": item.get("category", ""), "subcategory": item.get("subcategory", "")}
                for item in (extracted.revenue_breakdown + extracted.cogs_breakdown)
            ],
            "balance_sheet": extracted.balance_sheet or {},
            "revenue_breakdown": extracted.revenue_breakdown,
            "cogs_breakdown": extracted.cogs_breakdown,
            "company_name": extracted.company_name or state.get("company_name", "Unknown"),
            "period": extracted.period or state.get("period", "Unknown"),
            "data_type": dtype,
            "status": "extracted",
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[DataExtractor] FAILED: {e}")
        return {"data_type": "error", "status": "extraction_failed",
                "stages_failed": list(state.get("stages_failed", [])) + ["data_extraction"],
                "reasoning_trace": trace}


# ═══════════════════════════════════════════════════════════════════
# NODE 2: CALCULATOR (pure Decimal math)
# ═══════════════════════════════════════════════════════════════════

def calculator_node(state: FinAIState) -> Dict[str, Any]:
    """Deterministic financial calculations. NEVER uses LLM."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))
    fin = state.get("extracted_financials", {})

    rev = Decimal(str(fin.get("revenue", 0) or 0))
    cogs = Decimal(str(fin.get("cogs", 0) or 0))
    selling = Decimal(str(fin.get("selling_expenses", 0) or 0))
    admin = Decimal(str(fin.get("admin_expenses", 0) or 0))
    ga = Decimal(str(fin.get("ga_expenses", 0) or 0))
    other_inc = Decimal(str(fin.get("other_income", 0) or 0))
    other_exp = Decimal(str(fin.get("other_expense", 0) or 0))
    total_opex = selling + admin + ga

    metrics = {"total_opex": float(total_opex)}

    if rev > 0:
        gp = rev - cogs
        ebitda = gp - total_opex
        net = ebitda + other_inc - other_exp
        metrics.update({
            "revenue": float(rev),
            "cogs": float(cogs),
            "gross_profit": float(gp),
            "gross_margin_pct": float(round(gp / rev * 100, 2)),
            "cogs_to_revenue_pct": float(round(cogs / rev * 100, 2)),
            "ebitda": float(ebitda),
            "ebitda_margin_pct": float(round(ebitda / rev * 100, 2)),
            "net_profit": float(net),
            "net_margin_pct": float(round(net / rev * 100, 2)),
        })
        if total_opex > 0:
            metrics["opex_to_revenue_pct"] = float(round(total_opex / rev * 100, 2))

    if total_opex > 0 and other_exp > 0:
        metrics["interest_to_total_costs_pct"] = float(round(other_exp / (total_opex + other_exp) * 100, 2))

    # BS ratios
    bs = state.get("balance_sheet", {})
    ca = Decimal(str(bs.get("total_current_assets", 0) or 0))
    cl = Decimal(str(bs.get("total_current_liabilities", 0) or 0))
    equity = Decimal(str(bs.get("total_equity", 0) or 0))
    total_liab = Decimal(str(bs.get("total_liabilities", 0) or 0))

    if cl > 0:
        metrics["current_ratio"] = float(round(ca / cl, 2))
    if equity > 0:
        metrics["debt_to_equity"] = float(round(total_liab / equity, 2))
    metrics["working_capital"] = float(ca - cl)

    count = len([v for v in metrics.values() if v != 0])
    trace.append(f"[Calculator] Computed {count} metrics (Decimal precision)")
    stages.append("calculation")

    return {
        "calculated_metrics": metrics,
        "status": "calculated",
        "stages_completed": stages,
        "reasoning_trace": trace,
    }


# ═══════════════════════════════════════════════════════════════════
# NODE 3: INSIGHT ENGINE (reconstruction + company profiling)
# ═══════════════════════════════════════════════════════════════════

def insight_engine_node(state: FinAIState) -> Dict[str, Any]:
    """Financial reconstruction, weighted company profiling, cross-signal reasoning."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    try:
        from app.services.financial_reconstruction import reconstruction_engine

        fin = state.get("extracted_financials", {})
        items = state.get("line_items", [])
        result = reconstruction_engine.reconstruct(fin, items)

        trace.append(f"[InsightEngine] {result.completeness.data_type}, "
                     f"{len(result.insights)} insights, "
                     f"industry={result.company_character.industry} "
                     f"({result.company_character.industry_confidence*100:.0f}%)")
        stages.append("insight_engine")

        return {
            "completeness": result.completeness.to_dict(),
            "insights": [i.to_dict() for i in result.insights],
            "company_character": result.company_character.to_dict(),
            "suggestions": result.suggestions,
            "revenue_estimate": result.revenue_estimate.to_dict() if result.revenue_estimate else None,
            "status": "analyzed",
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[InsightEngine] FAILED: {e}")
        return {"stages_failed": list(state.get("stages_failed", [])) + ["insight_engine"],
                "reasoning_trace": trace}


# ═══════════════════════════════════════════════════════════════════
# NODE 4: MEMORY (cross-period comparison)
# ═══════════════════════════════════════════════════════════════════

def memory_node(state: FinAIState) -> Dict[str, Any]:
    """Inject cross-period context from CompanyMemory."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    try:
        from app.memory.company_memory import company_memory

        company_id = state.get("company_id", 0)
        period = state.get("period", "")
        previous = company_memory.get_previous_period(company_id, period)
        deltas = company_memory.compute_deltas(state.get("extracted_financials", {}), previous)

        if deltas.get("has_comparison"):
            n = len(deltas.get("changes", {}))
            trace.append(f"[Memory] Compared with {deltas['previous_period']}: {n} metrics tracked")
        else:
            trace.append("[Memory] No previous period for comparison")
        stages.append("memory")

        return {
            "period_deltas": deltas,
            "previous_periods": company_memory.get_last_n(company_id, 5),
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[Memory] FAILED: {e}")
        return {"reasoning_trace": trace}


# ═══════════════════════════════════════════════════════════════════
# NODE 5: LLM REASONER (the ONLY node that uses LLM)
# ═══════════════════════════════════════════════════════════════════

def reasoner_node(state: FinAIState) -> Dict[str, Any]:
    """CFO-level reasoning using LLM chain. NEVER generates numbers."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    try:
        import asyncio
        from app.services.llm_chain import llm_chain

        context = {
            "company": state.get("company_name", "Unknown"),
            "period": state.get("period", ""),
            "data_type": state.get("data_type", "unknown"),
            "pnl": state.get("calculated_metrics", {}),
            "balance_sheet": state.get("balance_sheet", {}),
            "company_character": state.get("company_character", {}),
            "expense_breakdown": {item.get("subcategory", item.get("category", "Other")): item.get("amount", 0)
                                  for item in state.get("line_items", []) if item.get("amount", 0) > 0},
            "insights": state.get("insights", []),
            "missing_data": state.get("completeness", {}).get("missing_for_pl", []),
            "period_deltas": state.get("period_deltas", {}).get("changes", {}),
        }

        # Try Claude first (best quality), then Ollama (free fallback)
        import httpx
        import os
        prompt = llm_chain._build_prompt(context)
        system_msg = "You are a world-class CFO and Big4 audit partner. Analyze the financial data. Be precise and concise. Output JSON with keys: summary (string), confidence (float 0-1). Never invent numbers."

        # Tier 1: Claude API
        api_key = os.getenv("ANTHROPIC_API_KEY", "")
        if not api_key:
            try:
                from app.config import settings
                api_key = str(settings.ANTHROPIC_API_KEY or "")
            except Exception:
                pass

        if api_key and len(api_key) > 20:
            try:
                claude_model = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-20250514")
                with httpx.Client(timeout=30) as client:
                    r = client.post(
                        "https://api.anthropic.com/v1/messages",
                        headers={
                            "x-api-key": api_key,
                            "anthropic-version": "2023-06-01",
                            "content-type": "application/json",
                        },
                        json={
                            "model": claude_model,
                            "max_tokens": 2048,
                            "system": system_msg,
                            "messages": [{"role": "user", "content": prompt}],
                        },
                    )
                    if r.status_code == 200:
                        data = r.json()
                        text = data.get("content", [{}])[0].get("text", "")
                        import json as _json
                        import re as _re
                        # Parse response
                        result = None
                        for attempt in [
                            lambda: _json.loads(text.strip()),
                            lambda: _json.loads(_re.search(r'```(?:json)?\s*([\s\S]*?)```', text).group(1)),
                            lambda: _json.loads(_re.search(r'\{[\s\S]*\}', text).group()),
                        ]:
                            try:
                                result = attempt()
                                break
                            except Exception:
                                pass
                        if not result:
                            result = {"summary": text[:500], "confidence": 0.85}
                        if isinstance(result.get("summary"), dict):
                            result["summary"] = ". ".join(f"{k}: {v}" for k, v in result["summary"].items() if v)
                        model = f"Claude ({claude_model})"
                        trace.append(f"[LLMReasoner] Model: {model}, Confidence: {result.get('confidence', '?')}")
                        stages.append("llm_reasoning")
                        return {
                            "llm_narrative": result.get("summary", ""),
                            "llm_model_used": model,
                            "llm_confidence": result.get("confidence", 0.85),
                            "status": "reasoned",
                            "stages_completed": stages,
                            "reasoning_trace": trace,
                        }
            except Exception as claude_err:
                trace.append(f"[LLMReasoner] Claude failed: {str(claude_err)[:80]}, trying Ollama")

        # Tier 2: Ollama (free fallback)
        try:
            with httpx.Client(timeout=60) as client:
                r = client.post("http://localhost:11434/api/chat", json={
                    "model": "qwen2.5:3b",
                    "messages": [
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": prompt},
                    ],
                    "stream": False,
                })
                data = r.json()
                text = data.get("message", {}).get("content", "")
                # Parse response — handle JSON, markdown code blocks, or plain text
                import json as _json
                import re as _re
                clean_text = text.strip()
                result = None
                # Try direct JSON
                try:
                    result = _json.loads(clean_text)
                except _json.JSONDecodeError:
                    pass
                # Try extracting from markdown code block
                if not result:
                    json_match = _re.search(r'```(?:json)?\s*([\s\S]*?)```', clean_text)
                    if json_match:
                        try:
                            result = _json.loads(json_match.group(1))
                        except _json.JSONDecodeError:
                            pass
                # Try finding JSON object
                if not result:
                    brace_match = _re.search(r'\{[\s\S]*\}', clean_text)
                    if brace_match:
                        try:
                            result = _json.loads(brace_match.group())
                        except _json.JSONDecodeError:
                            pass
                # Extract summary from nested JSON
                if result and isinstance(result, dict):
                    summary = result.get("summary", "")
                    # If summary is itself a dict, flatten it
                    if isinstance(summary, dict):
                        parts = []
                        for k, v in summary.items():
                            if v and v is not None:
                                parts.append(f"{k.replace('_',' ').title()}: {v}" if isinstance(v, str) else str(v))
                        result["summary"] = ". ".join(parts) if parts else str(summary)
                else:
                    # Plain text fallback — use the raw text as summary
                    result = {"summary": clean_text[:500], "confidence": 0.7}
                model = "Nemotron 3 Nano (NVIDIA)"
        except Exception as ollama_err:
            # Template fallback
            result = llm_chain._template_fallback(context)
            model = "Template (deterministic)"

        trace.append(f"[LLMReasoner] Model: {model}, Confidence: {result.get('confidence', '?')}")
        stages.append("llm_reasoning")

        return {
            "llm_narrative": result.get("summary", ""),
            "llm_model_used": model,
            "llm_confidence": result.get("confidence", 0),
            "status": "reasoned",
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[LLMReasoner] FAILED: {e}")
        return {
            "llm_narrative": "LLM reasoning unavailable.",
            "llm_model_used": "none",
            "llm_confidence": 0,
            "stages_failed": list(state.get("stages_failed", [])) + ["llm_reasoning"],
            "reasoning_trace": trace,
        }


# ═══════════════════════════════════════════════════════════════════
# NODE 6: ORCHESTRATOR (legacy 7-stage pipeline — only if full data)
# ═══════════════════════════════════════════════════════════════════

def orchestrator_node(state: FinAIState) -> Dict[str, Any]:
    """Run the existing 7-stage pipeline if full data is available."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    if state.get("data_type") not in ("full_financials", "basic_pl"):
        trace.append("[Orchestrator] SKIPPED — incomplete data")
        return {"reasoning_trace": trace}

    try:
        from app.services.orchestrator import orchestrator

        fin = state.get("extracted_financials", {})
        bs = state.get("balance_sheet")

        result = orchestrator.run(
            current_financials=fin,
            balance_sheet=bs or None,
            industry_id="fuel_distribution",
            monte_carlo_iterations=100,
        )

        trace.append(f"[Orchestrator] Health={result.health_score:.0f}/100 ({result.health_grade}), "
                     f"Strategy={result.strategy_name}")
        stages.append("orchestrator_7stage")

        return {
            "orchestrator_result": result.to_dict(),
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[Orchestrator] FAILED: {e}")
        return {"stages_failed": list(state.get("stages_failed", [])) + ["orchestrator"],
                "reasoning_trace": trace}


# ═══════════════════════════════════════════════════════════════════
# NODE 7: ALERT GENERATOR
# ═══════════════════════════════════════════════════════════════════

def alert_node(state: FinAIState) -> Dict[str, Any]:
    """Generate alerts from current data."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))
    alerts = []

    fin = state.get("extracted_financials", {})
    metrics = state.get("calculated_metrics", {})

    if not fin.get("revenue"):
        alerts.append({"severity": "critical", "type": "missing_data",
                       "title": "Revenue Missing", "message": "Upload revenue data for profitability analysis."})

    interest_pct = metrics.get("interest_to_total_costs_pct", 0)
    if interest_pct > 25:
        alerts.append({"severity": "critical", "type": "leverage",
                       "title": "Critical Leverage", "message": f"Interest is {interest_pct:.0f}% of total costs."})

    margin = metrics.get("gross_margin_pct", 100)
    if margin < 10 and fin.get("revenue"):
        alerts.append({"severity": "critical", "type": "profitability",
                       "title": "Thin Margins", "message": f"Gross margin only {margin:.1f}%."})

    trace.append(f"[AlertGenerator] {len(alerts)} alerts")
    stages.append("alerts")

    return {
        "alerts": alerts,
        "status": "done",
        "stages_completed": stages,
        "reasoning_trace": trace,
    }


# ═══════════════════════════════════════════════════════════════════
# NODE 8: ANOMALY DETECTOR
# ═══════════════════════════════════════════════════════════════════

def anomaly_detector_node(state: FinAIState) -> Dict[str, Any]:
    """Detect financial anomalies and outliers."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))
    anomalies = []
    metrics = state.get("calculated_metrics", {})

    # Margin anomalies
    margin = metrics.get("gross_margin_pct", 0)
    if margin < 0:
        anomalies.append({"severity": "critical", "type": "margin",
                          "title": "Negative Gross Margin", "detail": f"Margin is {margin:.1f}% — company is selling below cost"})
    elif margin < 10 and metrics.get("revenue", 0) > 0:
        anomalies.append({"severity": "warning", "type": "margin",
                          "title": "Critically Thin Margins", "detail": f"Margin {margin:.1f}% — minimal buffer for cost increases"})

    # Revenue concentration
    rev_items = state.get("revenue_breakdown", [])
    if len(rev_items) > 3:
        sorted_items = sorted(rev_items, key=lambda x: x.get("net_revenue", 0), reverse=True)
        top3_total = sum(i.get("net_revenue", 0) for i in sorted_items[:3])
        all_total = sum(i.get("net_revenue", 0) for i in sorted_items) or 1
        concentration = top3_total / all_total * 100
        if concentration > 80:
            anomalies.append({"severity": "warning", "type": "concentration",
                              "title": "Revenue Concentration Risk",
                              "detail": f"Top 3 products = {concentration:.0f}% of revenue — high dependency"})

    # Cross-period delta anomalies
    deltas = state.get("period_deltas", {}).get("changes", {})
    for key, delta in deltas.items():
        if isinstance(delta, dict) and abs(delta.get("change_pct", 0)) > 50:
            anomalies.append({"severity": "warning", "type": "volatility",
                              "title": f"Large {key} change",
                              "detail": f"{key} changed {delta['change_pct']:+.1f}% vs previous period"})

    trace.append(f"[AnomalyDetector] {len(anomalies)} anomalies found")
    stages.append("anomaly_detection")

    return {
        "anomalies": anomalies,
        "stages_completed": stages,
        "reasoning_trace": trace,
    }


# ═══════════════════════════════════════════════════════════════════
# NODE 9: WHAT-IF SIMULATOR
# ═══════════════════════════════════════════════════════════════════

def whatif_simulator_node(state: FinAIState) -> Dict[str, Any]:
    """Run what-if scenarios: revenue +/-10%, +/-20%, COGS +5%."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))
    metrics = state.get("calculated_metrics", {})

    rev = Decimal(str(metrics.get("revenue", 0) or 0))
    cogs = Decimal(str(metrics.get("cogs", 0) or 0))
    opex = Decimal(str(metrics.get("total_opex", 0) or 0))

    scenarios = {}
    if rev > 0:
        for pct in [-20, -10, 10, 20]:
            new_rev = rev * (1 + Decimal(str(pct)) / 100)
            new_gp = new_rev - cogs
            new_net = new_gp - opex
            scenarios[f"revenue_{'+' if pct > 0 else ''}{pct}pct"] = {
                "revenue": float(new_rev),
                "gross_profit": float(new_gp),
                "gross_margin_pct": float(round(new_gp / new_rev * 100, 2)) if new_rev > 0 else 0,
                "net_profit": float(new_net),
                "impact": f"{'Gain' if pct > 0 else 'Loss'} of {abs(float(new_rev - rev)):,.0f} GEL",
            }

        # COGS +5% scenario
        new_cogs = cogs * Decimal("1.05")
        new_gp = rev - new_cogs
        scenarios["cogs_plus_5pct"] = {
            "cogs": float(new_cogs),
            "gross_profit": float(new_gp),
            "gross_margin_pct": float(round(new_gp / rev * 100, 2)),
            "impact": f"Margin drops from {float(rev - cogs) / float(rev) * 100:.1f}% to {float(new_gp / rev * 100):.1f}%",
        }

    trace.append(f"[WhatIfSimulator] {len(scenarios)} scenarios computed")
    stages.append("whatif_simulation")

    return {
        "whatif_scenarios": scenarios,
        "stages_completed": stages,
        "reasoning_trace": trace,
    }


# ═══════════════════════════════════════════════════════════════════
# NODE 10: REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════

def report_generator_node(state: FinAIState) -> Dict[str, Any]:
    """Generate Excel report with charts and formulas."""
    trace = list(state.get("reasoning_trace", []))
    stages = list(state.get("stages_completed", []))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.chart import BarChart, Reference
        from pathlib import Path
        from datetime import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "Executive Summary"

        company = state.get("company_name", "Unknown")
        period = state.get("period", "")
        metrics = state.get("calculated_metrics", {})

        # Header
        ws["A1"] = f"{company} — Financial Intelligence Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")
        ws["A2"] = f"Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(color="666666", size=10)

        # KPIs
        row = 4
        kpi_data = [
            ("Revenue", metrics.get("revenue", 0)),
            ("COGS", metrics.get("cogs", 0)),
            ("Gross Profit", metrics.get("gross_profit", 0)),
            ("Gross Margin %", metrics.get("gross_margin_pct", 0)),
            ("EBITDA", metrics.get("ebitda", 0)),
            ("Net Profit", metrics.get("net_profit", 0)),
            ("Net Margin %", metrics.get("net_margin_pct", 0)),
        ]
        ws.cell(row=row, column=1, value="Metric").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Amount (GEL)").font = Font(bold=True)
        for label, val in kpi_data:
            row += 1
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=round(val, 2) if val else 0)
            ws.cell(row=row, column=2).number_format = '#,##0.00'

        # Revenue Breakdown sheet
        rev_items = state.get("revenue_breakdown", [])
        if rev_items:
            ws_rev = wb.create_sheet("Revenue Breakdown")
            ws_rev.append(["Product", "Net Revenue (GEL)", "Category"])
            for item in rev_items:
                ws_rev.append([item.get("product", ""), item.get("net_revenue", 0), item.get("category", "")])

            # Bar chart
            chart = BarChart()
            chart.title = "Revenue by Product (Top 10)"
            chart.y_axis.title = "GEL"
            data = Reference(ws_rev, min_col=2, min_row=1, max_row=min(11, len(rev_items) + 1))
            cats = Reference(ws_rev, min_col=1, min_row=2, max_row=min(11, len(rev_items) + 1))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 25
            chart.height = 15
            ws_rev.add_chart(chart, "E2")

        # COGS Breakdown sheet
        cogs_items = state.get("cogs_breakdown", [])
        if cogs_items:
            ws_cogs = wb.create_sheet("COGS Breakdown")
            ws_cogs.append(["Product", "Amount (GEL)"])
            for item in cogs_items:
                ws_cogs.append([item.get("product", ""), item.get("amount", 0)])

        # Insights sheet
        insights = state.get("insights", [])
        if insights:
            ws_ins = wb.create_sheet("Intelligence")
            ws_ins.append(["Severity", "Title", "Explanation", "Action"])
            for i in insights:
                ws_ins.append([i.get("severity", ""), i.get("title", ""),
                              i.get("explanation", "")[:200], i.get("action", "")])

        # Save
        folder = Path("reports")
        folder.mkdir(exist_ok=True)
        fname = f"FinAI_{company.replace(' ', '_')}_{period}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        path = str(folder / fname)
        wb.save(path)

        trace.append(f"[ReportGenerator] Excel saved: {path}")
        stages.append("report_generation")

        return {
            "report_path": path,
            "stages_completed": stages,
            "reasoning_trace": trace,
        }
    except Exception as e:
        trace.append(f"[ReportGenerator] FAILED: {e}")
        return {"stages_failed": list(state.get("stages_failed", [])) + ["report_generation"],
                "reasoning_trace": trace}
