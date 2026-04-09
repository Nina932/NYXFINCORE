"""
FinAI Agent Services Sub-Router
================================
Extracted from agent.py to reduce file size.
Covers: currency, workflow, activity, documents, company360,
causal graph, alert resolution, translation, flywheel,
email reports, advanced AI, data agent, AIP evals.
"""
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any
from app.database import get_db
from app.config import settings
import logging

logger = logging.getLogger(__name__)
router = APIRouter()  # NO prefix — parent adds /api/agent


# ═══════════════════════════════════════════════════════════════════
# CURRENCY ENGINE API
# ═══════════════════════════════════════════════════════════════════

@router.get("/currency/rates")
async def currency_rates(date: Optional[str] = None, fetch_live: bool = False):
    """Get current exchange rates (all currencies vs GEL).

    Query params:
        date: YYYY-MM-DD for historical rates
        fetch_live: if true, fetch fresh rates from NBG API
    """
    try:
        from app.services.currency_engine import currency_engine

        if fetch_live:
            nbg_rates = await currency_engine.fetch_rates_nbg(date)
            return {
                "base_currency": "GEL",
                "source": "nbg",
                "date": date or "current",
                "rates": nbg_rates,
                "supported_currencies": currency_engine.get_supported_currencies(),
            }

        return {
            "base_currency": "GEL",
            "source": "default",
            "date": date or "current",
            "rates": currency_engine.get_all_rates(),
            "supported_currencies": currency_engine.get_supported_currencies(),
        }
    except Exception as e:
        return {"error": str(e)}


@router.post("/currency/convert")
async def currency_convert(body: dict):
    """Convert amount between currencies.

    Body: { "amount": 1000, "from": "USD", "to": "GEL", "date": "2025-01-15" }
    """
    try:
        from app.services.currency_engine import currency_engine

        amount = float(body.get("amount", 0))
        from_ccy = body.get("from", "USD").upper()
        to_ccy = body.get("to", "GEL").upper()
        date = body.get("date")

        if amount == 0:
            return {"error": "amount must be non-zero"}

        converted = currency_engine.convert(amount, from_ccy, to_ccy, date)
        rate = currency_engine.get_rate(from_ccy, to_ccy, date)

        return {
            "original_amount": amount,
            "from_currency": from_ccy,
            "to_currency": to_ccy,
            "rate": float(rate),
            "converted_amount": float(converted),
            "date": date or "current",
        }
    except ValueError as e:
        return {"error": str(e)}
    except Exception as e:
        return {"error": str(e)}


@router.post("/currency/fetch-nbg")
async def currency_fetch_nbg(body: dict = {}):
    """Fetch fresh rates from National Bank of Georgia API.

    Body (optional): { "date": "2025-01-15" }
    """
    try:
        from app.services.currency_engine import currency_engine
        date = body.get("date")
        rates = await currency_engine.fetch_rates_nbg(date)
        return {
            "source": "nbg",
            "date": date or "today",
            "rates": rates,
            "supported_currencies": currency_engine.get_supported_currencies(),
        }
    except Exception as e:
        return {"error": str(e)}


@router.post("/currency/translate-ias21")
async def currency_translate_ias21(body: dict):
    """IAS 21 currency translation for consolidation.

    Body: {
        "amounts": {"assets": 1000000, "liabilities": 600000, "equity": 400000, "revenue": 500000},
        "from_currency": "USD",
        "to_currency": "GEL",
        "closing_rate": 2.75,
        "average_rate": 2.70,
        "historical_rate": 2.60
    }
    """
    try:
        from app.services.currency_engine import currency_engine

        amounts = body.get("amounts", {})
        from_ccy = body.get("from_currency", "USD").upper()
        to_ccy = body.get("to_currency", "GEL").upper()

        result = currency_engine.translate_ias21(
            amounts=amounts,
            from_currency=from_ccy,
            to_currency=to_ccy,
            closing_rate=body.get("closing_rate"),
            average_rate=body.get("average_rate"),
            historical_rate=body.get("historical_rate"),
        )
        return result
    except Exception as e:
        return {"error": str(e)}


# =============================================================================
# DATA CONNECTORS — Georgian market connectors (NBG, rs.ge, REST, SFTP)
# =============================================================================

# NOTE: GET /connectors is already in agent.py — not duplicated here.

@router.post("/connectors/nbg/fetch")
async def fetch_nbg_rates(currencies: str = "USD,EUR,GBP,TRY,RUB,AZN"):
    """Fetch live exchange rates from the National Bank of Georgia.

    Returns official NBG rates for the specified currencies (comma-separated).
    Rates are cached for 1 hour.
    """
    from app.services.data_connectors import NBGConnector, ConnectorConfig

    config = ConnectorConfig(connector_type="nbg", name="nbg_live")
    connector = NBGConnector(config)
    result = await connector.fetch_data(
        data_type="exchange_rates",
        params={"currencies": currencies},
    )
    return result.to_dict()


@router.get("/connectors/rs/company/{tax_id}")
async def lookup_georgian_company(tax_id: str):
    """Look up a Georgian company by tax identification number on rs.ge.

    Returns company registration info from the Georgian Revenue Service.
    """
    from app.services.data_connectors import RevenueServiceConnector, ConnectorConfig

    config = ConnectorConfig(connector_type="rs_ge", name="rs_ge_lookup")
    connector = RevenueServiceConnector(config)
    result = await connector.fetch_data(
        data_type="company_lookup",
        params={"tax_id": tax_id},
    )
    return result.to_dict()


class ConnectorTestRequest(BaseModel):
    connector_type: str
    server_url: str = ""
    host: str = ""
    username: str = ""
    password: str = ""
    database: str = ""
    extra: Dict[str, Any] = {}


@router.post("/connectors/test")
async def test_connector(req: ConnectorTestRequest):
    """Test connectivity for any connector configuration.

    Supports connector types: 1c, bank_csv, sap, excel, nbg, rs_ge, rest, sftp
    """
    from app.services.data_connectors import ConnectorConfig, connector_registry, CONNECTOR_CLASSES

    if req.connector_type not in CONNECTOR_CLASSES:
        return {
            "connected": False,
            "error": f"Unknown connector type: {req.connector_type}",
            "available_types": list(CONNECTOR_CLASSES.keys()),
        }

    config = ConnectorConfig(
        connector_type=req.connector_type,
        name=f"test_{req.connector_type}",
        server_url=req.server_url,
        host=req.host,
        username=req.username,
        password=req.password,
        database=req.database,
        extra=req.extra,
    )

    cls = CONNECTOR_CLASSES[req.connector_type]
    connector = cls(config)

    try:
        result = await connector.test_connection()
        return result
    except Exception as e:
        return {"connected": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════════════
#   WORKFLOW ENGINE — Composable AI Workflows (Palantir AIP Pattern)
# ═══════════════════════════════════════════════════════════════════

@router.get("/workflows")
async def list_workflows():
    """List all registered workflows."""
    from app.services.workflow_engine import workflow_engine
    return {"workflows": workflow_engine.list_workflows()}


@router.get("/workflows/{workflow_id}")
async def get_workflow(workflow_id: str):
    """Get workflow definition with step details."""
    from app.services.workflow_engine import workflow_engine
    wf = workflow_engine.get_workflow(workflow_id)
    if not wf:
        return {"error": f"Workflow not found: {workflow_id}"}
    return {
        "workflow_id": wf.workflow_id,
        "name": wf.name,
        "description": wf.description,
        "trigger_events": wf.trigger_events,
        "steps": [
            {
                "step_id": s.step_id,
                "name": s.name,
                "type": s.step_type.value,
                "tool": s.tool_name,
                "description": s.description,
            }
            for s in wf.steps
        ],
    }


@router.post("/workflows/{workflow_id}/execute")
async def execute_workflow(workflow_id: str, body: dict = None):
    """Execute a workflow with trigger data."""
    from app.services.workflow_engine import workflow_engine
    try:
        execution = await workflow_engine.execute(workflow_id, trigger_data=body or {})
        return {
            "execution_id": execution.execution_id,
            "status": execution.status.value,
            "steps_completed": sum(1 for s in execution.step_results.values() if s.status.value == "completed"),
            "steps_total": len(workflow_engine.get_workflow(workflow_id).steps),
            "results": {
                sid: {
                    "status": sr.status.value,
                    "output": sr.output if sr.status.value == "completed" else None,
                    "error": sr.error,
                    "duration_ms": sr.duration_ms,
                }
                for sid, sr in execution.step_results.items()
            },
            "error": execution.error,
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/workflows/executions/list")
async def list_executions(workflow_id: str = None, limit: int = 20):
    """List recent workflow executions."""
    from app.services.workflow_engine import workflow_engine
    return {"executions": workflow_engine.list_executions(workflow_id, limit)}


@router.post("/workflows/executions/{execution_id}/resume")
async def resume_execution(execution_id: str, body: dict = None):
    """Resume a paused workflow after human approval."""
    from app.services.workflow_engine import workflow_engine
    result = workflow_engine.resume_execution(execution_id, body)
    if result:
        return {"resumed": True, "execution_id": result}
    return {"error": "Execution not found or not paused"}


# ═══════════════════════════════════════════════════════════════
# ACTIVITY FEED & DISTRIBUTED TRACING (Palantir AIP Observability)
# ═══════════════════════════════════════════════════════════════

@router.get("/activity/feed")
async def activity_feed_list(
    resource_type: Optional[str] = None,
    event_type: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """Activity feed with optional filters (resource_type, event_type, status)."""
    try:
        from app.services.activity_feed import activity_feed
        events = activity_feed.get_feed(
            resource_type=resource_type,
            event_type=event_type,
            status=status,
            limit=limit,
            offset=offset,
        )
        return {"events": events, "count": len(events), "offset": offset, "limit": limit}
    except Exception as e:
        return {"error": str(e), "events": []}


@router.get("/activity/trace/{trace_id}")
async def activity_trace(trace_id: str):
    """Distributed trace view — hierarchical span tree for a trace_id."""
    try:
        from app.services.activity_feed import activity_feed
        return activity_feed.get_trace(trace_id)
    except Exception as e:
        return {"error": str(e), "trace_id": trace_id, "spans": []}


@router.get("/activity/metrics")
async def activity_metrics(hours: int = 24):
    """Aggregated activity metrics per resource type for the last N hours."""
    try:
        from app.services.activity_feed import activity_feed
        return activity_feed.get_metrics(hours=hours)
    except Exception as e:
        return {"error": str(e)}


@router.get("/activity/timeline")
async def activity_timeline(hours: int = 24):
    """Hourly event counts for timeline/chart visualization."""
    try:
        from app.services.activity_feed import activity_feed
        return activity_feed.get_timeline(hours=hours)
    except Exception as e:
        return {"error": str(e)}


# [Lineage and other advanced routes moved to standalone routers]


# ═══════════════════════════════════════════════════════════════════════════
#  Document Intelligence Engine Endpoints
# ═══════════════════════════════════════════════════════════════════════════


@router.post("/documents/process")
async def process_document(
    file: UploadFile = File(...),
    document_type: Optional[str] = None,
):
    """Upload and process a document through the Document Intelligence pipeline.

    Two-stage pipeline:
      Stage 1: Text extraction (PDF, image, Excel, CSV)
      Stage 2: AI structured extraction with schema validation

    Supports: invoice, contract, bank_statement, receipt (or auto-detect).
    """
    from app.services.document_processor import document_processor

    file_bytes = await file.read()
    filename = file.filename or "unknown"

    doc = await document_processor.process_document(
        file_bytes=file_bytes,
        filename=filename,
        document_type=document_type,
    )
    return doc.to_dict()


@router.get("/documents/queue")
async def document_review_queue():
    """List documents needing human review."""
    from app.services.document_processor import document_processor

    return {
        "queue": document_processor.get_review_queue(),
        "stats": document_processor.get_stats(),
    }


@router.get("/documents/{doc_id}")
async def get_document(doc_id: str):
    """Get a processed document with extraction results."""
    from app.services.document_processor import document_processor

    doc = document_processor.get_document(doc_id)
    if not doc:
        raise HTTPException(404, f"Document {doc_id} not found")
    return doc.to_dict()


@router.post("/documents/{doc_id}/approve")
async def approve_document(doc_id: str, body: dict = None):
    """Approve extracted data for a document."""
    from app.services.document_processor import document_processor

    body = body or {}
    approved_by = body.get("approved_by", "user")
    doc = document_processor.approve_document(doc_id, approved_by=approved_by)
    if not doc:
        raise HTTPException(404, f"Document {doc_id} not found")
    return doc.to_dict()


# [AP Automation routes moved to standalone router]


# ═══════════════════════════════════════════════════════════════════
#   COMPANY 360° VIEW
# ═══════════════════════════════════════════════════════════════════

@router.get("/company/360")
@router.get("/agents/company360")
async def company_360_view(period: str = None, company_id: int = None):
    """Unified 360-degree company view from all data sources."""
    try:
        from app.services.company_360 import company_360
        from app.services.data_store import data_store

        # Find the best company (same logic as dashboard)
        if not company_id:
            companies = data_store.list_companies()
            if companies:
                best = companies[-1]
                best_score = -1
                for c in reversed(companies):
                    periods = data_store.get_all_periods(c["id"])
                    if not periods:
                        continue
                    score = len(periods)
                    fin = data_store.get_financials(c["id"], periods[-1])
                    if fin and fin.get("revenue"):
                        score += 50
                    if score > best_score:
                        best_score = score
                        best = c
                company_id = best["id"]
            else:
                company_id = 1

        result = company_360.generate(company_id=company_id, period=period)
        return result
    except Exception as e:
        logger.error("company_360 error: %s", e, exc_info=True)
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════
#   CAUSAL GRAPH
# ═══════════════════════════════════════════════════════════════════

@router.post("/causal/graph")
async def causal_graph_endpoint(body: dict):
    """Generate a visual causal graph showing financial metric causation."""
    try:
        from app.services.causal_graph import causal_graph
        financials = body.get("financials", body.get("current", {}))
        previous = body.get("previous", None)
        health_score = body.get("health_score", None)
        health_grade = body.get("health_grade", None)
        result = causal_graph.generate(
            financials=financials,
            previous=previous,
            health_score=health_score,
            health_grade=health_grade,
        )
        return result
    except Exception as e:
        logger.error("causal_graph error: %s", e, exc_info=True)
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════
#   ALERT RESOLUTION WORKFLOW
# ═══════════════════════════════════════════════════════════════════

@router.post("/alerts/{alert_id}/resolve")
async def resolve_alert_endpoint(alert_id: int, body: dict = None):
    """Record a resolution decision for an alert."""
    try:
        from app.services.alert_resolution import alert_resolution_manager
        body = body or {}
        rid = alert_resolution_manager.resolve_alert(
            alert_id=alert_id,
            decision=body.get("decision", "resolve"),
            explanation=body.get("explanation", ""),
            resolution_type=body.get("resolution_type", "root_cause_fixed"),
            resolved_by=body.get("resolved_by", "user"),
        )
        return {"resolution_id": rid, "alert_id": alert_id, "status": "resolved"}
    except Exception as e:
        logger.error("resolve_alert error: %s", e)
        return {"error": str(e)}


@router.post("/alerts/{alert_id}/escalate")
async def escalate_alert_endpoint(alert_id: int, body: dict = None):
    """Escalate an alert to higher-level review."""
    try:
        from app.services.alert_resolution import alert_resolution_manager
        body = body or {}
        rid = alert_resolution_manager.escalate_alert(
            alert_id=alert_id,
            explanation=body.get("explanation", ""),
            escalated_by=body.get("escalated_by", "user"),
        )
        return {"resolution_id": rid, "alert_id": alert_id, "status": "escalated"}
    except Exception as e:
        logger.error("escalate_alert error: %s", e)
        return {"error": str(e)}


@router.get("/alerts/resolution-stats")
async def alert_resolution_stats():
    """Get alert resolution statistics."""
    try:
        from app.services.alert_resolution import alert_resolution_manager
        stats = alert_resolution_manager.get_resolution_stats()
        return stats
    except Exception as e:
        logger.error("resolution_stats error: %s", e)
        return {"error": str(e)}


@router.get("/alerts/{alert_id}/impact")
async def alert_impact(alert_id: int):
    """Get impact tracking for a resolved alert."""
    try:
        from app.services.alert_resolution import alert_resolution_manager
        impact = alert_resolution_manager.get_alert_impact(alert_id)
        return impact
    except Exception as e:
        logger.error("alert_impact error: %s", e)
        return {"error": str(e)}


# ═══════════════════════════════════════════════════════════════════
#   TRANSLATION — Georgian content via Gemini
# ═══════════════════════════════════════════════════════════════════

@router.post("/translate")
async def translate_text(body: dict):
    """Translate text to Georgian using Gemini 2.5 Flash."""
    text = body.get("text", "")
    target_lang = body.get("target_lang", "ka")
    if not text:
        return {"translated": ""}
    if target_lang != "ka":
        return {"translated": text}

    from app.services.local_llm import translate_to_georgian
    translated = await translate_to_georgian(text)
    return {"translated": translated, "source_lang": "en", "target_lang": "ka"}


# ═══════════════════════════════════════════════════════════════════
#   DATA FLYWHEEL — Continuous AI Model Improvement
# ═══════════════════════════════════════════════════════════════════

@router.get("/flywheel/stats")
async def flywheel_stats():
    """Get data flywheel statistics and optimization opportunities."""
    from app.services.data_flywheel import data_flywheel
    return data_flywheel.get_stats()


@router.get("/flywheel/workloads")
async def flywheel_workloads():
    """Analyze workload profiles for model optimization."""
    from app.services.data_flywheel import data_flywheel
    return {"workloads": data_flywheel.classify_workloads()}


@router.get("/flywheel/training-data")
async def flywheel_training_data():
    """Get training data candidates for model fine-tuning."""
    from app.services.data_flywheel import data_flywheel
    candidates = data_flywheel.get_training_candidates()
    return {"candidates": candidates, "count": len(candidates)}


@router.post("/flywheel/feedback")
async def flywheel_feedback(body: dict):
    """Record user feedback on an AI response."""
    from app.services.data_flywheel import data_flywheel
    interaction_id = body.get("interaction_id", "")
    feedback = body.get("feedback", "")  # thumbs_up, thumbs_down, corrected
    corrected = body.get("corrected_response")
    success = data_flywheel.record_feedback(interaction_id, feedback, corrected)
    return {"recorded": success}


@router.get("/flywheel/export")
async def flywheel_export(format: str = "jsonl"):
    """Export training data in JSONL format for fine-tuning."""
    from app.services.data_flywheel import data_flywheel
    from fastapi.responses import Response
    data = data_flywheel.export_training_data(format)
    return Response(
        content=data,
        media_type="application/jsonl" if format == "jsonl" else "application/json",
        headers={"Content-Disposition": f"attachment; filename=finai_training_data.{format}"},
    )


# ── Period Aggregation Endpoints ─────────────────────────────────────────


@router.get("/agents/periods/quarterly")
async def periods_quarterly(year: int = 2025, company_id: int = 1):
    """Get quarterly aggregates (Q1-Q4) for a given year."""
    from app.services.period_aggregation import period_aggregator
    quarters = period_aggregator.aggregate_quarters(company_id, year)
    return {
        "company_id": company_id,
        "year": year,
        "quarters": [
            {
                "period_label": q.period_label,
                "periods_included": q.periods_included,
                "financials": q.financials,
            }
            for q in quarters
        ],
        "count": len(quarters),
    }


@router.get("/agents/periods/ytd")
async def periods_ytd(year: int = 2025, month: int = 12, company_id: int = 1):
    """Get year-to-date aggregate through a specific month."""
    from app.services.period_aggregation import period_aggregator
    ytd = period_aggregator.aggregate_ytd(company_id, year, month)
    return {
        "company_id": company_id,
        "year": year,
        "through_month": month,
        "period_label": ytd.period_label,
        "periods_included": ytd.periods_included,
        "financials": ytd.financials,
    }


@router.get("/agents/periods/annual")
async def periods_annual(year: int = 2025, company_id: int = 1):
    """Get full year aggregate."""
    from app.services.period_aggregation import period_aggregator
    annual = period_aggregator.aggregate_full_year(company_id, year)
    return {
        "company_id": company_id,
        "year": year,
        "period_label": annual.period_label,
        "periods_included": annual.periods_included,
        "financials": annual.financials,
    }


@router.get("/agents/periods/summary")
async def periods_summary(company_id: int = 1):
    """Get all available period aggregations for a company."""
    from app.services.period_aggregation import period_aggregator
    return period_aggregator.get_summary(company_id)


# ═══════════════════════════════════════════════════════════════════
# MODERN EMAIL REPORTS
# ═══════════════════════════════════════════════════════════════════

class EmailReportRequest(BaseModel):
    """Request model for sending email reports."""
    recipients: List[str]
    report_type: str  # "cash_runway", "pl_comparison", "bs_comparison", etc.
    company_name: str = "Company"
    period: str = "Current Period"
    custom_message: Optional[str] = None
    dataset_id: Optional[int] = None
    prior_dataset_id: Optional[int] = None


@router.post("/agents/email-report")
async def send_modern_email_report(body: EmailReportRequest, db: AsyncSession = Depends(get_db)):
    """
    Send a modern, professionally formatted Excel report via email.

    Generates a top-quality Excel report with modern styling and sends it
    as an attachment in a beautifully designed HTML email.
    """
    try:
        from app.services.modern_email import modern_email_service
        from app.services.modern_excel import modern_excel_generator
        from app.models.all_models import Dataset

        # Generate Excel report based on type
        excel_bytes = None
        filename = f"{body.report_type}_{body.company_name}_{body.period}.xlsx".replace(" ", "_")

        if body.report_type == "cash_runway":
            # Generate cash runway report from real financial data
            from app.models.all_models import Dataset, BalanceSheetItem, Transaction
            from sqlalchemy import select as sa_select, func as sa_func

            try:
                # Get active dataset
                ds_result = await db.execute(
                    sa_select(Dataset).where(Dataset.is_active == True).limit(1)
                )
                active_ds = ds_result.scalar_one_or_none()
                ds_id = body.dataset_id or (active_ds.id if active_ds else None)

                cash_balance = None
                monthly_burn = None
                data_source = "real"

                # Try to get cash from balance sheet items
                if ds_id:
                    bs_result = await db.execute(
                        sa_select(BalanceSheetItem).where(
                            BalanceSheetItem.dataset_id == ds_id,
                            BalanceSheetItem.line_item.ilike("%cash%"),
                        )
                    )
                    cash_items = bs_result.scalars().all()
                    if cash_items:
                        cash_balance = sum(
                            float(getattr(item, "amount", 0) or 0) for item in cash_items
                        )

                    # Calculate monthly burn from expense transactions
                    burn_result = await db.execute(
                        sa_select(sa_func.sum(Transaction.amount)).where(
                            Transaction.dataset_id == ds_id,
                            Transaction.type == "Expense",
                        )
                    )
                    total_expenses = burn_result.scalar() or 0
                    if total_expenses > 0:
                        monthly_burn = float(total_expenses)

                # Fall back to placeholders only if real data unavailable
                if cash_balance is None or cash_balance == 0:
                    cash_balance = 1_000_000
                    data_source = "placeholder"
                    logger.warning(
                        "Cash runway: no real cash balance found (dataset_id=%s). "
                        "Using placeholder ₾1,000,000.",
                        ds_id,
                    )
                if monthly_burn is None or monthly_burn == 0:
                    monthly_burn = 150_000
                    data_source = "placeholder"
                    logger.warning(
                        "Cash runway: no real expense data found (dataset_id=%s). "
                        "Using placeholder ₾150,000/month.",
                        ds_id,
                    )

                runway_months = round(cash_balance / monthly_burn, 1) if monthly_burn > 0 else 99
                risk_level = (
                    "critical" if runway_months < 3
                    else "high" if runway_months < 6
                    else "medium" if runway_months < 12
                    else "low"
                )

                # Generate projection data
                projection_data = []
                balance = cash_balance
                for m in range(13):  # 12 months projection
                    projection_data.append({
                        "month": m,
                        "cash": max(balance, 0),
                        "burn": monthly_burn,
                        "risk": "High" if balance <= 0 else "Medium" if balance < monthly_burn * 3 else "Low",
                        "data_source": data_source,
                    })
                    balance -= monthly_burn

                excel_bytes = modern_excel_generator.generate_cash_runway_report(
                    company_name=body.company_name,
                    period=body.period,
                    cash_balance=cash_balance,
                    monthly_burn=monthly_burn,
                    runway_months=runway_months,
                    projection_data=projection_data,
                    risk_level=risk_level
                )
            except Exception as e:
                logger.warning(f"Failed to compute cash runway: {e}, using placeholder defaults")
                excel_bytes = modern_excel_generator.generate_cash_runway_report(
                    company_name=body.company_name,
                    period=body.period,
                    cash_balance=1_000_000,
                    monthly_burn=150_000,
                    runway_months=6.7,
                    projection_data=[],
                    risk_level="medium"
                )

        elif body.report_type == "pl_comparison":
            from app.services.excel_report_bytes import generate_pl_excel_bytes
            excel_bytes, filename = await generate_pl_excel_bytes(
                body.dataset_id, body.prior_dataset_id, db
            )

        elif body.report_type == "bs_comparison":
            from app.services.excel_report_bytes import generate_bs_excel_bytes
            excel_bytes, filename = await generate_bs_excel_bytes(
                body.dataset_id, body.prior_dataset_id, db
            )

        elif body.report_type == "revenue_comparison":
            from app.services.excel_report_bytes import generate_revenue_excel_bytes
            excel_bytes, filename = await generate_revenue_excel_bytes(
                body.dataset_id, body.prior_dataset_id, db
            )

        elif body.report_type == "cogs_comparison":
            from app.services.excel_report_bytes import generate_cogs_excel_bytes
            excel_bytes, filename = await generate_cogs_excel_bytes(
                body.dataset_id, body.prior_dataset_id, db
            )

        elif body.report_type == "modern_excel":
            # Use the EXACT same MR report generator as the download button
            try:
                from app.routers.mr_reports import generate_mr_excel, MRExcelRequest
                mr_payload = MRExcelRequest(
                    company_name=body.company_name or settings.COMPANY_NAME,
                    period=body.period or "January 2026",
                    dataset_id=body.dataset_id,
                    gel_usd_rate=2.72,
                )
                mr_response = await generate_mr_excel(mr_payload, db)
                # Extract bytes from StreamingResponse
                excel_bytes = b''
                async for chunk in mr_response.body_iterator:
                    excel_bytes += chunk
                filename = f"FinAI_MR_Report_{body.company_name}_{body.period}.xlsx".replace(" ", "_")
            except Exception as e:
                logger.error(f"MR report email failed: {e}", exc_info=True)
                from app.services.excel_report_bytes import generate_pl_excel_bytes
                excel_bytes, filename = await generate_pl_excel_bytes(body.dataset_id, body.prior_dataset_id, db)
                filename = f"FinAI_Report_{body.company_name}_{body.period}.xlsx".replace(" ", "_")

        elif body.report_type == "modern_excel_DISABLED_OLD":
            # OLD CODE — kept for reference, replaced by direct MR report call above
            try:
                from app.routers.mr_reports import generate_mr_excel
                from io import BytesIO
                from fastapi.testclient import TestClient

                # Call the MR report generator directly (same code as POST /api/mr/generate-excel)
                from app.services.v2.pl_comparison import pl_comparison
                from app.routers.journal_router import _auto_find_prior
                from app.models.all_models import Dataset

                ds_id = body.dataset_id
                if not ds_id:
                    ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
                    if ds: ds_id = ds.id
                prior_id = await _auto_find_prior(ds_id, db) if ds_id else None

                # Import the actual MR Excel generation logic
                from app.services.excel_report_bytes import generate_pl_excel_bytes as _fallback
                # Generate using all 4 comparison services + full styling from mr_reports.py
                pl_data = await pl_comparison.full_pl(ds_id, prior_id, db)
                bs_data = await pl_comparison.balance_sheet_comparison(ds_id, prior_id, db)
                rev_data = await pl_comparison.revenue_comparison(ds_id, prior_id, db)
                cogs_data = await pl_comparison.cogs_comparison(ds_id, prior_id, db)

                # Use the styled MR report builder from mr_reports.py
                import importlib
                mr_mod = importlib.import_module("app.routers.mr_reports")
                # The generate_mr_excel function is async and returns StreamingResponse
                # Instead, replicate the workbook building with proper styling
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                summary = pl_data.get("summary", {})
                period_label = pl_data.get("period", body.period) or body.period
                rate = 2.72  # Default GEL/USD

                PRIMARY = "1B3A5C"; ACCENT = "2563EB"; LIGHT_BG = "F0F4F8"
                header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
                accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
                bold_font = Font(name="Calibri", bold=True, size=11)
                title_font = Font(name="Calibri", bold=True, size=14, color=PRIMARY)
                subtitle_font = Font(name="Calibri", bold=True, size=12, color=PRIMARY)
                normal_font = Font(name="Calibri", size=10)
                thin_border = Border(left=Side(style="thin", color="D0D5DD"), right=Side(style="thin", color="D0D5DD"),
                                     top=Side(style="thin", color="D0D5DD"), bottom=Side(style="thin", color="D0D5DD"))
                num_fmt = '#,##0'; pct_fmt = '0.0%'

                wb = openpyxl.Workbook()

                # Sheet 1: Executive Summary
                ws1 = wb.active; ws1.title = "Executive Summary"; ws1.sheet_properties.tabColor = PRIMARY
                ws1.merge_cells("A1:F1"); ws1.cell(row=1, column=1, value=f"MR Report — {body.company_name}").font = title_font; ws1.row_dimensions[1].height = 30
                ws1.merge_cells("A2:F2"); ws1.cell(row=2, column=1, value=f"Period: {period_label}  |  FX Rate: {rate} GEL/USD").font = subtitle_font; ws1.row_dimensions[2].height = 22
                ws1.cell(row=4, column=1, value="Key Financial Indicators").font = subtitle_font
                for ci, h in enumerate(["Metric", "GEL", "USD", "Prior Year (GEL)", "Variance (GEL)", "Var %"], 1):
                    c = ws1.cell(row=5, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = thin_border
                pl_rows_data = pl_data.get("rows", [])
                prior_rev = summary.get("prior_revenue", 0) or 0
                prior_np = summary.get("prior_net_profit", 0) or 0
                prior_cogs = next((r.get("pr", 0) for r in pl_rows_data if r.get("c") == "COGS"), 0)
                prior_gp = next((r.get("pr", 0) for r in pl_rows_data if r.get("c") == "TGP"), 0)
                prior_ebitda = next((r.get("pr", 0) for r in pl_rows_data if r.get("c") == "EBITDA"), 0)
                kpis = [("Revenue", summary.get("revenue", 0), prior_rev), ("COGS", summary.get("cogs", 0), prior_cogs),
                        ("Gross Profit", summary.get("gross_profit", 0), prior_gp), ("EBITDA", summary.get("ebitda", 0), prior_ebitda),
                        ("Net Profit", summary.get("net_profit", 0), prior_np)]
                for ri, (name, val, prior) in enumerate(kpis, 6):
                    val = float(val or 0); prior = float(prior or 0); variance = val - prior; var_pct = (variance / abs(prior)) if prior else 0
                    ws1.cell(row=ri, column=1, value=name).font = bold_font
                    ws1.cell(row=ri, column=2, value=val).number_format = num_fmt
                    ws1.cell(row=ri, column=3, value=round(val / rate, 2)).number_format = num_fmt
                    ws1.cell(row=ri, column=4, value=prior).number_format = num_fmt
                    ws1.cell(row=ri, column=5, value=variance).number_format = num_fmt
                    ws1.cell(row=ri, column=6, value=var_pct).number_format = pct_fmt
                    for col in range(1, 7): ws1.cell(row=ri, column=col).border = thin_border
                for col in ['A','B','C','D','E','F']: ws1.column_dimensions[col].width = 22

                # Sheet 2: P&L Statement
                ws2 = wb.create_sheet("P&L Statement"); ws2.sheet_properties.tabColor = ACCENT
                ws2.merge_cells("A1:G1"); ws2.cell(row=1, column=1, value=f"Profit & Loss Statement — {period_label}").font = title_font; ws2.row_dimensions[1].height = 28
                for ci, h in enumerate(["Code", "Line Item", "Prior Year", "Actual", "Plan", "Variance", "Var %"], 1):
                    c = ws2.cell(row=3, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.border = thin_border
                for ri, r in enumerate(pl_rows_data, 4):
                    ws2.cell(row=ri, column=1, value=r.get("c", "")).font = Font(name="Calibri", size=9, color="9CA3AF")
                    indent = "  " * r.get("lvl", 0)
                    ws2.cell(row=ri, column=2, value=f"{indent}{r.get('l', '')}").font = bold_font if r.get("bold") else normal_font
                    ws2.cell(row=ri, column=3, value=r.get("pr") or None); ws2.cell(row=ri, column=3).number_format = num_fmt
                    ws2.cell(row=ri, column=4, value=r.get("ac") or None); ws2.cell(row=ri, column=4).number_format = num_fmt
                    ws2.cell(row=ri, column=6, value=r.get("var") or None); ws2.cell(row=ri, column=6).number_format = num_fmt
                    ws2.cell(row=ri, column=7, value=(r.get("var_pct", 0) or 0) / 100 if r.get("var_pct") else None); ws2.cell(row=ri, column=7).number_format = pct_fmt
                    for col in range(1, 8): ws2.cell(row=ri, column=col).border = thin_border
                ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 42
                for col in ['C','D','E','F']: ws2.column_dimensions[col].width = 16
                ws2.column_dimensions['G'].width = 10; ws2.freeze_panes = "A4"

                # Sheet 3: Balance Sheet
                ws3 = wb.create_sheet("Balance Sheet"); ws3.sheet_properties.tabColor = "10B981"
                ws3.merge_cells("A1:E1"); ws3.cell(row=1, column=1, value=f"Balance Sheet — {period_label}").font = title_font
                for ci, h in enumerate(["Section", "IFRS Line Item", "Prior Year", "Actual", "Variance"], 1):
                    c = ws3.cell(row=3, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.border = thin_border
                for ri, r in enumerate(bs_data.get("rows", []), 4):
                    ws3.cell(row=ri, column=1, value=r.get("section", "")).font = normal_font
                    ws3.cell(row=ri, column=2, value=r.get("ifrs_line", "")).font = bold_font if r.get("bold") else normal_font
                    ws3.cell(row=ri, column=3, value=r.get("prior") or None); ws3.cell(row=ri, column=3).number_format = num_fmt
                    ws3.cell(row=ri, column=4, value=r.get("actual") or None); ws3.cell(row=ri, column=4).number_format = num_fmt
                    ws3.cell(row=ri, column=5, value=r.get("variance") or None); ws3.cell(row=ri, column=5).number_format = num_fmt
                    for col in range(1, 6): ws3.cell(row=ri, column=col).border = thin_border
                ws3.column_dimensions['A'].width = 22; ws3.column_dimensions['B'].width = 35
                for col in ['C','D','E']: ws3.column_dimensions[col].width = 16

                # Sheet 4: Revenue
                ws4 = wb.create_sheet("Revenue"); ws4.sheet_properties.tabColor = "3B82F6"
                ws4.merge_cells("A1:G1"); ws4.cell(row=1, column=1, value=f"Revenue by Product — {period_label}").font = title_font
                for ci, h in enumerate(["Product", "Segment", "Prior Year", "Actual (Net)", "Variance", "Var %", "% of Total"], 1):
                    c = ws4.cell(row=3, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.border = thin_border
                for ri, r in enumerate(rev_data.get("rows", []), 4):
                    ws4.cell(row=ri, column=1, value=r.get("product", "")).font = bold_font
                    ws4.cell(row=ri, column=2, value=r.get("segment", "")).font = normal_font
                    ws4.cell(row=ri, column=3, value=r.get("prior_net") or None); ws4.cell(row=ri, column=3).number_format = num_fmt
                    ws4.cell(row=ri, column=4, value=r.get("actual_net") or None); ws4.cell(row=ri, column=4).number_format = num_fmt
                    ws4.cell(row=ri, column=5, value=r.get("variance") or None); ws4.cell(row=ri, column=5).number_format = num_fmt
                    ws4.cell(row=ri, column=6, value=(r.get("variance_pct", 0) or 0) / 100 if r.get("variance_pct") else None); ws4.cell(row=ri, column=6).number_format = pct_fmt
                    ws4.cell(row=ri, column=7, value=(r.get("pct_of_total", 0) or 0) / 100 if r.get("pct_of_total") else None); ws4.cell(row=ri, column=7).number_format = pct_fmt
                    for col in range(1, 8): ws4.cell(row=ri, column=col).border = thin_border
                for col in ['A','B','C','D','E','F','G']: ws4.column_dimensions[col].width = 18

                # Sheet 5: COGS
                ws5 = wb.create_sheet("COGS"); ws5.sheet_properties.tabColor = "EF4444"
                ws5.merge_cells("A1:F1"); ws5.cell(row=1, column=1, value=f"Cost of Goods Sold — {period_label}").font = title_font
                for ci, h in enumerate(["Category", "Segment", "Prior Year", "Actual", "Variance", "Var %"], 1):
                    c = ws5.cell(row=3, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.border = thin_border
                for ri, r in enumerate(cogs_data.get("rows", []), 4):
                    ws5.cell(row=ri, column=1, value=r.get("category", "")).font = bold_font
                    ws5.cell(row=ri, column=2, value=r.get("segment", "")).font = normal_font
                    ws5.cell(row=ri, column=3, value=r.get("prior") or None); ws5.cell(row=ri, column=3).number_format = num_fmt
                    ws5.cell(row=ri, column=4, value=r.get("actual") or None); ws5.cell(row=ri, column=4).number_format = num_fmt
                    ws5.cell(row=ri, column=5, value=r.get("variance") or None); ws5.cell(row=ri, column=5).number_format = num_fmt
                    ws5.cell(row=ri, column=6, value=(r.get("variance_pct", 0) or 0) / 100 if r.get("variance_pct") else None); ws5.cell(row=ri, column=6).number_format = pct_fmt
                    for col in range(1, 7): ws5.cell(row=ri, column=col).border = thin_border
                for col in ['A','B','C','D','E','F']: ws5.column_dimensions[col].width = 18

                # Sheet 6: KPI Dashboard
                ws6 = wb.create_sheet("KPI Dashboard"); ws6.sheet_properties.tabColor = "8B5CF6"
                ws6.merge_cells("A1:D1"); ws6.cell(row=1, column=1, value=f"KPI Dashboard — {period_label}").font = title_font
                for ci, h in enumerate(["KPI", "Value", "Prior Year", "Change"], 1):
                    c = ws6.cell(row=3, column=ci, value=h); c.font = header_font; c.fill = header_fill; c.border = thin_border
                revenue = float(summary.get("revenue", 0) or 1)
                _gp = float(summary.get("gross_profit", 0) or 0); _np = float(summary.get("net_profit", 0) or 0)
                _ebitda = float(summary.get("ebitda", 0) or 0)
                kpi_rows = [
                    ("Revenue (GEL)", revenue, prior_rev, (revenue - prior_rev) / abs(prior_rev) if prior_rev else 0),
                    ("Gross Profit (GEL)", _gp, float(prior_gp or 0), (_gp - float(prior_gp or 0)) / abs(float(prior_gp or 1)) if prior_gp else 0),
                    ("EBITDA (GEL)", _ebitda, float(prior_ebitda or 0), (_ebitda - float(prior_ebitda or 0)) / abs(float(prior_ebitda or 1)) if prior_ebitda else 0),
                    ("Net Profit (GEL)", _np, float(prior_np or 0), (_np - float(prior_np or 0)) / abs(float(prior_np or 1)) if prior_np else 0),
                    ("Gross Margin %", _gp / revenue if revenue else 0, float(prior_gp or 0) / float(prior_rev or 1) if prior_rev else 0, 0),
                    ("Net Margin %", _np / revenue if revenue else 0, float(prior_np or 0) / float(prior_rev or 1) if prior_rev else 0, 0),
                ]
                for ri, (kpi, val, prior, change) in enumerate(kpi_rows, 4):
                    ws6.cell(row=ri, column=1, value=kpi).font = bold_font
                    is_pct = '%' in kpi
                    ws6.cell(row=ri, column=2, value=val); ws6.cell(row=ri, column=2).number_format = pct_fmt if is_pct else num_fmt
                    ws6.cell(row=ri, column=3, value=prior); ws6.cell(row=ri, column=3).number_format = pct_fmt if is_pct else num_fmt
                    ws6.cell(row=ri, column=4, value=change); ws6.cell(row=ri, column=4).number_format = pct_fmt
                    for col in range(1, 5): ws6.cell(row=ri, column=col).border = thin_border
                for col in ['A','B','C','D']: ws6.column_dimensions[col].width = 22

                buf = BytesIO(); wb.save(buf)
                excel_bytes = buf.getvalue()
            except Exception as e:
                logger.error(f"MR email generation failed: {e}", exc_info=True)
                from app.services.excel_report_bytes import generate_pl_excel_bytes
                excel_bytes, _ = await generate_pl_excel_bytes(body.dataset_id, body.prior_dataset_id, db)
            filename = f"FinAI_MR_Report_{body.company_name}_{body.period}.xlsx".replace(" ", "_")

        elif body.report_type in ("full_pdf", "brief_pdf"):
            # For PDF reports, still send Excel attachment (PDF email attachment is complex)
            from app.services.excel_report_bytes import generate_pl_excel_bytes
            excel_bytes, _ = await generate_pl_excel_bytes(body.dataset_id, body.prior_dataset_id, db)
            report_label = "Full_Report" if body.report_type == "full_pdf" else "Executive_Brief"
            filename = f"FinAI_{report_label}_{body.company_name}_{body.period}.xlsx".replace(" ", "_")

        else:
            raise HTTPException(400, f"Unsupported report type: {body.report_type}")

        if not excel_bytes:
            raise HTTPException(500, "Failed to generate Excel report")

        # Try to get summary data for KPI display in email
        summary_data = None
        try:
            if body.report_type in ("pl_comparison", "revenue_comparison", "cogs_comparison"):
                from app.services.v2.pl_comparison import pl_comparison
                from app.models.all_models import Dataset
                from sqlalchemy import select as sa_select
                ds_id = body.dataset_id
                if not ds_id:
                    r = await db.execute(sa_select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))
                    ds = r.scalar_one_or_none()
                    if ds: ds_id = ds.id
                if ds_id:
                    pl_data = await pl_comparison.full_pl(ds_id, body.prior_dataset_id, db)
                    summary_data = pl_data.get("summary")
        except Exception:
            pass  # Non-critical, email sends without KPIs

        # Send email with attachment
        subject = f"[FinAI] {body.report_type.replace('_', ' ').title()} — {body.period} ({body.company_name})"

        success = modern_email_service.send_report_email(
            recipients=body.recipients,
            subject=subject,
            report_type=body.report_type.replace('_', ' ').title(),
            company_name=body.company_name,
            period=body.period,
            excel_attachment=excel_bytes,
            filename=filename,
            custom_message=body.custom_message,
            summary_data=summary_data,
        )

        if success:
            return {"message": "Modern email report sent successfully", "recipients": body.recipients}
        else:
            raise HTTPException(500, "Failed to send email - check SMTP configuration")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Email report error: {e}")
        raise HTTPException(500, f"Failed to send email report: {str(e)}")


# ── ReAct Reasoning Agent ──────────────────────────────────────────

@router.post("/agents/react/solve")
async def react_solve(body: dict, db: AsyncSession = Depends(get_db)):
    """Run the ReAct reasoning loop to solve a financial analysis goal."""
    from app.agents.react_agent import react_agent
    trace = await react_agent.solve(body.get("goal", ""), body.get("context", {}), db)
    return {
        "goal": trace.goal,
        "success": trace.success,
        "steps": [
            {
                "step": s.step_num,
                "thought": s.thought,
                "action": s.action,
                "observation": s.observation,
                "is_final": s.is_final,
            }
            for s in trace.steps
        ],
        "final_answer": trace.final_answer,
        "time_ms": trace.total_time_ms,
    }


# ── Planning Agent ──────────────────────────────────────────────────

@router.post("/agents/planner/create")
async def create_plan(body: dict):
    """Create an execution plan for a financial goal without executing it."""
    from app.agents.planner import planning_agent
    plan = planning_agent.create_plan(body.get("goal", ""))
    return {
        "goal": plan.goal,
        "complexity": plan.complexity,
        "estimated_seconds": plan.estimated_time_seconds,
        "steps": [
            {
                "num": s.step_num,
                "action": s.action,
                "description": s.description,
                "endpoint": s.endpoint,
            }
            for s in plan.steps
        ],
    }


@router.post("/agents/planner/execute")
async def execute_plan(body: dict, db: AsyncSession = Depends(get_db)):
    """Create and execute a plan for a financial goal."""
    from app.agents.planner import planning_agent
    plan = planning_agent.create_plan(body.get("goal", ""))
    result = await planning_agent.execute_plan(plan, db)
    return result


# ── Flywheel Retrainer ──────────────────────────────────────────────

@router.post("/agents/flywheel/retrain")
async def flywheel_retrain(db: AsyncSession = Depends(get_db)):
    """Run a full retrain cycle across all learnable components."""
    from app.services.v2.flywheel_retrain import flywheel_retrainer
    return await flywheel_retrainer.run_retrain_cycle(db)


@router.get("/agents/flywheel/retrain-history")
async def flywheel_retrain_history():
    """Get retrain history summary."""
    from app.services.v2.flywheel_retrain import flywheel_retrainer
    return flywheel_retrainer.get_history()


# ── Data Agent (AI Query Engine) ───────────────────────────────────

PREBUILT_QUERIES = {
    "revenue_by_month": {
        "title": "Revenue by Month",
        "sql": "SELECT period, SUM(net_amount) as revenue FROM dw_revenue_items GROUP BY period ORDER BY period",
        "description": "Monthly revenue breakdown across all periods",
    },
    "top_products": {
        "title": "Top 10 Products by Revenue",
        "sql": "SELECT product, SUM(net_amount) as revenue FROM dw_revenue_items GROUP BY product ORDER BY revenue DESC LIMIT 10",
        "description": "Highest revenue-generating products",
    },
    "expense_trends": {
        "title": "Expense Trends",
        "sql": "SELECT period, SUM(CASE WHEN account_code LIKE '7%' THEN amount ELSE 0 END) as expenses FROM dw_transactions GROUP BY period ORDER BY period",
        "description": "Monthly expense trends",
    },
    "profitability": {
        "title": "Product Profitability",
        "sql": "SELECT r.product, SUM(r.net_amount) as revenue, r.category FROM dw_revenue_items r GROUP BY r.product, r.category ORDER BY revenue DESC LIMIT 15",
        "description": "Revenue by product with category",
    },
    "balance_summary": {
        "title": "Financial Snapshots",
        "sql": "SELECT period, revenue, cogs, gross_profit, net_profit FROM dw_financial_snapshots ORDER BY period DESC LIMIT 12",
        "description": "Monthly P&L summary snapshots",
    },
    "segment_analysis": {
        "title": "Revenue by Segment",
        "sql": "SELECT segment, category, SUM(net_amount) as total FROM dw_revenue_items GROUP BY segment, category ORDER BY total DESC",
        "description": "Revenue breakdown by business segment",
    },
    "data_quality": {
        "title": "Data Quality Check",
        "sql": "SELECT 'revenue_items' as source, COUNT(*) as rows FROM dw_revenue_items UNION ALL SELECT 'transactions', COUNT(*) FROM dw_transactions UNION ALL SELECT 'financial_snapshots', COUNT(*) FROM dw_financial_snapshots UNION ALL SELECT 'trial_balance', COUNT(*) FROM dw_trial_balance",
        "description": "Row counts across all warehouse tables",
    },
}

# Knowledge Graph queries (ontology tools)
KG_QUERIES = {
    "kg_accounts": {
        "title": "Browse Accounts (Knowledge Graph)",
        "type": "ontology", "tool": "search_objects",
        "params": {"type_id": "Account", "limit": 30},
        "description": "All accounts in the financial knowledge graph (489 accounts)",
    },
    "kg_kpis": {
        "title": "KPI Overview (Knowledge Graph)",
        "type": "ontology", "tool": "search_objects",
        "params": {"type_id": "KPI", "limit": 50},
        "description": "All computed financial KPIs from the ontology (74 KPIs)",
    },
    "kg_risks": {
        "title": "Risk Signals (Knowledge Graph)",
        "type": "ontology", "tool": "search_objects",
        "params": {"type_id": "RiskSignal", "limit": 20},
        "description": "AI-detected risk signals and anomalies (47 signals)",
    },
    "kg_benchmarks": {
        "title": "Industry Benchmarks (Knowledge Graph)",
        "type": "ontology", "tool": "search_objects",
        "params": {"type_id": "Benchmark", "limit": 30},
        "description": "Industry benchmark metrics for comparison (23 benchmarks)",
    },
    "kg_standards": {
        "title": "IFRS Standards (Knowledge Graph)",
        "type": "ontology", "tool": "search_objects",
        "params": {"type_id": "Standard", "limit": 20},
        "description": "IFRS/GAAP standards linked to accounts (15 standards)",
    },
    "kg_graph_stats": {
        "title": "Graph Statistics",
        "type": "graph_api", "endpoint": "/api/graph/stats",
        "description": "Knowledge graph structure: nodes, edges, density, most connected entities",
    },
    "kg_anomalies": {
        "title": "Graph Anomalies",
        "type": "graph_api", "endpoint": "/api/graph/anomalies",
        "description": "Structural anomalies detected in the knowledge graph",
    },
}

# Keyword mapping: user words -> prebuilt key
_KEYWORD_MAP = {
    "revenue": "revenue_by_month",
    "monthly revenue": "revenue_by_month",
    "revenue by month": "revenue_by_month",
    "sales by month": "revenue_by_month",
    "top products": "top_products",
    "best products": "top_products",
    "best selling": "top_products",
    "product revenue": "top_products",
    "expense": "expense_trends",
    "expenses": "expense_trends",
    "cost": "expense_trends",
    "spending": "expense_trends",
    "expense trends": "expense_trends",
    "profit": "profitability",
    "profitability": "profitability",
    "margin": "profitability",
    "product profit": "profitability",
    "balance": "balance_summary",
    "snapshot": "balance_summary",
    "p&l": "balance_summary",
    "income statement": "balance_summary",
    "financial summary": "balance_summary",
    "segment": "segment_analysis",
    "segments": "segment_analysis",
    "category": "segment_analysis",
    "business segment": "segment_analysis",
    "data quality": "data_quality",
    "row count": "data_quality",
    "table stats": "data_quality",
    "how much data": "data_quality",
}


def _match_prebuilt(question: str) -> Optional[str]:
    """Match a natural-language question to a prebuilt query key."""
    q = question.lower().strip()
    # Direct key match
    if q in PREBUILT_QUERIES:
        return q
    # Keyword phrase match (longest first for specificity)
    for phrase in sorted(_KEYWORD_MAP, key=len, reverse=True):
        if phrase in q:
            return _KEYWORD_MAP[phrase]
    return None


def _generate_sql_from_question(question: str) -> Optional[dict]:
    """Simple keyword-based SQL generation for unmatched questions."""
    q = question.lower()
    # Detect table references
    tables = {
        "revenue_items": "dw_revenue_items",
        "transactions": "dw_transactions",
        "snapshots": "dw_financial_snapshots",
        "trial_balance": "dw_trial_balance",
        "trial balance": "dw_trial_balance",
    }
    target_table = None
    for kw, tbl in tables.items():
        if kw in q:
            target_table = tbl
            break
    if not target_table:
        # Default to revenue_items for revenue-sounding queries
        if any(w in q for w in ["sell", "sales", "revenue", "product", "item"]):
            target_table = "dw_revenue_items"
        elif any(w in q for w in ["account", "debit", "credit", "balance"]):
            target_table = "dw_trial_balance"
        elif any(w in q for w in ["transaction", "journal", "entry"]):
            target_table = "dw_transactions"
        else:
            target_table = "dw_financial_snapshots"

    # Detect aggregation intent
    if any(w in q for w in ["total", "sum", "how much", "aggregate"]):
        if target_table == "dw_revenue_items":
            sql = f"SELECT SUM(net_amount) as total FROM {target_table}"
        elif target_table == "dw_trial_balance":
            sql = f"SELECT SUM(debit) as total_debit, SUM(credit) as total_credit FROM {target_table}"
        else:
            sql = f"SELECT COUNT(*) as total_rows FROM {target_table}"
    elif any(w in q for w in ["count", "how many", "number of"]):
        sql = f"SELECT COUNT(*) as count FROM {target_table}"
    elif any(w in q for w in ["average", "avg", "mean"]):
        if target_table == "dw_revenue_items":
            sql = f"SELECT AVG(net_amount) as average_amount FROM {target_table}"
        else:
            sql = f"SELECT COUNT(*) as rows FROM {target_table}"
    elif any(w in q for w in ["group by", "by period", "by month", "monthly"]):
        if target_table == "dw_revenue_items":
            sql = f"SELECT period, SUM(net_amount) as total FROM {target_table} GROUP BY period ORDER BY period"
        else:
            sql = f"SELECT * FROM {target_table} ORDER BY period LIMIT 20"
    else:
        sql = f"SELECT * FROM {target_table} LIMIT 20"

    return {
        "title": f"Generated: {target_table}",
        "sql": sql,
        "description": f"Auto-generated query for: {question}",
    }


@router.get("/agents/data-agent/prebuilt")
async def data_agent_prebuilt():
    """List all pre-built analytical queries (warehouse + knowledge graph)."""
    queries = []
    for key, q in PREBUILT_QUERIES.items():
        queries.append({"key": key, "title": q["title"], "description": q["description"], "category": "warehouse"})
    for key, q in KG_QUERIES.items():
        queries.append({"key": key, "title": q["title"], "description": q["description"], "category": "knowledge_graph"})
    return {"queries": queries, "warehouse_count": len(PREBUILT_QUERIES), "kg_count": len(KG_QUERIES)}


@router.post("/agents/data-agent/query")
async def data_agent_query(body: dict):
    """
    AI Data Agent -- understands database schema, writes SQL, explains results.
    Accepts natural language questions about financial data.
    """
    from app.services.warehouse import warehouse

    question = body.get("question", "").strip()
    if not question:
        raise HTTPException(400, "question is required")

    # 0. Check if this is a Knowledge Graph query
    q_lower = question.lower()
    kg_keywords = {"account": "kg_accounts", "kpi": "kg_kpis", "risk": "kg_risks",
                   "benchmark": "kg_benchmarks", "standard": "kg_standards", "ifrs": "kg_standards",
                   "graph stat": "kg_graph_stats", "graph anomal": "kg_anomalies",
                   "knowledge graph": "kg_accounts", "ontology": "kg_accounts"}
    kg_match = None
    for kw, kg_key in kg_keywords.items():
        if kw in q_lower:
            kg_match = kg_key
            break

    # Also match direct KG query keys from body
    if body.get("key") and body["key"] in KG_QUERIES:
        kg_match = body["key"]

    if kg_match and kg_match in KG_QUERIES:
        kg_query = KG_QUERIES[kg_match]
        try:
            if kg_query.get("type") == "ontology":
                from app.services.ontology_tools import ontology_tool_executor
                result = await ontology_tool_executor.execute(kg_query["tool"], kg_query["params"])
                objects = result.get("objects", [])
                return {
                    "title": kg_query["title"],
                    "sql": f"ontology.{kg_query['tool']}({kg_query['params']})",
                    "results": objects[:50],
                    "row_count": result.get("count", len(objects)),
                    "description": kg_query["description"],
                    "source": "knowledge_graph",
                }
            elif kg_query.get("type") == "graph_api":
                import aiohttp
                # Call internal graph API
                from app.services.v2.graph_analytics import graph_analytics
                if "stats" in kg_query["endpoint"]:
                    from app.services.ontology_engine import ontology_registry
                    from app.services.v2.graph_analytics import GraphNode, GraphEdge
                    from collections import defaultdict
                    if not graph_analytics._nodes:
                        for oid, obj in ontology_registry._objects.items():
                            label = obj.properties.get("name_en") or obj.properties.get("name") or obj.properties.get("code") or oid
                            graph_analytics._nodes[oid] = GraphNode(id=oid, type=obj.object_type, label=str(label), properties={})
                    stats = graph_analytics.get_stats()
                    return {"title": kg_query["title"], "sql": "graph.stats()", "results": [stats], "row_count": 1, "description": kg_query["description"], "source": "knowledge_graph"}
                elif "anomal" in kg_query["endpoint"]:
                    anomalies = graph_analytics.find_anomalies()
                    return {"title": kg_query["title"], "sql": "graph.anomalies()", "results": anomalies.get("anomalies", []), "row_count": anomalies.get("total_anomalies", 0), "description": kg_query["description"], "source": "knowledge_graph"}
        except Exception as e:
            return {"title": kg_query["title"], "sql": "", "results": [], "row_count": 0, "description": kg_query["description"], "error": str(e), "source": "knowledge_graph"}

    # 1. Try matching to a prebuilt SQL query
    matched_key = _match_prebuilt(question)
    if body.get("key") and body["key"] in PREBUILT_QUERIES:
        matched_key = body["key"]
    if matched_key:
        query_info = PREBUILT_QUERIES[matched_key]
        suggestion = None
    else:
        # 2. Generate SQL from keywords
        query_info = _generate_sql_from_question(question)
        suggestion = "This query was auto-generated. For better results, try asking about: revenue, expenses, products, segments, profitability, accounts, KPIs, or risks."

    sql = query_info["sql"]
    title = query_info["title"]
    description = query_info["description"]

    # 3. Execute against DuckDB
    try:
        results = warehouse.execute(sql)
        if results and isinstance(results[0], dict) and "error" in results[0]:
            return {
                "title": title,
                "sql": sql,
                "results": [],
                "row_count": 0,
                "description": description,
                "error": results[0]["error"],
                "suggestion": "The warehouse tables may be empty. Try syncing from SQLite first.",
            }
    except Exception as e:
        return {
            "title": title,
            "sql": sql,
            "results": [],
            "row_count": 0,
            "description": description,
            "error": str(e),
            "suggestion": "Query execution failed. The warehouse may need to be synced first.",
        }

    return {
        "title": title,
        "sql": sql,
        "results": results,
        "row_count": len(results),
        "description": description,
        "suggestion": suggestion,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# AIP EVALS — AI Output Quality Evaluation Framework
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/agents/aip-evals/list")
async def list_aip_evals():
    """List all registered AIP evaluation test cases."""
    from app.services.v2.aip_evals import aip_eval_runner
    return {
        "evals": aip_eval_runner.list_evals(),
        "total": len(aip_eval_runner.evals),
    }


@router.post("/agents/aip-evals/run")
async def run_aip_evals(category: str = None):
    """Run all (or filtered) AIP evaluation test cases and return results."""
    from app.services.v2.aip_evals import aip_eval_runner
    if category:
        return await aip_eval_runner.run_by_category(category)
    return await aip_eval_runner.run_all()
