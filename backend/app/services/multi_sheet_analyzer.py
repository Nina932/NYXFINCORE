"""
Phase T: Multi-Sheet Financial Analyzer
========================================
Intelligently detects what each sheet in a workbook contains and extracts
all possible financial data: P&L, Balance Sheet, Trial Balance, Revenue/COGS
breakdowns, account mappings.

Sheet type detection is rule-based (deterministic, no LLM):
  - Revenue Breakdown: has columns like Amount, VAT, Net Revenue + product rows
  - COGS Breakdown: references account 1610/7110, has inventory items
  - Trial Balance (TDSheet): has Debit/Credit columns with account codes
  - Balance Sheet: has Start Dr/Cr, Turnover Dr/Cr, End Dr/Cr
  - Mapping: has account codes mapped to IFRS categories
  - BS Summary: condensed balance sheet with Assets/Liabilities/Equity sections

All numbers are deterministic. No LLM calls.
"""

import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

from app.config import settings

logger = logging.getLogger(__name__)

try:
    import openpyxl
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════
# SHEET TYPE DETECTION
# ═══════════════════════════════════════════════════════════════════

SHEET_TYPES = {
    "revenue_breakdown": {
        "keywords": ["revenue", "vat", "net revenue", "product", "amount gel",
                      "turnover", "sales", "retail", "wholesale"],
        "patterns": [r"revenue\s*(retail|wholesale)", r"vat", r"net\s*revenue"],
    },
    "cogs_breakdown": {
        "keywords": ["1610", "7110", "cogs", "cost of", "inventory",
                      "себестоимость", "номенклатура", "субконто"],
        "patterns": [r"1610|7110", r"себестоимость|cost\s*of\s*(goods|sales)"],
    },
    "trial_balance": {
        "keywords": ["оборотно-сальдовая", "trial balance", "сальдо на начало",
                      "дебет", "кредит", "оборот", "tdsheet"],
        "patterns": [r"оборотно.*сальдов", r"trial\s*balance", r"сальдо\s*на\s*начало"],
    },
    "balance_detail": {
        "keywords": ["start dr", "start cr", "starting balance", "turnover dr",
                      "end dr", "end cr", "ending balance", "company", "sgp"],
        "patterns": [r"start\s*dr", r"turnover\s*dr", r"ending?\s*balance"],
    },
    "pl_statement": {
        "keywords": ["code", "line item", "actual", "plan", "variance", "var%",
                      "prior year", "revenue", "cogs", "gross margin", "ebitda",
                      "net profit", "rev.w", "rev.r", "cogs.w", "cogs.r"],
        "patterns": [r"rev\.\w", r"cogs\.\w", r"line\s*item", r"actual.*plan",
                     r"prior\s*year", r"var(iance)?%?"],
    },
    "budget_pl": {
        "keywords": ["budget", "revenue", "cogs", "gross margin", "gr. margin",
                      "ebitda", "net profit", "opex", "revenue wholesale",
                      "revenue retail", "cogs wholesale", "cogs retail",
                      "budget article", "actual", "plan"],
        "patterns": [r"revenue\s*(wholesale|retail)", r"cogs\s*(wholesale|retail)",
                     r"gr\.?\s*margin", r"budget\s*article"],
    },
    "pl_mapping": {
        "keywords": ["mapping", "revenue from sale", "cogs", "operating expenses",
                      "other revenue", "admin", "selling"],
        "patterns": [r"revenue\s*from\s*sale", r"cogs|cost\s*of\s*goods"],
    },
    "bs_summary": {
        "keywords": ["assets", "liabilities", "equity", "balance sheet",
                      "აქტივები", "ვალდებულებები", "კაპიტალი", "ბალანსი"],
        "patterns": [r"total\s*assets", r"total\s*equity", r"balance\s*sheet"],
    },
    "transactions": {
        "keywords": ["transaction", "journal", "entry", "date", "debit", "credit",
                      "reference", "description"],
        "patterns": [r"journal\s*entry", r"transaction\s*id"],
    },
}


@dataclass
class SheetAnalysis:
    """Result of analyzing a single sheet."""
    sheet_name: str
    sheet_type: str  # one of SHEET_TYPES keys or "unknown"
    confidence: float  # 0-1
    row_count: int
    col_count: int
    header_row: int  # 0-indexed
    headers: List[str]
    data_preview: List[Dict[str, Any]] = field(default_factory=list)


@dataclass
class ExtractedFinancials:
    """All financials extracted from a multi-sheet workbook."""
    current_financials: Dict[str, float] = field(default_factory=dict)
    balance_sheet: Dict[str, float] = field(default_factory=dict)
    revenue_breakdown: List[Dict[str, Any]] = field(default_factory=list)
    cogs_breakdown: List[Dict[str, Any]] = field(default_factory=list)
    pl_line_items: List[Dict[str, Any]] = field(default_factory=list)
    trial_balance_accounts: List[Dict[str, Any]] = field(default_factory=list)
    account_mapping: Dict[str, str] = field(default_factory=dict)  # code -> IFRS category
    sheet_analyses: List[Dict[str, Any]] = field(default_factory=list)
    company_name: str = ""
    period: str = ""
    warnings: List[str] = field(default_factory=list)
    transaction_count: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "current_financials": self.current_financials,
            "balance_sheet": self.balance_sheet,
            "revenue_breakdown": self.revenue_breakdown[:20],
            "cogs_breakdown": self.cogs_breakdown[:30],
            "pl_line_items": self.pl_line_items[:60],
            "trial_balance_summary": {
                "total_accounts": len(self.trial_balance_accounts),
                "top_accounts": self.trial_balance_accounts[:10],
            },
            "account_mapping": dict(list(self.account_mapping.items())[:30]),
            "sheet_analyses": self.sheet_analyses,
            "company_name": self.company_name,
            "period": self.period,
            "warnings": self.warnings,
            "transaction_count": self.transaction_count,
        }


# ═══════════════════════════════════════════════════════════════════
# MULTI-SHEET ANALYZER
# ═══════════════════════════════════════════════════════════════════

class MultiSheetAnalyzer:
    """
    Analyzes multi-sheet financial workbooks and extracts all possible data.

    Detection is rule-based (keyword + pattern matching on headers and first rows).
    All financial computation is deterministic.
    """

    def analyze_file(self, file_path: str) -> ExtractedFinancials:
        """Analyze a workbook file and extract all financials."""
        if not OPENPYXL:
            raise RuntimeError("openpyxl required")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        return self._analyze_workbook(wb)

    def analyze_bytes(self, data: bytes) -> ExtractedFinancials:
        """Analyze workbook from bytes."""
        if not OPENPYXL:
            raise RuntimeError("openpyxl required")
        import io
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        return self._analyze_workbook(wb)

    def _analyze_workbook(self, wb) -> ExtractedFinancials:
        """Core analysis: detect sheet types, extract data from each."""
        result = ExtractedFinancials()

        # Phase 1: Detect sheet types
        sheet_types: List[Tuple[str, str, float]] = []  # (name, type, confidence)
        for sname in wb.sheetnames:
            ws = wb[sname]
            stype, conf = self._detect_sheet_type(ws, sname)
            sheet_types.append((sname, stype, conf))
            result.sheet_analyses.append({
                "sheet_name": sname,
                "sheet_type": stype,
                "confidence": round(conf, 2),
                "rows": ws.max_row or 0,
                "cols": ws.max_column or 0,
            })
            logger.info("Sheet '%s': detected as %s (%.0f%%)", sname, stype, conf * 100)

        # Phase 2: Extract data per sheet type
        # Process in priority order: mapping first (most accurate P&L), then revenue,
        # then balance, then trial balance, then COGS (lowest priority for amounts)
        priority_order = ["pl_statement", "pl_mapping", "budget_pl", "revenue_breakdown", "balance_detail",
                          "bs_summary", "trial_balance", "cogs_breakdown", "transactions"]
        ordered = sorted(sheet_types, key=lambda x: (
            priority_order.index(x[1]) if x[1] in priority_order else 99
        ))

        for sname, stype, conf in ordered:
            ws = wb[sname]
            try:
                if stype == "pl_statement":
                    self._extract_pl_statement(ws, result)
                elif stype == "revenue_breakdown":
                    self._extract_revenue_breakdown(ws, result)
                elif stype == "cogs_breakdown":
                    self._extract_cogs_breakdown(ws, result)
                elif stype == "trial_balance":
                    self._extract_trial_balance(ws, result)
                elif stype == "balance_detail":
                    self._extract_balance_detail(ws, result)
                elif stype == "budget_pl":
                    self._extract_budget_pl(ws, result)
                elif stype == "pl_mapping":
                    self._extract_pl_mapping(ws, result)
                elif stype == "bs_summary":
                    self._extract_bs_summary(ws, result)
                elif stype == "transactions":
                    result.transaction_count += (ws.max_row or 1) - 1
            except Exception as e:
                result.warnings.append(f"Error extracting '{sname}' ({stype}): {str(e)}")
                logger.warning("Extraction error on '%s': %s", sname, e)

        # Phase 3: Compute derived metrics
        self._compute_derived_metrics(result)

        # Phase 4: Detect company name and period
        self._detect_metadata(wb, result)

        return result

    # ── Sheet type detection ──────────────────────────────────────

    def _detect_sheet_type(self, ws, sheet_name: str) -> Tuple[str, float]:
        """Detect what type of financial data a sheet contains."""
        # Collect text from first 10 rows for analysis
        text_pool = sheet_name.lower() + " "
        for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row or 1), values_only=True):
            for cell in row:
                if cell is not None:
                    text_pool += str(cell).lower() + " "

        best_type = "unknown"
        best_score = 0.0

        for stype, rules in SHEET_TYPES.items():
            score = 0.0
            # Keyword matches
            keyword_hits = sum(1 for kw in rules["keywords"] if kw.lower() in text_pool)
            if rules["keywords"]:
                score += (keyword_hits / len(rules["keywords"])) * 0.6

            # Pattern matches
            pattern_hits = sum(1 for p in rules["patterns"]
                               if re.search(p, text_pool, re.IGNORECASE))
            if rules["patterns"]:
                score += (pattern_hits / len(rules["patterns"])) * 0.4

            if score > best_score:
                best_score = score
                best_type = stype

        # If score too low, mark unknown
        if best_score < 0.15:
            return "unknown", best_score

        return best_type, min(best_score, 1.0)

    # ── P&L Statement extraction (hierarchical codes) ─────────────

    def _extract_pl_statement(self, ws, result: ExtractedFinancials):
        """Extract hierarchical P&L statement with codes (REV.W.P, COGS.R.D, etc.)."""
        # Find header row (Code | Line Item | ... | Actual | Plan)
        header_row = None
        col_map = {}
        for r in range(1, min(10, (ws.max_row or 1) + 1)):
            cells = [str(ws.cell(r, c).value or '').strip().lower() for c in range(1, (ws.max_column or 1) + 1)]
            if 'code' in cells and ('line item' in cells or 'actual' in cells):
                header_row = r
                for ci, h in enumerate(cells):
                    if 'code' in h: col_map['code'] = ci
                    elif 'line item' in h: col_map['label'] = ci
                    elif 'prior' in h: col_map['prior'] = ci
                    elif 'actual' in h: col_map['actual'] = ci
                    elif 'plan' in h or 'budget' in h: col_map['plan'] = ci
                    elif 'variance' == h or 'var' in h and '%' not in h: col_map['variance'] = ci
                    elif 'var%' in h or ('var' in h and '%' in h): col_map['var_pct'] = ci
                break

        if not header_row or 'code' not in col_map:
            return

        # CODE HIERARCHY MAP for financial classification
        CODE_TO_SECTION = {
            'REV': 'revenue', 'COGS': 'cogs', 'GM': 'gross_margin',
            'GA': 'ga_expenses', 'EBITDA': 'ebitda', 'DA': 'depreciation',
            'EBIT': 'ebit', 'OE': 'other_expense', 'OI': 'other_income',
            'EBT': 'profit_before_tax', 'TAX': 'tax', 'NP': 'net_profit',
        }

        pl_items = []
        rev_items = []
        cogs_items = []

        for r in range(header_row + 1, (ws.max_row or 1) + 1):
            code = str(ws.cell(r, col_map['code'] + 1).value or '').strip()
            label = str(ws.cell(r, col_map.get('label', col_map['code']) + 1).value or '').strip()
            actual = 0.0
            plan = 0.0
            prior = 0.0

            if 'actual' in col_map:
                v = ws.cell(r, col_map['actual'] + 1).value
                actual = float(v) if v and v != '' else 0.0
            if 'plan' in col_map:
                v = ws.cell(r, col_map['plan'] + 1).value
                plan = float(v) if v and v != '' else 0.0
            if 'prior' in col_map:
                v = ws.cell(r, col_map['prior'] + 1).value
                prior = float(v) if v and v != '' else 0.0

            if not code and not label:
                continue

            # Determine hierarchy level from code dots
            level = code.count('.')
            top_code = code.split('.')[0] if code else ''
            section = CODE_TO_SECTION.get(top_code, 'other')
            is_total = level == 0 and top_code in CODE_TO_SECTION

            pl_items.append({
                "code": code, "label": label, "actual": actual,
                "plan": plan, "prior_year": prior,
                "variance": actual - plan, "var_pct": ((actual - plan) / plan * 100) if plan else 0,
                "level": level, "section": section, "is_total": is_total,
            })

            # Build revenue/COGS product breakdowns from detail lines
            if top_code == 'REV' and level >= 2:
                cat = 'Revenue Wholesale' if '.W.' in code else 'Revenue Retail' if '.R.' in code else 'Other Revenue'
                rev_items.append({
                    "product": label, "code": code,
                    "gross_amount": actual, "vat": 0, "net_revenue": actual,
                    "category": cat, "plan": plan, "prior_year": prior,
                })
            elif top_code == 'COGS' and level >= 2:
                cat = 'COGS Wholesale' if '.W.' in code else 'COGS Retail' if '.R.' in code else 'COGS'
                cogs_items.append({
                    "product": label, "code": code,
                    "amount": abs(actual), "category": cat,
                    "plan": abs(plan), "prior_year": abs(prior),
                })

        # Store in result
        if pl_items:
            result.pl_line_items = pl_items

        # Set top-level financials from totals
        for item in pl_items:
            c = item['code']
            a = item['actual']
            if c == 'REV': result.current_financials['revenue'] = a
            elif c == 'REV.W': result.current_financials['revenue_wholesale'] = a
            elif c == 'REV.R': result.current_financials['revenue_retail'] = a
            elif c == 'COGS': result.current_financials['cogs'] = abs(a)
            elif c == 'GM': result.current_financials['gross_profit'] = a
            elif c == 'GA': result.current_financials['ga_expenses'] = abs(a)
            elif c == 'GA.S' or c == 'SELL': result.current_financials['selling_expenses'] = abs(a)
            elif c == 'GA.A' or c == 'ADM': result.current_financials['admin_expenses'] = abs(a)
            elif c == 'EBITDA': result.current_financials['ebitda'] = a
            elif c == 'DA': result.current_financials['depreciation'] = abs(a)
            elif c == 'EBIT': result.current_financials['ebit'] = a
            elif c == 'EBT': result.current_financials['profit_before_tax'] = a
            elif c == 'NP': result.current_financials['net_profit'] = a

        # Add revenue/COGS product details
        if rev_items:
            result.revenue_breakdown = rev_items
        if cogs_items:
            result.cogs_breakdown = cogs_items

        logger.info("P&L Statement: %d lines, %d rev products, %d COGS products, Revenue=%s",
                     len(pl_items), len(rev_items), len(cogs_items),
                     result.current_financials.get('revenue', 0))

    # ── Revenue Breakdown extraction ──────────────────────────────

    def _extract_revenue_breakdown(self, ws, result: ExtractedFinancials):
        """Extract product-level revenue data."""
        total_gross = 0.0
        total_vat = 0.0
        total_net = 0.0
        categories = {}

        # Summary/total row indicators (EN/RU/GE) — these are NOT real products
        _TOTAL_KEYWORDS = {"итог", "итого", "всего", "total", "grand total", "subtotal",
                           "სულ", "ჯამი", "ჯამური", "sum", "zusammen"}

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            product = str(row[0]).strip()
            amount = self._to_float(row[1] if len(row) > 1 else None)
            vat = self._to_float(row[2] if len(row) > 2 else None)
            net = self._to_float(row[3] if len(row) > 3 else None)
            category = str(row[4]).strip() if len(row) > 4 and row[4] else "Other"

            if amount == 0 and net == 0:
                continue

            # Skip summary/total rows — they double-count real products
            if product.lower().strip() in _TOTAL_KEYWORDS:
                logger.info("Skipping total row in revenue: '%s' = %.2f", product, net)
                continue

            # Skip date rows (e.g., "01.01.2025 0:00:00") and bare account codes (e.g., "7310.02.1")
            import re
            if re.match(r'^\d{2}\.\d{2}\.\d{4}', product):
                continue
            if re.match(r'^\d{4}\.\d{2}', product):
                continue

            total_gross += amount
            total_vat += vat
            total_net += net
            categories[category] = categories.get(category, 0) + net

            result.revenue_breakdown.append({
                "product": product[:60],
                "gross_amount": round(amount, 2),
                "vat": round(vat, 2),
                "net_revenue": round(net, 2),
                "category": category,
            })

        # Store in financials (don't overwrite if Budget/Mapping already set a higher-priority value)
        if total_net > 0 and "revenue" not in result.current_financials:
            result.current_financials["revenue_gross"] = round(total_gross, 2)
            result.current_financials["vat_collected"] = round(total_vat, 2)
            for cat, amt in categories.items():
                key = "revenue_" + re.sub(r'\s+', '_', cat.lower())
                result.current_financials[key] = round(amt, 2)

        logger.info("Revenue: %.2f GEL net (%d products, %d categories)",
                     total_net, len(result.revenue_breakdown), len(categories))

    # ── COGS Breakdown extraction ─────────────────────────────────

    def _extract_cogs_breakdown(self, ws, result: ExtractedFinancials):
        """Extract cost of goods data."""
        total_cogs = 0.0
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            if not row:
                continue
            row_text = " ".join(str(c or "").lower() for c in row)
            if "субконто" in row_text or "item" in row_text.lower() or "product" in row_text.lower():
                header_row = row_idx
                continue

            if header_row and row_idx > header_row:
                product = str(row[0] or "").strip() if row[0] else ""
                if not product:
                    continue
                # Find the debit turnover (usually the largest numeric column)
                debit = 0.0
                for cell in row[1:]:
                    val = self._to_float(cell)
                    if val > debit:
                        debit = val

                if debit > 0:
                    total_cogs += debit
                    result.cogs_breakdown.append({
                        "product": product[:60],
                        "amount": round(debit, 2),
                    })

        # IMPORTANT: This sheet is account 1610 (inventory movements), NOT account 7110 (COGS)
        # Account 1610 = Balance Sheet / Current Assets / Inventory — NOT a P&L expense
        # NEVER use this as COGS. The real COGS comes from Mapping sheet (account 7110)
        # Store as informational breakdown only
        result.current_financials["inventory_turnover_total"] = round(total_cogs, 2)
        logger.info("COGS breakdown (inventory 1610 turnovers): %.2f GEL (%d items) — NOT used as P&L COGS",
                     total_cogs, len(result.cogs_breakdown))

    # ── Trial Balance extraction ──────────────────────────────────

    def _extract_trial_balance(self, ws, result: ExtractedFinancials):
        """Extract trial balance accounts."""
        header_found = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row:
                continue
            row_text = " ".join(str(c or "").lower() for c in row[:8])

            if not header_found:
                if "код" in row_text or "счет" in row_text or "code" in row_text.lower():
                    header_found = True
                continue

            # Extract account code and balances
            code = None
            name = ""
            for cell in row:
                val = str(cell or "").strip()
                if re.match(r'^\d{2,4}[X/]?\d*', val):
                    code = val
                elif code and not name and len(val) > 3 and not val.replace(",", "").replace(".", "").replace("-", "").isdigit():
                    name = val[:60]

            if not code:
                continue

            # Collect numeric values
            nums = []
            for cell in row:
                n = self._to_float(cell)
                if n != 0:
                    nums.append(n)

            if nums:
                result.trial_balance_accounts.append({
                    "code": code,
                    "name": name,
                    "values": nums[:6],  # up to 6 numeric columns
                })

        logger.info("Trial balance: %d accounts extracted", len(result.trial_balance_accounts))

    # ── Balance detail extraction ─────────────────────────────────

    def _extract_balance_detail(self, ws, result: ExtractedFinancials):
        """Extract balance sheet data from detailed account listing.
        Expected format: Index | Company | Code | Name | Start Dr | Start Cr | Starting Balance | ...
        """
        aggregates = {
            "cash": (["11", "12"], 0.0),
            "receivables": (["14"], 0.0),
            "inventory": (["16"], 0.0),
            "fixed_assets_gross": (["21"], 0.0),
            "accumulated_depreciation": (["22"], 0.0),
            "investments": (["24"], 0.0),
            "short_term_liabilities": (["31"], 0.0),
            "long_term_debt": (["41"], 0.0),
            "equity_items": (["51", "52", "53"], 0.0),
            "revenue_accounts": (["61"], 0.0),
            "cogs_accounts": (["71"], 0.0),
        }

        # Detect code and balance columns from header
        code_col = 2  # default column C
        bal_col = 6   # default column G (Starting Balance)
        header_row = None

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            if not row:
                continue
            for i, cell in enumerate(row):
                val = str(cell or "").lower()
                if val in ("code", "код", "счет"):
                    code_col = i
                if "starting" in val or "начало" in val or "start dr" in val.replace(" ", ""):
                    bal_col = i
                    header_row = row_idx

        for row in ws.iter_rows(min_row=(header_row or 1) + 1, max_row=ws.max_row, values_only=True):
            if not row or len(row) <= code_col:
                continue

            code = str(row[code_col] or "").strip()
            # Match XX-level codes like 11XX, 12XX, 21XX etc.
            if not re.match(r'^\d{2}XX$', code):
                continue

            # Get starting balance — try the detected column and nearby
            start_bal = 0.0
            for col_offset in [bal_col, 4, 5, 6]:
                if col_offset < len(row):
                    v = self._to_float(row[col_offset])
                    if v != 0:
                        start_bal = v
                        break

            code_prefix = code[:2]
            for key, (prefixes, _) in aggregates.items():
                if code_prefix in prefixes:
                    aggregates[key] = (prefixes, aggregates[key][1] + start_bal)

        # Build balance sheet
        cash = aggregates["cash"][1]
        recv = aggregates["receivables"][1]
        inv = aggregates["inventory"][1]
        fa_gross = aggregates["fixed_assets_gross"][1]
        accum_dep = abs(aggregates["accumulated_depreciation"][1])
        investments = aggregates["investments"][1]
        st_liab = abs(aggregates["short_term_liabilities"][1])
        lt_debt = abs(aggregates["long_term_debt"][1])

        total_current = cash + recv + inv
        total_fixed = fa_gross - accum_dep + investments
        total_assets = total_current + total_fixed
        total_liabilities = st_liab + lt_debt
        total_equity = total_assets - total_liabilities

        result.balance_sheet = {
            "cash": round(cash, 2),
            "receivables": round(recv, 2),
            "inventory": round(inv, 2),
            "total_current_assets": round(total_current, 2),
            "fixed_assets_gross": round(fa_gross, 2),
            "accumulated_depreciation": round(accum_dep, 2),
            "fixed_assets_net": round(fa_gross - accum_dep, 2),
            "investments": round(investments, 2),
            "total_fixed_assets": round(total_fixed, 2),
            "total_assets": round(total_assets, 2),
            "total_current_liabilities": round(st_liab, 2),
            "long_term_debt": round(lt_debt, 2),
            "total_liabilities": round(total_liabilities, 2),
            "total_equity": round(total_equity, 2),
        }

        # Extract P&L from revenue/cogs accounts if present
        rev_total = abs(aggregates["revenue_accounts"][1])
        cogs_total = abs(aggregates["cogs_accounts"][1])
        if rev_total > 0 and "revenue" not in result.current_financials:
            result.current_financials["revenue"] = round(rev_total, 2)
        if cogs_total > 0 and "cogs" not in result.current_financials:
            result.current_financials["cogs"] = round(cogs_total, 2)

        logger.info("Balance sheet extracted: assets=%.0f, liabilities=%.0f, equity=%.0f",
                     total_assets, total_liabilities, total_equity)

    # ── Budget / P&L extraction (label-value format) ───────────────

    def _extract_budget_pl(self, ws, result: ExtractedFinancials):
        """Extract P&L from Budget/Management Report sheet.
        Format: Column A = label (e.g. 'Revenue Wholesale'), Column B = amount.
        Understands hierarchical structure: Revenue > Revenue Wholesale > Revenue Whsale Petrol."""
        label_map = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or len(row) < 2:
                continue
            label = str(row[0] or "").strip()
            amount = self._to_float(row[1])
            if not label or amount is None:
                continue
            label_lower = label.lower()
            label_map[label_lower] = {"label": label, "amount": amount}

        # Extract P&L line items using fuzzy label matching
        def find(keys):
            for k in keys:
                for lbl, data in label_map.items():
                    if k in lbl and data["amount"] != 0:
                        return data["amount"]
            return 0

        revenue = find(["revenue "]) or find(["revenue"])  # "Revenue " with space to avoid "Revenue Wholesale"
        # More precise: look for exact "Revenue" label (total)
        for lbl, data in label_map.items():
            if lbl.strip() == "revenue" and data["amount"] > 0:
                revenue = data["amount"]
                break

        cogs = find(["cogs "]) or find(["cost of goods"])
        for lbl, data in label_map.items():
            if lbl.strip() == "cogs" and data["amount"] > 0:
                cogs = data["amount"]
                break

        gross_margin = find(["gr. margin", "gross margin", "gross profit"])
        for lbl, data in label_map.items():
            if lbl.strip() in ("gr. margin", "gross margin") and data["amount"] != 0:
                gross_margin = data["amount"]
                break

        ga_exp = find(["general and admin", "g&a", "opex", "operating exp"])
        ebitda = find(["ebitda"])
        depreciation = find(["depreciation", "amortization", "d&a"])
        net_profit = find(["net profit", "net income"])
        other_revenue = find(["other revenue"])
        interest = find(["interest", "finance cost", "finance expense"])
        tax = find(["income tax", "tax expense"])

        # Revenue sub-categories
        rev_wholesale = find(["revenue wholesale"])
        rev_retail = find(["revenue retail", "revenue retial"])
        cogs_wholesale = find(["cogs wholesale"])
        cogs_retail = find(["cogs retail", "cogs retial"])

        # Set financials (ONLY if not already set by higher-priority sheet)
        fin = result.current_financials
        if revenue > 0 and "revenue" not in fin:
            fin["revenue"] = round(revenue, 2)
        if cogs > 0 and "cogs" not in fin:
            fin["cogs"] = round(cogs, 2)
        if gross_margin != 0:
            fin["gross_profit"] = round(gross_margin, 2)
        elif revenue > 0 and cogs > 0 and "gross_profit" not in fin:
            fin["gross_profit"] = round(revenue - cogs, 2)
        if ga_exp > 0:
            fin["ga_expenses"] = round(ga_exp, 2)
        if ebitda != 0:
            fin["ebitda"] = round(ebitda, 2)
        if depreciation > 0:
            fin["depreciation"] = round(depreciation, 2)
        if net_profit != 0:
            fin["net_profit"] = round(net_profit, 2)
        if other_revenue != 0:
            fin["other_income"] = round(abs(other_revenue), 2)
        if interest > 0:
            fin["finance_expense"] = round(interest, 2)
        if tax > 0:
            fin["tax_expense"] = round(tax, 2)

        # Revenue sub-categories
        if rev_wholesale > 0:
            fin["revenue_wholesale"] = round(rev_wholesale, 2)
        if rev_retail > 0:
            fin["revenue_retail"] = round(rev_retail, 2)
        if cogs_wholesale > 0:
            fin["cogs_wholesale"] = round(cogs_wholesale, 2)
        if cogs_retail > 0:
            fin["cogs_retail"] = round(cogs_retail, 2)

        # Compute margins if we have revenue and cogs
        if "revenue" in fin and "cogs" in fin:
            rev_val = fin["revenue"]
            cogs_val = fin["cogs"]
            if rev_val > 0:
                fin.setdefault("gross_margin_pct", round((rev_val - cogs_val) / rev_val * 100, 2))
                fin.setdefault("cogs_to_revenue_pct", round(cogs_val / rev_val * 100, 2))

        # Also extract product-level detail from Budget rows
        product_categories = {"wholesale": [], "retail": []}
        for lbl, data in label_map.items():
            if "whsale" in lbl or "wholesale" in lbl:
                if any(p in lbl for p in ["petrol", "diesel", "bitumen", "cng", "lpg", "fuel oil", "urea", "petroleum"]):
                    is_cogs = "cogs" in lbl
                    is_margin = "margin" in lbl
                    if not is_cogs and not is_margin:
                        product_categories["wholesale"].append({"product": data["label"], "net_revenue": data["amount"], "category": "Revenue Wholesale"})
            elif "retial" in lbl or "retail" in lbl:
                if any(p in lbl for p in ["petrol", "diesel", "cng", "lpg"]):
                    is_cogs = "cogs" in lbl
                    is_margin = "margin" in lbl
                    if not is_cogs and not is_margin:
                        product_categories["retail"].append({"product": data["label"], "net_revenue": data["amount"], "category": "Revenue Retail"})

        # Add to revenue breakdown if not already populated
        if not result.revenue_breakdown and (product_categories["wholesale"] or product_categories["retail"]):
            result.revenue_breakdown = product_categories["wholesale"] + product_categories["retail"]

        logger.info("Budget P&L extracted: revenue=%.0f, cogs=%.0f, gm=%.0f, items=%d",
                     revenue, cogs, gross_margin, len(result.revenue_breakdown))

    # ── P&L Mapping extraction ────────────────────────────────────

    def _extract_pl_mapping(self, ws, result: ExtractedFinancials):
        """Extract P&L from account mapping sheet.
        Processes XX-level and 4-digit codes for P&L totals.
        Also scans sub-accounts for depreciation using column F (Detail Classification).
        """
        revenue = 0.0
        cogs = 0.0
        selling_exp = 0.0
        admin_exp = 0.0
        other_income = 0.0
        other_expense = 0.0
        depreciation = 0.0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row:
                continue

            code = str(row[1] or "").strip() if len(row) > 1 else ""
            amount = self._to_float(row[3] if len(row) > 3 else None)
            category = str(row[4] or "").strip().lower() if len(row) > 4 else ""
            detail_f = str(row[5] or "").strip() if len(row) > 5 else ""

            if not code or amount == 0:
                continue

            # Store mapping
            if category:
                result.account_mapping[code] = category

            # ── Depreciation from sub-accounts (column F) ──
            # Sub-accounts like 7310.02.1/1, 7410.01/1 may have "Depreciation" in column F
            if ('/' in code or '.' in code) and detail_f:
                detail_lower = detail_f.lower()
                if ('depreci' in detail_lower or 'amortiz' in detail_lower or
                    'ამორტიზ' in detail_lower or 'ცვეთა' in detail_lower):
                    depreciation += abs(amount)

            # Process XX-level aggregates AND 4-digit primary accounts
            code_clean = code.replace(" ", "").split("/")[0].split(".")[0]  # strip sub-accounts AND sub-codes
            is_xx = "XX" in code_clean
            is_4digit = bool(re.match(r'^\d{4}$', code_clean))
            if not is_xx and not is_4digit:
                continue
            code_prefix = code_clean[:2]

            # Accept: XX-level, OR 4-digit parent codes (with or without category)
            # Using max() avoids double-counting parent+child (XX is always the aggregate)
            if is_xx or is_4digit:
                if code_prefix == "61":
                    revenue = max(revenue, abs(amount))
                elif code_prefix == "71":
                    cogs = max(cogs, abs(amount))
                elif code_prefix == "73":
                    selling_exp = max(selling_exp, abs(amount))
                elif code_prefix == "74":
                    admin_exp = max(admin_exp, abs(amount))
                elif code_prefix == "81":
                    other_income = max(other_income, abs(amount))
                elif code_prefix in ("82", "83"):
                    other_expense = max(other_expense, abs(amount))

        # Merge into financials (don't overwrite if revenue breakdown was more detailed)
        if revenue > 0 and "revenue" not in result.current_financials:
            result.current_financials["revenue"] = round(revenue, 2)
        if cogs > 0:
            result.current_financials["cogs"] = round(cogs, 2)
        if selling_exp > 0:
            result.current_financials["selling_expenses"] = round(selling_exp, 2)
        if admin_exp > 0:
            result.current_financials["admin_expenses"] = round(admin_exp, 2)
        if selling_exp + admin_exp > 0:
            result.current_financials["ga_expenses"] = round(selling_exp + admin_exp, 2)
        if depreciation > 0:
            result.current_financials["depreciation"] = round(depreciation, 2)
        if other_income > 0:
            result.current_financials["other_income"] = round(other_income, 2)
        if other_expense > 0:
            result.current_financials["other_expense"] = round(other_expense, 2)

        # Compute EBITDA and EBIT
        gp = revenue - cogs
        ebitda = gp - selling_exp - admin_exp
        ebit = ebitda - depreciation
        result.current_financials["gross_profit"] = round(gp, 2)
        result.current_financials["ebitda"] = round(ebitda, 2)
        result.current_financials["ebit"] = round(ebit, 2)
        result.current_financials["net_profit"] = round(ebit + other_income - other_expense, 2)

        logger.info("P&L mapping: revenue=%.0f, cogs=%.0f, selling=%.0f, admin=%.0f, "
                     "depreciation=%.0f, ebitda=%.0f, net_profit=%.0f",
                     revenue, cogs, selling_exp, admin_exp, depreciation, ebitda,
                     ebit + other_income - other_expense)

    # ── BS Summary extraction ─────────────────────────────────────

    def _extract_bs_summary(self, ws, result: ExtractedFinancials):
        """Extract from a condensed balance sheet summary."""
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row:
                continue
            label = str(row[0] or "").strip().lower()
            value = self._to_float(row[1] if len(row) > 1 else None)

            if not label or value == 0:
                continue

            if any(k in label for k in ["total asset", "სულ აქტივ"]):
                result.balance_sheet["total_assets"] = round(abs(value), 2)
            elif any(k in label for k in ["total liabilit", "სულ ვალდებულ"]):
                result.balance_sheet["total_liabilities"] = round(abs(value), 2)
            elif any(k in label for k in ["total equity", "სულ კაპიტალ"]):
                result.balance_sheet["total_equity"] = round(abs(value), 2)

    # ── Derived metrics computation ───────────────────────────────

    def _compute_derived_metrics(self, result: ExtractedFinancials):
        """Compute all derived financial metrics deterministically."""
        fin = result.current_financials
        rev = fin.get("revenue", 0)
        cogs = fin.get("cogs", 0)

        if rev > 0:
            # Gross profit
            gp = rev - cogs
            fin.setdefault("gross_profit", round(gp, 2))
            fin["gross_margin_pct"] = round(gp / rev * 100, 2)
            fin["cogs_to_revenue_pct"] = round(cogs / rev * 100, 2)

            # Operating metrics
            ga = fin.get("ga_expenses", 0)
            ebitda = gp - ga
            fin.setdefault("ebitda", round(ebitda, 2))
            fin["ebitda_margin_pct"] = round(ebitda / rev * 100, 2)

            # Net profit
            dep = fin.get("depreciation", 0)
            fin_exp = fin.get("finance_expense", 0) + fin.get("other_expense", 0)
            other_inc = fin.get("other_income", 0)
            np_ = ebitda - dep - fin_exp + other_inc
            fin.setdefault("net_profit", round(np_, 2))
            fin["net_margin_pct"] = round(np_ / rev * 100, 2)

            fin.setdefault("tax_rate", 0.15)

        # Balance sheet ratios
        bs = result.balance_sheet
        if bs:
            ca = bs.get("total_current_assets", 0)
            cl = bs.get("total_current_liabilities", 0)
            eq = bs.get("total_equity", 0)
            tl = bs.get("total_liabilities", 0)

            if cl > 0:
                fin["current_ratio"] = round(ca / cl, 2)
            if eq > 0:
                fin["debt_to_equity"] = round(tl / eq, 2)
            if ca > 0:
                fin["working_capital"] = round(ca - cl, 2)

    # ── Metadata detection ────────────────────────────────────────

    def _detect_metadata(self, wb, result: ExtractedFinancials):
        """Detect company name and period from the workbook."""
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                for cell in row:
                    text = str(cell or "")
                    # Company name detection
                    if not result.company_name:
                        if "сокар" in text.lower() or "nyx" in text.lower():
                            result.company_name = settings.COMPANY_NAME
                        elif "შპს" in text or "LLC" in text:
                            result.company_name = text[:80]
                    # Period detection
                    if not result.period:
                        period_match = re.search(
                            r'(январь|февраль|март|апрель|май|июнь|июль|август|'
                            r'сентябрь|октябрь|ноябрь|декабрь|'
                            r'january|february|march|april|may|june|july|august|'
                            r'september|october|november|december)\s*(\d{4})',
                            text, re.IGNORECASE
                        )
                        if period_match:
                            month_map = {
                                "январь": "01", "февраль": "02", "март": "03",
                                "апрель": "04", "май": "05", "июнь": "06",
                                "июль": "07", "август": "08", "сентябрь": "09",
                                "октябрь": "10", "ноябрь": "11", "декабрь": "12",
                                "january": "01", "february": "02", "march": "03",
                                "april": "04", "may": "05", "june": "06",
                                "july": "07", "august": "08", "september": "09",
                                "october": "10", "november": "11", "december": "12",
                            }
                            m = month_map.get(period_match.group(1).lower(), "01")
                            y = period_match.group(2)
                            result.period = f"{y}-{m}"

        if not result.company_name:
            result.company_name = "Unknown Company"
        if not result.period:
            from datetime import datetime
            result.period = datetime.now().strftime("%Y-%m")

    # ── Helpers ───────────────────────────────────────────────────

    @staticmethod
    def _to_float(val) -> float:
        """Safely convert any value to float."""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            cleaned = str(val).replace(",", "").replace(" ", "").replace("\xa0", "")
            return float(cleaned)
        except (ValueError, TypeError):
            return 0.0


# Module-level singleton
multi_sheet_analyzer = MultiSheetAnalyzer()
