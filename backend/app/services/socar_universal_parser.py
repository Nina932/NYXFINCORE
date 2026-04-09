"""
FinAI Universal 1C Excel Parser
================================
Handles any NYX Core Thinker 1C export format:
  - File type A (Jan 2025 style): Budget, Revenue Breakdown, COGS Breakdown, Base
  - File type B (Jan 2026 style): Mapping, Revenue Breakdown, COGS Breakdown, TDSheet, Balance, BS

Every number must be correct. Every P&L line must be present.
"""

import openpyxl
import re
import json
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field, asdict
from datetime import datetime

from app.config import settings


# ─── Russian month names for period detection ───
RUSSIAN_MONTHS = {
    'январь': '01', 'февраль': '02', 'март': '03', 'апрель': '04',
    'май': '05', 'июнь': '06', 'июль': '07', 'август': '08',
    'сентябрь': '09', 'октябрь': '10', 'ноябрь': '11', 'декабрь': '12',
}

ENGLISH_MONTHS = {
    'january': '01', 'february': '02', 'march': '03', 'april': '04',
    'may': '05', 'june': '06', 'july': '07', 'august': '08',
    'september': '09', 'october': '10', 'november': '11', 'december': '12',
}


@dataclass
class DataQualityFlag:
    code: str
    severity: str  # CRITICAL, HIGH, MEDIUM, LOW
    message: str


@dataclass
class RevenueItem:
    product: str
    gross_amount: float
    vat: float
    net_revenue: float
    category: str  # Revenue Retail, Revenue Wholesale, Other Revenue


@dataclass
class COGSItem:
    product: str
    amount: float
    category: str  # COGS Retail, COGS Wholesale


@dataclass
class ExpenseCategory:
    name: str
    amount: float


@dataclass
class PnL:
    revenue: float = 0
    revenue_wholesale: float = 0
    revenue_retail: float = 0
    revenue_other: float = 0
    cogs: float = 0
    cogs_wholesale: float = 0
    cogs_retail: float = 0
    gross_profit: float = 0
    selling_expenses: float = 0
    admin_expenses: float = 0
    total_opex: float = 0
    ebitda: float = 0
    depreciation: float = 0
    amortization: float = 0
    ebit: float = 0
    non_operating_income: float = 0
    non_operating_expense: float = 0
    interest_income: float = 0
    interest_expense: float = 0
    fx_gain_loss: float = 0
    profit_before_tax: float = 0
    income_tax: float = 0
    net_profit: float = 0


@dataclass
class BalanceSheet:
    # Assets
    total_assets: float = 0
    non_current_assets: float = 0
    current_assets: float = 0
    cash: float = 0
    inventories: float = 0
    trade_receivables: float = 0
    trade_receivables_lt: float = 0
    tax_assets: float = 0
    prepayments: float = 0
    short_term_loans_receivable: float = 0
    ppe_net: float = 0
    ppe_cost: float = 0
    ppe_depreciation: float = 0
    right_of_use: float = 0
    investment_properties: float = 0
    investments: float = 0
    intangible_assets: float = 0
    # Liabilities
    total_liabilities: float = 0
    non_current_liabilities: float = 0
    current_liabilities: float = 0
    trade_payables: float = 0
    advances_received: float = 0
    long_term_loans: float = 0
    short_term_loans: float = 0
    lease_liability_nc: float = 0
    lease_liability_c: float = 0
    other_taxes_payable: float = 0
    # Equity
    total_equity: float = 0
    share_capital: float = 0
    retained_earnings: float = 0
    net_income_period: float = 0
    revaluation_reserve: float = 0


@dataclass
class ParseResult:
    success: bool = False
    company: str = ""
    period: str = ""
    period_source: str = ""
    file_type: str = ""  # "A" (Budget+Base) or "B" (Mapping+TDSheet)
    
    pnl: PnL = field(default_factory=PnL)
    balance_sheet: BalanceSheet = field(default_factory=BalanceSheet)
    revenue_breakdown: List[RevenueItem] = field(default_factory=list)
    cogs_breakdown: List[COGSItem] = field(default_factory=list)
    selling_expense_detail: List[ExpenseCategory] = field(default_factory=list)
    admin_expense_detail: List[ExpenseCategory] = field(default_factory=list)
    
    sheets_available: List[str] = field(default_factory=list)
    sheets_used: List[str] = field(default_factory=list)
    data_quality_score: int = 0
    data_quality_flags: List[DataQualityFlag] = field(default_factory=list)
    
    def to_dict(self) -> dict:
        d = {}
        for k, v in self.__dict__.items():
            if isinstance(v, (PnL, BalanceSheet)):
                d[k] = v.__dict__
            elif isinstance(v, list) and v and hasattr(v[0], '__dict__'):
                d[k] = [item.__dict__ for item in v]
            else:
                d[k] = v
        return d


def safe_float(val) -> float:
    """Convert any value to float, returning 0 for None/empty/non-numeric."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(',', '').replace(' ', ''))
    except (ValueError, TypeError):
        return 0.0


def detect_period_from_text(text: str) -> Optional[str]:
    """Extract YYYY-MM from Russian period text like 'Период: Январь 2026 г.'"""
    if not text:
        return None
    text_lower = text.lower().strip()
    for ru_month, num in RUSSIAN_MONTHS.items():
        if ru_month in text_lower:
            year_match = re.search(r'20\d{2}', text)
            if year_match:
                return f"{year_match.group()}-{num}"
    return None


def detect_period_from_filename(filename: str) -> Optional[str]:
    """Extract period from filename like 'Report- January 2025.xlsx'"""
    if not filename:
        return None
    fn_lower = filename.lower()
    for en_month, num in ENGLISH_MONTHS.items():
        if en_month in fn_lower:
            year_match = re.search(r'20\d{2}', filename)
            if year_match:
                return f"{year_match.group()}-{num}"
    return None


# ══════════════════════════════════════════════════════════════
# SHEET PARSERS
# ══════════════════════════════════════════════════════════════

def parse_mapping_sheet(ws) -> Tuple[PnL, List[ExpenseCategory], List[ExpenseCategory], List[DataQualityFlag]]:
    """
    Parse the Mapping sheet — the BEST source for complete P&L.
    Account codes: 61XX=Revenue, 7110=COGS, 7310=Selling, 7410=Admin, 81XX=NonOpIncome, 8220=NonOpExpense
    """
    pnl = PnL()
    selling_detail = {}
    admin_detail = {}
    flags = []
    
    depreciation_total = 0.0
    interest_income = 0.0
    interest_expense = 0.0
    fx_total = 0.0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        code = str(row[1].value or '').strip() if len(row) > 1 else ''
        name = str(row[2].value or '').strip() if len(row) > 2 else ''
        amount = safe_float(row[3].value) if len(row) > 3 else 0
        mapping_e = str(row[4].value or '').strip() if len(row) > 4 else ''
        detail_f = str(row[5].value or '').strip() if len(row) > 5 else ''
        
        if not code or not amount:
            continue
        
        # ── Top-level P&L accounts (XX suffix = aggregated) ──
        if code == '61XX' or (code.startswith('61') and 'XX' in code):
            pnl.revenue = abs(amount)  # Revenue is negative in 1C (credit), flip to positive
        
        elif code == '7110':
            pnl.cogs = abs(amount)
        
        elif code == '7310':
            pnl.selling_expenses = abs(amount)
        
        elif code == '7410':
            pnl.admin_expenses = abs(amount)
        
        elif code == '81XX' or (code.startswith('81') and 'XX' in code):
            pnl.non_operating_income = abs(amount)
        
        elif code == '8220':
            pnl.non_operating_expense = abs(amount)
        
        # ── Sub-account detail (*/1 pattern = leaf level with classifications) ──
        elif '/1' in code or '/1' in code:
            # Depreciation detection
            if detail_f and 'Depreciation' in detail_f:
                depreciation_total += abs(amount)
            
            # Interest
            if detail_f and 'Interest income' in detail_f:
                interest_income += abs(amount)
            elif detail_f and ('Interest expense' in detail_f or 'interest expense' in detail_f):
                interest_expense += abs(amount)
            
            # FX
            if detail_f and 'Net FX' in detail_f:
                fx_total += amount  # Keep sign — negative = gain, positive = loss
            
            # Selling expense detail (7310.xx.x/1)
            if code.startswith('7310') and detail_f and amount:
                cat = detail_f if detail_f and detail_f != '???' else 'Other'
                selling_detail[cat] = selling_detail.get(cat, 0) + abs(amount)
            
            # Admin expense detail (7410.xx/1)
            elif code.startswith('7410') and detail_f and amount:
                cat = detail_f if detail_f and detail_f != '???' else 'Other'
                admin_detail[cat] = admin_detail.get(cat, 0) + abs(amount)
    
    # ── Computed P&L lines ──
    pnl.gross_profit = pnl.revenue - pnl.cogs
    pnl.total_opex = pnl.selling_expenses + pnl.admin_expenses
    pnl.ebitda = pnl.gross_profit - pnl.total_opex
    pnl.depreciation = depreciation_total
    pnl.ebit = pnl.ebitda - pnl.depreciation
    pnl.interest_income = interest_income
    pnl.interest_expense = interest_expense
    pnl.fx_gain_loss = fx_total
    pnl.profit_before_tax = pnl.ebit + pnl.non_operating_income - pnl.non_operating_expense
    pnl.net_profit = pnl.profit_before_tax  # No tax line in current data
    
    # ── Validation flags ──
    if pnl.selling_expenses == 0 and pnl.admin_expenses == 0:
        flags.append(DataQualityFlag('MISSING_OPEX', 'CRITICAL', 'No operating expenses found in Mapping sheet'))
    if depreciation_total == 0:
        flags.append(DataQualityFlag('MISSING_DEPRECIATION', 'HIGH', 'No depreciation found in Mapping sheet'))
    
    selling_list = [ExpenseCategory(k, v) for k, v in sorted(selling_detail.items(), key=lambda x: -x[1])]
    admin_list = [ExpenseCategory(k, v) for k, v in sorted(admin_detail.items(), key=lambda x: -x[1])]
    
    return pnl, selling_list, admin_list, flags


def parse_budget_sheet(ws) -> Tuple[PnL, List[DataQualityFlag]]:
    """
    Parse the Budget sheet — pre-computed P&L summary in English.
    LIMITATION: Does NOT contain G&A, Selling Expenses, or Depreciation.
    """
    pnl = PnL()
    flags = []
    data = {}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = str(row[0] or '').strip() if row[0] else ''
        value = safe_float(row[1]) if len(row) > 1 else 0
        if label and value:
            data[label] = value
    
    pnl.revenue = data.get('Revenue ', data.get('Revenue', 0))
    pnl.revenue_wholesale = data.get('Revenue Wholesale', 0)
    pnl.revenue_retail = data.get('Revenue Retial', data.get('Revenue Retail', 0))
    pnl.cogs = data.get('COGS', 0)
    pnl.cogs_wholesale = data.get('COGS Wholesale', 0)
    pnl.cogs_retail = data.get('COGS Retial', data.get('COGS Retail', 0))
    pnl.gross_profit = pnl.revenue - pnl.cogs
    
    # Budget sheet has NO G&A/Selling/Depreciation → flag it
    flags.append(DataQualityFlag('BUDGET_NO_OPEX', 'HIGH', 
        'Budget sheet has Revenue and COGS only — G&A and Selling expenses must come from Base sheet'))
    
    return pnl, flags


def parse_base_sheet(ws) -> Tuple[float, float, float, float, float, List[ExpenseCategory], List[ExpenseCategory], List[DataQualityFlag]]:
    """
    Parse the Base sheet (GL journal entries) to extract G&A, Selling, Depreciation.
    Returns: (selling_total, admin_total, depreciation, non_op_income, non_op_expense, selling_detail, admin_detail, flags)
    """
    selling_total = 0.0
    admin_total = 0.0
    depreciation = 0.0
    non_op_income = 0.0
    non_op_expense = 0.0
    
    selling_cats = {}
    admin_cats = {}
    flags = []
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 36:
            continue
        
        acct_dr = str(row[4] or '') if row[4] else ''   # Column E = Account Dr
        amount = safe_float(row[18])                      # Column S = Сумма
        classification = str(row[28] or '') if len(row) > 28 and row[28] else ''  # Column AC
        cost_class = str(row[35] or '') if len(row) > 35 and row[35] else ''      # Column AJ
        
        if not acct_dr or amount == 0:
            continue
        
        # Selling expenses (account 7310.xx)
        if acct_dr.startswith('7310'):
            selling_total += amount
            if cost_class:
                selling_cats[cost_class] = selling_cats.get(cost_class, 0) + amount
        
        # Admin expenses (account 7410.xx)
        elif acct_dr.startswith('7410'):
            admin_total += amount
            if cost_class:
                admin_cats[cost_class] = admin_cats.get(cost_class, 0) + amount
        
        # Non-operating expenses (account 8220.xx)
        elif acct_dr.startswith('8220'):
            non_op_expense += amount
        
        # Non-operating income (account 8110)  
        elif acct_dr.startswith('8110'):
            non_op_income += amount
        
        # Depreciation detection from cost classification
        if cost_class == 'Depreciation and Amortization':
            depreciation += amount
    
    selling_list = [ExpenseCategory(k, v) for k, v in sorted(selling_cats.items(), key=lambda x: -x[1])]
    admin_list = [ExpenseCategory(k, v) for k, v in sorted(admin_cats.items(), key=lambda x: -x[1])]
    
    return selling_total, admin_total, depreciation, non_op_income, non_op_expense, selling_list, admin_list, flags


def parse_revenue_breakdown(ws) -> Tuple[List[RevenueItem], float, float, float, List[DataQualityFlag]]:
    """
    Parse Revenue Breakdown sheet.
    Returns: (items, wholesale_total, retail_total, other_total, flags)
    """
    items = []
    flags = []
    wholesale = 0.0
    retail = 0.0
    other = 0.0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        product = str(row[0] or '').strip() if row[0] else ''
        gross = safe_float(row[1])
        vat = safe_float(row[2]) if len(row) > 2 else 0
        net = safe_float(row[3]) if len(row) > 3 else 0
        category = str(row[4] or '').strip() if len(row) > 4 else ''
        eliminated = str(row[5] or '').strip() if len(row) > 5 else ''
        
        # Skip total row
        if product.lower() in ('итог', 'итого', 'total', ''):
            continue
        
        # Skip eliminated (intercompany)
        if eliminated.lower() == 'eliminated':
            continue
        
        # Skip if no revenue
        if net == 0 and gross == 0:
            continue
        
        # Detect junk rows: dates, account codes
        if re.match(r'^\d{2}\.\d{2}\.\d{4}', product):  # Date pattern
            flags.append(DataQualityFlag('JUNK_ROW', 'LOW', f'Date-like row skipped: {product[:30]}'))
            continue
        if re.match(r'^\d{4}\.\d{2}', product):  # Account code pattern
            flags.append(DataQualityFlag('JUNK_ROW', 'LOW', f'Account code row skipped: {product[:30]}'))
            continue
        
        items.append(RevenueItem(
            product=product,
            gross_amount=gross,
            vat=vat,
            net_revenue=net if net else gross - vat,
            category=category
        ))
        
        if 'Wholesale' in category:
            wholesale += (net if net else gross - vat)
        elif 'Retail' in category:
            retail += (net if net else gross - vat)
        elif 'Other' in category:
            other += (net if net else gross - vat)
    
    return items, wholesale, retail, other, flags


"""
PRODUCT → CATEGORY MAPPING (from Word document instructions)
These define which products belong to which Revenue/COGS sub-categories.
"""
PRODUCT_CATEGORY_MAP = {
    # Revenue/COGS Wholesale Petrol
    'ევრო რეგულარი (იმპორტი)': 'Wholesale Petrol',
    'პრემიუმი (რეექსპორტი)': 'Wholesale Petrol',
    'სუპერი (რეექსპორტი)': 'Wholesale Petrol',
    'ევრო რეგულარი (საბითუმო)': 'Wholesale Petrol',
    # Revenue/COGS Wholesale Diesel
    'დიზელი (საბითუმო)': 'Wholesale Diesel',
    'ევროდიზელი (ექსპორტი)': 'Wholesale Diesel',
    'ევროდიზელი  (ექსპორტი)': 'Wholesale Diesel',
    # Revenue/COGS Wholesale Bitumen
    'ბიტუმი (საბითუმო)': 'Wholesale Bitumen',
    # Revenue/COGS Retail Petrol
    'ევრო რეგულარი': 'Retail Petrol',
    'პრემიუმი': 'Retail Petrol',
    'პრემიუმი ': 'Retail Petrol',
    'სუპერი': 'Retail Petrol',
    'სუპერი ': 'Retail Petrol',
    # Revenue/COGS Retail Diesel
    'დიზელი': 'Retail Diesel',
    'ევრო დიზელი': 'Retail Diesel',
    # Revenue/COGS Retail CNG
    'ბუნებრივი აირი': 'Retail CNG',
    'ბუნებრივი აირი (საბითუმო)': 'Retail CNG',
    # Revenue/COGS Retail LPG
    'თხევადი აირი (მხოლოდ SGP !!!)': 'Retail LPG',
    # Wholesale intermediate
    'პრემიუმი (საბითუმო)': 'Wholesale Petrol',
}


def parse_cogs_breakdown(ws) -> Tuple[List[COGSItem], float, float, List[DataQualityFlag]]:
    """
    Parse COGS Breakdown sheet using the correct 1C accounting rule:
    COGS per product = Account 6 + Account 7310 + Account 8230 (from credit turnover columns)
    
    The header row contains account codes as column headers (e.g., 6, 7310, 8230).
    Different file formats have different column positions, so we detect dynamically.
    """
    items = []
    flags = []
    cogs_retail = 0.0
    cogs_wholesale = 0.0
    
    # Step 1: Find header row and identify column positions by account code
    header_row = 1
    acct_6_col = None
    acct_7310_col = None
    acct_8230_col = None
    category_col = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        if any(str(v or '') == 'Субконто' for v in row):
            header_row = row_idx
            for col_idx, v in enumerate(row):
                sv = str(v or '').strip()
                if sv == '6':
                    acct_6_col = col_idx
                elif sv == '7310':
                    acct_7310_col = col_idx
                elif sv == '8230':
                    acct_8230_col = col_idx
                elif sv == 'q':
                    category_col = col_idx
            break
    
    # Find category column from data if not in header
    if category_col is None:
        for row in ws.iter_rows(min_row=header_row + 1, max_row=min(header_row + 5, ws.max_row), values_only=True):
            for ci, v in enumerate(row):
                if str(v or '').startswith('COGS '):
                    category_col = ci
                    break
            if category_col:
                break
    
    if acct_6_col is None:
        flags.append(DataQualityFlag('COGS_NO_ACCT6', 'HIGH', 'Cannot find account 6 column in COGS Breakdown header'))
        return items, cogs_retail, cogs_wholesale, flags
    
    # Step 2: Parse product rows
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
        product = str(row[0] or '').strip() if row[0] else ''
        if not product or product.lower() in ('итого', 'итог', 'total'):
            continue
        
        category = str(row[category_col] or '').strip() if category_col and len(row) > category_col else ''
        if category not in ('COGS Retail', 'COGS Wholesale'):
            continue
        
        # COGS = account 6 + account 7310 + account 8230
        amt_6 = safe_float(row[acct_6_col]) if acct_6_col and len(row) > acct_6_col else 0
        amt_7310 = safe_float(row[acct_7310_col]) if acct_7310_col and len(row) > acct_7310_col else 0
        amt_8230 = safe_float(row[acct_8230_col]) if acct_8230_col and len(row) > acct_8230_col else 0
        
        total_cogs = amt_6 + amt_7310 + amt_8230
        
        if total_cogs > 0:
            # Determine sub-category from PRODUCT_CATEGORY_MAP
            sub_cat = PRODUCT_CATEGORY_MAP.get(product, PRODUCT_CATEGORY_MAP.get(product.strip(), ''))
            
            items.append(COGSItem(product=product, amount=total_cogs, category=category))
            if 'Retail' in category:
                cogs_retail += total_cogs
            elif 'Wholesale' in category:
                cogs_wholesale += total_cogs
    
    return items, cogs_retail, cogs_wholesale, flags


def parse_bs_sheet(ws) -> Tuple[BalanceSheet, List[DataQualityFlag]]:
    """
    Parse the clean BS sheet (English line items, two columns).
    """
    bs = BalanceSheet()
    flags = []
    data = {}
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = str(row[0] or '').strip() if row[0] else ''
        value = safe_float(row[1]) if len(row) > 1 else 0
        if label:
            data[label] = value
    
    # Assets
    bs.ppe_net = abs(data.get('Property, plant and equipment', 0))
    bs.ppe_cost = abs(data.get('PPE Cost', 0))
    bs.ppe_depreciation = data.get('PPE Depreciation', 0)  # Negative
    bs.right_of_use = abs(data.get('Right of use asset', 0))
    bs.investment_properties = abs(data.get('Investment properties', 0))
    bs.investments = abs(data.get('Investments', 0))
    bs.intangible_assets = abs(data.get('Intangible assets', 0))
    bs.non_current_assets = abs(data.get('Total non-current assets', 0))
    
    bs.inventories = abs(data.get('Inventories', 0))
    bs.trade_receivables = abs(data.get('Trade receivables', 0))
    bs.tax_assets = data.get('Tax assets', 0)
    bs.prepayments = abs(data.get('Prepayments and other receivables', 0))
    bs.short_term_loans_receivable = abs(data.get('Short term loans receivable', 0))
    bs.cash = abs(data.get('Cash and cash equivalents', 0))
    bs.current_assets = abs(data.get('Total current assets', 0))
    
    bs.total_assets = abs(data.get('Total assets', 0))
    
    # Liabilities (stored as negative in BS sheet)
    bs.lease_liability_nc = abs(data.get('Lease liability non current portion', 0))
    bs.non_current_liabilities = abs(data.get('Total non-current liabilities', 0))
    
    bs.short_term_loans = abs(data.get('Short-term loans and borrowings', 0))
    bs.other_taxes_payable = abs(data.get('Other taxes payable', 0))
    bs.trade_payables = abs(data.get('Trade and other payables', 0))
    bs.advances_received = abs(data.get('Advances received', 0))
    bs.long_term_loans = abs(data.get('Long-Term Loans Payable', 0))
    bs.current_liabilities = abs(data.get('Total short-term liabilities', 0))
    
    bs.total_liabilities = abs(data.get('Total liabilities', 0))
    
    # Equity
    bs.share_capital = abs(data.get('Share capital', 0))
    bs.retained_earnings = data.get('Retained earnings', 0)  # Keep sign
    bs.net_income_period = data.get('Net income for the Period', 0)
    bs.revaluation_reserve = data.get('Revaluation reserve', 0)
    bs.total_equity = abs(data.get('Total equity attributable to shareholders', 
                          data.get('Total equity', 0)))
    
    # Validation
    check = data.get('Check', None)
    if check is not None and abs(safe_float(check)) > 1:
        flags.append(DataQualityFlag('BS_CHECK_FAIL', 'CRITICAL', 
            f'Balance Sheet Check cell = {check} (should be 0)'))
    
    if bs.total_assets == 0:
        flags.append(DataQualityFlag('MISSING_BS', 'HIGH', 'BS sheet has no asset data'))
    
    return bs, flags


# ══════════════════════════════════════════════════════════════
# MAIN PARSER
# ══════════════════════════════════════════════════════════════

def parse_nyx_excel(filepath: str, filename: str = "") -> ParseResult:
    """
    Universal parser for any NYX Core Thinker 1C Excel export.
    Detects file type, routes to correct parsers, validates, returns complete financial data.
    """
    result = ParseResult()
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        result.data_quality_flags.append(
            DataQualityFlag('OPEN_FAILED', 'CRITICAL', f'Cannot open file: {str(e)}'))
        return result
    
    result.sheets_available = wb.sheetnames
    
    # ── Detect company ──
    result.company = settings.COMPANY_NAME  # Default
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=5, values_only=True):
            for cell in row:
                s = str(cell or '')
                if 'სოკარ' in s.lower() or 'nyx' in s.lower() or 'SGP' in str(cell or ''):
                    if 'პეტროლეუმი' in s or 'Петролеум' in s:
                        result.company = settings.COMPANY_NAME
                    elif 'ენერჯი' in s or 'Энерджи' in s:
                        result.company = settings.COMPANY_NAME
                    break
    
    # ── Detect period ──
    # Try TDSheet first
    if 'TDSheet' in wb.sheetnames:
        ws = wb['TDSheet']
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=10, values_only=True):
            for cell in row:
                p = detect_period_from_text(str(cell or ''))
                if p:
                    result.period = p
                    result.period_source = "TDSheet"
                    break
            if result.period:
                break
    
    # Try COGS Breakdown
    if not result.period and 'COGS Breakdown' in wb.sheetnames:
        ws = wb['COGS Breakdown']
        for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
            for cell in row:
                p = detect_period_from_text(str(cell or ''))
                if p:
                    result.period = p
                    result.period_source = "COGS Breakdown"
                    break
            if result.period:
                break
    
    # Try Base sheet month column
    if not result.period and 'Base' in wb.sheetnames:
        ws = wb['Base']
        for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
            date_val = str(row[0] or '')
            month_val = str(row[30] or '') if len(row) > 30 else ''  # Column AE
            date_match = re.match(r'(\d{2})\.(\d{2})\.(\d{4})', date_val)
            if date_match:
                result.period = f"{date_match.group(3)}-{date_match.group(2)}"
                result.period_source = "Base sheet date"
                break
    
    # Try filename
    if not result.period:
        p = detect_period_from_filename(filename or filepath.split('/')[-1])
        if p:
            result.period = p
            result.period_source = "filename"
    
    if not result.period:
        result.data_quality_flags.append(
            DataQualityFlag('NO_PERIOD', 'HIGH', 'Could not detect period from file'))
        result.period = datetime.now().strftime('%Y-%m')
        result.period_source = "default (current month)"
    
    # ══════════════════════════════════════════════
    # P&L EXTRACTION — Route based on available sheets
    # ══════════════════════════════════════════════
    
    has_mapping = 'Mapping' in wb.sheetnames
    has_budget = any('Budget' in s or 'budget' in s.lower() for s in wb.sheetnames)
    has_base = 'Base' in wb.sheetnames
    
    if has_mapping:
        # ── ROUTE A: Mapping sheet = complete P&L ──
        result.file_type = "B"
        pnl, selling_detail, admin_detail, flags = parse_mapping_sheet(wb['Mapping'])
        result.pnl = pnl
        result.selling_expense_detail = selling_detail
        result.admin_expense_detail = admin_detail
        result.data_quality_flags.extend(flags)
        result.sheets_used.append('Mapping')
        
    elif has_budget:
        # ── ROUTE B: Budget + Base ──
        result.file_type = "A"
        budget_name = next(s for s in wb.sheetnames if 'Budget' in s or 'budget' in s.lower())
        pnl, flags = parse_budget_sheet(wb[budget_name])
        result.pnl = pnl
        result.data_quality_flags.extend(flags)
        result.sheets_used.append(budget_name)
        
        # Supplement with Base sheet for G&A, Selling, Depreciation
        if has_base:
            selling, admin, depr, noi, noe, sell_det, adm_det, base_flags = parse_base_sheet(wb['Base'])
            result.pnl.selling_expenses = selling
            result.pnl.admin_expenses = admin
            result.pnl.depreciation = depr
            result.pnl.non_operating_income = noi
            result.pnl.non_operating_expense = noe
            result.pnl.total_opex = selling + admin
            result.pnl.gross_profit = result.pnl.revenue - result.pnl.cogs
            result.pnl.ebitda = result.pnl.gross_profit - result.pnl.total_opex
            result.pnl.ebit = result.pnl.ebitda - depr
            result.pnl.profit_before_tax = result.pnl.ebit + noi - noe
            result.pnl.net_profit = result.pnl.profit_before_tax
            result.selling_expense_detail = sell_det
            result.admin_expense_detail = adm_det
            result.data_quality_flags.extend(base_flags)
            result.sheets_used.append('Base')
        else:
            # No G&A data available — critical flag
            result.pnl.ebitda = result.pnl.gross_profit
            result.pnl.net_profit = result.pnl.gross_profit
            result.data_quality_flags.append(
                DataQualityFlag('MISSING_GA', 'CRITICAL', 
                    'No Base or Mapping sheet — G&A expenses unknown, EBITDA = Gross Profit'))
    else:
        result.data_quality_flags.append(
            DataQualityFlag('NO_PNL_SOURCE', 'CRITICAL', 'No Mapping or Budget sheet found'))
    
    # ══════════════════════════════════════════════
    # REVENUE BREAKDOWN
    # ══════════════════════════════════════════════
    
    if 'Revenue Breakdown' in wb.sheetnames:
        items, wholesale, retail, other_rev, flags = parse_revenue_breakdown(wb['Revenue Breakdown'])
        result.revenue_breakdown = items
        result.pnl.revenue_wholesale = wholesale
        result.pnl.revenue_retail = retail
        result.pnl.revenue_other = other_rev
        result.data_quality_flags.extend(flags)
        result.sheets_used.append('Revenue Breakdown')
        
        # Cross-validate revenue
        rev_bd_total = wholesale + retail + other_rev
        if result.pnl.revenue > 0 and abs(rev_bd_total - result.pnl.revenue) > 1000:
            result.data_quality_flags.append(
                DataQualityFlag('REVENUE_MISMATCH', 'MEDIUM',
                    f'Revenue Breakdown total ({rev_bd_total:,.0f}) ≠ P&L revenue ({result.pnl.revenue:,.0f})'))
    
    # ══════════════════════════════════════════════
    # COGS BREAKDOWN
    # ══════════════════════════════════════════════
    
    if 'COGS Breakdown' in wb.sheetnames:
        items, cogs_ret, cogs_whl, flags = parse_cogs_breakdown(wb['COGS Breakdown'])
        result.cogs_breakdown = items
        result.pnl.cogs_retail = cogs_ret
        result.pnl.cogs_wholesale = cogs_whl
        result.data_quality_flags.extend(flags)
        result.sheets_used.append('COGS Breakdown')
    
    # ══════════════════════════════════════════════
    # BALANCE SHEET
    # ══════════════════════════════════════════════
    
    if 'BS' in wb.sheetnames:
        bs, flags = parse_bs_sheet(wb['BS'])
        result.balance_sheet = bs
        result.data_quality_flags.extend(flags)
        result.sheets_used.append('BS')
    elif 'Balance' in wb.sheetnames:
        # TODO: Parse detailed Balance sheet by MAPPING GRP aggregation
        result.data_quality_flags.append(
            DataQualityFlag('BS_FROM_BALANCE', 'MEDIUM', 
                'Using detailed Balance sheet — aggregation may need review'))
        result.sheets_used.append('Balance')
    else:
        result.data_quality_flags.append(
            DataQualityFlag('MISSING_BS', 'HIGH', 'No Balance Sheet data in file'))
    
    # ══════════════════════════════════════════════
    # DATA QUALITY SCORE
    # ══════════════════════════════════════════════
    
    score = 100
    for flag in result.data_quality_flags:
        if flag.severity == 'CRITICAL':
            score -= 25
        elif flag.severity == 'HIGH':
            score -= 15
        elif flag.severity == 'MEDIUM':
            score -= 5
        elif flag.severity == 'LOW':
            score -= 1
    
    result.data_quality_score = max(0, min(100, score))
    result.success = result.pnl.revenue > 0
    
    wb.close()
    return result


# ══════════════════════════════════════════════════════════════
# TEST
# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys
    
    files = [
        ("/mnt/user-data/uploads/Report-_January_2025.xlsx", "Report- January 2025.xlsx"),
        ("/mnt/user-data/uploads/January_2026__1_.xlsx", "January 2026 (1).xlsx"),
    ]
    
    for fpath, fname in files:
        print(f"\n{'='*80}")
        print(f"PARSING: {fname}")
        print(f"{'='*80}")
        
        r = parse_nyx_excel(fpath, fname)
        
        print(f"\nCompany:  {r.company}")
        print(f"Period:   {r.period} (from {r.period_source})")
        print(f"Type:     File type {r.file_type}")
        print(f"Sheets:   {r.sheets_available}")
        print(f"Used:     {r.sheets_used}")
        print(f"Quality:  {r.data_quality_score}/100")
        
        if r.data_quality_flags:
            print(f"\nFlags:")
            for f in r.data_quality_flags:
                print(f"  [{f.severity}] {f.code}: {f.message}")
        
        p = r.pnl
        print(f"\n{'─'*50}")
        print(f"INCOME STATEMENT")
        print(f"{'─'*50}")
        print(f"  Revenue:                    {p.revenue:>15,.0f}")
        print(f"    Wholesale:                {p.revenue_wholesale:>15,.0f}")
        print(f"    Retail:                   {p.revenue_retail:>15,.0f}")
        print(f"    Other:                    {p.revenue_other:>15,.0f}")
        print(f"  COGS:                      -{p.cogs:>15,.0f}")
        print(f"    Wholesale:               -{p.cogs_wholesale:>15,.0f}")
        print(f"    Retail:                  -{p.cogs_retail:>15,.0f}")
        print(f"  ─────────────────────────────────────────")
        print(f"  GROSS PROFIT:               {p.gross_profit:>15,.0f}  ({p.gross_profit/p.revenue*100 if p.revenue else 0:.1f}%)")
        print(f"  Selling Expenses (7310):   -{p.selling_expenses:>15,.0f}")
        print(f"  Admin Expenses (7410):     -{p.admin_expenses:>15,.0f}")
        print(f"  ─────────────────────────────────────────")
        print(f"  EBITDA:                     {p.ebitda:>15,.0f}  ({p.ebitda/p.revenue*100 if p.revenue else 0:.1f}%)")
        print(f"  Depreciation & Amort.:     -{p.depreciation:>15,.0f}")
        print(f"  ─────────────────────────────────────────")
        print(f"  EBIT:                       {p.ebit:>15,.0f}")
        print(f"  Non-Op Income:              {p.non_operating_income:>15,.0f}")
        print(f"  Non-Op Expense:            -{p.non_operating_expense:>15,.0f}")
        print(f"  ─────────────────────────────────────────")
        print(f"  PROFIT BEFORE TAX:          {p.profit_before_tax:>15,.0f}  ({p.profit_before_tax/p.revenue*100 if p.revenue else 0:.1f}%)")
        print(f"  NET PROFIT:                 {p.net_profit:>15,.0f}")
        
        if r.selling_expense_detail:
            print(f"\n  Selling Expense Detail:")
            for cat in r.selling_expense_detail[:10]:
                print(f"    {cat.name:40s} {cat.amount:>12,.0f}")
        
        if r.admin_expense_detail:
            print(f"\n  Admin Expense Detail:")
            for cat in r.admin_expense_detail[:10]:
                print(f"    {cat.name:40s} {cat.amount:>12,.0f}")
        
        b = r.balance_sheet
        if b.total_assets > 0:
            print(f"\n{'─'*50}")
            print(f"BALANCE SHEET")
            print(f"{'─'*50}")
            print(f"  ASSETS")
            print(f"    Non-current assets:       {b.non_current_assets:>15,.0f}")
            print(f"      PPE (net):              {b.ppe_net:>15,.0f}")
            print(f"      Right of use:           {b.right_of_use:>15,.0f}")
            print(f"      Investments:            {b.investments:>15,.0f}")
            print(f"    Current assets:           {b.current_assets:>15,.0f}")
            print(f"      Cash:                   {b.cash:>15,.0f}")
            print(f"      Inventories:            {b.inventories:>15,.0f}")
            print(f"      Trade receivables:      {b.trade_receivables:>15,.0f}")
            print(f"    TOTAL ASSETS:             {b.total_assets:>15,.0f}")
            print(f"  LIABILITIES")
            print(f"    Non-current:              {b.non_current_liabilities:>15,.0f}")
            print(f"    Current:                  {b.current_liabilities:>15,.0f}")
            print(f"      Trade payables:         {b.trade_payables:>15,.0f}")
            print(f"      Long-term loans:        {b.long_term_loans:>15,.0f}")
            print(f"    TOTAL LIABILITIES:        {b.total_liabilities:>15,.0f}")
            print(f"  EQUITY")
            print(f"    Share capital:            {b.share_capital:>15,.0f}")
            print(f"    Retained earnings:        {b.retained_earnings:>15,.0f}")
            print(f"    Net income (period):      {b.net_income_period:>15,.0f}")
            print(f"    TOTAL EQUITY:             {b.total_equity:>15,.0f}")
            print(f"  ─────────────────────────────────────────")
            de = b.total_liabilities / b.total_equity if b.total_equity else 0
            cr = b.current_assets / b.current_liabilities if b.current_liabilities else 0
            print(f"  D/E Ratio:                  {de:>15.2f}x")
            print(f"  Current Ratio:              {cr:>15.2f}x")
        
        print(f"\n  Revenue products: {len(r.revenue_breakdown)}")
        print(f"  COGS products:   {len(r.cogs_breakdown)}")
