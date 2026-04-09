"""
onec_interpreter.py — Deep 1C Accounting System Interpreter
=============================================================
Parses any 1C Chart of Accounts export (AccountN / ПланСчетов) and
produces a fully-structured, IFRS-mapped account tree with:
  • bilingual name splitting (Georgian // Russian)
  • Georgian boolean flag decoding (დიახ / არა)
  • account type → normal balance mapping (А / П / АП)
  • 44 subkonto dimension classification
  • parent-child account hierarchy (4-digit tree)
  • IFRS classification with BS/PL routing
  • off-balance, currency, quantity flag tracking
  • corrupted-data detection and sanitisation

Key classes:
  OneCAccount           — single parsed account (dataclass)
  OneCInterpreter       — file loader + full interpreter
  AccountHierarchyTree  — tree of OneCAccounts with navigation helpers

Usage:
    interp = OneCInterpreter()
    tree   = interp.parse_file("1c AccountN.xlsx")
    pl_accounts  = tree.filter(ifrs_section="income_statement")
    bs_accounts  = tree.filter(ifrs_section="balance_sheet")
    children_of  = tree.children("6110")
"""
from __future__ import annotations

import re
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ── Georgian boolean values ────────────────────────────────────────────────────
_GEO_YES = {"დიახ", "yes", "да", "true", "1", "дa"}
_GEO_NO  = {"არა",  "no",  "нет", "false", "0"}

# ── Account type → normal balance ─────────────────────────────────────────────
# 1C uses А = Active (debit-normal), П = Passive (credit-normal),
# АП = Mixed (context-dependent). We also track Russian/English variants.
_ACCT_TYPE_MAP: Dict[str, str] = {
    "А":  "Active",  "A": "Active",
    "П":  "Passive", "P": "Passive",
    "АП": "Mixed",   "AP": "Mixed",
}

_TYPE_TO_NORMAL_BALANCE: Dict[str, str] = {
    "Active":  "debit",
    "Passive": "credit",
    "Mixed":   "both",
    "Unknown": "unknown",
}

# ── IFRS classification by first account digit ────────────────────────────────
# Based on the Georgian National IFRS COA structure observed in the real data.
_CLASS_IFRS: Dict[str, Dict] = {
    "0": {"section": "balance_sheet",     "bs_side": "asset",    "bs_sub": "off_balance",  "pl_line": None,              "bs_line": "Off-Balance Assets"},
    "1": {"section": "balance_sheet",     "bs_side": "asset",    "bs_sub": "current",      "pl_line": None,              "bs_line": "Current Assets"},
    "2": {"section": "balance_sheet",     "bs_side": "asset",    "bs_sub": "noncurrent",   "pl_line": None,              "bs_line": "Non-Current Assets"},
    "3": {"section": "balance_sheet",     "bs_side": "liability","bs_sub": "current",      "pl_line": None,              "bs_line": "Current Liabilities"},
    "4": {"section": "balance_sheet",     "bs_side": "liability","bs_sub": "noncurrent",   "pl_line": None,              "bs_line": "Non-Current Liabilities"},
    "5": {"section": "balance_sheet",     "bs_side": "equity",   "bs_sub": "equity",       "pl_line": None,              "bs_line": "Equity"},
    "6": {"section": "income_statement",  "bs_side": None,       "bs_sub": None,           "pl_line": "Revenue",         "bs_line": None},
    "7": {"section": "income_statement",  "bs_side": None,       "bs_sub": None,           "pl_line": "Expenses",        "bs_line": None},
    "8": {"section": "income_statement",  "bs_side": None,       "bs_sub": None,           "pl_line": "Other Income",    "bs_line": None},
    "9": {"section": "income_statement",  "bs_side": None,       "bs_sub": None,           "pl_line": "Tax/Deferred",    "bs_line": None},
}

# Standard 1C Russian 2-digit accounts — mapped to IFRS
_RUSSIAN_1C_IFRS: Dict[str, Dict] = {
    "01": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Fixed Assets"},
    "02": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Accumulated Depreciation"},
    "03": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Investment Property"},
    "04": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Intangible Assets"},
    "05": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Amortisation"},
    "07": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Equipment in Installation"},
    "08": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "noncurrent",  "pl_line": None, "bs_line": "Capital WIP"},
    "10": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "current",     "pl_line": None, "bs_line": "Inventory"},
    "20": {"section": "income_statement", "bs_side": None,       "bs_sub": None,          "pl_line": "COGS", "bs_line": None},
    "41": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "current",     "pl_line": None, "bs_line": "Goods for Sale"},
    "50": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "current",     "pl_line": None, "bs_line": "Cash"},
    "51": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "current",     "pl_line": None, "bs_line": "Bank Accounts"},
    "60": {"section": "balance_sheet",    "bs_side": "liability","bs_sub": "current",     "pl_line": None, "bs_line": "Accounts Payable"},
    "62": {"section": "balance_sheet",    "bs_side": "asset",    "bs_sub": "current",     "pl_line": None, "bs_line": "Accounts Receivable"},
    "68": {"section": "balance_sheet",    "bs_side": "liability","bs_sub": "current",     "pl_line": None, "bs_line": "Tax Payable"},
    "70": {"section": "balance_sheet",    "bs_side": "liability","bs_sub": "current",     "pl_line": None, "bs_line": "Payroll Payable"},
    "76": {"section": "balance_sheet",    "bs_side": "liability","bs_sub": "current",     "pl_line": None, "bs_line": "Other Payables"},
    "80": {"section": "balance_sheet",    "bs_side": "equity",   "bs_sub": "equity",      "pl_line": None, "bs_line": "Share Capital"},
    "84": {"section": "balance_sheet",    "bs_side": "equity",   "bs_sub": "equity",      "pl_line": None, "bs_line": "Retained Earnings"},
    "90": {"section": "income_statement", "bs_side": None,       "bs_sub": None,          "pl_line": "Revenue", "bs_line": None},
    "91": {"section": "income_statement", "bs_side": None,       "bs_sub": None,          "pl_line": "Other Income", "bs_line": None},
    "99": {"section": "income_statement", "bs_side": None,       "bs_sub": None,          "pl_line": "Net P&L", "bs_line": None},
}

# ── Subkonto dimension semantic classification ─────────────────────────────────
_SUBKONTO_SEMANTICS: Dict[str, str] = {
    "Контрагенты":                          "counterparty",
    "Договоры":                             "contract",
    "Номенклатура":                         "product_item",
    "Номенклатурные группы":                "product_group",
    "Подразделения":                        "department",
    "Обособленные подразделения":           "subdivision",
    "Склады":                               "warehouse",
    "Основные средства":                    "fixed_asset",
    "Нематериальные активы":               "intangible_asset",
    "Работники организации":               "employee",
    "Банковские счета":                    "bank_account",
    "Статьи затрат":                       "cost_item",
    "Статьи движения денежных средств":    "cash_flow_item",
    "Прочие доходы и расходы":             "other_income_expense",
    "Прибыли и убытки":                    "profit_loss",
    "Расходы будущих периодов":            "deferred_expense",
    "Доходы будущих периодов":             "deferred_income",
    "Расходы на НИОКР":                    "rnd_expense",
    "Объекты строительства":               "construction_object",
    "Способы строительства":               "construction_method",
    "Документы реализации":                "sales_document",
    "Счета-фактуры выданные":              "issued_invoice",
    "Счета-фактуры полученные":            "received_invoice",
    "Исполнительные документы":            "legal_document",
    "Виды активов и обязательств":         "asset_liability_type",
    "Виды платежей в бюджет (фонды)":      "budget_payment_type",
    "Ставки НДС":                          "vat_rate",
    "Денежные документы":                  "cash_document",
    "Ценные бумаги":                       "security",
    "Комиссионеры":                        "consignee",
    "Бланки строгой отчетности":           "strict_form",
    "Уровни бюджетов":                     "budget_level",
    "Назначение целевых средств":          "target_fund_purpose",
    "Направления использования прибыли":   "profit_use",
    "(не используется)":                   "unused",
    "(об) Номенклатура":                   "product_item_rev",
    "(об) Номенклатурные группы":          "product_group_rev",
    "(об) Подразделения":                  "department_rev",
    "(об) Статьи затрат":                  "cost_item_rev",
    "(об) Прочие доходы и расходы":        "other_income_expense_rev",
    "(об) Прибыли и убытки":               "profit_loss_rev",
    "(об) Ставки НДС":                     "vat_rate_rev",
    "(об) Статьи движения денежных средств": "cash_flow_item_rev",
    "(об) Виды стоимости":                 "value_type_rev",
    "(об) Движения целевых средств":       "target_fund_movement_rev",
    "(об) Виды расчетов по средствам ФСС": "fss_payment_type_rev",
    "(об) Оценочные обязательства":        "provision_rev",
    "Оценочные обязательства и резервы":   "provision_reserve",
}


@dataclass
class OneCAccount:
    """Single parsed account from 1C Chart of Accounts."""
    code: str                                # e.g. "6110", "50.01", "11XX"
    code_normalized: str                     # digits only, e.g. "6110"
    name_ka: str                             # Georgian name
    name_ru: str                             # Russian name
    name_full: str                           # original bilingual string
    is_group: bool                           # True if this is a group/parent header (e.g. "11XX")
    account_type: str                        # "Active", "Passive", "Mixed", "Unknown"
    normal_balance: str                      # "debit", "credit", "both", "unknown"
    is_off_balance: bool                     # Забалансовый
    tracks_currency: bool                    # Вал.
    tracks_quantity: bool                    # Кол.
    subkonto: List[str]                      # raw subkonto dimension names (up to 3)
    subkonto_semantics: List[str]            # semantic types: "counterparty", "contract", etc.
    # IFRS
    ifrs_section: str                        # "balance_sheet" or "income_statement"
    ifrs_bs_side: Optional[str]             # "asset", "liability", "equity"
    ifrs_bs_sub: Optional[str]              # "current", "noncurrent", "equity", "off_balance"
    ifrs_bs_line: Optional[str]             # e.g. "Current Assets"
    ifrs_pl_line: Optional[str]             # e.g. "Revenue", "COGS"
    # Hierarchy
    class_digit: str                         # "1" ... "9" or "0"
    parent_code: Optional[str]              # inferred parent account code
    depth: int                               # 0=top-level class, 1=group, 2=postable
    # Metadata
    quick_select: Optional[str]             # Быстрый выбор alias
    is_corrupted: bool = False               # name looks garbled


@dataclass
class AccountHierarchyTree:
    """Tree of parsed 1C accounts with navigation helpers."""
    accounts: List[OneCAccount] = field(default_factory=list)
    _by_code: Dict[str, OneCAccount] = field(default_factory=dict, repr=False)

    def __post_init__(self):
        self._by_code = {a.code: a for a in self.accounts}

    def get(self, code: str) -> Optional[OneCAccount]:
        return self._by_code.get(code)

    def children(self, parent_code: str) -> List[OneCAccount]:
        return [a for a in self.accounts if a.parent_code == parent_code]

    def filter(self, **kwargs) -> List[OneCAccount]:
        """Filter accounts by any field. E.g. filter(ifrs_section='income_statement')"""
        result = self.accounts
        for k, v in kwargs.items():
            result = [a for a in result if getattr(a, k, None) == v]
        return result

    def postable(self) -> List[OneCAccount]:
        """Return only leaf/postable accounts (not group headers)."""
        return [a for a in self.accounts if not a.is_group]

    def by_pl_line(self, line: str) -> List[OneCAccount]:
        """Get all P&L accounts for a given line (Revenue, COGS, Expenses, etc.)"""
        return [a for a in self.accounts if a.ifrs_pl_line == line and not a.is_group]

    def by_bs_side(self, side: str, sub: Optional[str] = None) -> List[OneCAccount]:
        """Get BS accounts: side='asset'/'liability'/'equity', sub='current'/'noncurrent'/etc."""
        result = [a for a in self.accounts if a.ifrs_bs_side == side and not a.is_group]
        if sub:
            result = [a for a in result if a.ifrs_bs_sub == sub]
        return result

    def dimensions_for(self, code: str) -> List[Dict]:
        """Return analytical dimension types for an account."""
        acct = self._by_code.get(code)
        if not acct:
            return []
        return [
            {"raw": raw, "semantic": sem}
            for raw, sem in zip(acct.subkonto, acct.subkonto_semantics)
        ]

    def summary(self) -> Dict:
        """Return count summary by section."""
        return {
            "total": len(self.accounts),
            "postable": sum(1 for a in self.accounts if not a.is_group),
            "groups": sum(1 for a in self.accounts if a.is_group),
            "off_balance": sum(1 for a in self.accounts if a.is_off_balance),
            "with_dimensions": sum(1 for a in self.accounts if a.subkonto),
            "income_statement": sum(1 for a in self.accounts if a.ifrs_section == "income_statement" and not a.is_group),
            "balance_sheet": sum(1 for a in self.accounts if a.ifrs_section == "balance_sheet" and not a.is_group),
            "currency_tracked": sum(1 for a in self.accounts if a.tracks_currency),
            "quantity_tracked": sum(1 for a in self.accounts if a.tracks_quantity),
        }

    def to_kg_entities(self) -> List[Dict]:
        """Convert tree to KG entity dicts for knowledge_graph ingestion."""
        entities = []
        for a in self.postable():
            entities.append({
                "entity_id":    f"account_{a.code_normalized}",
                "entity_type":  "account",
                "name":         a.name_ka or a.name_ru or a.code,
                "description":  _build_description(a),
                "attributes": {
                    "code": a.code,
                    "name_ka": a.name_ka,
                    "name_ru": a.name_ru,
                    "account_type": a.account_type,
                    "normal_balance": a.normal_balance,
                    "ifrs_section": a.ifrs_section,
                    "ifrs_bs_side": a.ifrs_bs_side,
                    "ifrs_bs_sub": a.ifrs_bs_sub,
                    "ifrs_pl_line": a.ifrs_pl_line,
                    "dimensions": a.subkonto_semantics,
                },
            })
        return entities


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_bool(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return s in _GEO_YES or s.lower() in {"yes", "да", "true", "1"}


def _split_bilingual(name: str) -> Tuple[str, str]:
    """'Georgian // Russian' → (ka, ru). Returns (name, '') if no separator."""
    if not name:
        return ("", "")
    if "//" in name:
        parts = name.split("//", 1)
        return (parts[0].strip(), parts[1].strip())
    return (name.strip(), "")


def _is_group_code(code: str) -> bool:
    """Detect group/header codes like '11XX', '141X', '62.X'."""
    return bool(re.search(r'[Xx]', code))


def _normalize_code(code: str) -> str:
    """Keep only digits and dots for prefix matching."""
    return re.sub(r'[^0-9.]', '', code)


def _is_corrupted_name(name: str) -> bool:
    """Detect keyboard-layout-shifted Cyrillic (qwerty typed as Russian)."""
    if not name:
        return False
    # Heuristic: many consecutive lowercase Latin chars that look like shifted Cyrillic
    latin_runs = re.findall(r'[a-z]{4,}', name)
    return len(latin_runs) >= 2


def _class_digit(code: str) -> str:
    """Extract the leading class digit from a code."""
    norm = _normalize_code(code)
    if norm:
        return norm[0]
    return ""


def _infer_ifrs(code: str, account_type: str, is_off_balance: bool) -> Dict:
    """Derive IFRS classification from code + type."""
    norm = _normalize_code(code)
    digit = norm[0] if norm else ""

    # Two-digit Russian 1C accounts (01, 02, ... 99)
    if len(norm) <= 2 and not is_off_balance:
        two_digit = norm[:2]
        for prefix, mapping in _RUSSIAN_1C_IFRS.items():
            if two_digit == prefix or norm.startswith(prefix):
                return dict(mapping)

    # Off-balance accounts
    if is_off_balance:
        return _CLASS_IFRS.get("0", {}).copy()

    # Georgian 4-digit IFRS accounts
    ifrs = _CLASS_IFRS.get(digit, {
        "section": "balance_sheet", "bs_side": "asset",
        "bs_sub": "current", "pl_line": None, "bs_line": "Other"
    }).copy()

    # Override: Active account in liability class → treat as asset (e.g. Input VAT 3340)
    if account_type == "Active" and ifrs.get("bs_side") == "liability":
        ifrs["bs_side"] = "asset"
    # Override: Passive account in asset class → treat as liability (e.g. contra-asset)
    if account_type == "Passive" and ifrs.get("bs_side") == "asset" and digit in ("1", "2"):
        ifrs["bs_side"] = "liability"
        ifrs["bs_sub"] = "current"

    return ifrs


def _infer_parent(code: str, all_codes: List[str]) -> Optional[str]:
    """Find closest parent group code for a given code."""
    norm = _normalize_code(code)
    if not norm:
        return None
    # Look for group codes whose digit prefix matches, e.g. 6110 → "61XX" or "6XXX"
    candidates = []
    for c in all_codes:
        if _is_group_code(c):
            pattern = _normalize_code(c).replace(".", r"\.")
            regex   = pattern.replace("x", r"\d").replace("X", r"\d")
            try:
                if re.fullmatch(regex, norm):
                    candidates.append(c)
            except re.error:
                pass
    if not candidates:
        return None
    # Return most specific (longest numeric prefix)
    return max(candidates, key=lambda c: len(_normalize_code(c)))


def _classify_subkonto(raw: Optional[str]) -> str:
    if not raw:
        return ""
    return _SUBKONTO_SEMANTICS.get(raw, "custom_dimension")


def _build_description(a: OneCAccount) -> str:
    parts = []
    if a.name_ka:
        parts.append(a.name_ka)
    if a.name_ru and a.name_ru != a.name_ka:
        parts.append(f"({a.name_ru})")
    parts.append(f"Account {a.code}")
    if a.ifrs_pl_line:
        parts.append(f"P&L: {a.ifrs_pl_line}")
    if a.ifrs_bs_line:
        parts.append(f"BS: {a.ifrs_bs_line}")
    if a.subkonto_semantics:
        parts.append(f"Dimensions: {', '.join(a.subkonto_semantics)}")
    return " | ".join(parts)


# ── Main Interpreter ──────────────────────────────────────────────────────────

class OneCInterpreter:
    """
    Parses a 1C Chart of Accounts Excel file and produces a fully-structured
    AccountHierarchyTree with IFRS mappings and analytical dimension types.

    Supported formats:
      - Standard 1C AccountN export (ПланСчетов)
      - Any variant with columns: Код, Наименование, Заб., Акт., Вал., Кол., Субконто 1-3
      - Bilingual Georgian // Russian names
      - Georgian boolean flags (დიახ / არა)

    Usage:
        interp = OneCInterpreter()
        tree   = interp.parse_file("1c AccountN.xlsx")
        # or from raw list of dicts:
        tree   = interp.parse_records([{"code": "6110", "name": "Revenue", ...}])
    """

    # ── Column auto-detection keywords ────────────────────────────────────────
    _COL_KEYWORDS = {
        "code":      ["код",    "code",   "account",  "счет"],
        "quick":     ["быстрый", "quick", "alias"],
        "name":      ["наименование", "name",  "назва"],
        "off_bal":   ["заб",    "offbal", "забалансов"],
        "type":      ["акт",    "type",   "тип",  "активн"],
        "currency":  ["вал",    "currenc", "валют"],
        "quantity":  ["кол",    "qty",    "колич"],
        "sub1":      ["субконто 1", "subkonto1", "sub1"],
        "sub2":      ["субконто 2", "subkonto2", "sub2"],
        "sub3":      ["субконто 3", "subkonto3", "sub3"],
    }

    def parse_file(self, path: str) -> AccountHierarchyTree:
        """Load and parse a 1C COA Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required: pip install openpyxl")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            logger.warning("OneCInterpreter: empty file at %s", path)
            return AccountHierarchyTree()

        return self._parse_rows(rows)

    def parse_file_bytes(self, data: bytes, filename: str = "coa.xlsx") -> AccountHierarchyTree:
        """Parse from raw bytes (e.g., from HTTP upload)."""
        import io
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required")
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return self._parse_rows(rows)

    def _parse_rows(self, rows: list) -> AccountHierarchyTree:
        """Core parser: detect header, map columns, produce AccountHierarchyTree."""
        if not rows:
            return AccountHierarchyTree()

        # Find header row (first row with 'Код' or 'code' etc.)
        header_idx, col_map = self._detect_header(rows)
        if col_map is None:
            logger.warning("OneCInterpreter: could not detect header row")
            return AccountHierarchyTree()

        data_rows = rows[header_idx + 1:]
        raw_accounts = []
        all_codes: List[str] = []

        for row in data_rows:
            record = self._parse_row(row, col_map)
            if record is None:
                continue
            raw_accounts.append(record)
            all_codes.append(record["code"])

        accounts = [self._build_account(r, all_codes) for r in raw_accounts]
        tree = AccountHierarchyTree(accounts=accounts)

        logger.info(
            "OneCInterpreter: parsed %d accounts (%d postable, %d groups) from %d rows",
            len(accounts),
            sum(1 for a in accounts if not a.is_group),
            sum(1 for a in accounts if a.is_group),
            len(data_rows),
        )
        return tree

    def _detect_header(self, rows: list) -> Tuple[int, Optional[Dict[str, int]]]:
        """Find the header row and build column → index mapping."""
        for idx, row in enumerate(rows[:10]):
            col_map = self._map_columns(row)
            if col_map.get("code") is not None and col_map.get("name") is not None:
                return idx, col_map
        return 0, None

    def _map_columns(self, header_row) -> Dict[str, Optional[int]]:
        """Map semantic column names to column indices."""
        col_map: Dict[str, Optional[int]] = {k: None for k in self._COL_KEYWORDS}
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            cell_lower = str(cell).lower().strip()
            for field_name, keywords in self._COL_KEYWORDS.items():
                if any(kw in cell_lower for kw in keywords):
                    if col_map[field_name] is None:  # First match wins
                        col_map[field_name] = col_idx
        return col_map

    def _parse_row(self, row, col_map: Dict[str, Optional[int]]) -> Optional[Dict]:
        """Extract one raw record dict from a data row."""
        def _get(field: str):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        code_raw = _get("code")
        if code_raw is None:
            return None
        code = str(code_raw).strip()
        if not code:
            return None

        name_raw  = _get("name") or ""
        name_full = str(name_raw).strip()

        acct_type_raw = _get("type") or ""
        off_bal_raw   = _get("off_bal")
        cur_raw       = _get("currency")
        qty_raw       = _get("quantity")
        quick_raw     = _get("quick")
        sub1_raw      = _get("sub1")
        sub2_raw      = _get("sub2")
        sub3_raw      = _get("sub3")

        return {
            "code":       code,
            "name_full":  name_full,
            "acct_type":  str(acct_type_raw).strip(),
            "is_off_bal": _parse_bool(off_bal_raw),
            "tracks_cur": _parse_bool(cur_raw),
            "tracks_qty": _parse_bool(qty_raw),
            "quick":      str(quick_raw).strip() if quick_raw else None,
            "sub1":       str(sub1_raw).strip() if sub1_raw else None,
            "sub2":       str(sub2_raw).strip() if sub2_raw else None,
            "sub3":       str(sub3_raw).strip() if sub3_raw else None,
        }

    def _build_account(self, r: Dict, all_codes: List[str]) -> OneCAccount:
        """Convert a raw dict to a fully-classified OneCAccount."""
        code      = r["code"]
        norm      = _normalize_code(code)
        is_group  = _is_group_code(code)
        name_ka, name_ru = _split_bilingual(r["name_full"])

        acct_type_key = r["acct_type"]
        acct_type     = _ACCT_TYPE_MAP.get(acct_type_key, "Unknown")
        normal_bal    = _TYPE_TO_NORMAL_BALANCE[acct_type]

        is_off_bal  = r["is_off_bal"]
        tracks_cur  = r["tracks_cur"]
        tracks_qty  = r["tracks_qty"]
        quick       = r["quick"]

        subkonto_raw = [s for s in [r["sub1"], r["sub2"], r["sub3"]] if s]
        subkonto_sem = [_classify_subkonto(s) for s in subkonto_raw]

        ifrs = _infer_ifrs(code, acct_type, is_off_bal)
        section  = ifrs.get("section", "balance_sheet")
        bs_side  = ifrs.get("bs_side")
        bs_sub   = ifrs.get("bs_sub")
        bs_line  = ifrs.get("bs_line")
        pl_line  = ifrs.get("pl_line")

        parent = _infer_parent(code, all_codes) if not is_group else None

        # Depth: 0=class digit only (e.g. "6"), 1=group (11XX), 2=postable (6110)
        if len(norm) <= 1:
            depth = 0
        elif is_group:
            depth = 1
        else:
            depth = 2

        is_corrupted = _is_corrupted_name(r["name_full"])
        class_dig = _class_digit(code)

        return OneCAccount(
            code=code,
            code_normalized=norm,
            name_ka=name_ka,
            name_ru=name_ru,
            name_full=r["name_full"],
            is_group=is_group,
            account_type=acct_type,
            normal_balance=normal_bal,
            is_off_balance=is_off_bal,
            tracks_currency=tracks_cur,
            tracks_quantity=tracks_qty,
            subkonto=subkonto_raw,
            subkonto_semantics=subkonto_sem,
            ifrs_section=section,
            ifrs_bs_side=bs_side,
            ifrs_bs_sub=bs_sub,
            ifrs_bs_line=bs_line,
            ifrs_pl_line=pl_line,
            class_digit=class_dig,
            parent_code=parent,
            depth=depth,
            quick_select=quick,
            is_corrupted=is_corrupted,
        )

    def parse_records(self, records: List[Dict]) -> AccountHierarchyTree:
        """Build tree from a pre-parsed list of dicts (e.g., from DB or API)."""
        all_codes = [r.get("code", "") for r in records]
        accounts  = [self._build_account(r, all_codes) for r in records]
        return AccountHierarchyTree(accounts=accounts)


# ── Module-level singleton ────────────────────────────────────────────────────
onec_interpreter = OneCInterpreter()
