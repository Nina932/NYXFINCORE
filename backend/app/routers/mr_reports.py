"""
FinAI MR Reports Router — Management Reporting with USD conversion
Exchange rate CRUD, MR report generation from FinAI data, snapshot history,
13-sheet NYX Core Thinker Baku format Excel export.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from pydantic import BaseModel
from typing import Optional, List, Any
from app.database import get_db
from app.models.all_models import (
    ExchangeRate, MRReportSnapshot, Report,
    BalanceSheetItem, TrialBalanceItem, RevenueItem, COGSItem, GAExpenseItem, Dataset,
    COAMasterAccount, BudgetLine,
)
from app.services.mr_engine import MREngine, enrich_with_account_names
from app.services.mr_template import (
    MR_SHEETS, BS_COLUMNS, PL_COLUMNS, CFS_COLUMNS, PRODUCTS_COLUMNS,
    OPEX_COLUMNS, BORROWINGS_COLUMNS, RECEIVABLES_COLUMNS,
    PAYABLES_COLUMNS, PREPAYMENTS_COLUMNS,
)
from app.services.mr_mapping import (
    seed_coa_mr_mappings, populate_tb_mr_mappings,
    build_mapping_grp_baku_reference, build_full_mr_mapping,
    resolve_mr_code_for_account,
)
from app.config import settings
import logging, os
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/mr", tags=["mr-reports"])

UPLOAD_DIR = Path("./uploads")


def _load_mapping_sheet_rows(filename: str) -> list[list]:
    """
    Load the Mapping sheet from the uploaded Excel file.

    The Mapping sheet contains pre-classified P&L expense lines with:
      Col B (idx 1) = account code (e.g. "7310.01.1/1")
      Col C (idx 2) = account name in Georgian/Russian
      Col D (idx 3) = turnover amount
      Col E (idx 4) = MAPING ST classification (e.g. "Wages, benefits and payroll taxes")

    Returns the raw rows (list of lists) for the MREngine to process.
    Returns empty list if Mapping sheet not found.
    """
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        logger.warning(f"Mapping sheet: file not found at {file_path}")
        return []

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        mapping_sheet = None
        for name in wb.sheetnames:
            if name.lower().strip() == "mapping":
                mapping_sheet = wb[name]
                break

        if not mapping_sheet:
            wb.close()
            logger.info(f"Mapping sheet: no 'Mapping' sheet found in {filename}")
            return []

        rows = [list(r) for r in mapping_sheet.iter_rows(values_only=True)]
        wb.close()

        # Filter to only P&L rows with data
        pl_rows = []
        for row in rows:
            if not row or len(row) < 5:
                continue
            code = str(row[1] or "").strip()
            if not code or code[:1] not in "6789":
                continue
            pl_rows.append(row)

        logger.info(f"Mapping sheet: loaded {len(pl_rows)} P&L rows from {filename}")
        return pl_rows

    except Exception as e:
        logger.error(f"Mapping sheet load failed: {e}")
        return []


# ════════════════════════════════════════════════════════════════
# Exchange Rate CRUD
# ════════════════════════════════════════════════════════════════

class ExchangeRateCreate(BaseModel):
    rate: float
    rate_date: str           # "2025-01-15"
    period: Optional[str] = None  # "January 2025" — convenience field
    from_currency: str = "GEL"
    to_currency: str = "USD"
    source: str = "manual"


@router.post("/exchange-rates")
async def save_exchange_rate(payload: ExchangeRateCreate, db: AsyncSession = Depends(get_db)):
    """Save or update an exchange rate for a specific date."""
    er = ExchangeRate(
        from_currency=payload.from_currency,
        to_currency=payload.to_currency,
        rate=payload.rate,
        rate_date=payload.rate_date,
        source=payload.source,
    )
    db.add(er)
    await db.commit()
    await db.refresh(er)
    return er.to_dict()


@router.get("/exchange-rates")
async def list_exchange_rates(
    from_currency: str = "GEL",
    to_currency: str = "USD",
    db: AsyncSession = Depends(get_db)
):
    """List all exchange rates, most recent first."""
    q = select(ExchangeRate).where(
        ExchangeRate.from_currency == from_currency,
        ExchangeRate.to_currency == to_currency,
    ).order_by(ExchangeRate.rate_date.desc())
    result = await db.execute(q)
    rates = result.scalars().all()
    return [r.to_dict() for r in rates]


@router.delete("/exchange-rates/{rate_id}")
async def delete_exchange_rate(rate_id: int, db: AsyncSession = Depends(get_db)):
    """Delete an exchange rate."""
    r = (await db.execute(select(ExchangeRate).where(ExchangeRate.id == rate_id))).scalar_one_or_none()
    if not r:
        raise HTTPException(404, "Exchange rate not found")
    await db.delete(r)
    await db.commit()
    return {"message": "Deleted", "id": rate_id}


# ════════════════════════════════════════════════════════════════
# MR Report Generation — Uses MREngine
# ════════════════════════════════════════════════════════════════

class MRGenerateRequest(BaseModel):
    dataset_id: int
    period: str = "January 2025"
    exchange_rate: float          # GEL per 1 USD
    rate_date: Optional[str] = None
    prior_dataset_id: Optional[int] = None   # Optional: dataset ID for prior year comparison
    skip_prior_year: bool = False             # If True, skip prior year entirely (leave blank)
    include_budget: bool = True               # Auto-include budget data for plan column


@router.post("/generate")
async def generate_mr_report(payload: MRGenerateRequest, db: AsyncSession = Depends(get_db)):
    """
    Generate MR Report from existing FinAI data using the Baku Template Engine.
    Populates all 13-sheet Baku MR structure from TB/Revenue/COGS/GA data.
    """
    ds_id = payload.dataset_id
    rate = payload.exchange_rate
    period = payload.period

    # Verify dataset exists
    ds = (await db.execute(select(Dataset).where(Dataset.id == ds_id))).scalar_one_or_none()
    if not ds:
        raise HTTPException(404, "Dataset not found")

    # ── DatasetIntelligence: Analyze before fetching ────────────
    # Use the manifest to understand what data exists before querying
    from app.services.financial_intelligence import DatasetIntelligence
    intel = DatasetIntelligence(db)
    manifest = await intel.analyze_quick(ds_id)

    # ── Fetch source data (only what exists per manifest) ─────
    tb_items = []
    if not manifest or manifest.has_trial_balance:
        tb_items = (await db.execute(
            select(TrialBalanceItem).where(TrialBalanceItem.dataset_id == ds_id)
        )).scalars().all()

    rev_items = []
    if not manifest or manifest.has_revenue_detail:
        rev_items = (await db.execute(
            select(RevenueItem).where(RevenueItem.dataset_id == ds_id)
        )).scalars().all()

    cogs_items = []
    if not manifest or manifest.has_cogs_detail:
        cogs_items = (await db.execute(
            select(COGSItem).where(COGSItem.dataset_id == ds_id)
        )).scalars().all()

    ga_items = []
    if not manifest or manifest.has_ga_expenses:
        ga_items = (await db.execute(
            select(GAExpenseItem).where(GAExpenseItem.dataset_id == ds_id)
        )).scalars().all()

    logger.info(f"MR Generate: TB={len(tb_items)}, Rev={len(rev_items)}, "
                f"COGS={len(cogs_items)}, GA={len(ga_items)}, Rate={rate}"
                f" | Manifest: {manifest.summary if manifest else 'N/A'}")

    # ── Seed COA MR mappings if not already done ─────────────────
    coa_check = (await db.execute(
        select(COAMasterAccount).where(COAMasterAccount.baku_mr_code.isnot(None)).limit(1)
    )).scalar_one_or_none()
    if not coa_check:
        seed_result = await seed_coa_mr_mappings(db)
        logger.info(f"COA MR mapping seeded: {seed_result}")

    # ── Populate TB items with MR mappings from COA ─────────────
    tb_mapping_result = await populate_tb_mr_mappings(db, ds_id)
    logger.info(f"TB MR mapping: {tb_mapping_result}")

    # Re-fetch TB items with updated mappings
    tb_items = (await db.execute(
        select(TrialBalanceItem).where(TrialBalanceItem.dataset_id == ds_id)
    )).scalars().all()

    # ── Load Mapping sheet for P&L expense sub-classification ──
    # The Mapping sheet has pre-classified expense lines with IFRS categories
    # (Wages, Depreciation, Rent, etc.) that the MR engine uses to populate
    # P&L sub-items (02.A.01 Wages, 02.B.02 Depreciation, etc.)
    mapping_rows = _load_mapping_sheet_rows(ds.name)
    logger.info(f"MR Generate: Mapping sheet rows={len(mapping_rows)} for P&L sub-classification")

    # ── Fetch BSI items as fallback (if no Mapping sheet) ──
    bsi_items = []
    if not mapping_rows:
        bsi_items = (await db.execute(
            select(BalanceSheetItem).where(BalanceSheetItem.dataset_id == ds_id)
        )).scalars().all()
        logger.info(f"MR Generate: BSI={len(bsi_items)} items (Mapping sheet not found, using BSI fallback)")

    # ── AUTO-DETECT PRIOR YEAR DATASET (via manifest intelligence) ──
    prior_values = {}
    prior_period_label = ""

    prior_dataset_id = payload.prior_dataset_id
    if payload.skip_prior_year:
        prior_dataset_id = None
        logger.info("MR Generate: Prior year skipped by user request")
    elif not prior_dataset_id and manifest:
        # Use manifest's pre-computed prior year link (no extra queries)
        if manifest.prior_year_dataset_id and manifest.prior_year_has_data:
            prior_dataset_id = manifest.prior_year_dataset_id
            prior_period_label = manifest.prior_year_period or ""
            logger.info(f"MR Generate: Intelligence auto-linked prior year: DS #{prior_dataset_id} ({prior_period_label})")
        elif manifest.prior_year_dataset_id:
            logger.info(f"MR Generate: Prior year DS #{manifest.prior_year_dataset_id} exists but has no usable data")
    elif not prior_dataset_id:
        # Fallback if manifest not available
        from app.services.financial_intelligence import DatasetDiscovery
        discovery = DatasetDiscovery(db)
        prior_ds = await discovery.find_prior_year_dataset(ds.period)
        if prior_ds:
            prior_dataset_id = prior_ds.id
            prior_period_label = prior_ds.period
            logger.info(f"MR Generate: Fallback auto-detected prior year: {prior_ds.name} ({prior_ds.period})")

    if prior_dataset_id and prior_dataset_id != ds_id:
        try:
            # Fetch prior dataset's TB, Revenue, COGS, GA
            prior_tb = (await db.execute(
                select(TrialBalanceItem).where(TrialBalanceItem.dataset_id == prior_dataset_id)
            )).scalars().all()
            prior_rev = (await db.execute(
                select(RevenueItem).where(RevenueItem.dataset_id == prior_dataset_id)
            )).scalars().all()
            prior_cogs = (await db.execute(
                select(COGSItem).where(COGSItem.dataset_id == prior_dataset_id)
            )).scalars().all()
            prior_ga = (await db.execute(
                select(GAExpenseItem).where(GAExpenseItem.dataset_id == prior_dataset_id)
            )).scalars().all()

            # Load mapping sheet for prior dataset if available
            prior_ds_obj = (await db.execute(
                select(Dataset).where(Dataset.id == prior_dataset_id)
            )).scalar_one_or_none()
            prior_mapping_rows = _load_mapping_sheet_rows(prior_ds_obj.name) if prior_ds_obj else []
            prior_bsi = []
            if not prior_mapping_rows:
                prior_bsi = (await db.execute(
                    select(BalanceSheetItem).where(BalanceSheetItem.dataset_id == prior_dataset_id)
                )).scalars().all()

            # Run a second MREngine on prior-year data if any useful data exists
            has_prior_data = prior_tb or prior_rev or prior_cogs or prior_ga
            if has_prior_data:
                prior_engine = MREngine(
                    tb_items=prior_tb, bsi_items=prior_bsi, mapping_rows=prior_mapping_rows,
                    rev_items=prior_rev, cogs_items=prior_cogs, ga_items=prior_ga,
                    rate=rate,  # Use same exchange rate for comparability
                )
                prior_sections = prior_engine.populate_all()

                # Extract {code: actual_usd_k} from all prior sections
                for section_key in ["bs", "pl", "cfs", "products_wholesale",
                                     "products_retail", "products_gas_distr", "opex"]:
                    for row in prior_sections.get(section_key, []):
                        if isinstance(row, dict) and row.get("code"):
                            val = row.get("actual_usd_k", 0.0)
                            if val:  # Only store non-zero values
                                prior_values[row["code"]] = val

                if not prior_period_label and prior_ds_obj:
                    prior_period_label = prior_ds_obj.period
                data_types = []
                if prior_tb: data_types.append(f"{len(prior_tb)} TB")
                if prior_rev: data_types.append(f"{len(prior_rev)} Rev")
                if prior_cogs: data_types.append(f"{len(prior_cogs)} COGS")
                if prior_ga: data_types.append(f"{len(prior_ga)} GA")
                logger.info(f"MR Generate: Prior year values extracted for {len(prior_values)} codes from {prior_period_label} ({', '.join(data_types)})")
            else:
                logger.warning(f"MR Generate: Prior dataset {prior_dataset_id} has no usable data (TB/Rev/COGS/GA all empty)")
        except Exception as e:
            logger.warning(f"MR Generate: Failed to load prior year data: {e}")

    # ── LOAD BUDGET DATA FOR PLAN COLUMN (manifest-driven) ──
    budget_values = {}
    if payload.include_budget:
        try:
            # Use manifest to know WHERE budget data lives (current or prior dataset)
            budget_source_id = ds_id
            budget_source = f"DS {ds_id}"

            if manifest and manifest.budget_source_dataset_id:
                budget_source_id = manifest.budget_source_dataset_id
                budget_source = f"DS {budget_source_id} ({manifest.budget_source})"
                logger.info(f"MR Generate: Intelligence resolved budget source: {budget_source}")

            budget_lines = (await db.execute(
                select(BudgetLine).where(BudgetLine.dataset_id == budget_source_id)
            )).scalars().all()

            # Fallback if manifest missed it — try prior dataset directly
            if not budget_lines and prior_dataset_id and prior_dataset_id != budget_source_id:
                budget_lines = (await db.execute(
                    select(BudgetLine).where(BudgetLine.dataset_id == prior_dataset_id)
                )).scalars().all()
                budget_source = f"prior DS {prior_dataset_id} (fallback)"

            # Map BudgetLine.line_item to MR template codes
            BUDGET_LINE_TO_MR_CODE = {
                "Revenue": "01",
                "Revenue Retial": "01.A",
                "Revenue Retail": "01.A",
                "Revenue Wholesale": "01.A",
                "COGS": "02.A",
                "COGS Retial": "02.A",
                "COGS Retail": "02.A",
                "COGS Wholesale": "02.A",
                "GA Expenses": "02.B",
                "G&A Expenses": "02.B",
                "D&A Expenses": "02.B.02",
                "Labour Costs": "02.B.01",
            }

            for bl in budget_lines:
                line = (bl.line_item or "").strip()
                budget_amt = bl.budget_amount or 0
                mr_code = BUDGET_LINE_TO_MR_CODE.get(line)
                if mr_code and budget_amt:
                    budget_values[mr_code] = budget_values.get(mr_code, 0) + budget_amt

            if budget_values:
                logger.info(f"MR Generate: Budget values mapped for {len(budget_values)} MR codes from {budget_source}")
        except Exception as e:
            logger.warning(f"MR Generate: Failed to load budget data: {e}")

    # ── Run MR Engine (with prior year + budget intelligence) ─
    engine = MREngine(
        tb_items=tb_items,
        bsi_items=bsi_items,
        mapping_rows=mapping_rows,
        rev_items=rev_items,
        cogs_items=cogs_items,
        ga_items=ga_items,
        rate=rate,
        prior_values=prior_values,
        budget_values=budget_values,
    )
    sections = engine.populate_all()

    # Enrich TB extract sections with account names
    for key in ["borrowings", "receivables", "payables", "prepayments"]:
        if key in sections:
            sections[key] = enrich_with_account_names(sections[key], tb_items)

    # ── Build MAPPING GRP / MAPPING BAKU cross-reference ────────
    mapping_ref = build_mapping_grp_baku_reference(tb_items)
    sections["mapping_grp_baku"] = mapping_ref

    # ── Intelligence Metadata ────────────────────────────────────
    # Tell the user exactly what data sources were used and what's missing
    sections["_intelligence"] = {
        "data_used": {
            "trial_balance": len(tb_items),
            "revenue": len(rev_items),
            "cogs": len(cogs_items),
            "ga_expenses": len(ga_items),
        },
        "prior_year": {
            "dataset_id": prior_dataset_id,
            "period": prior_period_label,
            "values_extracted": len(prior_values),
        } if prior_dataset_id else None,
        "budget": {
            "source": budget_source if budget_values else None,
            "codes_mapped": len(budget_values),
        } if payload.include_budget else None,
        "report_capabilities": manifest.report_capabilities if manifest else {},
        "warnings": manifest.missing if manifest else [],
        "revenue_source": manifest.revenue_source if manifest else "unknown",
        "cogs_source": manifest.cogs_source if manifest else "unknown",
        "richness_score": manifest.data_richness_score if manifest else 0,
    }

    # ── Save exchange rate if not already saved ─────────────────
    rate_date = payload.rate_date or datetime.now().strftime("%Y-%m-%d")
    existing_rate = (await db.execute(
        select(ExchangeRate).where(
            ExchangeRate.rate_date == rate_date,
            ExchangeRate.from_currency == "GEL",
            ExchangeRate.to_currency == "USD",
        )
    )).scalars().first()
    if not existing_rate:
        db.add(ExchangeRate(
            from_currency="GEL", to_currency="USD",
            rate=rate, rate_date=rate_date, source="mr_generation"
        ))

    # ── Save Report ────────────────────────────────────────────
    summary = sections.get("summary", {})
    report = Report(
        title=f"MR Report - {period}",
        report_type="mr",
        period=period,
        currency="USD",
        company=settings.COMPANY_NAME,
        kpis={
            "revenue_usd_k": summary.get("total_revenue_usd_k", 0),
            "gross_profit_usd_k": summary.get("gross_profit_usd_k", 0),
            "operating_profit_usd_k": summary.get("operating_profit_usd_k", 0),
            "net_income_usd_k": summary.get("net_income_usd_k", 0),
            "exchange_rate": rate,
        },
        summary=f"MR for {period}. Rate: {rate} GEL/USD. Revenue: {summary.get('total_revenue_usd_k', 0)}K USD",
        source_dataset_id=ds_id,
        generated_by="mr_module",
    )
    db.add(report)
    await db.flush()

    # ── Save MR Snapshot ──────────────────────────────────────
    snapshot = MRReportSnapshot(
        dataset_id=ds_id,
        report_id=report.id,
        period=period,
        exchange_rate=rate,
        rate_date=rate_date,
        currency="USD",
        sections=sections,
        generated_by="user",
    )
    db.add(snapshot)
    await db.commit()
    await db.refresh(snapshot)

    return snapshot.to_dict()


# ════════════════════════════════════════════════════════════════
# MR Snapshots — List / Get / Delete
# ════════════════════════════════════════════════════════════════

@router.get("/snapshots")
async def list_snapshots(
    dataset_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db)
):
    """List all MR report snapshots, newest first."""
    q = select(MRReportSnapshot).order_by(MRReportSnapshot.created_at.desc())
    if dataset_id:
        q = q.where(MRReportSnapshot.dataset_id == dataset_id)
    result = await db.execute(q)
    return [s.to_dict() for s in result.scalars().all()]


@router.get("/snapshots/{snapshot_id}")
async def get_snapshot(snapshot_id: int, db: AsyncSession = Depends(get_db)):
    """Get a single MR report snapshot with full sections."""
    s = (await db.execute(
        select(MRReportSnapshot).where(MRReportSnapshot.id == snapshot_id)
    )).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Snapshot not found")
    return s.to_dict()


@router.delete("/snapshots/{snapshot_id}")
async def delete_snapshot(snapshot_id: int, db: AsyncSession = Depends(get_db)):
    """Delete an MR report snapshot (and its associated Report)."""
    s = (await db.execute(
        select(MRReportSnapshot).where(MRReportSnapshot.id == snapshot_id)
    )).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Snapshot not found")
    if s.report_id:
        rpt = (await db.execute(select(Report).where(Report.id == s.report_id))).scalar_one_or_none()
        if rpt:
            await db.delete(rpt)
    await db.delete(s)
    await db.commit()
    return {"message": "Deleted", "id": snapshot_id}


# ════════════════════════════════════════════════════════════════
# Excel Export — 13-Sheet NYX Core Thinker Baku Format
# ════════════════════════════════════════════════════════════════

@router.get("/snapshots/{snapshot_id}/export")
async def export_snapshot_excel(snapshot_id: int, db: AsyncSession = Depends(get_db)):
    """Export MR Report snapshot to Excel in full 13-sheet NYX Core Thinker Baku format."""
    s = (await db.execute(
        select(MRReportSnapshot).where(MRReportSnapshot.id == snapshot_id)
    )).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Snapshot not found")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

        wb = openpyxl.Workbook()
        sections = s.sections or {}
        period = s.period
        rate = s.exchange_rate

        # ── Style Definitions ──────────────────────────────────
        title_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
        subtitle_font = Font(name="Calibri", size=10, color="4472C4", italic=True)
        header_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        bold_row_font = Font(name="Calibri", size=9, bold=True)
        normal_font = Font(name="Calibri", size=9)
        total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        section_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7'),
        )
        num_fmt = '#,##0'
        num_fmt_1 = '#,##0.0'
        pct_fmt = '0.0%'

        def write_entity_header(ws, title_text, col_count=9):
            """Write standard entity header rows."""
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
            ws.cell(row=1, column=1, value=settings.COMPANY_NAME).font = title_font
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
            ws.cell(row=2, column=1, value=title_text).font = title_font
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
            ws.cell(row=3, column=1,
                    value=f"Period: {period}  |  Exchange rate: {rate} GEL/USD  |  Currency: USD (thousands)"
            ).font = subtitle_font

        def write_column_headers(ws, headers, start_row=5):
            """Write column header row with styling."""
            for ci, hdr in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=ci, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

        def write_baku_data_row(ws, row_num, row_data, value_col=6, has_code=True):
            """Write a Baku-format data row (Code | Line | +/- | ... | Actual | Plan | Dev | Dev%)."""
            code = row_data.get("code", "")
            line = row_data.get("line", "")
            sign = row_data.get("sign", "")
            is_bold = row_data.get("bold", False)
            level = row_data.get("level", 0)
            actual = row_data.get("actual_usd_k", 0)

            font = bold_row_font if is_bold else normal_font
            fill = total_fill if is_bold and level == 0 else (section_fill if is_bold else None)

            # Indentation
            indent_prefix = "  " * level

            col = 1
            if has_code:
                c = ws.cell(row=row_num, column=col, value=code)
                c.font = font
                c.border = thin_border
                if fill: c.fill = fill
                col += 1

            # Line item with indentation
            c = ws.cell(row=row_num, column=col, value=f"{indent_prefix}{line}")
            c.font = font
            c.border = thin_border
            if fill: c.fill = fill
            col += 1

            # +/- sign
            c = ws.cell(row=row_num, column=col, value=sign)
            c.font = font
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border
            if fill: c.fill = fill
            col += 1

            # Opening / Previous year — NOW POPULATED from prior year engine
            prev_year = row_data.get("prev_year_usd_k", 0) or 0
            opening = row_data.get("opening_usd_k", 0) or 0
            # For BS: use opening balance. For P&L/CFS/Products: use prior year
            prev_val = prev_year if prev_year else opening
            for i in range(value_col - col):
                if i == (value_col - col - 1) and prev_val:
                    # Last column before actual = the comparison column
                    c = ws.cell(row=row_num, column=col, value=prev_val)
                else:
                    c = ws.cell(row=row_num, column=col)
                c.border = thin_border
                c.number_format = num_fmt
                if fill: c.fill = fill
                col += 1

            # Current period actual
            c = ws.cell(row=row_num, column=col, value=actual if actual else None)
            c.font = font
            c.number_format = num_fmt
            c.border = thin_border
            if fill: c.fill = fill
            col += 1

            # Plan — NOW POPULATED from budget data
            plan = row_data.get("plan_usd_k", 0) or 0
            c = ws.cell(row=row_num, column=col, value=plan if plan else None)
            c.border = thin_border
            c.number_format = num_fmt
            if fill: c.fill = fill
            col += 1

            # Deviation absolute — NOW POPULATED
            dev_abs = row_data.get("deviation_abs", 0) or 0
            c = ws.cell(row=row_num, column=col, value=dev_abs if dev_abs else None)
            c.border = thin_border
            c.number_format = num_fmt
            if fill: c.fill = fill
            col += 1

            # Deviation % — NOW POPULATED
            dev_pct = row_data.get("deviation_pct", 0) or 0
            c = ws.cell(row=row_num, column=col, value=(dev_pct / 100) if dev_pct else None)
            c.border = thin_border
            c.number_format = pct_fmt
            if fill: c.fill = fill

        def set_column_widths(ws, widths):
            """Set column widths from a dict of {col_letter: width}."""
            for letter, width in widths.items():
                ws.column_dimensions[letter].width = width

        # ════════════════════════════════════════════════════════
        # Sheet 1: BS (Balance Sheet)
        # ════════════════════════════════════════════════════════
        ws = wb.active
        ws.title = "BS"
        ws.sheet_properties.tabColor = "38BDF8"
        write_entity_header(ws, "Statement of Financial Position (Balance Sheet)")
        write_column_headers(ws, BS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 50, "C": 6, "D": 18, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        bs_rows = sections.get("bs", [])
        if isinstance(bs_rows, list):
            for i, row_data in enumerate(bs_rows):
                write_baku_data_row(ws, 6 + i, row_data, value_col=6)
                # Also fill opening balance column (col D)
                opening = row_data.get("opening_usd_k", 0)
                if opening:
                    ws.cell(row=6 + i, column=4, value=opening).number_format = num_fmt

        # ════════════════════════════════════════════════════════
        # Sheet 2: P&L
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("P&L")
        ws.sheet_properties.tabColor = "22C55E"
        write_entity_header(ws, "Profit and Loss Statement", col_count=8)
        write_column_headers(ws, PL_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 50, "C": 6, "D": 18, "E": 18, "F": 18, "G": 16, "H": 12})

        pl_rows = sections.get("pl", [])
        if isinstance(pl_rows, list):
            for i, row_data in enumerate(pl_rows):
                write_baku_data_row(ws, 6 + i, row_data, value_col=5)

        # ════════════════════════════════════════════════════════
        # Sheet 3: CFS (ind) — Cash Flow Statement (V8 codes)
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("CFS (ind)")
        ws.sheet_properties.tabColor = "A855F7"
        write_entity_header(ws, "Cash Flow Statement (Indirect Method)", col_count=9)
        write_column_headers(ws, CFS_COLUMNS)
        set_column_widths(ws, {"A": 18, "B": 55, "C": 6, "D": 18, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        cfs_rows = sections.get("cfs", [])
        if isinstance(cfs_rows, list):
            for i, row_data in enumerate(cfs_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                level = row_data.get("level", 0)
                font = bold_row_font if is_bold else normal_font
                fill = total_fill if is_bold and level == 0 else (section_fill if is_bold else None)
                indent = "  " * level
                actual = row_data.get("actual_usd_k", 0)

                # Code column (V8 has CFS codes)
                c = ws.cell(row=row_num, column=1, value=row_data.get("code", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                # Line item with indentation
                c = ws.cell(row=row_num, column=2, value=f"{indent}{row_data.get('line', '')}")
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                # +/- sign
                c = ws.cell(row=row_num, column=3, value=row_data.get("sign", ""))
                c.font = font; c.border = thin_border; c.alignment = Alignment(horizontal="center")
                if fill: c.fill = fill

                # Previous year (empty)
                c = ws.cell(row=row_num, column=4); c.border = thin_border; c.number_format = num_fmt
                if fill: c.fill = fill

                # Actual
                c = ws.cell(row=row_num, column=5, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Plan, Dev abs, Dev%
                for col in [6, 7, 8]:
                    c = ws.cell(row=row_num, column=col); c.border = thin_border
                    c.number_format = num_fmt if col < 8 else pct_fmt
                    if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 4: Service revenue&COGS (empty template)
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Service revenue&COGS")
        ws.sheet_properties.tabColor = "F59E0B"
        write_entity_header(ws, "Service Revenue & COGS")
        write_column_headers(ws, PRODUCTS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 40, "C": 12, "D": 6, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})
        ws.cell(row=6, column=2, value="(Not applicable for SGP)").font = Font(name="Calibri", size=9, italic=True, color="808080")

        # ════════════════════════════════════════════════════════
        # Sheet 5: Products revenue&COGS_WSales_Tr (Wholesale)
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Products revenue&COGS_WSales_Tr")
        ws.sheet_properties.tabColor = "EF4444"
        write_entity_header(ws, "Products Revenue & COGS (Wholesale / Trading)")
        write_column_headers(ws, PRODUCTS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 40, "C": 12, "D": 6, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        ws_rows = sections.get("products_wholesale", [])
        if isinstance(ws_rows, list):
            for i, row_data in enumerate(ws_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font

                ws.cell(row=row_num, column=1, value=row_data.get("code", "")).font = font
                ws.cell(row=row_num, column=1).border = thin_border
                ws.cell(row=row_num, column=2, value=row_data.get("line", "")).font = font
                ws.cell(row=row_num, column=2).border = thin_border
                # Units
                ws.cell(row=row_num, column=3, value="thsd USD").font = normal_font
                ws.cell(row=row_num, column=3).border = thin_border
                # +/-
                ws.cell(row=row_num, column=4).border = thin_border
                # Prev year (empty)
                ws.cell(row=row_num, column=5).border = thin_border
                # Actual
                actual = row_data.get("actual_usd_k", 0)
                ws.cell(row=row_num, column=6, value=actual if actual else None).font = font
                ws.cell(row=row_num, column=6).number_format = num_fmt
                ws.cell(row=row_num, column=6).border = thin_border
                # Plan, Dev, Dev%
                for col in [7, 8, 9]:
                    ws.cell(row=row_num, column=col).border = thin_border

        # ════════════════════════════════════════════════════════
        # Sheet 6: Products revenue & COGS_Retail
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Products revenue & COGS_Retail")
        ws.sheet_properties.tabColor = "EC4899"
        write_entity_header(ws, "Products Revenue & COGS (Retail)")
        write_column_headers(ws, PRODUCTS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 40, "C": 12, "D": 6, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        rt_rows = sections.get("products_retail", [])
        if isinstance(rt_rows, list):
            for i, row_data in enumerate(rt_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font

                ws.cell(row=row_num, column=1, value=row_data.get("code", "")).font = font
                ws.cell(row=row_num, column=1).border = thin_border
                ws.cell(row=row_num, column=2, value=row_data.get("line", "")).font = font
                ws.cell(row=row_num, column=2).border = thin_border
                ws.cell(row=row_num, column=3, value="thsd USD").font = normal_font
                ws.cell(row=row_num, column=3).border = thin_border
                ws.cell(row=row_num, column=4).border = thin_border
                ws.cell(row=row_num, column=5).border = thin_border
                actual = row_data.get("actual_usd_k", 0)
                ws.cell(row=row_num, column=6, value=actual if actual else None).font = font
                ws.cell(row=row_num, column=6).number_format = num_fmt
                ws.cell(row=row_num, column=6).border = thin_border
                for col in [7, 8, 9]:
                    ws.cell(row=row_num, column=col).border = thin_border

        # ════════════════════════════════════════════════════════
        # Sheet 7: Products revenue&COGS_Gas Distr
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Products revenue&COGS_Gas Distr")
        ws.sheet_properties.tabColor = "10B981"
        write_entity_header(ws, "Products Revenue & COGS (Gas Distribution)")
        write_column_headers(ws, PRODUCTS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 40, "C": 12, "D": 6, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        gd_rows = sections.get("products_gas_distr", [])
        if isinstance(gd_rows, list):
            for i, row_data in enumerate(gd_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font

                ws.cell(row=row_num, column=1, value=row_data.get("code", "")).font = font
                ws.cell(row=row_num, column=1).border = thin_border
                ws.cell(row=row_num, column=2, value=row_data.get("line", "")).font = font
                ws.cell(row=row_num, column=2).border = thin_border
                ws.cell(row=row_num, column=3, value=row_data.get("units", "thsd USD")).font = normal_font
                ws.cell(row=row_num, column=3).border = thin_border
                ws.cell(row=row_num, column=4).border = thin_border
                ws.cell(row=row_num, column=5).border = thin_border
                actual = row_data.get("actual_usd_k", 0)
                ws.cell(row=row_num, column=6, value=actual if actual else None).font = font
                ws.cell(row=row_num, column=6).number_format = num_fmt
                ws.cell(row=row_num, column=6).border = thin_border
                for col in [7, 8, 9]:
                    ws.cell(row=row_num, column=col).border = thin_border

        # ════════════════════════════════════════════════════════
        # Sheet 8: CAPEX&INVESTMENT (empty template)
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("CAPEX&INVESTMENT")
        ws.sheet_properties.tabColor = "6366F1"
        write_entity_header(ws, "CAPEX & Investments")
        headers = ["#", "Project name", "Status", "Budget (thsd USD)", "Actual spent (thsd USD)",
                   "Remaining (thsd USD)", "Completion %", "Expected completion"]
        write_column_headers(ws, headers, start_row=5)
        set_column_widths(ws, {"A": 6, "B": 40, "C": 14, "D": 18, "E": 18, "F": 18, "G": 14, "H": 18})
        ws.cell(row=6, column=2, value="(Project-level CAPEX data to be populated)").font = Font(name="Calibri", size=9, italic=True, color="808080")

        # ════════════════════════════════════════════════════════
        # Sheet 8: CAPEX_cash basis (empty template)
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("CAPEX_cash basis")
        ws.sheet_properties.tabColor = "8B5CF6"
        write_entity_header(ws, "CAPEX (Cash Basis)")
        write_column_headers(ws, headers, start_row=5)
        set_column_widths(ws, {"A": 6, "B": 40, "C": 14, "D": 18, "E": 18, "F": 18, "G": 14, "H": 18})
        ws.cell(row=6, column=2, value="(Cash basis CAPEX data to be populated)").font = Font(name="Calibri", size=9, italic=True, color="808080")

        # ════════════════════════════════════════════════════════
        # Sheet 9: OPEX_breakdown
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("OPEX_breakdown")
        ws.sheet_properties.tabColor = "F97316"
        write_entity_header(ws, "OPEX Breakdown")
        write_column_headers(ws, OPEX_COLUMNS)
        set_column_widths(ws, {"A": 22, "B": 45, "C": 12, "D": 6, "E": 18, "F": 18, "G": 18, "H": 16, "I": 12})

        opex_rows = sections.get("opex", [])
        if isinstance(opex_rows, list):
            for i, row_data in enumerate(opex_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font
                fill = section_fill if is_bold else None

                c = ws.cell(row=row_num, column=1, value=row_data.get("segment", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=2, value=row_data.get("line", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=3, value="thsd USD")
                c.font = normal_font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=4); c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=5); c.border = thin_border
                if fill: c.fill = fill

                actual = row_data.get("actual_usd_k", 0)
                c = ws.cell(row=row_num, column=6, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                for col in [7, 8, 9]:
                    c = ws.cell(row=row_num, column=col); c.border = thin_border
                    if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 10: Borrowings
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Borrowings")
        ws.sheet_properties.tabColor = "14B8A6"
        write_entity_header(ws, "Borrowings", col_count=10)
        write_column_headers(ws, BORROWINGS_COLUMNS)
        set_column_widths(ws, {"A": 14, "B": 30, "C": 12, "D": 14, "E": 14, "F": 14, "G": 18, "H": 18, "I": 18, "J": 18})

        borrow_rows = sections.get("borrowings", [])
        if isinstance(borrow_rows, list):
            for i, row_data in enumerate(borrow_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font
                fill = total_fill if is_bold else None

                c = ws.cell(row=row_num, column=1, value=row_data.get("account_code", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=2, value=row_data.get("line", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                # Currency, dates, rate (cols 3-6 — empty for now)
                for col in range(3, 7):
                    c = ws.cell(row=row_num, column=col); c.border = thin_border
                    if fill: c.fill = fill

                # Outstanding balance original (col 7)
                c = ws.cell(row=row_num, column=7); c.border = thin_border
                if fill: c.fill = fill

                # Outstanding USD (col 8 = actual)
                actual = row_data.get("actual_usd_k", 0)
                c = ws.cell(row=row_num, column=8, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Principal, Interest (cols 9-10)
                for col in [9, 10]:
                    c = ws.cell(row=row_num, column=col); c.border = thin_border
                    if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 11: Receivables
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Receivables")
        ws.sheet_properties.tabColor = "06B6D4"
        write_entity_header(ws, "Receivables")
        write_column_headers(ws, RECEIVABLES_COLUMNS)
        set_column_widths(ws, {"A": 6, "B": 30, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16})

        recv_rows = sections.get("receivables", [])
        if isinstance(recv_rows, list):
            for i, row_data in enumerate(recv_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font
                fill = total_fill if is_bold else None

                c = ws.cell(row=row_num, column=1, value=i + 1 if not is_bold else "")
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=2, value=row_data.get("line", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=3); c.border = thin_border
                if fill: c.fill = fill

                # Year start
                opening = row_data.get("opening_usd_k", 0)
                c = ws.cell(row=row_num, column=4, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Period start (same as opening for now)
                c = ws.cell(row=row_num, column=5, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Supplied (empty)
                c = ws.cell(row=row_num, column=6); c.border = thin_border
                if fill: c.fill = fill

                # Payments (empty)
                c = ws.cell(row=row_num, column=7); c.border = thin_border
                if fill: c.fill = fill

                # Period end
                actual = row_data.get("actual_usd_k", 0)
                c = ws.cell(row=row_num, column=8, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Overdue
                c = ws.cell(row=row_num, column=9); c.border = thin_border
                if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 12: Payables
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Payables")
        ws.sheet_properties.tabColor = "D946EF"
        write_entity_header(ws, "Payables")
        write_column_headers(ws, PAYABLES_COLUMNS)
        set_column_widths(ws, {"A": 6, "B": 30, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16})

        pay_rows = sections.get("payables", [])
        if isinstance(pay_rows, list):
            for i, row_data in enumerate(pay_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font
                fill = total_fill if is_bold else None

                c = ws.cell(row=row_num, column=1, value=i + 1 if not is_bold else "")
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=2, value=row_data.get("line", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=3); c.border = thin_border
                if fill: c.fill = fill

                # Annual plan (empty)
                c = ws.cell(row=row_num, column=4); c.border = thin_border
                if fill: c.fill = fill

                # Year start balance
                opening = row_data.get("opening_usd_k", 0)
                c = ws.cell(row=row_num, column=5, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Period start
                c = ws.cell(row=row_num, column=6, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Purchases (empty)
                c = ws.cell(row=row_num, column=7); c.border = thin_border
                if fill: c.fill = fill

                # Payments (empty)
                c = ws.cell(row=row_num, column=8); c.border = thin_border
                if fill: c.fill = fill

                # Period end
                actual = row_data.get("actual_usd_k", 0)
                c = ws.cell(row=row_num, column=9, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 13: Prepayments
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("Prepayments")
        ws.sheet_properties.tabColor = "84CC16"
        write_entity_header(ws, "Prepayments")
        write_column_headers(ws, PREPAYMENTS_COLUMNS)
        set_column_widths(ws, {"A": 6, "B": 30, "C": 18, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16})

        prep_rows = sections.get("prepayments", [])
        if isinstance(prep_rows, list):
            for i, row_data in enumerate(prep_rows):
                row_num = 6 + i
                is_bold = row_data.get("bold", False)
                font = bold_row_font if is_bold else normal_font
                fill = total_fill if is_bold else None

                c = ws.cell(row=row_num, column=1, value=i + 1 if not is_bold else "")
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=2, value=row_data.get("line", ""))
                c.font = font; c.border = thin_border
                if fill: c.fill = fill

                c = ws.cell(row=row_num, column=3); c.border = thin_border
                if fill: c.fill = fill

                # Year start
                opening = row_data.get("opening_usd_k", 0)
                c = ws.cell(row=row_num, column=4, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Period start
                c = ws.cell(row=row_num, column=5, value=opening if opening else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Prepayments made (empty)
                c = ws.cell(row=row_num, column=6); c.border = thin_border
                if fill: c.fill = fill

                # Acceptance (empty)
                c = ws.cell(row=row_num, column=7); c.border = thin_border
                if fill: c.fill = fill

                # Period end
                actual = row_data.get("actual_usd_k", 0)
                c = ws.cell(row=row_num, column=8, value=actual if actual else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

                # Change
                change = row_data.get("change_usd_k", 0)
                c = ws.cell(row=row_num, column=9, value=change if change else None)
                c.font = font; c.number_format = num_fmt; c.border = thin_border
                if fill: c.fill = fill

        # ════════════════════════════════════════════════════════
        # Sheet 14: MAPPING GRP / MAPPING BAKU (BS)
        # Cross-reference: Account Code → IFRS Category → Baku MR Code
        # ════════════════════════════════════════════════════════
        ws = wb.create_sheet("MAPPING GRP_BAKU")
        ws.sheet_properties.tabColor = "0EA5E9"
        mapping_col_count = 8
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=mapping_col_count)
        ws.cell(row=1, column=1, value=settings.COMPANY_NAME).font = title_font
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mapping_col_count)
        ws.cell(row=2, column=1, value="MAPPING GRP → MAPPING BAKU (BS/PL) Cross-Reference").font = title_font
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=mapping_col_count)
        ws.cell(row=3, column=1,
                value=f"Period: {period}  |  Account Code → IFRS Classification → Baku MR Report Line"
        ).font = subtitle_font

        mapping_headers = [
            "Account Code", "Account Name", "MAPPING GRP (IFRS)",
            "MAPPING BAKU (MR Code)", "MR Line Item", "Statement",
            "Side", "Closing Balance (GEL)"
        ]
        write_column_headers(ws, mapping_headers, start_row=5)
        set_column_widths(ws, {
            "A": 16, "B": 40, "C": 30, "D": 22,
            "E": 40, "F": 10, "G": 12, "H": 20
        })

        mapping_rows = sections.get("mapping_grp_baku", [])
        if isinstance(mapping_rows, list):
            for i, row_data in enumerate(mapping_rows):
                row_num = 6 + i
                font = normal_font
                has_mr = bool(row_data.get("baku_mr_code"))

                # Account Code
                c = ws.cell(row=row_num, column=1, value=row_data.get("account_code", ""))
                c.font = font; c.border = thin_border

                # Account Name
                c = ws.cell(row=row_num, column=2, value=row_data.get("account_name", ""))
                c.font = font; c.border = thin_border

                # MAPPING GRP (IFRS)
                c = ws.cell(row=row_num, column=3, value=row_data.get("ifrs_mapping_grp", ""))
                c.font = font; c.border = thin_border
                if row_data.get("ifrs_mapping_grp"):
                    c.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

                # MAPPING BAKU (MR Code)
                c = ws.cell(row=row_num, column=4, value=row_data.get("baku_mr_code", ""))
                c.font = bold_row_font if has_mr else font
                c.border = thin_border
                if has_mr:
                    c.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

                # MR Line Item
                c = ws.cell(row=row_num, column=5, value=row_data.get("baku_mr_line", ""))
                c.font = font; c.border = thin_border
                if has_mr:
                    c.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

                # Statement
                c = ws.cell(row=row_num, column=6, value=row_data.get("statement", ""))
                c.font = font; c.border = thin_border; c.alignment = Alignment(horizontal="center")

                # Side
                c = ws.cell(row=row_num, column=7, value=row_data.get("side", ""))
                c.font = font; c.border = thin_border; c.alignment = Alignment(horizontal="center")

                # Closing Balance
                balance = row_data.get("closing_balance", 0)
                c = ws.cell(row=row_num, column=8, value=balance if balance else None)
                c.font = font; c.number_format = '#,##0'; c.border = thin_border

        # Add summary at bottom
        if mapping_rows:
            summary_row = 6 + len(mapping_rows) + 1
            ws.cell(row=summary_row, column=1, value="Summary:").font = bold_row_font
            mapped_count = sum(1 for r in mapping_rows if r.get("baku_mr_code"))
            ws.cell(row=summary_row + 1, column=1, value=f"Total accounts: {len(mapping_rows)}").font = normal_font
            ws.cell(row=summary_row + 2, column=1, value=f"Mapped to MR: {mapped_count}").font = normal_font
            ws.cell(row=summary_row + 3, column=1, value=f"Unmapped: {len(mapping_rows) - mapped_count}").font = normal_font

        # ── Save to exports directory ──────────────────────────
        export_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        filename = f"MR_Report_{s.period.replace(' ', '_')}_{s.id}.xlsx"
        filepath = os.path.join(export_dir, filename)
        wb.save(filepath)

        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )

    except Exception as e:
        logger.error(f"MR Excel export error: {e}", exc_info=True)
        raise HTTPException(500, f"Export failed: {str(e)}")


# ════════════════════════════════════════════════════════════════
# MR Mapping Management — Seed, View, Update
# ════════════════════════════════════════════════════════════════

@router.post("/mappings/seed")
async def seed_mr_mappings(db: AsyncSession = Depends(get_db)):
    """
    Seed/refresh Baku MR report codes on all COA Master accounts.
    This populates baku_mr_code, baku_mr_line, baku_mr_statement
    from the template-derived account prefix mappings.
    """
    result = await seed_coa_mr_mappings(db)
    await db.commit()
    return {"message": "COA MR mappings seeded", **result}


@router.post("/mappings/populate-tb/{dataset_id}")
async def populate_tb_mappings(dataset_id: int, db: AsyncSession = Depends(get_db)):
    """
    Populate MR mappings on all TB items for a dataset.
    Inherits from COA Master (if seeded) or falls back to template matching.
    """
    result = await populate_tb_mr_mappings(db, dataset_id)
    await db.commit()
    return {"message": f"TB MR mappings populated for dataset {dataset_id}", **result}


@router.get("/mappings/coa")
async def list_coa_mr_mappings(
    mapped_only: bool = False,
    statement: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """
    List all COA accounts with their MR mappings.
    Optionally filter by mapped_only=true or statement=BS/PL.
    """
    q = select(COAMasterAccount).order_by(COAMasterAccount.account_code)
    if mapped_only:
        q = q.where(COAMasterAccount.baku_mr_code.isnot(None))
    if statement:
        q = q.where(COAMasterAccount.baku_mr_statement == statement.upper())
    result = await db.execute(q)
    accounts = result.scalars().all()
    return [
        {
            "account_code": a.account_code,
            "name_ka": a.name_ka,
            "ifrs_bs_line": a.ifrs_bs_line,
            "ifrs_pl_line": a.ifrs_pl_line,
            "ifrs_side": a.ifrs_side,
            "baku_mr_code": a.baku_mr_code,
            "baku_mr_line": a.baku_mr_line,
            "baku_mr_statement": a.baku_mr_statement,
        }
        for a in accounts
    ]


@router.get("/mappings/tb/{dataset_id}")
async def list_tb_mr_mappings(dataset_id: int, db: AsyncSession = Depends(get_db)):
    """
    List all TB items for a dataset with their MR mappings.
    Shows the flow: Account Code → IFRS Classification → Baku MR Code.
    """
    tb_items = (await db.execute(
        select(TrialBalanceItem)
        .where(TrialBalanceItem.dataset_id == dataset_id)
        .order_by(TrialBalanceItem.account_code)
    )).scalars().all()

    return [
        {
            "account_code": t.account_code,
            "account_name": t.account_name,
            "mr_mapping": t.mr_mapping,
            "mr_mapping_line": t.mr_mapping_line,
            "ifrs_line_item": t.ifrs_line_item,
            "closing_dr": t.closing_debit,
            "closing_cr": t.closing_credit,
        }
        for t in tb_items
    ]


@router.put("/mappings/coa/{account_code}")
async def update_coa_mr_mapping(
    account_code: str,
    mr_code: str = Query(..., description="Baku MR code (e.g., '10.B.03.01')"),
    mr_line: Optional[str] = Query(None, description="MR line name"),
    mr_statement: Optional[str] = Query(None, description="BS or PL"),
    db: AsyncSession = Depends(get_db),
):
    """
    Manually update the MR mapping for a specific COA account.
    This allows user overrides for accounts that don't match template patterns.
    """
    acct = (await db.execute(
        select(COAMasterAccount).where(COAMasterAccount.account_code == account_code)
    )).scalar_one_or_none()
    if not acct:
        raise HTTPException(404, f"COA account {account_code} not found")

    acct.baku_mr_code = mr_code
    if mr_line:
        acct.baku_mr_line = mr_line
    if mr_statement:
        acct.baku_mr_statement = mr_statement.upper()

    await db.commit()
    return {"message": f"Updated MR mapping for {account_code}", "baku_mr_code": mr_code}


@router.get("/mappings/reference/{dataset_id}")
async def get_mapping_reference(dataset_id: int, db: AsyncSession = Depends(get_db)):
    """
    Get the complete MAPPING GRP → MAPPING BAKU cross-reference for a dataset.
    This is the data that appears in the MAPPING GRP_BAKU sheet of the Excel export.
    """
    tb_items = (await db.execute(
        select(TrialBalanceItem).where(TrialBalanceItem.dataset_id == dataset_id)
    )).scalars().all()

    reference = build_mapping_grp_baku_reference(tb_items)
    mapped_count = sum(1 for r in reference if r.get("baku_mr_code"))

    return {
        "total_accounts": len(reference),
        "mapped_to_mr": mapped_count,
        "unmapped": len(reference) - mapped_count,
        "reference": reference,
    }


# ═══════════════════════════════════════════════════════════════════
# Excel MR Report Generation (from P&L Comparison service — correct data)
# ═══════════════════════════════════════════════════════════════════

class MRExcelRequest(BaseModel):
    """Request for Excel MR report generation."""
    period: str = "2026-01"
    gel_usd_rate: float = 2.72           # Average rate for the month
    gel_usd_rate_eop: float = 2.69       # End-of-period rate (for BS)
    gel_usd_rate_boy: float = 2.70       # Beginning-of-year rate
    company_name: str = None
    include_sheets: Optional[List[str]] = None
    dataset_id: Optional[int] = None


@router.post("/generate-excel")
async def generate_mr_excel(payload: MRExcelRequest, db: AsyncSession = Depends(get_db)):
    """
    Generate MR Report Excel workbook using the P&L Comparison service.

    Produces 6 sheets: Executive Summary, P&L Statement, Balance Sheet,
    Revenue, COGS, KPI Dashboard — all sourced from the verified
    pl_comparison service (same data as /api/analytics/pl-comparison).
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from app.services.v2.pl_comparison import pl_comparison
    from app.routers.journal_router import _auto_find_prior
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    import re

    # Apply default company name from settings if not provided
    if not payload.company_name:
        payload.company_name = settings.COMPANY_NAME

    # ── Resolve dataset ──────────────────────────────────────────
    dataset_id = payload.dataset_id
    if not dataset_id:
        ds = (await db.execute(
            select(Dataset)
            .where(Dataset.record_count > 0, Dataset.record_count < 10000)
            .order_by(Dataset.id.desc()).limit(1)
        )).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(
                select(Dataset)
                .where(Dataset.record_count > 0)
                .order_by(Dataset.id.desc()).limit(1)
            )).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset with financial data found")
        dataset_id = ds.id

    # Auto-find prior year
    prior_dataset_id = await _auto_find_prior(dataset_id, db)

    # ── Fetch all data from pl_comparison service ────────────────
    try:
        pl_data = await pl_comparison.full_pl(dataset_id, prior_dataset_id, db)
    except Exception as e:
        logger.error(f"P&L data fetch failed: {e}", exc_info=True)
        raise HTTPException(500, f"Failed to fetch P&L data: {e}")

    try:
        bs_data = await pl_comparison.balance_sheet_comparison(dataset_id, prior_dataset_id, db)
    except Exception as e:
        logger.warning(f"Balance Sheet data fetch failed: {e}")
        bs_data = {"rows": [], "period": ""}

    try:
        rev_data = await pl_comparison.revenue_comparison(dataset_id, prior_dataset_id, db)
    except Exception as e:
        logger.warning(f"Revenue data fetch failed: {e}")
        rev_data = {"rows": [], "period": ""}

    try:
        cogs_data = await pl_comparison.cogs_comparison(dataset_id, prior_dataset_id, db)
    except Exception as e:
        logger.warning(f"COGS data fetch failed: {e}")
        cogs_data = {"rows": [], "period": ""}

    # ── Styling constants ────────────────────────────────────────
    PRIMARY = "1B3A5C"
    ACCENT = "2563EB"
    LIGHT_BG = "F0F4F8"
    WHITE = "FFFFFF"
    GREEN_BG = "E6F4EA"
    RED_BG = "FDE8E8"

    header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    header_fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
    accent_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    light_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    red_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
    bold_font = Font(name="Calibri", bold=True, size=11)
    normal_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", bold=True, size=14, color=PRIMARY)
    subtitle_font = Font(name="Calibri", bold=True, size=12, color=PRIMARY)
    thin_border = Border(
        left=Side(style="thin", color="D0D5DD"),
        right=Side(style="thin", color="D0D5DD"),
        top=Side(style="thin", color="D0D5DD"),
        bottom=Side(style="thin", color="D0D5DD"),
    )
    num_fmt = '#,##0'
    pct_fmt = '0.0%'

    def _apply_header(ws, row, cols):
        for col_idx in range(1, cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border

    def _apply_row_style(ws, row, cols, is_bold=False, is_alt=False):
        for col_idx in range(1, cols + 1):
            c = ws.cell(row=row, column=col_idx)
            c.font = bold_font if is_bold else normal_font
            c.border = thin_border
            if is_alt:
                c.fill = light_fill

    def _fmt_number(val):
        """Return numeric value or 0."""
        if val is None:
            return 0
        try:
            return round(float(val), 2)
        except (ValueError, TypeError):
            return 0

    # ── Build workbook ───────────────────────────────────────────
    wb = openpyxl.Workbook()
    summary = pl_data.get("summary", {})
    rate = payload.gel_usd_rate or 1
    period_label = pl_data.get("period", payload.period) or payload.period

    # ── Sheet 1: Executive Summary ───────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_properties.tabColor = PRIMARY

    ws1.merge_cells("A1:F1")
    ws1.cell(row=1, column=1, value=f"MR Report — {payload.company_name}").font = title_font
    ws1.merge_cells("A2:F2")
    ws1.cell(row=2, column=1, value=f"Period: {period_label}  |  FX Rate: {rate} GEL/USD").font = subtitle_font
    ws1.row_dimensions[1].height = 30
    ws1.row_dimensions[2].height = 22

    # KPI cards
    kpis = [
        ("Revenue", summary.get("revenue", 0)),
        ("COGS", summary.get("cogs", 0)),
        ("Gross Profit", summary.get("gross_profit", 0)),
        ("EBITDA", summary.get("ebitda", 0)),
        ("Net Profit", summary.get("net_profit", 0)),
    ]

    ws1.cell(row=4, column=1, value="Key Financial Indicators").font = subtitle_font
    headers = ["Metric", "GEL", "USD", "Prior Year (GEL)", "Variance (GEL)", "Var %"]
    for ci, h in enumerate(headers, 1):
        ws1.cell(row=5, column=ci, value=h)
    _apply_header(ws1, 5, 6)

    # Build prior year values from P&L rows (sum of prior column by category)
    pl_rows = pl_data.get("rows", [])
    prior_rev = summary.get("prior_revenue", 0) or 0
    prior_np = summary.get("prior_net_profit", 0) or 0
    prior_cogs = next((r.get("pr", 0) for r in pl_rows if r.get("c") == "COGS"), 0)
    prior_gp = next((r.get("pr", 0) for r in pl_rows if r.get("c") == "TGP"), 0)
    prior_ebitda = next((r.get("pr", 0) for r in pl_rows if r.get("c") == "EBITDA"), 0)
    prior_map = {
        "Revenue": prior_rev,
        "COGS": prior_cogs,
        "Gross Profit": prior_gp,
        "EBITDA": prior_ebitda,
        "Net Profit": prior_np,
    }

    for ri, (name, val) in enumerate(kpis, 6):
        val = _fmt_number(val)
        prior_val = _fmt_number(prior_map.get(name, 0))
        variance = val - prior_val if prior_val else 0
        var_pct = (variance / abs(prior_val)) if prior_val else 0

        ws1.cell(row=ri, column=1, value=name).font = bold_font
        ws1.cell(row=ri, column=2, value=val).number_format = num_fmt
        ws1.cell(row=ri, column=3, value=round(val / rate, 2)).number_format = num_fmt
        ws1.cell(row=ri, column=4, value=prior_val).number_format = num_fmt
        ws1.cell(row=ri, column=5, value=variance).number_format = num_fmt
        pct_cell = ws1.cell(row=ri, column=6, value=var_pct)
        pct_cell.number_format = pct_fmt
        if variance > 0:
            pct_cell.fill = green_fill
        elif variance < 0:
            pct_cell.fill = red_fill
        _apply_row_style(ws1, ri, 6, is_bold=True, is_alt=(ri % 2 == 0))

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 22

    # ── Sheet 2: P&L Statement ───────────────────────────────────
    ws2 = wb.create_sheet("P&L Statement")
    ws2.sheet_properties.tabColor = ACCENT

    ws2.merge_cells("A1:G1")
    ws2.cell(row=1, column=1, value=f"Profit & Loss Statement — {period_label}").font = title_font
    ws2.row_dimensions[1].height = 28

    pl_headers = ["Code", "Line Item", "Prior Year", "Actual", "Plan", "Variance", "Var %"]
    for ci, h in enumerate(pl_headers, 1):
        ws2.cell(row=3, column=ci, value=h)
    _apply_header(ws2, 3, 7)

    pl_rows = pl_data.get("rows", [])
    for ri, row in enumerate(pl_rows, 4):
        code = row.get("c", "")
        label = row.get("l", "")
        actual = _fmt_number(row.get("ac", 0))
        prior = _fmt_number(row.get("pr", 0))
        plan = _fmt_number(row.get("pl", 0))
        variance = _fmt_number(row.get("var", 0))
        var_pct = row.get("var_pct", 0) or 0
        is_bold = row.get("bold", False)
        level = row.get("lvl", 0)

        indent = "  " * level
        ws2.cell(row=ri, column=1, value=code).font = bold_font if is_bold else normal_font
        ws2.cell(row=ri, column=2, value=f"{indent}{label}").font = bold_font if is_bold else normal_font
        ws2.cell(row=ri, column=3, value=prior).number_format = num_fmt
        ws2.cell(row=ri, column=4, value=actual).number_format = num_fmt
        ws2.cell(row=ri, column=5, value=plan).number_format = num_fmt
        ws2.cell(row=ri, column=6, value=variance).number_format = num_fmt
        pct_cell = ws2.cell(row=ri, column=7, value=var_pct / 100 if var_pct else 0)
        pct_cell.number_format = pct_fmt

        is_separator = row.get("sep", False)
        _apply_row_style(ws2, ri, 7, is_bold=is_bold, is_alt=(is_separator or level == 0))

        # Conditional color for variance
        if variance > 0:
            ws2.cell(row=ri, column=6).fill = green_fill
        elif variance < 0:
            ws2.cell(row=ri, column=6).fill = red_fill

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 38
    for cl in ["C", "D", "E", "F"]:
        ws2.column_dimensions[cl].width = 18
    ws2.column_dimensions["G"].width = 12

    # ── Sheet 3: Balance Sheet ───────────────────────────────────
    ws3 = wb.create_sheet("Balance Sheet")
    ws3.sheet_properties.tabColor = "10B981"

    ws3.merge_cells("A1:F1")
    ws3.cell(row=1, column=1, value=f"Balance Sheet — {period_label}").font = title_font
    ws3.row_dimensions[1].height = 28

    bs_headers = ["Section", "IFRS Line Item", "Prior Year", "Actual", "Variance", "Var %"]
    for ci, h in enumerate(bs_headers, 1):
        ws3.cell(row=3, column=ci, value=h)
    _apply_header(ws3, 3, 6)

    bs_rows = bs_data.get("rows", [])
    for ri, row in enumerate(bs_rows, 4):
        section = row.get("section", "")
        ifrs_line = row.get("ifrs_line", "")
        actual = _fmt_number(row.get("actual", 0))
        prior = _fmt_number(row.get("prior", 0))
        variance = _fmt_number(row.get("variance", 0))
        var_pct = row.get("variance_pct", 0) or 0
        is_bold = row.get("bold", False)
        level = row.get("level", 1)

        ws3.cell(row=ri, column=1, value=section).font = bold_font if is_bold else normal_font
        indent = "" if is_bold else "  "
        ws3.cell(row=ri, column=2, value=f"{indent}{ifrs_line}").font = bold_font if is_bold else normal_font
        ws3.cell(row=ri, column=3, value=prior).number_format = num_fmt
        ws3.cell(row=ri, column=4, value=actual).number_format = num_fmt
        ws3.cell(row=ri, column=5, value=variance).number_format = num_fmt
        pct_cell = ws3.cell(row=ri, column=6, value=var_pct / 100 if var_pct else 0)
        pct_cell.number_format = pct_fmt

        _apply_row_style(ws3, ri, 6, is_bold=is_bold, is_alt=(level == 0))
        if variance > 0:
            ws3.cell(row=ri, column=5).fill = green_fill
        elif variance < 0:
            ws3.cell(row=ri, column=5).fill = red_fill

    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 36
    for cl in ["C", "D", "E"]:
        ws3.column_dimensions[cl].width = 18
    ws3.column_dimensions["F"].width = 12

    # ── Sheet 4: Revenue ─────────────────────────────────────────
    ws4 = wb.create_sheet("Revenue")
    ws4.sheet_properties.tabColor = "F59E0B"

    ws4.merge_cells("A1:G1")
    ws4.cell(row=1, column=1, value=f"Revenue by Product — {period_label}").font = title_font
    ws4.row_dimensions[1].height = 28

    rev_headers = ["Product", "Segment", "Prior Year", "Actual (Net)", "Variance", "Var %", "% of Total"]
    for ci, h in enumerate(rev_headers, 1):
        ws4.cell(row=3, column=ci, value=h)
    _apply_header(ws4, 3, 7)

    rev_rows = rev_data.get("rows", [])
    for ri, row in enumerate(rev_rows, 4):
        product = row.get("product", "")
        segment = row.get("segment", "")
        actual_net = _fmt_number(row.get("actual_net", 0))
        prior_net = _fmt_number(row.get("prior_net", 0))
        variance = _fmt_number(row.get("variance", 0))
        var_pct = row.get("variance_pct", 0) or 0
        pct_of_total = row.get("pct_of_total", 0) or 0

        ws4.cell(row=ri, column=1, value=product).font = normal_font
        ws4.cell(row=ri, column=2, value=segment).font = normal_font
        ws4.cell(row=ri, column=3, value=prior_net).number_format = num_fmt
        ws4.cell(row=ri, column=4, value=actual_net).number_format = num_fmt
        ws4.cell(row=ri, column=5, value=variance).number_format = num_fmt
        ws4.cell(row=ri, column=6, value=var_pct / 100 if var_pct else 0).number_format = pct_fmt
        ws4.cell(row=ri, column=7, value=pct_of_total / 100 if pct_of_total else 0).number_format = pct_fmt

        _apply_row_style(ws4, ri, 7, is_alt=(ri % 2 == 0))
        if variance > 0:
            ws4.cell(row=ri, column=5).fill = green_fill
        elif variance < 0:
            ws4.cell(row=ri, column=5).fill = red_fill

    # Totals row
    total_row = 4 + len(rev_rows)
    ws4.cell(row=total_row, column=1, value="TOTAL").font = bold_font
    ws4.cell(row=total_row, column=3, value=_fmt_number(rev_data.get("total_revenue_prior", 0))).number_format = num_fmt
    ws4.cell(row=total_row, column=4, value=_fmt_number(rev_data.get("total_revenue_actual", 0))).number_format = num_fmt
    ws4.cell(row=total_row, column=5, value=_fmt_number(rev_data.get("total_variance", 0))).number_format = num_fmt
    _apply_row_style(ws4, total_row, 7, is_bold=True)

    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 16
    for cl in ["C", "D", "E"]:
        ws4.column_dimensions[cl].width = 18
    ws4.column_dimensions["F"].width = 12
    ws4.column_dimensions["G"].width = 12

    # ── Sheet 5: COGS ────────────────────────────────────────────
    ws5 = wb.create_sheet("COGS")
    ws5.sheet_properties.tabColor = "EF4444"

    ws5.merge_cells("A1:F1")
    ws5.cell(row=1, column=1, value=f"Cost of Goods Sold — {period_label}").font = title_font
    ws5.row_dimensions[1].height = 28

    cogs_headers = ["Category", "Segment", "Prior Year", "Actual", "Variance", "Var %"]
    for ci, h in enumerate(cogs_headers, 1):
        ws5.cell(row=3, column=ci, value=h)
    _apply_header(ws5, 3, 6)

    cogs_rows = cogs_data.get("rows", [])
    for ri, row in enumerate(cogs_rows, 4):
        category = row.get("category", "")
        segment = row.get("segment", "")
        actual = _fmt_number(row.get("actual", 0))
        prior = _fmt_number(row.get("prior", 0))
        variance = _fmt_number(row.get("variance", 0))
        var_pct = row.get("variance_pct", 0) or 0

        ws5.cell(row=ri, column=1, value=category).font = normal_font
        ws5.cell(row=ri, column=2, value=segment).font = normal_font
        ws5.cell(row=ri, column=3, value=prior).number_format = num_fmt
        ws5.cell(row=ri, column=4, value=actual).number_format = num_fmt
        ws5.cell(row=ri, column=5, value=variance).number_format = num_fmt
        ws5.cell(row=ri, column=6, value=var_pct / 100 if var_pct else 0).number_format = pct_fmt

        _apply_row_style(ws5, ri, 6, is_alt=(ri % 2 == 0))
        # For COGS, negative variance (lower costs) is good
        if variance < 0:
            ws5.cell(row=ri, column=5).fill = green_fill
        elif variance > 0:
            ws5.cell(row=ri, column=5).fill = red_fill

    # Totals row
    cogs_total_row = 4 + len(cogs_rows)
    ws5.cell(row=cogs_total_row, column=1, value="TOTAL COGS").font = bold_font
    ws5.cell(row=cogs_total_row, column=3, value=_fmt_number(cogs_data.get("total_prior", 0))).number_format = num_fmt
    ws5.cell(row=cogs_total_row, column=4, value=_fmt_number(cogs_data.get("total_actual", 0))).number_format = num_fmt
    ws5.cell(row=cogs_total_row, column=5, value=_fmt_number(cogs_data.get("total_variance", 0))).number_format = num_fmt
    _apply_row_style(ws5, cogs_total_row, 6, is_bold=True)

    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 16
    for cl in ["C", "D", "E"]:
        ws5.column_dimensions[cl].width = 18
    ws5.column_dimensions["F"].width = 12

    # ── Sheet 6: KPI Dashboard ───────────────────────────────────
    ws6 = wb.create_sheet("KPI Dashboard")
    ws6.sheet_properties.tabColor = "8B5CF6"

    ws6.merge_cells("A1:D1")
    ws6.cell(row=1, column=1, value=f"KPI Dashboard — {period_label}").font = title_font
    ws6.row_dimensions[1].height = 28

    kpi_headers = ["KPI", "Value", "Prior Year", "Change"]
    for ci, h in enumerate(kpi_headers, 1):
        ws6.cell(row=3, column=ci, value=h)
    _apply_header(ws6, 3, 4)

    revenue = _fmt_number(summary.get("revenue", 0))
    gross_profit = _fmt_number(summary.get("gross_profit", 0))
    ebitda = _fmt_number(summary.get("ebitda", 0))
    net_profit = _fmt_number(summary.get("net_profit", 0))
    prior_revenue = _fmt_number(summary.get("prior_revenue", 0))
    prior_net = _fmt_number(summary.get("prior_net_profit", 0))

    # Get prior year values from P&L rows
    _prior_gp = _fmt_number(prior_map.get("Gross Profit", 0))
    _prior_ebitda = _fmt_number(prior_map.get("EBITDA", 0))
    _prior_cogs = _fmt_number(prior_map.get("COGS", 0))

    gross_margin = (gross_profit / revenue) if revenue else 0
    net_margin = (net_profit / revenue) if revenue else 0
    ebitda_margin = (ebitda / revenue) if revenue else 0
    rev_growth = ((revenue - prior_revenue) / abs(prior_revenue)) if prior_revenue else 0
    gp_change = ((gross_profit - _prior_gp) / abs(_prior_gp)) if _prior_gp else 0
    ebitda_change = ((ebitda - _prior_ebitda) / abs(_prior_ebitda)) if _prior_ebitda else 0
    np_change = ((net_profit - prior_net) / abs(prior_net)) if prior_net else 0

    prior_gross_margin = (_prior_gp / prior_revenue) if prior_revenue else 0
    prior_net_margin = (prior_net / prior_revenue) if prior_revenue else 0
    prior_ebitda_margin = (_prior_ebitda / prior_revenue) if prior_revenue else 0

    kpi_rows = [
        ("Revenue (GEL)", revenue, prior_revenue, rev_growth, num_fmt, pct_fmt),
        ("Gross Profit (GEL)", gross_profit, _prior_gp, gp_change, num_fmt, pct_fmt),
        ("EBITDA (GEL)", ebitda, _prior_ebitda, ebitda_change, num_fmt, pct_fmt),
        ("Net Profit (GEL)", net_profit, prior_net, np_change, num_fmt, pct_fmt),
        ("Gross Margin %", gross_margin, prior_gross_margin, gross_margin - prior_gross_margin, pct_fmt, pct_fmt),
        ("Net Margin %", net_margin, prior_net_margin, net_margin - prior_net_margin, pct_fmt, pct_fmt),
        ("EBITDA Margin %", ebitda_margin, prior_ebitda_margin, ebitda_margin - prior_ebitda_margin, pct_fmt, pct_fmt),
        ("Revenue Growth %", rev_growth, 0, 0, pct_fmt, pct_fmt),
    ]

    for ri, (name, val, prior_val, change, val_fmt, chg_fmt) in enumerate(kpi_rows, 4):
        ws6.cell(row=ri, column=1, value=name).font = bold_font
        ws6.cell(row=ri, column=2, value=val).number_format = val_fmt
        ws6.cell(row=ri, column=3, value=prior_val).number_format = val_fmt
        chg_cell = ws6.cell(row=ri, column=4, value=change)
        chg_cell.number_format = chg_fmt
        if change > 0:
            chg_cell.fill = green_fill
        elif change < 0:
            chg_cell.fill = red_fill
        _apply_row_style(ws6, ri, 4, is_bold=True, is_alt=(ri % 2 == 0))

    for cl in ["A", "B", "C", "D"]:
        ws6.column_dimensions[cl].width = 22

    # ── Save and return ──────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    filename = f"MR_Report_{payload.period}_{payload.company_name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
