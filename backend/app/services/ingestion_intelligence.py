"""
ingestion_intelligence.py — Autonomous Financial File Ingestion Intelligence
=============================================================================
Detects, classifies, and routes any Excel/CSV financial file without
requiring prior knowledge of its format.

Supported file types detected:
  CHART_OF_ACCOUNTS    — 1C AccountN / ПланСчетов
  GENERAL_LEDGER       — Journal / transaction records with account pairs
  TRIAL_BALANCE        — Period-end debit/credit balance summary
  INCOME_STATEMENT     — Pre-built P&L report
  BALANCE_SHEET        — Pre-built BS report
  CASH_FLOW            — Cash flow statement
  BUDGET               — Budget / plan data
  KPI_DASHBOARD        — KPI or metrics dashboard
  UNKNOWN              — Unrecognised structure

Key classes:
  FileStructureDetector   — detect headers, table regions, merged cells
  ColumnSemanticClassifier— classify columns by semantic role
  SchemaDetector          — combine signals → schema_type + confidence
  IngestionPipeline       — orchestrate detect → parse → route

Usage:
    pipeline = IngestionPipeline()
    result   = pipeline.process_file("January 2025 export.xlsx")
    # result.schema_type → "GENERAL_LEDGER"
    # result.columns     → {"account_code": 2, "debit": 4, "credit": 5, ...}
    # result.data        → parsed records as list of dicts
"""
from __future__ import annotations

import re
import logging
import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ── Schema type constants ──────────────────────────────────────────────────────
class SchemaType:
    CHART_OF_ACCOUNTS = "CHART_OF_ACCOUNTS"
    GENERAL_LEDGER    = "GENERAL_LEDGER"
    TRIAL_BALANCE     = "TRIAL_BALANCE"
    INCOME_STATEMENT  = "INCOME_STATEMENT"
    BALANCE_SHEET     = "BALANCE_SHEET"
    CASH_FLOW         = "CASH_FLOW"
    BUDGET            = "BUDGET"
    KPI_DASHBOARD     = "KPI_DASHBOARD"
    UNKNOWN           = "UNKNOWN"


# ── Semantic column role keywords (multilingual) ──────────────────────────────
_COLUMN_ROLE_KEYWORDS: Dict[str, List[str]] = {
    "account_code": [
        "код", "code", "account", "счет", "acct", "account_code",
        "accounts", "account no", "acc", "хесანგარიშო",
    ],
    "account_name": [
        "наименование", "name", "назва", "description", "account name",
        "label", "title", "item", "სახელი",
    ],
    "debit": [
        "дебет", "debit", "dr", "debit turnover", "dt", "დებეტი",
        "debet", "дт", "приход", "turnover dr",
    ],
    "credit": [
        "кредит", "credit", "cr", "credit turnover", "ct", "კრედიტი",
        "kredit", "кт", "расход", "turnover cr",
    ],
    "balance": [
        "остаток", "balance", "ending balance", "closing balance",
        "balance c/f", "net balance", "ნაშთი",
    ],
    "amount": [
        "сумма", "amount", "value", "turnover", "total",
        "оборот", "sum", "ოდენობა",
    ],
    "period": [
        "период", "period", "month", "quarter", "year", "date range",
        "reporting period", "პერიოდი",
    ],
    "date": [
        "дата", "date", "doc date", "transaction date",
        "posting date", "value date", "თარიღი",
    ],
    "counterparty": [
        "контрагент", "counterparty", "partner", "vendor", "customer",
        "supplier", "client", "კონტრაგენტი",
    ],
    "document_id": [
        "номер документа", "document", "doc no", "reference",
        "doc id", "transaction id", "entry no", "документი",
    ],
    "cost_center": [
        "подразделение", "department", "cost center", "division",
        "unit", "branch", "cost_center", "დეპარტამენტი",
    ],
    "currency": [
        "валюта", "currency", "ccy", "curr", "fx", "ვალუტა",
    ],
    "product": [
        "номенклатура", "product", "item", "goods", "sku",
        "product name", "description", "საქონელი",
    ],
    "quantity": [
        "количество", "quantity", "qty", "units", "volume",
        "კოლიჩ", "რაოდენობა",
    ],
    "subkonto1": ["субконто 1", "subkonto1", "sub1", "dimension1"],
    "subkonto2": ["субконто 2", "subkonto2", "sub2", "dimension2"],
    "subkonto3": ["субконто 3", "subkonto3", "sub3", "dimension3"],
    "account_type": ["акт.", "type", "тип", "active", "passive", "акт"],
    "off_balance":  ["заб.", "off-balance", "забалансов"],
    "is_currency":  ["вал.", "currency flag", "foreign"],
    "is_quantity":  ["кол.", "quantity flag"],
}

# ── Schema detection signal weights ──────────────────────────────────────────
# Each schema type has required and bonus column roles.
_SCHEMA_SIGNALS: Dict[str, Dict] = {
    SchemaType.CHART_OF_ACCOUNTS: {
        "required": {"account_code", "account_name"},
        "bonus":    {"account_type", "off_balance", "subkonto1"},
        "keywords": ["план счетов", "chart of accounts", "accountn",
                     "ანგარიშები", "субконто", "акт."],
        "row_keyword": ["заб.", "субконто", "акт."],
    },
    SchemaType.GENERAL_LEDGER: {
        "required": {"account_code", "debit", "credit"},
        "bonus":    {"date", "counterparty", "document_id", "amount"},
        "keywords": ["журнал", "general ledger", "transactions",
                     "оборот", "обороты", "проводк"],
        "row_keyword": [],
    },
    SchemaType.TRIAL_BALANCE: {
        "required": {"account_code", "debit", "credit"},
        "bonus":    {"balance", "account_name"},
        "keywords": ["оборотно-сальдовая", "trial balance", "oborotka",
                     "sal", "obb"],
        "row_keyword": [],
    },
    SchemaType.INCOME_STATEMENT: {
        "required": {"amount"},
        "bonus":    {"period"},
        "keywords": ["income statement", "p&l", "profit", "loss",
                     "revenue", "cogs", "gross profit", "ebitda",
                     "отчет о прибылях", "отчет прибылях и убытках",
                     "მოგება-ზარალი"],
        "row_keyword": ["revenue", "gross profit", "ebitda", "net income",
                        "cogs", "revenues", "expenses"],
    },
    SchemaType.BALANCE_SHEET: {
        "required": {"amount"},
        "bonus":    {"period"},
        "keywords": ["balance sheet", "бухгалтерский баланс", "баланс",
                     "assets", "liabilities", "equity",
                     "ბალანსი"],
        "row_keyword": ["total assets", "total liabilities", "equity",
                        "fixed assets", "current assets"],
    },
    SchemaType.CASH_FLOW: {
        "required": {"amount"},
        "bonus":    {"period"},
        "keywords": ["cash flow", "движение денежных", "operating activities",
                     "investing", "financing", "fcf"],
        "row_keyword": ["operating activities", "investing activities",
                        "financing activities", "net cash"],
    },
    SchemaType.BUDGET: {
        "required": {"amount"},
        "bonus":    {"period", "account_code"},
        "keywords": ["budget", "бюджет", "plan", "план", "forecast",
                     "прогноз"],
        "row_keyword": ["budget", "plan", "actual", "variance"],
    },
    SchemaType.KPI_DASHBOARD: {
        "required": set(),
        "bonus":    {"period", "amount"},
        "keywords": ["kpi", "dashboard", "metrics", "indicator",
                     "показатели", "performance"],
        "row_keyword": ["kpi", "target", "actual", "metric"],
    },
}


@dataclass
class ColumnMapping:
    """Detected column roles for a sheet."""
    roles: Dict[str, Optional[int]] = field(default_factory=dict)
    header_row: int = 0
    detected_columns: List[str] = field(default_factory=list)

    def has(self, *roles) -> bool:
        return all(self.roles.get(r) is not None for r in roles)

    def get_idx(self, role: str) -> Optional[int]:
        return self.roles.get(role)


@dataclass
class DetectionResult:
    """Result of schema detection for one sheet."""
    schema_type: str = SchemaType.UNKNOWN
    confidence: float = 0.0
    filename: str = ""
    sheet_name: str = ""
    header_row: int = 0
    columns: Dict[str, Optional[int]] = field(default_factory=dict)
    row_count: int = 0
    signals: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class ParsedFile:
    """Full result after detection + parsing."""
    detection: DetectionResult
    records: List[Dict[str, Any]] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

    @property
    def schema_type(self) -> str:
        return self.detection.schema_type

    @property
    def confidence(self) -> float:
        return self.detection.confidence


# ── FileStructureDetector ─────────────────────────────────────────────────────

class FileStructureDetector:
    """
    Detects table structure in raw Excel/CSV rows.
    Finds header rows, table boundaries, skips decoration rows.
    """

    def find_header_row(self, rows: List[tuple], max_search: int = 20) -> int:
        """Return 0-based index of the header row."""
        best_score = -1
        best_idx   = 0
        for idx, row in enumerate(rows[:max_search]):
            score = self._header_score(row)
            if score > best_score:
                best_score = score
                best_idx   = idx
        return best_idx

    # Known header keywords — includes abbreviations for robust detection
    _HEADER_KEYWORDS = {
        # Russian
        "код", "дебет", "кредит", "наим", "период", "валют", "акт",
        "сумма", "дата", "счет", "номер", "контрагент", "оборот",
        "остаток", "итого", "субконто", "подразделение", "количество",
        # English (full + abbreviations)
        "amount", "date", "account", "debit", "credit", "balance",
        "code", "name", "value", "description", "total", "turnover",
        "period", "month", "quarter", "year", "currency", "type",
        "acct", "amt", "qty", "dr", "cr", "ref", "doc",
        "item", "product", "vendor", "customer", "counterparty",
        # Georgian
        "კოდი", "თარიღი", "თანხა", "ანგარიში", "სახელი",
    }

    def _header_score(self, row: tuple) -> float:
        """Score a row for likelihood of being a header.

        Scoring signals:
            - text_ratio: headers are mostly text, not numbers (weight: 2)
            - keyword_hits: cells that EXACTLY match or CONTAIN header keywords (weight: 3)
            - exact_match_bonus: cells that ARE a keyword (not substring) (weight: 2)
            - num_penalty: numeric columns unlikely in headers (weight: -1)
            - short_text_bonus: headers tend to be short labels (weight: 0.5)
        """
        if not any(row):
            return -1
        non_null  = sum(1 for c in row if c is not None)
        text_cols = sum(1 for c in row if c is not None and isinstance(c, str))
        num_cols  = sum(1 for c in row if c is not None and isinstance(c, (int, float)))
        if non_null == 0:
            return -1

        text_ratio = text_cols / non_null

        # Check for known header keywords with exact-match bonus
        keyword_hits = 0
        exact_matches = 0
        for c in row:
            if c is None or not isinstance(c, str):
                continue
            cell_lower = str(c).lower().strip()

            # Exact match: the cell IS a keyword (or close to it)
            # Strip common decorators
            cell_clean = re.sub(r'[#.\-_/\\()\[\]:]', ' ', cell_lower).strip()
            cell_words = cell_clean.split()

            # Check if any word in the cell is exactly a header keyword
            for word in cell_words:
                if word in self._HEADER_KEYWORDS:
                    exact_matches += 1
                    break

            # Substring containment (looser match)
            for kw in self._HEADER_KEYWORDS:
                if kw in cell_lower:
                    keyword_hits += 1
                    break

        # Short text bonus: header cells are typically < 30 chars
        short_texts = sum(
            1 for c in row
            if c is not None and isinstance(c, str) and len(str(c)) < 30
        )
        short_ratio = short_texts / max(text_cols, 1) if text_cols > 0 else 0

        return (
            text_ratio * 2
            + keyword_hits * 3
            + exact_matches * 2
            + short_ratio * 0.5
            - num_cols * 1.0
        )

    def detect_table_region(self, rows: List[tuple], header_idx: int) -> Tuple[int, int]:
        """Return (first_data_row, last_data_row) indices."""
        start = header_idx + 1
        end   = len(rows) - 1
        # Find last non-empty row
        for i in range(len(rows) - 1, header_idx, -1):
            if any(rows[i]):
                end = i
                break
        return start, end

    def estimate_column_count(self, rows: List[tuple], header_idx: int) -> int:
        """Return the maximum non-null column index in the table."""
        max_col = 0
        for row in rows[header_idx:header_idx + 10]:
            for i, c in enumerate(row):
                if c is not None:
                    max_col = max(max_col, i)
        return max_col + 1


# ── ColumnSemanticClassifier ──────────────────────────────────────────────────

class ColumnSemanticClassifier:
    """
    Classifies each column into a semantic role based on:
    1. Header keyword matching (primary)
    2. Data pattern analysis (secondary)
    3. Position heuristics (tertiary)
    """

    def classify(self, header_row: tuple, data_rows: List[tuple]) -> ColumnMapping:
        """Classify all columns in a row."""
        roles: Dict[str, Optional[int]] = {role: None for role in _COLUMN_ROLE_KEYWORDS}
        detected: List[str] = []

        # Pass 1: Header keyword matching
        for col_idx, cell in enumerate(header_row):
            if cell is None:
                continue
            cell_str = str(cell).lower().strip()
            best_role, best_score = None, 0
            for role, keywords in _COLUMN_ROLE_KEYWORDS.items():
                for kw in keywords:
                    if kw in cell_str:
                        score = len(kw)  # Longer match = more specific
                        if score > best_score:
                            best_score = score
                            best_role  = role
            if best_role and roles.get(best_role) is None:
                roles[best_role] = col_idx
                detected.append(best_role)

        # Pass 2: Data pattern analysis for unresolved columns
        if data_rows:
            sample = data_rows[:min(20, len(data_rows))]
            for col_idx in range(len(header_row)):
                if col_idx in roles.values():
                    continue  # Already classified
                vals = [row[col_idx] for row in sample if col_idx < len(row) and row[col_idx] is not None]
                if not vals:
                    continue
                inferred = self._infer_role_from_data(vals)
                if inferred and roles.get(inferred) is None:
                    roles[inferred] = col_idx

        return ColumnMapping(roles=roles, detected_columns=detected)

    def _infer_role_from_data(self, values: List[Any]) -> Optional[str]:
        """Infer column role from sample values."""
        num_count  = sum(1 for v in values if isinstance(v, (int, float)))
        str_count  = sum(1 for v in values if isinstance(v, str))
        total      = len(values)
        if total == 0:
            return None

        # Mostly numeric → could be amount, debit, credit, balance
        if num_count / total > 0.7:
            # Large positive numbers → amount
            avg = sum(abs(float(v)) for v in values if isinstance(v, (int, float))) / max(num_count, 1)
            if avg > 100:
                return "amount"

        # Mostly short strings → account_code or account_type
        if str_count / total > 0.7:
            avg_len = sum(len(str(v)) for v in values if isinstance(v, str)) / max(str_count, 1)
            if avg_len <= 8:  # Short code-like strings
                # Check if looks like account codes (digits/dots)
                code_like = sum(1 for v in values if isinstance(v, str) and re.match(r'^[\d.]+$', v.strip()))
                if code_like / max(str_count, 1) > 0.5:
                    return "account_code"
            elif avg_len > 8:
                return "account_name"

        return None


# ── SchemaDetector ────────────────────────────────────────────────────────────

class SchemaDetector:
    """
    Combines column roles + file content signals → schema_type + confidence.
    Uses a scoring model:
      - Required columns present → base score
      - Bonus columns → additional score
      - Filename keywords → additional score
      - Row content keywords → additional score
    """

    def detect(
        self,
        filename: str,
        sheet_name: str,
        header_row: tuple,
        data_rows: List[tuple],
        column_mapping: ColumnMapping,
    ) -> DetectionResult:
        filename_lower = filename.lower()
        sheet_lower    = sheet_name.lower()

        # Sample text from first 30 data rows
        all_text = " ".join(
            str(c).lower()
            for row in data_rows[:30]
            for c in row
            if c is not None and isinstance(c, str)
        )
        header_text = " ".join(str(c).lower() for c in header_row if c is not None)

        best_type  = SchemaType.UNKNOWN
        best_score = 0.0
        signals    = []
        warnings   = []

        for schema_type, config in _SCHEMA_SIGNALS.items():
            score  = 0.0
            schema_signals = []

            # Required columns check
            required = config.get("required", set())
            has_required = all(column_mapping.has(r) for r in required)
            if required and not has_required:
                missing = [r for r in required if not column_mapping.has(r)]
                schema_signals.append(f"missing_required:{','.join(missing)}")
            elif required:
                score += 30.0
                schema_signals.append("required_cols_present")

            # Bonus columns check
            bonus    = config.get("bonus", set())
            bonus_ct = sum(1 for r in bonus if column_mapping.has(r))
            score   += bonus_ct * 5.0
            if bonus_ct:
                schema_signals.append(f"bonus_cols:{bonus_ct}/{len(bonus)}")

            # File/sheet name keywords
            for kw in config.get("keywords", []):
                if kw in filename_lower or kw in sheet_lower:
                    score += 20.0
                    schema_signals.append(f"filename_kw:{kw}")
                    break

            # Row content keywords
            row_kws = config.get("row_keyword", [])
            hits = sum(1 for kw in row_kws if kw in all_text)
            score += hits * 3.0
            if hits:
                schema_signals.append(f"row_content_kws:{hits}")

            # Header content keywords
            for kw in config.get("keywords", []):
                if kw in header_text:
                    score += 10.0
                    schema_signals.append(f"header_kw:{kw}")
                    break

            if score > best_score:
                best_score = score
                best_type  = schema_type
                signals    = schema_signals

        # Confidence: normalise score
        confidence = min(round(best_score / 60.0, 2), 1.0)
        if best_score < 15:
            best_type  = SchemaType.UNKNOWN
            confidence = 0.0
            warnings.append("Low confidence detection — schema unclear")

        return DetectionResult(
            schema_type=best_type,
            confidence=confidence,
            filename=filename,
            sheet_name=sheet_name,
            header_row=column_mapping.header_row,
            columns={k: v for k, v in column_mapping.roles.items() if v is not None},
            row_count=len(data_rows),
            signals=signals,
            warnings=warnings,
        )


# ── IngestionPipeline ──────────────────────────────────────────────────────────

class IngestionPipeline:
    """
    Full ingestion pipeline with Hypothesis-Driven Parsing (HDP).

    Architecture (Phase H upgrade):
        Primary path:  HDP engine (multi-hypothesis search + scoring)
        Fallback path: Legacy sequential pipeline (if HDP fails)

    HDP flow:
        1. Generate 8 candidate interpretations (one per schema type)
        2. Parse data under EACH hypothesis
        3. Validate using accounting invariants (debit=credit, COA match, etc.)
        4. Score and rank hypotheses
        5. Apply Schema Memory boost (learned patterns)
        6. Select best interpretation

    This architecture makes ingestion self-correcting:
        - Column misclassification is caught by validation scoring
        - Schema Memory enables instant recognition of known formats
        - Dropped rows are tracked with reasons (never silent)
        - Runner-up hypothesis warnings flag ambiguous files

    For CHART_OF_ACCOUNTS files, also runs OneCInterpreter.
    """

    def __init__(self, use_hdp: bool = True):
        self.structure_detector = FileStructureDetector()
        self.column_classifier  = ColumnSemanticClassifier()
        self.schema_detector    = SchemaDetector()
        self._use_hdp = use_hdp
        self._hdp_parser = None  # Lazy init to avoid circular imports

    @property
    def hdp_parser(self):
        """Lazy-init HDP parser with schema memory."""
        if self._hdp_parser is None and self._use_hdp:
            try:
                from app.services.hypothesis_parser import HypothesisDrivenParser
                from app.services.schema_memory import schema_memory
                self._hdp_parser = HypothesisDrivenParser(schema_memory=schema_memory)
            except Exception as e:
                logger.warning("HDP parser init failed, using legacy pipeline: %s", e)
                self._use_hdp = False
        return self._hdp_parser

    def process_file(self, path: str) -> "List[ParsedFile]":
        """Process all sheets in an Excel file. Returns one ParsedFile per sheet."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl required")

        p = Path(path)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        results = []

        for sheet_name in wb.sheetnames:
            ws    = wb[sheet_name]
            rows  = list(ws.iter_rows(values_only=True))
            parsed = self._process_sheet(rows, p.name, sheet_name)
            results.append(parsed)

        wb.close()
        return results

    def process_bytes(self, data: bytes, filename: str) -> "List[ParsedFile]":
        """Process Excel bytes (HTTP upload)."""
        import io
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl required")

        wb  = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        p   = Path(filename)
        results = []

        for sheet_name in wb.sheetnames:
            ws    = wb[sheet_name]
            rows  = list(ws.iter_rows(values_only=True))
            parsed = self._process_sheet(rows, p.name, sheet_name)
            results.append(parsed)

        wb.close()
        return results

    def _process_sheet(self, rows: List[tuple], filename: str, sheet_name: str) -> ParsedFile:
        if not rows:
            det = DetectionResult(schema_type=SchemaType.UNKNOWN, confidence=0.0,
                                  filename=filename, sheet_name=sheet_name)
            return ParsedFile(detection=det)

        # ── PRIMARY PATH: Hypothesis-Driven Parsing ──────────────────
        if self._use_hdp and self.hdp_parser is not None:
            try:
                hdp_result = self.hdp_parser.parse_sheet(rows, filename, sheet_name)

                # Convert to ParsedFile (backward compatible)
                metadata = hdp_result.metadata.copy()

                # If COA detected, also run OneCInterpreter
                if hdp_result.schema_type == SchemaType.CHART_OF_ACCOUNTS:
                    self._run_onec_interpreter(rows, metadata)

                logger.info(
                    "IngestionPipeline[HDP]: %s / %s -> %s (conf=%.2f, rows=%d, "
                    "hypotheses=%d, runner_up=%s)",
                    filename, sheet_name,
                    hdp_result.schema_type,
                    hdp_result.confidence,
                    len(hdp_result.records),
                    len(hdp_result.all_hypotheses),
                    metadata.get("runner_up", {}).get("schema", "N/A"),
                )

                return ParsedFile(
                    detection=hdp_result.detection,
                    records=hdp_result.records,
                    metadata=metadata,
                )
            except Exception as e:
                logger.warning(
                    "HDP parsing failed for %s/%s, falling back to legacy: %s",
                    filename, sheet_name, e
                )

        # ── FALLBACK PATH: Legacy sequential pipeline ────────────────
        return self._legacy_process_sheet(rows, filename, sheet_name)

    def _legacy_process_sheet(self, rows: List[tuple], filename: str, sheet_name: str) -> ParsedFile:
        """Original sequential pipeline (kept as fallback)."""
        # Step 1: find header row
        header_idx = self.structure_detector.find_header_row(rows)
        header_row = rows[header_idx]
        start, end = self.structure_detector.detect_table_region(rows, header_idx)
        data_rows  = rows[start:end + 1]

        # Step 2: classify columns
        col_map = self.column_classifier.classify(header_row, data_rows)
        col_map.header_row = header_idx

        # Step 3: detect schema
        detection = self.schema_detector.detect(filename, sheet_name, header_row, data_rows, col_map)

        # Step 4: parse records
        records = self._parse_records(data_rows, col_map, detection.schema_type)

        logger.info(
            "IngestionPipeline[legacy]: %s / %s -> %s (conf=%.2f, rows=%d)",
            filename, sheet_name, detection.schema_type, detection.confidence, len(records),
        )

        # Step 5: if COA, also run OneCInterpreter
        metadata: Dict[str, Any] = {"pipeline": "legacy"}
        if detection.schema_type == SchemaType.CHART_OF_ACCOUNTS:
            self._run_onec_interpreter(rows, metadata)

        return ParsedFile(detection=detection, records=records, metadata=metadata)

    def _run_onec_interpreter(self, rows: List[tuple], metadata: Dict[str, Any]) -> None:
        """Run OneCInterpreter for COA files (shared by HDP and legacy)."""
        try:
            from app.services.onec_interpreter import OneCInterpreter
            interp  = OneCInterpreter()
            tree    = interp._parse_rows(rows)
            summary = tree.summary()
            metadata["account_tree"] = summary
            metadata["coa_parsed"]   = True
        except Exception as e:
            logger.warning("IngestionPipeline: OneCInterpreter failed: %s", e)
            metadata["coa_parsed"] = False
            metadata["coa_error"] = str(e)

    def _parse_records(
        self,
        data_rows: List[tuple],
        col_map: ColumnMapping,
        schema_type: str,
    ) -> List[Dict[str, Any]]:
        """Convert data rows to list of dicts using detected column roles."""
        records = []
        roles = col_map.roles

        for row in data_rows:
            if not any(row):
                continue  # skip blank rows
            record: Dict[str, Any] = {}
            for role, idx in roles.items():
                if idx is None or idx >= len(row):
                    continue
                val = row[idx]
                if val is not None:
                    record[role] = val
            if record:
                records.append(record)

        return records

    def detect_from_sample(self, rows: List[tuple], filename: str = "unknown.xlsx") -> DetectionResult:
        """Quick schema detection from sample rows (no full parse).

        Uses HDP if available for more accurate detection.
        """
        if not rows:
            return DetectionResult(schema_type=SchemaType.UNKNOWN, confidence=0.0)

        # Try HDP first for better accuracy
        if self._use_hdp and self.hdp_parser is not None:
            try:
                result = self.hdp_parser.parse_sheet(rows, filename)
                return result.detection
            except Exception:
                pass  # Fall through to legacy

        # Legacy detection
        header_idx = self.structure_detector.find_header_row(rows)
        header_row = rows[header_idx]
        data_rows  = rows[header_idx + 1: header_idx + 30]
        col_map    = self.column_classifier.classify(header_row, data_rows)
        return self.schema_detector.detect(filename, "", header_row, data_rows, col_map)


# ── Module-level singleton ────────────────────────────────────────────────────
ingestion_pipeline = IngestionPipeline(use_hdp=True)
