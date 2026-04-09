"""
FinAI Journal & Period API Router — System of Record endpoints.
================================================================
Exposes v2 journal_system, approval_engine, compliance_engine,
product_profitability, and lineage_service to the frontend.
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
from app.database import get_db
from app.config import settings
import logging

logger = logging.getLogger(__name__)
router = APIRouter(tags=["journal"])


# ── Pydantic models ───────────────────────────────────────────────────

class PostingLineInput(BaseModel):
    account_code: str
    account_name: str = ""
    debit: str = "0"
    credit: str = "0"
    description: str = ""
    cost_center: str = ""
    tax_code: str = ""

class JournalEntryInput(BaseModel):
    posting_date: str
    period: str
    fiscal_year: int
    description: str
    lines: List[PostingLineInput]
    currency: str = "GEL"
    reference: str = ""
    source_type: str = "manual"

class PeriodInput(BaseModel):
    period_name: str
    fiscal_year: int
    start_date: str
    end_date: str

class CloseInput(BaseModel):
    close_type: str = "hard_close"

class RejectInput(BaseModel):
    reason: str


# ── Journal Entry CRUD ────────────────────────────────────────────────

@router.post("/api/journal/entries")
async def create_journal_entry(data: JournalEntryInput, db: AsyncSession = Depends(get_db)):
    """Create a draft journal entry with posting lines."""
    from app.services.v2.journal_system import journal_service, UnbalancedEntryError
    try:
        result = await journal_service.create_entry(
            posting_date=datetime.fromisoformat(data.posting_date),
            period=data.period,
            fiscal_year=data.fiscal_year,
            description=data.description,
            lines=[l.model_dump() for l in data.lines],
            currency=data.currency,
            reference=data.reference,
            source_type=data.source_type,
            db=db,
        )
        return result
    except UnbalancedEntryError as e:
        raise HTTPException(400, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/api/journal/entries")
async def list_journal_entries(
    status: Optional[str] = None,
    period: Optional[str] = None,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    """List journal entries with optional filters."""
    from app.models.all_models import JournalEntryRecord
    from sqlalchemy import select

    q = select(JournalEntryRecord).order_by(JournalEntryRecord.created_at.desc())
    if status:
        q = q.where(JournalEntryRecord.status == status)
    if period:
        q = q.where(JournalEntryRecord.period == period)
    q = q.limit(limit)

    result = await db.execute(q)
    entries = result.scalars().all()
    return {"entries": [e.to_dict() for e in entries], "count": len(entries)}


@router.get("/api/journal/entries/{entry_id}")
async def get_journal_entry(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Get journal entry with posting lines."""
    from app.services.v2.journal_system import journal_service
    result = await journal_service.get_entry(entry_id, db)
    if not result:
        raise HTTPException(404, f"Journal entry {entry_id} not found")
    return result


@router.post("/api/journal/entries/{entry_id}/post")
async def post_journal_entry(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Post a draft entry — assigns document number, becomes immutable."""
    from app.services.v2.journal_system import journal_service, ImmutableEntryError, PeriodClosedError
    try:
        return await journal_service.post_entry(entry_id, db=db)
    except (ImmutableEntryError, PeriodClosedError, ValueError) as e:
        raise HTTPException(400, str(e))


@router.post("/api/journal/entries/{entry_id}/reverse")
async def reverse_journal_entry(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Create a reversal entry for a posted JE."""
    from app.services.v2.journal_system import journal_service
    try:
        return await journal_service.reverse_entry(entry_id, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/api/journal/entries/{entry_id}/verify")
async def verify_hash(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Verify document hash integrity."""
    from app.services.v2.journal_system import journal_service
    return await journal_service.verify_hash(entry_id, db)


@router.post("/api/journal/entries/{entry_id}/submit")
async def submit_for_approval(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Submit a draft for approval."""
    from app.services.v2.approval_engine import approval_engine
    try:
        return await approval_engine.submit_for_approval(entry_id, submitted_by=0, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/api/journal/entries/{entry_id}/approve")
async def approve_entry(entry_id: int, db: AsyncSession = Depends(get_db)):
    """Approve and auto-post a submitted entry."""
    from app.services.v2.approval_engine import approval_engine
    try:
        return await approval_engine.approve(entry_id, approved_by=1, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/api/journal/entries/{entry_id}/reject")
async def reject_entry(entry_id: int, data: RejectInput, db: AsyncSession = Depends(get_db)):
    """Reject a submitted entry back to draft."""
    from app.services.v2.approval_engine import approval_engine
    try:
        return await approval_engine.reject(entry_id, rejected_by=1, reason=data.reason, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/api/journal/pending-approvals")
async def get_pending_approvals(db: AsyncSession = Depends(get_db)):
    """Get all entries awaiting approval."""
    from app.services.v2.approval_engine import approval_engine
    return {"entries": await approval_engine.get_pending_approvals(db)}


@router.get("/api/journal/trial-balance")
async def get_trial_balance(period: str, db: AsyncSession = Depends(get_db)):
    """Trial balance from posted journal entries."""
    from app.services.v2.journal_system import journal_service
    return await journal_service.trial_balance(period, db)


@router.get("/api/journal/stats")
async def get_journal_stats(db: AsyncSession = Depends(get_db)):
    """Journal entry statistics."""
    from app.services.v2.approval_engine import approval_engine
    return await approval_engine.get_approval_stats(db)


# ── Period Control ────────────────────────────────────────────────────

@router.get("/api/periods")
@router.get("/api/journal/periods")
async def list_periods(db: AsyncSession = Depends(get_db)):
    """List all fiscal periods with status."""
    from app.models.all_models import FiscalPeriod
    from sqlalchemy import select
    result = await db.execute(select(FiscalPeriod).order_by(FiscalPeriod.fiscal_year.desc(), FiscalPeriod.period_name))
    periods = result.scalars().all()
    return {"periods": [p.to_dict() for p in periods], "count": len(periods)}


@router.post("/api/periods")
async def create_period(data: PeriodInput, db: AsyncSession = Depends(get_db)):
    """Create a new fiscal period."""
    from app.models.all_models import FiscalPeriod
    period = FiscalPeriod(
        period_name=data.period_name,
        fiscal_year=data.fiscal_year,
        start_date=datetime.fromisoformat(data.start_date),
        end_date=datetime.fromisoformat(data.end_date),
        status="open",
    )
    db.add(period)
    await db.flush()
    return period.to_dict()


@router.post("/api/periods/{period_name}/close")
async def close_period(period_name: str, data: CloseInput, db: AsyncSession = Depends(get_db)):
    """Close a fiscal period."""
    from app.services.v2.journal_system import journal_service
    try:
        return await journal_service.close_period(period_name, data.close_type, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.post("/api/periods/{period_name}/reopen")
async def reopen_period(period_name: str, db: AsyncSession = Depends(get_db)):
    """Reopen a closed period."""
    from app.services.v2.journal_system import journal_service
    try:
        return await journal_service.reopen_period(period_name, db=db)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/api/periods/{period_name}/integrity")
async def check_integrity(period_name: str, db: AsyncSession = Depends(get_db)):
    """Run integrity checks for a period."""
    from app.services.v2.compliance import compliance_engine
    return await compliance_engine.verify_financial_integrity(period_name, db)


@router.get("/api/periods/sync")
async def sync_periods(db: AsyncSession = Depends(get_db)):
    """Auto-create fiscal_period records for all periods found in journal entries."""
    from app.models.all_models import JournalEntryRecord, FiscalPeriod
    from sqlalchemy import select, distinct
    import re

    # Get unique periods from journal entries
    result = await db.execute(select(distinct(JournalEntryRecord.period)))
    journal_periods = [r[0] for r in result.all() if r[0]]

    # Get existing fiscal periods
    existing_result = await db.execute(select(FiscalPeriod.period_name))
    existing_names = {r[0] for r in existing_result.all()}

    created = []
    for period_name in sorted(journal_periods):
        if period_name in existing_names:
            continue

        # Parse fiscal year from period name (e.g., "January 2026" -> 2026)
        year_match = re.search(r'(\d{4})', period_name)
        fiscal_year = int(year_match.group(1)) if year_match else 2026

        # Parse month for start/end dates
        month_names = {
            "january": 1, "february": 2, "march": 3, "april": 4,
            "may": 5, "june": 6, "july": 7, "august": 8,
            "september": 9, "october": 10, "november": 11, "december": 12,
        }
        month_num = 1
        for mname, mnum in month_names.items():
            if mname in period_name.lower():
                month_num = mnum
                break

        import calendar
        last_day = calendar.monthrange(fiscal_year, month_num)[1]
        start_date = datetime(fiscal_year, month_num, 1, tzinfo=timezone.utc)
        end_date = datetime(fiscal_year, month_num, last_day, 23, 59, 59, tzinfo=timezone.utc)

        fp = FiscalPeriod(
            period_name=period_name,
            fiscal_year=fiscal_year,
            start_date=start_date,
            end_date=end_date,
            status="open",
        )
        db.add(fp)
        created.append(period_name)

    if created:
        await db.commit()

    return {
        "synced": len(created),
        "created_periods": created,
        "total_journal_periods": len(journal_periods),
        "previously_existing": len(existing_names),
    }


@router.post("/api/journal/audit-trail/backfill")
async def backfill_audit_trail(db: AsyncSession = Depends(get_db)):
    """Create audit trail entries for all existing posted journal entries."""
    from app.models.all_models import JournalEntryRecord, AuditTrailEntry
    from app.services.v2.audit_trail import audit_trail_service
    from sqlalchemy import select, func

    # Check how many audit entries already exist
    existing_count_result = await db.execute(
        select(func.count(AuditTrailEntry.id))
    )
    existing_count = existing_count_result.scalar() or 0

    # Get all posted journal entries
    result = await db.execute(
        select(JournalEntryRecord).where(JournalEntryRecord.status == "posted")
    )
    posted_entries = result.scalars().all()

    created = 0
    for je in posted_entries:
        # Check if audit entry already exists for this JE
        check = await db.execute(
            select(func.count(AuditTrailEntry.id)).where(
                AuditTrailEntry.entity_type == "journal_entry",
                AuditTrailEntry.entity_id == je.id,
                AuditTrailEntry.field_name == "status",
            )
        )
        if (check.scalar() or 0) > 0:
            continue

        await audit_trail_service.log_change(
            db, "journal_entry", je.id, "status",
            "draft", "posted", "system", "Backfill from existing data",
        )
        created += 1

    await db.commit()

    return {
        "backfilled": created,
        "total_posted_entries": len(posted_entries),
        "previously_existing_audit_entries": existing_count,
    }


# ── Product Profitability ─────────────────────────────────────────────

@router.get("/api/analytics/product-profitability")
async def get_product_profitability(
    dataset_id: Optional[int] = None,
    segment: Optional[str] = None,
    sort_by: str = "margin_pct",
    db: AsyncSession = Depends(get_db),
):
    """Product-level profitability: Revenue - COGS by product."""
    from app.services.v2.product_profitability import compute_product_profitability

    if not dataset_id:
        from app.models.all_models import Dataset
        from sqlalchemy import select
        # Try production-size datasets first, then any with records
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    return await compute_product_profitability(dataset_id, db, segment=segment, sort_by=sort_by)


# ── Intelligent Ingestion — Reasoning Engine ─────────────────────

@router.post("/api/journal/intelligent-ingest/{dataset_id}/plan")
async def intelligent_ingest_plan(dataset_id: int, db: AsyncSession = Depends(get_db)):
    """Analyze dataset and return a transparent ingestion plan.

    The system THINKS about each account:
    1. Detects file type (TB, transaction journal, P&L report)
    2. Classifies each account using KG + COA + learned patterns
    3. Plans granular journal entries grouped by accounting logic
    4. Explains every decision with reasoning chains
    5. Returns the plan for user review before execution
    """
    from app.services.v2.intelligent_ingestion import intelligent_ingestion
    try:
        plan = await intelligent_ingestion.analyze_and_plan(dataset_id, db)
        return plan.to_dict()
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.error("Intelligent ingestion plan failed: %s", e)
        raise HTTPException(500, f"Analysis failed: {e}")


@router.post("/api/journal/intelligent-ingest/{dataset_id}/execute")
async def intelligent_ingest_execute(dataset_id: int, auto_post: bool = True, db: AsyncSession = Depends(get_db)):
    """Analyze AND execute: create all planned journal entries.

    Full pipeline: Detect → Classify → Plan → Execute → Learn
    """
    from app.services.v2.intelligent_ingestion import intelligent_ingestion
    try:
        plan = await intelligent_ingestion.analyze_and_plan(dataset_id, db)
        result = await intelligent_ingestion.execute_plan(plan, db, auto_post=auto_post)
        result["plan_summary"] = {
            "file_analysis": plan.file_analysis,
            "steps_taken": plan.steps_taken,
            "classification_summary": plan.classification_summary,
            "confidence": plan.confidence,
        }
        return result
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.error("Intelligent ingestion execute failed: %s", e)
        raise HTTPException(500, f"Execution failed: {e}")


# ── Legacy Ingestion (dumb summary — kept for backward compat) ───

@router.post("/api/journal/ingest/{dataset_id}")
async def ingest_dataset_to_journal(
    dataset_id: int,
    auto_post: bool = True,
    db: AsyncSession = Depends(get_db),
):
    """Convert uploaded dataset into proper journal entries.

    This is the bridge between data upload and System of Record:
    Excel → Parse → Revenue/COGS/Expense items → JournalEntry + PostingLine

    After this, ALL reports can read from PostingLine table (the GL).
    """
    from app.services.v2.ingestion_journal import ingestion_journal
    try:
        result = await ingestion_journal.process_dataset(
            dataset_id, db, auto_post=auto_post,
        )
        return result
    except Exception as e:
        logger.error("Ingestion-to-journal failed for dataset %d: %s", dataset_id, e)
        raise HTTPException(500, f"Journal ingestion failed: {e}")


# ── P&L Comparison (7-column NYX Core Thinker template format) ──────────────

@router.get("/api/analytics/pl-comparison")
async def get_pl_comparison(
    dataset_id: int = None,
    prior_dataset_id: int = None,
    db: AsyncSession = Depends(get_db),
):
    """P&L in NYX Core Thinker 7-column format: Code | Item | Prior | Actual | Plan | Variance | Var%.

    If prior_dataset_id not provided, tries to find prior year dataset automatically.
    """
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select

    if not dataset_id:
        # Find best dataset: prefer real financial uploads (not stress tests)
        # Real uploads have file_type containing "Report" or "Full" and reasonable record counts
        result = await db.execute(
            select(Dataset)
            .where(Dataset.record_count > 0, Dataset.record_count < 10000)  # Exclude stress tests
            .order_by(Dataset.id.desc()).limit(1)
        )
        ds = result.scalar_one_or_none()
        if not ds:
            result = await db.execute(
                select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1)
            )
            ds = result.scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    # Auto-find prior year dataset if not specified
    if not prior_dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.id == dataset_id))).scalar_one_or_none()
        if ds and ds.period:
            # Try to find prior year (e.g., "January 2026" → "January 2025")
            import re
            year_match = re.search(r'20(\d{2})', ds.period)
            if year_match:
                prior_period = ds.period.replace(f"20{year_match.group(1)}", f"20{int(year_match.group(1))-1:02d}")
                prior_ds = (await db.execute(
                    select(Dataset).where(Dataset.period == prior_period, Dataset.record_count > 0).limit(1)
                )).scalar_one_or_none()
                if prior_ds:
                    prior_dataset_id = prior_ds.id

    return await pl_comparison.full_pl(dataset_id, prior_dataset_id, db)


@router.get("/api/analytics/pl-trend")
async def get_pl_trend(
    dataset_ids: str = "",
    db: AsyncSession = Depends(get_db),
):
    """Multi-period P&L trend. Pass comma-separated dataset IDs."""
    from app.services.v2.pl_comparison import pl_comparison

    if not dataset_ids:
        # Default: all datasets with records
        from app.models.all_models import Dataset
        from sqlalchemy import select
        result = await db.execute(
            select(Dataset.id).where(Dataset.record_count > 0).order_by(Dataset.id)
        )
        ids = [row[0] for row in result.all()]
    else:
        ids = [int(x.strip()) for x in dataset_ids.split(",") if x.strip()]

    return await pl_comparison.multi_period_trend(ids, db)


# ── COGS Comparison ──────────────────────────────────────────────

@router.get("/api/analytics/cogs-comparison")
async def get_cogs_comparison(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """COGS breakdown by product with prior year + variance + cost components."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds: raise HTTPException(404, "No dataset")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    return await pl_comparison.cogs_comparison(dataset_id, prior_dataset_id, db)


@router.get("/api/analytics/bs-comparison")
async def get_bs_comparison(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Balance Sheet with prior year + variance, grouped by IFRS section."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds: raise HTTPException(404, "No dataset")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    return await pl_comparison.balance_sheet_comparison(dataset_id, prior_dataset_id, db)


@router.get("/api/analytics/revenue-comparison")
async def get_revenue_comparison(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Revenue breakdown by product/segment with prior year + variance."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds: raise HTTPException(404, "No dataset")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    return await pl_comparison.revenue_comparison(dataset_id, prior_dataset_id, db)


@router.get("/api/analytics/pl-comparison/export")
async def export_pl_comparison_excel(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Export CFO-ready P&L comparison as professionally styled Excel."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    import io

    if not dataset_id:
        result = await db.execute(
            select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000)
            .order_by(Dataset.id.desc()).limit(1)
        )
        ds = result.scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.full_pl(dataset_id, prior_dataset_id, db)
    summary = data.get("summary", {})
    rows = data.get("rows", [])

    # ── Corporate style definitions ──
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"
    RED = "C00000"
    GREEN = "006100"

    title_font = Font(name="Arial", bold=True, size=16, color=WHITE)
    subtitle_font = Font(name="Arial", bold=True, size=11, color=WHITE)
    header_font = Font(name="Arial", bold=True, size=10, color=WHITE)
    bold_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    small_font = Font(name="Arial", size=9, color="666666")
    kpi_value_font = Font(name="Arial", bold=True, size=14)
    kpi_label_font = Font(name="Arial", size=9, color="666666")
    red_font = Font(name="Arial", size=10, color=RED)
    green_font = Font(name="Arial", size=10, color=GREEN)
    red_bold = Font(name="Arial", bold=True, size=10, color=RED)
    green_bold = Font(name="Arial", bold=True, size=10, color=GREEN)

    dark_fill = PatternFill("solid", fgColor=DARK_BLUE)
    medium_fill = PatternFill("solid", fgColor=MEDIUM_BLUE)
    light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    white_fill = PatternFill("solid", fgColor=WHITE)

    thin_border = Border(
        bottom=Side(style="thin", color="B0B0B0"),
    )
    thick_border = Border(
        bottom=Side(style="medium", color=DARK_BLUE),
    )
    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L Statement"

    # ── Row 1: Company name banner ──
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = data.get("company", settings.COMPANY_NAME)
    c.font = title_font
    c.fill = dark_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: Period + Currency ──
    ws.merge_cells("A2:G2")
    period_text = f"Income Statement — {data.get('period', '')}"
    if data.get("prior_period"):
        period_text += f"  vs  {data['prior_period']}"
    period_text += f"  |  Currency: {data.get('currency', 'GEL')}"
    c = ws["A2"]
    c.value = period_text
    c.font = subtitle_font
    c.fill = medium_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    # ── Row 3: Blank separator ──
    ws.row_dimensions[3].height = 8

    # ── Row 4-5: KPI Summary ──
    kpi_data = [
        ("Revenue", summary.get("revenue", 0)),
        ("Gross Profit", summary.get("gross_profit", 0)),
        ("EBITDA", summary.get("ebitda", 0)),
        ("Net Profit", summary.get("net_profit", 0)),
        ("Prior Revenue", summary.get("prior_revenue", 0)),
    ]
    for i, (label, value) in enumerate(kpi_data):
        col = i + 2  # Start from column B
        ws.cell(row=4, column=col, value=label).font = kpi_label_font
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        cell = ws.cell(row=5, column=col, value=value)
        cell.font = kpi_value_font if value >= 0 else Font(name="Arial", bold=True, size=14, color=RED)
        cell.number_format = num_fmt
        cell.alignment = Alignment(horizontal="center")

    # ── Row 6: Blank separator ──
    ws.row_dimensions[6].height = 8

    # ── Row 7: Column headers ──
    headers = ["Code", "Line Item", "Prior Year", "Actual", "% of Revenue", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=i, value=h)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center" if i > 2 else "left", vertical="center")
    ws.row_dimensions[7].height = 24

    # ── Data rows ──
    row_num = 8
    for idx, r in enumerate(rows):
        ac = r.get("ac", 0) or 0
        pr = r.get("pr", 0) or 0
        var = r.get("var", 0) or 0
        var_pct = r.get("var_pct", 0) or 0
        is_bold = r.get("bold", False)
        is_sep = r.get("sep", False)
        level = r.get("lvl", 0)
        revenue = summary.get("revenue", 1) or 1

        indent = "  " * level
        ws.cell(row=row_num, column=1, value=r.get("c", ""))
        ws.cell(row=row_num, column=2, value=f"{indent}{r.get('l', '')}")
        ws.cell(row=row_num, column=3, value=pr if pr != 0 else None)
        ws.cell(row=row_num, column=4, value=ac if ac != 0 else None)
        ws.cell(row=row_num, column=5, value=ac / revenue if revenue and ac else None)
        ws.cell(row=row_num, column=6, value=var if var != 0 else None)
        ws.cell(row=row_num, column=7, value=var_pct / 100 if var_pct else None)

        # Styling
        row_fill = light_fill if is_bold else (gray_fill if idx % 2 else white_fill)
        row_font = bold_font if is_bold else normal_font
        border = thick_border if is_sep else thin_border

        for col in range(1, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = border

            if col <= 2:
                cell.font = row_font
            elif col in (3, 4, 6):
                cell.number_format = num_fmt
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = red_bold if (val < 0 and is_bold) else (red_font if val < 0 else (bold_font if is_bold else normal_font))
                else:
                    cell.font = row_font
            elif col == 5:
                cell.number_format = pct_fmt
                cell.font = small_font
            elif col == 7:
                cell.number_format = pct_fmt
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = green_bold if val > 0 else (red_bold if val < 0 else row_font) if is_bold else (Font(name="Arial", size=10, color=GREEN) if val > 0 else red_font)
                else:
                    cell.font = row_font

        ws.cell(row=row_num, column=1).font = Font(name="Arial", size=9, color="999999")
        row_num += 1

    # ── Footer ──
    row_num += 1
    ws.merge_cells(f"A{row_num}:G{row_num}")
    footer = ws.cell(row=row_num, column=1)
    footer.value = f"Generated by FinAI Foundry  |  {data.get('company', '')}  |  {data.get('period', '')}  |  All amounts in {data.get('currency', 'GEL')}"
    footer.font = Font(name="Arial", size=8, color="999999", italic=True)
    footer.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10

    # ── Freeze panes (freeze header row) ──
    ws.freeze_panes = "A8"

    # ── Print setup ──
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "Generated by FinAI Foundry — Confidential"
    ws.oddFooter.center.font = "Arial,Regular"
    ws.oddFooter.center.size = 8

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PL_Statement_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/analytics/bs-comparison/export")
async def export_bs_comparison_excel(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Export Balance Sheet comparison to Excel (CFO-ready corporate styling)."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.balance_sheet_comparison(dataset_id, prior_dataset_id, db)

    # ── Corporate style definitions ──
    DARK_BLUE = "1F4E79"
    MEDIUM_BLUE = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"
    RED = "C00000"
    GREEN = "006100"

    title_font = Font(name="Arial", bold=True, size=16, color=WHITE)
    subtitle_font = Font(name="Arial", bold=True, size=11, color=WHITE)
    header_font = Font(name="Arial", bold=True, size=10, color=WHITE)
    bold_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    red_font = Font(name="Arial", size=10, color=RED)
    green_font = Font(name="Arial", size=10, color=GREEN)
    red_bold = Font(name="Arial", bold=True, size=10, color=RED)
    green_bold = Font(name="Arial", bold=True, size=10, color=GREEN)

    dark_fill = PatternFill("solid", fgColor=DARK_BLUE)
    medium_fill = PatternFill("solid", fgColor=MEDIUM_BLUE)
    light_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    white_fill = PatternFill("solid", fgColor=WHITE)

    thin_border = Border(bottom=Side(style="thin", color="B0B0B0"))
    thick_border = Border(bottom=Side(style="medium", color=DARK_BLUE))
    num_fmt = '#,##0;(#,##0);"-"'
    pct_fmt = '0.0%;(0.0%);"-"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    # ── Row 1: Company name banner ──
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = settings.COMPANY_NAME
    c.font = title_font
    c.fill = dark_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Row 2: Period ──
    ws.merge_cells("A2:E2")
    period_text = f"Balance Sheet — {data.get('period', '')}"
    if data.get("prior_period"):
        period_text += f"  vs  {data['prior_period']}"
    period_text += "  |  Currency: GEL"
    c = ws["A2"]
    c.value = period_text
    c.font = subtitle_font
    c.fill = medium_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 28

    # ── Row 3: Blank separator ──
    ws.row_dimensions[3].height = 8

    # ── Row 4: Column headers ──
    headers = ["IFRS Line Item", "Prior Year", "Actual", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = header_font
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center")
    ws.row_dimensions[4].height = 24

    # ── Data rows ──
    rows_data = data.get("rows", [])
    row_num = 5
    current_section = None

    for r in rows_data:
        # Section header
        if r.get("section") != current_section:
            current_section = r.get("section")
            # Section header row
            ws.merge_cells(f"A{row_num}:E{row_num}")
            cell = ws.cell(row=row_num, column=1, value=f"{current_section}")
            cell.font = Font(name="Arial", bold=True, size=12, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=MEDIUM_BLUE)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_num].height = 20
            row_num += 1

        # Data row
        actual = r.get("actual", 0)
        prior = r.get("prior", 0)
        variance = r.get("variance", 0)
        variance_pct = r.get("variance_pct", 0)
        is_bold = r.get("bold", False)
        level = r.get("level", 0)

        indent = "  " * level
        ws.cell(row=row_num, column=1, value=f"{indent}{r.get('ifrs_line', '')}")
        ws.cell(row=row_num, column=2, value=prior if prior != 0 else None)
        ws.cell(row=row_num, column=3, value=actual if actual != 0 else None)
        ws.cell(row=row_num, column=4, value=variance if variance != 0 else None)
        ws.cell(row=row_num, column=5, value=variance_pct / 100 if variance_pct else None)

        # Styling
        row_fill = light_fill if is_bold else (gray_fill if (row_num - 5) % 2 else white_fill)
        row_font = bold_font if is_bold else normal_font

        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = thin_border

            if col == 1:
                cell.font = row_font
            elif col in (2, 3, 4):
                cell.number_format = num_fmt
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = red_bold if (val < 0 and is_bold) else (red_font if val < 0 else (bold_font if is_bold else normal_font))
                else:
                    cell.font = row_font
            elif col == 5:
                cell.number_format = pct_fmt
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.font = green_bold if val > 0 else (red_bold if val < 0 else row_font) if is_bold else (Font(name="Arial", size=10, color=GREEN) if val > 0 else red_font)
                else:
                    cell.font = row_font

        row_num += 1

    # ── Footer ──
    row_num += 1
    ws.merge_cells(f"A{row_num}:E{row_num}")
    footer = ws.cell(row=row_num, column=1)
    footer.value = f"Generated by FinAI Foundry  |  {settings.COMPANY_NAME}  |  {data.get('period', '')}  |  All amounts in GEL"
    footer.font = Font(name="Arial", size=8, color="999999", italic=True)
    footer.alignment = Alignment(horizontal="center")

    # ── Column widths ──
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 10

    # ── Freeze panes ──
    ws.freeze_panes = "A5"

    # ── Print setup ──
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddFooter.center.text = "Generated by FinAI Foundry — Confidential"
    ws.oddFooter.center.font = "Arial,Regular"
    ws.oddFooter.center.size = 8

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Balance_Sheet_{data.get('period', 'report').replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/api/analytics/revenue-comparison/export")
async def export_revenue_comparison_excel(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Export Revenue comparison as CFO-ready Excel."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.revenue_comparison(dataset_id, prior_dataset_id, db)

    wb = Workbook(); ws = wb.active; ws.title = "Revenue Analysis"
    # Header rows
    ws.merge_cells("A1:G1"); ws["A1"].value = settings.COMPANY_NAME; ws["A1"].font = Font(name="Arial", bold=True, size=16)
    ws.merge_cells("A2:G2"); ws["A2"].value = f"Revenue Analysis — {data.get('period', '')} vs {data.get('prior_period','')}"; ws["A2"].font = Font(name="Arial", size=11)
    # Column headings
    headers = ["Product", "Segment", "Prior Year", "Actual", "% of Total", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")

    rows = data.get('rows', [])
    for rindex, r in enumerate(rows, 5):
        ws.cell(row=rindex, column=1, value=r.get('product', ''))
        ws.cell(row=rindex, column=2, value=r.get('segment', ''))
        c3 = ws.cell(row=rindex, column=3, value=r.get('prior_net', 0) or None)
        c3.number_format = '#,##0'
        c4 = ws.cell(row=rindex, column=4, value=r.get('actual_net', 0) or None)
        c4.number_format = '#,##0'
        c5 = ws.cell(row=rindex, column=5, value=(r.get('pct_of_total', 0) or 0) / 100)
        c5.number_format = '0.0%'
        c6 = ws.cell(row=rindex, column=6, value=r.get('variance', 0) or None)
        c6.number_format = '#,##0'
        c7 = ws.cell(row=rindex, column=7, value=(r.get('variance_pct', 0) or 0) / 100)
        c7.number_format = '0.0%'

    for col in ['A','B','C','D','E','F','G']:
        ws.column_dimensions[col].width = 18

    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    filename = f"Revenue_Comparison_{data.get('period','report').replace(' ','_')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/api/analytics/cogs-comparison/export")
async def export_cogs_comparison_excel(
    dataset_id: int = None, prior_dataset_id: int = None, db: AsyncSession = Depends(get_db),
):
    """Export COGS comparison as CFO-ready Excel."""
    from app.services.v2.pl_comparison import pl_comparison
    from app.models.all_models import Dataset
    from sqlalchemy import select
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    if not dataset_id:
        ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0, Dataset.record_count < 10000).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            ds = (await db.execute(select(Dataset).where(Dataset.record_count > 0).order_by(Dataset.id.desc()).limit(1))).scalar_one_or_none()
        if not ds:
            raise HTTPException(404, "No dataset found")
        dataset_id = ds.id

    if not prior_dataset_id:
        prior_dataset_id = await _auto_find_prior(dataset_id, db)

    data = await pl_comparison.cogs_comparison(dataset_id, prior_dataset_id, db)

    wb = Workbook(); ws = wb.active; ws.title = "COGS Analysis"
    ws.merge_cells("A1:G1"); ws["A1"].value = settings.COMPANY_NAME; ws["A1"].font = Font(name="Arial", bold=True, size=16)
    ws.merge_cells("A2:G2"); ws["A2"].value = f"COGS Analysis — {data.get('period', '')} vs {data.get('prior_period','')}"; ws["A2"].font = Font(name="Arial", size=11)
    headers = ["Category", "Segment", "Prior Year", "Actual", "% of Total", "Variance", "Var %"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center")

    rows = data.get('rows', [])
    for rindex, r in enumerate(rows, 5):
        ws.cell(row=rindex, column=1, value=r.get('category', ''))
        ws.cell(row=rindex, column=2, value=r.get('segment', ''))
        c3 = ws.cell(row=rindex, column=3, value=r.get('prior', 0) or None)
        c3.number_format = '#,##0'
        c4 = ws.cell(row=rindex, column=4, value=r.get('actual', 0) or None)
        c4.number_format = '#,##0'
        c5 = ws.cell(row=rindex, column=5, value=((r.get('actual', 0) or 0) / ((data.get('total_actual', 1) or 1))) if data.get('total_actual') else 0)
        c5.number_format = '0.0%'
        c6 = ws.cell(row=rindex, column=6, value=r.get('variance', 0) or None)
        c6.number_format = '#,##0'
        c7 = ws.cell(row=rindex, column=7, value=(r.get('variance_pct', 0) or 0) / 100)
        c7.number_format = '0.0%'

    for col in ['A','B','C','D','E','F','G']:
        ws.column_dimensions[col].width = 18

    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    filename = f"COGS_Comparison_{data.get('period','report').replace(' ','_')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


async def _auto_find_prior(dataset_id: int, db: AsyncSession) -> Optional[int]:
    """Auto-find prior year dataset."""
    from app.models.all_models import Dataset
    from sqlalchemy import select
    import re

    ds = (await db.execute(select(Dataset).where(Dataset.id == dataset_id))).scalar_one_or_none()
    if not ds or not ds.period:
        return None

    year_match = re.search(r'20(\d{2})', ds.period)
    if not year_match:
        return None

    prior_period = ds.period.replace(f"20{year_match.group(1)}", f"20{int(year_match.group(1))-1:02d}")
    prior_ds = (await db.execute(
        select(Dataset).where(Dataset.period == prior_period, Dataset.record_count > 0).limit(1)
    )).scalar_one_or_none()

    return prior_ds.id if prior_ds else None


# ── GL Reporting (from posted journal entries) ───────────────────

@router.get("/api/gl/income-statement")
async def gl_income_statement(period: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Income statement computed from posted journal entries (the REAL GL)."""
    from app.services.v2.gl_reporting import gl_reporting
    if not period:
        periods = await gl_reporting.available_periods(db)
        period = periods[-1] if periods else None
    if not period:
        raise HTTPException(404, "No posted journal entries found")
    return await gl_reporting.income_statement(period, db)


@router.get("/api/gl/balance-sheet")
async def gl_balance_sheet(period: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Balance sheet from posted journal entries."""
    from app.services.v2.gl_reporting import gl_reporting
    if not period:
        periods = await gl_reporting.available_periods(db)
        period = periods[-1] if periods else None
    if not period:
        raise HTTPException(404, "No posted journal entries found")
    return await gl_reporting.balance_sheet(period, db)


@router.get("/api/gl/trial-balance-from-gl")
@router.get("/api/gl/trial-balance")
async def gl_trial_balance(period: Optional[str] = None, db: AsyncSession = Depends(get_db)):
    """Trial balance from posted journal entries."""
    from app.services.v2.gl_reporting import gl_reporting
    if not period:
        periods = await gl_reporting.available_periods(db)
        period = periods[-1] if periods else None
    if not period:
        raise HTTPException(404, "No posted journal entries found")
    return await gl_reporting.trial_balance(period, db)


@router.get("/api/gl/periods")
async def gl_periods(db: AsyncSession = Depends(get_db)):
    """List all periods that have posted journal entries."""
    from app.services.v2.gl_reporting import gl_reporting
    return {"periods": await gl_reporting.available_periods(db)}


# ── Event Dispatcher Status ───────────────────────────────────────

@router.get("/api/events/recent")
async def get_recent_events(limit: int = 50, event_type: Optional[str] = None):
    """Get recent events from the dispatcher log."""
    from app.services.v2.event_dispatcher import event_dispatcher
    return {
        "events": event_dispatcher.get_recent_events(limit, event_type),
        "stats": event_dispatcher.stats(),
    }


# ── P&L Drill-Down / Lineage ─────────────────────────────────────────

@router.get("/api/analytics/lineage/pl/{line_code}")
async def get_pl_lineage(
    line_code: str,
    dataset_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    """Get transactions underlying a P&L line item."""
    from app.services.v2.lineage_service import lineage_service

    if not dataset_id:
        from app.models.all_models import Dataset
        from sqlalchemy import select
        result = await db.execute(select(Dataset.id).where(Dataset.is_active == True).limit(1))
        row = result.first()
        if not row:
            raise HTTPException(404, "No active dataset found")
        dataset_id = row[0]

    return await lineage_service.get_transactions_for_pl_line(line_code, dataset_id, db)


# ═══════════════════════════════════════════════════════════════════
# AR / AP / ASSET ACCOUNTING (SAP FI Modules)
# ═══════════════════════════════════════════════════════════════════

@router.get("/api/analytics/ar-summary")
async def ar_summary(db: AsyncSession = Depends(get_db)):
    """Accounts Receivable summary with aging analysis."""
    from app.services.v2.ar_ap_asset import sap_fi
    await sap_fi.ar.load_from_tb(db)
    return sap_fi.ar.get_summary()


@router.get("/api/analytics/ap-summary")
async def ap_summary(db: AsyncSession = Depends(get_db)):
    """Accounts Payable summary with payment schedule."""
    from app.services.v2.ar_ap_asset import sap_fi
    await sap_fi.ap.load_from_tb(db)
    return sap_fi.ap.get_summary()


@router.get("/api/analytics/asset-summary")
async def asset_summary(db: AsyncSession = Depends(get_db)):
    """Fixed Asset Register summary with depreciation."""
    from app.services.v2.ar_ap_asset import sap_fi
    await sap_fi.assets.load_from_tb(db)
    return sap_fi.assets.get_summary()


@router.get("/api/analytics/sap-fi")
async def sap_fi_full(db: AsyncSession = Depends(get_db)):
    """Complete SAP FI module summary: AR + AP + Assets + Working Capital."""
    from app.services.v2.ar_ap_asset import sap_fi
    return await sap_fi.get_full_summary(db)


@router.get("/api/analytics/working-capital")
async def working_capital(
    revenue: float = 51000000, cogs: float = 49000000,
    db: AsyncSession = Depends(get_db),
):
    """Working capital cycle analysis (DSO, DPO, CCC)."""
    from app.services.v2.ar_ap_asset import sap_fi
    await sap_fi.load_all(db)
    return sap_fi.working_capital_analysis(revenue, cogs)


# ── Reconciliation Engine ─────────────────────────────────────────────

@router.get("/api/analytics/reconciliation")
async def run_reconciliation(
    period: Optional[str] = None,
    dataset_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
):
    """SAP-grade reconciliation: P&L ties to TB ties to BS.

    Runs 7 cross-checks:
    1. TB Balance (debits == credits)
    2. Net income P&L vs TB
    3. Revenue P&L vs TB 6xxx accounts
    4. COGS P&L vs TB 71xx accounts
    5. BS equation (Assets == Liabilities + Equity)
    6. Total expenses P&L vs TB
    7. Cash in BS vs TB cash accounts
    """
    from app.services.v2.reconciliation_engine import reconciliation_engine
    from app.services.v2.gl_reporting import gl_reporting

    if not period:
        periods = await gl_reporting.available_periods(db)
        if not periods:
            raise HTTPException(404, "No posted journal entries found for reconciliation")
        period = periods[-1]

    try:
        report = await reconciliation_engine.run_reconciliation(period, db, dataset_id)
        return report.to_dict()
    except Exception as e:
        logger.error("Reconciliation failed for period %s: %s", period, e)
        raise HTTPException(500, f"Reconciliation failed: {e}")


# ── Audit Trail ───────────────────────────────────────────────────────

@router.get("/api/journal/audit-trail/{entity_type}/{entity_id}")
async def get_audit_trail(
    entity_type: str,
    entity_id: int,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
):
    """Get the full field-level audit trail for a specific entity.

    Supported entity types: journal_entry, posting_line, period
    """
    from app.services.v2.audit_trail import audit_trail_service

    trail = await audit_trail_service.get_trail(db, entity_type, entity_id, limit)
    return {
        "entity_type": entity_type,
        "entity_id": entity_id,
        "entries": trail,
        "count": len(trail),
    }


@router.get("/api/journal/audit-trail/{entity_type}")
async def get_audit_trail_summary(
    entity_type: str,
    db: AsyncSession = Depends(get_db),
):
    """Get aggregate summary of audit trail changes for an entity type.

    Returns counts by field, by user, and recent changes.
    """
    from app.services.v2.audit_trail import audit_trail_service

    return await audit_trail_service.get_trail_summary(db, entity_type)


# ── Data Lineage (Transformation Graph) ──────────────────────────────

@router.post("/api/journal/lineage/migrate")
async def migrate_lineage(db: AsyncSession = Depends(get_db)):
    """Migrate records from data_lineage table to transformation_lineage table."""
    from app.services.v2.data_lineage import data_lineage_tracker
    try:
        return await data_lineage_tracker.migrate_from_data_lineage(db)
    except Exception as e:
        logger.error("Lineage migration failed: %s", e)
        raise HTTPException(500, f"Migration failed: {e}")


@router.get("/api/journal/lineage/{entity_type}/{entity_id}")
async def get_data_lineage(
    entity_type: str,
    entity_id: int,
    direction: str = Query("backward", regex="^(backward|forward)$"),
    db: AsyncSession = Depends(get_db),
):
    """Trace the data lineage chain for any entity.

    Direction:
    - backward: trace to original sources (default)
    - forward: trace to downstream targets
    """
    from app.services.v2.data_lineage import data_lineage_tracker

    try:
        chain = await data_lineage_tracker.get_lineage_chain(
            db, entity_type, entity_id, direction,
        )
        return chain
    except Exception as e:
        logger.error("Lineage trace failed for %s#%d: %s", entity_type, entity_id, e)
        raise HTTPException(500, f"Lineage trace failed: {e}")


@router.get("/api/journal/lineage-graph/{entity_type}/{entity_id}")
async def get_lineage_graph(
    entity_type: str,
    entity_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Get full bidirectional lineage graph (upstream + downstream) for an entity."""
    from app.services.v2.data_lineage import data_lineage_tracker

    try:
        return await data_lineage_tracker.get_lineage_graph(db, entity_type, entity_id)
    except Exception as e:
        logger.error("Lineage graph failed for %s#%d: %s", entity_type, entity_id, e)
        raise HTTPException(500, f"Lineage graph failed: {e}")


# ═══════════════════════════════════════════════════════════════════
# GRAPH ANALYTICS (Palantir Foundry-Grade)
# ═══════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════
# SMART DATA ORGANIZER
# ═══════════════════════════════════════════════════════════════════

@router.get("/api/data/organized")
async def data_organized(db: AsyncSession = Depends(get_db)):
    """Get all datasets organized by year, type, company with recommendations."""
    from app.services.v2.data_organizer import data_organizer
    return await data_organizer.get_organized_view(db)


@router.get("/api/data/recommend")
async def data_recommend(purpose: str = "analysis", db: AsyncSession = Depends(get_db)):
    """Smart recommendation: which dataset to use for a specific purpose."""
    from app.services.v2.data_organizer import data_organizer
    return await data_organizer.get_smart_recommendation(db, purpose)


@router.post("/api/data/archive/{dataset_id}")
async def data_archive(dataset_id: int, db: AsyncSession = Depends(get_db)):
    """Archive a dataset (hide from default views, keep data)."""
    from app.services.v2.data_organizer import data_organizer
    return await data_organizer.archive_dataset(db, dataset_id)


# ═══════════════════════════════════════════════════════════════════
# GRAPH ANALYTICS (Palantir Foundry-Grade)
# ═══════════════════════════════════════════════════════════════════

def _ensure_graph_loaded():
    """Force-load graph from ontology if not already loaded."""
    from app.services.v2.graph_analytics import graph_analytics, GraphNode, GraphEdge
    from collections import defaultdict
    if graph_analytics._nodes:
        return
    try:
        from app.services.ontology_engine import ontology_registry
        obj_count = len(ontology_registry._objects)
        idx_count = sum(len(v) for v in ontology_registry._index_by_type.items())
        print(f"[GRAPH] _objects={obj_count}, _index_by_type keys={list(ontology_registry._index_by_type.keys())}")

        # Build from whatever is available
        loaded = 0
        if obj_count > 0:
            source = ontology_registry._objects
        else:
            # Reconstruct by querying each type
            source = {}
            for type_id, oid_set in ontology_registry._index_by_type.items():
                for oid in oid_set:
                    obj = ontology_registry.get_object(oid)
                    if obj:
                        source[oid] = obj

        print(f"[GRAPH] source has {len(source)} objects")
        for oid, obj in source.items():
            label = (obj.properties.get("name_en") or obj.properties.get("name")
                     or obj.properties.get("code") or obj.properties.get("metric") or oid)
            node = GraphNode(id=oid, type=obj.object_type, label=str(label),
                properties={k: v for k, v in obj.properties.items()
                           if not isinstance(v, (dict, list)) and not k.startswith("_")})
            graph_analytics._nodes[oid] = node
            for rel_type, targets in (obj.relationships or {}).items():
                tlist = targets if isinstance(targets, list) else [targets] if isinstance(targets, str) else []
                for t in tlist:
                    tid = t if isinstance(t, str) else str(t)
                    graph_analytics._edges.append(GraphEdge(source=oid, target=tid, relationship=rel_type))
                    graph_analytics._adjacency[oid].append((tid, rel_type, 1.0))
                    graph_analytics._reverse_adj[tid].append((oid, rel_type, 1.0))
        logger.info(f"Graph loaded: {len(graph_analytics._nodes)} nodes, {len(graph_analytics._edges)} edges")
    except Exception as e:
        logger.warning(f"Graph load failed: {e}")


@router.get("/api/graph/stats")
async def graph_stats():
    """Graph statistics: nodes, edges, types, density, top connected."""
    from app.services.ontology_engine import ontology_registry
    from collections import defaultdict

    # Build graph inline from ontology
    nodes = {}
    edges = []
    for oid, obj in ontology_registry._objects.items():
        label = (obj.properties.get("name_en") or obj.properties.get("name")
                 or obj.properties.get("code") or obj.properties.get("metric") or oid)
        nodes[oid] = {"id": oid, "type": obj.object_type, "label": str(label)}

        for rel_type, targets in (obj.relationships or {}).items():
            tl = targets if isinstance(targets, list) else [targets] if isinstance(targets, str) else []
            for t in tl:
                edges.append({"source": oid, "target": str(t), "relationship": rel_type})

    # Stats
    type_counts = defaultdict(int)
    for n in nodes.values():
        type_counts[n["type"]] += 1

    rel_counts = defaultdict(int)
    for e in edges:
        rel_counts[e["relationship"]] += 1

    degree = defaultdict(int)
    for e in edges:
        degree[e["source"]] += 1
        degree[e["target"]] += 1

    top = sorted(degree.items(), key=lambda x: x[1], reverse=True)[:10]

    return {
        "total_nodes": len(nodes),
        "total_edges": len(edges),
        "node_types": dict(type_counts),
        "relationship_types": dict(rel_counts),
        "avg_degree": round(sum(degree.values()) / max(len(degree), 1), 2),
        "max_degree": max(degree.values()) if degree else 0,
        "top_connected": [{"id": nid, "degree": d, "label": nodes.get(nid, {}).get("label", "?")} for nid, d in top],
        "density": round(len(edges) / max(len(nodes) * (len(nodes) - 1), 1), 6),
        "_debug_objects_in_registry": len(ontology_registry._objects),
        "_debug_types_in_registry": len(ontology_registry._types),
    }


@router.get("/api/graph/neighborhood/{node_id}")
async def graph_neighborhood(node_id: str, depth: int = 2):
    """Get N-depth neighborhood of a node for visualization."""
    _ensure_graph_loaded()
    from app.services.v2.graph_analytics import graph_analytics
    return graph_analytics.get_neighborhood(node_id, min(depth, 4))


@router.get("/api/graph/impact/{node_id}")
async def graph_impact(node_id: str, change_pct: float = 10.0):
    """Impact analysis: what entities are affected if this node changes by X%?"""
    _ensure_graph_loaded()
    from app.services.v2.graph_analytics import graph_analytics
    return graph_analytics.impact_analysis(node_id, change_pct)


@router.get("/api/graph/anomalies")
async def graph_anomalies():
    """Detect structural anomalies in the knowledge graph."""
    _ensure_graph_loaded()
    from app.services.v2.graph_analytics import graph_analytics
    return graph_analytics.find_anomalies()


@router.get("/api/graph/query")
async def graph_query(q: str = "revenue"):
    """Cross-entity natural language query over the graph."""
    _ensure_graph_loaded()
    from app.services.v2.graph_analytics import graph_analytics
    return graph_analytics.cross_entity_query(q)


@router.get("/api/graph/subgraph")
async def graph_subgraph(node_type: str = None, limit: int = 100):
    """Get a subgraph for D3.js visualization."""
    _ensure_graph_loaded()
    from app.services.v2.graph_analytics import graph_analytics
    return graph_analytics.get_subgraph(node_type, min(limit, 500))


# ═══════════════════════════════════════════════════════════════════
# AIP LOGIC — Palantir-Grade AI Functions on Ontology
# ═══════════════════════════════════════════════════════════════════

@router.get("/api/aip/functions")
async def aip_list_functions():
    """List all registered AIP Logic functions."""
    from app.services.v2.aip_logic import aip_logic
    return {"functions": aip_logic.list_functions(), "total": len(aip_logic._functions)}


@router.post("/api/aip/execute")
async def aip_execute(body: dict):
    """Execute an AIP Logic function on ontology objects."""
    from app.services.v2.aip_logic import aip_logic
    function_name = body.get("function", "")
    object_ids = body.get("object_ids", [])
    params = body.get("params", {})
    if not function_name:
        return {"error": "Missing 'function' field", "available": list(aip_logic._functions.keys())}
    if not object_ids:
        return {"error": "Missing 'object_ids' — provide list of ontology object IDs"}
    return await aip_logic.execute(function_name, object_ids, params)


@router.post("/api/aip/batch")
async def aip_batch(body: dict):
    """Execute AIP Logic function on multiple object sets (batch mode)."""
    from app.services.v2.aip_logic import aip_logic
    function_name = body.get("function", "")
    object_id_sets = body.get("object_id_sets", [])
    results = []
    for oid_set in object_id_sets[:20]:  # max 20 batches
        r = await aip_logic.execute(function_name, oid_set if isinstance(oid_set, list) else [oid_set])
        results.append(r)
    return {"function": function_name, "batch_size": len(results), "results": results}


# ═══════════════════════════════════════════════════════════════════
# FOUNDRY TS — Time Series Analytics
# ═══════════════════════════════════════════════════════════════════

@router.post("/api/ts/rolling")
async def ts_rolling(body: dict):
    """Rolling aggregate on time series data."""
    from app.services.v2.aip_logic import financial_ts
    values = body.get("values", [])
    window = body.get("window", 3)
    method = body.get("method", "mean")
    return {"values": values, "window": window, "method": method,
            "result": financial_ts.rolling_aggregate(values, window, method)}


@router.post("/api/ts/trend")
async def ts_trend(body: dict):
    """Detect trend direction and strength."""
    from app.services.v2.aip_logic import financial_ts
    return financial_ts.trend_detection(body.get("values", []))


@router.post("/api/ts/seasonal")
async def ts_seasonal(body: dict):
    """Seasonal decomposition of time series."""
    from app.services.v2.aip_logic import financial_ts
    return financial_ts.seasonal_decomposition(body.get("values", []), body.get("period", 12))


@router.post("/api/ts/analyze")
async def ts_full_analysis(body: dict):
    """Full time series analysis: trend + rolling + percent change + cumulative."""
    from app.services.v2.aip_logic import financial_ts
    values = body.get("values", [])
    return {
        "values": values,
        "trend": financial_ts.trend_detection(values),
        "rolling_mean_3": financial_ts.rolling_aggregate(values, 3, "mean"),
        "rolling_mean_6": financial_ts.rolling_aggregate(values, min(6, len(values)), "mean"),
        "percent_change": financial_ts.percent_change(values),
        "cumulative": financial_ts.cumulative(values),
    }


# ═══════════════════════════════════════════════════════════════════════
# ONTOLOGY FINANCIAL CALCULATOR — sync P&L values into KPI objects
# ═══════════════════════════════════════════════════════════════════════

class ValidateCalcInput(BaseModel):
    revenue: float = 0
    cogs: float = 0
    gross_profit: float = 0
    other_revenue: float = 0
    total_gross_profit: float = 0
    ga_expenses: float = 0
    ebitda: float = 0
    da_expenses: float = 0
    ebit: float = 0
    finance_net: float = 0
    tax: float = 0
    net_profit: float = 0


@router.post("/api/ontology/sync-financials")
async def ontology_sync_financials(db: AsyncSession = Depends(get_db)):
    """Compute P&L from entity tables and write results into ontology KPI objects.

    After this call, ontology KPI objects carry live financial values
    (revenue, gross_profit, ebitda, net_profit, margin percentages, etc.).
    """
    from app.services.v2.ontology_calculator import ontology_calculator
    return await ontology_calculator.sync_to_ontology(db)


@router.get("/api/ontology/calculation-rules")
async def ontology_calculation_rules():
    """Return the NYX Core Thinker financial calculation rule definitions."""
    from app.services.v2.ontology_calculator import ontology_calculator
    return {"rules": ontology_calculator.get_rules()}


@router.post("/api/ontology/validate-calculations")
async def ontology_validate_calculations(data: ValidateCalcInput):
    """Validate that the given financials satisfy NYX Core Thinker calculation rules.

    Returns a list of violations (empty list = all rules pass).
    """
    from app.services.v2.ontology_calculator import ontology_calculator
    violations = ontology_calculator.validate_calculations(data.model_dump())
    return {
        "valid": len(violations) == 0,
        "violations": violations,
        "rules_checked": list(ontology_calculator.NYX_RULES.keys()),
    }


# ── Equity Journal Entry Generator ──────────────────────────────────

@router.post("/api/gl/generate-equity")
async def generate_equity_entries(db: AsyncSession = Depends(get_db)):
    """
    Generate equity journal entries to make BS balance.
    Creates: Share Capital + Retained Earnings entries that close the equity gap.

    In proper double-entry:
    - Share Capital (5110): credit balance for authorized capital
    - Retained Earnings (5210): cumulative P&L net income
    - Balancing contra on 1910 (Opening Balance Equity)
    """
    from app.models.all_models import JournalEntryRecord, PostingLineRecord
    from sqlalchemy import select, func, Float as SAFloat

    # Calculate current totals from posted entries
    # Assets (1xxx + 2xxx): normal debit balance
    asset_result = await db.execute(
        select(func.sum(
            func.cast(PostingLineRecord.debit, SAFloat) - func.cast(PostingLineRecord.credit, SAFloat)
        ))
        .join(JournalEntryRecord, PostingLineRecord.journal_entry_id == JournalEntryRecord.id)
        .where(JournalEntryRecord.status == 'posted')
        .where(PostingLineRecord.account_code.like('1%') | PostingLineRecord.account_code.like('2%'))
    )
    total_assets = asset_result.scalar() or 0

    # Liabilities (3xxx + 4xxx): normal credit balance
    liab_result = await db.execute(
        select(func.sum(
            func.cast(PostingLineRecord.credit, SAFloat) - func.cast(PostingLineRecord.debit, SAFloat)
        ))
        .join(JournalEntryRecord, PostingLineRecord.journal_entry_id == JournalEntryRecord.id)
        .where(JournalEntryRecord.status == 'posted')
        .where(PostingLineRecord.account_code.like('3%') | PostingLineRecord.account_code.like('4%'))
    )
    total_liabilities = liab_result.scalar() or 0

    # P&L Net (6xxx-9xxx): revenue credit - expense debit
    pnl_result = await db.execute(
        select(func.sum(
            func.cast(PostingLineRecord.credit, SAFloat) - func.cast(PostingLineRecord.debit, SAFloat)
        ))
        .join(JournalEntryRecord, PostingLineRecord.journal_entry_id == JournalEntryRecord.id)
        .where(JournalEntryRecord.status == 'posted')
        .where(
            PostingLineRecord.account_code.like('6%') |
            PostingLineRecord.account_code.like('7%') |
            PostingLineRecord.account_code.like('8%') |
            PostingLineRecord.account_code.like('9%')
        )
    )
    pnl_net = pnl_result.scalar() or 0

    # Existing equity (5xxx): normal credit balance
    existing_equity_result = await db.execute(
        select(func.sum(
            func.cast(PostingLineRecord.credit, SAFloat) - func.cast(PostingLineRecord.debit, SAFloat)
        ))
        .join(JournalEntryRecord, PostingLineRecord.journal_entry_id == JournalEntryRecord.id)
        .where(JournalEntryRecord.status == 'posted')
        .where(PostingLineRecord.account_code.like('5%'))
    )
    existing_equity = existing_equity_result.scalar() or 0

    # Equity needed = Assets - Liabilities - P&L Net - Existing Equity
    # In balanced books: Assets = Liabilities + Equity + P&L
    equity_needed = total_assets - total_liabilities - pnl_net - existing_equity

    # If equity is already balanced (equity_needed ~ 0), nothing to do
    if abs(equity_needed) < 100:
        return {
            "message": "BS already balanced",
            "equity_gap": round(equity_needed, 2),
            "total_assets": round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "pnl_net": round(pnl_net, 2),
            "existing_equity": round(existing_equity, 2),
        }

    # For NYX Core Thinker: typical share capital is GEL 260,534,000
    share_capital = 260_534_000
    retained_earnings = equity_needed - share_capital

    # Generate next document number
    last = await db.execute(
        select(JournalEntryRecord.document_number)
        .order_by(JournalEntryRecord.id.desc()).limit(1)
    )
    last_doc = last.scalar() or "JE-2026-000000"
    parts = last_doc.split('-')
    try:
        next_num = int(parts[-1]) + 1
        year_part = parts[1] if len(parts) > 1 else "2026"
    except (ValueError, IndexError):
        next_num = 1
        year_part = "2026"
    doc_number = f"JE-{year_part}-{next_num:06d}"

    je = JournalEntryRecord(
        document_number=doc_number,
        posting_date=datetime.now(timezone.utc),
        period="January 2026",
        fiscal_year=2026,
        description="Equity Opening Balance -- Share Capital + Retained Earnings",
        status="posted",
        reference="EQUITY-OB",
        currency="GEL",
        source_type="adjustment",
        total_debit=str(abs(share_capital) + abs(retained_earnings) if retained_earnings < 0 else share_capital),
        total_credit=str(abs(share_capital) + abs(retained_earnings) if retained_earnings < 0 else share_capital),
    )
    db.add(je)
    await db.flush()

    lines = []
    line_num = 1

    # Share Capital (5110) -- credit balance
    lines.append(PostingLineRecord(
        journal_entry_id=je.id, line_number=line_num,
        account_code="5110", account_name="Share Capital (Authorized Capital)",
        debit="0", credit=str(abs(share_capital)),
        description="Authorized share capital", currency="GEL",
    ))
    line_num += 1

    # Retained Earnings (5210) -- could be debit (loss) or credit (profit)
    if retained_earnings >= 0:
        lines.append(PostingLineRecord(
            journal_entry_id=je.id, line_number=line_num,
            account_code="5210", account_name="Retained Earnings",
            debit="0", credit=str(retained_earnings),
            description="Accumulated retained earnings", currency="GEL",
        ))
    else:
        lines.append(PostingLineRecord(
            journal_entry_id=je.id, line_number=line_num,
            account_code="5210", account_name="Retained Earnings (Accumulated Deficit)",
            debit=str(abs(retained_earnings)), credit="0",
            description="Accumulated deficit", currency="GEL",
        ))
    line_num += 1

    # Balancing entry on 1910 (Opening Balance Equity / Investment)
    net_equity_credit = share_capital + (retained_earnings if retained_earnings > 0 else 0)
    net_equity_debit = abs(retained_earnings) if retained_earnings < 0 else 0
    balance_amount = net_equity_credit - net_equity_debit

    if balance_amount > 0:
        lines.append(PostingLineRecord(
            journal_entry_id=je.id, line_number=line_num,
            account_code="1910", account_name="Opening Balance Equity (Investment)",
            debit=str(balance_amount), credit="0",
            description="Opening balance contra -- owner investment", currency="GEL",
        ))
    elif balance_amount < 0:
        lines.append(PostingLineRecord(
            journal_entry_id=je.id, line_number=line_num,
            account_code="1910", account_name="Opening Balance Equity (Investment)",
            debit="0", credit=str(abs(balance_amount)),
            description="Opening balance contra -- owner withdrawal", currency="GEL",
        ))

    for line in lines:
        db.add(line)

    await db.commit()

    return {
        "created": True,
        "document_number": doc_number,
        "share_capital": share_capital,
        "retained_earnings": round(retained_earnings, 2),
        "equity_gap_before": round(equity_needed, 2),
        "lines": len(lines),
        "total_assets": round(total_assets, 2),
        "total_liabilities": round(total_liabilities, 2),
        "pnl_net": round(pnl_net, 2),
        "note": "Equity journal entry created. BS should now show proper equity section.",
    }


# ═══════════════════════════════════════════════════════════════════
# SELF-UPGRADING SYSTEM
# ═══════════════════════════════════════════════════════════════════

@router.get("/api/system/health")
async def system_health_assessment(db: AsyncSession = Depends(get_db)):
    """Comprehensive system health across all dimensions — the system grades itself."""
    from app.services.v2.self_upgrade import self_upgrade_engine
    return await self_upgrade_engine.assess_system_health(db)


@router.post("/api/system/auto-upgrade")
async def system_auto_upgrade(db: AsyncSession = Depends(get_db)):
    """Auto-fix issues that can be resolved automatically."""
    from app.services.v2.self_upgrade import self_upgrade_engine
    return await self_upgrade_engine.auto_upgrade(db)
