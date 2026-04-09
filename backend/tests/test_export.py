import os
from openpyxl import load_workbook
from app.utils.excel_export import excel_exporter


class DummyReport:
    def __init__(self):
        self.report_type = "pl"
        self.period = "TestPeriod"
        self.id = 9999
        self.company = "TestCo"
        self.title = "Test Report"
        self.currency = "GEL"
        self.kpis = {"revenue": 1000, "gross_margin": 500}
        # rows include provenance meta
        self.rows = [
            {"c": "1000", "l": "Revenue", "ac": 1000, "pl": 900, "meta": {"tx_ids": ["t1", "t2"], "dataset_id": 1}},
            {"c": "2000", "l": "COGS", "ac": 500, "pl": 400, "meta": {"tx_ids": ["t3"], "dataset_id": 1}},
        ]


def test_export_with_provenance(tmp_path):
    # Temporarily point exporter to tmp path
    old_dir = excel_exporter.export_dir
    excel_exporter.export_dir = tmp_path

    rep = DummyReport()
    path = excel_exporter.export_report_with_provenance(rep)
    assert os.path.exists(path)

    wb = load_workbook(path)
    assert "Provenance" in wb.sheetnames
    ws = wb["Provenance"]
    # header + two rows
    assert ws.max_row >= 3

    # cleanup
    excel_exporter.export_dir = old_dir
    try:
        os.remove(path)
    except Exception:
        pass
