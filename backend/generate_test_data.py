"""
Generate multi-year financial test data for FinAI -Georgian fuel distribution company.
Period: January 2023 → March 2026 (39 monthly files)
Currency: GEL | Company: NYXCoreThinker LLC

Each file: test_data_complex_MonthName_Year.xlsx with 7 sheets matching FinAI parser.
"""
import calendar
import math
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ═══════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════
START_YEAR, START_MONTH = 2023, 1
END_YEAR, END_MONTH = 2026, 3  # inclusive
CURRENCY = "GEL"
MONEY_FMT = '#,##0'
OUTPUT_DIR = Path(__file__).parent / "test_data"

# Styling
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Consolas", size=9)
THIN_BORDER = Border(
    left=Side(style='thin', color='333333'), right=Side(style='thin', color='333333'),
    top=Side(style='thin', color='333333'), bottom=Side(style='thin', color='333333'),
)

# Seasonal fuel demand index (1.0 = baseline Dec 2025)
# Winter high, summer lower
SEASONAL = {
    1: 1.08, 2: 1.05, 3: 0.98, 4: 0.92, 5: 0.88, 6: 0.85,
    7: 0.87, 8: 0.90, 9: 0.95, 10: 1.00, 11: 1.04, 12: 1.10,
}

# Annual growth rate (compounding from Jan 2023 baseline)
ANNUAL_GROWTH = 0.10  # 10% per year

# Base period reference: December 2025 (month_index=35 from Jan 2023)
BASE_MONTH_INDEX = 35


def style_header(ws, row=1, cols=10):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER


def growth_factor(month_index):
    """Scale factor relative to Dec 2025 base amounts. month_index 0=Jan 2023."""
    years_from_base = (month_index - BASE_MONTH_INDEX) / 12.0
    return (1 + ANNUAL_GROWTH) ** years_from_base


def scale(base_amount, month_index, month_num, noise=0.05):
    """Scale a base Dec-2025 amount to the target month with growth + season + noise."""
    g = growth_factor(month_index)
    s = SEASONAL[month_num]
    n = 1.0 + random.uniform(-noise, noise)
    return round(base_amount * g * s * n)


# ═══════════════════════════════════════════════════════════════════════
# PRODUCT DEFINITIONS (base = Dec 2025 monthly gross revenue)
# ═══════════════════════════════════════════════════════════════════════
RETAIL_PRODUCTS = [
    ("Euro Regular A-92", 8_200_000, 0.08),
    ("Euro Premium A-95", 6_800_000, 0.10),
    ("Euro Super A-98", 2_100_000, 0.12),
    ("Euro Diesel", 11_500_000, 0.09),
    ("Euro Diesel Premium", 3_400_000, 0.11),
    ("LPG (Propane-Butane)", 4_200_000, 0.07),
    ("CNG (Compressed Natural Gas)", 1_800_000, 0.06),
    ("Motor Oil & Lubricants", 650_000, 0.35),
    ("AdBlue (DEF)", 280_000, 0.25),
    ("Car Wash Services", 420_000, 0.70),
    ("Convenience Store Sales", 1_100_000, 0.40),
]
WHOLESALE_PRODUCTS = [
    ("Diesel Bulk (Wholesale)", 14_200_000, 0.04),
    ("Petrol Bulk A-92 (Wholesale)", 5_600_000, 0.035),
    ("Petrol Bulk A-95 (Wholesale)", 3_800_000, 0.045),
    ("Bitumen BND 60/90", 2_400_000, 0.05),
    ("Jet Fuel A-1 (Wholesale)", 1_900_000, 0.03),
    ("Fuel Oil M100 (Wholesale)", 1_200_000, 0.025),
    ("LPG Bulk (Wholesale)", 1_600_000, 0.04),
]
OTHER_PRODUCTS = [
    ("Transit Card Revenue", 380_000, 0.90),
    ("BP Transportation Services", 520_000, 0.60),
    ("Lease Revenue (Station Land)", 290_000, 0.85),
    ("Electricity Resale", 180_000, 0.15),
    ("Security Services", 95_000, 0.50),
    ("Auto Maintenance Services", 340_000, 0.55),
]
ELIMINATED_PRODUCTS = [
    ("Intercompany Fuel Transfer - Batumi", -1_800_000, 0.0),
    ("Intercompany Service Charge - Tbilisi HQ", -450_000, 0.0),
]

# ═══════════════════════════════════════════════════════════════════════
# TRIAL BALANCE BASE (Dec 2025 values)
# ═══════════════════════════════════════════════════════════════════════
TB_BASE = [
    ("11", "ფულადი სახსრები", 0, 0, 0, 0),
    ("1110", "სალარო (Petty Cash)", 450_000, 0, 2_800_000, 2_650_000),
    ("1120", "ეროვნული ვალუტის ანგარიში (GEL)", 5_200_000, 0, 48_000_000, 45_800_000),
    ("1130", "უცხოური ვალუტის ანგარიში (USD)", 2_100_000, 0, 8_500_000, 8_200_000),
    ("12", "მოკლევადიანი მოთხოვნები", 0, 0, 0, 0),
    ("1210", "მყიდველთა მოთხოვნები", 12_400_000, 0, 72_000_000, 69_500_000),
    ("1220", "გაცემული ავანსები", 1_800_000, 0, 3_200_000, 3_600_000),
    ("1230", "საეჭვო მოთხოვნების რეზერვი", 0, 680_000, 0, 120_000),
    ("14", "მარაგები", 0, 0, 0, 0),
    ("1410", "საწვავის მარაგი", 8_900_000, 0, 62_000_000, 61_200_000),
    ("1420", "საქონლის მარაგი", 1_200_000, 0, 4_500_000, 4_300_000),
    ("1430", "მასალების მარაგი", 650_000, 0, 1_800_000, 1_700_000),
    ("16", "სხვა მოკლევადიანი აქტივები", 0, 0, 0, 0),
    ("1610", "წინასწარ გადახდილი ხარჯები", 420_000, 0, 1_200_000, 1_100_000),
    ("1620", "დღგ-ის ჩასათვლელი", 1_800_000, 0, 12_960_000, 12_400_000),
    ("21", "ძირითადი საშუალებები", 0, 0, 0, 0),
    ("2110", "მიწის ნაკვეთები", 12_000_000, 0, 0, 0),
    ("2120", "შენობა-ნაგებობები", 28_500_000, 0, 2_400_000, 0),
    ("2130", "მანქანა-დანადგარები", 8_200_000, 0, 1_800_000, 350_000),
    ("2140", "სატრანსპორტო საშუალებები", 4_600_000, 0, 1_200_000, 800_000),
    ("2150", "ცვეთის ჯამი", 0, 18_400_000, 680_000, 1_850_000),
    ("22", "არამატერიალური აქტივები", 0, 0, 0, 0),
    ("2210", "პროგრამული უზრუნველყოფა", 1_200_000, 0, 350_000, 0),
    ("2220", "ლიცენზიები", 800_000, 0, 0, 150_000),
    ("23", "გრძელვადიანი ინვესტიციები", 0, 0, 0, 0),
    ("2310", "შვილობილი კომპანიები", 3_500_000, 0, 0, 0),
    ("31", "მოკლევადიანი ვალდებულებები", 0, 0, 0, 0),
    ("3110", "მომწოდებელთა ვალდებულებები", 0, 14_200_000, 58_000_000, 61_800_000),
    ("3120", "მიღებული ავანსები", 0, 2_100_000, 1_800_000, 2_200_000),
    ("3130", "საგადასახადო ვალდებულებები", 0, 3_400_000, 12_800_000, 13_200_000),
    ("3140", "ხელფასის ვალდებულება", 0, 1_600_000, 4_800_000, 5_200_000),
    ("3150", "დარიცხული ვალდებულებები", 0, 980_000, 2_400_000, 2_600_000),
    ("3160", "მოკლევადიანი სესხები", 0, 8_000_000, 2_000_000, 4_000_000),
    ("3170", "დღგ-ის ვალდებულება", 0, 2_800_000, 11_500_000, 12_960_000),
    ("41", "გრძელვადიანი ვალდებულებები", 0, 0, 0, 0),
    ("4110", "გრძელვადიანი საბანკო სესხი (BoG)", 0, 15_000_000, 1_500_000, 0),
    ("4120", "გრძელვადიანი სესხი (TBC)", 0, 10_000_000, 800_000, 0),
    ("4130", "საპენსიო ვალდებულება", 0, 1_200_000, 0, 150_000),
    ("51", "საკუთარი კაპიტალი", 0, 0, 0, 0),
    ("5110", "საწესდებო კაპიტალი", 0, 20_000_000, 0, 0),
    ("5120", "სარეზერვო კაპიტალი", 0, 2_000_000, 0, 0),
    ("5130", "გაუნაწილებელი მოგება", 0, 5_800_000, 0, 0),
    ("5140", "საკურსო სხვაობა", 0, 320_000, 180_000, 45_000),
    ("61", "შემოსავლები რეალიზაციიდან", 0, 0, 0, 0),
    ("6110", "საცალო რეალიზაცია", 0, 0, 0, 40_550_000),
    ("6120", "საბითუმო რეალიზაცია", 0, 0, 0, 30_700_000),
    ("6130", "სხვა შემოსავალი", 0, 0, 0, 1_805_000),
    ("71", "თვითღირებულება", 0, 0, 0, 0),
    ("7110", "საწვავის თვითღირებულება", 0, 0, 54_200_000, 0),
    ("7120", "ტრანსპორტირების ხარჯი", 0, 0, 3_800_000, 0),
    ("7130", "სხვა პირდაპირი ხარჯი", 0, 0, 1_400_000, 0),
    ("72", "შრომის ხარჯები", 0, 0, 0, 0),
    ("7210", "ხელფასები", 0, 0, 3_200_000, 0),
    ("7220", "სოციალური გადასახადი", 0, 0, 640_000, 0),
    ("7230", "ბონუსები", 0, 0, 280_000, 0),
    ("73", "გაყიდვების ხარჯები", 0, 0, 0, 0),
    ("7310", "რეკლამა და მარკეტინგი", 0, 0, 850_000, 0),
    ("7320", "საკომისიო ხარჯი", 0, 0, 420_000, 0),
    ("74", "ადმინისტრაციული ხარჯები", 0, 0, 0, 0),
    ("7410", "ოფისის ქირა", 0, 0, 480_000, 0),
    ("7420", "კომუნალური", 0, 0, 320_000, 0),
    ("7430", "დაზღვევა", 0, 0, 280_000, 0),
    ("7440", "აუდიტი და კონსულტაცია", 0, 0, 180_000, 0),
    ("7450", "IT ხარჯი", 0, 0, 240_000, 0),
    ("7460", "სამოგზაურო", 0, 0, 120_000, 0),
    ("7470", "საოფისე მასალები", 0, 0, 65_000, 0),
    ("75", "ფინანსური ხარჯები", 0, 0, 0, 0),
    ("7510", "საპროცენტო ხარჯი", 0, 0, 1_250_000, 0),
    ("7520", "საბანკო მომსახურება", 0, 0, 180_000, 0),
    ("76", "ფინანსური შემოსავალი", 0, 0, 0, 0),
    ("7610", "საპროცენტო შემოსავალი", 0, 0, 0, 320_000),
    ("7620", "საკურსო მოგება", 0, 0, 0, 85_000),
    ("77", "საშემოსავლო გადასახადი", 0, 0, 0, 0),
    ("7710", "მიმდინარე მოგების გადასახადი", 0, 0, 980_000, 0),
    ("7720", "გადავადებული გადასახადი", 0, 0, 120_000, 0),
    ("79", "ცვეთა და ამორტიზაცია", 0, 0, 0, 0),
    ("7910", "ძირითადი საშუალებების ცვეთა", 0, 0, 1_450_000, 0),
    ("7920", "არამატერიალურის ამორტიზაცია", 0, 0, 180_000, 0),
    ("81", "არასაოპერაციო შემოსავალი", 0, 0, 0, 0),
    ("8110", "ძირითადი საშუალებების რეალიზაცია", 0, 0, 0, 250_000),
    ("82", "არასაოპერაციო ხარჯი", 0, 0, 0, 0),
    ("8210", "ჩამოწერები", 0, 0, 180_000, 0),
    ("8230", "დანაკლისი", 0, 0, 95_000, 0),
]

# ═══════════════════════════════════════════════════════════════════════
# BALANCE SHEET BASE (Dec 2025)
# ═══════════════════════════════════════════════════════════════════════
BS_BASE = [
    ("1110", "Cash & Equivalents", "Current Assets", "CA.Cash", 5_200_000, 51_300_000, 48_450_000),
    ("1120", "Bank GEL", "Current Assets", "CA.Bank", 5_200_000, 48_000_000, 45_800_000),
    ("1130", "Bank USD", "Current Assets", "CA.BankFX", 2_100_000, 8_500_000, 8_200_000),
    ("1210", "Trade Receivables", "Current Assets", "CA.AR", 12_400_000, 72_000_000, 69_500_000),
    ("1220", "Prepaid Advances", "Current Assets", "CA.Prepaid", 1_800_000, 3_200_000, 3_600_000),
    ("1230", "Bad Debt Provision", "Current Assets", "CA.BadDebt", -680_000, 0, 120_000),
    ("1410", "Fuel Inventory", "Current Assets", "CA.InvFuel", 8_900_000, 62_000_000, 61_200_000),
    ("1420", "Merchandise", "Current Assets", "CA.InvMerch", 1_200_000, 4_500_000, 4_300_000),
    ("1430", "Spare Parts", "Current Assets", "CA.InvParts", 650_000, 1_800_000, 1_700_000),
    ("1610", "Prepaid Expenses", "Current Assets", "CA.PrepaidExp", 420_000, 1_200_000, 1_100_000),
    ("1620", "Input VAT", "Current Assets", "CA.VAT", 1_800_000, 12_960_000, 12_400_000),
    ("2110", "Land", "Non-current Assets", "NCA.Land", 12_000_000, 0, 0),
    ("2120", "Buildings", "Non-current Assets", "NCA.Build", 28_500_000, 2_400_000, 0),
    ("2130", "Equipment", "Non-current Assets", "NCA.Equip", 8_200_000, 1_800_000, 350_000),
    ("2140", "Vehicles", "Non-current Assets", "NCA.Vehicle", 4_600_000, 1_200_000, 800_000),
    ("2150", "Accum. Depreciation", "Non-current Assets", "NCA.AccDepr", -18_400_000, 680_000, 1_850_000),
    ("2210", "Software", "Non-current Assets", "NCA.SW", 1_200_000, 350_000, 0),
    ("2220", "Licenses", "Non-current Assets", "NCA.Lic", 800_000, 0, 150_000),
    ("2310", "Subsidiaries", "Non-current Assets", "NCA.Invest", 3_500_000, 0, 0),
    ("3110", "Trade Payables", "Current Liabilities", "CL.AP", -14_200_000, 58_000_000, 61_800_000),
    ("3120", "Advances Received", "Current Liabilities", "CL.Adv", -2_100_000, 1_800_000, 2_200_000),
    ("3130", "Tax Payables", "Current Liabilities", "CL.Tax", -3_400_000, 12_800_000, 13_200_000),
    ("3140", "Salary Payable", "Current Liabilities", "CL.Salary", -1_600_000, 4_800_000, 5_200_000),
    ("3150", "Accrued Expenses", "Current Liabilities", "CL.Accrued", -980_000, 2_400_000, 2_600_000),
    ("3160", "Short-term Loans", "Current Liabilities", "CL.STLoan", -8_000_000, 2_000_000, 4_000_000),
    ("3170", "Output VAT", "Current Liabilities", "CL.VAT", -2_800_000, 11_500_000, 12_960_000),
    ("4110", "LT Loan (BoG)", "Non-current Liabilities", "NCL.Loan1", -15_000_000, 1_500_000, 0),
    ("4120", "LT Loan (TBC)", "Non-current Liabilities", "NCL.Loan2", -10_000_000, 800_000, 0),
    ("4130", "Pension", "Non-current Liabilities", "NCL.Pension", -1_200_000, 0, 150_000),
    ("5110", "Share Capital", "Equity", "EQ.Cap", -20_000_000, 0, 0),
    ("5120", "Legal Reserve", "Equity", "EQ.Reserve", -2_000_000, 0, 0),
    ("5130", "Retained Earnings", "Equity", "EQ.RE", -5_800_000, 0, 0),
    ("5140", "FX Reserve", "Equity", "EQ.FX", -320_000, 180_000, 45_000),
]

# ═══════════════════════════════════════════════════════════════════════
# G&A BASE (Dec 2025)
# ═══════════════════════════════════════════════════════════════════════
GA_BASE = [
    ("7210", "Salaries & Wages", 3_200_000), ("7220", "Social Security", 640_000),
    ("7230", "Bonuses", 280_000), ("7310", "Advertising & Marketing", 850_000),
    ("7320", "Sales Commissions", 420_000), ("7410", "Office Rent", 480_000),
    ("7420", "Utilities", 320_000), ("7430", "Insurance", 280_000),
    ("7440", "Audit & Consulting", 180_000), ("7450", "IT & Software", 240_000),
    ("7460", "Travel & Entertainment", 120_000), ("7470", "Office Supplies", 65_000),
    ("7510", "Interest Expense", 1_250_000), ("7520", "Bank Charges", 180_000),
    ("7610", "Interest Income", -320_000), ("7620", "FX Gain", -85_000),
    ("7710", "Income Tax", 980_000), ("7720", "Deferred Tax", 120_000),
    ("7910", "Depreciation", 1_450_000), ("7920", "Amortization", 180_000),
    ("7120", "Transport Costs", 3_800_000), ("7130", "Other Direct Costs", 1_400_000),
    ("8110", "Asset Disposal Gain", -250_000), ("8210", "Write-offs", 180_000),
    ("8230", "Shortages", 95_000),
]

# ═══════════════════════════════════════════════════════════════════════
# BUDGET BASE (Dec 2025)
# ═══════════════════════════════════════════════════════════════════════
BUD_BASE = [
    ("Revenue - Retail", 42_000_000, 40_550_000),
    ("Revenue - Wholesale", 32_000_000, 30_700_000),
    ("Revenue - Other", 2_000_000, 1_805_000),
    ("COGS - Fuel", 55_000_000, 54_200_000),
    ("COGS - Transport", 4_000_000, 3_800_000),
    ("COGS - Other Direct", 1_500_000, 1_400_000),
    ("Labour Costs", 4_200_000, 4_120_000),
    ("Selling Expenses", 1_400_000, 1_270_000),
    ("G&A Expenses", 1_900_000, 1_822_000),
    ("D&A", 1_700_000, 1_630_000),
    ("Finance Costs (net)", 1_200_000, 1_025_000),
    ("Income Tax", 1_000_000, 1_100_000),
    ("Capital Expenditure", 6_000_000, 5_400_000),
]

# Transaction templates
TEMPLATES = [
    ("Revenue", "1210", "6110", (50_000, 800_000), "Retail Operations", "Direct Variable",
     ["SOCAR Energy Georgia", "Wissol Petroleum", "Gulf Georgia"]),
    ("Revenue", "1210", "6120", (200_000, 2_000_000), "Wholesale", "Direct Variable",
     ["Rompetrol Georgia", "GOGC", "Black Sea Transit Corp"]),
    ("Purchase", "7110", "3110", (100_000, 1_500_000), "Logistics & Supply", "Direct Variable",
     ["SOCAR Energy Georgia", "Batumi Oil Terminal", "Kulevi Terminal"]),
    ("Purchase", "1410", "3110", (50_000, 500_000), "Logistics & Supply", "Direct Variable",
     ["SOCAR Energy Georgia", "Rompetrol Georgia"]),
    ("Salary", "7210", "3140", (200_000, 400_000), "HR", "Direct Fixed", ["Internal Payroll"]),
    ("Tax", "3130", "1120", (100_000, 500_000), "Finance & Treasury", "Indirect Fixed",
     ["RS.GE (Revenue Service)"]),
    ("Loan", "4110", "1120", (100_000, 250_000), "Finance & Treasury", "Capital",
     ["Bank of Georgia"]),
    ("Interest", "7510", "1120", (80_000, 200_000), "Finance & Treasury", "Indirect Fixed",
     ["Bank of Georgia", "TBC Bank"]),
    ("Depreciation", "7910", "2150", (100_000, 200_000), "Administration", "Indirect Fixed",
     ["Internal"]),
    ("Utility", "7420", "3110", (15_000, 45_000), "Administration", "Indirect Variable",
     ["Energo-Pro Georgia"]),
    ("Rent", "7410", "3110", (35_000, 60_000), "Administration", "Indirect Fixed",
     ["Various Landlords"]),
    ("Marketing", "7310", "3110", (30_000, 120_000), "Retail Operations", "Indirect Variable",
     ["Various Media"]),
    ("Bank Fee", "7520", "1120", (5_000, 25_000), "Finance & Treasury", "Indirect Fixed",
     ["Bank of Georgia", "TBC Bank"]),
    ("Cash Collection", "1120", "1210", (100_000, 1_200_000), "Finance & Treasury",
     "Direct Variable", ["Bank of Georgia"]),
    ("Supplier Payment", "3110", "1120", (150_000, 2_000_000), "Finance & Treasury",
     "Direct Variable", ["SOCAR Energy Georgia", "Batumi Oil Terminal"]),
    ("VAT Payment", "3170", "1120", (200_000, 800_000), "Finance & Treasury", "Indirect Fixed",
     ["RS.GE (Revenue Service)"]),
]
RECORDERS = ["N. Kapanadze", "G. Tsereteli", "M. Gvazava", "T. Lomidze", "S. Beridze",
             "I. Kvirikashvili"]


# ═══════════════════════════════════════════════════════════════════════
# GENERATOR
# ═══════════════════════════════════════════════════════════════════════

def generate_month(year, month, month_index, prev_bs_closing):
    """Generate one monthly Excel file. Returns BS closing balances for next month."""
    random.seed(year * 100 + month)  # reproducible per month

    period_name = f"{calendar.month_name[month]} {year}"
    mi = month_index
    mn = month

    wb = Workbook()

    # ── Sheet 1: Revenue ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Revenue"
    for i, h in enumerate(["Product", "Gross", "VAT", "Net", "Segment", "Eliminated"], 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, 6)

    row = 2
    total_retail = total_wholesale = total_other = 0
    for products, segment in [
        (RETAIL_PRODUCTS, "Revenue Retail"),
        (WHOLESALE_PRODUCTS, "Revenue Wholesale"),
        (OTHER_PRODUCTS, "Other Revenue"),
    ]:
        for name, base_gross, _ in products:
            gross = scale(base_gross, mi, mn)
            vat = round(gross * 18 / 118)
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=gross).number_format = MONEY_FMT
            ws.cell(row=row, column=3, value=vat).number_format = MONEY_FMT
            ws.cell(row=row, column=4, value=gross - vat).number_format = MONEY_FMT
            ws.cell(row=row, column=5, value=segment)
            ws.cell(row=row, column=6, value="")
            if segment == "Revenue Retail":
                total_retail += gross
            elif segment == "Revenue Wholesale":
                total_wholesale += gross
            else:
                total_other += gross
            row += 1

    # Credit note anomaly (every ~6 months)
    if month % 6 == 0:
        cn_amt = scale(320_000, mi, mn)
        ws.cell(row=row, column=1, value="CREDIT NOTE: Diesel Returns (Defective Batch)")
        ws.cell(row=row, column=2, value=-cn_amt).number_format = MONEY_FMT
        ws.cell(row=row, column=3, value=-round(cn_amt * 18 / 118)).number_format = MONEY_FMT
        ws.cell(row=row, column=4, value=-cn_amt + round(cn_amt * 18 / 118)).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value="Revenue Retail")
        total_retail -= cn_amt
        row += 1

    for name, base_gross, _ in ELIMINATED_PRODUCTS:
        gross = scale(base_gross, mi, mn, noise=0.02)
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=gross).number_format = MONEY_FMT
        ws.cell(row=row, column=3, value=0)
        ws.cell(row=row, column=4, value=gross).number_format = MONEY_FMT
        ws.cell(row=row, column=5, value="Eliminated")
        ws.cell(row=row, column=6, value="TRUE")
        row += 1

    n_rev = row - 2
    style_data(ws, 2, row - 1, 6)
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20

    # ── Sheet 2: COGS ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("COGS")
    for i, h in enumerate(["Субконто", "6", "7310", "8230", "1610", "Кред. оборот"], 1):
        ws2.cell(row=1, column=i, value=h)
    style_header(ws2, 1, 6)

    row = 2
    for name, base_gross, margin in RETAIL_PRODUCTS + WHOLESALE_PRODUCTS:
        total_cost = round(scale(base_gross, mi, mn) * (1 - margin))
        a6 = round(total_cost * 0.82)
        a73 = round(total_cost * 0.12)
        a82 = round(total_cost * 0.06)
        a16 = round(total_cost * 0.03) if random.random() > 0.6 else 0
        ws2.cell(row=row, column=1, value=name)
        ws2.cell(row=row, column=2, value=a6).number_format = MONEY_FMT
        ws2.cell(row=row, column=3, value=a73).number_format = MONEY_FMT
        ws2.cell(row=row, column=4, value=a82).number_format = MONEY_FMT
        ws2.cell(row=row, column=5, value=a16).number_format = MONEY_FMT
        ws2.cell(row=row, column=6, value=a6 + a73 + a82).number_format = MONEY_FMT
        row += 1

    n_cogs = row - 2
    style_data(ws2, 2, row - 1, 6)
    ws2.column_dimensions['A'].width = 40

    # ── Sheet 3: Trial Balance (TDSheet) ──────────────────────────────
    ws3 = wb.create_sheet("TDSheet")
    tb_h = ["Счет", "Субконто", "Нач. дебет", "Нач. кредит",
            "Дебет. оборот", "Кредит. оборот", "Сальдо дебет", "Сальдо кредит"]
    for i, h in enumerate(tb_h, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header(ws3, 1, 8)

    # Build TB with scaled turnovers, opening from prev closing
    tb_rows = []
    for code, name, base_od, base_oc, base_td, base_tc in TB_BASE:
        is_parent = len(code) <= 2
        if is_parent:
            tb_rows.append((code, name, 0, 0, 0, 0, True))
            continue

        # Scale turnovers (P&L accounts scale with season; BS accounts scale less)
        is_pnl = code[0] in ('6', '7', '8')
        if is_pnl:
            td = scale(base_td, mi, mn) if base_td else 0
            tc = scale(base_tc, mi, mn) if base_tc else 0
            od = 0
            oc = 0
        else:
            td = scale(base_td, mi, mn, noise=0.08) if base_td else 0
            tc = scale(base_tc, mi, mn, noise=0.08) if base_tc else 0
            # Opening balances: use prev_bs_closing if available, else scale from base
            if prev_bs_closing and code in prev_bs_closing:
                bal = prev_bs_closing[code]
                od = max(0, bal)
                oc = max(0, -bal)
            else:
                od = scale(base_od, mi, mn, noise=0.03) if base_od else 0
                oc = scale(base_oc, mi, mn, noise=0.03) if base_oc else 0

        tb_rows.append((code, name, od, oc, td, tc, False))

    # Compute parent aggregates and closing balances
    row = 2
    for idx, (code, name, od, oc, td, tc, is_parent) in enumerate(tb_rows):
        if is_parent:
            children = [(c, n, a, b, d, e) for c, n, a, b, d, e, p in tb_rows
                        if c.startswith(code) and len(c) == 4 and not p]
            od = sum(x[2] for x in children)
            oc = sum(x[3] for x in children)
            td = sum(x[4] for x in children)
            tc = sum(x[5] for x in children)
        cd = max(0, od + td - tc)
        cc = max(0, oc + tc - td)
        ws3.cell(row=row, column=1, value=code)
        ws3.cell(row=row, column=2, value=name)
        for ci, v in enumerate([od, oc, td, tc, cd, cc], 3):
            ws3.cell(row=row, column=ci, value=v).number_format = MONEY_FMT
        if is_parent:
            for c in range(1, 9):
                ws3.cell(row=row, column=c).font = Font(name="Consolas", bold=True, size=10)
        row += 1

    n_tb = row - 2
    style_data(ws3, 2, row - 1, 8)
    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 45

    # ── Sheet 4: Balance Sheet ────────────────────────────────────────
    ws4 = wb.create_sheet("BS")
    for i, h in enumerate(["Счет", "Наименование", "MAPPING GRP", "MAPING BAKU",
                            "Остаток начисл.", "Дебет. оборот", "Кредит. оборот",
                            "Остаток конечн."], 1):
        ws4.cell(row=1, column=i, value=h)
    style_header(ws4, 1, 8)

    new_bs_closing = {}
    row = 2
    for code, name, mg, mb, base_op, base_td, base_tc in BS_BASE:
        is_pnl_related = code[0] in ('6', '7', '8')
        if prev_bs_closing and code in prev_bs_closing:
            op = prev_bs_closing[code]
        else:
            op = scale(base_op, mi, mn, noise=0.03)

        td = scale(base_td, mi, mn, noise=0.06)
        tc = scale(base_tc, mi, mn, noise=0.06)
        closing = op + td - tc

        # Fixed assets: land/buildings don't fluctuate much
        if code in ("2110",):
            td, tc = 0, 0
            closing = op
        # Equity: share capital/reserve are stable
        if code in ("5110", "5120"):
            td, tc = 0, 0
            closing = op

        new_bs_closing[code] = closing

        ws4.cell(row=row, column=1, value=code)
        ws4.cell(row=row, column=2, value=name)
        ws4.cell(row=row, column=3, value=mg)
        ws4.cell(row=row, column=4, value=mb)
        ws4.cell(row=row, column=5, value=op).number_format = MONEY_FMT
        ws4.cell(row=row, column=6, value=td).number_format = MONEY_FMT
        ws4.cell(row=row, column=7, value=tc).number_format = MONEY_FMT
        ws4.cell(row=row, column=8, value=closing).number_format = MONEY_FMT
        row += 1

    style_data(ws4, 2, row - 1, 8)
    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 28
    ws4.column_dimensions['C'].width = 22

    # ── Sheet 5: Transactions ─────────────────────────────────────────
    ws5 = wb.create_sheet("Transactions")
    txn_h = ["Date", "Recorder", "Account DR", "Account CR", "Department Eng",
             "Counterparty Eng", "Cost Classification", "Type", "Amount GEL", "VAT", "Currency"]
    for i, h in enumerate(txn_h, 1):
        ws5.cell(row=1, column=i, value=h)
    style_header(ws5, 1, len(txn_h))

    days_in_month = calendar.monthrange(year, month)[1]
    amt_scale = growth_factor(mi) * SEASONAL[mn]

    row = 2
    for day in range(1, days_in_month + 1):
        dt = datetime(year, month, day)
        if dt.weekday() >= 6:
            continue
        n = random.randint(12, 25) if 10 <= day <= 25 else random.randint(5, 12)
        for _ in range(n):
            t = random.choice(TEMPLATES)
            typ, dr, cr, (lo, hi), dept, cc, cps = t
            lo_s = round(lo * amt_scale)
            hi_s = round(hi * amt_scale)
            amt = round(random.randint(lo_s, hi_s) / 100) * 100
            vat = round(amt * 18 / 118) if typ in ("Revenue", "Purchase") else 0
            ws5.cell(row=row, column=1, value=dt.strftime("%Y-%m-%d"))
            ws5.cell(row=row, column=2, value=random.choice(RECORDERS))
            ws5.cell(row=row, column=3, value=dr)
            ws5.cell(row=row, column=4, value=cr)
            ws5.cell(row=row, column=5, value=dept)
            ws5.cell(row=row, column=6, value=random.choice(cps))
            ws5.cell(row=row, column=7, value=cc)
            ws5.cell(row=row, column=8, value=typ)
            ws5.cell(row=row, column=9, value=amt).number_format = MONEY_FMT
            ws5.cell(row=row, column=10, value=vat).number_format = MONEY_FMT
            ws5.cell(row=row, column=11, value=CURRENCY)
            row += 1

    n_txn = row - 2
    style_data(ws5, 2, row - 1, len(txn_h))
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['F'].width = 28

    # ── Sheet 6: Mapping (G&A) ────────────────────────────────────────
    ws6 = wb.create_sheet("Mapping")
    for i, h in enumerate(["Account Code", "Account Name", "Amount"], 1):
        ws6.cell(row=1, column=i, value=h)
    style_header(ws6, 1, 3)

    row = 2
    for code, name, base_amt in GA_BASE:
        amt = scale(abs(base_amt), mi, mn, noise=0.04)
        if base_amt < 0:
            amt = -amt
        ws6.cell(row=row, column=1, value=code)
        ws6.cell(row=row, column=2, value=name)
        ws6.cell(row=row, column=3, value=amt).number_format = MONEY_FMT
        row += 1
    style_data(ws6, 2, row - 1, 3)
    ws6.column_dimensions['A'].width = 14
    ws6.column_dimensions['B'].width = 30

    # ── Sheet 7: Budget ───────────────────────────────────────────────
    ws7 = wb.create_sheet("Budget")
    for i, h in enumerate(["Category", "Budget Amount", "Actual Amount", "Variance",
                            "Variance %"], 1):
        ws7.cell(row=1, column=i, value=h)
    style_header(ws7, 1, 5)

    row = 2
    for cat, base_bud, base_act in BUD_BASE:
        bud = scale(base_bud, mi, mn, noise=0.02)
        # Actual deviates from budget by -8% to +5%
        act = round(bud * (1 + random.uniform(-0.08, 0.05)))
        var = act - bud
        ws7.cell(row=row, column=1, value=cat)
        ws7.cell(row=row, column=2, value=bud).number_format = MONEY_FMT
        ws7.cell(row=row, column=3, value=act).number_format = MONEY_FMT
        ws7.cell(row=row, column=4, value=var).number_format = MONEY_FMT
        ws7.cell(row=row, column=5, value=round(var / bud * 100, 1) if bud else 0)
        row += 1
    style_data(ws7, 2, row - 1, 5)
    ws7.column_dimensions['A'].width = 28

    # ── Save ──────────────────────────────────────────────────────────
    fname = f"test_data_complex_{calendar.month_name[month]}_{year}.xlsx"
    out = OUTPUT_DIR / fname
    wb.save(str(out))

    return n_rev, n_cogs, n_tb, n_txn, new_bs_closing


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    months = []
    y, m = START_YEAR, START_MONTH
    while (y, m) <= (END_YEAR, END_MONTH):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1

    print(f"Generating {len(months)} monthly files: "
          f"{calendar.month_name[months[0][1]]} {months[0][0]} -> "
          f"{calendar.month_name[months[-1][1]]} {months[-1][0]}")
    print(f"Output: {OUTPUT_DIR}\n")

    prev_bs = None
    total_txn = 0

    for idx, (year, month) in enumerate(months):
        n_rev, n_cogs, n_tb, n_txn, prev_bs = generate_month(
            year, month, idx, prev_bs
        )
        total_txn += n_txn
        period = f"{calendar.month_name[month]} {year}"
        print(f"  [{idx + 1:2d}/{len(months)}] {period:20s} -"
              f"Rev:{n_rev:3d} COGS:{n_cogs:2d} TB:{n_tb:2d} Txn:{n_txn:4d}")

    print(f"\n{'=' * 60}")
    print(f"  Total files: {len(months)}")
    print(f"  Total transactions: {total_txn:,}")
    print(f"  Period: {calendar.month_name[months[0][1]]} {months[0][0]} -> "
          f"{calendar.month_name[months[-1][1]]} {months[-1][0]}")
    print(f"  Output directory: {OUTPUT_DIR}")
    print(f"  Sheets per file: Revenue | COGS | TDSheet | BS | Transactions | Mapping | Budget")


if __name__ == "__main__":
    main()
