"""
Phase P-2: Professional Excel Report Export
=============================================
Generates multi-sheet Excel workbooks with:
  - Executive Summary (formatted)
  - P&L Statement (with formulas)
  - Financial Ratios
  - Diagnosis & Recommendations
  - Raw Data

Uses openpyxl for formatting, formulas, and conditional formatting.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from app.config import settings

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════

_NAVY_FILL = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid") if OPENPYXL_AVAILABLE else None
_LIGHT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") if OPENPYXL_AVAILABLE else None
_GOLD_FILL = PatternFill(start_color="C4A35A", end_color="C4A35A", fill_type="solid") if OPENPYXL_AVAILABLE else None
_WHITE_FONT = Font(color="FFFFFF", bold=True) if OPENPYXL_AVAILABLE else None
_NAVY_FONT = Font(color="1B2A4A", bold=True) if OPENPYXL_AVAILABLE else None
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11) if OPENPYXL_AVAILABLE else None
_THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
) if OPENPYXL_AVAILABLE else None


def _auto_width(ws, min_width=12, max_width=40):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value or ""))
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ═══════════════════════════════════════════════════════════════════
# EXCEL REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════════

class ExcelReportGenerator:
    """Generates professional multi-sheet Excel workbooks from orchestrator output."""

    def generate(
        self,
        result_dict: Dict[str, Any],
        company_name: str = None,
        period: str = "FY 2025",
    ) -> bytes:
        """
        Generate Excel workbook.

        Returns:
            Excel file bytes
        """
        company_name = company_name or settings.COMPANY_NAME
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl required. Install: pip install openpyxl")

        wb = openpyxl.Workbook()

        exec_summary = result_dict.get("executive_summary", {})
        diagnosis = result_dict.get("diagnosis", {})
        decision = result_dict.get("decision", {})
        strategy = result_dict.get("strategy", {})

        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        self._build_summary_sheet(ws1, company_name, period, exec_summary)

        # Sheet 2: P&L Statement
        ws2 = wb.create_sheet("P&L Statement")
        self._build_pl_sheet(ws2, diagnosis, period)

        # Sheet 3: Financial Ratios
        ws3 = wb.create_sheet("Financial Ratios")
        self._build_ratios_sheet(ws3, diagnosis)

        # Sheet 4: Strategy & Actions
        ws4 = wb.create_sheet("Strategy")
        self._build_strategy_sheet(ws4, strategy, decision)

        # Sheet 5: Raw Data
        ws5 = wb.create_sheet("Raw Data")
        self._build_raw_sheet(ws5, diagnosis)

        # Freeze panes
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _build_summary_sheet(self, ws, company, period, summary):
        """Build executive summary sheet."""
        # Header
        ws.merge_cells("A1:D1")
        ws["A1"] = f"{company} - Financial Intelligence Summary"
        ws["A1"].font = Font(size=16, bold=True, color="1B2A4A")
        ws["A2"] = f"Period: {period}"
        ws["A2"].font = Font(size=11, color="888888")

        # Metrics
        metrics = [
            ("Health Score", summary.get("health_score", 0), "/100"),
            ("Health Grade", summary.get("health_grade", "?"), ""),
            ("Strategy", summary.get("strategy_name", "N/A"), ""),
            ("Conviction Grade", summary.get("conviction_grade", "N/A"), ""),
            ("Cash Runway", summary.get("cash_runway_months", 0), " months"),
            ("Active Alerts", summary.get("active_alerts", 0), ""),
            ("KPIs On Track", summary.get("kpi_on_track", 0), ""),
            ("Stages Completed", summary.get("stages_completed", 0), ""),
            ("System Health", summary.get("system_health", "unknown"), ""),
        ]

        row = 4
        ws.cell(row=row, column=1, value="Metric").font = _HEADER_FONT
        ws.cell(row=row, column=2, value="Value").font = _HEADER_FONT
        for c in [1, 2]:
            ws.cell(row=row, column=c).fill = _NAVY_FILL

        for label, value, suffix in metrics:
            row += 1
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=f"{value}{suffix}")
            if row % 2 == 0:
                for c in [1, 2]:
                    ws.cell(row=row, column=c).fill = _LIGHT_FILL

        _auto_width(ws)

    def _build_pl_sheet(self, ws, diagnosis, period):
        """Build P&L with formulas."""
        data = diagnosis.get("financial_data", {})

        ws["A1"] = "Income Statement"
        ws["A1"].font = Font(size=14, bold=True, color="1B2A4A")

        headers = ["Line Item", "Amount (GEL)", "% of Revenue"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = _HEADER_FONT
            c.fill = _NAVY_FILL

        pl_items = [
            ("Revenue", data.get("revenue", 0)),
            ("Cost of Goods Sold", data.get("cogs", 0)),
            ("Gross Profit", None),  # formula
            ("G&A Expenses", data.get("ga_expenses", 0)),
            ("EBITDA", None),  # formula
            ("Depreciation", data.get("depreciation", 0)),
            ("Finance Expense", data.get("finance_expense", 0)),
            ("Tax Expense", data.get("tax_expense", 0)),
            ("Net Profit", None),  # formula
        ]

        row = 4
        for label, value in pl_items:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                c = ws.cell(row=row, column=2, value=value)
                c.number_format = '#,##0'
            elif label == "Gross Profit":
                ws.cell(row=row, column=2, value=f"=B4-B5")  # Revenue - COGS
                ws.cell(row=row, column=2).number_format = '#,##0'
            elif label == "EBITDA":
                ws.cell(row=row, column=2, value=f"=B6-B7")  # GP - G&A
                ws.cell(row=row, column=2).number_format = '#,##0'
            elif label == "Net Profit":
                ws.cell(row=row, column=2, value=f"=B8-B9-B10-B11")  # EBITDA - Dep - Fin - Tax
                ws.cell(row=row, column=2).number_format = '#,##0'

            # % of Revenue formula
            if row > 3:
                ws.cell(row=row, column=3, value=f"=IF(B4=0,0,B{row}/B4)")
                ws.cell(row=row, column=3).number_format = '0.0%'

            # Bold subtotals
            if label in ("Gross Profit", "EBITDA", "Net Profit"):
                for c in range(1, 4):
                    ws.cell(row=row, column=c).font = _NAVY_FONT
                    ws.cell(row=row, column=c).fill = _LIGHT_FILL

            # Borders
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = _THIN_BORDER

            row += 1

        _auto_width(ws)

    def _build_ratios_sheet(self, ws, diagnosis):
        """Build financial ratios sheet."""
        data = diagnosis.get("financial_data", {})

        ws["A1"] = "Financial Ratios"
        ws["A1"].font = Font(size=14, bold=True, color="1B2A4A")

        headers = ["Ratio", "Value", "Benchmark"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = _HEADER_FONT
            c.fill = _NAVY_FILL

        ratios = [
            ("Gross Margin %", data.get("gross_margin_pct"), "15%+"),
            ("Net Margin %", data.get("net_margin_pct"), "5%+"),
            ("EBITDA Margin %", data.get("ebitda_margin_pct"), "8%+"),
            ("COGS/Revenue %", data.get("cogs_to_revenue_pct"), "<85%"),
        ]

        for i, (label, value, bench) in enumerate(ratios, 4):
            ws.cell(row=i, column=1, value=label)
            c = ws.cell(row=i, column=2, value=value)
            c.number_format = '0.0'
            ws.cell(row=i, column=3, value=bench)

        _auto_width(ws)

    def _build_strategy_sheet(self, ws, strategy, decision):
        """Build strategy sheet."""
        strategy = strategy or {}
        decision = decision or {}
        ws["A1"] = "Strategic Plan"
        ws["A1"].font = Font(size=14, bold=True, color="1B2A4A")

        ws["A3"] = "Strategy"
        ws["B3"] = strategy.get("name", "N/A")
        ws["A4"] = "Duration"
        ws["B4"] = f"{strategy.get('total_duration_days', 0)} days"
        ws["A5"] = "Investment"
        ws["B5"] = strategy.get("total_investment", 0)
        ws["B5"].number_format = '#,##0'

        phases = strategy.get("phases", [])
        if phases:
            row = 7
            headers = ["Phase", "Name", "Duration", "Expected Impact"]
            for i, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=i, value=h)
                c.font = _HEADER_FONT
                c.fill = _NAVY_FILL

            for p in phases:
                row += 1
                ws.cell(row=row, column=1, value=p.get("phase_number", ""))
                ws.cell(row=row, column=2, value=p.get("phase_name", "").title())
                ws.cell(row=row, column=3, value=f"{p.get('duration_days', 0)} days")
                ws.cell(row=row, column=4, value=p.get("expected_profit_delta", 0))
                ws.cell(row=row, column=4).number_format = '#,##0'

        _auto_width(ws)

    def _build_raw_sheet(self, ws, diagnosis):
        """Build raw data sheet."""
        data = diagnosis.get("financial_data", {})

        ws["A1"] = "Raw Financial Data"
        ws["A1"].font = Font(size=14, bold=True, color="1B2A4A")

        headers = ["Field", "Value"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.font = _HEADER_FONT
            c.fill = _NAVY_FILL

        row = 4
        for key, value in sorted(data.items()):
            ws.cell(row=row, column=1, value=key)
            c = ws.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)):
                c.number_format = '#,##0.00'
            row += 1

        _auto_width(ws)


# Module-level singleton
excel_report = ExcelReportGenerator()
